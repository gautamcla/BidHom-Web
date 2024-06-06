# -*- coding: utf-8 -*-
"""This file contains general view functions for this module
"""
import json
import datetime
import time
import io
from datetime import timedelta
from django.shortcuts import render, redirect
from django.conf import settings
from django.http import JsonResponse
from django.http import HttpResponse
# from home.services import *
from django.core.cache import cache
from django.conf import settings
from subdomain.constants import *
from subdomain.services import call_api_get_method, call_api_post_method
from django.http import HttpResponseRedirect
from packages.context_processors import subdomain_site_details
from packages.globalfunction import *
from django.core.cache.backends.base import DEFAULT_TIMEOUT
from geopy.geocoders import Nominatim
import urllib.request
from urllib.request import urlopen, unquote
from urllib.parse import parse_qs, urlparse
from pyzipcode import ZipCodeDatabase
import math
from wsgiref.util import FileWrapper
CACHE_TTL = getattr(settings, 'CACHE_TTL', DEFAULT_TIMEOUT)
from django.views.decorators.csrf import csrf_exempt
from django.template.loader import get_template
from io import BytesIO
from io import StringIO
from xhtml2pdf import pisa
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from django.template import Context
from packages.constants import *


def dashboard(request):
    """Home page function
        """
    try:
        try:
            theme = 1
            user_id = None
            try:
                token = request.session['token']['access_token']
                user_id = request.session['user_id']

            except Exception as exp:
                print(exp)

            try:
                is_first_login = request.session['is_first_login'] if 'is_first_login' in request.session else 0
            except Exception as exp:
                is_first_login = 0

            try:
                site_detail = subdomain_site_details(request)
                site_id = site_detail['site_detail']['site_id']
                templete_dir = site_detail['site_detail']['theme_directory']
            except Exception as exp:
                site_id = ""
                templete_dir = 'theme-1'
            static_dir = templete_dir


            try:
                dashboard_data_params = {'site_id': site_id, 'user_id': user_id}
                dashboard_data_url = settings.API_URL + '/api-home/detail/'
                response_data = call_api_post_method(dashboard_data_params, dashboard_data_url)
                dashboard_data = response_data['data']
                social_account_list = dashboard_data['social_account']

                auction_name = []
                expertise_name = []
                if templete_dir == "theme-4":
                    if dashboard_data['active_auction'] > 0:
                        for i in dashboard_data['auctions']:
                            if i['status'] == 1:
                                temp = {}
                                auction_data = i['auction_name'].split(" ")
                                if len(auction_data) == 2:
                                    temp['auction_name'] = "<span>" + auction_data[0] + "</span><br>" + auction_data[1]
                                    temp['row'] = i['auction_name'].lower()
                                elif len(auction_data) >= 4:
                                    str_1 = auction_data[0] + " " + auction_data[1]
                                    str_2 = ""
                                    for j, k in enumerate(auction_data):
                                        if j > 1:
                                            str_2 += k + " "
                                    temp['auction_name'] = "<span>" + str_1 + "</span><br>" + str_2
                                    temp['row'] = i['auction_name'].lower()
                                else:
                                    str_1 = auction_data[0]
                                    str_2 = ""
                                    for j, k in enumerate(auction_data):
                                        if j > 0:
                                            str_2 += k + " "
                                    temp['auction_name'] = "<span>" + str_1 + "</span><br>" + str_2
                                    temp['row'] = i['auction_name'].lower()

                                if i['auction_name'].lower() == 'classic online auction':
                                    # temp['img_path'] = 'theme-4/images/rethink-icon-1.png'
                                    temp['img_path'] = 'theme-4/images/classic-icon.svg'
                                    temp['text'] = 'An English auction, also known as an ascending-bid auction, is a type of auction ...'
                                elif i['auction_name'].lower() == 'live event auction':
                                    # temp['img_path'] = 'theme-4/images/rethink-icon-4.png'
                                    temp['img_path'] = 'theme-4/images/live-icon.svg'
                                    temp['text'] = 'A live auction is an event where items or properties are presented for sale to the ...'
                                elif i['auction_name'].lower() == 'insider online auction':
                                    # temp['img_path'] = 'theme-4/images/rethink-icon-1.png'
                                    temp['img_path'] = 'theme-4/images/insider-icon.svg'
                                    temp['text'] = 'A hybrid auction is a type of auction that combines both traditional auction methods ...'
                                elif i['auction_name'].lower() == 'multiple parcel online auction':
                                    # temp['img_path'] = 'theme-4/images/rethink-icon-3.png'
                                    temp['img_path'] = 'theme-4/images/multiple-icon.svg'
                                    temp['text'] = 'Multiple auctions refer to the occurrence of several auctions taking place ...'
                                elif i['auction_name'].lower() == 'highest and best offer':
                                    # temp['img_path'] = 'theme-4/images/rethink-icon-1.png'
                                    temp['img_path'] = 'theme-4/images/highest-icon.svg'
                                    temp['text'] = 'A direct offer in the context of a real estate website refers to a feature or option ...'
                                elif i['auction_name'].lower() == 'traditional listing':
                                    # temp['img_path'] = 'theme-4/images/rethink-icon-2.png'
                                    temp['img_path'] = 'theme-4/images/traditinoal-icon.svg'
                                    temp['text'] = 'A traditional auction is a method of buying and selling goods or services through ...'
                                auction_name.append(temp)
                    else:
                        for i in range(6):
                            temp = {}
                            if i == 0:
                                temp['auction_name'] = "<span>Classic</span><br>Online Auction "
                                temp['row'] = 'classic online auction'
                                # temp['img_path'] = 'theme-4/images/rethink-icon-1.png'
                                temp['img_path'] = 'theme-4/images/classic-icon.svg'
                                temp['text'] = 'An English auction, also known as an ascending-bid auction, is a type of auction ...'
                            elif i == 1:
                                temp['auction_name'] = "<span>Live</span><br>Event Auction"
                                temp['row'] = 'live event auction'
                                # temp['img_path'] = 'theme-4/images/rethink-icon-4.png'
                                temp['img_path'] = 'theme-4/images/live-icon.svg'
                                temp['text'] = 'A live auction is an event where items or properties are presented for sale to the ...'
                            elif i == 2:
                                temp['auction_name'] = "<span>Insider</span><br>Online Auction"
                                temp['row'] = 'insider online auction'
                                # temp['img_path'] = 'theme-4/images/rethink-icon-1.png'
                                temp['img_path'] = 'theme-4/images/insider-icon.svg'
                                temp['text'] = 'A hybrid auction is a type of auction that combines both traditional auction methods ...'
                            elif i == 3:
                                temp['auction_name'] = "<span>Multiple Parcel</span><br>Online Auction"
                                temp['row'] = 'multiple parcel online auction'
                                # temp['img_path'] = 'theme-4/images/rethink-icon-3.png'
                                temp['img_path'] = 'theme-4/images/multiple-icon.svg'
                                temp['text'] = 'Multiple auctions refer to the occurrence of several auctions taking place ...'
                            elif i == 4:
                                temp['auction_name'] = "<span>Highest and</span><br>Best Offer"
                                temp['row'] = 'highest and best offer'
                                # temp['img_path'] = 'theme-4/images/rethink-icon-1.png'
                                temp['img_path'] = 'theme-4/images/highest-icon.svg'
                                temp['text'] = 'A direct offer in the context of a real estate website refers to a feature or option ...'
                            elif i == 5:
                                temp['auction_name'] = "<span>Traditional</span><br>Listing"
                                temp['row'] = 'traditional listing'
                                # temp['img_path'] = 'theme-4/images/rethink-icon-2.png'
                                temp['img_path'] = 'theme-4/images/traditinoal-icon.svg'
                                temp['text'] = 'A traditional auction is a method of buying and selling goods or services through ...'
                            auction_name.append(temp)

                    if dashboard_data['active_expertise'] > 0:
                        for i in dashboard_data['expertise']:
                            if i['status'] == 1:
                                temp = {}
                                temp['expertise_name'] = i['expertise_name']
                                # if i['expertise_name'] == "Residential":
                                #     temp['img_path'] = 'theme-4/images/exp-2.jpg'
                                #     temp['text'] = 'Find affordable listed residential properties. Choose the best-suited residential property effortlessly.'
                                # elif i['expertise_name'] == "Condo":
                                #     temp['img_path'] = 'theme-4/images/exp-3.jpg'
                                #     temp['text'] = 'Looking for a condo property? Find the best location and all the amenities with fare rules for your condo property.'
                                # elif i['expertise_name'] == "Commercial":
                                #     temp['img_path'] = 'theme-4/images/exp-6.jpg'
                                #     temp['text'] = 'Unlock your potential income by investing in commercial properties. Purchase property with lease agreements and the assistance of legal professionals.'
                                # elif i['expertise_name'] == "Multifamily":
                                #     temp['img_path'] = 'theme-4/images/exp-4.jpg'
                                #     temp['text'] = 'Find the best multifamily properties such as apartment buildings or multi-unit complexes under one roof. Filter properties based on reference.'
                                if i['expertise_name'] == "House":
                                    temp['img_path'] = 'theme-4/images/exp-2.jpg'
                                    temp['text'] = 'Find affordable listed residential properties. Choose the best-suited residential property effortlessly.'
                                elif i['expertise_name'] == "Land":
                                    temp['img_path'] = 'theme-4/images/exp-4.jpg'
                                    temp['text'] = 'Find the best lands & plots for sale. Plots are available with better value than other property options. It is a great money-saving investment that will guarantee higher returns in the future.'
                                elif i['expertise_name'] == "Commercial":
                                    temp['img_path'] = 'theme-4/images/exp-6.jpg'
                                    temp['text'] = 'Unlock your potential income by investing in commercial properties. Purchase property with lease agreements and the assistance of legal professionals.'
                                elif i['expertise_name'] == "Auction":
                                    temp['img_path'] = 'theme-4/images/exp-3.jpg'
                                    temp['text'] = 'Find the latest listing of properties for auction. Buyers can easily search and view details of properties and participate in the auction process.'
                                expertise_name.append(temp)
                    else:
                        for i in range(4):
                            temp = {}
                            if i == 0:
                                temp['expertise_name'] = 'House'
                                temp['img_path'] = 'theme-4/images/exp-2.jpg'
                                temp['text'] = 'Find affordable listed residential properties. Choose the best-suited residential property effortlessly.'
                            elif i == 1:
                                temp['expertise_name'] = 'Land'
                                temp['img_path'] = 'theme-4/images/exp-4.jpg'
                                temp['text'] = 'Find the best lands & plots for sale. Plots are available with better value than other property options. It is a great money-saving investment that will guarantee higher returns in the future.'
                            elif i == 2:
                                temp['expertise_name'] = 'Commercial'
                                temp['img_path'] = 'theme-4/images/exp-6.jpg'
                                temp['text'] = 'Unlock your potential income by investing in commercial properties. Purchase property with lease agreements and the assistance of legal professionals.'
                            elif i == 3:
                                temp['expertise_name'] = 'Auction'
                                temp['img_path'] = 'theme-4/images/exp-3.jpg'
                                temp['text'] = 'Find the latest listing of properties for auction. Buyers can easily search and view details of properties and participate in the auction process.'
                            expertise_name.append(temp)

            except Exception as exp:
                print(exp)
                dashboard_data = {}
                social_account_list = []
                auction_name = []

        except Exception as exp:
            print(exp)
        pre_auction_type_list = []
        for i in range(6):
            type_name = ''
            if i == 0:
                type_name = 'Classic Online Auction'
                image_path = 'static/'+static_dir+'/images/real-1.jpg'
            elif i == 1:
                type_name = 'Live Event Auction'
                image_path = 'static/'+static_dir+'/images/real-4.jpg'
            elif i == 2:
                type_name = 'Insider Online Auction'
                image_path = 'static/'+static_dir+'/images/real-3.jpg'
            elif i == 3:
                type_name = 'Multiple Parcel Online Auction'
                image_path = 'static/'+static_dir+'/images/real-4.jpg'
            elif i == 4:
                type_name = 'Highest and Best Offer'
                image_path = 'static/'+static_dir+'/images/real-2.jpg'
            elif i == 5:
                type_name = 'Traditional Listing'
                image_path = 'static/'+static_dir+'/images/real-4.jpg'

            auction_params = {
                'type_name': type_name,
                'image': image_path,
            }

            pre_auction_type_list.append(auction_params)

        pre_expertise_list = []
        for i in range(4):
            type_name = ''
            if i == 0:
                type_name = 'House'
                image_path = 'static/'+static_dir+'/images/experties-1.jpg'
                icon_class = 'icon resiicon-1'
            elif i == 1:
                type_name = 'Land'
                image_path = 'static/'+static_dir+'/images/experties-2.jpg'
                icon_class = ' icon land-1'
            elif i == 2:
                type_name = 'Commercial'
                image_path = 'static/'+static_dir+'/images/experties-3.jpg'
                icon_class = 'icon commercial-1'
            elif i == 3:
                type_name = 'Auction'
                image_path = 'static/'+static_dir+'/images/experties-4.jpg'
                icon_class = 'icon auction-1'

            expertise_params = {
                'type_name': type_name,
                'image': image_path,
                'icon_class': icon_class,
            }

            pre_expertise_list.append(expertise_params)

        pre_number_list = []
        for i in range(4):
            type_name = ''
            value = 0
            if i == 0:
                type_name = 'Best Downtown Locations'
                value = 10
            elif i == 1:
                type_name = 'Exceptional Properties'
                value = 26
            elif i == 2:
                type_name = 'Satisfied Buyer-Seller'
                value = 200
            elif i == 3:
                type_name = 'Email Subscriber'
                value = '+20K'

            pre_num_params = {
                'title': type_name,
                'value': value
            }

            pre_number_list.append(pre_num_params)

        pre_social_account_list = []
        if len(social_account_list) > 0:
            if templete_dir == "theme-4":
                for item in social_account_list:
                    title = ''
                    icon = ''
                    url = ''
                    if item['status'] == 1 and item['url'] != "":
                        if item['account_type'] == 1:
                            title = 'facebook'
                            icon = 'theme-4/images/facebook-icon.png'
                        elif item['account_type'] == 2:
                            title = 'twitter'
                            icon = 'theme-4/images/twitter-icon.png'
                        elif item['account_type'] == 3:
                            title = 'youtube'
                            icon = 'theme-4/images/youtube-icon.png'
                        elif item['account_type'] == 4:
                            title = 'linkedin'
                            icon = 'theme-4/images/linkedin-icon.png'
                        elif item['account_type'] == 5:
                            title = 'instagram'
                            icon = 'theme-4/images/instagram-icon.png'

                        pre_social_params = {
                            'title': title,
                            'icon': icon,
                            'account_type': item['account_type'],
                            'url': item['url'],
                        }

                        pre_social_account_list.append(pre_social_params)
            else:
                for item in social_account_list:
                    title = ''
                    icon = ''
                    url = ''
                    if item['status'] == 1 and item['url'] != "":
                        if item['account_type'] == 1:
                            title = 'facebook'
                            icon = 'fab fa-facebook-f'
                        elif item['account_type'] == 2:
                            title = 'twitter'
                            icon = 'fab fa-twitter'
                        elif item['account_type'] == 3:
                            title = 'youtube'
                            icon = 'fab fa-youtube'
                        elif item['account_type'] == 4:
                            title = 'linkedin'
                            icon = 'fab fa-linkedin-in'
                        elif item['account_type'] == 5:
                            title = 'instagram'
                            icon = 'fab fa-instagram'

                        pre_social_params = {
                            'title': title,
                            'icon': icon,
                            'account_type': item['account_type'],
                            'url': item['url'],
                        }

                        pre_social_account_list.append(pre_social_params)
        context = {"data": "COMING SOON...", "dashboard_data": dashboard_data, 'pre_auction_type_list': pre_auction_type_list, 'pre_expertise_list': pre_expertise_list, 'pre_number_list': pre_number_list, 'pre_social_account_list': pre_social_account_list, 'is_home_page': True, 'site_id': site_id, 'node_url': settings.NODE_URL, 'auction_name': auction_name, 'expertise_name': expertise_name, 'is_first_login': is_first_login}
        # print(context)
        # settings_value = settings_data(token, user_id)
        # print(context)
        theme_path = 'home/{}/dashboard/dashboard.html'.format(templete_dir)
        return render(request, theme_path, context)
    except Exception as exp:
        print(exp)
        return HttpResponse("Issue in views")


def login(request):
    """This function is used to login
    subdomain users
    """
    try:
        next = ""
        # check if any redirection requested

        if request.method == "GET":  
            next = request.get_full_path().rsplit('?next=', 1)[-1] if 'next' in request.GET else ''

        if 'user_id' in request.session and request.session['user_id']:
            return redirect('/')
        try:
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']
            templete_dir = site_detail['site_detail']['theme_directory']
        except Exception as exp:
            templete_dir = 'theme-1'
            site_id = ''

        if (request.is_ajax() and request.method == 'POST') and ("user_id" not in request.session or request.session['user_id'] <= 0):
            email = request.POST['email']
            password = request.POST['password']
            api_url = settings.API_URL + '/api-users/subdomain-login/'
            payload = {
                'domain_id': site_id,
                'email': email,
                'password': password
            }
            try:
                response = call_api_post_method(payload, api_url)
                if "error" in response and response['error'] == 0:
                    response_data = response['data']
                    expires_in = int(response_data['auth_token']['expires_in'])
                    expiry_date_time = datetime.datetime.now() + timedelta(seconds=expires_in)
                    expiry_time = expiry_date_time.timestamp() * 1000
                    request.session['token_expiry_time'] = expiry_time
                    request.session['user_id'] = response_data['user_id']
                    request.session['token'] = response_data['auth_token']
                    request.session['first_name'] = response_data['first_name']
                    request.session['user_type'] = response_data['user_type']
                    request.session['is_admin'] = response_data['is_admin']
                    request.session['is_broker'] = response_data['is_broker']
                    request.session['profile_image'] = response_data['profile_image']
                    request.session['user_type_name'] = response_data['user_type_name']
                    request.session['site_id'] = response_data['site_id']
                    request.session['email'] = response_data['email']
                    # request.session['is_first_login'] = response_data['is_first_login']
                    # request.session['is_first_admin_login'] = response_data['is_first_login']
                    request.session['is_first_login'] = 0
                    request.session['is_first_admin_login'] = 0
                    request.session['is_free_plan'] = response_data['is_free_plan']
                    # append next link if any
                    response['next'] = request.POST['next'] if 'next' in request.POST and request.POST['next'] else ''
                    if not response_data['stripe_customer_id'] and response_data['customer_site_id'] == response_data['site_id']:
                        # request.session['demo_tour'] = 0
                        # response['next'] = "/demo/"
                        request.session['demo_tour'] = 1
                        response['next'] = "/admin/dashboard/"
                    elif response_data['is_free_plan'] and response_data['is_broker']:
                        # ---------------------Save Payment Detail Data------------------
                        # plan_price_id = int(site_detail['site_detail']['current_plan_price_id'])
                        plan_price_id = int(site_detail['site_detail']['previous_plan_price_id']) if site_detail['site_detail']['previous_plan_price_id'] else int(site_detail['site_detail']['current_plan_price_id'])
                        theme_id = int(site_detail['site_detail']['current_theme_id'])
                        api_url = settings.API_URL + "/api-payments/create-payment-detail/"
                        params = {
                            "domain_id": site_id,
                            "user_id": response_data['user_id'],
                            "plan_price_id": plan_price_id,
                            "theme_id": theme_id
                        }
                        token = request.session['token']['access_token']
                        api_response = call_api_post_method(params, api_url, token)
                        if 'error' in api_response and api_response['error'] == 0:
                            request.session['auto_payment_redirection'] = 1
                        # response['next'] = "/admin/"
                        response['next'] = "/admin/pay-now/"
                    elif response_data['is_admin']:
                        response['next'] = "/admin/dashboard/"
                return JsonResponse(response)

            except Exception as exp:
                print(exp)
                data = {'status': 403, 'msg': 'invalid request.'}
                return JsonResponse(data)

        http_host = request.META['HTTP_HOST']
        site_url = settings.URL_SCHEME + str(http_host) 
        context = {
            'is_home_page': False,
            'aws_url': settings.AWS_URL,
            'SITE_URL': site_url,
            'hide_header': True,
            'hide_footer': True,
            'next': next
        }
        theme_path = 'home/{}/login.html'.format(templete_dir)
        return render(request, theme_path, context)
    except Exception as exp:
        print(exp)
        return HttpResponse("Issue in views")


def set_login(request):
    try:
        token = request.GET.get('token', None)
        try:
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']
        except:
            site_id = ""
        if token is not None and 'user_id' not in request.session:
            try:
                site_detail = subdomain_site_details(request)
                site_id = site_detail['site_detail']['site_id']
            except:
                site_id = ""

            api_url = settings.API_URL + '/api-users/login-details/'
            payload = {
                'token': token,
                'site_id': site_id
            }
            response = call_api_post_method(payload, api_url)
            if "error" in response and response['error'] == 0:
                try:
                    response = response['data']
                    expires_in = int(response['auth_token']['expires_in'])
                    expiry_date_time = datetime.datetime.now() + timedelta(seconds=expires_in)
                    expiry_time = expiry_date_time.timestamp() * 1000
                    request.session['token_expiry_time'] = expiry_time
                    request.session['user_id'] = response['user_id']
                    request.session['token'] = response['auth_token']
                    request.session['first_name'] = response['first_name']
                    request.session['user_type'] = response['user_type']
                    request.session['is_admin'] = response['is_admin']
                    request.session['is_broker'] = response['is_broker']
                    request.session['profile_image'] = response['profile_image']
                    request.session['user_type_name'] = response['user_type_name']
                    request.session['site_id'] = response['site_id']
                    request.session['email'] = response['email']
                    # request.session['is_first_login'] = response['is_first_login']
                    # request.session['is_first_admin_login'] = response['is_first_login']
                    request.session['is_first_login'] = 0
                    request.session['is_first_admin_login'] = 0
                    request.session['is_free_plan'] = response['is_free_plan']
                    if not response['stripe_customer_id'] and response['customer_site_id'] == response['site_id']:
                        # request.session['demo_tour'] = 0
                        # return HttpResponseRedirect('/demo/')
                        request.session['demo_tour'] = 1
                        return HttpResponseRedirect('/admin/dashboard/')
                    elif response['is_free_plan'] and response['is_broker']:
                        # ---------------------Save Payment Detail Data------------------
                        # plan_price_id = int(site_detail['site_detail']['current_plan_price_id'])
                        plan_price_id = int(site_detail['site_detail']['previous_plan_price_id']) if site_detail['site_detail']['previous_plan_price_id'] else int(site_detail['site_detail']['current_plan_price_id'])
                        theme_id = int(site_detail['site_detail']['current_theme_id'])
                        api_url = settings.API_URL + "/api-payments/create-payment-detail/"
                        params = {
                            "domain_id": site_id,
                            "user_id": response['user_id'],
                            "plan_price_id": plan_price_id,
                            "theme_id": theme_id
                        }
                        token = request.session['token']['access_token']
                        api_response = call_api_post_method(params, api_url, token)
                        if 'error' in api_response and api_response['error'] == 0:
                            request.session['auto_payment_redirection'] = 1

                        # return HttpResponseRedirect('/admin/')
                        return HttpResponseRedirect('/admin/pay-now/')
                    elif response['is_admin']:
                        return HttpResponseRedirect('/admin/dashboard/')
                    return HttpResponseRedirect('/')
                except Exception as exp:
                    print(exp)
                    return HttpResponseRedirect('/')
        else:
            return HttpResponseRedirect('/')
    except Exception as exp:
        return HttpResponse("Issue in views")


def register(request):
    """This function is used to show registration page
    for subdomain users
    """
    try:
        if 'user_id' in request.session and request.session['user_id']:
            return redirect('/')
        try:
            site_detail = subdomain_site_details(request)
            templete_dir = site_detail['site_detail']['theme_directory']
            site_id = int(site_detail['site_detail']['site_id'])
        except Exception as exp:
            templete_dir = 'theme-1'
            site_id = None

        http_host = request.META['HTTP_HOST']
        site_url = settings.URL_SCHEME + str(http_host) 
        captcha_site_key = settings.CAPTCHA_SITE_KEY
        context = {
            'is_home_page': False,
            'aws_url': settings.AWS_URL,
            'SITE_URL': site_url,
            'hide_header': True,
            'hide_footer': True,
            'captcha_site_key': captcha_site_key,
            'site_id': site_id
        }
        theme_path = 'home/{}/register.html'.format(templete_dir)
        return render(request, theme_path, context)
    except Exception as exp:
        print(exp)
        return HttpResponse("Issue in views")


def save_user(request):
    """
        Method : Post
        Description : Register User From Subdomain
        Url : /save-user/
        Params:
        :param 1: user_id :: integer
        :param 2: first_name :: String
        :param 3: last_name :: String
        :param 4: email :: String
        :param 5: password :: String
        :param 6: phone :: String
        :param 7: agree_term :: Boolean(1/0)
        :param 8: brokerage_name :: String
        :param 9: licence_number :: String
        :param 10: described_by :: integer
    """
    try:
        if 'user_id' in request.session and request.session['user_id']:
            return redirect('/')
        if request.is_ajax() and request.method == 'POST':
            reg_success_html = ''
            try:
                site_detail = subdomain_site_details(request)
                site_id = site_detail['site_detail']['site_id']
                templete_dir = site_detail['site_detail']['theme_directory']
            except Exception as exp:
                site_id = ""

            params = {
                'domain_id': site_id,
                'described_by': request.POST['described_by'],
                'first_name': request.POST['first_name'].strip(),
                'last_name': request.POST['last_name'].strip(),
                'email': request.POST['email'],
                'password': request.POST['password'],
                'phone_no': request.POST['phone'],
                'agree_term': request.POST['agree_term'],
                'brokerage_name': request.POST['brokerage_name'].strip() if 'brokerage_name' in request.POST and request.POST['brokerage_name'] else '',
                'licence_number': request.POST['licence_number'] if 'licence_number' in request.POST and request.POST['licence_number'] else ''
            }

            recaptcha_response = request.POST.get('g-recaptcha-response')
            result = check_recaptcha(recaptcha_response)

            if result['success']:
                url = settings.API_URL + '/api-users/subdomain-registration/'
                data = call_api_post_method(params, url)
                if "error" in data and data['error'] == 0:
                    reg_succes_path = 'home/{}/registration-successful.html'.format(templete_dir)
                    reg_success_template = get_template(reg_succes_path)
                    reg_success_html = reg_success_template.render({'email': request.POST['email'], 'admin_email':data['data']['admin_email'] })
                    data['reg_success_html'] = reg_success_html
            else:
                data = {'status': 403, 'msg': 'Invalid reCAPTCHA, Please try again'}

        else:
            data = {'status': 403, 'msg': 'Forbidden'}
        return JsonResponse(data)
    except Exception as exp:
        print(exp)
        data = {'status': 403, 'msg': 'invalid request.'}
        return JsonResponse(data)


def activation(request):
    try:
        activation_code = request.GET.get('token', None)
        if activation_code is None:
            return redirect('/')
        
        try:
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']
        except:
            site_id = ""

        api_url = settings.API_URL + '/api-users/subdomain-registration-verification/'
        payload = {
            'domain_id' : site_id,
            'activation_code': activation_code,
        }
        data = call_api_post_method(payload, api_url)
        if "error" in data and data['error'] == 0:

            api_url = settings.API_URL + '/api-users/login-details/'
            payload = {
                'token': data['data']['auth_token']['access_token'],
                'site_id': site_id
            }
            response = call_api_post_method(payload, api_url)
            if "error" in response and response['error'] == 0:
                try:
                    response = response['data']
                    expires_in = int(response['auth_token']['expires_in'])
                    expiry_date_time = datetime.datetime.now() + timedelta(seconds=expires_in)
                    expiry_time = expiry_date_time.timestamp() * 1000
                    request.session['token_expiry_time'] = expiry_time
                    request.session['user_id'] = response['user_id']
                    request.session['token'] = response['auth_token']
                    request.session['first_name'] = response['first_name']
                    request.session['user_type'] = response['user_type']
                    request.session['is_admin'] = response['is_admin']
                    request.session['is_broker'] = response['is_broker']
                    request.session['profile_image'] = response['profile_image']
                    request.session['edit_profile_modal'] = True
                    
                    return redirect('/edit-profile')
                except Exception as exp:
                    print(exp)
                    return HttpResponseRedirect('/')

        return redirect('/') 
    except Exception as exp:
        print(exp)
        return redirect('/')



def forgot_password(request):
    """[summary]

    Args:
        request ([type]): [description]
    """
    try:
        if 'user_id' in request.session and request.session['user_id']:
            return redirect('/')
        try:
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']
            templete_dir = site_detail['site_detail']['theme_directory']
        except Exception as exp:
            site_id = ""
            templete_dir = 'theme-1'

        if request.is_ajax() and request.method == 'POST':
            params = {
                'email': request.POST['email'],
                'domain_id': site_id
            }
            url = settings.API_URL + '/api-users/subdomain-forgot-password/'
            data = call_api_post_method(params, url)
            return JsonResponse(data)

        http_host = request.META['HTTP_HOST']
        site_url = settings.URL_SCHEME + str(http_host) 
        context = {
            'is_home_page': False,
            'aws_url': settings.AWS_URL,
            'SITE_URL': site_url,
            'hide_header': True,
            'hide_footer': True
        }
        theme_path = 'home/{}/forgot-password.html'.format(templete_dir)
        return render(request, theme_path, context)
    except Exception as exp:
        print(exp)
        return HttpResponse("Issue in views") 


@csrf_exempt
def reset_password(request):
    try:
        try:
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']
            templete_dir = site_detail['site_detail']['theme_directory']
        except Exception as exp:
            site_id = ""
            templete_dir = 'theme-1'

        reset_pass_token = request.GET.get('token', None)
        if not reset_pass_token and not request.is_ajax():
            return HttpResponseRedirect(settings.BASE_URL)

        if request.is_ajax() and request.method == 'POST':
            params = {
                'password': request.POST['password'],
                'reset_token': request.POST['reset_pass_token'],
                'domain_id': site_id
            }
            url = settings.API_URL + '/api-users/subdomain-reset-password/'
            data = call_api_post_method(params, url)
            return JsonResponse(data)


        http_host = request.META['HTTP_HOST']
        site_url = settings.URL_SCHEME + str(http_host) 
        context = {
            'is_home_page': False,
            'aws_url': settings.AWS_URL,
            'SITE_URL': site_url,
            'hide_header': True,
            'hide_footer': True
        }
        theme_path = 'home/{}/reset-password.html'.format(templete_dir)
        return render(request, theme_path, context)
    except Exception as exp:
        print(exp)
        return redirect('/')


@csrf_exempt
def check_user_exists(request):
    """This function is used to check
         users existance
        """
    try:
        if request.is_ajax() and request.method == 'POST':
            if 'email' in request.POST and request.POST['email'] != "":
                email = request.POST['email']
            elif 'business_email' in request.POST and request.POST['business_email'] != "":
                email = request.POST['business_email']
            else:
                email = ""
            if 'phone' in request.POST and request.POST['phone'] != "":
                phone = request.POST['phone']
            elif 'business_phone' in request.POST and request.POST['business_phone'] != "":
                phone = request.POST['business_phone']
            else:
                phone = ""
            params = {
                'email': email,
                'phone': phone,
                'check_type': request.POST['check_type']
            }
            url = settings.API_URL + '/api-users/check-user-exists/'
            data = call_api_post_method(params, url)
        else:
            data = {'status': 403, 'msg': 'Forbidden'}
        return JsonResponse(data)
    except Exception as exp:
        print(exp)
        data = {'status': 403, 'msg': 'invalid request.'}
        return JsonResponse(data)

# def login(request):
#     try:
#         token = request.GET.get('token', None)
#         try:
#             site_detail = subdomain_site_details(request)
#             site_id = site_detail['site_detail']['site_id']
#         except:
#             site_id = ""
#         if token is not None  and 'user_id' not in request.session:
#             try:
#                 site_detail = subdomain_site_details(request)
#                 site_id = site_detail['site_detail']['site_id']
#             except:
#                 site_id = ""

#             api_url = settings.API_URL + '/api-users/login-details/'
#             payload = {
#                 'token': token,
#                 'site_id': site_id
#             }
#             response = call_api_post_method(payload, api_url)
#             if "error" in response and response['error'] == 0:
#                 try:
#                     response = response['data']
#                     expires_in = int(response['auth_token']['expires_in'])
#                     expiry_date_time = datetime.datetime.now() + timedelta(seconds=expires_in)
#                     expiry_time = expiry_date_time.timestamp() * 1000
#                     request.session['token_expiry_time'] = expiry_time
#                     request.session['user_id'] = response['user_id']
#                     # request.session['site_id'] = response['site_id']
#                     request.session['token'] = response['auth_token']
#                     request.session['first_name'] = response['first_name']
#                     request.session['user_type'] = response['user_type']
#                     request.session['is_admin'] = response['is_admin']
#                     request.session['is_broker'] = response['is_broker']
#                     request.session['profile_image'] = response['profile_image']
#                     return HttpResponseRedirect('/')
#                 except Exception as exp:
#                     print(exp)
#                     return HttpResponseRedirect('/')
#     except Exception as exp:
#         return HttpResponse("Issue in views")


def logout(request):
    try:
        request.session['is_first_login'] = 0
        api_url = settings.API_URL + '/api-users/revoke-token/'
        payload = {
            'user_id': request.session['user_id'],
            'token': request.session['token']['access_token']
        }
        call_api_post_method(payload, api_url, request.session['token']['access_token'])
        del request.session['user_id']
        del request.session['profile_image']
        del request.session['token']
        del request.session['first_name']
        del request.session['token_expiry_time']
        del request.session['edit_profile_modal']
        request.session.modified = True
        request.session.flush()
        http_host = request.META['HTTP_HOST']
        redirect_url = settings.URL_SCHEME + str(http_host)
        return HttpResponseRedirect(redirect_url)
    except Exception as exp:
        http_host = request.META['HTTP_HOST']
        redirect_url = settings.URL_SCHEME + str(http_host)
        return HttpResponseRedirect(redirect_url)


def profile(request):
    try:
        return HttpResponse("Profile Page")
    except Exception as exp:
        return HttpResponse("Issue in views")

def asset_details(request):
    """
        Method : Get
        Description : Get property details
        Url : /asset-details/
        Params:
        :param 1: property_id :: integer
    """
    try:
        property_id = request.GET.get('property_id', None)
        if not property_id:
            http_host = request.META['HTTP_HOST']
            redirect_url = settings.URL_SCHEME + str(http_host)
            return HttpResponseRedirect(redirect_url)
        # try:
        #     token = request.session['token']['access_token']
        #     user_id = request.session['user_id']
        # except Exception as exp:
        #     print(exp)
        try:
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']
            templete_dir = site_detail['site_detail']['theme_directory']

        except Exception as exp:
            print(exp)
            site_id = ""
            templete_dir = 'theme-1'
        static_dir = templete_dir

        user_id = ''
        token = None
        if 'user_id' in request.session and request.session['user_id']:
            user_id =  request.session['user_id']
            token = request.session['token']['access_token']
        else:
            user_id = ''

        try:
            state_param = {}
            state_api_url = settings.API_URL + '/api-settings/get-state/'
            state_data = call_api_post_method(state_param, state_api_url, token)
            state_list = state_data['data']
        except:
            state_list = []


        try:
            asset_details_params = {'site_id': site_id, 'property_id': property_id, 'user_id': user_id}

            asset_details_url = settings.API_URL + '/api-property/property-detail/'
            asset_details_data = call_api_post_method(asset_details_params, asset_details_url)

            if 'error' in asset_details_data and asset_details_data['error'] == 1:
                http_host = request.META['HTTP_HOST']
                redirect_url = settings.URL_SCHEME + str(http_host)
                return HttpResponseRedirect(redirect_url)

            asset_details = asset_details_data['data']

            property_sub_types = ''
            property_address = ''
            term_accepted_name = ''
            if 'property_subtype' in asset_details and len(asset_details['property_subtype']) > 0:
                for pst in asset_details['property_subtype']:
                    property_sub_types = property_sub_types+', '+pst['feature_name']

            if 'address_one' in asset_details and asset_details['address_one'] != "":
                property_address = asset_details['address_one']

            if 'city' in asset_details and asset_details['city'] != "":
                property_address = property_address+', '+str(asset_details['city'])

            if 'state' in asset_details and asset_details['state'] != "":
                property_address = property_address+', '+str(asset_details['state'])

            if 'postal_code' in asset_details and asset_details['postal_code'] != "":
                property_address = property_address+', '+str(asset_details['postal_code'])


            current_bid_amount = 0

            if 'property_current_price' in asset_details and asset_details['property_current_price']:
                current_bid_amount = current_bid_amount + float(asset_details['property_current_price'])


            earnest_money_deposit = ''
            earnest_money_deposit_percent = ''
            custom_start_price = 0

            if asset_details['property_auction_data'] and 'start_price' in asset_details['property_auction_data'][
                0] and asset_details['property_auction_data'][0]['start_price']:
                custom_start_price = asset_details['property_auction_data'][0]['start_price']


            try:
                if 'sale_by_type_id' in asset_details and int(asset_details['sale_by_type_id']) == 7 and custom_start_price:

                    if 'earnest_deposit_type' in asset_details and asset_details['earnest_deposit_type'] == 1:
                        earnest_money_deposit = asset_details['earnest_deposit']
                        earnest_money_deposit = format_currency(earnest_money_deposit)
                        earnest_money_deposit_percent = (float(earnest_money_deposit)*100)/float(custom_start_price)
                        earnest_money_deposit_percent = format_currency(earnest_money_deposit_percent)
                    elif 'earnest_deposit_type' in asset_details and asset_details['earnest_deposit_type'] == 2:
                        earnest_money_deposit_percent = asset_details['earnest_deposit']
                        earnest_money_deposit_percent = format_currency(earnest_money_deposit_percent)
                        earnest_money_deposit = (float(custom_start_price)*float(asset_details['earnest_deposit']))/float(100)
                        earnest_money_deposit = format_currency(earnest_money_deposit)
                    else:
                        earnest_money_deposit = ''
                        earnest_money_deposit_percent = ''
                else:
                    earnest_money_deposit = ''
                    earnest_money_deposit_percent = ''
            except Exception as exp:
                print(exp)
                earnest_money_deposit = ''
                earnest_money_deposit_percent = ''

            

            asset_details['earnest_money_deposit'] = earnest_money_deposit
            asset_details['earnest_money_deposit_percent'] = earnest_money_deposit_percent
            next_bid_amount = float(current_bid_amount)
            bide_increments = 0
            if asset_details['property_auction_data'] and 'bid_increments' in asset_details['property_auction_data'][0] and asset_details['property_auction_data'][0]['bid_increments']:
                bide_increments = asset_details['property_auction_data'][0]['bid_increments']

            if bide_increments and asset_details['bid_count'] > 0:
                next_bid_amount = float(next_bid_amount) + float(bide_increments)


            if 'latitude' not in asset_details or not asset_details['latitude'] or 'longitude' not in asset_details or not asset_details['longitude']:

                try:
                    geolocator = Nominatim(user_agent=settings.GEOLOCATOR_EMAIL, timeout=10)
                    location = geolocator.geocode(property_address)
                    try:
                        asset_details['latitude'] = location.latitude
                        asset_details['longitude'] = location.longitude
                    except:
                        asset_details['latitude'] = 0
                        asset_details['longitude'] = 0
                except:
                    geolocator = Nominatim(user_agent=settings.GEOLOCATOR_EMAIL, timeout=10)
                    location = geolocator.geocode("47 W 13th St, New York, NY 10011, USA")
                    try:
                        asset_details['latitude'] = location.latitude
                        asset_details['longitude'] = location.longitude
                    except:
                        asset_details['latitude'] = 0
                        asset_details['longitude'] = 0


            if 'terms_accepted' in asset_details and len(asset_details['terms_accepted']) > 0:
                for pst in asset_details['terms_accepted']:
                    term_accepted_name = term_accepted_name+', '+str(pst['feature_name'])

            occupied_by_name = ''
            if 'occupied_by' in asset_details and len(asset_details['occupied_by']) > 0:
                for pst in asset_details['occupied_by']:
                    occupied_by_name = occupied_by_name+', '+str(pst['feature_name'])

            ownership_name = ''
            if 'ownership' in asset_details and len(asset_details['ownership']) > 0:
                for pst in asset_details['ownership']:
                    ownership_name = ownership_name+', '+str(pst['feature_name'])

            possession_name = ''
            if 'possession' in asset_details and len(asset_details['possession']) > 0:
                for pst in asset_details['possession']:
                    possession_name = possession_name + ', ' + str(pst['feature_name'])

            style_name = ''
            if 'style' in asset_details and len(asset_details['style']) > 0:
                for pst in asset_details['style']:
                    style_name = style_name + ', ' + str(pst['feature_name'])

            stories_name = ''
            if 'stories' in asset_details and len(asset_details['stories']) > 0:
                for pst in asset_details['stories']:
                    stories_name = stories_name + ', ' + str(pst['feature_name'])

            recent_updates_name = ''
            if 'recent_updates' in asset_details and len(asset_details['recent_updates']) > 0:
                for pst in asset_details['recent_updates']:
                    recent_updates_name = recent_updates_name + ', ' + str(pst['feature_name'])

            security_features_name = ''
            if 'security_features' in asset_details and len(asset_details['security_features']) > 0:
                for pst in asset_details['security_features']:
                    security_features_name = security_features_name + ', ' + str(pst['feature_name'])

            cooling_name = ''
            if 'cooling' in asset_details and len(asset_details['cooling']) > 0:
                for pst in asset_details['cooling']:
                    cooling_name = cooling_name + ', ' + str(pst['feature_name'])

            heating_name = ''
            if 'cooling' in asset_details and len(asset_details['heating']) > 0:
                for pst in asset_details['heating']:
                    heating_name = heating_name + ', ' + str(pst['feature_name'])

            electric_name = ''
            if 'electric' in asset_details and len(asset_details['electric']) > 0:
                for pst in asset_details['electric']:
                    electric_name = electric_name + ', ' + str(pst['feature_name'])

            gas_name = ''
            if 'gas' in asset_details and len(asset_details['gas']) > 0:
                for pst in asset_details['gas']:
                    gas_name = gas_name + ', ' + str(pst['feature_name'])

            water_name = ''
            if 'water' in asset_details and len(asset_details['water']) > 0:
                for pst in asset_details['water']:
                    water_name = water_name + ', ' + str(pst['feature_name'])

            sewer_name = ''
            if 'sewer' in asset_details and len(asset_details['sewer']) > 0:
                for pst in asset_details['sewer']:
                    sewer_name = sewer_name + ', ' + str(pst['feature_name'])

            tax_exemptions_name = ''
            if 'tax_exemptions' in asset_details and len(asset_details['tax_exemptions']) > 0:
                for pst in asset_details['tax_exemptions']:
                    tax_exemptions_name = tax_exemptions_name + ', ' + str(pst['feature_name'])


            hoa_amenities_name = ''
            if 'hoa_amenities' in asset_details and len(asset_details['hoa_amenities']) > 0:
                for pst in asset_details['hoa_amenities']:
                    hoa_amenities_name = hoa_amenities_name + ', ' + str(pst['feature_name'])

            zoning_name = ''
            if 'zoning' in asset_details and len(asset_details['zoning']) > 0:
                for pst in asset_details['zoning']:
                    zoning_name = zoning_name + ', ' + str(pst['feature_name'])

            kitchen_features_name = ''
            if 'kitchen_features' in asset_details and len(asset_details['kitchen_features']) > 0:
                for pst in asset_details['kitchen_features']:
                    kitchen_features_name = kitchen_features_name + ', ' + str(pst['feature_name'])

            appliances_name = ''
            if 'appliances' in asset_details and len(asset_details['appliances']) > 0:
                for pst in asset_details['appliances']:
                    appliances_name = appliances_name + ', ' + str(pst['feature_name'])

            flooring_name = ''
            if 'flooring' in asset_details and len(asset_details['flooring']) > 0:
                for pst in asset_details['flooring']:
                    flooring_name = flooring_name + ', ' + str(pst['feature_name'])

            windows_name = ''
            if 'windows' in asset_details and len(asset_details['windows']) > 0:
                for pst in asset_details['windows']:
                    windows_name = windows_name + ', ' + str(pst['feature_name'])

            bedroom_features_name = ''
            if 'bedroom_features' in asset_details and len(asset_details['bedroom_features']) > 0:
                for pst in asset_details['bedroom_features']:
                    bedroom_features_name = bedroom_features_name + ', ' + str(pst['feature_name'])

            other_rooms_name = ''
            if 'other_rooms' in asset_details and len(asset_details['other_rooms']) > 0:
                for pst in asset_details['other_rooms']:
                    other_rooms_name = other_rooms_name + ', ' + str(pst['feature_name'])

            bathroom_features_name = ''
            if 'bathroom_features' in asset_details and len(asset_details['bathroom_features']) > 0:
                for pst in asset_details['bathroom_features']:
                    bathroom_features_name = bathroom_features_name + ', ' + str(pst['feature_name'])

            master_bedroom_features_name = ''
            if 'master_bedroom_features' in asset_details and len(asset_details['master_bedroom_features']) > 0:
                for pst in asset_details['master_bedroom_features']:
                    master_bedroom_features_name = master_bedroom_features_name + ', ' + str(pst['feature_name'])

            basement_features_name = ''
            if 'basement_features' in asset_details and len(asset_details['basement_features']) > 0:
                for pst in asset_details['basement_features']:
                    basement_features_name = basement_features_name + ', ' + str(pst['feature_name'])

            other_features_name = ''
            if 'other_features' in asset_details and len(asset_details['other_features']) > 0:
                for pst in asset_details['other_features']:
                    other_features_name = other_features_name + ', ' + str(pst['feature_name'])

            fireplace_type_name = ''
            if 'fireplace_type' in asset_details and len(asset_details['fireplace_type']) > 0:
                for pst in asset_details['fireplace_type']:
                    fireplace_type_name = fireplace_type_name + ', ' + str(pst['feature_name'])

            handicap_amenities_name = ''
            if 'handicap_amenities' in asset_details and len(asset_details['handicap_amenities']) > 0:
                for pst in asset_details['handicap_amenities']:
                    handicap_amenities_name = handicap_amenities_name + ', ' + str(pst['feature_name'])

            construction_name = ''
            if 'construction' in asset_details and len(asset_details['construction']) > 0:
                for pst in asset_details['construction']:
                    construction_name = construction_name + ', ' + str(pst['feature_name'])

            exterior_features_name = ''
            if 'exterior_features' in asset_details and len(asset_details['exterior_features']) > 0:
                for pst in asset_details['exterior_features']:
                    exterior_features_name = exterior_features_name + ', ' + str(pst['feature_name'])

            roof_name = ''
            if 'roof' in asset_details and len(asset_details['roof']) > 0:
                for pst in asset_details['roof']:
                    roof_name = roof_name + ', ' + str(pst['feature_name'])

            foundation_name = ''
            if 'foundation' in asset_details and len(asset_details['foundation']) > 0:
                for pst in asset_details['foundation']:
                    foundation_name = foundation_name + ', ' + str(pst['feature_name'])

            fence_name = ''
            if 'fence' in asset_details and len(asset_details['fence']) > 0:
                for pst in asset_details['fence']:
                    fence_name = fence_name + ', ' + str(pst['feature_name'])

            pool_name = ''
            if 'pool' in asset_details and len(asset_details['pool']) > 0:
                for pst in asset_details['pool']:
                    pool_name = pool_name + ', ' + str(pst['feature_name'])

            garage_parking_name = ''
            if 'garage_parking' in asset_details and len(asset_details['garage_parking']) > 0:
                for pst in asset_details['garage_parking']:
                    garage_parking_name = garage_parking_name + ', ' + str(pst['feature_name'])

            garage_features_name = ''
            if 'garage_features' in asset_details and len(asset_details['garage_features']) > 0:
                for pst in asset_details['garage_features']:
                    garage_features_name = garage_features_name + ', ' + str(pst['feature_name'])

            outbuildings_name = ''
            if 'outbuildings' in asset_details and len(asset_details['outbuildings']) > 0:
                for pst in asset_details['outbuildings']:
                    outbuildings_name = outbuildings_name + ', ' + str(pst['feature_name'])

            location_features_name = ''
            if 'location_features' in asset_details and len(asset_details['location_features']) > 0:
                for pst in asset_details['location_features']:
                    location_features_name = location_features_name + ', ' + str(pst['feature_name'])

            road_frontage_name = ''
            if 'road_frontage' in asset_details and len(asset_details['road_frontage']) > 0:
                for pst in asset_details['road_frontage']:
                    road_frontage_name = road_frontage_name + ', ' + str(pst['feature_name'])

            property_faces_name = ''
            if 'property_faces' in asset_details and len(asset_details['property_faces']) > 0:
                for pst in asset_details['property_faces']:
                    property_faces_name = property_faces_name + ', ' + str(pst['feature_name'])

            property_faces_name = ''
            if 'property_faces' in asset_details and len(asset_details['property_faces']) > 0:
                for pst in asset_details['property_faces']:
                    property_faces_name = property_faces_name + ', ' + str(pst['feature_name'])

            property_pic_urls = ''
            if 'property_pic' in asset_details and len(asset_details['property_pic']) > 0:
                for pst in asset_details['property_pic']:
                    doc_file_name = pst['doc_file_name']
                    bucket_name = pst['bucket_name']
                    image_url = settings.AWS_URL+''+bucket_name+'/'+doc_file_name
                    property_pic_urls = property_pic_urls+','+image_url

            property_video_urls = []
            add_demo_video = 0
            if 'property_video' in asset_details and len(asset_details['property_video']) > 0:
                if len(asset_details['property_video']) < 5:
                    add_demo_video = 5 - len(asset_details['property_video'])
                for pst in asset_details['property_video']:
                    embeded_code = pst['doc_file_name'].split('/')[-1]
                    pst['embeded_code'] = embeded_code



            property_lease_type = ''
            if 'lease_type' in asset_details and len(asset_details['lease_type']) > 0:
                for pst in asset_details['lease_type']:
                    property_lease_type = property_lease_type + ', ' + str(pst['feature_name'])

            tenant_pays_name = ''
            if 'tenant_pays' in asset_details and len(asset_details['tenant_pays']) > 0:
                for pst in asset_details['tenant_pays']:
                    tenant_pays_name = tenant_pays_name + ', ' + str(pst['feature_name'])

            inclusions_name = ''
            if 'inclusions' in asset_details and len(asset_details['inclusions']) > 0:
                for pst in asset_details['inclusions']:
                    inclusions_name = inclusions_name + ', ' + str(pst['feature_name'])

            building_class_name = ''
            if 'building_class' in asset_details and len(asset_details['building_class']) > 0:
                for pst in asset_details['building_class']:
                    building_class_name = building_class_name + ', ' + str(pst['feature_name'])

            interior_features_name = ''
            if 'interior_features' in asset_details and len(asset_details['interior_features']) > 0:
                for pst in asset_details['interior_features']:
                    interior_features_name = interior_features_name + ', ' + str(pst['feature_name'])

            mineral_rights_name = ''
            if 'mineral_rights' in asset_details and len(asset_details['mineral_rights']) > 0:
                for pst in asset_details['mineral_rights']:
                    mineral_rights_name = mineral_rights_name + ', ' + str(pst['feature_name'])

            easements_name = ''
            if 'easements' in asset_details and len(asset_details['easements']) > 0:
                for pst in asset_details['easements']:
                    easements_name = easements_name + ', ' + str(pst['feature_name'])

            survey_name = ''
            if 'survey' in asset_details and len(asset_details['survey']) > 0:
                for pst in asset_details['survey']:
                    survey_name = survey_name + ', ' + str(pst['feature_name'])

            utilities_name = ''
            if 'utilities' in asset_details and len(asset_details['utilities']) > 0:
                for pst in asset_details['utilities']:
                    utilities_name = utilities_name + ', ' + str(pst['feature_name'])

            improvements_name = ''
            if 'improvements' in asset_details and len(asset_details['improvements']) > 0:
                for pst in asset_details['improvements']:
                    improvements_name = improvements_name + ', ' + str(pst['feature_name'])

            irrigation_system_name = ''
            if 'irrigation_system' in asset_details and len(asset_details['irrigation_system']) > 0:
                for pst in asset_details['irrigation_system']:
                    irrigation_system_name = irrigation_system_name + ', ' + str(pst['feature_name'])

            topography_name = ''
            if 'topography' in asset_details and len(asset_details['topography']) > 0:
                for pst in asset_details['topography']:
                    topography_name = topography_name + ', ' + str(pst['feature_name'])

            wildlife_name = ''
            if 'wildlife' in asset_details and len(asset_details['wildlife']) > 0:
                for pst in asset_details['wildlife']:
                    wildlife_name = wildlife_name + ', ' + str(pst['feature_name'])

            fish_name = ''
            if 'fish' in asset_details and len(asset_details['fish']) > 0:
                for pst in asset_details['fish']:
                    fish_name = fish_name + ', ' + str(pst['feature_name'])

            recreation_name = ''
            if 'recreation' in asset_details and len(asset_details['recreation']) > 0:
                for pst in asset_details['recreation']:
                    recreation_name = recreation_name + ', ' + str(pst['feature_name'])


            property_lease_type = property_lease_type.lstrip(', ')
            property_sub_types = property_sub_types.lstrip(', ')
            term_accepted_name = term_accepted_name.lstrip(', ')
            occupied_by_name = occupied_by_name.lstrip(', ')
            ownership_name = ownership_name.lstrip(', ')
            possession_name = possession_name.lstrip(', ')
            style_name = style_name.lstrip(', ')
            stories_name = stories_name.lstrip(', ')
            recent_updates_name = recent_updates_name.lstrip(', ')
            security_features_name = security_features_name.lstrip(', ')
            cooling_name = cooling_name.lstrip(', ')
            heating_name = heating_name.lstrip(', ')
            electric_name = electric_name.lstrip(', ')
            gas_name = gas_name.lstrip(', ')
            water_name = water_name.lstrip(', ')
            sewer_name = sewer_name.lstrip(', ')
            tax_exemptions_name = tax_exemptions_name.lstrip(', ')
            hoa_amenities_name = hoa_amenities_name.lstrip(', ')
            zoning_name = zoning_name.lstrip(', ')
            kitchen_features_name = kitchen_features_name.lstrip(', ')
            appliances_name = appliances_name.lstrip(', ')
            flooring_name = flooring_name.lstrip(', ')
            windows_name = windows_name.lstrip(', ')
            other_rooms_name = other_rooms_name.lstrip(', ')
            bedroom_features_name = bedroom_features_name.lstrip(', ')
            bathroom_features_name = bathroom_features_name.lstrip(', ')
            master_bedroom_features_name = master_bedroom_features_name.lstrip(', ')
            basement_features_name = basement_features_name.lstrip(', ')
            other_features_name = other_features_name.lstrip(', ')
            fireplace_type_name = fireplace_type_name.lstrip(', ')
            handicap_amenities_name = handicap_amenities_name.lstrip(', ')
            construction_name = construction_name.lstrip(', ')
            exterior_features_name = exterior_features_name.lstrip(', ')
            foundation_name = foundation_name.lstrip(', ')
            roof_name = roof_name.lstrip(', ')
            fence_name = fence_name.lstrip(', ')
            pool_name = pool_name.lstrip(', ')
            interior_features_name = interior_features_name.lstrip(', ')
            garage_parking_name = garage_parking_name.lstrip(', ')
            garage_features_name = garage_features_name.lstrip(', ')
            outbuildings_name = outbuildings_name.lstrip(', ')
            road_frontage_name = road_frontage_name.lstrip(', ')
            property_faces_name = property_faces_name.lstrip(', ')
            property_pic_urls = property_pic_urls.lstrip(', ')
            location_features_name = location_features_name.lstrip(', ')
            tenant_pays_name = tenant_pays_name.lstrip(', ')
            inclusions_name = inclusions_name.lstrip(', ')
            building_class_name = building_class_name.lstrip(', ')
            mineral_rights_name = mineral_rights_name.lstrip(', ')
            easements_name = easements_name.lstrip(', ')
            survey_name = survey_name.lstrip(', ')
            utilities_name = utilities_name.lstrip(', ')
            improvements_name = improvements_name.lstrip(', ')
            irrigation_system_name = irrigation_system_name.lstrip(', ')
            topography_name = topography_name.lstrip(', ')
            wildlife_name = wildlife_name.lstrip(', ')
            fish_name = fish_name.lstrip(', ')
            recreation_name = recreation_name.lstrip(', ')

            asset_details['property_sub_types'] = property_sub_types
            asset_details['term_accepted_name'] = term_accepted_name
            asset_details['occupied_by_name'] = occupied_by_name
            asset_details['ownership_name'] = ownership_name
            asset_details['possession_name'] = possession_name
            asset_details['style_name'] = style_name
            asset_details['stories_name'] = stories_name
            asset_details['property_address'] = property_address
            asset_details['recent_updates_name'] = recent_updates_name
            asset_details['security_features_name'] = security_features_name
            asset_details['cooling_name'] = cooling_name
            asset_details['heating_name'] = heating_name
            asset_details['electric_name'] = electric_name
            asset_details['gas_name'] = gas_name
            asset_details['water_name'] = water_name
            asset_details['sewer_name'] = sewer_name
            asset_details['tax_exemptions_name'] = tax_exemptions_name
            asset_details['hoa_amenities_name'] = hoa_amenities_name
            asset_details['zoning_name'] = zoning_name
            asset_details['kitchen_features_name'] = kitchen_features_name
            asset_details['appliances_name'] = appliances_name
            asset_details['flooring_name'] = flooring_name
            asset_details['windows_name'] = windows_name
            asset_details['bedroom_features_name'] = bedroom_features_name
            asset_details['other_rooms_name'] = other_rooms_name
            asset_details['bathroom_features_name'] = bathroom_features_name
            asset_details['basement_features_name'] = basement_features_name
            asset_details['other_features_name'] = other_features_name
            asset_details['fireplace_type_name'] = fireplace_type_name
            asset_details['handicap_amenities_name'] = handicap_amenities_name
            asset_details['construction_name'] = construction_name
            asset_details['exterior_features_name'] = exterior_features_name
            asset_details['roof_name'] = roof_name
            asset_details['foundation_name'] = foundation_name
            asset_details['fence_name'] = fence_name
            asset_details['pool_name'] = pool_name
            asset_details['garage_parking_name'] = garage_parking_name
            asset_details['garage_features_name'] = garage_features_name
            asset_details['outbuildings_name'] = outbuildings_name
            asset_details['road_frontage_name'] = road_frontage_name
            asset_details['property_faces_name'] = property_faces_name
            asset_details['location_features_name'] = location_features_name
            asset_details['property_pic_urls'] = property_pic_urls
            asset_details['property_video_urls'] = property_video_urls
            asset_details['current_bid_amount'] = current_bid_amount
            asset_details['bide_increments'] = bide_increments
            asset_details['next_bid_amount'] = next_bid_amount
            asset_details['property_lease_type'] = property_lease_type
            asset_details['tenant_pays_name'] = tenant_pays_name
            asset_details['inclusions_name'] = inclusions_name
            asset_details['building_class_name'] = building_class_name
            asset_details['interior_features_name'] = interior_features_name
            asset_details['mineral_rights_name'] = mineral_rights_name
            asset_details['easements_name'] = easements_name
            asset_details['survey_name'] = survey_name
            asset_details['utilities_name'] = utilities_name
            asset_details['improvements_name'] = improvements_name
            asset_details['irrigation_system_name'] = irrigation_system_name
            asset_details['topography_name'] = topography_name
            asset_details['wildlife_name'] = wildlife_name
            asset_details['fish_name'] = fish_name
            asset_details['recreation_name'] = recreation_name
            asset_details['master_bedroom_features_name'] = master_bedroom_features_name
            asset_details['add_demo_video'] = add_demo_video


        except Exception as exp:
            print(exp)
            asset_details = {}



        try:
            user_id = request.session['user_id']
            token = request.session['token']['access_token']
            user_viewed_params = {
                "user_id": user_id,
                "site_id": site_id,
                "property_id": property_id
            }

            user_viewed_url = settings.API_URL + '/api-property/add-property-view/'
            user_viewed_data = call_api_post_method(user_viewed_params, user_viewed_url, token)
        except:
            pass

        sale_type = int(asset_details['sale_by_type_id'])
        today_date = datetime.datetime.now()
        today = time.time()
        start_date_time = 0
        end_date_time = 0
        try:
            start_date_str = asset_details['property_auction_data'][0]['start_date']
            end_date_str = asset_details['property_auction_data'][0]['end_date']
        except:
            pass

        try:
            start_date_time = time.mktime(datetime.datetime.strptime(start_date_str, "%Y-%m-%dT%H:%M:%SZ").timetuple())
        except ValueError:
            start_date_time = time.mktime(datetime.datetime.strptime(start_date_str, "%Y-%m-%dT%H:%M:%S.%fZ").timetuple())
        except:
            start_date_time = 0

        try:
            end_date = datetime.datetime.strptime(end_date_str, "%Y-%m-%dT%H:%M:%SZ")
            end_date_time = time.mktime(datetime.datetime.strptime(end_date_str, "%Y-%m-%dT%H:%M:%SZ").timetuple())
        except ValueError:
            end_date = datetime.datetime.strptime(end_date_str, "%Y-%m-%dT%H:%M:%S.%fZ")
            end_date_time = time.mktime(datetime.datetime.strptime(end_date_str, "%Y-%m-%dT%H:%M:%S.%fZ").timetuple())
        except:
            end_date = ''
            end_date_time = 0

        # try:
        #     end_date = datetime.datetime.strptime(end_date_str, "%Y-%m-%dT%H:%M:%SZ")
        #     end_date_time = time.mktime(datetime.datetime.strptime(end_date_str, "%Y-%m-%dT%H:%M:%SZ").timetuple())
        # except ValueError:
        #     end_date = datetime.datetime.strptime(end_date_str, "%Y-%m-%dT%H:%M:%S.%fZ")
        #     end_date_time = time.mktime(datetime.datetime.strptime(end_date_str, "%Y-%m-%dT%H:%M:%S.%fZ").timetuple())
        # except:
        #     end_date = ''
        #     end_date_time = 0

        try:
            dutch_end_time_str = asset_details['property_auction_data'][0]['dutch_end_time']
            dutch_end_date = datetime.datetime.strptime(dutch_end_time_str, "%Y-%m-%dT%H:%M:%SZ")
            dutch_end_date_time = time.mktime(datetime.datetime.strptime(dutch_end_time_str, "%Y-%m-%dT%H:%M:%SZ").timetuple())
        except ValueError:
            dutch_end_time_str = asset_details['property_auction_data'][0]['dutch_end_time']
            dutch_end_date = datetime.datetime.strptime(dutch_end_time_str, "%Y-%m-%dT%H:%M:%S.%fZ")
            dutch_end_date_time = time.mktime(datetime.datetime.strptime(dutch_end_time_str, "%Y-%m-%dT%H:%M:%S.%fZ").timetuple())
        except:
            dutch_end_date = ''
            dutch_end_date_time = 0

        try:
            sealed_start_time_str = asset_details['property_auction_data'][0]['sealed_start_time']
            sealed_start_date = datetime.datetime.strptime(sealed_start_time_str, "%Y-%m-%dT%H:%M:%SZ")
            sealed_start_date_time = time.mktime(datetime.datetime.strptime(sealed_start_time_str, "%Y-%m-%dT%H:%M:%SZ").timetuple())
        except ValueError:
            sealed_start_time_str = asset_details['property_auction_data'][0]['sealed_start_time']
            sealed_start_date = datetime.datetime.strptime(sealed_start_time_str, "%Y-%m-%dT%H:%M:%S.%fZ")
            sealed_start_date_time = time.mktime(datetime.datetime.strptime(sealed_start_time_str, "%Y-%m-%dT%H:%M:%S.%fZ").timetuple())
        except:
            sealed_start_date = ''
            sealed_start_date_time = 0

        try:
            sealed_end_time_str = asset_details['property_auction_data'][0]['sealed_end_time']
            sealed_end_date = datetime.datetime.strptime(sealed_end_time_str, "%Y-%m-%dT%H:%M:%SZ")
            sealed_end_date_time = time.mktime(datetime.datetime.strptime(sealed_end_time_str, "%Y-%m-%dT%H:%M:%SZ").timetuple())
        except ValueError:
            sealed_end_time_str = asset_details['property_auction_data'][0]['sealed_end_time']
            sealed_end_date = datetime.datetime.strptime(sealed_end_time_str, "%Y-%m-%dT%H:%M:%S.%fZ")
            sealed_end_date_time = time.mktime(datetime.datetime.strptime(sealed_end_time_str, "%Y-%m-%dT%H:%M:%S.%fZ").timetuple())
        except:
            sealed_end_date = ''
            sealed_end_date_time = 0

        try:
            english_start_time_str = asset_details['property_auction_data'][0]['english_start_time']
            english_start_date = datetime.datetime.strptime(english_start_time_str, "%Y-%m-%dT%H:%M:%SZ")
            english_start_date_time = time.mktime(datetime.datetime.strptime(english_start_time_str, "%Y-%m-%dT%H:%M:%SZ").timetuple())
        except ValueError:
            english_start_time_str = asset_details['property_auction_data'][0]['english_start_time']
            english_start_date = datetime.datetime.strptime(english_start_time_str, "%Y-%m-%dT%H:%M:%S.%fZ")
            english_start_date_time = time.mktime(datetime.datetime.strptime(english_start_time_str, "%Y-%m-%dT%H:%M:%S.%fZ").timetuple())
        except:
            english_start_date = ''
            english_start_date_time = 0

        try:
            added_on_str = asset_details['added_on']
            added_on = datetime.datetime.strptime(added_on_str, "%Y-%m-%dT%H:%M:%SZ")
            #added_on = added_on + datetime.timedelta(days=15)
            traditional_status_time = time.mktime(added_on.timetuple())
        except ValueError:
            added_on_str = asset_details['added_on']
            added_on = datetime.datetime.strptime(added_on_str, "%Y-%m-%dT%H:%M:%S.%fZ")
            #added_on = added_on + datetime.timedelta(days=15)
            traditional_status_time = time.mktime(added_on.timetuple())
        except:
            traditional_status_time = 0

        try:
            delta = end_date - today_date
            no_days = delta.days
            hours_remain = no_days * 24
        except:
            hours_remain = 0

        try:
            is_approved = asset_details['is_approved']
            is_reviewed = asset_details['is_reviewed']
        except:
            is_approved = 0
            is_reviewed = False

        try:
            reserve_amount = asset_details['property_auction_data'][0]['reserve_amount']
        except:
            reserve_amount = 0

        bid_count = int(asset_details['bid_count']) if 'bid_count' in asset_details and int(asset_details['bid_count']) > 0 else 0
        total_bid_till_sealed = int(asset_details['total_bid_till_sealed']) if 'total_bid_till_sealed' in asset_details and int(asset_details['total_bid_till_sealed']) > 0 else 0
        sealed_user_bid_count = int(asset_details['sealed_user_bid_count']) if 'sealed_user_bid_count' in asset_details and int(asset_details['sealed_user_bid_count']) > 0 else 0
        english_user_bid_count = int(asset_details['english_user_bid_count']) if 'english_user_bid_count' in asset_details and int(asset_details['english_user_bid_count']) > 0 else 0
        my_bid_count = int(asset_details['my_bid_count']) if 'my_bid_count' in asset_details and asset_details['my_bid_count'] != "" else 0
        high_bidder_id = int(asset_details['high_bidder_id']) if 'high_bidder_id' in asset_details and asset_details['high_bidder_id'] !="" else 0
        flash_timer = asset_details['property_setting']['time_flash'] if 'property_setting' in asset_details and asset_details['property_setting'] and 'time_flash' in asset_details['property_setting'] and asset_details['property_setting']['time_flash'] else 0
        flash_timer = flash_timer*60*1000
        reverse_not_met = asset_details['property_setting']['show_reverse_not_met'] if 'property_setting' in asset_details and asset_details['property_setting'] and 'show_reverse_not_met' in asset_details['property_setting'] and asset_details['property_setting']['show_reverse_not_met'] else False

        # condition for button text
        bid_now_style = {
            'id': 'bidNowFormSection',
            'style': 'display:none;'
        }
        pre_final_style = {
            'id': 'preFinalBiddingSection',
            'style': 'display: inline-block;'
        }
        final_style = {
            'id': 'finalBiddingSection',
            'style': 'display:none;'
        }
        final_lost_style = {
            'id': 'finalBiddingLostSection',
            'style': 'display:none;'
        }
        reg_bid_btn_style = {
            'id': 'registerBidBtn',
            'style': 'display:inline-block;',
            'btn_url': '/bid-registration/?property=' + str(asset_details['id'])
        }
        bidder_msg_section_style = {
            'id': 'bidder_msg_section',
            'style': 'display:none;',
            'text': '',
            'class': 'badge-success',
            'icon': '',
        }

        time_remaining_style = {
            'class': 'time_remaining',
            'style': 'display:block;'
        }
        countdown_text_style = {
            'class': 'countdown_text',
            'style': 'display:inline-block;'
        }
        make_offer_btn_style = {
            'id': 'make_offer_btn',
            'style': 'display:none;'
        }
        traditional_style = {
            'id': 'TraditionalOfferSection',
            'style': 'display:none;'
        }
        live_auction_style = {
            'id': 'liveAuctionSection',
            'style': 'display:none;'
        }
        best_offer_style = {
            'id': 'bestOfferSection',
            'style': 'display:none;'
        }
        # insider auction functionality
        insider_auction_style = {
            'id': 'insiderAuctionSection',
            'style': 'display:none;'
        }
        insider_reg_bid_btn_style = {
            'id': 'insiderRegisterBidBtn',
            'style': 'display:none;',
            'btn_url': '/bid-registration/?property=' + str(asset_details['id'])
        }
        insider_bidder_msg_section_style = {
            'id': 'insider_bidder_msg_section',
            'style': 'display:none;',
            'text': '',
            'class': 'badge-success',
            'icon': '',
        }
        insider_dutch_bid_btn_style = {
            'id': 'insiderDutchBidBtn',
            'style': 'display:none;',
        }
        insider_lost_section_btn_style = {
            'id': 'insiderLostSectionBtn',
            'style': 'display:none;',
        }
        insider_bid_now_style = {
            'id': 'insiderbidNowFormSection',
            'style': 'display:none;'
        }
        insider_bid_won_style = {
            'id': 'insiderBiddingWonSection',
            'style': 'display:none;'
        }
        insider_bid_inc_style = {
            'id': 'englishAuctionBidIncSection',
            'style': 'display:none;'
        }
        insider_bid_count_style = {
            'id': 'insiderBiddingLostSectionBidCount',
            'style': 'display:none;'
        }
        insider_bidder_btn_txt = ''
        if int(sale_type) == 1:
            if reverse_not_met:
                if not reserve_amount:
                    reserve_text_style = {
                        'id': 'reserve_text',
                        'style': 'display:inline-block;',
                        'text': 'No Reserve',
                        'class': 'badge-danger',
                        'tooltip_text': '1. At the conclusion of the auction the Highest Bid will be sent to the seller for approval. <br>2. Sells to the highest bidder. Not subject to seller approval.'
                    }
                elif int(current_bid_amount) >= int(reserve_amount):
                    reserve_text_style = {
                        'id': 'reserve_text',
                        'style': 'display:inline-block;',
                        'text': 'Reserve Met',
                        'class': 'badge-success',
                        'tooltip_text': 'Sells to the highest bidder. Not subject to seller approval.'
                    }
                else:
                    reserve_text_style = {
                        'id': 'reserve_text',
                        'style': 'display:inline-block;',
                        'text': 'Reserve Not Met',
                        'class': 'badge-danger',
                        'tooltip_text': 'The reserve price is confidential and will not be shown. Once the reserve price is met it will be displayed. At the conclusion of the auction, the seller may negotiate with the bidders if the reserve price isnt met.'
                    }
            else:
                reserve_text_style = {
                    'id': 'reserve_text',
                    'style': 'display:inline-block;',
                    'text': 'No Reserve',
                    'class': 'badge-danger',
                    'tooltip_text': '1. At the conclusion of the auction the Highest Bid will be sent to the seller for approval. <br>2. Sells to the highest bidder. Not subject to seller approval.'
                }
        else:
            reserve_text_style = {
                'id': 'reserve_text',
                'style': 'display:none;',
                'text': 'No Reserve',
                'class': 'badge-danger',
                'tooltip_text': '1. At the conclusion of the auction the Highest Bid will be sent to the seller for approval. <br>2. Sells to the highest bidder. Not subject to seller approval.'
            }
        bidder_lost_section_success_msg = ''
        if user_id:
            if int(sale_type) == 1:
                if 'property_status_id' in asset_details and int(asset_details['property_status_id']) == 1 and int(asset_details['auction_status']) == 1:
                    if is_approved == 3 or is_approved == 4:
                        if is_approved == 4:
                            bidder_btn_txt = 'Not Interested'
                            reg_bid_btn_style = {
                                'id': 'registerBidBtn',
                                'style': 'display:inline-block;cursor:initial;',
                                'btn_url': 'javascript:void(0)'
                            }
                        else:
                            bidder_btn_txt = 'Registration Declined'
                            reg_bid_btn_style = {
                                'id': 'registerBidBtn',
                                'style': 'display:inline-block;cursor:initial;',
                                'btn_url': 'javascript:void(0)'
                            }
                        if ((today >= start_date_time and today <= end_date_time) or today < start_date_time):
                            time_remaining_style = {
                                'class': 'time_remaining',
                                'style': 'display:block;'
                            }
                            countdown_text_style = {
                                'class': 'countdown_text',
                                'style': 'display:inline-block;'
                            }
                        else:
                            time_remaining_style = {
                                'class': 'time_remaining',
                                'style': 'display:block;'
                            }
                            countdown_text_style = {
                                'class': 'countdown_text',
                                'style': 'display:inline-block;'
                            }


                        bid_now_style = {
                            'id': 'bidNowFormSection',
                            'style': 'display:none;'
                        }
                        pre_final_style = {
                            'id': 'preFinalBiddingSection',
                            'style': 'display:inline-block;'
                        }
                        final_style = {
                            'id': 'finalBiddingSection',
                            'style': 'display:none;'
                        }
                        make_offer_btn_style = {
                            'id': 'make_offer_btn',
                            'style': 'display:none;'
                        }
                    elif (today >= start_date_time and today <= end_date_time and is_approved == 1 and (
                            is_reviewed == False or is_reviewed == True)) or (
                            today < start_date_time and is_approved == 1 and (
                            is_reviewed == False or is_reviewed == True)) or (
                            today < start_date_time and is_approved == 2 and is_reviewed == False) or (
                            today >= start_date_time and today <= end_date_time and is_approved == 2 and is_reviewed == False):

                        bidder_btn_txt = 'Registration Pending Approval'
                        if templete_dir == "theme-4":
                            bidder_btn_txt = 'You are registered to bid. <strong>Approval Pending</strong>'

                        reg_bid_btn_style = {
                            'id': 'registerBidBtn',
                            'style': 'display:inline-block;text-transform:uppercase;cursor:initial;',
                            'btn_url': 'javascript:void(0)'
                        }
                        if templete_dir == "theme-4":
                            reg_bid_btn_style = {
                                'id': 'registerBidBtn',
                                'style': 'display:inline-block;',
                                'btn_url': 'javascript:void(0)'
                            }
                        time_remaining_style = {
                            'class': 'time_remaining',
                            'style': 'display:block;'
                        }
                        countdown_text_style = {
                            'class': 'countdown_text',
                            'style': 'display:inline-block;'
                        }

                        bid_now_style = {
                            'id': 'bidNowFormSection',
                            'style': 'display:none;'
                        }
                        pre_final_style = {
                            'id': 'preFinalBiddingSection',
                            'style': 'display:inline-block;'
                        }
                        final_style = {
                            'id': 'finalBiddingSection',
                            'style': 'display:none;'
                        }
                        make_offer_btn_style = {
                            'id': 'make_offer_btn',
                            'style': 'display:none;'
                        }
                    elif today >= start_date_time and today <= end_date_time  and is_approved == 2 and is_reviewed == True:
                        bidder_btn_txt = ''
                        bid_now_style = {
                            'id': 'bidNowFormSection',
                            'style': 'display:block;'
                        }


                        reg_bid_btn_style = {
                            'id': 'registerBidBtn',
                            'style': 'display:none;text-transform:uppercase;cursor:initial;',
                            'btn_url': 'javascript:void(0)'
                        }
                        if my_bid_count > 0:
                            if high_bidder_id == int(user_id):
                                bidder_msg_section_style = {
                                    'id': 'bidder_msg_section',
                                    'style': 'display:inline-block;',
                                    'text': 'You Have The Highest Bid.',
                                    'class': 'badge-success',
                                    'icon': 'fas fa-check-circle'
                                }
                            else:
                                bidder_msg_section_style = {
                                    'id': 'bidder_msg_section',
                                    'style': 'display:inline-block;',
                                    'text': 'You Are Out Bidded.',
                                    'class': 'badge-danger',
                                    'icon': 'fas fa-exclamation-triangle'
                                }
                        else:
                            if templete_dir == "theme-4":
                                bidder_msg_section_style = {
                                    'id': 'bidder_msg_section',
                                    'style': 'display:inline-block;',
                                    'text': 'You are registered and approved to bid. <strong>Place your first bid now.</strong>',
                                    'class': 'badge-success',
                                    'icon': 'fas fa-check-circle'
                                }
                            else:
                                bidder_msg_section_style = {
                                    'id': 'bidder_msg_section',
                                    'style': 'display:inline-block;',
                                    'text': 'You are registered and approved to bid. Place your first bid now.',
                                    'class': 'badge-success',
                                    'icon': 'fas fa-check-circle'
                                }

                        time_remaining_style = {
                            'class': 'time_remaining',
                            'style': 'display:block;'
                        }
                        countdown_text_style = {
                            'class': 'countdown_text',
                            'style': 'display:inline-block;'
                        }
                        pre_final_style = {
                            'id': 'preFinalBiddingSection',
                            'style': 'display:inline-block;'
                        }
                        final_style = {
                            'id': 'finalBiddingSection',
                            'style': 'display:none;'
                        }
                        make_offer_btn_style = {
                            'id': 'make_offer_btn',
                            'style': 'display:none;'
                        }
                    elif today < start_date_time and is_approved == 2 and is_reviewed == True:
                        bidder_btn_txt = 'Auction Has Not Started'
                        if templete_dir == "theme-4":
                            bidder_btn_txt = 'You are registered and <strong>Approved To Bid</strong>'
                        reg_bid_btn_style = {
                            'id': 'registerBidBtn',
                            'style': 'display:inline-block;cursor:initial;',
                            'btn_url': 'javascript:void(0)'
                        }
                        bidder_msg_section_style = {
                            'id': 'bidder_msg_section',
                            'style': 'display:block;',
                            'text': 'Registration Approved But Auction Not Started Yet.',
                            'class': 'badge-success',
                            'icon': 'fas fa-check-circle'
                        }

                        time_remaining_style = {
                            'class': 'time_remaining',
                            'style': 'display:block;'
                        }
                        countdown_text_style = {
                            'class': 'countdown_text',
                            'style': 'display:inline-block;'
                        }

                        bid_now_style = {
                            'id': 'bidNowFormSection',
                            'style': 'display:none;'
                        }
                        pre_final_style = {
                            'id': 'preFinalBiddingSection',
                            'style': 'display:inline-block;'
                        }
                        final_style = {
                            'id': 'finalBiddingSection',
                            'style': 'display:none;'
                        }
                        make_offer_btn_style = {
                            'id': 'make_offer_btn',
                            'style': 'display:none;'
                        }
                    elif today > end_date_time:
                        bidder_btn_txt = ''
                        if int(user_id) == (high_bidder_id):
                            final_style = {
                                'id': 'finalBiddingSection',
                                'style': 'display:block;'
                            }
                            final_lost_style = {
                                'id': 'finalBiddingLostSection',
                                'style': 'display:none;'
                            }
                        else:
                            final_style = {
                                'id': 'finalBiddingSection',
                                'style': 'display:none;'
                            }
                            final_lost_style = {
                                'id': 'finalBiddingLostSection',
                                'style': 'display:block;'
                            }

                        pre_final_style = {
                            'id': 'preFinalBiddingSection',
                            'style': 'display:none;'
                        }
                        reg_bid_btn_style = {
                            'id': 'registerBidBtn',
                            'style': 'display:none;',
                            'btn_url': 'javascript:void(0)'
                        }
                        make_offer_btn_style = {
                            'id': 'make_offer_btn',
                            'style': 'display:none;'
                        }

                        bidder_msg_section_style = {
                            'id': 'bidder_msg_section',
                            'style': 'display:none;',
                            'text': '',
                            'class': 'badge-success',
                            'icon': 'fas fa-check-circle'
                        }

                        time_remaining_style = {
                            'class': 'time_remaining',
                            'style': 'display:block;'
                        }
                        countdown_text_style = {
                            'class': 'countdown_text',
                            'style': 'display:inline-block;'
                        }

                        bid_now_style = {
                            'id': 'bidNowFormSection',
                            'style': 'display:none;'
                        }
                    else:
                        bidder_btn_txt = 'Register For Auction'
                        if templete_dir == "theme-4":
                            if user_id == "":
                                bidder_btn_txt = "Login/Register to Bid"
                            else:
                                bidder_btn_txt = "Register to Bid"
                        reg_bid_btn_style = {
                            'id': 'registerBidBtn',
                            'style': 'display:inline-block;',
                            'btn_url': '/bid-registration/?property=' + str(asset_details['id'])
                        }

                        final_style = {
                            'id': 'finalBiddingSection',
                            'style': 'display:none;'
                        }

                        pre_final_style = {
                            'id': 'preFinalBiddingSection',
                            'style': 'display:block;'
                        }
                        make_offer_btn_style = {
                            'id': 'make_offer_btn',
                            'style': 'display:none;'
                        }

                        bidder_msg_section_style = {
                            'id': 'bidder_msg_section',
                            'style': 'display:none;',
                            'text': '',
                            'class': 'badge-success',
                            'icon': 'fas fa-check-circle'
                        }

                        if ((today >= start_date_time and today <= end_date_time) or today < start_date_time):
                            time_remaining_style = {
                                'class': 'time_remaining',
                                'style': 'display:block;'
                            }
                            countdown_text_style = {
                                'class': 'countdown_text',
                                'style': 'display:inline-block;'
                            }
                        else:
                            time_remaining_style = {
                                'class': 'time_remaining',
                                'style': 'display:block;'
                            }
                            countdown_text_style = {
                                'class': 'countdown_text',
                                'style': 'display:inline-block;'
                            }

                        bid_now_style = {
                            'id': 'bidNowFormSection',
                            'style': 'display:none;'
                        }
                elif int(asset_details['property_status_id']) == 9:
                    bidder_btn_txt = ''
                    if int(user_id) == (high_bidder_id):

                        final_style = {
                            'id': 'finalBiddingSection',
                            'style': 'display:block;'
                        }
                        final_lost_style = {
                            'id': 'finalBiddingLostSection',
                            'style': 'display:none;'
                        }
                    else:
                        final_style = {
                            'id': 'finalBiddingSection',
                            'style': 'display:none;'
                        }
                        final_lost_style = {
                            'id': 'finalBiddingLostSection',
                            'style': 'display:block;'
                        }
                        if int(my_bid_count) > 0:
                            bidder_lost_section_success_msg = '<i class="fas fa-exclamation-triangle"></i> Out Bid.'
                            bidder_msg_section_style = {
                                'id': 'bidder_msg_section',
                                'style': 'display:block;',
                                'text': '',
                                'class': 'badge-danger',
                                'icon': 'fas fa-check-circle'
                            }
                        else:
                            bidder_msg_section_style = {
                                'id': 'bidder_msg_section',
                                'style': 'display:none;',
                                'text': '',
                                'class': 'badge-success',
                                'icon': 'fas fa-check-circle'
                            }

                    pre_final_style = {
                        'id': 'preFinalBiddingSection',
                        'style': 'display:none;'
                    }
                    reg_bid_btn_style = {
                        'id': 'registerBidBtn',
                        'style': 'display:none;',
                        'btn_url': 'javascript:void(0)'
                    }
                    make_offer_btn_style = {
                        'id': 'make_offer_btn',
                        'style': 'display:none;'
                    }



                    time_remaining_style = {
                        'class': 'time_remaining',
                        'style': 'display:block;'
                    }
                    countdown_text_style = {
                        'class': 'countdown_text',
                        'style': 'display:inline-block;'
                    }

                    bid_now_style = {
                        'id': 'bidNowFormSection',
                        'style': 'display:none;'
                    }
                else:
                    bidder_btn_txt = ''
                    if int(my_bid_count) > 0:
                        if int(user_id) == (high_bidder_id):
                            bidder_lost_section_success_msg = '<i class="fas fa-check-circle"></i> You Have Highest Bid.'
                        else:
                            bidder_lost_section_success_msg = '<i class="fas fa-exclamation-triangle"></i> Out Bid.'
                    else:
                        bidder_lost_section_success_msg = ''
                    final_style = {
                        'id': 'finalBiddingSection',
                        'style': 'display:none;'
                    }
                    final_lost_style = {
                        'id': 'finalBiddingLostSection',
                        'style': 'display:block;'
                    }

                    pre_final_style = {
                        'id': 'preFinalBiddingSection',
                        'style': 'display:none;'
                    }
                    reg_bid_btn_style = {
                        'id': 'registerBidBtn',
                        'style': 'display:none;',
                        'btn_url': 'javascript:void(0)'
                    }
                    make_offer_btn_style = {
                        'id': 'make_offer_btn',
                        'style': 'display:none;'
                    }
                    if int(my_bid_count) > 0:
                        bidder_msg_section_style = {
                            'id': 'bidder_msg_section',
                            'style': 'display:block;',
                            'text': '',
                            'class': 'badge-danger',
                            'icon': 'fas fa-check-circle'
                        }
                    else:
                        bidder_msg_section_style = {
                            'id': 'bidder_msg_section',
                            'style': 'display:none;',
                            'text': '',
                            'class': 'badge-success',
                            'icon': 'fas fa-check-circle'
                        }

                    time_remaining_style = {
                        'class': 'time_remaining',
                        'style': 'display:block;'
                    }
                    countdown_text_style = {
                        'class': 'countdown_text',
                        'style': 'display:inline-block;'
                    }

                    bid_now_style = {
                        'id': 'bidNowFormSection',
                        'style': 'display:none;'
                    }

            elif int(sale_type) == 2:
                if 'property_status_id' in asset_details and int(asset_details['property_status_id']) == 1 and int(asset_details['auction_status']) == 1:
                    if is_approved == 3 or is_approved == 4:
                        if is_approved == 4:
                            bidder_btn_txt = ''
                            reg_bid_btn_style = {
                                'id': 'registerBidBtn',
                                'style': 'display:none;cursor:initial;',
                                'btn_url': 'javascript:void(0)'
                            }
                            insider_bidder_btn_txt = 'Not Interested'
                            insider_reg_bid_btn_style = {
                                'id': 'insiderRegisterBidBtn',
                                'style': 'display:inline-block;cursor:initial;',
                                'btn_url': 'javascript:void(0)'
                            }
                        else:
                            bidder_btn_txt = ''
                            reg_bid_btn_style = {
                                'id': 'registerBidBtn',
                                'style': 'display:none;cursor:initial;',
                                'btn_url': 'javascript:void(0)'
                            }
                            insider_bidder_btn_txt = 'Registration Declined'
                            insider_reg_bid_btn_style = {
                                'id': 'insiderRegisterBidBtn',
                                'style': 'display:inline-block;cursor:initial;',
                                'btn_url': 'javascript:void(0)'
                            }
                        if ((today >= start_date_time and today <= end_date_time) or today < start_date_time):
                            time_remaining_style = {
                                'class': 'time_remaining',
                                'style': 'display:block;'
                            }
                            countdown_text_style = {
                                'class': 'countdown_text',
                                'style': 'display:inline-block;'
                            }
                        else:
                            time_remaining_style = {
                                'class': 'time_remaining',
                                'style': 'display:block;'
                            }
                            countdown_text_style = {
                                'class': 'countdown_text',
                                'style': 'display:inline-block;'
                            }


                        bid_now_style = {
                            'id': 'bidNowFormSection',
                            'style': 'display:none;'
                        }
                        pre_final_style = {
                            'id': 'preFinalBiddingSection',
                            'style': 'display:none;'
                        }
                        insider_auction_style = {
                            'id': 'insiderAuctionSection',
                            'style': 'display:inline-block;'
                        }
                        final_style = {
                            'id': 'finalBiddingSection',
                            'style': 'display:none;'
                        }
                        make_offer_btn_style = {
                            'id': 'make_offer_btn',
                            'style': 'display:none;'
                        }
                    elif (today >= start_date_time and today <= end_date_time and is_approved == 1 and (
                            is_reviewed == False or is_reviewed == True)) or (
                            today < start_date_time and is_approved == 1 and (
                            is_reviewed == False or is_reviewed == True)) or (
                            today < start_date_time and is_approved == 2 and is_reviewed == False) or (
                            today >= start_date_time and today <= end_date_time and is_approved == 2 and is_reviewed == False):

                        bidder_btn_txt = ''
                        reg_bid_btn_style = {
                            'id': 'registerBidBtn',
                            'style': 'display:none;text-transform:uppercase;cursor:initial;',
                            'btn_url': 'javascript:void(0)'
                        }
                        insider_bidder_btn_txt = 'Registration Pending Approval'
                        insider_reg_bid_btn_style = {
                            'id': 'insiderRegisterBidBtn',
                            'style': 'display:inline-block;text-transform:uppercase;cursor:initial;',
                            'btn_url': 'javascript:void(0)'
                        }
                        time_remaining_style = {
                            'class': 'time_remaining',
                            'style': 'display:block;'
                        }
                        countdown_text_style = {
                            'class': 'countdown_text',
                            'style': 'display:inline-block;'
                        }

                        bid_now_style = {
                            'id': 'bidNowFormSection',
                            'style': 'display:none;'
                        }
                        insider_auction_style = {
                            'id': 'insiderAuctionSection',
                            'style': 'display:inline-block;'
                        }
                        pre_final_style = {
                            'id': 'preFinalBiddingSection',
                            'style': 'display:none;'
                        }
                        final_style = {
                            'id': 'finalBiddingSection',
                            'style': 'display:none;'
                        }
                        make_offer_btn_style = {
                            'id': 'make_offer_btn',
                            'style': 'display:none;'
                        }
                    elif today >= start_date_time and today <= end_date_time  and is_approved == 2 and is_reviewed == True:
                        bidder_btn_txt = ''
                        bid_now_style = {
                            'id': 'bidNowFormSection',
                            'style': 'display:none;'
                        }

                        reg_bid_btn_style = {
                            'id': 'registerBidBtn',
                            'style': 'display:none;text-transform:uppercase;cursor:initial;',
                            'btn_url': 'javascript:void(0)'
                        }
                        if not asset_details['dutch_winning_amount'] and today >= start_date_time and today < dutch_end_date_time:
                            insider_dutch_bid_btn_style = {
                                'id': 'insiderDutchBidBtn',
                                'style': 'display:inline-block;',
                            }
                        if today > sealed_start_date_time and today < sealed_end_date_time and 'dutch_winning_user' in asset_details and asset_details['dutch_winning_user'] and int(user_id) != int(asset_details['dutch_winning_user']):
                            insider_bid_now_style = {
                                'id': 'insiderbidNowFormSection',
                                'style': 'display:block;'
                            }
                        else:
                            insider_bid_now_style = {
                                'id': 'insiderbidNowFormSection',
                                'style': 'display:none;'
                            }

                        if (((today > sealed_start_date_time and today < sealed_end_date_time) or (today > dutch_end_date_time and today < sealed_start_date_time)) and 'dutch_winning_user' in asset_details and asset_details['dutch_winning_user'] and int(user_id) == int(asset_details['dutch_winning_user'])):

                            insider_bidder_msg_section_style = {
                                'id': 'insider_bidder_msg_section',
                                'style': 'display:inline-block;',
                                'text': 'You are the winner of Round One so you can\'t participate in Sealed Bid Auction.Please wait for English Auction.',
                                'class': 'badge-success',
                                'icon': 'fas fa-check-circle'
                            }
                        else:
                            insider_bidder_msg_section_style = {
                                'id': 'insider_bidder_msg_section',
                                'style': 'display:none;',
                                'text': '',
                                'class': 'badge-success',
                                'icon': 'fas fa-check-circle'
                            }

                        if ((today > sealed_end_date_time and today < english_start_date_time) and 'seal_winning_user' in asset_details and asset_details['seal_winning_user'] and int(user_id) == int(asset_details['seal_winning_user'])):

                            insider_bidder_msg_section_style = {
                                'id': 'insider_bidder_msg_section',
                                'style': 'display:inline-block;',
                                'text': 'You are the winner of Round Two.Please wait for English Auction.',
                                'class': 'badge-success',
                                'icon': 'fas fa-check-circle'
                            }
                        else:
                            insider_bidder_msg_section_style = {
                                'id': 'insider_bidder_msg_section',
                                'style': 'display:none;',
                                'text': '',
                                'class': 'badge-success',
                                'icon': 'fas fa-check-circle'
                            }

                        # if ((today > sealed_end_date_time and today < english_start_date_time and 'dutch_winning_user' in asset_details and asset_details['dutch_winning_user'] and 'seal_winning_user' in asset_details and asset_details['seal_winning_user']) and (int(user_id) == int(asset_details['seal_winning_user']) or int(user_id) == int(asset_details['dutch_winning_user']))):
                        #     insider_bid_now_style = {
                        #         'id': 'insiderbidNowFormSection',
                        #         'style': 'display:block;'
                        #     }
                        #     insider_bid_inc_style = {
                        #         'id': 'englishAuctionBidIncSection',
                        #         'style': 'display:none;'
                        #     }
                        #     if my_bid_count > 0:
                        #         if high_bidder_id == int(user_id) and int(bid_count) > int(bid_count):
                        #             insider_bidder_msg_section_style = {
                        #                 'id': 'insider_bidder_msg_section',
                        #                 'style': 'display:inline-block;',
                        #                 'text': 'You Have The Highest Bid.',
                        #                 'class': 'badge-success',
                        #                 'icon': 'fas fa-check-circle'
                        #             }
                        #         else:
                        #             insider_bidder_msg_section_style = {
                        #                 'id': 'insider_bidder_msg_section',
                        #                 'style': 'display:inline-block;',
                        #                 'text': 'You Are Out Bidded.',
                        #                 'class': 'badge-danger',
                        #                 'icon': 'fas fa-exclamation-triangle'
                        #             }

                        if ((today > english_start_date_time and today < end_date_time and 'dutch_winning_user' in asset_details and asset_details['dutch_winning_user'] and 'seal_winning_user' in asset_details and asset_details['seal_winning_user']) and (int(user_id) == int(asset_details['seal_winning_user']) or int(user_id) == int(asset_details['dutch_winning_user']))):
                            insider_bid_now_style = {
                                'id': 'insiderbidNowFormSection',
                                'style': 'display:block;'
                            }
                            insider_bid_inc_style = {
                                'id': 'englishAuctionBidIncSection',
                                'style': 'display:block;'
                            }
                            if bid_count > 0:
                                if high_bidder_id == int(user_id) and int(bid_count) > int(total_bid_till_sealed):
                                    insider_bidder_msg_section_style = {
                                        'id': 'insider_bidder_msg_section',
                                        'style': 'display:inline-block;',
                                        'text': 'You Have The Highest Bid.',
                                        'class': 'badge-success',
                                        'icon': 'fas fa-check-circle'
                                    }
                                elif int(user_id) == int(asset_details['seal_winning_user']) and int(bid_count) == int(total_bid_till_sealed):
                                    insider_bidder_msg_section_style = {
                                        'id': 'insider_bidder_msg_section',
                                        'style': 'display:inline-block;',
                                        'text': 'You Have The Highest Bid.',
                                        'class': 'badge-success',
                                        'icon': 'fas fa-check-circle'
                                    }
                                else:
                                    insider_bidder_msg_section_style = {
                                        'id': 'insider_bidder_msg_section',
                                        'style': 'display:inline-block;',
                                        'text': 'You Are Out Bidded.',
                                        'class': 'badge-danger',
                                        'icon': 'fas fa-exclamation-triangle'
                                    }

                        if templete_dir == "theme-4":
                            bidder_msg_section_style = {
                                'id': 'bidder_msg_section',
                                'style': 'display:none;',
                                'text': 'You are registered and approved to bid. <strong>Place your first bid now.</strong>',
                                'class': 'badge-success',
                                'icon': 'fas fa-check-circle'
                            }
                        else:
                            bidder_msg_section_style = {
                                'id': 'bidder_msg_section',
                                'style': 'display:none;',
                                'text': 'You are registered and approved to bid. Place your first bid now.',
                                'class': 'badge-success',
                                'icon': 'fas fa-check-circle'
                            }

                        time_remaining_style = {
                            'class': 'time_remaining',
                            'style': 'display:block;'
                        }
                        countdown_text_style = {
                            'class': 'countdown_text',
                            'style': 'display:inline-block;'
                        }
                        pre_final_style = {
                            'id': 'preFinalBiddingSection',
                            'style': 'display:none;'
                        }
                        insider_auction_style = {
                            'id': 'insiderAuctionSection',
                            'style': 'display:inline-block;'
                        }
                        final_style = {
                            'id': 'finalBiddingSection',
                            'style': 'display:none;'
                        }
                        make_offer_btn_style = {
                            'id': 'make_offer_btn',
                            'style': 'display:none;'
                        }
                    elif today < start_date_time and is_approved == 2 and is_reviewed == True:
                        bidder_btn_txt = ''
                        reg_bid_btn_style = {
                            'id': 'registerBidBtn',
                            'style': 'display:none;cursor:initial;',
                            'btn_url': 'javascript:void(0)'
                        }
                        bidder_msg_section_style = {
                            'id': 'bidder_msg_section',
                            'style': 'display:none;',
                            'text': 'Registration Approved But Auction Not Started Yet.',
                            'class': 'badge-success',
                            'icon': 'fas fa-check-circle'
                        }
                        insider_bidder_btn_txt = 'Auction Has Not Started'
                        insider_reg_bid_btn_style = {
                            'id': 'insiderRegisterBidBtn',
                            'style': 'display:inline-block;cursor:initial;',
                            'btn_url': 'javascript:void(0)'
                        }
                        insider_bidder_msg_section_style = {
                            'id': 'insider_bidder_msg_section',
                            'style': 'display:block;',
                            'text': 'Registration Approved But Auction Not Started Yet.',
                            'class': 'badge-success',
                            'icon': 'fas fa-check-circle'
                        }

                        time_remaining_style = {
                            'class': 'time_remaining',
                            'style': 'display:block;'
                        }
                        countdown_text_style = {
                            'class': 'countdown_text',
                            'style': 'display:inline-block;'
                        }

                        bid_now_style = {
                            'id': 'bidNowFormSection',
                            'style': 'display:none;'
                        }
                        pre_final_style = {
                            'id': 'preFinalBiddingSection',
                            'style': 'display:none;'
                        }
                        insider_auction_style = {
                            'id': 'insiderAuctionSection',
                            'style': 'display:inline-block;'
                        }
                        final_style = {
                            'id': 'finalBiddingSection',
                            'style': 'display:none;'
                        }
                        make_offer_btn_style = {
                            'id': 'make_offer_btn',
                            'style': 'display:none;'
                        }
                    elif today > end_date_time:
                        bidder_btn_txt = ''
                        insider_bidder_btn_txt = ''
                        insider_auction_style = {
                            'id': 'insiderAuctionSection',
                            'style': 'display:inline-block;'
                        }
                        if int(user_id) == (high_bidder_id):
                            final_style = {
                                'id': 'finalBiddingSection',
                                'style': 'display:none;'
                            }
                            final_lost_style = {
                                'id': 'finalBiddingLostSection',
                                'style': 'display:none;'
                            }
                            insider_bid_won_style = {
                                'id': 'insiderBiddingWonSection',
                                'style': 'display:block;'
                            }
                        else:
                            final_style = {
                                'id': 'finalBiddingSection',
                                'style': 'display:none;'
                            }
                            final_lost_style = {
                                'id': 'finalBiddingLostSection',
                                'style': 'display:none;'
                            }
                            insider_lost_section_btn_style = {
                                'id': 'insiderLostSectionBtn',
                                'style': 'display:inline-block;',
                            }
                            # if int(my_bid_count) > 0:
                            #     if int(user_id) == (high_bidder_id):
                            #         insider_bidder_msg_section_style = {
                            #             'id': 'insider_bidder_msg_section',
                            #             'style': 'display:inline-block;',
                            #             'text': 'You Have Highest Bid.',
                            #             'class': 'badge-success',
                            #             'icon': 'fas fa-check-circle',
                            #         }
                            #     else:
                            #         insider_bidder_msg_section_style = {
                            #             'id': 'insider_bidder_msg_section',
                            #             'style': 'display:inline-block;',
                            #             'text': 'Out Bid.',
                            #             'class': 'badge-danger',
                            #             'icon': 'fas fa-exclamation-triangle ',
                            #         }
                            # else:
                            #     insider_bidder_msg_section_style = {
                            #         'id': 'insider_bidder_msg_section',
                            #         'style': 'display:inline-block;',
                            #         'text': '',
                            #         'class': 'badge-success',
                            #         'icon': '',
                            #     }


                        time_remaining_style = {
                            'class': 'time_remaining',
                            'style': 'display:block;'
                        }
                        countdown_text_style = {
                            'class': 'countdown_text',
                            'style': 'display:inline-block;'
                        }

                        bid_now_style = {
                            'id': 'bidNowFormSection',
                            'style': 'display:none;'
                        }
                    else:
                        bidder_btn_txt = ''
                        reg_bid_btn_style = {
                            'id': 'registerBidBtn',
                            'style': 'display:none;',
                            'btn_url': '/bid-registration/?property=' + str(asset_details['id'])
                        }
                        insider_bidder_btn_txt = 'Register For Auction'
                        if templete_dir == "theme-4":
                            if user_id == "":
                                insider_bidder_btn_txt = "Login/Register to Bid"
                            else:
                                insider_bidder_btn_txt = "Register to Bid"
                        insider_reg_bid_btn_style = {
                            'id': 'insiderRegisterBidBtn',
                            'style': 'display:inline-block;',
                            'btn_url': '/bid-registration/?property=' + str(asset_details['id'])
                        }
                        insider_dutch_bid_btn_style = {
                            'id': 'insiderDutchBidBtn',
                            'style': 'display:none;',
                        }

                        final_style = {
                            'id': 'finalBiddingSection',
                            'style': 'display:none;'
                        }

                        insider_auction_style = {
                            'id': 'insiderAuctionSection',
                            'style': 'display:inline-block;'
                        }
                        pre_final_style = {
                            'id': 'preFinalBiddingSection',
                            'style': 'display:none;'
                        }
                        make_offer_btn_style = {
                            'id': 'make_offer_btn',
                            'style': 'display:none;'
                        }

                        bidder_msg_section_style = {
                            'id': 'bidder_msg_section',
                            'style': 'display:none;',
                            'text': '',
                            'class': 'badge-success',
                            'icon': 'fas fa-check-circle'
                        }
                        insider_bidder_msg_section_style = {
                            'id': 'insider_bidder_msg_section',
                            'style': 'display:none;',
                            'text': '',
                            'class': 'badge-success',
                            'icon': 'fas fa-check-circle'
                        }

                        if ((today >= start_date_time and today <= end_date_time) or today < start_date_time):
                            time_remaining_style = {
                                'class': 'time_remaining',
                                'style': 'display:block;'
                            }
                            countdown_text_style = {
                                'class': 'countdown_text',
                                'style': 'display:inline-block;'
                            }
                        else:
                            time_remaining_style = {
                                'class': 'time_remaining',
                                'style': 'display:block;'
                            }
                            countdown_text_style = {
                                'class': 'countdown_text',
                                'style': 'display:inline-block;'
                            }

                        bid_now_style = {
                            'id': 'bidNowFormSection',
                            'style': 'display:none;'
                        }
                elif int(asset_details['property_status_id']) == 9:
                    bidder_btn_txt = ''
                    insider_bidder_btn_txt = ''
                    insider_auction_style = {
                        'id': 'insiderAuctionSection',
                        'style': 'display:inline-block;'
                    }
                    if int(user_id) == (high_bidder_id):

                        final_style = {
                            'id': 'finalBiddingSection',
                            'style': 'display:none;'
                        }
                        final_lost_style = {
                            'id': 'finalBiddingLostSection',
                            'style': 'display:none;'
                        }
                        insider_bid_won_style = {
                            'id': 'insiderBiddingWonSection',
                            'style': 'display:block;'
                        }
                    else:
                        final_style = {
                            'id': 'finalBiddingSection',
                            'style': 'display:none;'
                        }
                        final_lost_style = {
                            'id': 'finalBiddingLostSection',
                            'style': 'display:none;'
                        }
                        insider_lost_section_btn_style = {
                            'id': 'insiderLostSectionBtn',
                            'style': 'display:inline-block;',
                        }

                        insider_reg_bid_btn_style = {
                            'id': 'insiderRegisterBidBtn',
                            'style': 'display:none;',
                            'btn_url': 'javascript:void(0)'
                        }
                        insider_bidder_msg_section_style = {
                            'id': 'insider_bidder_msg_section',
                            'style': 'display:none;',
                            'text': '',
                            'class': 'badge-success',
                            'icon': '',
                        }
                        insider_dutch_bid_btn_style = {
                            'id': 'insiderDutchBidBtn',
                            'style': 'display:none;',
                        }
                        insider_lost_section_btn_style = {
                            'id': 'insiderLostSectionBtn',
                            'style': 'display:inline-block;',
                        }
                        insider_bid_count_style = {
                            'id': 'insiderBiddingLostSectionBidCount',
                            'style': 'display:block;'
                        }
                        # if int(my_bid_count) > 0:
                        #     if int(user_id) == (high_bidder_id):
                        #         insider_bidder_msg_section_style = {
                        #             'id': 'insider_bidder_msg_section',
                        #             'style': 'display:inline-block;',
                        #             'text': 'You Have Highest Bid.',
                        #             'class': 'badge-success',
                        #             'icon': 'fas fa-check-circle ',
                        #         }
                        #     else:
                        #         insider_bidder_msg_section_style = {
                        #             'id': 'insider_bidder_msg_section',
                        #             'style': 'display:inline-block;',
                        #             'text': 'Out Bid.',
                        #             'class': 'badge-danger',
                        #             'icon': 'fas fa-exclamation-triangle ',
                        #         }
                        # else:
                        #     insider_bidder_msg_section_style = {
                        #         'id': 'insider_bidder_msg_section',
                        #         'style': 'display:inline-block;',
                        #         'text': 'Out Bid.',
                        #         'class': 'badge-danger',
                        #         'icon': 'fas fa-exclamation-triangle ',
                        #     }
                    bidder_msg_section_style = {
                        'id': 'bidder_msg_section',
                        'style': 'display:none;',
                        'text': '',
                        'class': 'badge-success',
                        'icon': 'fas fa-check-circle'
                    }
                    pre_final_style = {
                        'id': 'preFinalBiddingSection',
                        'style': 'display:none;'
                    }
                    reg_bid_btn_style = {
                        'id': 'registerBidBtn',
                        'style': 'display:none;',
                        'btn_url': 'javascript:void(0)'
                    }
                    make_offer_btn_style = {
                        'id': 'make_offer_btn',
                        'style': 'display:none;'
                    }
                    time_remaining_style = {
                        'class': 'time_remaining',
                        'style': 'display:block;'
                    }
                    countdown_text_style = {
                        'class': 'countdown_text',
                        'style': 'display:inline-block;'
                    }
                    bid_now_style = {
                        'id': 'bidNowFormSection',
                        'style': 'display:none;'
                    }
                else:
                    bidder_btn_txt = ''
                    if int(my_bid_count) > 0:
                        if int(user_id) == (high_bidder_id):
                            bidder_lost_section_success_msg = '<i class="fas fa-check-circle"></i> You Have Highest Bid.'
                        else:
                            bidder_lost_section_success_msg = '<i class="fas fa-exclamation-triangle"></i> Out Bid.'
                    else:
                        bidder_lost_section_success_msg = ''
                    final_style = {
                        'id': 'finalBiddingSection',
                        'style': 'display:none;'
                    }
                    final_lost_style = {
                        'id': 'finalBiddingLostSection',
                        'style': 'display:none;'
                    }

                    pre_final_style = {
                        'id': 'preFinalBiddingSection',
                        'style': 'display:none;'
                    }
                    reg_bid_btn_style = {
                        'id': 'registerBidBtn',
                        'style': 'display:none;',
                        'btn_url': 'javascript:void(0)'
                    }
                    make_offer_btn_style = {
                        'id': 'make_offer_btn',
                        'style': 'display:none;'
                    }
                    # if int(my_bid_count) > 0:
                    #     bidder_msg_section_style = {
                    #         'id': 'bidder_msg_section',
                    #         'style': 'display:block;',
                    #         'text': '',
                    #         'class': 'badge-danger',
                    #         'icon': 'fas fa-check-circle'
                    #     }
                    # else:
                    bidder_msg_section_style = {
                        'id': 'bidder_msg_section',
                        'style': 'display:none;',
                        'text': '',
                        'class': 'badge-success',
                        'icon': 'fas fa-check-circle'
                    }
                    insider_auction_style = {
                        'id': 'insiderAuctionSection',
                        'style': 'display:inline-block;'
                    }
                    insider_reg_bid_btn_style = {
                        'id': 'insiderRegisterBidBtn',
                        'style': 'display:none;',
                        'btn_url': 'javascript:void(0)'
                    }
                    insider_bidder_msg_section_style = {
                        'id': 'insider_bidder_msg_section',
                        'style': 'display:none;',
                        'text': '',
                        'class': 'badge-success',
                        'icon': '',
                    }
                    insider_dutch_bid_btn_style = {
                        'id': 'insiderDutchBidBtn',
                        'style': 'display:none;',
                    }
                    insider_lost_section_btn_style = {
                        'id': 'insiderLostSectionBtn',
                        'style': 'display:inline-block;',
                    }
                    insider_bid_count_style = {
                        'id': 'insiderBiddingLostSectionBidCount',
                        'style': 'display:block;'
                    }

                    time_remaining_style = {
                        'class': 'time_remaining',
                        'style': 'display:block;'
                    }
                    countdown_text_style = {
                        'class': 'countdown_text',
                        'style': 'display:inline-block;'
                    }

                    bid_now_style = {
                        'id': 'bidNowFormSection',
                        'style': 'display:none;'
                    }

            elif int(sale_type) == 7:
                bidder_btn_txt = ''
                reg_bid_btn_style = {
                    'id': 'registerBidBtn',
                    'style': 'display:none;',
                    'btn_url': 'javascript:void(0)'
                }
                time_remaining_style = {
                    'class': 'time_remaining',
                    'style': 'display:block;'
                }
                countdown_text_style = {
                    'class': 'countdown_text',
                    'style': 'display:inline-block;'
                }

                bid_now_style = {
                    'id': 'bidNowFormSection',
                    'style': 'display:none;'
                }
                pre_final_style = {
                    'id': 'preFinalBiddingSection',
                    'style': 'display:none;'
                }
                final_style = {
                    'id': 'finalBiddingSection',
                    'style': 'display:none;'
                }
                make_offer_btn_style = {
                    'id': 'make_offer_btn',
                    'style': 'display:none;'
                }
                best_offer_style = {
                    'id': 'bestOfferSection',
                    'style': 'display:inline-block;'
                }
            elif int(sale_type) == 6:
                bidder_btn_txt = ''
                reg_bid_btn_style = {
                    'id': 'registerBidBtn',
                    'style': 'display:none;',
                    'btn_url': 'javascript:void(0)'
                }
                time_remaining_style = {
                    'class': 'time_remaining',
                    'style': 'display:none;'
                }
                countdown_text_style = {
                    'class': 'countdown_text',
                    'style': 'display:none;'
                }

                bid_now_style = {
                    'id': 'bidNowFormSection',
                    'style': 'display:none;'
                }
                pre_final_style = {
                    'id': 'preFinalBiddingSection',
                    'style': 'display:none;'
                }
                final_style = {
                    'id': 'finalBiddingSection',
                    'style': 'display:none;'
                }
                traditional_style = {
                    'id': 'TraditionalOfferSection',
                    'style': 'display:none;'
                }
                make_offer_btn_style = {
                    'id': 'make_offer_btn',
                    'style': 'display:none;'
                }
                live_auction_style = {
                    'id': 'liveAuctionSection',
                    'style': 'display:inline-block;'
                }
            else:
                bidder_btn_txt = ''
                reg_bid_btn_style = {
                    'id': 'registerBidBtn',
                    'style': 'display:none;',
                    'btn_url': 'javascript:void(0)'
                }
                time_remaining_style = {
                    'class': 'time_remaining',
                    'style': 'display:none;'
                }
                countdown_text_style = {
                    'class': 'countdown_text',
                    'style': 'display:none;'
                }

                bid_now_style = {
                    'id': 'bidNowFormSection',
                    'style': 'display:none;'
                }
                pre_final_style = {
                    'id': 'preFinalBiddingSection',
                    'style': 'display:none;'
                }
                final_style = {
                    'id': 'finalBiddingSection',
                    'style': 'display:none;'
                }
                traditional_style = {
                    'id': 'TraditionalOfferSection',
                    'style': 'display:inline-block;'
                }
                make_offer_btn_style = {
                    'id': 'make_offer_btn',
                    'style': 'display:inline-block;'
                }


        else:
            if int(sale_type) == 1:
                if 'property_status_id' in asset_details and int(asset_details['property_status_id']) == 1  and int(asset_details['auction_status']) == 1:
                    if today < start_date_time:
                        bidder_btn_txt = 'Register For Auction'
                        if templete_dir == "theme-4":
                            if user_id == "":
                                bidder_btn_txt = "Login/Register to Bid"
                            else:
                                bidder_btn_txt = "Register to Bid"
                        reg_bid_btn_style = {
                            'id': 'registerBidBtn',
                            'style': 'display:inline-block;cursor:pointer;',
                            'btn_url': '/login/?next=/asset-details/?property_id='+str(asset_details['id'])
                        }
                        time_remaining_style = {
                            'class': 'time_remaining',
                            'style': 'display:block;'
                        }
                        countdown_text_style = {
                            'class': 'countdown_text',
                            'style': 'display:inline-block;',
                        }

                        bid_now_style = {
                            'id': 'bidNowFormSection',
                            'style': 'display:none;'
                        }
                        pre_final_style = {
                            'id': 'preFinalBiddingSection',
                            'style': 'display:inline-block;'
                        }
                        final_style = {
                            'id': 'finalBiddingSection',
                            'style': 'display:none;'
                        }
                        make_offer_btn_style = {
                            'id': 'make_offer_btn',
                            'style': 'display:none;'
                        }
                    elif today >= start_date_time and today <= end_date_time:
                        bidder_btn_txt = 'Register For Auction'
                        if templete_dir == "theme-4":
                            if user_id == "":
                                bidder_btn_txt = "Login/Register to Bid"
                            else:
                                bidder_btn_txt = "Register to Bid"
                        reg_bid_btn_style = {
                            'id': 'registerBidBtn',
                            'style': 'display:inline-block;cursor:pointer;',
                            'btn_url': '/login/?next=/asset-details/?property_id=' + str(asset_details['id'])
                        }
                        time_remaining_style = {
                            'class': 'time_remaining',
                            'style': 'display:block;'
                        }
                        countdown_text_style = {
                            'class': 'countdown_text',
                            'style': 'display:inline-block;'
                        }

                        bid_now_style = {
                            'id': 'bidNowFormSection',
                            'style': 'display:none;'
                        }
                        pre_final_style = {
                            'id': 'preFinalBiddingSection',
                            'style': 'display:inline-block;'
                        }
                        final_style = {
                            'id': 'finalBiddingSection',
                            'style': 'display:none;'
                        }
                        make_offer_btn_style = {
                            'id': 'make_offer_btn',
                            'style': 'display:none;'
                        }
                    else:
                        bidder_btn_txt = 'Register For Auction'
                        if templete_dir == "theme-4":
                            if user_id == "":
                                bidder_btn_txt = "Login/Register to Bid"
                            else:
                                bidder_btn_txt = "Register to Bid"
                        reg_bid_btn_style = {
                            'id': 'registerBidBtn',
                            'style': 'display:inline-block;cursor:pointer;',
                            'btn_url': '/bid-registration/?property='+str(asset_details['id'])
                        }
                        time_remaining_style = {
                            'class': 'time_remaining',
                            'style': 'display:none;'
                        }
                        countdown_text_style = {
                            'class': 'countdown_text',
                            'style': 'display:none;'
                        }

                        bid_now_style = {
                            'id': 'bidNowFormSection',
                            'style': 'display:none;'
                        }
                        pre_final_style = {
                            'id': 'preFinalBiddingSection',
                            'style': 'display:inline-block;'
                        }
                        final_style = {
                            'id': 'finalBiddingSection',
                            'style': 'display:none;'
                        }
                        make_offer_btn_style = {
                            'id': 'make_offer_btn',
                            'style': 'display:none;'
                        }
                else:
                    bidder_btn_txt = ''
                    bidder_lost_section_success_msg = ''
                    final_style = {
                        'id': 'finalBiddingSection',
                        'style': 'display:none;'
                    }
                    final_lost_style = {
                        'id': 'finalBiddingLostSection',
                        'style': 'display:block;'
                    }

                    pre_final_style = {
                        'id': 'preFinalBiddingSection',
                        'style': 'display:none;'
                    }
                    reg_bid_btn_style = {
                        'id': 'registerBidBtn',
                        'style': 'display:none;',
                        'btn_url': 'javascript:void(0)'
                    }
                    make_offer_btn_style = {
                        'id': 'make_offer_btn',
                        'style': 'display:none;'
                    }

                    bidder_msg_section_style = {
                        'id': 'bidder_msg_section',
                        'style': 'display:none;',
                        'text': '',
                        'class': 'badge-success',
                        'icon': 'fas fa-check-circle'
                    }

                    time_remaining_style = {
                        'class': 'time_remaining',
                        'style': 'display:block;'
                    }
                    countdown_text_style = {
                        'class': 'countdown_text',
                        'style': 'display:inline-block;'
                    }

                    bid_now_style = {
                        'id': 'bidNowFormSection',
                        'style': 'display:none;'
                    }

            elif int(sale_type) == 2:
                if 'property_status_id' in asset_details and int(asset_details['property_status_id']) == 1  and int(asset_details['auction_status']) == 1:
                    if today < start_date_time:
                        bidder_btn_txt = ''
                        reg_bid_btn_style = {
                            'id': 'registerBidBtn',
                            'style': 'display:none;cursor:pointer;',
                            'btn_url': '/login/?next=/asset-details/?property_id='+str(asset_details['id'])
                        }
                        insider_bidder_btn_txt = 'Register For Auction'
                        if templete_dir == "theme-4":
                            if user_id == "":
                                insider_bidder_btn_txt = "Login/Register to Bid"
                            else:
                                insider_bidder_btn_txt = "Register to Bid"
                        insider_reg_bid_btn_style = {
                            'id': 'insiderRegisterBidBtn',
                            'style': 'display:inline-block;cursor:pointer;',
                            'btn_url': '/login/?next=/asset-details/?property_id=' + str(asset_details['id'])
                        }
                        time_remaining_style = {
                            'class': 'time_remaining',
                            'style': 'display:block;'
                        }
                        countdown_text_style = {
                            'class': 'countdown_text',
                            'style': 'display:inline-block;',
                        }

                        bid_now_style = {
                            'id': 'bidNowFormSection',
                            'style': 'display:none;'
                        }
                        insider_auction_style = {
                            'id': 'insiderAuctionSection',
                            'style': 'display:inline-block;'
                        }
                        pre_final_style = {
                            'id': 'preFinalBiddingSection',
                            'style': 'display:none;'
                        }
                        final_style = {
                            'id': 'finalBiddingSection',
                            'style': 'display:none;'
                        }
                        make_offer_btn_style = {
                            'id': 'make_offer_btn',
                            'style': 'display:none;'
                        }
                    elif today >= start_date_time and today <= end_date_time:
                        bidder_btn_txt = ''
                        reg_bid_btn_style = {
                            'id': 'registerBidBtn',
                            'style': 'display:none;cursor:pointer;',
                            'btn_url': '/login/?next=/asset-details/?property_id=' + str(asset_details['id'])
                        }
                        insider_bidder_btn_txt = 'Register For Auction'
                        if templete_dir == "theme-4":
                            if user_id == "":
                                insider_bidder_btn_txt = "Login/Register to Bid"
                            else:
                                insider_bidder_btn_txt = "Register to Bid"
                        insider_reg_bid_btn_style = {
                            'id': 'insiderRegisterBidBtn',
                            'style': 'display:inline-block;cursor:pointer;',
                            'btn_url': '/login/?next=/asset-details/?property_id=' + str(asset_details['id'])
                        }
                        time_remaining_style = {
                            'class': 'time_remaining',
                            'style': 'display:block;'
                        }
                        countdown_text_style = {
                            'class': 'countdown_text',
                            'style': 'display:inline-block;'
                        }

                        bid_now_style = {
                            'id': 'bidNowFormSection',
                            'style': 'display:none;'
                        }
                        insider_auction_style = {
                            'id': 'insiderAuctionSection',
                            'style': 'display:inline-block;'
                        }
                        pre_final_style = {
                            'id': 'preFinalBiddingSection',
                            'style': 'display:none;'
                        }
                        final_style = {
                            'id': 'finalBiddingSection',
                            'style': 'display:none;'
                        }
                        make_offer_btn_style = {
                            'id': 'make_offer_btn',
                            'style': 'display:none;'
                        }
                    else:
                        bidder_btn_txt = ''
                        reg_bid_btn_style = {
                            'id': 'registerBidBtn',
                            'style': 'display:none;cursor:pointer;',
                            'btn_url': '/bid-registration/?property='+str(asset_details['id'])
                        }
                        insider_bidder_btn_txt = 'Register For Auction'
                        if templete_dir == "theme-4":
                            if user_id == "":
                                insider_bidder_btn_txt = "Login/Register to Bid"
                            else:
                                insider_bidder_btn_txt = "Register to Bid"
                        insider_reg_bid_btn_style = {
                            'id': 'insiderRegisterBidBtn',
                            'style': 'display:inline-block;cursor:pointer;',
                            'btn_url': '/login/?next=/asset-details/?property_id=' + str(asset_details['id'])
                        }
                        time_remaining_style = {
                            'class': 'time_remaining',
                            'style': 'display:none;'
                        }
                        countdown_text_style = {
                            'class': 'countdown_text',
                            'style': 'display:none;'
                        }

                        bid_now_style = {
                            'id': 'bidNowFormSection',
                            'style': 'display:none;'
                        }
                        pre_final_style = {
                            'id': 'preFinalBiddingSection',
                            'style': 'display:none;'
                        }
                        insider_auction_style = {
                            'id': 'insiderAuctionSection',
                            'style': 'display:inline-block;'
                        }
                        final_style = {
                            'id': 'finalBiddingSection',
                            'style': 'display:none;'
                        }
                        make_offer_btn_style = {
                            'id': 'make_offer_btn',
                            'style': 'display:none;'
                        }
                else:
                    bidder_btn_txt = ''
                    bidder_lost_section_success_msg = ''
                    final_style = {
                        'id': 'finalBiddingSection',
                        'style': 'display:none;'
                    }
                    final_lost_style = {
                        'id': 'finalBiddingLostSection',
                        'style': 'display:none;'
                    }
                    insider_auction_style = {
                        'id': 'insiderAuctionSection',
                        'style': 'display:inline-block;'
                    }
                    insider_lost_section_btn_style = {
                        'id': 'insiderLostSectionBtn',
                        'style': 'display:inline-block;',
                    }
                    insider_bid_count_style = {
                        'id': 'insiderBiddingLostSectionBidCount',
                        'style': 'display:block;'
                    }

                    pre_final_style = {
                        'id': 'preFinalBiddingSection',
                        'style': 'display:none;'
                    }
                    reg_bid_btn_style = {
                        'id': 'registerBidBtn',
                        'style': 'display:none;',
                        'btn_url': 'javascript:void(0)'
                    }
                    make_offer_btn_style = {
                        'id': 'make_offer_btn',
                        'style': 'display:none;'
                    }

                    bidder_msg_section_style = {
                        'id': 'bidder_msg_section',
                        'style': 'display:none;',
                        'text': '',
                        'class': 'badge-success',
                        'icon': 'fas fa-check-circle'
                    }

                    time_remaining_style = {
                        'class': 'time_remaining',
                        'style': 'display:block;'
                    }
                    countdown_text_style = {
                        'class': 'countdown_text',
                        'style': 'display:inline-block;'
                    }

                    bid_now_style = {
                        'id': 'bidNowFormSection',
                        'style': 'display:none;'
                    }
            elif int(sale_type) == 7:
                bidder_btn_txt = 'Register For Auction'
                if templete_dir == "theme-4":
                    if user_id == "":
                        bidder_btn_txt = "Login/Register to Bid"
                    else:
                        bidder_btn_txt = "Register to Bid"
                reg_bid_btn_style = {
                    'id': 'registerBidBtn',
                    'style': 'display:none;',
                    'btn_url': '/login/?next=/asset-details/?property_id=' + str(asset_details['id'])
                }
                time_remaining_style = {
                    'class': 'time_remaining',
                    'style': 'display:block;'
                }
                countdown_text_style = {
                    'class': 'countdown_text',
                    'style': 'display:inline-block;'
                }

                bid_now_style = {
                    'id': 'bidNowFormSection',
                    'style': 'display:none;'
                }
                pre_final_style = {
                    'id': 'preFinalBiddingSection',
                    'style': 'display:none;'
                }
                final_style = {
                    'id': 'finalBiddingSection',
                    'style': 'display:none;'
                }
                make_offer_btn_style = {
                    'id': 'make_offer_btn',
                    'style': 'display:none;'
                }
                best_offer_style = {
                    'id': 'bestOfferSection',
                    'style': 'display:inline-block;'
                }
            elif int(sale_type) == 6:
                bidder_btn_txt = ''
                reg_bid_btn_style = {
                    'id': 'registerBidBtn',
                    'style': 'display:none;',
                    'btn_url': 'javascript:void(0)'
                }
                time_remaining_style = {
                    'class': 'time_remaining',
                    'style': 'display:none;'
                }
                countdown_text_style = {
                    'class': 'countdown_text',
                    'style': 'display:none;'
                }

                bid_now_style = {
                    'id': 'bidNowFormSection',
                    'style': 'display:none;'
                }
                pre_final_style = {
                    'id': 'preFinalBiddingSection',
                    'style': 'display:none;'
                }
                final_style = {
                    'id': 'finalBiddingSection',
                    'style': 'display:none;'
                }
                traditional_style = {
                    'id': 'TraditionalOfferSection',
                    'style': 'display:none;'
                }
                make_offer_btn_style = {
                    'id': 'make_offer_btn',
                    'style': 'display:none;'
                }
                live_auction_style = {
                    'id': 'liveAuctionSection',
                    'style': 'display:inline-block;'
                }
            else:
                bidder_btn_txt = 'Register For Auction'
                if templete_dir == "theme-4":
                    if user_id == "":
                        bidder_btn_txt = "Login/Register to Bid"
                    else:
                        bidder_btn_txt = "Register to Bid"
                reg_bid_btn_style = {
                    'id': 'registerBidBtn',
                    'style': 'display:none;',
                    'btn_url': '/login/?next=/asset-details/?property_id=' + str(asset_details['id'])
                }
                time_remaining_style = {
                    'class': 'time_remaining',
                    'style': 'display:none;'
                }
                countdown_text_style = {
                    'class': 'countdown_text',
                    'style': 'display:none;'
                }

                bid_now_style = {
                    'id': 'bidNowFormSection',
                    'style': 'display:none;'
                }
                pre_final_style = {
                    'id': 'preFinalBiddingSection',
                    'style': 'display:none;'
                }
                final_style = {
                    'id': 'finalBiddingSection',
                    'style': 'display:none;'
                }
                traditional_style = {
                    'id': 'TraditionalOfferSection',
                    'style': 'display:inline-block;'
                }
                make_offer_btn_style = {
                    'id': 'make_offer_btn',
                    'style': 'display:inline-block;'
                }

        if int(sale_type) == 1:
            if today >= start_date_time and today <= end_date_time:
                start_bid_style = {
                    'id': 'starting_bid',
                    'style': 'display:block;'
                }
                bid_inc_style = {
                    'id': 'bid_increment',
                    'style': 'display:block;'
                }
                bid_cnt_style = {
                    'id': 'bid_count',
                    'style': 'display:block;'
                }
            elif today < start_date_time:
                start_bid_style = {
                    'id': 'starting_bid',
                    'style': 'display:none;'
                }
                bid_inc_style = {
                    'id': 'bid_increment',
                    'style': 'display:block;'
                }
                bid_cnt_style = {
                    'id': 'bid_count',
                    'style': 'display:block;'
                }
            else:
                start_bid_style = {
                    'id': 'starting_bid',
                    'style': 'display:none;'
                }
                bid_inc_style = {
                    'id': 'bid_increment',
                    'style': 'display:none;'
                }
                bid_cnt_style = {
                    'id': 'bid_count',
                    'style': 'display:none;'
                }
        elif int(sale_type) == 4:
            start_bid_style = {
                'id': 'starting_bid',
                'style': 'display:none;'
            }
            bid_inc_style = {
                'id': 'bid_increment',
                'style': 'display:none;'
            }
            bid_cnt_style = {
                'id': 'bid_count',
                'style': 'display:none;'
            }
        else:
            start_bid_style = {
                'id': 'starting_bid',
                'style': 'display:none;'
            }
            bid_inc_style = {
                'id': 'bid_increment',
                'style': 'display:none;'
            }
            bid_cnt_style = {
                'id': 'bid_count',
                'style': 'display:none;'
            }

        x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
        if x_forwarded_for:
            ip = x_forwarded_for.split(',')[0]
        else:
            ip = request.META.get('REMOTE_ADDR')

        if 'property_doc' in asset_details and len(asset_details['property_doc']) > 0:
            for doc in asset_details['property_doc']:
                document_name = doc['doc_file_name']
                original_doc_arr = document_name.split('_')
                doc_name_length = len(original_doc_arr)
                #original_doc_name = document_name.split('_')[1]
                original_doc_name = ''
                for i in range(doc_name_length):
                    if i > 0:
                        original_doc_name = original_doc_name+'_'+original_doc_arr[i]
                original_doc_name = original_doc_name.lstrip('_')
                doc['original_doc_name'] = original_doc_name

        try:
            country_param = {}
            country_api_url = settings.API_URL + '/api-settings/get-country/'
            country_data = call_api_post_method(country_param, country_api_url)
            country_list = country_data['data']
        except:
            country_list = []
        context = {
            'is_home_page': False,
            'assets': asset_details,
            'node_url': settings.NODE_URL,
            'site_id': site_id,
            'start_bid_style': start_bid_style,
            'bid_inc_style': bid_inc_style,
            'bid_cnt_style': bid_cnt_style,
            'bid_now_style': bid_now_style,
            'pre_final_style': pre_final_style,
            'final_style': final_style,
            'reg_bid_btn_style': reg_bid_btn_style,
            'bidder_btn_txt': bidder_btn_txt,
            'bidder_msg_section_style': bidder_msg_section_style,
            'reserve_text_style': reserve_text_style,
            'time_remaining_style': time_remaining_style,
            'countdown_text_style': countdown_text_style,
            'make_offer_btn_style': make_offer_btn_style,
            'final_lost_style': final_lost_style,
            'ip_address': ip,
            'flash_timer': flash_timer,
            'hours_remain': hours_remain,
            'bidder_lost_section_success_msg': bidder_lost_section_success_msg,
            'traditional_style': traditional_style,
            'live_auction_style': live_auction_style,
            'templete_dir': templete_dir,
            'DOMAIN_NAME_URL': settings.DOMAIN_NAME_URL,
            'best_offer_style': best_offer_style,
            'current_time': time.time(),
            'state_list': state_list,
            'insider_auction_style': insider_auction_style,
            'insider_reg_bid_btn_style': insider_reg_bid_btn_style,
            'insider_bidder_msg_section_style': insider_bidder_msg_section_style,
            'insider_bidder_btn_txt': insider_bidder_btn_txt,
            'insider_dutch_bid_btn_style': insider_dutch_bid_btn_style,
            'insider_lost_section_btn_style': insider_lost_section_btn_style,
            'insider_bid_now_style': insider_bid_now_style,
            'insider_bid_inc_style': insider_bid_inc_style,
            'insider_bid_won_style': insider_bid_won_style,
            'insider_bid_count_style': insider_bid_count_style,
            'country_list': country_list,
            'my_bid_count': my_bid_count
        }
        theme_path = 'home/{}/listings/asset-details.html'.format(templete_dir)
        return render(request, theme_path, context)
    except Exception as exp:
        print(exp)
        return HttpResponse("Issue in views")


@csrf_exempt
def our_listing(request):
    """
        Method : GET/POST
        Description : Get property listing page
        Url : /our-listing/
        Params:
        :param 1: page :: integer
        :param 1: page_size :: integer
        :param 1: listing_filter :: String
        :param 1: sort_column :: String
        :param 1: search :: String
    """
    try:
        try:
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']
            templete_dir = site_detail['site_detail']['theme_directory']

        except Exception as exp:
            site_id = ""
            templete_dir = 'theme-1'

        static_dir = templete_dir

        token = None
        user_id = None
        if 'user_id' in request.session and request.session['user_id']:
            user_id = request.session['user_id']
            token = request.session['token']['access_token']

        if request.is_ajax() and request.method == 'POST':
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']
            
            search = ''
            if 'search' in request.POST and request.POST['search']:
                search = request.POST['search']

            page = 1
            if 'page' in request.POST and request.POST['page']:
                page = int(request.POST['page'])

            page_size = 10
            if 'perpage' in request.POST and request.POST['perpage']:
                page_size = int(request.POST['perpage'])

            listing_filter = ''
            if 'listing_filter' in request.POST and request.POST['listing_filter'] != "":
                listing_filter = request.POST['listing_filter']

            agent_id = ''
            if 'agent_id' in request.POST and request.POST['agent_id']:
                agent_id = int(request.POST['agent_id'])

            filter_asset_type = ''
            if 'filter_asset_type' in request.POST and request.POST['filter_asset_type'] != "":
                filter_asset_type = request.POST['filter_asset_type']

            filter_property_type = ''
            if 'filter_property_type' in request.POST and request.POST['filter_property_type'] != "":
                filter_property_type = request.POST['filter_property_type']    

            filter_auction_type = ''
            if 'filter_auction_type' in request.POST and request.POST['filter_auction_type'] != "":
                filter_auction_type = request.POST['filter_auction_type']

            filter_beds = ''
            if 'filter_beds' in request.POST and request.POST['filter_beds'] != "":
                filter_beds = request.POST['filter_beds']

            filter_baths = ''
            if 'filter_baths' in request.POST and request.POST['filter_baths'] != "":
                filter_baths = request.POST['filter_baths']

            filter_mls_property = ''
            if 'filter_mls_property' in request.POST and request.POST['filter_mls_property'] != "":
                filter_mls_property = request.POST['filter_mls_property']

            filter_min_price = ''
            if 'filter_min_price' in request.POST and request.POST['filter_min_price'] != "":
                filter_min_price = request.POST['filter_min_price']

            filter_max_price = ''
            if 'filter_max_price' in request.POST and request.POST['filter_max_price'] != "":
                filter_max_price = request.POST['filter_max_price']

            sort_column = ''
            short_by = ''
            sort_order = ''
            if 'sort_column' in request.POST and request.POST['sort_column'] != "":
                sort_column = request.POST['sort_column']
                if sort_column == 'auction_start_date_asc':
                    short_by = 'auction_start'
                    sort_order = 'asc'
                elif sort_column == 'auction_start_date_desc':
                    short_by = 'auction_start'
                    sort_order = 'desc'
                elif sort_column == 'auction_end_date_asc':
                    short_by = 'auction_end'
                    sort_order = 'asc'
                elif sort_column == 'auction_end_date_desc':
                    short_by = 'auction_end'
                    sort_order = 'desc'
                else:
                    short_by = sort_column
                    sort_order = ''

            listing_params = {
                "site_id": site_id,
                "page_size": page_size,
                "page": page,
                "search": search,
                "filter": listing_filter,
                "short_by": short_by,
                "agent_id": agent_id,
                "filter_asset_type": filter_asset_type,
                "filter_property_type": filter_property_type,
                "filter_auction_type": filter_auction_type,
                "filter_beds": filter_beds,
                "filter_baths": filter_baths,
                "filter_mls_property": filter_mls_property,
                "filter_min_price": filter_min_price,
                "filter_max_price": filter_max_price
            }

            if sort_order !="":
                listing_params['sort_order'] = sort_order
            if 'user_id' in request.session and request.session['user_id']:
                listing_params['user_id'] = user_id

            listing_url = settings.API_URL + '/api-property/front-property-listing/'

            listing_data = call_api_post_method(listing_params, listing_url, token)

            http_host = request.META['HTTP_HOST']
            site_url = settings.URL_SCHEME + str(http_host)

            if 'error' in listing_data and listing_data['error'] == 0:
                property_list = listing_data['data']['data']
                property_total = listing_data['data']['total']
                total_page = math.ceil(property_total / page_size)
                property_list_data_path = 'home/{}/listings/our_listing_content.html'.format(templete_dir)

                listing_template = get_template(property_list_data_path)
                listing_html = listing_template.render({'property_list': property_list, 'total': property_total, 'aws_url': settings.AWS_URL, 'listing_filter': listing_filter, 'sort_column': sort_column, 'sess_user_id': user_id, 'site_url': site_url,'page': page, 'no_page': total_page})

                pagination_path = 'home/{}/listings/pagination.html'.format(templete_dir)
                pagination_template = get_template(pagination_path)

                pagination_data = {"no_page": total_page, "total_page": range(total_page), "current_page": page,
                                   "pagination_id": "prop_listing_pagination_list"}
                pagination_html = pagination_template.render(pagination_data)

                data = {'listing_html': listing_html, 'property_list': property_list, 'status': 200, 'msg': 'Successfully received', 'error': 0, 'total': property_total, 'listing_filter': listing_filter, 'pagination_html': pagination_html,'no_page': total_page}
            else:
                data = {'status': 403, 'msg': 'Server error, Please try again', 'property_list': [], 'error': 1, 'total': 0}     
            
            return JsonResponse(data)
        else:
            page = 1
            page_size = 12
            search = ''
            short_by = 'ending_soonest'
            sort_order = 'asc'
            sort_column = 'auction_start_date_asc'
            listing_filter = request.GET.get('listing_type', '')
            agent_id = request.GET.get('agent_id', '')

            listing_params = {
                "site_id": site_id,
                "page_size": page_size,
                "page": page,
                "search": search,
                "filter": listing_filter,
                "short_by": short_by,
                "sort_order": sort_order,
                "agent_id": agent_id,

            }
            if 'user_id' in request.session and request.session['user_id']:
                listing_params['user_id'] = user_id
            listing_url = settings.API_URL + '/api-property/front-property-listing/'
            listing_data = call_api_post_method(listing_params, listing_url, token)
            
            if 'error' in listing_data and listing_data['error'] == 0:
                property_list = listing_data['data']['data']
                property_total = listing_data['data']['total']
            else:
                property_list = []
                property_total = 0

            # ---------Property Type---------
            api_url = settings.API_URL + "/api-property/property-type/"
            params = {
                "asset_id": "",
            }
            api_response = call_api_post_method(params, api_url, token)
            if 'error' in api_response and api_response['error'] == 0:
                property_type = api_response['data']
            else:
                property_type = []    

            # ---------------Pagination--------
            pagination_path = 'home/{}/listings/pagination.html'.format(templete_dir)
            pagination_template = get_template(pagination_path)
            total_page = math.ceil(property_total / page_size)
            pagination_data = {"no_page": total_page, "total_page": range(total_page), "current_page": page,
                               "pagination_id": "prop_listing_pagination_list"}
            pagination_html = pagination_template.render(pagination_data)

            http_host = request.META['HTTP_HOST']
            site_url = settings.URL_SCHEME + str(http_host)
            context = {
                'is_home_page': False,
                'data': 'Listing comming soon',
                'property_list': property_list,
                'total': property_total,
                'aws_url': settings.AWS_URL,
                'listing_filter': listing_filter,
                'sort_column': sort_column,
                'sess_user_id': user_id,
                'site_url': site_url,
                'pagination_html': pagination_html,
                'node_url': settings.NODE_URL,
                'site_id': site_id,
                'page': page,
                'no_page': total_page,
                'property_type': property_type,
            }
            theme_path = 'home/{}/listings/our-listing.html'.format(templete_dir)
            return render(request, theme_path, context)
    except Exception as exp:
        print(exp)
        return HttpResponse("Issue in views")

def about_us(request):
    """
            Method : GET/POST
            Description : About us page
            Url : /about-us/
        """
    try:
        try:
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']
            templete_dir = site_detail['site_detail']['theme_directory']
        except Exception as exp:
            site_id = ""
            templete_dir = 'theme-1'

        static_dir = templete_dir

        try:
            about_params = {
                'site_id': site_id
            }
            about_url = settings.API_URL + '/api-cms/about-detail/'
            response_data = call_api_post_method(about_params, about_url)
            if 'error' in response_data and response_data['error'] == 0:
                about_data = response_data['data']
            else:
                about_data = []
        except Exception as exp:
            print(exp)
            about_data = []
        
        user_id = None
        if 'user_id' in request.session and request.session['user_id']:
            user_id = request.session['user_id']
    
        http_host = request.META['HTTP_HOST']
        site_url = settings.URL_SCHEME + str(http_host)   

        context = {'is_home_page': True, 'data': about_data, 'aws_url': settings.AWS_URL, 'sess_user_id': user_id, 'SITE_URL': site_url}
        theme_path = 'home/{}/cms/about-us.html'.format(templete_dir)
        return render(request, theme_path, context)
    except Exception as exp:
        print(exp)
        return HttpResponse("Issue in views")    

@csrf_exempt
def contact_us(request):
    """
                Method : GET/POST
                Description : Contact us page
                Url : /contact-us/
            """
    try:
        try:
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']
            templete_dir = site_detail['site_detail']['theme_directory']
        except Exception as exp:
            site_id = ""
            templete_dir = 'theme-1'

        static_dir = templete_dir

        user_id = None
        if 'user_id' in request.session and request.session['user_id']:
            user_id = request.session['user_id']

        http_host = request.META['HTTP_HOST']
        site_url = settings.URL_SCHEME + str(http_host)    

        try:
            if request.is_ajax() and request.method == 'POST':
                if 'first_name' in request.POST and request.POST['first_name']:
                    first_name = request.POST['first_name']
                
                if 'email' in request.POST and request.POST['email']:
                    email = request.POST['email']

                if 'phone_no' in request.POST and request.POST['phone_no']:
                    phone_no = re.sub('\D', '', request.POST['phone_no'])
                
                if 'user_type' in request.POST and request.POST['user_type']:
                    user_type = request.POST['user_type']

                if 'message' in request.POST and request.POST['message']:
                    message = request.POST['message']            
                
                save_contact_params = {
                    'site_id': site_id,
                    'first_name': first_name,
                    'email': email,
                    'phone_no': phone_no,
                    'user_type': user_type,
                    'message': message
                }
                save_contact_url = settings.API_URL + '/api-cms/save-contact/'
                save_response = call_api_post_method(save_contact_params, save_contact_url)
                return JsonResponse(save_response, safe=False)     
            
        except Exception as exp:
            print(exp)

        try:
            contact_params = {
                'site_id': site_id
            }
            contact_url = settings.API_URL + '/api-cms/contact-detail/'
            response_data = call_api_post_method(contact_params, contact_url)
            if 'error' in response_data and response_data['error'] == 0:
                contact_data = response_data['data']
                try:
                    address_list = contact_data['profile']['address']
                except:
                    address_list = []

                if address_list:
                    try:
                        for addr in contact_data['profile']['address']:
                            try:
                                address_first = addr['address_first']
                                state = addr['state']
                                postal_code = addr['postal_code']
                                name = '{}, {}, {}'.format(address_first, state, postal_code)
                                address = name
                                geolocator = Nominatim(user_agent=settings.GEOLOCATOR_EMAIL, timeout=10)
                                location = geolocator.geocode(address)
                                addr['name'] = address
                                addr['latitude'] = location.latitude
                                addr['longitude'] = location.longitude
                            except Exception as exp:
                                print(exp)
                                addr['name'] = ''
                                addr['latitude'] = 0
                                addr['longitude'] = 0
                    except:
                        pass
                    pass
            else:
                contact_data = []
        except Exception as exp:
            print(exp)
            contact_data = []

        address_list = json.dumps(address_list)

        context = {'is_home_page': True, 'data': contact_data, 'address_list': address_list, 'sess_user_id': user_id, 'SITE_URL': site_url}
        theme_path = 'home/{}/cms/contact-us.html'.format(templete_dir)
        return render(request, theme_path, context)
    except Exception as exp:
        print(exp)
        return HttpResponse("Issue in views")

@csrf_exempt
def our_listing_map(request):
    """
            Method : GET/POST
            Description : Get property listing map page
            Url : /our-listing-map/
            Params:
            :param 1: page :: integer
            :param 1: page_size :: integer
            :param 1: listing_filter :: String
            :param 1: sort_column :: String
            :param 1: search :: String
        """
    try:
        try:
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']
            templete_dir = site_detail['site_detail']['theme_directory']
        except Exception as exp:
            site_id = ""
            templete_dir = 'theme-1'

        static_dir = templete_dir

        token = None
        user_id = None
        if 'user_id' in request.session and request.session['user_id']:
            user_id = request.session['user_id']
            token = request.session['token']['access_token']

        if request.is_ajax() and request.method == 'POST':
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']

            search = ''
            if 'search' in request.POST and request.POST['search']:
                search = request.POST['search']

            page = 1
            if 'page' in request.POST and request.POST['page']:
                page = int(request.POST['page'])


            page_size = 10
            if 'perpage' in request.POST and request.POST['perpage']:
                page_size = int(request.POST['perpage'])

            listing_filter = ''
            if 'listing_filter' in request.POST and request.POST['listing_filter'] != "":
                listing_filter = request.POST['listing_filter']

            filter_asset_type = ''
            if 'filter_asset_type' in request.POST and request.POST['filter_asset_type'] != "":
                filter_asset_type = request.POST['filter_asset_type']

            filter_auction_type = ''
            if 'filter_auction_type' in request.POST and request.POST['filter_auction_type'] != "":
                filter_auction_type = request.POST['filter_auction_type']

            filter_beds = ''
            if 'filter_beds' in request.POST and request.POST['filter_beds'] != "":
                filter_beds = request.POST['filter_beds']

            filter_baths = ''
            if 'filter_baths' in request.POST and request.POST['filter_baths'] != "":
                filter_baths = request.POST['filter_baths']

            filter_mls_property = ''
            if 'filter_mls_property' in request.POST and request.POST['filter_mls_property'] != "":
                filter_mls_property = request.POST['filter_mls_property']

            filter_min_price = ''
            if 'filter_min_price' in request.POST and request.POST['filter_min_price'] != "":
                filter_min_price = request.POST['filter_min_price']

            filter_max_price = ''
            if 'filter_max_price' in request.POST and request.POST['filter_max_price'] != "":
                filter_max_price = request.POST['filter_max_price']

            sort_column = ''
            short_by = ''
            sort_order = ''
            if 'sort_column' in request.POST and request.POST['sort_column'] != "":
                sort_column = request.POST['sort_column']
                if sort_column == 'auction_start_date_asc':
                    short_by = 'auction_start'
                    sort_order = 'asc'
                elif sort_column == 'auction_start_date_desc':
                    short_by = 'auction_start'
                    sort_order = 'desc'
                elif sort_column == 'auction_end_date_asc':
                    short_by = 'auction_end'
                    sort_order = 'asc'
                elif sort_column == 'auction_end_date_desc':
                    short_by = 'auction_end'
                    sort_order = 'desc'
                else:
                    short_by = sort_column
                    sort_order = ''

            listing_params = {
                "site_id": site_id,
                "page_size": page_size,
                "page": page,
                "search": search,
                "filter": listing_filter,
                "short_by": short_by,
                "filter_asset_type": filter_asset_type,
                "filter_auction_type": filter_auction_type,
                "filter_beds": filter_beds,
                "filter_baths": filter_baths,
                "filter_mls_property": filter_mls_property,
                "filter_min_price": filter_min_price,
                "filter_max_price": filter_max_price
            }
            if sort_order !="":
                listing_params['sort_order'] = sort_order
            if 'user_id' in request.session and request.session['user_id']:
                listing_params['user_id'] = user_id

            listing_url = settings.API_URL + '/api-property/front-property-listing/'
            listing_data = call_api_post_method(listing_params, listing_url, token)

            http_host = request.META['HTTP_HOST']
            site_url = settings.URL_SCHEME + str(http_host)

            if 'error' in listing_data and listing_data['error'] == 0:
                property_list = listing_data['data']['data']
                try:
                    for property in property_list:
                        try:
                            property_address = property['name']
                            geolocator = Nominatim(user_agent=settings.GEOLOCATOR_EMAIL, timeout=10)
                            location = geolocator.geocode(property_address)
                            property['latitude'] = location.latitude
                            property['longitude'] = location.longitude
                        except:
                            property['latitude'] = 0
                            property['longitude'] = 0
                except:
                    pass

                property_list_json = json.dumps(property_list)

                property_total = listing_data['data']['total']
                total_page = math.ceil(property_total / page_size)
                property_list_data_path = 'home/{}/listings/our-listing-map-content.html'.format(templete_dir)

                listing_template = get_template(property_list_data_path)


                listing_html = listing_template.render(
                    {'property_list': property_list, 'total': property_total, 'aws_url': settings.AWS_URL,
                     'listing_filter': listing_filter, 'sort_column': sort_column, 'property_list_json': property_list_json, 'sess_user_id': user_id, 'site_url': site_url,'page': page, 'no_page': total_page})

                pagination_path = 'home/{}/listings/map-pagination.html'.format(templete_dir)
                pagination_template = get_template(pagination_path)

                pagination_data = {"no_page": total_page, "total_page": range(total_page), "current_page": page,
                                   "pagination_id": "map_prop_listing_pagination_list"}
                pagination_html = pagination_template.render(pagination_data)

                data = {'listing_html': listing_html, 'property_list': property_list, 'status': 200,
                        'msg': 'Successfully received', 'error': 0, 'total': property_total,
                        'listing_filter': listing_filter, 'pagination_html': pagination_html, "page": page,"no_page": total_page,}
            else:
                data = {'status': 403, 'msg': 'Server error, Please try again', 'property_list': [], 'error': 1,
                        'total': 0}

            return JsonResponse(data)
        else:
            search = ''
            page = 1
            page_size = 10
            sort_column = 'auction_start_date_asc'
            short_by = 'ending_soonest'
            sort_order = 'asc'
            listing_filter = request.GET.get('listing_type', '')
            agent_id = request.GET.get('agent_id', '')
            listing_params = {
                "site_id": site_id,
                "page_size": page_size,
                "page": page,
                "search": search,
                "filter": listing_filter,
                "short_by": short_by,
                "sort_order": sort_order,
            }

            if 'user_id' in request.session and request.session['user_id']:
                listing_params['user_id'] = user_id

            listing_url = settings.API_URL + '/api-property/front-property-listing/'
            listing_data = call_api_post_method(listing_params, listing_url, token)

            if 'error' in listing_data and listing_data['error'] == 0:
                property_list = listing_data['data']['data']
                try:
                    for property in property_list:
                        try:
                            property_address = property['name']
                            geolocator = Nominatim(user_agent=settings.GEOLOCATOR_EMAIL, timeout=10)
                            location = geolocator.geocode(property_address)
                            property['latitude'] = location.latitude
                            property['longitude'] = location.longitude
                        except:
                            property['latitude'] = 0
                            property['longitude'] = 0
                except:
                    pass

                property_list_json = json.dumps(property_list)

                property_total = listing_data['data']['total']
            else:
                property_list = []
                property_total = 0
                property_list_json = ''

            pagination_path = 'home/{}/listings/map-pagination.html'.format(templete_dir)
            pagination_template = get_template(pagination_path)
            total_page = math.ceil(property_total / page_size)
            pagination_data = {"no_page": total_page, "total_page": range(total_page), "current_page": page,
                               "pagination_id": "map_prop_listing_pagination_list"}
            pagination_html = pagination_template.render(pagination_data)
        http_host = request.META['HTTP_HOST']
        site_url = settings.URL_SCHEME + str(http_host)
        context = {
            'is_home_page': False,
            'data': 'Listing comming soon',
            'property_list': property_list,
            'total': property_total,
            'aws_url': settings.AWS_URL,
            'listing_filter': listing_filter,
            'sort_column': sort_column,
            'property_list_json': property_list_json,
            'sess_user_id': user_id,
            'site_url': site_url,
            'pagination_html': pagination_html,
            'node_url': settings.NODE_URL,
            'site_id': site_id,
            "page": page,
            "no_page": total_page,
        }
        theme_path = 'home/{}/listings/our-listing-map.html'.format(templete_dir)
        return render(request, theme_path, context)
    except Exception as exp:
        print(exp)
        return HttpResponse("Issue in views")

@csrf_exempt
def front_listing_search_suggestion(request):
    """
                Method : GET/POST
                Description : Get Property listing Search input suggestion
                Url : /front-listing-search-suggestion/
                Params:
                :param 1: search :: String
            """
    try:
        if request.is_ajax() and request.method == 'POST':
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']
            token = None
            if 'user_id' in request.session:
                token = request.session['token']['access_token']

            params = {
                'site_id': site_id,
                'search': request.POST['search']
            }

            api_url = settings.API_URL + '/api-property/front-property-suggestion/'

            suggestion_data = call_api_post_method(params, api_url, token)

            if 'error' in suggestion_data and suggestion_data['error'] == 0:
                data = {'status': 200, 'suggestion_list': suggestion_data['data'], 'error': 0}
            else:
                data = {'status': 403, 'suggestion_list': [], 'error': 1}
        else:
            data = {'status': 403, 'suggestion_list': [], 'error': 1}

        return JsonResponse(data)
    except Exception as exp:
        print(exp)
        data = {'status': 403, 'suggestion_list': [], 'error': 1}
        return JsonResponse(data)

@csrf_exempt
def make_favourite_listing(request):
    """
                    Method : GET/POST
                    Description : Make listing favourite
                    Url : /make-favourite-listing/
                    Params:
                    :param 1: property :: Integer
                """
    try:
        if request.is_ajax() and request.method == 'POST':
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']
            token = request.session['token']['access_token']
            user_id = request.session['user_id']

            params = {
                "domain": site_id,
                "property": request.POST['property'],
                "user": user_id
            }

            api_url = settings.API_URL + '/api-property/make-favourite-property/'

            data = call_api_post_method(params, api_url, token)
            if 'error' in data and data['error'] == 0:
                data = {'status': 200, 'error': 0, 'msg': 'Added to favorite listing.'}
            else:
                data = {'status': 403, 'error': 1, 'msg': data['msg']}
        else:
            data = {'status': 403, 'error': 1, 'msg': 'Forbidden.'}

        return JsonResponse(data)
    except Exception as exp:
        print(exp)
        data = {'status': 403, 'error': 1, 'msg': 'Forbidden.'}
        return JsonResponse(data)

@csrf_exempt
def favourite_listings(request):
    """
                        Method : GET/POST
                        Description : User Dashboard Favourite Listing page
                        Url : /make-favourite-listing/
                        Params:
                        :param 1: search :: String
                        :param 2: page :: Integer
                        :param 3: page_size :: Integer
                    """
    try:
        try:
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']
            templete_dir = site_detail['site_detail']['theme_directory']

        except Exception as exp:
            print(exp)
            site_id = ""
            templete_dir = 'theme-1'

        static_dir = templete_dir

        token = request.session['token']['access_token']
        user_id = request.session['user_id']
        page_size = 10
        if request.is_ajax() and request.method == 'POST':
            search = ''
            if 'search' in request.POST and request.POST['search']:
                search = request.POST['search']
            page = 1
            if 'page' in request.POST and request.POST['page'] != "":
                page = request.POST['page']

            if 'page_size' in request.POST and request.POST['page_size'] != "":
                page_size = request.POST['page_size']
            params = {
                'domain': site_id,
                'user': user_id,
                'page_size': page_size,
                'search': search,
                'page': page
            }
            sno = (int(page) - 1) * int(page_size) + 1
            api_url = settings.API_URL + '/api-property/favourite-property-listing/'
            favorite_data = call_api_post_method(params, api_url, token=token)
            if 'error' in favorite_data and favorite_data['error'] == 0:
                favorite_listing = favorite_data['data']['data']
                total = favorite_data['data']['total']
            else:
                favorite_listing = []
                total = 0
            context = {'favorite_listing': favorite_listing, 'total': total, "active_menu": "favourite_listing", "aws_url": settings.AWS_URL, 'sno': sno}

            favorite_listing_path = 'home/{}/user-dashboard/favourite/favorite-listing-content.html'.format(templete_dir)
            favorite_listing_template = get_template(favorite_listing_path)
            favorite_listing_html = favorite_listing_template.render(context)
            # ---------------Pagination--------
            pagination_path = 'home/{}/user-dashboard/favourite/favourite-pagination.html'.format(templete_dir)
            pagination_template = get_template(pagination_path)
            total_page = math.ceil(int(total) / int(page_size))
            pagination_data = {"no_page": int(total_page), "total_page": range(total_page), "current_page": int(page)}
            pagination_html = pagination_template.render(pagination_data)
            data = {'favorite_listing_html': favorite_listing_html, 'status': 200, 'msg': '', 'error': 0, 'total': total, "pagination_html": pagination_html}
            return JsonResponse(data)
        else:
            page = 1
            params = {
                'domain': site_id,
                'user': user_id,
                'page_size': page_size,
                'search': '',
                'page': 1
            }
            api_url = settings.API_URL + '/api-property/favourite-property-listing/'
            favourite_data = call_api_post_method(params, api_url, token=token)

            if 'error' in favourite_data and favourite_data['error'] == 0:
                try:
                    favorite_listing = favourite_data['data']['data']
                    total = favourite_data['data']['total']
                except:
                    favorite_listing = []
                    total = 0
            else:
                favorite_listing = []
                total = 0
            sno = (int(page) - 1) * int(page_size) + 1
            # ---------------Pagination--------
            pagination_path = 'home/{}/user-dashboard/favourite/favourite-pagination.html'.format(templete_dir)
            pagination_template = get_template(pagination_path)
            total_page = math.ceil(total/page_size)
            pagination_data = {"no_page": total_page, "total_page": range(total_page), "current_page": 1}
            pagination_html = pagination_template.render(pagination_data)
            context = {'favorite_listing': favorite_listing, 'total': total, "active_menu": "favourite_listing", "pagination_html": pagination_html, "sno":sno}

            theme_path = 'home/{}/user-dashboard/favourite/favorite-listing.html'.format(templete_dir)
            return render(request, theme_path, context)
    except Exception as exp:
        print(exp)
        return HttpResponse("Issue in views")

@csrf_exempt
def my_bids(request):
    try:
        try:
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']
            templete_dir = site_detail['site_detail']['theme_directory']

        except Exception as exp:
            print(exp)
            site_id = ""
            templete_dir = 'theme-1'

        static_dir = templete_dir

        token = request.session['token']['access_token']
        user_id = request.session['user_id']
        page_size = 10
        if request.is_ajax() and request.method == 'POST':
            search = ''
            if 'search' in request.POST and request.POST['search']:
                search = request.POST['search']
            page = 1
            if 'page' in request.POST and request.POST['page'] != "":
                page = request.POST['page']

            if 'page_size' in request.POST and request.POST['page_size'] != "":
                page_size = request.POST['page_size']
            params = {
                'site_id': site_id,
                'user_id': user_id,
                'page_size': page_size,
                'search': search,
                'page': page
            }
            sno = (int(page) - 1) * int(page_size) + 1
            api_url = settings.API_URL + '/api-bid/bid-registration-listing/'
            bid_data = call_api_post_method(params, api_url, token=token)
            if 'error' in bid_data and bid_data['error'] == 0:
                bid_listing = bid_data['data']['data']
                total = bid_data['data']['total']
            else:
                bid_listing = []
                total = 0
            context = {'bid_listing': bid_listing, 'total': total, "active_menu": "my_bid", "aws_url": settings.AWS_URL, 'sno': sno}

            bid_listing_path = 'home/{}/user-dashboard/bids/bid-listing.html'.format(templete_dir)
            bid_listing_template = get_template(bid_listing_path)
            bid_listing_html = bid_listing_template.render(context)
            # ---------------Pagination--------
            pagination_path = 'home/{}/user-dashboard/bids/pagination.html'.format(templete_dir)
            pagination_template = get_template(pagination_path)
            total_page = math.ceil(int(total) / int(page_size))
            pagination_data = {"no_page": int(total_page), "total_page": range(total_page), "current_page": int(page)}
            pagination_html = pagination_template.render(pagination_data)
            data = {'bid_listing_html': bid_listing_html, 'status': 200, 'msg': '', 'error': 0, 'total': total, "pagination_html": pagination_html}
            return JsonResponse(data)
        else:
            page = 1
            params = {
                'site_id': site_id,
                'user_id': user_id,
                'page_size': page_size,
                'search': '',
                'page': page
            }
            api_url = settings.API_URL + '/api-bid/bid-registration-listing/'
            bid_data = call_api_post_method(params, api_url, token=token)

            if 'error' in bid_data and bid_data['error'] == 0:
                bid_listing = bid_data['data']['data']
                total = bid_data['data']['total']
            else:
                bid_listing = []
                total = 0
            sno = (int(page) - 1) * int(page_size) + 1
            # ---------------Pagination--------
            pagination_path = 'home/{}/user-dashboard/bids/pagination.html'.format(templete_dir)
            pagination_template = get_template(pagination_path)
            total_page = math.ceil(total/page_size)
            pagination_data = {"no_page": total_page, "total_page": range(total_page), "current_page": 1}
            pagination_html = pagination_template.render(pagination_data)

            x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
            if x_forwarded_for:
                ip_address = x_forwarded_for.split(',')[0]
            else:
                ip_address = request.META.get('REMOTE_ADDR')

            context = {'bid_listing': bid_listing, 'total': total, "active_menu": "my_bid", "pagination_html": pagination_html, "sno":sno, "node_url": settings.NODE_URL, "user_id": user_id, "domain_id": site_id, "ip_address": ip_address}

            theme_path = 'home/{}/user-dashboard/bids/my-bid.html'.format(templete_dir)
            return render(request, theme_path, context)
    except Exception as exp:
        print(exp)
        return HttpResponse("Issue in views")



@csrf_exempt
def property_bid_history(request):
    try:
        user_id = user_id = None
        page_size = 10
        try:
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']
            templete_dir = site_detail['site_detail']['theme_directory']

        except Exception as exp:
            print(exp)
            site_id = ""

        if 'user_id' in request.session and request.session['user_id']:
            token = request.session['token']['access_token']
            user_id = request.session['user_id']

        if request.is_ajax() and request.method == 'POST':
            page = 1
            if 'page' in request.POST and request.POST['page'] != "":
                page = request.POST['page']

            if 'page_size' in request.POST and request.POST['page_size'] != "":
                page_size = request.POST['page_size']

            property_id = request.POST['property_id']
            params = {
                'site_id': site_id,
                'user_id': user_id,
                'page_size': page_size,
                'page': page,
                'property_id': property_id,
                'register_user':user_id
            }

            api_url = settings.API_URL + '/api-bid/subdomain-bid-history/'
            bid_history = call_api_post_method(params, api_url, token=token)

            try:
                prop_detail = bid_history['data']['property_detail']
                image = prop_detail['property_image']
                if 'image' in image and image['image'] != "":
                    image_url = settings.AWS_URL + image['bucket_name'] + '/' + image['image']
                else:
                    image_url = ''
                property_address = prop_detail['address_one']
                property_city = prop_detail['city']
                property_state = prop_detail['state']
                property_postal_code = prop_detail['postal_code']
                property_image = image_url
                auction_type = prop_detail['auction_type']
                bid_increment = prop_detail['bid_increment']
            except Exception as exp:
                print(exp)
                property_address = property_city = property_state = property_postal_code = property_image=auction_type=bid_increment=''

            
            if 'error' in bid_history and bid_history['error'] == 0:
                total = bid_history['data']['total']
                bid_history = bid_history['data']['data']
            else:
                bid_history = []
                total = 0


            context = {'bid_history': bid_history, 'total': total, "aws_url": settings.AWS_URL, 'start_index': (int(page) - 1) * int(page_size)}

            bidder_listing_path = 'home/{}/user-dashboard/bids/property-bid-history.html'.format(templete_dir)
            bidder_listing_template = get_template(bidder_listing_path)
            bid_history_html = bidder_listing_template.render(context)
            # ---------------Pagination--------
            pagination_path = 'home/{}/user-dashboard/bids/bid-history-pagination.html'.format(templete_dir)
            pagination_template = get_template(pagination_path)
            total_page = math.ceil(int(total) / int(page_size))
            pagination_html = ''
            if total_page > 1:
                pagination_data = {"no_page": int(total_page), "total_page": range(total_page), "current_page": int(page),
                                   "pagination_id": "bidHistoryPaginationList", "property_id": property_id}
                pagination_html = pagination_template.render(pagination_data)


            data = {
                'bid_history_html': bid_history_html,
                'status': 200,
                'msg': '',
                'error': 0,
                'total': total,
                "pagination_html": pagination_html,
                'pagination_id': 'bidHistoryPaginationList',
                'property_address': property_address,
                'property_city': property_city,
                'property_state': property_state,
                'property_postal_code': property_postal_code,
                'property_image': property_image,
                'auction_type': auction_type,
                'bid_increment': f"{bid_increment:,}" if bid_increment else 0,
                'page': page,
                'page_size': page_size,
            }
        else:
            data = {'status': 403, 'msg': 'Forbidden', 'error': 1}
        return JsonResponse(data)
    except Exception as exp:
        print(exp)
        return HttpResponse("Issue in views")



@csrf_exempt
def delete_favourite_listing(request):
    try:
        if request.is_ajax() and request.method == 'POST':
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']
            token = request.session['token']['access_token']
            user_id = request.session['user_id']
            params = {
                'domain': site_id,
                'user': user_id,
                'property': request.POST['property']
            }
            url = settings.API_URL + '/api-property/delete-favourite-property/'
            data = call_api_post_method(params, url, token)

            if 'error' in data and data['error'] == 0:
                data = {'error': 0, 'status': 200, 'msg': "deleted successfully"}
            else:
                data = {'error': 1, 'status': 403, 'msg': 'Server error, Please try again'}
        else:
            data = {'error': 1, 'status': 403, 'msg': 'Forbidden'}

        return JsonResponse(data)
    except Exception as exp:
        print(exp)
        data = {'status': 403, 'msg': 'invalid request.'}
        return JsonResponse(data)

@csrf_exempt
def favourite_search_suggestion(request):
    try:
        if request.is_ajax() and request.method == 'POST':
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']
            token = request.session['token']['access_token']
            user_id = request.session['user_id']

            params = {
                'site_id': site_id,
                'user_id': user_id,
                'search': request.POST['search']
            }

            api_url = settings.API_URL + '/api-property/favourite-suggestion/'

            suggestion_data = call_api_post_method(params, api_url, token)
            if 'error' in suggestion_data and suggestion_data['error'] == 0:
                data = {'status': 200, 'suggestion_list': suggestion_data['data'], 'error': 0}
            else:
                data = {'status': 403, 'suggestion_list': [], 'error': 1}
        else:
            data = {'status': 403, 'suggestion_list': [], 'error': 1}

        return JsonResponse(data)
    except Exception as exp:
        print(exp)
        data = {'status': 403, 'suggestion_list': [], 'error': 1}
        return JsonResponse(data)

@csrf_exempt
def blog_listings(request):
    try:
        try:
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']
            templete_dir = site_detail['site_detail']['theme_directory']

        except Exception as exp:
            print(exp)
            site_id = ""
            templete_dir = 'theme-1'

        static_dir = templete_dir
        user_id = None
        token = None
        if 'user_id' in request.session and request.session['user_id']:
            token = request.session['token']['access_token']
            user_id = request.session['user_id']

        sidebar_api_url = settings.API_URL + '/api-blog/blog-sidebar/'
        blog_sidebar_data = call_api_post_method({"site_id": site_id}, sidebar_api_url)
        if 'error' in blog_sidebar_data and blog_sidebar_data['error'] == 0:
            blog_category = blog_sidebar_data['data']['blog_category']
            blog_recent_post = blog_sidebar_data['data']['recent_post']
        else:
            blog_category = []
            blog_recent_post = []

        page_size = 10
        if request.is_ajax() and request.method == 'POST':
            search = ''
            if 'search' in request.POST and request.POST['search']:
                search = request.POST['search']
            page = 1
            if 'page' in request.POST and request.POST['page'] != "":
                page = request.POST['page']

            if 'page_size' in request.POST and request.POST['page_size'] != "":
                page_size = request.POST['page_size']

            category_id = ''
            if 'category_id' in request.POST and request.POST['category_id'] != "":
                category_id = request.POST['category_id']

            sort_column = ''
            sort_by = ''
            sort_order = ''
            if 'sort_column' in request.POST and request.POST['sort_column'] != "":
                sort_column = request.POST['sort_column']
                if sort_column == 'publish_date_asc':
                    sort_by = 'publish_date'
                    sort_order = 'asc'
                elif sort_column == 'publish_date_desc':
                    sort_by = 'publish_date'
                    sort_order = 'desc'

            params = {
                'site_id': site_id,
                'page_size': page_size,
                'search': search,
                'page': page,
                'category_id': category_id,
                'sort_by': sort_by,
                'sort_order': sort_order,
            }

            api_url = settings.API_URL + '/api-blog/front-article-listing/'
            blog_data = call_api_post_method(params, api_url, token=token)
            if 'error' in blog_data and blog_data['error'] == 0:
                blog_listing = blog_data['data']['data']
                total = blog_data['data']['total']
            else:
                blog_listing = []
                total = 0
            context = {'blog_listing': blog_listing, 'total': total, "aws_url": settings.AWS_URL}

            blog_listing_path = 'home/{}/blogs/blog-listing.html'.format(templete_dir)
            blog_listing_template = get_template(blog_listing_path)
            blog_listing_html = blog_listing_template.render(context)
            # ---------------Pagination--------
            pagination_path = 'home/{}/blogs/pagination.html'.format(templete_dir)
            pagination_template = get_template(pagination_path)
            total_page = math.ceil(int(total) / int(page_size))
            pagination_data = {"no_page": int(total_page), "total_page": range(total_page), "current_page": int(page), "pagination_id": "blog_listing_pagination_list"}
            pagination_html = pagination_template.render(pagination_data)

            sidebar_path = 'home/{}/blogs/blog-sidebar.html'.format(templete_dir)
            sidebar_template = get_template(sidebar_path)
            total_page = math.ceil(int(total) / int(page_size))
            sidebar_data = {"blog_category": blog_category, "blog_recent_post": blog_recent_post}
            sidebar_html = sidebar_template.render(sidebar_data)

            data = {'blog_listing_html': blog_listing_html, 'status': 200, 'msg': '', 'error': 0, 'total': total, "pagination_html": pagination_html, "pagination_id": "blog_listing_pagination_list", "blog_sidebar_html": sidebar_html}
            return JsonResponse(data)
        else:
            category = request.GET.get('category', None)
            params = {
                'site_id': site_id,
                'page_size': page_size,
                'search': '',
                'page': 1
            }
            if category:
                params['category_id'] = category


            api_url = settings.API_URL + '/api-blog/front-article-listing/'
            blog_data = call_api_post_method(params, api_url, token=token)
            if 'error' in blog_data and blog_data['error'] == 0:
                blog_listing = blog_data['data']['data']
                total = blog_data['data']['total']
            else:
                blog_listing = []
                total = 0
            # ---------------Pagination--------
            pagination_path = 'home/{}/blogs/pagination.html'.format(templete_dir)
            pagination_template = get_template(pagination_path)
            total_page = math.ceil(total/page_size)
            pagination_data = {"no_page": total_page, "total_page": range(total_page), "current_page": 1, "pagination_id": "blog_listing_pagination_list"}
            pagination_html = pagination_template.render(pagination_data)

            context = {'blog_listing': blog_listing, 'total': total, "pagination_html": pagination_html, "blog_category": blog_category, "blog_recent_post": blog_recent_post, "pagination_id": "blog_listing_pagination_list", "is_home_page": True}

            theme_path = 'home/{}/blogs/blogs.html'.format(templete_dir)
            return render(request, theme_path, context)
    except Exception as exp:
        print(exp)
        return HttpResponse("Issue in views")

@csrf_exempt
def blog_details(request):
    try:
        blog_id = request.GET.get('id', None)
        try:
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']
            templete_dir = site_detail['site_detail']['theme_directory']

        except Exception as exp:
            print(exp)
            site_id = ""
            templete_dir = 'theme-1'

        static_dir = templete_dir
        user_id = None
        token = None
        if 'user_id' in request.session and request.session['user_id']:
            token = request.session['token']['access_token']
            user_id = request.session['user_id']

        sidebar_api_url = settings.API_URL + '/api-blog/blog-sidebar/'
        blog_sidebar_data = call_api_post_method({"site_id": site_id}, sidebar_api_url)
        if 'error' in blog_sidebar_data and blog_sidebar_data['error'] == 0:
            blog_category = blog_sidebar_data['data']['blog_category']
            blog_recent_post = blog_sidebar_data['data']['recent_post']
        else:
            blog_category = []
            blog_recent_post = []

        params = {
            'site_id': site_id,
            'article_id': blog_id
        }
        api_url = settings.API_URL + '/api-blog/front-article-detail/'
        blog_data = call_api_post_method(params, api_url, token=token)
        if 'error' in blog_data and blog_data['error'] == 0:
            blog = blog_data['data']
        else:
            blog = {}

        context = {'blog': blog, "blog_category": blog_category, "blog_recent_post": blog_recent_post, 'is_home_page': True}

        theme_path = 'home/{}/blogs/blog-details.html'.format(templete_dir)
        return render(request, theme_path, context)
    except Exception as exp:
        print(exp)
        return HttpResponse("Issue in views")

@csrf_exempt
def edit_profile(request):
    try:
        try:
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']
            user_id = request.session['user_id']
            templete_dir = site_detail['site_detail']['theme_directory']
            token = request.session['token']['access_token']

            if user_id and site_id and token:
                params = {
                   "site_id": site_id,
                   "user_id": user_id,
                }
                profile_api_url = settings.API_URL + '/api-users/user-profile-detail/'
                profile_response_data = call_api_post_method(params, profile_api_url, token)

                if 'error' in profile_response_data and profile_response_data['error'] == 0:
                    profile_data = profile_response_data['data']
                    # country_id = profile_data['address']['country'] if profile_data['address']['country'] is not None else 1
                    country_id = profile_data['address']['country'] if 'country' in profile_data['address'] and profile_data['address']['country'] is not None else 1
                else:
                   profile_data = []
                   country_id = 1
            else:
                profile_data = []
                country_id = 1

            if request.is_ajax() and request.method == 'POST' and request.POST['change_password']:
                
                pwd_params = {
                    "site_id": site_id,
                    "user_id": user_id,
                    "password": request.POST['password'],
                    "new_password":  request.POST['new_password']
                }

                pwd_url =  settings.API_URL + '/api-users/front-user-change-password/'
                pwd_response = call_api_post_method(pwd_params, pwd_url, token)
                if 'error' in pwd_response and pwd_response['error'] == 0:
                    data = {'status': 200, 'msg': pwd_response['msg'], 'error': 0}
                else:
                    data = {'status': 403, 'msg': pwd_response['msg'], 'error': 1}
                return JsonResponse(data)

        except Exception as exp:
            site_id = ""
            templete_dir = 'theme-1'

        try:
            state_param = {"country_id": country_id}
            state_api_url = settings.API_URL + '/api-settings/get-state/'
            state_data = call_api_post_method(state_param, state_api_url, token)
            state_list = state_data['data']
        except:
            state_list = []
            
        # delete edit profile open modal session once visited
        show_begin_setup = False
        try:
            if 'edit_profile_modal' in request.session:
                del request.session['edit_profile_modal']
                show_begin_setup = True
        except:
            pass

        static_dir = templete_dir
        try:
            country_param = {}
            country_api_url = settings.API_URL + '/api-settings/get-country/'
            country_data = call_api_post_method(country_param, country_api_url)
            country_list = country_data['data']
        except:
            country_list = []
        context = {'active_menu': 'edit_profile', 'profile_data': profile_data, 'aws_url': settings.AWS_URL, 'state_list': state_list, 'show_begin_setup': show_begin_setup, "country_list": country_list}
        theme_path = 'home/{}/user-dashboard/profile/edit-profile.html'.format(templete_dir)
        return render(request, theme_path, context)
    except Exception as exp:
        print(exp)
        return HttpResponse("Issue in views")

@csrf_exempt
def blog_search_suggestion(request):
    try:
        if request.is_ajax() and request.method == 'POST':
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']
            token = None
            try:
                token = request.session['token']['access_token']
            except Exception as exp:
                token = None

            params = {
                'site_id': site_id,
                'search': request.POST['search']
            }

            api_url = settings.API_URL + '/api-blog/front-article-suggestion/'
            suggestion_data = call_api_post_method(params, api_url, token=token)
            if 'error' in suggestion_data and suggestion_data['error'] == 0:
                data = {'status': 200, 'suggestion_list': suggestion_data['data'], 'error': 0}
            else:
                data = {'status': 403, 'suggestion_list': [], 'error': 1}
        else:
            data = {'status': 403, 'suggestion_list': [], 'error': 1}

        return JsonResponse(data)
    except Exception as exp:
        print(exp)
        data = {'status': 403, 'suggestion_list': [], 'error': 1}
        return JsonResponse(data)

@csrf_exempt
def save_enquiry(request):
    try:
        if request.is_ajax() and request.method == 'POST':
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']

            user_phone_no = ''
            if 'user_phone_no' in request.POST and request.POST['user_phone_no'] != "":
                user_phone_no = re.sub('\D', '', request.POST['user_phone_no'])
            params = {
                "site_id": site_id,
                "first_name": request.POST['user_first_name'],
                "email": request.POST['usr_email'],
                "phone_no": user_phone_no,
                "message": request.POST['user_message']
            }

            api_url = settings.API_URL + '/api-contact/front-enquiry/'

            enq_data = call_api_post_method(params, api_url)
            if 'error' in enq_data and enq_data['error'] == 0:
                data = {'status': 200, 'msg': 'Send Successfully.', 'error': 0}
            else:
                data = {'status': 403, 'msg': 'Some error occurs, please try again.', 'error': 1}
        else:
            data = {'status': 403, 'msg': 'Forbidden', 'error': 1}


        return JsonResponse(data)
    except Exception as exp:
        print(exp)
        data = {'status': 403, 'msg': 'Invalid request', 'error': 1}
        return JsonResponse(data)


@csrf_exempt
def profile_update(request):
    try:
        if request.is_ajax() and request.method == 'POST':
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']
            token = request.session['token']['access_token']
            user_id = request.session['user_id']
            
            phone_no = ''
            if 'phone_no' in request.POST and request.POST['phone_no'] != "":
                phone_no = re.sub('\D', '', request.POST['phone_no']) 
            
            first_name = ''
            if 'first_name' in request.POST and request.POST['first_name'] != "":
                first_name = request.POST['first_name'].strip()
            
            last_name = ''
            if 'last_name' in request.POST and request.POST['last_name'] != "":
                last_name = request.POST['last_name'].strip()
            
            email = ''
            if 'email' in request.POST and request.POST['email'] != "":
                email = request.POST['email'].strip()

            
            brokerage_name = licence_number = ''
            try:
                if 'brokerage_name' in request.POST and request.POST['brokerage_name']:
                    brokerage_name = request.POST['brokerage_name'].strip()
                if 'licence_number' in request.POST and request.POST['licence_number']:
                    licence_number = request.POST['licence_number'].strip()
            except:
                pass

            address = {
                'address_first': request.POST['address_first'],
                'state': request.POST['state'],
                'postal_code': request.POST['zip_code'],
                'city': request.POST['address_city'],
                'country': request.POST['country'],
            }
            params = {
                'site_id': site_id,
                'user_id': user_id,
                'first_name': first_name,
                'last_name': last_name,
                'email': email,
                'phone_no': phone_no,
                'address': address,
                'brokerage_name': brokerage_name,
                'licence_number': licence_number

            }
            if 'profile_image' in request.POST and request.POST['profile_image'] != "":
                params['profile_image'] = request.POST['profile_image']

            api_url = settings.API_URL + '/api-users/user-profile-update/'
            update_data = call_api_post_method(params, api_url, token)

            if 'error' in update_data and update_data['error'] == 0:
                if 'pic_name' in request.POST and request.POST['pic_name'] != "":
                    api_url = settings.API_URL + '/api-users/login-details/'
                    payload = {
                        'token': token,
                        'site_id': site_id
                    }
                    response = call_api_post_method(payload, api_url)
                    if "error" in response and response['error'] == 0:
                        response = response['data']
                        request.session.modified = True
                        request.session['profile_image'] = response['profile_image']
                                            
                data = {'status': 200, 'msg': update_data['msg'], 'error': 0}
            else:
                data = {'status': 403, 'msg': update_data['msg'], 'error': 1}
        else:
            data = {'status': 403, 'msg': update_data['msg'], 'error': 1}

        return JsonResponse(data)
    except Exception as exp:
        print(exp)
        data = {'status': 403, 'msg': 'Invalid data', 'error': 1}
        return JsonResponse(data)

@csrf_exempt
def chat(request):
    try:
        try:
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']
            templete_dir = site_detail['site_detail']['theme_directory']

        except Exception as exp:
            print(exp)
            site_id = ""
            templete_dir = 'theme-1'

        static_dir = templete_dir

        token = request.session['token']['access_token']
        user_id = request.session['user_id']

        context = {"active_menu": "chat"}

        theme_path = 'home/{}/user-dashboard/chat/chat.html'.format(templete_dir)
        return render(request, theme_path, context)
    except Exception as exp:
        print(exp)
        return HttpResponse("Issue in views")


@csrf_exempt
def image_upload(request):
    try:
        if request.is_ajax() and request.method == 'POST':
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']
            token = request.session['token']['access_token']
            user_id = request.session['user_id']
            file_urls = ""
            upload_to = ""
            uploaded_file_list = []
            file_size = request.POST['file_size']
            try:
                for key, value in request.FILES.items():
                    params = {}
                    if 'profile_image' in key.lower():
                        upload_to = 'profile_image'
                        doc_type = 9
                        file_urls = request.FILES[key]
            except:
                pass

            if int(request.POST['file_length']) == 1:
                for key, value in request.FILES.items():
                    params = {}
                    if 'profile_image' in key.lower():
                        upload_to = 'profile_image'
                        doc_type = 9
                        file_res = request.FILES[key]
                        response = save_to_s3(file_res, upload_to)            
                        if 'error' in response and response['error'] == 0:
                            try:
                                upload_param = {
                                    "site_id": site_id,
                                    "user_id": user_id,
                                    "doc_file_name": response['file_name'],
                                    "document_type": doc_type,
                                    "bucket_name": upload_to,
                                    "added_by": user_id,
                                    "file_size": str(file_size)+'MB'
                                    }
                                url = settings.API_URL + '/api-users/file-upload/'
                                upload_data = call_api_post_method(upload_param, url, token)
                                upload_id = upload_data['data']['upload_id']
                                upload_size = upload_data['data']['file_size']
                                upload_date = upload_data['data']['added_date']
                            except:
                                upload_id = 0
                                upload_size = '0MB'
                                upload_date = ''
                            params['file_name'] = response['file_name']
                            params['error'] = 0
                            params['msg'] = response['msg']
                            params['upload_id'] = upload_id
                            params['file_size'] = upload_size
                            params['upload_date'] = upload_date
                            params['upload_to'] = upload_to
                        else:
                            params['file_name'] = response['file_name']
                            params['error'] = 1
                            params['msg'] = response['msg']
                            params['upload_id'] = 0
                            params['file_size'] = '0MB'
                            params['upload_date'] = ''
                            params['upload_to'] = upload_to
                        uploaded_file_list.append(params)
                    return JsonResponse({'status': 200, 'uploaded_file_list': uploaded_file_list, 'error': 0})
        else:
            data = {'status': 201, 'msg': 'Invalid data', 'error': 1}
            return JsonResponse(data)    

    except Exception as exp:
        print(exp)
        data = {'status': 403, 'msg': 'Invalid data', 'error': 1}
        return JsonResponse(data)

@csrf_exempt
def send_enquiry_message(request):
    try:
        if request.is_ajax() and request.method == 'POST':
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']
            user_id = None
            token = None
            if 'user_id' in request.session:
                token = request.session['token']['access_token']
                user_id = request.session['user_id']
            
            doc_id_list =  dic_name_list = []
            if request.POST['chat_doc_id'] != "":
                doc_id_list = request.POST['chat_doc_id'].split(',')
            if request.POST['chat_doc_name'] != "":
                dic_name_list = request.POST['chat_doc_name'].split(',')

            params = {
                "site_id": site_id,
                "property_id": request.POST['property_id'],
                "user_id": user_id,
                "message": request.POST['user_message'],
                "property_image": request.POST['property_image'],
                "property_name": request.POST['property_name'],
                "property_address": request.POST['property_address'],
                "property_city": request.POST['property_city'],
                "property_state": request.POST['property_state'],
                "property_zipcode": request.POST['property_zipcode'],
                "doc_id_list": doc_id_list,
                "chat_doc_names": dic_name_list
            }

            api_url = settings.API_URL + '/api-contact/chat-to-seller/'
            enq_data = call_api_post_method(params, api_url, token=token)

            if 'error' in enq_data and enq_data['error'] == 0:
                data = {'status': 200, 'msg': 'Send Successfully.', 'error': 0}
            else:
                data = {'status': 403, 'msg': enq_data['msg'], 'error': 1}
        else:
            data = {'status': 403, 'msg': 'Forbidden', 'error': 1}


        return JsonResponse(data)
    except Exception as exp:
        print(exp)
        data = {'status': 403, 'msg': 'Invalid request', 'error': 1}
        return JsonResponse(data)


@csrf_exempt
def our_agents(request):
    try:
        try:
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']
            templete_dir = site_detail['site_detail']['theme_directory']

        except Exception as exp:
            print(exp)
            site_id = ""
            templete_dir = 'theme-1'

        static_dir = templete_dir
        user_id = None
        token = None
        if 'user_id' in request.session and request.session['user_id']:
            token = request.session['token']['access_token']
            user_id = request.session['user_id']

        sidebar_api_url = settings.API_URL + '/api-blog/blog-sidebar/'
        blog_sidebar_data = call_api_post_method({"site_id": site_id}, sidebar_api_url)
        if 'error' in blog_sidebar_data and blog_sidebar_data['error'] == 0:
            blog_category = blog_sidebar_data['data']['blog_category']
            blog_recent_post = blog_sidebar_data['data']['recent_post']
        else:
            blog_category = []
            blog_recent_post = []

        page_size = 12
        if request.is_ajax() and request.method == 'POST':
            search = ''
            if 'search' in request.POST and request.POST['search']:
                search = request.POST['search']

            page = 1
            if 'page' in request.POST and request.POST['page']:
                page = int(request.POST['page'])


            if 'perpage' in request.POST and request.POST['perpage']:
                page_size = int(request.POST['perpage'])

            sort_column = ''
            short_by = ''
            sort_order = ''
            if 'sort_column' in request.POST and request.POST['sort_column']:
                sort_column = request.POST['sort_column']
                if sort_column == 'featured_first':
                    short_by = 'agent_start'
                    sort_order = 'asc'
                elif sort_column == 'featured_last':
                    short_by = 'agent_end'
                    sort_order = 'desc'

            params = {
                "site_id": site_id,
                "page_size": page_size,
                "page": page,
                "search": search,
                "short_by": short_by,
                "sort_order": sort_order
            }

            api_url = settings.API_URL + '/api-users/agent-list/'

            agent_listing_response = call_api_post_method(params, api_url)


            if 'error' in agent_listing_response and agent_listing_response['error'] == 0:
                agent_list = agent_listing_response['data']['data']
                total = agent_listing_response['data']['total']
            else:
                agent_list = []
                total = 0
            context = {'agent_list': agent_list, 'total': total, "aws_url": settings.AWS_URL, 'sess_user_id': user_id}

            agent_list_data_path = 'home/{}/agents/our-agent-listing-content.html'.format(templete_dir)
            listing_template = get_template(agent_list_data_path)
            listing_html = listing_template.render(context)
            # ---------------Pagination--------
            pagination_path = 'home/{}/agents/agent_pagination.html'.format(templete_dir)
            pagination_template = get_template(pagination_path)
            total_page = math.ceil(int(total) / int(page_size))
            pagination_data = {"no_page": int(total_page), "total_page": range(total_page), "current_page": int(page), "pagination_id": "agent_listing_pagination_list"}
            pagination_html = pagination_template.render(pagination_data)


            data = {'agent_list': agent_list, 'listing_html': listing_html, 'status': 200, 'msg': '', 'error': 0, 'total': total, "pagination_html": pagination_html, "pagination_id": "agent_listing_pagination_list"}
            return JsonResponse(data)
        else:
            category = request.GET.get('category', None)
            params = {
                'site_id': site_id,
                'page_size': page_size,
                'search': '',
                'page': 1
            }

            api_url = settings.API_URL + '/api-users/agent-list/'
            agent_listing_response = call_api_post_method(params, api_url, token=token)
            if 'error' in agent_listing_response and agent_listing_response['error'] == 0:
                agent_list = agent_listing_response['data']['data']
                total = agent_listing_response['data']['total']
            else:
                agent_list = []
                total = 0
            # ---------------Pagination--------
            pagination_path = 'home/{}/agents/agent_pagination.html'.format(templete_dir)
            pagination_template = get_template(pagination_path)
            total_page = math.ceil(total/page_size)
            pagination_data = {"no_page": total_page, "total_page": range(total_page), "current_page": 1, "pagination_id": "blog_listing_pagination_list"}
            pagination_html = pagination_template.render(pagination_data)

            context = {'agent_list': agent_list, 'total': total, "pagination_html": pagination_html, "pagination_id": "agent_listing_pagination_list", "is_home_page": True, "sess_user_id": user_id}

            theme_path = 'home/{}/agents/our-agents.html'.format(templete_dir)
            return render(request, theme_path, context)
    except Exception as exp:
        print(exp)
        return HttpResponse("Issue in views")

@csrf_exempt
def agent_search_suggestion(request):
    try:
        if request.is_ajax() and request.method == 'POST':
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']
            token = None
            if 'user_id' in request.session:
                token = request.session['token']['access_token']

            params = {
                'site_id': site_id,
                'search': request.POST['search']
            }

            api_url = settings.API_URL + '/api-users/agent-search-suggestion/'

            suggestion_data = call_api_post_method(params, api_url, token)

            if 'error' in suggestion_data and suggestion_data['error'] == 0:
                data = {'status': 200, 'suggestion_list': suggestion_data['data'], 'error': 0}
            else:
                data = {'status': 403, 'suggestion_list': [], 'error': 1}
        else:
            data = {'status': 403, 'suggestion_list': [], 'error': 1}

        return JsonResponse(data)
    except Exception as exp:
        print(exp)
        data = {'status': 403, 'suggestion_list': [], 'error': 1}
        return JsonResponse(data)


@csrf_exempt
def chat_to_agent(request):
    try:
        if request.is_ajax() and request.method == 'POST':
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']
            user_id = None
            token = None
            if 'user_id' in request.session:
                token = request.session['token']['access_token']
                user_id = request.session['user_id']

            params = {
                "site_id": site_id,
                "agent_id": request.POST['agent_id'],
                "user_id": user_id,
                "message": request.POST['user_message']
            }

            api_url = settings.API_URL + '/api-contact/chat-to-agent/'

            enq_data = call_api_post_method(params, api_url, token=token)
            if 'error' in enq_data and enq_data['error'] == 0:
                data = {'status': 200, 'msg': enq_data['msg'], 'error': 0}
            else:
                data = {'status': 403, 'msg': enq_data['msg'], 'error': 1}
        else:
            data = {'status': 403, 'msg': 'Forbidden', 'error': 1}


        return JsonResponse(data)
    except Exception as exp:
        print(exp)
        data = {'status': 403, 'msg': 'Invalid request', 'error': 1}
        return JsonResponse(data)

@csrf_exempt
def bid_online(request):
    try:
        try:
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']
            templete_dir = site_detail['site_detail']['theme_directory']

        except Exception as exp:
            print(exp)
            site_id = ""
            templete_dir = 'theme-1'

        static_dir = templete_dir
        user_id = None
        token = None
        if 'user_id' in request.session and request.session['user_id']:
            token = request.session['token']['access_token']
            user_id = request.session['user_id']

        params = {
            'site_id': site_id,
            'slug': 'how-to-bid-online'
        }
        api_url = settings.API_URL + '/api-cms/get-page/'
        bid_online_data = call_api_post_method(params, api_url, token=token)
        if 'error' in bid_online_data and bid_online_data['error'] == 0:
            bid_online = bid_online_data['data']
        else:
            bid_online = {}

        context = {'bid': bid_online, 'is_home_page': True}

        theme_path = 'home/{}/cms/how-to-bid-online.html'.format(templete_dir)
        return render(request, theme_path, context)
    except Exception as exp:
        print(exp)
        return HttpResponse("Issue in views")

@csrf_exempt
def auction_type(request):
    try:
        try:
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']
            templete_dir = site_detail['site_detail']['theme_directory']

        except Exception as exp:
            print(exp)
            site_id = ""
            templete_dir = 'theme-1'

        static_dir = templete_dir
        user_id = None
        token = None
        if 'user_id' in request.session and request.session['user_id']:
            token = request.session['token']['access_token']
            user_id = request.session['user_id']

        params = {
            'site_id': site_id,
            'slug': 'auction-type'
        }
        api_url = settings.API_URL + '/api-cms/get-auction-type/'
        auction_data = call_api_post_method(params, api_url, token=token)
        if 'error' in auction_data and auction_data['error'] == 0:
            auction = auction_data['data']['data']
        else:
            auction = {}


        context = {'auction': auction, 'is_home_page': True}

        theme_path = 'home/{}/cms/auction-type.html'.format(templete_dir)
        return render(request, theme_path, context)
    except Exception as exp:
        print(exp)
        return HttpResponse("Issue in views")


@csrf_exempt
def faq(request):
    try:
        try:
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']
            templete_dir = site_detail['site_detail']['theme_directory']

        except Exception as exp:
            print(exp)
            site_id = ""
            templete_dir = 'theme-1'

        static_dir = templete_dir
        user_id = None
        token = None
        if 'user_id' in request.session and request.session['user_id']:
            token = request.session['token']['access_token']
            user_id = request.session['user_id']
        params = {
            'site_id': site_id,
            'user': user_id,
        }
        api_url = settings.API_URL + '/api-faq/faq-listing/'
        try:
            params['faq_type'] = 1

            buyer_faq_data = call_api_post_method(params, api_url, token=token)

            if 'error' in buyer_faq_data and buyer_faq_data['error'] == 0:
                buyer_faq = buyer_faq_data['data']['data']
            else:
                buyer_faq = []
        except:
            buyer_faq = []

        try:
            params['faq_type'] = 2
            seller_faq_data = call_api_post_method(params, api_url, token=token)
            if 'error' in seller_faq_data and seller_faq_data['error'] == 0:
                seller_faq = seller_faq_data['data']['data']
            else:
                seller_faq = []
        except:
            seller_faq = []

        try:
            params['faq_type'] = 3
            agent_faq_data = call_api_post_method(params, api_url, token=token)
            if 'error' in agent_faq_data and agent_faq_data['error'] == 0:
                agent_faq = agent_faq_data['data']['data']
            else:
                agent_faq = []
        except:
            agent_faq = []




        context = {'buyer_faq': buyer_faq, 'seller_faq': seller_faq, 'agent_faq': agent_faq, 'is_home_page': True}

        theme_path = 'home/{}/cms/faq.html'.format(templete_dir)
        return render(request, theme_path, context)
    except Exception as exp:
        print(exp)
        return HttpResponse("Issue in views")


@csrf_exempt
def video_tutorials(request):
    try:
        try:
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']
            templete_dir = site_detail['site_detail']['theme_directory']
        except Exception as exp:
            site_id = ""
            templete_dir = 'theme-1'

        static_dir = templete_dir

        http_host = request.META['HTTP_HOST']
        site_url = settings.URL_SCHEME + str(http_host)

        user_id = None
        if 'user_id' in request.session and request.session['user_id']:
            user_id = request.session['user_id']
        if request.is_ajax() and request.method == 'POST':

            search = ''
            if 'search' in request.POST and request.POST['search']:
                search = request.POST['search']
            page = 1
            if 'page' in request.POST and request.POST['page']:
                page = int(request.POST['page'])
            page_size = 10
            if 'perpage' in request.POST and request.POST['perpage']:
                page_size = int(request.POST['perpage'])
            sort_column = ''
            short_by = ''
            sort_order = ''
            if 'sort_column' in request.POST and request.POST['sort_column']:
                sort_column = request.POST['sort_column']
                if sort_column == 'featured_first':
                    short_by = 'agent_start'
                    sort_order = 'asc'
                elif sort_column == 'featured_last':
                    short_by = 'agent_end'
                    sort_order = 'desc'

            params = {
                "site_id": site_id,
                "page_size": page_size,
                "page": page,
                "short_by": short_by,
                "sort_order": sort_order
            }
            api_url = settings.API_URL + '/api-cms/video-tutorials/'
            video_list_response = call_api_post_method(params, api_url)
            if 'error' in video_list_response and video_list_response['error'] == 0:
                video_list = video_list_response['data']['data']
                total = video_list_response['data']['total']
                video_list_data_path = 'home/{}/video/video-tutorial-content.html'.format(templete_dir)
                listing_template = get_template(video_list_data_path)
                listing_html = listing_template.render(
                    {'video_list': video_list, 'total': total, 'aws_url': settings.AWS_URL, 'sort_column': sort_column,
                     'SITE_URL': site_url, 'sess_user_id': user_id})
                data = {'listing_html': listing_html, 'video_list': video_list, 'status': 200,
                        'msg': 'Video list successfully received.', 'error': 0, 'total': total}
            else:
                data = {'status': 403, 'msg': 'Fetching error', 'video_list': [], 'error': 1, 'total': 0}
            return JsonResponse(data)
        context = {'is_home_page': True, 'data': 'video tutorials comming soon', 'sess_user_id': user_id,
                   'SITE_URL': site_url}
        theme_path = 'home/{}/video/video-tutorial.html'.format(templete_dir)
        return render(request, theme_path, context)
    except Exception as exp:
        print(exp)
        return HttpResponse("Issue in views")

@csrf_exempt
def terms(request):
    try:
        try:
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']
            templete_dir = site_detail['site_detail']['theme_directory']
        except Exception as exp:
            print(exp)
            site_id = ""
            templete_dir = 'theme-1'

        static_dir = templete_dir
        user_id = None
        token = None
        if 'user_id' in request.session and request.session['user_id']:
            token = request.session['token']['access_token']
            user_id = request.session['user_id']

        params = {
            'site_id': site_id,
            'slug': 'terms'
        }
        api_url = settings.API_URL + '/api-cms/get-page/'
        terms_data = call_api_post_method(params, api_url, token=token)
        if 'error' in terms_data and terms_data['error'] == 0:
            terms = terms_data['data']
        else:
            terms = {}

        context = {'terms': terms, 'is_home_page': True}

        theme_path = 'home/{}/cms/terms.html'.format(templete_dir)
        return render(request, theme_path, context)

    except Exception as exp:
        print(exp)
        return HttpResponse("Issue in views")

@csrf_exempt
def get_chat_count(request):
    try:
        if request.is_ajax() and request.method == 'POST':
            try:
                site_detail = subdomain_site_details(request)
                site_id = site_detail['site_detail']['site_id']
            except Exception as exp:
                print(exp)
                site_id = ""

            user_id = None
            token = None
            if 'user_id' in request.session and request.session['user_id']:
                token = request.session['token']['access_token']
                user_id = request.session['user_id']

            params = {
                "site_id": site_id,
                "user_id": user_id
            }
            api_url = settings.API_URL + '/api-settings/chat-count/'
            chat_data = call_api_post_method(params, api_url, token=token)
            if 'error' in chat_data and chat_data['error'] == 0:
                chat = chat_data['data']
                data = {'chat_count': chat['user_msg_cnt'], 'error': 0, 'msg':'success'}
            else:
                chat = {}
                data = {'chat_count': 0, 'error': 1, 'msg':'failure'}

            return JsonResponse(data)

    except Exception as exp:
        print(exp)
        return HttpResponse("Issue in views")

@csrf_exempt
def privacy_policy(request):
    try:
        try:
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']
            templete_dir = site_detail['site_detail']['theme_directory']
        except Exception as exp:
            print(exp)
            site_id = ""
            templete_dir = 'theme-1'
        static_dir = templete_dir
        user_id = None
        token = None
        if 'user_id' in request.session and request.session['user_id']:
            token = request.session['token']['access_token']
            user_id = request.session['user_id']
        params = {
            'site_id': site_id,
            'slug': 'privacy-policy'
        }
        api_url = settings.API_URL + '/api-cms/get-page/'
        privacy_data = call_api_post_method(params, api_url, token=token)
        if 'error' in privacy_data and privacy_data['error'] == 0:
            privacy = privacy_data['data']
        else:
            privacy = {}
        context = {'privacy': privacy, 'is_home_page': True}
        theme_path = 'home/{}/cms/privacy-policy.html'.format(templete_dir)
        return render(request, theme_path, context)
    except Exception as exp:
        print(exp)
        return HttpResponse("Issue in views")

@csrf_exempt
def bid_registration(request):
    property = request.GET.get('property', None)
    try:
        try:
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']
            templete_dir = site_detail['site_detail']['theme_directory']
        except Exception as exp:
            print(exp)
            site_id = ""
            templete_dir = 'theme-1'

        static_dir = templete_dir
        user_id = None
        token = None
        if 'user_id' in request.session and request.session['user_id']:
            token = request.session['token']['access_token']
            user_id = request.session['user_id']

        try:
            state_param = {}
            state_api_url = settings.API_URL + '/api-settings/get-state/'
            state_data = call_api_post_method(state_param, state_api_url, token)
            state_list = state_data['data']
        except:
            state_list = []

        if request.is_ajax() and request.method == 'POST':
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']

            phone_no = ''
            if 'phone' in request.POST and request.POST['phone'] != "":
                phone_no = re.sub('\D', '', request.POST['phone'])

            agent_buyer_phone_no = ''
            if 'agent_buyer_phone' in request.POST and request.POST['agent_buyer_phone'] != "":
                agent_buyer_phone_no = re.sub('\D', '', request.POST['agent_buyer_phone'])

            email = request.POST['email']



            bidder_doc = []
            if 'is_submit_proof_fund' in request.POST and int(request.POST['is_submit_proof_fund']) == 1:
                bidder_doc_id = request.POST['bidder_doc_id']
                try:
                    bidder_doc = bidder_doc_id.split(',')

                except:
                    bidder_doc = []


            addresses = {
                "first_name": request.POST['first_name'],
                "last_name": request.POST['last_name'],
                "email": request.POST['email'],
                "phone_no": phone_no,
                "address_first": request.POST['address'],
                "city": request.POST['city'],
                "state": request.POST['state'],
                "postal_code": request.POST['bidder_zip_code'],
                "country": request.POST['buyer_country']
            }
            x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
            if x_forwarded_for:
                ip = x_forwarded_for.split(',')[0]
            else:
                ip = request.META.get('REMOTE_ADDR')

            bidder_params = {
                "domain": site_id,
                "user": user_id,
                "property_id": request.POST['property_id'],
                "user_type": request.POST['user_type'],
                "term_accepted": request.POST['term_accepted'],
                "age_accepted": request.POST['age_validate'],
                "uploads": bidder_doc,
                # "ip_address": request.META.get("REMOTE_ADDR")
                "ip_address": ip
            }
            if bidder_doc:
                bidder_params['upload_pof'] = 1
            else:
                bidder_params['upload_pof'] = 0
                bidder_params['reason_for_not_upload'] = request.POST['doc_reason']

            if 'user_type' in request.POST and int(request.POST['user_type']) == 2:
                try:
                    bidder_params['working_with_agent'] = request.POST['working_with_agent']
                except:
                    bidder_params['working_with_agent'] = 0

            if 'user_type' in request.POST and int(request.POST['user_type']) == 4:

                try:
                    bidder_params['property_yourself'] = request.POST['working_with_agent']
                except:
                    bidder_params['property_yourself'] = 0

            if 'user_type' in request.POST and int(request.POST['user_type']) == 2 and 'working_with_agent' in request.POST and int(request.POST['working_with_agent']) == 1:
                agent_buyer_addresses = {
                    "first_name": request.POST['agent_buyer_first_name'],
                    "last_name": request.POST['agent_buyer_last_name'],
                    "email": request.POST['agent_buyer_email'],
                    "phone_no": agent_buyer_phone_no,
                    "address_first": request.POST['agent_buyer_address'],
                    "state": request.POST['agent_buyer_state'],
                    "city": request.POST['agent_buyer_city'],
                    "postal_code": request.POST['agent_buyer_zipcode'],
                    "company_name": request.POST['agent_buyer_company_name'],
                    "country": request.POST['agent_buyer_country']
                }
                bidder_params['agent_address'] = agent_buyer_addresses
                bidder_params['buyer_address'] = addresses
            elif 'user_type' in request.POST and int(request.POST['user_type']) == 2 and 'working_with_agent' in request.POST and int(request.POST['working_with_agent']) == 0:
                bidder_params['buyer_address'] = addresses
            elif 'user_type' in request.POST and int(request.POST['user_type']) == 4 and 'working_with_agent' in request.POST and int(request.POST['working_with_agent']) == 1:
                agent_buyer_addresses = {
                    "first_name": request.POST['agent_buyer_first_name'],
                    "last_name": request.POST['agent_buyer_last_name'],
                    "email": request.POST['agent_buyer_email'],
                    "phone_no": agent_buyer_phone_no,
                    "address_first": request.POST['agent_buyer_address'],
                    "state": request.POST['agent_buyer_state'],
                    "city": request.POST['city'],
                    "postal_code": request.POST['agent_buyer_zipcode'],
                    "country": request.POST['agent_buyer_country']
                }
                addresses['company_name'] = request.POST['bidder_company_name']
                bidder_params['agent_address'] = addresses
                bidder_params['buyer_address'] = agent_buyer_addresses

            elif 'user_type' in request.POST and int(request.POST['user_type']) == 4 and 'working_with_agent' in request.POST and int(request.POST['working_with_agent']) == 0:

                addresses['company_name'] = request.POST['bidder_company_name']
                bidder_params['buyer_seller_address'] = addresses
            asset_type = request.POST['asset_type']
            if asset_type:
                bidder_api_url = settings.API_URL + '/api-bid/new-bid-registration/'
            else:
                bidder_api_url = settings.API_URL + '/api-bid/bid-registration/'
            save_data = call_api_post_method(bidder_params, bidder_api_url, token=token)
            if 'error' in save_data and save_data['error'] == 0:
                try:
                    reg_id = save_data['data']['registration_id']
                except:
                    reg_id = ''
                if bidder_doc:
                    doc_uploaded = 1
                    sucess_paragraph = 'The listing agent will either promptly approve your request to bid or contact you with any questions before granting you permission to bid. Once approved you will be sent an email to '+email+' granting you approval to bid. If you do not receive the approval email within the next few hours please check your junk/spam folder first or contact the listing agent for help.'
                else:
                    doc_uploaded = 0
                    sucess_paragraph = 'You may select the Proof of Funds and/or Bank Letter of Credit option and get approved in a matter of minutes.'
                data = {'status': 200, 'msg': 'Registered Successfully.', 'error': 0, 'data': save_data, 'doc_uploaded': doc_uploaded, 'sucess_paragraph': sucess_paragraph, 'reg_id': reg_id}
            else:
                data = {'status': 403, 'msg': save_data['msg'], 'error': 1, 'data': save_data, 'reg_id': ''}
            return JsonResponse(data)

        params = {
            "user": user_id,
            "domain": site_id,
            "property": property
        }
        api_url = settings.API_URL + '/api-bid/bid-registration-detail/'
        bidder_data = call_api_post_method(params, api_url, token=token)

        if 'error' in bidder_data and bidder_data['error'] == 0:
            bidder = bidder_data['data']
        else:
            bidder = {}

        try:
            country_param = {}
            country_api_url = settings.API_URL + '/api-settings/get-country/'
            country_data = call_api_post_method(country_param, country_api_url)
            country_list = country_data['data']
        except:
            country_list = []

        context = {'bidder': bidder, 'state_list': state_list, 'is_home_page': True, 'node_url': settings.NODE_URL, "country_list": country_list, "stripe_publishable_key": settings.STRIPE_PUBLIC_KEY}
        theme_path = 'home/{}/listings/bid-registration.html'.format(templete_dir)
        return render(request, theme_path, context)

    except Exception as exp:
        print(exp)
        return HttpResponse("Issue in views")


def deposit_bid_registration(request):
    property = request.GET.get('property', None)
    try:
        try:
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']
            templete_dir = site_detail['site_detail']['theme_directory']
        except Exception as exp:
            print(exp)
            site_id = ""
            templete_dir = 'theme-1'

        static_dir = templete_dir
        user_id = None
        token = None
        if 'user_id' in request.session and request.session['user_id']:
            token = request.session['token']['access_token']
            user_id = request.session['user_id']

        if request.is_ajax() and request.method == 'POST':
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']

            phone_no = ''
            if 'phone' in request.POST and request.POST['phone'] != "":
                phone_no = re.sub('\D', '', request.POST['phone'])

            agent_buyer_phone_no = ''
            if 'agent_buyer_phone' in request.POST and request.POST['agent_buyer_phone'] != "":
                agent_buyer_phone_no = re.sub('\D', '', request.POST['agent_buyer_phone'])

            email = request.POST['email']



            bidder_doc = []
            if 'is_submit_proof_fund' in request.POST and int(request.POST['is_submit_proof_fund']) == 1:
                bidder_doc_id = request.POST['bidder_doc_id']
                try:
                    bidder_doc = bidder_doc_id.split(',')

                except:
                    bidder_doc = []


            addresses = {
                "first_name": request.POST['first_name'],
                "last_name": request.POST['last_name'],
                "email": request.POST['email'],
                "phone_no": phone_no,
                "address_first": request.POST['address'],
                "city": request.POST['city'],
                "state": request.POST['state'],
                "postal_code": request.POST['bidder_zip_code'],
                "country": request.POST['buyer_country']
            }
            x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
            if x_forwarded_for:
                ip = x_forwarded_for.split(',')[0]
            else:
                ip = request.META.get('REMOTE_ADDR')

            bidder_params = {
                "domain": site_id,
                "user": user_id,
                "property_id": request.POST['property_id'],
                "user_type": request.POST['user_type'],
                "term_accepted": request.POST['term_accepted'],
                "age_accepted": request.POST['age_validate'],
                "uploads": bidder_doc,
                # "ip_address": request.META.get("REMOTE_ADDR")
                "ip_address": ip
            }
            if bidder_doc:
                bidder_params['upload_pof'] = 1
            else:
                bidder_params['upload_pof'] = 0
                bidder_params['reason_for_not_upload'] = request.POST['doc_reason']

            if 'user_type' in request.POST and int(request.POST['user_type']) == 2:
                try:
                    bidder_params['working_with_agent'] = request.POST['working_with_agent']
                except:
                    bidder_params['working_with_agent'] = 0

            if 'user_type' in request.POST and int(request.POST['user_type']) == 4:

                try:
                    bidder_params['property_yourself'] = request.POST['working_with_agent']
                except:
                    bidder_params['property_yourself'] = 0

            if 'user_type' in request.POST and int(request.POST['user_type']) == 2 and 'working_with_agent' in request.POST and int(request.POST['working_with_agent']) == 1:
                agent_buyer_addresses = {
                    "first_name": request.POST['agent_buyer_first_name'],
                    "last_name": request.POST['agent_buyer_last_name'],
                    "email": request.POST['agent_buyer_email'],
                    "phone_no": agent_buyer_phone_no,
                    "address_first": request.POST['agent_buyer_address'],
                    "state": request.POST['agent_buyer_state'],
                    "city": request.POST['agent_buyer_city'],
                    "postal_code": request.POST['agent_buyer_zipcode'],
                    "company_name": request.POST['agent_buyer_company_name'],
                    "country": request.POST['agent_buyer_country']
                }
                bidder_params['agent_address'] = agent_buyer_addresses
                bidder_params['buyer_address'] = addresses
            elif 'user_type' in request.POST and int(request.POST['user_type']) == 2 and 'working_with_agent' in request.POST and int(request.POST['working_with_agent']) == 0:
                bidder_params['buyer_address'] = addresses
            elif 'user_type' in request.POST and int(request.POST['user_type']) == 4 and 'working_with_agent' in request.POST and int(request.POST['working_with_agent']) == 1:
                agent_buyer_addresses = {
                    "first_name": request.POST['agent_buyer_first_name'],
                    "last_name": request.POST['agent_buyer_last_name'],
                    "email": request.POST['agent_buyer_email'],
                    "phone_no": agent_buyer_phone_no,
                    "address_first": request.POST['agent_buyer_address'],
                    "state": request.POST['agent_buyer_state'],
                    "city": request.POST['city'],
                    "postal_code": request.POST['agent_buyer_zipcode'],
                    "country": request.POST['agent_buyer_country']
                }
                addresses['company_name'] = request.POST['bidder_company_name']
                bidder_params['agent_address'] = addresses
                bidder_params['buyer_address'] = agent_buyer_addresses

            elif 'user_type' in request.POST and int(request.POST['user_type']) == 4 and 'working_with_agent' in request.POST and int(request.POST['working_with_agent']) == 0:

                addresses['company_name'] = request.POST['bidder_company_name']
                bidder_params['buyer_seller_address'] = addresses
            asset_type = request.POST['asset_type']

            bidder_api_url = settings.API_URL + '/api-bid/deposit-bid-registration/'
            
            save_data = call_api_post_method(bidder_params, bidder_api_url, token=token)
            if 'error' in save_data and save_data['error'] == 0:
                try:
                    reg_id = save_data['data']['registration_id']
                except:
                    reg_id = ''
                if bidder_doc:
                    doc_uploaded = 1
                    sucess_paragraph = 'The listing agent will either promptly approve your request to bid or contact you with any questions before granting you permission to bid. Once approved you will be sent an email to '+email+' granting you approval to bid. If you do not receive the approval email within the next few hours please check your junk/spam folder first or contact the listing agent for help.'
                else:
                    doc_uploaded = 0
                    sucess_paragraph = 'You may select the Proof of Funds and/or Bank Letter of Credit option and get approved in a matter of minutes.'
                data = {'status': 200, 'msg': 'Registered Successfully.', 'error': 0, 'data': save_data, 'doc_uploaded': doc_uploaded, 'sucess_paragraph': sucess_paragraph, 'reg_id': reg_id}
            else:
                data = {'status': 403, 'msg': save_data['msg'], 'error': 1, 'data': save_data, 'reg_id': ''}
            return JsonResponse(data)
    except Exception as exp:
        print(exp)
        data = {'status': 403, 'msg': "Invalid Request", 'error': 1, 'data': "", 'reg_id': ''}
        return JsonResponse(data)        


@csrf_exempt
def save_bidder_document(request):
    try:
        user_id = request.session['user_id']
        site_detail = subdomain_site_details(request)
        site_id = site_detail['site_detail']['site_id']
        token = request.session['token']['access_token']
        file_urls = ""
        upload_to = ""
        uploaded_file_list = []
        file_size = request.POST['file_size']

        try:
            for key, value in request.FILES.items():
                params = {}
                if 'bidder_document' in key.lower():
                    upload_to = 'bidder_document'
                    doc_type = 16
                file_urls = request.FILES[key]
        except:
            pass

        if int(request.POST['file_length']) > 1:
            for key, value in request.FILES.items():
                params = {}
                upload_to = 'bidder_document'
                doc_type = 16

                file_res = request.FILES[key]
                response = save_to_s3(file_res, upload_to)
                if 'error' in response and response['error'] == 0:
                    try:
                        upload_param = {
                            "site_id": site_id,
                            "user_id": user_id,
                            "doc_file_name": response['file_name'],
                            "document_type": doc_type,
                            "bucket_name": upload_to,
                            "added_by": user_id,
                            "file_size": str(file_size) + 'MB'
                        }
                        url = settings.API_URL + '/api-users/file-upload/'
                        upload_data = call_api_post_method(upload_param, url, token)
                        upload_id = upload_data['data']['upload_id']
                        upload_size = upload_data['data']['file_size']
                        upload_date = upload_data['data']['added_date']
                    except:
                        upload_id = 0
                        upload_size = '0MB'
                        upload_date = ''

                    params['file_name'] = response['file_name']
                    params['error'] = 0
                    params['msg'] = response['msg']
                    params['upload_id'] = upload_id
                    params['file_size'] = upload_size
                    params['upload_date'] = upload_date
                    params['upload_to'] = upload_to
                else:
                    params['file_name'] = response['file_name']
                    params['error'] = 1
                    params['msg'] = response['msg']
                    params['upload_id'] = 0
                    params['file_size'] = '0MB'
                    params['upload_date'] = ''
                    params['upload_to'] = upload_to

                uploaded_file_list.append(params)
        else:
            params = {}
            response = save_to_s3(file_urls, upload_to)
            if 'error' in response and response['error'] == 0:
                try:
                    upload_param = {
                        "site_id": site_id,
                        "user_id": user_id,
                        "doc_file_name": response['file_name'],
                        "document_type": doc_type,
                        "bucket_name": upload_to,
                        "added_by": user_id,
                        "file_size": str(file_size) + 'MB'
                    }
                    url = settings.API_URL + '/api-users/file-upload/'
                    upload_data = call_api_post_method(upload_param, url, token)
                    upload_id = upload_data['data']['upload_id']
                    upload_size = upload_data['data']['file_size']
                    upload_date = upload_data['data']['added_date']
                except Exception as exp:
                    print(exp)
                    upload_id = 0
                    upload_size = '0MB'
                    upload_date = ''

                params['file_name'] = response['file_name']
                params['error'] = 0
                params['msg'] = response['msg']
                params['upload_id'] = upload_id
                params['file_size'] = upload_size
                params['upload_date'] = upload_date
                params['upload_to'] = upload_to
            else:
                params['file_name'] = response['file_name']
                params['error'] = 1
                params['msg'] = response['msg']
                params['upload_id'] = 0
                params['file_size'] = '0MB'
                params['upload_date'] = ''
                params['upload_to'] = upload_to
            uploaded_file_list.append(params)

        return JsonResponse({'status': 200, 'uploaded_file_list': uploaded_file_list})
    except Exception as exp:
        print(exp)
        data = {'status': 403, 'msg': 'invalid request.'}
        return JsonResponse(data)


@csrf_exempt
def delete_bidder_documents(request):
    try:
        return_data = {
            'section': request.POST['section'],
            'image_id': request.POST['image_id'],
            'image_name': request.POST['image_name'],
        }
        if request.is_ajax() and request.method == 'POST':
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']
            token = request.session['token']['access_token']
            image_name = request.POST['image_name']
            bucket = request.POST['section']

            params = {
                'site_id': site_id,
                'user_id': request.session['user_id'],
                'property_id': request.POST['article_id'],
                'upload_id': request.POST['image_id'],
                'upload_type': 'property_video'
            }
            url = settings.API_URL + '/api-property/subdomain-property-document-delete/'

            delete_data = call_api_post_method(params, url, token)
            user_list = []
            if 'error' in delete_data and delete_data['error'] == 0:
                try:
                    delete = delete_s3_file(image_name, setting.AWS_BUCKET_NAME + '/' + bucket)
                except Exception as exp:
                    print(exp)
                return_data['msg'] = 'Image deleted successfully'
                return_data['error'] = 0
                return_data['status'] = 200
                data = return_data
            else:
                return_data['msg'] = 'Some error occurs, please try again'
                return_data['error'] = 1
                return_data['status'] = 403
                data = return_data
        else:
            return_data['msg'] = 'Invalid request'
            return_data['error'] = 1
            return_data['status'] = 403
            data = return_data

        return JsonResponse(data)
    except Exception as exp:
        print(exp)
        data = {'status': 403, 'msg': 'invalid request.', 'user_list': []}
        return JsonResponse(data)


def buyers(request):
    try:
        try:
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']
            templete_dir = site_detail['site_detail']['theme_directory']
            company_name = site_detail['site_detail']['company_name']
        except Exception as exp:
            site_id = ""
            templete_dir = 'theme-1'
            company_name = "BIDHOM"

        static_dir = templete_dir

        try:
            params = {
                'site_id': site_id
            }
            api_url = settings.API_URL + '/api-cms/about-detail/'
            response_data = call_api_post_method(params, api_url)
            if 'error' in response_data and response_data['error'] == 0:
                data = response_data['data']
            else:
                data = []
        except Exception as exp:
            print(exp)
            data = []

        try:
            params = {
                'site_id': site_id,
                'type': 'buyer'
            }
            api_url = settings.API_URL + '/api-users/front-testimonial/'
            token = request.session['token']['access_token']
            response_data = call_api_post_method(params, api_url, token)
            if 'error' in response_data and response_data['error'] == 0:
                testimonials = response_data['data']
            else:
                testimonials = []
        except Exception as exp:
            testimonials = []
        
        user_id = None
        if 'user_id' in request.session and request.session['user_id']:
            user_id = request.session['user_id']
    
        http_host = request.META['HTTP_HOST']
        site_url = settings.URL_SCHEME + str(http_host)   

        context = {'is_home_page': True, 'data': data, 'aws_url': settings.AWS_URL, 'sess_user_id': user_id, 'SITE_URL': site_url, 'company_name': company_name, 'testimonials': testimonials}
        theme_path = 'home/{}/cms/buyers.html'.format(templete_dir)
        return render(request, theme_path, context)
    except Exception as exp:
        print(exp)
        return HttpResponse("Issue in views")


def sellers(request):
    try:
        try:
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']
            templete_dir = site_detail['site_detail']['theme_directory']
            company_name = site_detail['site_detail']['company_name']
        except Exception as exp:
            site_id = ""
            templete_dir = 'theme-1'
            company_name = "BIDHOM"

        static_dir = templete_dir

        try:
            params = {
                'site_id': site_id
            }
            api_url = settings.API_URL + '/api-cms/about-detail/'
            response_data = call_api_post_method(params, api_url)
            if 'error' in response_data and response_data['error'] == 0:
                data = response_data['data']
            else:
                data = []
        except Exception as exp:
            print(exp)
            data = []

        try:
            params = {
                'site_id': site_id,
                'type': 'seller'
            }
            api_url = settings.API_URL + '/api-users/front-testimonial/'
            token = request.session['token']['access_token']
            response_data = call_api_post_method(params, api_url, token)
            if 'error' in response_data and response_data['error'] == 0:
                testimonials = response_data['data']
            else:
                testimonials = []
        except Exception as exp:
            testimonials = []
        
        user_id = None
        if 'user_id' in request.session and request.session['user_id']:
            user_id = request.session['user_id']
    
        http_host = request.META['HTTP_HOST']
        site_url = settings.URL_SCHEME + str(http_host)   

        context = {'is_home_page': True, 'data': data, 'aws_url': settings.AWS_URL, 'sess_user_id': user_id, 'SITE_URL': site_url, 'company_name': company_name, 'testimonials': testimonials}
        theme_path = 'home/{}/cms/sellers.html'.format(templete_dir)
        return render(request, theme_path, context)
    except Exception as exp:
        print(exp)
        return HttpResponse("Issue in views")

@csrf_exempt
def delete_bidder_documents(request):
    try:
        return_data = {
            'section': request.POST['section'],
            'image_id': request.POST['image_id'],
            'image_name': request.POST['image_name'],
        }
        if request.is_ajax() and request.method == 'POST':
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']
            token = request.session['token']['access_token']
            image_name = request.POST['image_name']
            bucket = request.POST['section']

            params = {
                "domain": site_id,
                "user": request.session['user_id'],
                "registration_id": request.POST['reg_id'],
                "upload_id": request.POST['image_id']
            }
            url = settings.API_URL + '/api-bid/subdomain-delete-registration-upload/'

            delete_data = call_api_post_method(params, url, token)

            user_list = []

            if 'error' in delete_data and delete_data['error'] == 0:
                try:
                    delete = delete_s3_file(bucket+'/'+image_name, settings.AWS_BUCKET_NAME)
                except Exception as exp:
                    print(exp)
                return_data['msg'] = 'Image deleted successfully'
                return_data['error'] = 0
                return_data['status'] = 200
                data = return_data
            else:
                return_data['msg'] = 'Some error occurs, please try again'
                return_data['error'] = 1
                return_data['status'] = 403
                data = return_data
        else:
            return_data['msg'] = 'Invalid request'
            return_data['error'] = 1
            return_data['status'] = 403
            data = return_data

        return JsonResponse(data)
    except Exception as exp:
        print(exp)
        data = {'status': 403, 'msg': 'invalid request.', 'user_list': []}
        return JsonResponse(data)

@csrf_exempt
def upload_bidder_document(request):
    try:

        if request.is_ajax() and request.method == 'POST':
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']

            user_id = None
            token = None
            if 'user_id' in request.session and request.session['user_id']:
                token = request.session['token']['access_token']
                user_id = request.session['user_id']

            bidder_doc = []
            if 'is_submit_proof_fund' in request.POST and int(request.POST['is_submit_proof_fund']) == 1:
                bidder_doc_id = request.POST['bidder_doc_id']
                try:
                    bidder_doc = bidder_doc_id.split(',')
                except:
                    bidder_doc = []
            reg_id = request.POST['reg_id']
            bidder_params = {
                "domain": site_id,
                "user": user_id,
                "uploads": bidder_doc,
                "registration_id": request.POST['reg_id']
            }
            if bidder_doc:
                bidder_params['upload_pof'] = 1
            else:
                bidder_params['upload_pof'] = 0
                bidder_params['reason_for_not_upload'] = request.POST['doc_reason']

            email = request.POST['email']

            bidder_api_url = settings.API_URL + '/api-bid/bid-registration-proof-upload/'
            save_data = call_api_post_method(bidder_params, bidder_api_url, token=token)

            if 'error' in save_data and save_data['error'] == 0:
                if bidder_doc:
                    doc_uploaded = 1
                    sucess_paragraph = 'The listing agent will either promptly approve your request to bid or contact you with any questions before granting you permission to bid. Once approved you will be sent an email to '+email+' granting you approval to bid. If you do not receive the approval email within the next few hours please check your junk/spam folder first or contact the listing agent for help.'
                else:
                    doc_uploaded = 0
                    sucess_paragraph = 'You may select the Proof of Funds and/or Bank Letter of Credit option and get approved in a matter of minutes.'
                data = {'status': 200, 'msg': 'Document Uploaded Successfully.', 'error': 0, 'data': save_data, 'action': 'exit', 'sucess_paragraph': sucess_paragraph, 'doc_uploaded': doc_uploaded, 'reg_id': reg_id}
            else:
                data = {'status': 403, 'msg': save_data['msg'], 'error': 1, 'data': save_data, 'reg_id': ''}
            return JsonResponse(data)



    except Exception as exp:
        print(exp)
        return HttpResponse("Issue in views")

@csrf_exempt
def get_address_detail_by_zipcode(request):
    try:
        if request.is_ajax() and request.method == 'POST':
            zip_code = request.POST['zip_code']
            try:
                zcdb = ZipCodeDatabase()
                address = zcdb[zip_code]
                data = {'error': 0, 'status': 200, 'state': address.state, 'city': address.city, 'zip_code': zip_code}
            except Exception as exp:
                print(exp)
                data = {'error': 1, 'status': 403, 'state': '', 'city': '', 'zip_code': zip_code}
        else:
            data = {'error': 0, 'status': 403, 'msg': 'Forbidden', 'state': '', 'city': ''}

        return JsonResponse(data)
    except Exception as exp:
        print(exp)
        data = {'status': 403, 'msg': 'invalid request.'}
        return JsonResponse(data)

@csrf_exempt
def save_watch_property(request):
    try:
        if request.is_ajax() and request.method == 'POST':
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']
            token = request.session['token']['access_token']
            user_id = request.session['user_id']

            params = {
                "domain": site_id,
                "property": request.POST['property'],
                "user": user_id
            }

            api_url = settings.API_URL + '/api-property/make-watch-property/'

            data = call_api_post_method(params, api_url, token)
            if 'error' in data and data['error'] == 0:
                data = {'status': 200, 'error': 0, 'msg': 'Added to Watch List.'}
            else:
                data = {'status': 403, 'error': 1, 'msg': data['msg']}
        else:
            data = {'status': 403, 'error': 1, 'msg': 'Forbidden.'}

        return JsonResponse(data)
    except Exception as exp:
        print(exp)
        data = {'status': 403, 'error': 1, 'msg': 'Forbidden.'}
        return JsonResponse(data)

@csrf_exempt
def save_schedule_tour(request):
    try:
        if request.is_ajax() and request.method == 'POST':
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']
            token = request.session['token']['access_token']
            user_id = request.session['user_id']
            params = {
                "domain": site_id,
                "property": request.POST['tour_property_id'],
                "user": user_id,
                "schedule_date": request.POST['tour_date']+' 00:00:00',
                "message": request.POST['tour_message'],
                "tour_type": request.POST['tour_type'],
                "first_name": request.POST['user_first_name'],
                "last_name": request.POST['user_last_name'],
                "email": request.POST['user_email'],
                "phone_no": re.sub('\D', '', request.POST['user_phone']),
                "availability": request.POST['availability'],
                "property_image": request.POST['tour_property_image'],
                "property_name": request.POST['tour_property_name'],
                "property_address": request.POST['tour_property_address'],
                "property_city": request.POST['tour_property_city'],
                "property_state": request.POST['tour_property_state'],
                "property_zipcode": request.POST['tour_property_zipcode']
            }

            api_url = settings.API_URL + '/api-property/schedule-tour/'

            data = call_api_post_method(params, api_url, token)

            if 'error' in data and data['error'] == 0:
                data = {'status': 200, 'error': 0, 'msg': 'Saved Successfully.'}
            else:
                data = {'status': 403, 'error': 1, 'msg': data['msg']}
        else:
            data = {'status': 403, 'error': 1, 'msg': 'Forbidden.'}

        return JsonResponse(data)
    except Exception as exp:
        print(exp)
        data = {'status': 403, 'error': 1, 'msg': 'Forbidden.'}
        return JsonResponse(data)

@csrf_exempt
def get_tour_person_details(request):
    try:
        if request.is_ajax() and request.method == 'POST':
            try:
                site_detail = subdomain_site_details(request)
                site_id = site_detail['site_detail']['site_id']
                templete_dir = site_detail['site_detail']['theme_directory']
            except Exception as exp:
                site_id = ""
                templete_dir = 'theme-1'

            static_dir = templete_dir


            params = {
                'user': request.session['user_id'],
                'domain': site_id
            }
            token = request.session['token']['access_token']
            url = settings.API_URL + '/api-property/schedule-tour-detail/'
            user_data = call_api_post_method(params, url, token)
            if 'error' in user_data and user_data['error'] == 0:
                user_info_data = user_data['data']
                # context = {
                #     'first_name': user_info_data['first_name'],
                #     'last_name': user_info_data['last_name'],
                #     'email': user_info_data['email'],
                #     'phone_no': user_info_data['phone_no'],
                # }
                # info_data_path = 'home/{}/listings/tour-person-info.html'.format(templete_dir)
                # user_info_template = get_template(info_data_path)
                # user_info_html = user_info_template.render(context)
                data = {'error': 0, 'status': 200, 'user_info_data': user_info_data}
            else:

                data = {'error': 1, 'status': 403, 'user_info_data': {}}

        else:
            data = {'error': 1,'status': 403, 'msg': 'Forbidden', 'user_info_data': {}}

        return JsonResponse(data)
    except Exception as exp:
        print(exp)
        data = {'status': 403, 'msg': 'invalid request.'}
        return JsonResponse(data)

@csrf_exempt
def track_advertisement(request):
    try:

        if request.is_ajax() and request.method == 'POST':
            try:
                site_detail = subdomain_site_details(request)
                site_id = site_detail['site_detail']['site_id']
            except:
                site_id = ''

            try:
                token = request.session['token']['access_token']
                user_id = request.session['user_id']
            except:
                token = None
                user_id = None

            params = {
                "domain": site_id,
                "property": request.POST['property_id'],
                "user": user_id,
                "advertisement": request.POST['advertisement'],
            }
            api_url = settings.API_URL + '/api-advertisement/track-advertisement/'

            data = call_api_post_method(params, api_url, token)

            if 'error' in data and data['error'] == 0:
                data = {'status': 200, 'error': 0, 'msg': 'Saved Successfully.'}
            else:
                data = {'status': 403, 'error': 1, 'msg': data['msg']}
        else:
            data = {'status': 403, 'error': 1, 'msg': 'Forbidden.'}

        return JsonResponse(data)
    except Exception as exp:
        print(exp)
        data = {'status': 403, 'error': 1, 'msg': 'Forbidden.'}
        return JsonResponse(data)

@csrf_exempt
def view_document(request):
    try:
        property = request.GET.get('property', None)
        document = request.GET.get('document', None)
        try:
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']
            templete_dir = site_detail['site_detail']['theme_directory']
        except Exception as exp:
            print(exp)
            site_id = ""
            templete_dir = 'theme-1'
        static_dir = templete_dir

        try:
            token = request.session['token']['access_token']
            user_id = request.session['user_id']
        except:
            token = None
            user_id = None

        params = {
            "domain": site_id,
            "property": property,
            "user": user_id,
            "upload_id": document,
        }

        api_url = settings.API_URL + '/api-property/property-upload-detail/'

        data = call_api_post_method(params, api_url, token)

        if 'error' in data and data['error'] == 0:
            data = data['data']
        else:
            data = {}

        context = data
        theme_path = 'home/{}/listings/view-document.html'.format(templete_dir)
        return render(request, theme_path, context)
    except Exception as exp:
        print(exp)
        return HttpResponse("Issue in views")

@csrf_exempt
def track_document_visitor(request):
    try:

        if request.is_ajax() and request.method == 'POST':
            try:
                site_detail = subdomain_site_details(request)
                site_id = site_detail['site_detail']['site_id']
            except:
                site_id = ''

            try:
                token = request.session['token']['access_token']
                user_id = request.session['user_id']
            except:
                token = None
                user_id = None

            property_id = request.POST['property_id']
            document = request.POST['document']
            upload_id = request.POST['upload_id']
            doc_url = request.POST['doc_url']
            params = {
                "domain": site_id,
                "property": request.POST['property_id'],
                "user": user_id,
                "documents": request.POST['document'],
            }

            api_url = settings.API_URL + '/api-property/document-vault-visit/'

            data = call_api_post_method(params, api_url, token)

            if 'error' in data and data['error'] == 0:
                data = {
                    'status': 200,
                    'error': 0,
                    'msg': 'Saved Successfully.',
                    'property_id': property_id,
                    'document': document,
                    'upload_id': upload_id,
                    'doc_url': doc_url,
                }
            else:
                data = {
                    'status': 403,
                    'error': 1,
                    'msg': data['msg'],
                    'property_id': property_id,
                    'document': document,
                    'upload_id': upload_id,
                    'doc_url': doc_url,
                }
        else:
            data = {
                'status': 403,
                'error': 1,
                'msg': 'Forbidden.'
            }

        return JsonResponse(data)
    except Exception as exp:
        print(exp)
        data = {'status': 403, 'error': 1, 'msg': 'Forbidden.'}
        return JsonResponse(data)


@csrf_exempt
def node(request):
    try:
        try:
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']
            templete_dir = site_detail['site_detail']['theme_directory']
        except Exception as exp:
            site_id = ""
            templete_dir = 'theme-1'
        context = {"node_url": settings.NODE_URL}
        theme_path = 'home/{}/listings/node.html'.format(templete_dir)
        return render(request, theme_path, context)
    except Exception as exp:
        return HttpResponse("Issue in views")


@csrf_exempt
def node_one(request):
    try:
        try:
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']
            templete_dir = site_detail['site_detail']['theme_directory']
        except Exception as exp:
            site_id = ""
            templete_dir = 'theme-1'
        context = {"node_url": settings.NODE_URL}
        theme_path = 'home/{}/listings/node_1.html'.format(templete_dir)
        return render(request, theme_path, context)
    except Exception as exp:
        return HttpResponse("Issue in views")

@csrf_exempt
def delete_profile_image(request):
    try:
        if request.is_ajax() and request.method == 'POST':
            try:
                site_detail = subdomain_site_details(request)
                site_id = site_detail['site_detail']['site_id']
                templete_dir = site_detail['site_detail']['theme_directory']
            except:
                site_id = ''
                templete_dir = 'theme-1'

            try:
                token = request.session['token']['access_token']
                user_id = request.session['user_id']
            except:
                token = None
                user_id = None


            section = request.POST['section']
            upload_id = request.POST['image_id']
            image_name = request.POST['image_name']
            return_data = {}
            params = {
                'site_id': site_id,
                'user_id': user_id,
                'upload_id': upload_id,
                'upload_type': 'profile_image'
            }

            url = settings.API_URL + '/api-users/delete-file/'

            delete_data = call_api_post_method(params, url, token)
            user_list = []
            if 'error' in delete_data and delete_data['error'] == 0:
                try:
                    delete = delete_s3_file(section + '/' + image_name, settings.AWS_BUCKET_NAME)

                except Exception as exp:
                    print(exp)

                try:
                    api_url = settings.API_URL + '/api-users/login-details/'
                    payload = {
                        'token': token,
                        'site_id': site_id
                    }
                    response = call_api_post_method(payload, api_url)
                    if "error" in response and response['error'] == 0:
                        response = response['data']
                        request.session.modified = True
                        request.session['profile_image'] = response['profile_image']
                except:
                    pass
                return_data['msg'] = 'Image deleted successfully'
                return_data['error'] = 0
                return_data['status'] = 200
                data = return_data
            else:
                return_data['msg'] = 'Some error occurs, please try again'
                return_data['error'] = 1
                return_data['status'] = 403
                data = return_data
        else:
            return_data = {
                'status': 403,
                'error': 1,
                'msg': 'Forbidden.'
            }
            data = return_data
        data['templete_dir'] = templete_dir
        return JsonResponse(data)
    except Exception as exp:
        print(exp)
        data = {'status': 403, 'error': 1, 'msg': 'Forbidden.'}
        return JsonResponse(data)

@csrf_exempt
def get_property_bids(request):
    try:
        if request.is_ajax() and request.method == 'POST':
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']
            try:
                token = request.session['token']['access_token']
                user_id = request.session['user_id']
                templete_dir = site_detail['site_detail']['theme_directory']
            except:
                token = None
                user_id = None
                templete_dir = 'theme-1'


            params = {
                "domain_id": site_id,
                "property_id": request.POST['property'],
            }

            api_url = settings.API_URL + '/api-bid/bid-history/'

            bidder_data = call_api_post_method(params, api_url, token)
            if 'error' in bidder_data and bidder_data['error'] == 0:
                bidder_listing = bidder_data['data']['data']
                context = {'bidder_listing': bidder_listing, "aws_url": settings.AWS_URL, "user_id": request.session['user_id']}

                bidder_listing_path = 'home/{}/listings/property-bidder-listing-content.html'.format(templete_dir)
                bidder_listing_template = get_template(bidder_listing_path)
                bid_history_html = bidder_listing_template.render(context)
                data = {'status': 200, 'error': 0, 'bid_history_html': bid_history_html}
            else:
                data = {'status': 403, 'error': 1, 'msg': bidder_data['msg'],'bid_history_html': ''}
        else:
            data = {'status': 403, 'error': 1, 'msg': 'Forbidden.', 'bid_history_html': ''}

        return JsonResponse(data)
    except Exception as exp:
        print(exp)
        data = {'status': 403, 'error': 1, 'msg': 'Forbidden.'}
        return JsonResponse(data)

@csrf_exempt
def get_offer_details(request):
    try:
        if request.is_ajax() and request.method == 'POST':
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']
            user_id = request.session['user_id']
            params = {
                "domain_id": site_id,
                "property_id": request.POST['property_id'],
                "user_id": user_id
            }
            token = request.session['token']['access_token']
            url = settings.API_URL + '/api-bid/offer-detail/'
            data = call_api_post_method(params, url, token)
        else:
            data = {'status': 403, 'msg': 'Forbidden'}

        return JsonResponse(data)
    except Exception as exp:
        print(exp)
        data = {'status': 403, 'msg': 'invalid request.'}
        return JsonResponse(data)

@csrf_exempt
def save_traditional_offer(request):
    try:
        if request.is_ajax() and request.method == 'POST':
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']
            token = request.session['token']['access_token']
            user_id = request.session['user_id']
            if 'user_offer' in request.POST and request.POST['user_offer'] != "":
                user_offer = request.POST['user_offer'].replace(',', '').replace('$', '')

            if 'offer_doc_id' in request.POST and request.POST['offer_doc_id']:
                offer_doc_id = request.POST['offer_doc_id']
                try:
                    offer_doc = offer_doc_id.split(',')

                except:
                    offer_doc = []
            else:
                offer_doc = []

            params = {
                "domain_id": site_id,
                "first_name": request.POST['first_name'],
                "last_name": request.POST['last_name'],
                "email": request.POST['trad_email'],
                "phone_no": re.sub('\D', '', request.POST['trad_phone']),
                "address_first": request.POST['address_1'],
                "city": request.POST['city'],
                "state": request.POST['state'],
                "postal_code": request.POST['zip_code'],
                "property_id": request.POST['property_id'],
                "offer_price": user_offer,
                "comment": request.POST['offer_comment'],
                "user_id": user_id,
                "user_type": request.POST['trad_user_type'],
                "document_id": offer_doc,
                "country": request.POST['traditional_country'],
            }

            if 'trad_user_type' in request.POST and int(request.POST['trad_user_type']) == 2:
                if 'working_with_agent' in request.POST and int(request.POST['working_with_agent']) == 0:
                    params['working_with_agent'] = False

                if 'property_in_person' in request.POST and request.POST['property_in_person'] != "":
                    params['property_in_person'] = request.POST['property_in_person']

                if 'buyer_pre_qualified' in request.POST and request.POST['buyer_pre_qualified'] != "":
                    params['pre_qualified_lender'] = request.POST['buyer_pre_qualified']

            else:
                if 'agent_property_in_person' in request.POST and request.POST['agent_property_in_person'] != "":
                    params['property_in_person'] = request.POST['agent_property_in_person']

                if 'agent_pre_qualified' in request.POST and request.POST['agent_pre_qualified'] != "":
                    params['pre_qualified_lender'] = request.POST['agent_pre_qualified']

            api_url = settings.API_URL + '/api-bid/buyer-make-offer/'

            data = call_api_post_method(params, api_url, token)
            if 'error' in data and data['error'] == 0:
                data = {'status': 200, 'error': 0, 'msg': 'Offer Submitted Successfully.'}
            else:
                data = {'status': 403, 'error': 1, 'msg': data['msg']}
        else:
            data = {'status': 403, 'error': 1, 'msg': 'Forbidden.'}

        return JsonResponse(data)
    except Exception as exp:
        print(exp)
        data = {'status': 403, 'error': 1, 'msg': 'Forbidden.'}
        return JsonResponse(data)

@csrf_exempt
def check_agent_user(request):
    try:
        if request.is_ajax() and request.method == 'POST':
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']
            token = request.session['token']['access_token']
            user_id = request.session['user_id']

            params = {
                "domain_id": site_id,
                "user_id": user_id,
            }
            api_url = settings.API_URL + '/api-bid/check-agent/'

            data = call_api_post_method(params, api_url, token)
            if 'error' in data and data['error'] == 0:
                data = {'status': 200, 'error': 0, 'data': data}
            else:
                data = {'status': 403, 'error': 1, 'data': data}
        else:
            data = {'status': 403, 'error': 1, 'msg': 'Forbidden.'}

        return JsonResponse(data)
    except Exception as exp:
        print(exp)
        data = {'status': 403, 'error': 1, 'msg': 'Forbidden.'}
        return JsonResponse(data)

@csrf_exempt
def my_offers(request):
    try:
        try:
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']
            templete_dir = site_detail['site_detail']['theme_directory']

        except Exception as exp:
            print(exp)
            site_id = ""
            templete_dir = 'theme-1'

        static_dir = templete_dir

        token = request.session['token']['access_token']
        user_id = request.session['user_id']
        page_size = 10
        if request.is_ajax() and request.method == 'POST':
            search = ''
            if 'search' in request.POST and request.POST['search']:
                search = request.POST['search']
            page = 1
            if 'page' in request.POST and request.POST['page'] != "":
                page = request.POST['page']

            if 'page_size' in request.POST and request.POST['page_size'] != "":
                page_size = request.POST['page_size']
            params = {
                'domain_id': site_id,
                'user_id': user_id,
                'page_size': page_size,
                'search': search,
                'page': page
            }
            sno = (int(page) - 1) * int(page_size) + 1
            api_url = settings.API_URL + '/api-bid/my-offer-listing/'
            offer_data = call_api_post_method(params, api_url, token=token)

            if 'error' in offer_data and offer_data['error'] == 0:
                offer_listing = offer_data['data']['data']
                total = offer_data['data']['total']
            else:
                offer_listing = []
                total = 0

            context = {'offer_listing': offer_listing, 'total': total, "active_menu": "traditional_offer", "aws_url": settings.AWS_URL, 'sno': sno}

            offer_listing_path = 'home/{}/user-dashboard/offers/offer-listing.html'.format(templete_dir)
            offer_listing_template = get_template(offer_listing_path)
            offer_listing_html = offer_listing_template.render(context)
            # ---------------Pagination--------
            pagination_path = 'home/{}/user-dashboard/offers/pagination.html'.format(templete_dir)
            pagination_template = get_template(pagination_path)
            total_page = math.ceil(int(total) / int(page_size))
            pagination_data = {"no_page": int(total_page), "total_page": range(total_page), "current_page": int(page)}
            pagination_html = pagination_template.render(pagination_data)
            data = {'offer_listing_html': offer_listing_html, 'status': 200, 'msg': '', 'error': 0, 'total': total, "pagination_html": pagination_html}
            return JsonResponse(data)
        else:
            page = 1
            params = {
                'domain_id': site_id,
                'user_id': user_id,
                'page_size': page_size,
                'search': '',
                'page': page
            }
            api_url = settings.API_URL + '/api-bid/my-offer-listing/'
            offer_data = call_api_post_method(params, api_url, token=token)

            if 'error' in offer_data and offer_data['error'] == 0:
                offer_listing = offer_data['data']['data']
                total = offer_data['data']['total']
            else:
                offer_listing = []
                total = 0
            sno = (int(page) - 1) * int(page_size) + 1
            # ---------------Pagination--------
            pagination_path = 'home/{}/user-dashboard/offers/pagination.html'.format(templete_dir)
            pagination_template = get_template(pagination_path)
            total_page = math.ceil(total/page_size)
            pagination_data = {"no_page": total_page, "total_page": range(total_page), "current_page": 1}
            pagination_html = pagination_template.render(pagination_data)

            x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
            if x_forwarded_for:
                ip_address = x_forwarded_for.split(',')[0]
            else:
                ip_address = request.META.get('REMOTE_ADDR')

            context = {'offer_listing': offer_listing, 'total': total, "active_menu": "traditional_offer", "pagination_html": pagination_html, "sno":sno, "node_url": settings.NODE_URL, "user_id": user_id, "domain_id": site_id, "ip_address": ip_address}

            theme_path = 'home/{}/user-dashboard/offers/my-offers.html'.format(templete_dir)
            return render(request, theme_path, context)
    except Exception as exp:
        print(exp)
        return HttpResponse("Issue in views")

@csrf_exempt
def property_offer_details(request):
    try:
        if request.is_ajax() and request.method == 'POST':
            try:
                site_detail = subdomain_site_details(request)
                site_id = site_detail['site_detail']['site_id']
                templete_dir = site_detail['site_detail']['theme_directory']

            except Exception as exp:
                print(exp)
                site_id = ""
                templete_dir = 'theme-1'

            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']
            token = request.session['token']['access_token']
            user_id = request.session['user_id']

            params = {
                "domain_id": site_id,
                "property_id": request.POST['property_id'],
                "user_id": user_id
            }
            from_page = request.POST['from_page']
            api_url = settings.API_URL + '/api-bid/buyer-offer-detail/'

            offer_data = call_api_post_method(params, api_url, token)

            if 'error' in offer_data and offer_data['error'] == 0:
                offer_details = offer_data['data']['data']

                context = {'offer': offer_details, "aws_url": settings.AWS_URL, "user_id": user_id}
                if from_page == 'asset_details':
                    offer_details_path = 'home/{}/listings/offer_details_content.html'.format(templete_dir)
                else:
                    offer_details_path = 'home/{}/user-dashboard/offers/offer_details_content.html'.format(templete_dir)
                offer_details_template = get_template(offer_details_path)
                offer_details_html = offer_details_template.render(context)
            else:
                offer_details_html = ''

            data = {'status': 200, 'offer_details_html': offer_details_html, 'error': 0}
        else:
            data = {'status': 403, 'offer_details_html': '', 'error': 1}

        return JsonResponse(data)
    except Exception as exp:
        print(exp)
        data = {'status': 403, 'offer_details_html': '', 'error': 1}
        return JsonResponse(data)


@csrf_exempt
def buyer_accept_offer(request):
    try:
        site_detail = subdomain_site_details(request)
        site_id = site_detail['site_detail']['site_id']
        token = request.session['token']['access_token']
        user_id = request.session['user_id']

        if request.is_ajax() and request.method == 'POST':
            params = {
                'domain_id': site_id,
                'user_id': user_id,
                'property_id': request.POST['property_id'],
                'negotiation_id': request.POST['negotiated_id'],
            }
            url = settings.API_URL + '/api-bid/buyer-accept-offer/'

            response = call_api_post_method(params, url, token)

            if 'error' in response and response['error'] == 0:
                data = {
                    'error': 0,
                    'msg': "Offer Accepted successfully."
                }
            else:
                data = {
                    'error': 1,
                    'msg': response['msg']
                }
            return JsonResponse(data)
    except Exception as exp:
        print(exp)
        return HttpResponse("Issue in views")

@csrf_exempt
def buyer_reject_offer(request):
    try:
        site_detail = subdomain_site_details(request)
        site_id = site_detail['site_detail']['site_id']
        token = request.session['token']['access_token']
        user_id = request.session['user_id']

        if request.is_ajax() and request.method == 'POST':
            params = {
                'domain_id': site_id,
                'user_id': user_id,
                'property_id': request.POST['property_id'],
                'negotiation_id': request.POST['negotiated_id'],
                'declined_reason': request.POST['reason'] if 'reason' in request.POST and request.POST['reason'] != "" else "",
            }
            url = settings.API_URL + '/api-bid/buyer-reject-offer/'

            response = call_api_post_method(params, url, token)

            if 'error' in response and response['error'] == 0:
                data = {
                    'error': 0,
                    'msg': "Offer Rejected successfully."
                }
            else:
                data = {
                    'error': 1,
                    'msg': response['msg']
                }
            return JsonResponse(data)
    except Exception as exp:
        print(exp)
        return HttpResponse("Issue in views")

@csrf_exempt
def counter_offer_details(request):
    try:
        if request.is_ajax() and request.method == 'POST':
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']
            token = request.session['token']['access_token']
            user_id = request.session['user_id']
            offer_status = request.POST['offer_status'] if 'offer_status' in request.POST else ""

            params = {
                "domain_id": site_id,
                "property_id": request.POST['property_id'],
                "user_id": user_id
            }

            api_url = settings.API_URL + '/api-bid/buyer-offer-detail/'

            offer_data = call_api_post_method(params, api_url, token)

            if 'error' in offer_data and offer_data['error'] == 0:
                offer_details = offer_data['data']['data']
                offer = {
                    'property_id': offer_details['property'],
                    'negotiated_id': offer_details['master_offer'],
                    'offer_amount': offer_details['offer_price'],
                    'offer_status': offer_status,
                }
                data = {'status': 200, 'offer': offer, 'error': 0}
            else:
                data = {'status': 200, 'offer': {}, 'error': 1,'msg': offer_data['msg']}


        else:
            data = {'status': 403, 'offer_details': {}, 'error': 1, 'msg': 'forbidden'}

        return JsonResponse(data)
    except Exception as exp:
        print(exp)
        data = {'status': 403, 'offer_details': {}, 'error': 1, 'msg': 'forbidden'}
        return JsonResponse(data)

@csrf_exempt
def buyer_counter_offer(request):
    try:
        site_detail = subdomain_site_details(request)
        site_id = site_detail['site_detail']['site_id']
        token = request.session['token']['access_token']
        user_id = request.session['user_id']

        if request.is_ajax() and request.method == 'POST':
            if 'offer_price' in request.POST and request.POST['offer_price']:
                offer_price = float(request.POST['offer_price'].replace(',', '').replace('$', ''))
            else:
                offer_price = ''

            params = {
                "domain_id": site_id,
                "property_id": request.POST['property_id'],
                "negotiation_id": request.POST['negotiated_id'],
                "offer_price": offer_price,
                "comment": request.POST['offer_comment'],
                "user_id": user_id
            }
            url = settings.API_URL + '/api-bid/buyer-make-offer/'

            response = call_api_post_method(params, url, token)

            if 'error' in response and response['error'] == 0:
                data = {
                    'error': 0,
                    'msg': "Counter Offer successfully."
                }
            else:
                data = {
                    'error': 1,
                    'msg': response['msg']
                }
            return JsonResponse(data)
    except Exception as exp:
        print(exp)
        return HttpResponse("Issue in views")

@csrf_exempt
def save_offer_document(request):
    try:
        user_id = request.session['user_id']
        site_detail = subdomain_site_details(request)
        site_id = site_detail['site_detail']['site_id']
        token = request.session['token']['access_token']
        file_urls = ""
        upload_to = ""
        uploaded_file_list = []
        file_size = request.POST['file_size']

        try:
            for key, value in request.FILES.items():
                params = {}
                if 'offer_document' in key.lower():
                    upload_to = 'banner'
                    doc_type = 18
                file_urls = request.FILES[key]
        except:
            pass

        if int(request.POST['file_length']) > 1:
            for key, value in request.FILES.items():
                params = {}
                upload_to = 'banner'
                doc_type = 18

                file_res = request.FILES[key]
                response = save_to_s3(file_res, upload_to)
                if 'error' in response and response['error'] == 0:
                    try:
                        upload_param = {
                            "site_id": site_id,
                            "user_id": user_id,
                            "doc_file_name": response['file_name'],
                            "document_type": doc_type,
                            "bucket_name": upload_to,
                            "added_by": user_id,
                            "file_size": str(file_size) + 'MB'
                        }
                        url = settings.API_URL + '/api-users/file-upload/'
                        upload_data = call_api_post_method(upload_param, url, token)
                        upload_id = upload_data['data']['upload_id']
                        upload_size = upload_data['data']['file_size']
                        upload_date = upload_data['data']['added_date']
                    except:
                        upload_id = 0
                        upload_size = '0MB'
                        upload_date = ''

                    params['file_name'] = response['file_name']
                    params['error'] = 0
                    params['msg'] = response['msg']
                    params['upload_id'] = upload_id
                    params['file_size'] = upload_size
                    params['upload_date'] = upload_date
                    params['upload_to'] = upload_to
                else:
                    params['file_name'] = response['file_name']
                    params['error'] = 1
                    params['msg'] = response['msg']
                    params['upload_id'] = 0
                    params['file_size'] = '0MB'
                    params['upload_date'] = ''
                    params['upload_to'] = upload_to

                uploaded_file_list.append(params)
        else:
            params = {}
            response = save_to_s3(file_urls, upload_to)
            if 'error' in response and response['error'] == 0:
                try:
                    upload_param = {
                        "site_id": site_id,
                        "user_id": user_id,
                        "doc_file_name": response['file_name'],
                        "document_type": doc_type,
                        "bucket_name": upload_to,
                        "added_by": user_id,
                        "file_size": str(file_size) + 'MB'
                    }
                    url = settings.API_URL + '/api-users/file-upload/'
                    upload_data = call_api_post_method(upload_param, url, token)
                    upload_id = upload_data['data']['upload_id']
                    upload_size = upload_data['data']['file_size']
                    upload_date = upload_data['data']['added_date']
                except Exception as exp:
                    print(exp)
                    upload_id = 0
                    upload_size = '0MB'
                    upload_date = ''

                params['file_name'] = response['file_name']
                params['error'] = 0
                params['msg'] = response['msg']
                params['upload_id'] = upload_id
                params['file_size'] = upload_size
                params['upload_date'] = upload_date
                params['upload_to'] = upload_to
            else:
                params['file_name'] = response['file_name']
                params['error'] = 1
                params['msg'] = response['msg']
                params['upload_id'] = 0
                params['file_size'] = '0MB'
                params['upload_date'] = ''
                params['upload_to'] = upload_to
            uploaded_file_list.append(params)

        return JsonResponse({'status': 200, 'uploaded_file_list': uploaded_file_list})
    except Exception as exp:
        print(exp)
        data = {'status': 403, 'msg': 'invalid request.'}
        return JsonResponse(data)


@csrf_exempt
def delete_offer_document(request):
    try:
        return_data = {
            'section': request.POST['section'],
            'image_id': request.POST['upload_id'],
            'image_name': request.POST['doc_name'],
        }
        if request.is_ajax() and request.method == 'POST':
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']
            token = request.session['token']['access_token']
            doc_name = request.POST['doc_name']
            bucket = request.POST['section']

            params = {
                "domain": site_id,
                "user_id": request.session['user_id'],
                "upload_type": "offer_upload",
                "upload_id": request.POST['upload_id']
            }

            url = settings.API_URL + '/api-users/delete-file/'

            delete_data = call_api_post_method(params, url, token)

            user_list = []

            if 'error' in delete_data and delete_data['error'] == 0:
                try:
                    delete = delete_s3_file(bucket + '/' + doc_name, settings.AWS_BUCKET_NAME)
                except Exception as exp:
                    print(exp)
                return_data['msg'] = 'Document deleted successfully'
                return_data['error'] = 0
                return_data['status'] = 200
                data = return_data
            else:
                return_data['msg'] = 'Some error occurs, please try again'
                return_data['error'] = 1
                return_data['status'] = 403
                data = return_data
        else:
            return_data['msg'] = 'Invalid request'
            return_data['error'] = 1
            return_data['status'] = 403
            data = return_data

        return JsonResponse(data)
    except Exception as exp:
        print(exp)
        data = {'status': 403, 'msg': 'invalid request.', 'user_list': []}
        return JsonResponse(data)

@csrf_exempt
def get_notification_count(request):
    try:
        if request.is_ajax() and request.method == 'POST':
            try:
                site_detail = subdomain_site_details(request)
                site_id = site_detail['site_detail']['site_id']
            except Exception as exp:
                print(exp)
                site_id = ""

            user_id = None
            token = None
            if 'user_id' in request.session and request.session['user_id']:
                token = request.session['token']['access_token']
                user_id = request.session['user_id']

            params = {
                "domain": site_id,
                "user_id": user_id
            }
            api_url = settings.API_URL + '/api-notifications/notification-count/'
            notification_data = call_api_post_method(params, api_url, token=token)
            if 'error' in notification_data and notification_data['error'] == 0:
                notification = notification_data['data']
                data = {'notification_count': notification['count'], 'error': 0, 'msg':'success'}
            else:
                chat = {}
                data = {'notification_count': 0, 'error': 1, 'msg':'failure'}

            return JsonResponse(data)

    except Exception as exp:
        print(exp)
        return HttpResponse("Issue in views")

@csrf_exempt
def get_notification_details(request):
    try:
        if request.is_ajax() and request.method == 'POST':
            try:
                site_detail = subdomain_site_details(request)
                site_id = site_detail['site_detail']['site_id']
                templete_dir = site_detail['site_detail']['theme_directory']

            except Exception as exp:
                print(exp)
                site_id = ""
                templete_dir = 'theme-1'

            user_id = None
            token = None
            if 'user_id' in request.session and request.session['user_id']:
                token = request.session['token']['access_token']
                user_id = request.session['user_id']

            params = {
                "domain": site_id,
                "user_id": user_id
            }
            api_url = settings.API_URL + '/api-notifications/notification-detail/'
            notification_data = call_api_post_method(params, api_url, token=token)
            if 'error' in notification_data and notification_data['error'] == 0:
                try:
                    url = settings.API_URL + '/api-notifications/notification-read/'
                    read_notification = call_api_post_method(params, url, token=token)
                except Exception as exp:
                    print(exp)
                notification = notification_data['data']
                context = {'notification_list': notification,"aws_url": settings.AWS_URL}

                noti_listing_path = 'layout/{}/notification_content.html'.format(templete_dir)
                noti_listing_template = get_template(noti_listing_path)
                noti_listing_html = noti_listing_template.render(context)
                data = {'notification': notification, 'noti_listing_html': noti_listing_html, 'error': 0, 'msg':'success'}
            else:
                data = {'notification': [], 'noti_listing_html': '', 'error': 1, 'msg':'failure'}

            return JsonResponse(data)

    except Exception as exp:
        print(exp)
        return HttpResponse("Issue in views")


@csrf_exempt
def notifications(request):
    try:
        try:
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']
            templete_dir = site_detail['site_detail']['theme_directory']

        except Exception as exp:
            site_id = ""
            templete_dir = 'theme-1'

        static_dir = templete_dir

        token = None
        user_id = None
        if 'user_id' in request.session and request.session['user_id']:
            user_id = request.session['user_id']
            token = request.session['token']['access_token']

        if request.is_ajax() and request.method == 'POST':

            page = 1
            if 'page' in request.POST and request.POST['page']:
                page = int(request.POST['page'])

            page_size = 10
            if 'perpage' in request.POST and request.POST['perpage']:
                page_size = int(request.POST['perpage'])

            listing_params = {
                "domain": site_id,
                "page_size": page_size,
                "page": page,
                "user_id": user_id
            }

            listing_url = settings.API_URL + '/api-notifications/notification-listing/'

            listing_data = call_api_post_method(listing_params, listing_url, token)

            http_host = request.META['HTTP_HOST']
            site_url = settings.URL_SCHEME + str(http_host)

            if 'error' in listing_data and listing_data['error'] == 0:
                notification_list = listing_data['data']['data']
                notification_total = listing_data['data']['total']
                notification_list_data_path = 'home/{}/notification/notification_listing.html'.format(templete_dir)

                listing_template = get_template(notification_list_data_path)
                listing_html = listing_template.render(
                    {'notification_list': notification_list, 'total': notification_total, 'aws_url': settings.AWS_URL, 'sess_user_id': user_id,
                     'site_url': site_url})

                pagination_path = 'home/{}/notification/pagination.html'.format(templete_dir)
                pagination_template = get_template(pagination_path)

                total_page = math.ceil(notification_total / page_size)
                pagination_data = {"no_page": total_page, "total_page": range(total_page), "current_page": page,
                                   "pagination_id": "notification_listing_pagination_list"}
                pagination_html = pagination_template.render(pagination_data)

                data = {'listing_html': listing_html, 'status': 200,
                        'msg': 'Successfully received', 'error': 0, 'total': notification_total, 'pagination_html': pagination_html}
            else:
                data = {'status': 403, 'msg': 'Server error, Please try again', 'listing_html': '', 'error': 1,
                        'total': 0}

            return JsonResponse(data)

        else:
            page = 1
            page_size = 10

            listing_params = {
                "domain": site_id,
                "page_size": page_size,
                "page": page,
                "user_id": user_id
            }

            listing_url = settings.API_URL + '/api-notifications/notification-listing/'

            listing_data = call_api_post_method(listing_params, listing_url, token)

            if 'error' in listing_data and listing_data['error'] == 0:
                notification_list = listing_data['data']['data']
                notification_total = listing_data['data']['total']
            else:
                notification_list = []
                notification_total = 0

            # ---------------Pagination--------
            pagination_path = 'home/{}/notification/pagination.html'.format(templete_dir)
            pagination_template = get_template(pagination_path)
            total_page = math.ceil(notification_total / page_size)
            pagination_data = {"no_page": total_page, "total_page": range(total_page), "current_page": page,
                               "pagination_id": "notification_listing_pagination_list"}
            pagination_html = pagination_template.render(pagination_data)

            http_host = request.META['HTTP_HOST']
            site_url = settings.URL_SCHEME + str(http_host)
            context = {
                'is_home_page': False,
                'notification_list': notification_list,
                'total': notification_total,
                'aws_url': settings.AWS_URL,
                'site_url': site_url,
                'pagination_html': pagination_html
            }
            theme_path = 'home/{}/notification/notification.html'.format(templete_dir)
            return render(request, theme_path, context)
    except Exception as exp:
        print(exp)
        return HttpResponse("Issue in views")

@csrf_exempt
def get_offer_document_details(request):
    try:
        if request.is_ajax() and request.method == 'POST':
            try:
                site_detail = subdomain_site_details(request)
                site_id = site_detail['site_detail']['site_id']
                templete_dir = site_detail['site_detail']['theme_directory']

            except Exception as exp:
                print(exp)
                site_id = ""
                templete_dir = 'theme-1'

            user_id = None
            token = None
            if 'user_id' in request.session and request.session['user_id']:
                token = request.session['token']['access_token']
                user_id = request.session['user_id']

            params = {
                "domain": site_id,
                "property_id": request.POST['property_id'],
                "negotiation_id": request.POST['negotiated_id'],
            }
            api_url = settings.API_URL + '/api-bid/get-offer-documents/'
            document_data = call_api_post_method(params, api_url, token=token)

            if 'error' in document_data and document_data['error'] == 0:
                document_list = document_data['data']
                if document_list:
                    for doc in document_list:
                        document_name = doc['doc_file_name']
                        original_doc_arr = document_name.split('_')
                        original_doc_ext = document_name.split('.')[-1]
                        doc_name_length = len(original_doc_arr)
                        # original_doc_name = document_name.split('_')[1]
                        original_doc_name = ''
                        for i in range(doc_name_length):
                            if i > 0:
                                original_doc_name = original_doc_name + '_' + original_doc_arr[i]
                        original_doc_name = original_doc_name.lstrip('_')
                        doc['original_doc_name'] = original_doc_name
                        doc['extension'] = original_doc_ext
                context = {'document_list': document_list,"aws_url": settings.AWS_URL}

                doc_listing_path = 'home/{}/listings/traditional_offer_document_content.html'.format(templete_dir)
                doc_listing_template = get_template(doc_listing_path)
                doc_listing_html = doc_listing_template.render(context)
                data = {'doc_listing_html': doc_listing_html, 'error': 0, 'msg':'success'}
            else:
                data = {'doc_listing_html': '', 'error': 1, 'msg':'failure'}

            return JsonResponse(data)

    except Exception as exp:
        print(exp)
        return HttpResponse("Issue in views")

@csrf_exempt
def submit_loi(request):
    try:
        property = request.GET.get('property_id', None)
        try:
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']
            templete_dir = site_detail['site_detail']['theme_directory']
        except Exception as exp:
            site_id = ""
            templete_dir = 'theme-1'

        static_dir = templete_dir

        token = None
        user_id = None
        if 'user_id' in request.session and request.session['user_id']:
            user_id = request.session['user_id']
            token = request.session['token']['access_token']

        # try:
        #     state_param = {}
        #     state_api_url = settings.API_URL + '/api-settings/get-state/'
        #     state_data = call_api_post_method(state_param, state_api_url, token)
        #     state_list = state_data['data']
        # except:
        #     state_list = []

        try:
            cnt = 0
            str_offer_doc_id = ''
            offer_doc_ids = ''
            str_offer_doc_name = ''
            offer_doc_name = ''
            best_offer_param = {"domain": site_id, "user": user_id, "property": property}
            best_offer_url = settings.API_URL + '/api-bid/enhanced-best-final-detail/'

            best_offer_data = call_api_post_method(best_offer_param, best_offer_url, token)

            offer_details = best_offer_data['data']
            country = offer_details['user_detail']['country'] if offer_details['user_detail']['country'] is not None else 1

            state_param = {'country_id': country}
            state_api_url = settings.API_URL + '/api-settings/get-state/'
            state_data = call_api_post_method(state_param, state_api_url, token)
            state_list = state_data['data']

            try:
                if 'property_asset' in offer_details and int(offer_details['property_asset']) == 1:
                    loan_type_list = land_loan_types
                elif 'property_asset' in offer_details and int(offer_details['property_asset']) == 2:
                    loan_type_list = commerical_loan_types
                elif 'property_asset' in offer_details and int(offer_details['property_asset']) == 3:
                    loan_type_list = residential_loan_types
                else:
                    loan_type_list = residential_loan_types
            except Exception as exp:
                print(exp)
                loan_type_list = residential_loan_types

            if 'documents' in offer_details and len(offer_details['documents']) > 0:
                for doc in offer_details['documents']:
                    document_name = doc['doc_file_name']
                    original_doc_arr = document_name.split('_')
                    original_doc_ext = document_name.split('.')[-1]
                    doc_name_length = len(original_doc_arr)
                    # original_doc_name = document_name.split('_')[1]
                    original_doc_name = ''
                    for i in range(doc_name_length):
                        if i > 0:
                            original_doc_name = original_doc_name + '_' + original_doc_arr[i]

                    if cnt == 0:
                        str_offer_doc_id = str(doc['document_id'])
                        str_offer_doc_name = str(doc['doc_file_name'])
                    else:
                        str_offer_doc_id = str_offer_doc_id+','+str(doc['document_id'])
                        str_offer_doc_name = str_offer_doc_name+','+str(doc['doc_file_name'])
                    cnt += 1

                    original_doc_name = original_doc_name.lstrip('_')
                    doc['original_doc_name'] = original_doc_name
                    doc['extension'] = original_doc_ext

                offer_doc_ids = str_offer_doc_id.rstrip(',')
                offer_doc_name = str_offer_doc_name.rstrip(',')
                offer_details['offer_doc_ids'] = offer_doc_ids
                offer_details['offer_doc_name'] = offer_doc_name
            try:
                if 'offer_detail' in offer_details and offer_details['offer_detail']:
                    offer_start_price = 0

                    if 'current_offer_price' in offer_details['offer_detail'] and \
                            offer_details['offer_detail']['current_offer_price']:
                        offer_start_price = offer_details['offer_detail']['current_offer_price']

                    if 'earnest_deposit_type' in offer_details['property'] and offer_details['property']['earnest_deposit_type'] == 1:
                        offer_earnest_deposit = offer_details['offer_detail']['earnest_money_deposit']
                        current_earnest_money_deposit = format_currency(offer_earnest_deposit)
                        offer_money_deposit_percent = (float(offer_earnest_deposit)*100)/float(offer_start_price)
                        current_earnest_money_deposit_percent = format_currency(offer_money_deposit_percent)
                    elif 'earnest_deposit_type' in offer_details['property'] and offer_details['property']['earnest_deposit_type'] == 2:
                        offer_money_deposit_percent = offer_details['offer_detail']['earnest_money_deposit']
                        current_earnest_money_deposit_percent = format_currency(offer_money_deposit_percent)
                        offer_earnest_deposit = (float(offer_start_price)*float(offer_money_deposit_percent))/float(100)
                        current_earnest_money_deposit = format_currency(offer_earnest_deposit)
                    else:
                        current_earnest_money_deposit = ''
                        current_earnest_money_deposit_percent = ''

                    offer_details['offer_detail']['current_earnest_money_deposit'] = current_earnest_money_deposit
                    offer_details['offer_detail']['current_earnest_money_deposit_percent'] = current_earnest_money_deposit_percent

            except:
                pass

            try:
                private_offer_start_price = 0
                if 'start_price' in offer_details['auction_data'] and \
                        offer_details['auction_data']['start_price']:
                    private_offer_start_price = offer_details['auction_data']['start_price']

                if 'earnest_deposit_type' in offer_details['property'] and offer_details['property'][
                    'earnest_deposit_type'] == 1:
                    private_offer_earnest_deposit = offer_details['property']['earnest_deposit']
                    private_current_earnest_money_deposit = format_currency(private_offer_earnest_deposit)
                    private_offer_money_deposit_percent = (float(private_offer_earnest_deposit) * 100) / float(
                        private_offer_start_price)
                    private_current_earnest_money_deposit_percent = format_currency(private_offer_money_deposit_percent)
                elif 'earnest_deposit_type' in offer_details['property'] and offer_details['property'][
                    'earnest_deposit_type'] == 2:
                    private_offer_money_deposit_percent = offer_details['property']['earnest_deposit']
                    private_current_earnest_money_deposit_percent = format_currency(private_offer_money_deposit_percent)
                    private_offer_earnest_deposit = (float(private_offer_start_price) * float(
                        private_offer_money_deposit_percent)) / float(100)
                    private_current_earnest_money_deposit = format_currency(private_offer_earnest_deposit)
                else:
                    private_current_earnest_money_deposit = ''
                    private_current_earnest_money_deposit_percent = ''

                offer_details['property'][
                    'private_current_earnest_money_deposit'] = private_current_earnest_money_deposit
                offer_details['property'][
                    'private_current_earnest_money_deposit_percent'] = private_current_earnest_money_deposit_percent

            except:
                pass

            try:
                if 'current_offer_detail' in offer_details and offer_details['current_offer_detail']:
                    offer_start_price = 0
                    if 'current_offer_amount' in offer_details['current_offer_detail'] and \
                            offer_details['current_offer_detail']['current_offer_amount']:
                        offer_start_price = offer_details['current_offer_detail']['current_offer_amount']

                    if 'earnest_deposit_type' in offer_details['property'] and offer_details['property']['earnest_deposit_type'] == 1:
                        offer_earnest_deposit = offer_details['current_offer_detail']['earnest_money_deposit']
                        current_earnest_money_deposit = format_currency(offer_earnest_deposit)
                        offer_money_deposit_percent = (float(offer_earnest_deposit)*100)/float(offer_start_price)
                        current_earnest_money_deposit_percent = format_currency(offer_money_deposit_percent)
                    elif 'earnest_deposit_type' in offer_details['property'] and offer_details['property']['earnest_deposit_type'] == 2:
                        offer_money_deposit_percent = offer_details['current_offer_detail']['earnest_money_deposit']
                        current_earnest_money_deposit_percent = format_currency(offer_money_deposit_percent)
                        offer_earnest_deposit = (float(offer_start_price)*float(offer_money_deposit_percent))/float(100)
                        current_earnest_money_deposit = format_currency(offer_earnest_deposit)
                    else:
                        current_earnest_money_deposit = ''
                        current_earnest_money_deposit_percent = ''

                    offer_details['current_offer_detail']['current_earnest_money_deposit'] = current_earnest_money_deposit
                    offer_details['current_offer_detail']['current_earnest_money_deposit_percent'] = current_earnest_money_deposit_percent
            except:
                pass



            try:
                if 'property' in offer_details and offer_details['property']:
                    prop_start_price = 0
                    if 'auction_data' in offer_details and offer_details['auction_data']['start_price']:
                        prop_start_price = offer_details['auction_data']['start_price']

                    if 'earnest_deposit_type' in offer_details['property'] and offer_details['property'][
                        'earnest_deposit_type'] == 1:
                        earnest_deposit = offer_details['property']['earnest_deposit']
                        prop_earnest_money_deposit = format_currency(earnest_deposit)
                        prop_money_deposit_percent = (float(earnest_deposit) * 100) / float(prop_start_price)
                        prop_money_deposit_percent = format_currency(prop_money_deposit_percent)
                    elif 'earnest_deposit_type' in offer_details['property'] and offer_details['property'][
                        'earnest_deposit_type'] == 2:
                        money_deposit_percent = offer_details['property']['earnest_deposit']
                        prop_money_deposit_percent = format_currency(money_deposit_percent)
                        earnest_deposit = (float(prop_start_price) * float(money_deposit_percent)) / float(
                            100)
                        prop_earnest_money_deposit = format_currency(earnest_deposit)
                    else:
                        prop_money_deposit_percent = ''
                        prop_earnest_money_deposit = ''

                    offer_details['property']['prop_earnest_money_deposit'] = prop_earnest_money_deposit
                    offer_details['property'][
                        'prop_money_deposit_percent'] = prop_money_deposit_percent

            except Exception as exp:
                print(exp)



            if offer_details['property']['status'] != 1:
                http_host = request.META['HTTP_HOST']
                redirect_url = settings.URL_SCHEME + str(http_host)
                return HttpResponseRedirect(redirect_url)

        except:
            offer_details = {}
            loan_type_list = []
            state_list = []

        try:
            params = {
                "domain_id": site_id,
                "user_id": user_id,
            }
            api_url = settings.API_URL + '/api-bid/check-agent/'

            data = call_api_post_method(params, api_url, token)
            if 'error' in data and data['error'] == 0:
                offer_details['is_agent'] = True
            else:
                offer_details['is_agent'] = False
        except:
            offer_details['is_agent'] = False


        if request.is_ajax() and request.method == 'POST':
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']
            user_id = request.session['user_id']
            token = request.session['token']['access_token']

            user_phone_no = ''
            if 'user_phone_no' in request.POST and request.POST['user_phone_no'] != "":
                user_phone_no = re.sub('\D', '', request.POST['user_phone_no'])

            offer_doc = []
            if 'offer_doc_id' in request.POST and request.POST['offer_doc_id']:
                offer_doc_id = request.POST['offer_doc_id']
                try:
                    offer_doc = offer_doc_id.split(',')
                except:
                    offer_doc = []

            offer_price = None
            if 'offer_price' in request.POST and request.POST['offer_price'] != "":
                offer_price = float(request.POST['offer_price'].replace(',', '').replace('$', ''))

            earnest_deposit = None
            if 'earnest_deposit' in request.POST and request.POST['earnest_deposit'] != "":
                earnest_deposit = float(request.POST['earnest_deposit'].replace(',', '').replace('$', ''))

            due_diligence = None
            if 'due_diligence' in request.POST and request.POST['due_diligence'] != "":
                due_diligence = request.POST['due_diligence'].replace(',', '')

            closing_period = None
            if 'closing_period' in request.POST and request.POST['closing_period'] != "":
                closing_period = request.POST['closing_period'].replace(',', '')


            offer_param = {
                "domain_id": site_id,
                "user_id": user_id,
                "property_id": request.POST['property_id'],
                "offer_price": offer_price,
                "first_name": request.POST['first_name'],
                "last_name": request.POST['last_name'],
                "email": request.POST['user_email'],
                "address_first": request.POST['address_1'],
                "city": request.POST['city'],
                "state": request.POST['state'],
                "phone_no": user_phone_no,
                "postal_code": request.POST['zip_code'],
                "earnest_money_deposit": earnest_deposit,
                "due_diligence_period": due_diligence,
                "closing_period": closing_period,
                "financing": request.POST['financing'],
                "document_id": offer_doc,
                "user_type": request.POST['trad_user_type'],
                "sale_contingency": request.POST['sale_contingency'],
                "offer_contingent": request.POST['offer_contingent'],
                "offer_comment": request.POST['offer_comment'],
                "is_update": 1 if request.POST['is_update'] else None,
            }
            if 'trad_user_type' in request.POST and int(request.POST['trad_user_type']) == 2:
                if 'working_with_agent' in request.POST and int(request.POST['working_with_agent']) == 0:
                    offer_param['working_with_agent'] = False

                if 'property_in_person' in request.POST and request.POST['property_in_person'] != "":
                    offer_param['property_in_person'] = request.POST['property_in_person']

                if 'buyer_pre_qualified' in request.POST and request.POST['buyer_pre_qualified'] != "":
                    offer_param['pre_qualified_lender'] = request.POST['buyer_pre_qualified']

            else:
                if 'agent_property_in_person' in request.POST and request.POST['agent_property_in_person'] != "":
                    offer_param['property_in_person'] = request.POST['agent_property_in_person']

                if 'agent_pre_qualified' in request.POST and request.POST['agent_pre_qualified'] != "":
                    offer_param['pre_qualified_lender'] = request.POST['agent_pre_qualified']
            offer_url = settings.API_URL + '/api-bid/buyer-make-loi/'
            offer_data = call_api_post_method(offer_param, offer_url, token)

            if 'error' in offer_data and offer_data['error'] == 0:
                data = {"status": 200, "msg": "Submitted Successfully.", "error": 0, "data": offer_data}
            else:
                data = {"status": 400, "msg": offer_data['msg'], "error": 1, "data": offer_data}

            return JsonResponse(data)


        http_host = request.META['HTTP_HOST']
        site_url = settings.URL_SCHEME + str(http_host)
        try:
            country_param = {}
            country_api_url = settings.API_URL + '/api-settings/get-country/'
            country_data = call_api_post_method(country_param, country_api_url)
            country_list = country_data['data']
        except:
            country_list = []

        context = {
            'is_home_page': False,
            'aws_url': settings.AWS_URL,
            'site_url': site_url,
            'state_list': state_list,
            'offer_details': offer_details,
            'loan_type_list': loan_type_list,
            'country_list': country_list,
            'country_id': country if country is not None else 1
        }
        theme_path = 'home/{}/listings/best-offer.html'.format(templete_dir)
        return render(request, theme_path, context)

    except Exception as exp:
        print(exp)
        return HttpResponse("Issue in views")

@csrf_exempt
def best_offers(request):
    try:
        try:
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']
            templete_dir = site_detail['site_detail']['theme_directory']

        except Exception as exp:
            print(exp)
            site_id = ""
            templete_dir = 'theme-1'

        static_dir = templete_dir

        token = request.session['token']['access_token']
        user_id = request.session['user_id']
        page_size = 10
        if request.is_ajax() and request.method == 'POST':
            search = ''
            if 'search' in request.POST and request.POST['search']:
                search = request.POST['search']
            page = 1
            if 'page' in request.POST and request.POST['page'] != "":
                page = request.POST['page']

            if 'page_size' in request.POST and request.POST['page_size'] != "":
                page_size = request.POST['page_size']
            params = {
                'domain_id': site_id,
                'user_id': user_id,
                'page_size': page_size,
                'search': search,
                'page': page
            }
            sno = (int(page) - 1) * int(page_size) + 1
            api_url = settings.API_URL + '/api-bid/enhanced-best-offer-listing/'
            offer_data = call_api_post_method(params, api_url, token=token)

            if 'error' in offer_data and offer_data['error'] == 0:
                offer_listing = offer_data['data']['data']
                total = offer_data['data']['total']
            else:
                offer_listing = []
                total = 0
            if offer_listing:
                for offer in offer_listing:
                    start_price = 0
                    try:
                        if 'your_offer_price' in offer and offer['your_offer_price']:
                            start_price = offer['your_offer_price']
                        if 'user_data' in offer and 'earnest_money_deposit' in offer['user_data'] and offer['user_data'][
                            'earnest_money_deposit'] and start_price:
                            if 'earnest_deposit_type' in offer and offer[
                                'earnest_deposit_type'] == 1:
                                earnest_money_deposit = offer['user_data']['earnest_money_deposit']
                                earnest_money_deposit = format_currency(earnest_money_deposit)
                                earnest_money_deposit_percent = (float(earnest_money_deposit) * 100) / float(start_price)
                                earnest_money_deposit_percent = format_currency(earnest_money_deposit_percent)
                            elif 'earnest_deposit_type' in offer and offer[
                                'earnest_deposit_type'] == 2:
                                earnest_money_deposit_percent = offer['user_data']['earnest_money_deposit']
                                earnest_money_deposit_percent = format_currency(earnest_money_deposit_percent)
                                earnest_money_deposit = (float(start_price) * float(
                                    offer['user_data']['earnest_money_deposit'])) / float(100)
                                earnest_money_deposit = format_currency(earnest_money_deposit)
                            else:
                                earnest_money_deposit = ''
                                earnest_money_deposit_percent = ''
                        else:
                            earnest_money_deposit = ''
                            earnest_money_deposit_percent = ''
                    except:
                        earnest_money_deposit = ''
                        earnest_money_deposit_percent = ''
                    offer['earnest_money_deposit'] = earnest_money_deposit
                    offer['earnest_money_deposit_percent'] = earnest_money_deposit_percent

            context = {'offer_listing': offer_listing, 'total': total, "active_menu": "best_offer", "aws_url": settings.AWS_URL, 'sno': sno}

            offer_listing_path = 'home/{}/user-dashboard/best_offers/best-offer-listing.html'.format(templete_dir)
            offer_listing_template = get_template(offer_listing_path)
            offer_listing_html = offer_listing_template.render(context)
            # ---------------Pagination--------
            pagination_path = 'home/{}/user-dashboard/best_offers/pagination.html'.format(templete_dir)
            pagination_template = get_template(pagination_path)
            total_page = math.ceil(int(total) / int(page_size))
            pagination_data = {"no_page": int(total_page), "total_page": range(total_page), "current_page": int(page)}
            pagination_html = pagination_template.render(pagination_data)
            data = {'offer_listing_html': offer_listing_html, 'status': 200, 'msg': '', 'error': 0, 'total': total, "pagination_html": pagination_html}
            return JsonResponse(data)
        else:
            page = 1
            params = {
                'domain_id': site_id,
                'user_id': user_id,
                'page_size': page_size,
                'search': '',
                'page': page
            }
            api_url = settings.API_URL + '/api-bid/enhanced-best-offer-listing/'
            offer_data = call_api_post_method(params, api_url, token=token)

            if 'error' in offer_data and offer_data['error'] == 0:
                offer_listing = offer_data['data']['data']
                total = offer_data['data']['total']
            else:
                offer_listing = []
                total = 0
            if offer_listing:
                for offer in offer_listing:
                    start_price = 0
                    try:
                        if 'your_offer_price' in offer and offer['your_offer_price']:
                            start_price = offer['your_offer_price']
                        if 'user_data' in offer and 'earnest_money_deposit' in offer['user_data'] and offer['user_data'][
                            'earnest_money_deposit'] and start_price:
                            if 'earnest_deposit_type' in offer and offer[
                                'earnest_deposit_type'] == 1:
                                earnest_money_deposit = offer['user_data']['earnest_money_deposit']
                                earnest_money_deposit = format_currency(earnest_money_deposit)
                                earnest_money_deposit_percent = (float(earnest_money_deposit) * 100) / float(start_price)
                                earnest_money_deposit_percent = format_currency(earnest_money_deposit_percent)
                            elif 'earnest_deposit_type' in offer and offer[
                                'earnest_deposit_type'] == 2:
                                earnest_money_deposit_percent = offer['user_data']['earnest_money_deposit']
                                earnest_money_deposit_percent = format_currency(earnest_money_deposit_percent)
                                earnest_money_deposit = (float(start_price) * float(
                                    offer['user_data']['earnest_money_deposit'])) / float(100)
                                earnest_money_deposit = format_currency(earnest_money_deposit)
                            else:
                                earnest_money_deposit = ''
                                earnest_money_deposit_percent = ''
                        else:
                            earnest_money_deposit = ''
                            earnest_money_deposit_percent = ''
                    except:
                        earnest_money_deposit = ''
                        earnest_money_deposit_percent = ''
                    offer['earnest_money_deposit'] = earnest_money_deposit
                    offer['earnest_money_deposit_percent'] = earnest_money_deposit_percent
            sno = (int(page) - 1) * int(page_size) + 1
            # ---------------Pagination--------
            pagination_path = 'home/{}/user-dashboard/best_offers/pagination.html'.format(templete_dir)
            pagination_template = get_template(pagination_path)
            total_page = math.ceil(total/page_size)
            pagination_data = {"no_page": total_page, "total_page": range(total_page), "current_page": 1}
            pagination_html = pagination_template.render(pagination_data)

            x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
            if x_forwarded_for:
                ip_address = x_forwarded_for.split(',')[0]
            else:
                ip_address = request.META.get('REMOTE_ADDR')

            #context = {'offer_listing': offer_listing, 'total': total, "active_menu": "my_offer", "pagination_html": pagination_html, "sno":sno, "node_url": settings.NODE_URL, "user_id": user_id, "domain_id": site_id, "ip_address": ip_address}
            context = {
                "active_menu": "best_offer",
                "node_url": settings.NODE_URL,
                "user_id": user_id,
                "domain_id": site_id,
                "ip_address": ip_address,
                "offer_listing": offer_listing,
                "pagination_html": pagination_html,
                "total": total,
                "sno": sno,
                "node_url": settings.NODE_URL,
                "aws_url": settings.AWS_URL
            }
            theme_path = 'home/{}/user-dashboard/best_offers/offers.html'.format(templete_dir)
            return render(request, theme_path, context)
    except Exception as exp:
        print(exp)
        return HttpResponse("Issue in views")

@csrf_exempt
def save_best_offer_document(request):
    try:
        user_id = request.session['user_id']
        site_detail = subdomain_site_details(request)
        site_id = site_detail['site_detail']['site_id']
        token = request.session['token']['access_token']
        file_urls = ""
        upload_to = ""
        uploaded_file_list = []
        file_size = request.POST['file_size']

        try:
            for key, value in request.FILES.items():
                params = {}
                if 'best_offer_document' in key.lower():
                    upload_to = 'best_and_final_offer'
                    doc_type = 19
                file_urls = request.FILES[key]
        except:
            pass

        if int(request.POST['file_length']) > 1:
            for key, value in request.FILES.items():
                params = {}
                upload_to = 'best_and_final_offer'
                doc_type = 19

                file_res = request.FILES[key]
                response = save_to_s3(file_res, upload_to)
                if 'error' in response and response['error'] == 0:
                    try:
                        upload_param = {
                            "site_id": site_id,
                            "user_id": user_id,
                            "doc_file_name": response['file_name'],
                            "document_type": doc_type,
                            "bucket_name": upload_to,
                            "added_by": user_id,
                            "file_size": str(file_size) + 'MB'
                        }
                        url = settings.API_URL + '/api-users/file-upload/'
                        upload_data = call_api_post_method(upload_param, url, token)
                        upload_id = upload_data['data']['upload_id']
                        upload_size = upload_data['data']['file_size']
                        upload_date = upload_data['data']['added_date']
                    except:
                        upload_id = 0
                        upload_size = '0MB'
                        upload_date = ''

                    params['file_name'] = response['file_name']
                    params['error'] = 0
                    params['msg'] = response['msg']
                    params['upload_id'] = upload_id
                    params['file_size'] = upload_size
                    params['upload_date'] = upload_date
                    params['upload_to'] = upload_to
                else:
                    params['file_name'] = response['file_name']
                    params['error'] = 1
                    params['msg'] = response['msg']
                    params['upload_id'] = 0
                    params['file_size'] = '0MB'
                    params['upload_date'] = ''
                    params['upload_to'] = upload_to

                uploaded_file_list.append(params)
        else:
            params = {}
            response = save_to_s3(file_urls, upload_to)
            if 'error' in response and response['error'] == 0:
                try:
                    upload_param = {
                        "site_id": site_id,
                        "user_id": user_id,
                        "doc_file_name": response['file_name'],
                        "document_type": doc_type,
                        "bucket_name": upload_to,
                        "added_by": user_id,
                        "file_size": str(file_size) + 'MB'
                    }
                    url = settings.API_URL + '/api-users/file-upload/'
                    upload_data = call_api_post_method(upload_param, url, token)
                    upload_id = upload_data['data']['upload_id']
                    upload_size = upload_data['data']['file_size']
                    upload_date = upload_data['data']['added_date']
                except Exception as exp:
                    print(exp)
                    upload_id = 0
                    upload_size = '0MB'
                    upload_date = ''

                params['file_name'] = response['file_name']
                params['error'] = 0
                params['msg'] = response['msg']
                params['upload_id'] = upload_id
                params['file_size'] = upload_size
                params['upload_date'] = upload_date
                params['upload_to'] = upload_to
            else:
                params['file_name'] = response['file_name']
                params['error'] = 1
                params['msg'] = response['msg']
                params['upload_id'] = 0
                params['file_size'] = '0MB'
                params['upload_date'] = ''
                params['upload_to'] = upload_to
            uploaded_file_list.append(params)

        return JsonResponse({'status': 200, 'uploaded_file_list': uploaded_file_list})
    except Exception as exp:
        print(exp)
        data = {'status': 403, 'msg': 'invalid request.'}
        return JsonResponse(data)

@csrf_exempt
def delete_best_offer_document(request):
    try:
        return_data = {
            'section': request.POST['section'],
            'image_id': request.POST['upload_id'],
            'image_name': request.POST['doc_name'],
        }
        if request.is_ajax() and request.method == 'POST':
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']
            token = request.session['token']['access_token']
            doc_name = request.POST['doc_name']
            bucket = request.POST['section']

            params = {
                "domain": site_id,
                "user_id": request.session['user_id'],
                "upload_type": "offer_upload",
                "upload_id": request.POST['upload_id']
            }

            url = settings.API_URL + '/api-users/delete-file/'

            delete_data = call_api_post_method(params, url, token)

            user_list = []

            if 'error' in delete_data and delete_data['error'] == 0:
                try:
                    delete = delete_s3_file(bucket + '/' + doc_name, settings.AWS_BUCKET_NAME)
                except Exception as exp:
                    print(exp)
                return_data['msg'] = 'Document deleted successfully'
                return_data['error'] = 0
                return_data['status'] = 200
                data = return_data
            else:
                return_data['msg'] = 'Some error occurs, please try again'
                return_data['error'] = 1
                return_data['status'] = 403
                data = return_data
        else:
            return_data['msg'] = 'Invalid request'
            return_data['error'] = 1
            return_data['status'] = 403
            data = return_data

        return JsonResponse(data)
    except Exception as exp:
        print(exp)
        data = {'status': 403, 'msg': 'invalid request.', 'user_list': []}
        return JsonResponse(data)

@csrf_exempt
def buyer_best_offer_counter(request):
    try:
        site_detail = subdomain_site_details(request)
        site_id = site_detail['site_detail']['site_id']
        token = request.session['token']['access_token']
        user_id = request.session['user_id']

        if request.is_ajax() and request.method == 'POST':
            if 'offer_price' in request.POST and request.POST['offer_price']:
                offer_price = float(request.POST['offer_price'].replace('$', '').replace(',', ''))
            else:
                offer_price = ''

            if 'earnest_deposit' in request.POST and request.POST['earnest_deposit'] and request.POST['earnest_deposit'] != "$":
                earnest_deposit = float(request.POST['earnest_deposit'].replace('$', '').replace(',', ''))
            else:
                earnest_deposit = ''

            if 'down_payment' in request.POST and request.POST['down_payment'].replace(',', '').replace('$', '') != "" and request.POST['down_payment'] != "$":
                down_payment = float(request.POST['down_payment'].replace('$', '').replace(',', ''))
            else:
                down_payment = ''

            params = {
                "domain_id": site_id,
                "property_id": request.POST['property_id'],
                "negotiation_id": request.POST['negotiated_id'],
                "offer_price": offer_price,
                "offer_comment": request.POST['offer_comment'],
                "user_id": user_id,
                "earnest_money_deposit": earnest_deposit,
                "down_payment": down_payment,
                "due_diligence_period": request.POST['due_diligence'],
                "closing_period": request.POST['closing_period'],
                "financing": request.POST['financing'],
                "offer_contingent": request.POST['offer_contingent'] if "offer_contingent" in request.POST else "",
                "sale_contingency": request.POST['sale_contingency'] if "sale_contingency" in request.POST else "",
                "appraisal_contingent": request.POST['appraisal_contingent'] if "appraisal_contingent" in request.POST else "",
                "closing_cost": request.POST['closing_cost'] if "closing_cost" in request.POST else "",
            }
            url = settings.API_URL + '/api-bid/enhanced-buyer-make-loi/'

            response = call_api_post_method(params, url, token)
            if 'error' in response and response['error'] == 0:
                data = {
                    'error': 0,
                    'msg': "Counter Offer successfully."
                }
            else:
                data = {
                    'error': 1,
                    'msg': response['msg']
                }
            return JsonResponse(data)
    except Exception as exp:
        print(exp)
        return HttpResponse("Issue in views")

@csrf_exempt
def buyer_reject_best_offer(request):
    try:
        site_detail = subdomain_site_details(request)
        site_id = site_detail['site_detail']['site_id']
        token = request.session['token']['access_token']
        user_id = request.session['user_id']

        if request.is_ajax() and request.method == 'POST':
            params = {
                'domain_id': site_id,
                'user_id': user_id,
                'property_id': request.POST['property_id'],
                'negotiation_id': request.POST['negotiated_id'],
                'declined_reason': request.POST['reason'] if 'reason' in request.POST and request.POST[
                    'reason'] != "" else "",
            }
            url = settings.API_URL + '/api-bid/enhanced-buyer-reject-offer/'

            response = call_api_post_method(params, url, token)

            if 'error' in response and response['error'] == 0:
                data = {
                    'error': 0,
                    'msg': "Offer Cancelled successfully."
                }
            else:
                data = {
                    'error': 1,
                    'msg': response['msg']
                }
            return JsonResponse(data)
    except Exception as exp:
        print(exp)
        return HttpResponse("Issue in views")

@csrf_exempt
def property_best_offer_details(request):
    try:
        if request.is_ajax() and request.method == 'POST':
            try:
                site_detail = subdomain_site_details(request)
                site_id = site_detail['site_detail']['site_id']
                templete_dir = site_detail['site_detail']['theme_directory']

            except Exception as exp:
                print(exp)
                site_id = ""
                templete_dir = 'theme-1'

            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']
            token = request.session['token']['access_token']
            user_id = request.session['user_id']

            params = {
                "domain_id": site_id,
                "property_id": request.POST['property_id'],
                "user_id": user_id
            }
            from_page = request.POST['from_page']
            api_url = settings.API_URL + '/api-bid/enhanced-best-buyer-offer-detail/'

            offer_data = call_api_post_method(params, api_url, token)

            if 'error' in offer_data and offer_data['error'] == 0:
                offer_details = offer_data['data']['data']
                start_price = 0
                if 'offer_price' in offer_details and offer_details['offer_price']:
                    start_price = offer_details['offer_price']
                if 'offer_detail' in offer_details and 'earnest_money_deposit' in offer_details['offer_detail'] and offer_details['offer_detail']['earnest_money_deposit'] and start_price:
                    if 'earnest_deposit_type' in offer_details['offer_detail'] and offer_details['offer_detail']['earnest_deposit_type'] == 1:
                        earnest_money_deposit = offer_details['offer_detail']['earnest_money_deposit']
                        earnest_money_deposit = format_currency(earnest_money_deposit)
                        earnest_money_deposit_percent = (float(earnest_money_deposit) * 100) / float(start_price)
                        earnest_money_deposit_percent = format_currency(earnest_money_deposit_percent)
                    elif 'earnest_deposit_type' in offer_details['offer_detail'] and offer_details['offer_detail']['earnest_deposit_type'] == 2:
                        earnest_money_deposit_percent = offer_details['offer_detail']['earnest_money_deposit']
                        earnest_money_deposit_percent = format_currency(earnest_money_deposit_percent)
                        earnest_money_deposit = (float(start_price) * float(
                            offer_details['offer_detail']['earnest_money_deposit'])) / float(100)
                        earnest_money_deposit = format_currency(earnest_money_deposit)
                    else:
                        earnest_money_deposit = ''
                        earnest_money_deposit_percent = ''
                else:
                    earnest_money_deposit = ''
                    earnest_money_deposit_percent = ''
                try:
                    offer_details['earnest_money_deposit'] = earnest_money_deposit
                    offer_details['earnest_money_deposit_percent'] = earnest_money_deposit_percent
                except:
                    offer_details['earnest_money_deposit'] = ''
                    offer_details['earnest_money_deposit_percent'] = ''
                context = {'offer': offer_details, "aws_url": settings.AWS_URL, "user_id": user_id}
                if from_page == 'asset_details':
                    offer_details_path = 'home/{}/listings/best_offer_details_content.html'.format(templete_dir)
                else:
                    offer_details_path = 'home/{}/user-dashboard/offers/offer_details_content.html'.format(templete_dir)
                offer_details_template = get_template(offer_details_path)
                offer_details_html = offer_details_template.render(context)
            else:
                offer_details_html = ''

            data = {'status': 200, 'offer_details_html': offer_details_html, 'error': 0}
        else:
            data = {'status': 403, 'offer_details_html': '', 'error': 1}

        return JsonResponse(data)
    except Exception as exp:
        print(exp)
        data = {'status': 403, 'offer_details_html': '', 'error': 1}
        return JsonResponse(data)

@csrf_exempt
def download_loi(request):
    try:
        """
            Downloads LOI
        """
        try:
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']
            templete_dir = site_detail['site_detail']['theme_directory']

        except Exception as exp:
            print(exp)
            site_id = ""
            templete_dir = 'theme-1'

        user_id = None
        token = None
        if 'user_id' in request.session and request.session['user_id']:
            token = request.session['token']['access_token']
            user_id = request.session['user_id']

        property = request.GET.get('property', '')
        loi_param = {
            "domain": site_id,
            "user": user_id,
            "property": property
        }
        loi_url = settings.API_URL + '/api-bid/get-loi/'
        loi_data = call_api_post_method(loi_param, loi_url, token)
        offer_details = loi_data['data']
        try:
            if 'offer_detail' in offer_details and offer_details['offer_detail']:
                offer_start_price = 0

                if 'last_offer_price' in offer_details['offer_detail'] and \
                        offer_details['offer_detail']['last_offer_price']:
                    offer_start_price = offer_details['offer_detail']['last_offer_price']

                if 'earnest_deposit_type' in offer_details['property_detail'] and offer_details['property_detail'][
                    'earnest_deposit_type'] == 1:
                    offer_earnest_deposit = offer_details['offer_detail']['earnest_money_deposit']
                    current_earnest_money_deposit = format_currency(offer_earnest_deposit)
                    offer_money_deposit_percent = (float(offer_earnest_deposit) * 100) / float(offer_start_price)
                    current_earnest_money_deposit_percent = format_currency(offer_money_deposit_percent)
                elif 'earnest_deposit_type' in offer_details['property_detail'] and offer_details['property_detail'][
                    'earnest_deposit_type'] == 2:
                    offer_money_deposit_percent = offer_details['offer_detail']['earnest_money_deposit']
                    current_earnest_money_deposit_percent = format_currency(offer_money_deposit_percent)
                    offer_earnest_deposit = (float(offer_start_price) * float(offer_money_deposit_percent)) / float(100)
                    current_earnest_money_deposit = format_currency(offer_earnest_deposit)
                else:
                    current_earnest_money_deposit = ''
                    current_earnest_money_deposit_percent = ''

                offer_details['offer_detail']['current_earnest_money_deposit'] = current_earnest_money_deposit
                offer_details['offer_detail'][
                    'current_earnest_money_deposit_percent'] = current_earnest_money_deposit_percent

        except:
            pass
        context = {"offer_details": offer_details}
        offer_listing_path = 'home/download/loi_pdf.html'
        offer_listing_template = get_template(offer_listing_path)
        offer_listing_html = offer_listing_template.render(context)

        result = BytesIO()

        # This part will create the pdf.
        pdf = pisa.pisaDocument(BytesIO(offer_listing_html.encode("UTF-8")), result)
        if not pdf.err:
            #return HttpResponse(result.getvalue(), content_type='application/pdf')
            response = HttpResponse(result.getvalue(), content_type='application/pdf')
            response['Content-Disposition'] = 'attachment; filename={date}-offer.pdf'.format(
                date=datetime.datetime.now().strftime('%Y-%m-%d'),
            )
            return response
        return None

        # context = {"offer_details": offer_details}
        # theme_path = 'home/{}/user-dashboard/best_offers/loi_pdf.html'.format(templete_dir)
        # return render(request, theme_path, context)

    except Exception as exp:
        print(exp)
        return HttpResponse("Issue in views")

@csrf_exempt
def export_bid_history(request):
    try:
        """
            Downloads all listings as Excel file with a single worksheet
        """
        try:
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']

        except Exception as exp:
            print(exp)
            site_id = ""

        user_id = None
        token = None
        is_broker = 0
        if 'user_id' in request.session and request.session['user_id']:
            token = request.session['token']['access_token']
            user_id = request.session['user_id']
            is_broker = 1 if request.session['is_broker'] == True else 0

        search = request.GET.get('search', '')
        page = request.GET.get('page', 1)
        page_size = request.GET.get('page_size', 10)
        property_id = request.GET.get('property', '')
        timezone = request.GET.get('timezone', '')

        list_param = {
            'site_id': site_id,
            'user_id': user_id,
            'page_size': page_size,
            'page': page,
            'property_id': property_id,
            'register_user': user_id
        }

        list_url = settings.API_URL + '/api-bid/subdomain-bid-history/'

        list_data = call_api_post_method(list_param, list_url, token=token)

        if 'error' in list_data and list_data['error'] == 0:
            bid_history = list_data['data']['data']
            total = list_data['data']['total'] if 'total' in list_data['data'] else 0
        else:
            bid_history = []
            total = 0
        sno = (int(page) - 1) * int(page_size) + 1


        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename={date}-bid-history.xlsx'.format(
            date=datetime.datetime.now().strftime('%Y-%m-%d'),
        )
        workbook = Workbook()

        # Get active worksheet/tab
        worksheet = workbook.active
        worksheet.title = 'Bid History'

        # Define the titles for columns
        columns = [
            '#',
            'Bidder Name',
            'Email',
            'Phone',
            'Bid Amount',
            'IP Address',
            'Bidding Date',
        ]
        row_num = 1
        header_font = Font(name='Calibri', bold=True)
        # Assign the titles for each cell of the header
        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title
            cell.font = header_font

        # Iterate through all history
        count = 0
        for bids in bid_history:
            row_num += 1

            phone_no = bids['bidder_detail']['phone_no']
            bid_amount = bids['bid_amount']
            formatted_phone_no = format_phone_number(phone_no)
            formatted_amount = format_currency(bid_amount)
            bid_date = bids['bid_date']
            bid_date_time = ''
            if timezone:
                try:
                    added_on_time = time.mktime(
                        datetime.datetime.strptime(bid_date, "%Y-%m-%dT%H:%M:%SZ").timetuple())
                except ValueError:
                    added_on_time = time.mktime(
                        datetime.datetime.strptime(bid_date, "%Y-%m-%dT%H:%M:%S.%fZ").timetuple())
                except:
                    added_on_time = 0
            if added_on_time:
                added_on_time = float(added_on_time) - (float(timezone)*60)
                bid_date_time = datetime.datetime.fromtimestamp(added_on_time)
                bid_date_time = datetime.datetime.strftime(bid_date_time, "%m-%d-%Y %I:%M %p")
            # Define the data for each cell in the row
            row = [
                sno,
                bids['bidder_detail']['first_name']+' '+bids['bidder_detail']['last_name'],
                bids['bidder_detail']['email'],
                formatted_phone_no,
                '$'+str(f"{formatted_amount:,}"),
                bids['ip_address'],
                bid_date_time,
            ]
            sno += 1

            # Assign the data for each cell of the row
            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value


        workbook.save(response)

        return response
    except Exception as exp:
        print(exp)
        return HttpResponse("Issue in views")

@csrf_exempt
def submit_best_offer(request):
    try:
        site_detail = subdomain_site_details(request)
        site_id = site_detail['site_detail']['site_id']
        token = request.session['token']['access_token']
        user_id = request.session['user_id']

        if request.is_ajax() and request.method == 'POST':
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']
            user_id = request.session['user_id']
            token = request.session['token']['access_token']
            step = request.POST['current_step']
            print(request.POST)
            offer_param = {
                "domain_id": site_id,
                "user_id": user_id,
                "property_id": request.POST['property_id'],
                "step": step,
            }
            if int(step) == 1:
                user_phone_no = ''
                if 'user_phone_no' in request.POST and request.POST['user_phone_no'] != "":
                    user_phone_no = re.sub('\D', '', request.POST['user_phone_no'])

                offer_step_param = {
                    "first_name": request.POST['first_name'],
                    "last_name": request.POST['last_name'],
                    "email": request.POST['user_email'],
                    "address_first": request.POST['address_1'],
                    "city": request.POST['city'],
                    "state": request.POST['state'],
                    "phone_no": user_phone_no,
                    "postal_code": request.POST['zip_code'],
                    "user_type": request.POST['trad_user_type'],
                    "country": request.POST['country'],

                }
                if 'trad_user_type' in request.POST and int(request.POST['trad_user_type']) == 2:
                    if 'working_with_agent' in request.POST and int(request.POST['working_with_agent']) == 0:
                        offer_step_param['working_with_agent'] = False

                    if 'property_in_person' in request.POST and request.POST['property_in_person'] != "":
                        offer_step_param['property_in_person'] = request.POST['property_in_person']

                    if 'buyer_pre_qualified' in request.POST and request.POST['buyer_pre_qualified'] != "":
                        offer_step_param['pre_qualified_lender'] = request.POST['buyer_pre_qualified']

                else:
                    if 'agent_property_in_person' in request.POST and request.POST['agent_property_in_person'] != "":
                        offer_step_param['property_in_person'] = request.POST['agent_property_in_person']

                    if 'agent_pre_qualified' in request.POST and request.POST['agent_pre_qualified'] != "":
                        offer_step_param['pre_qualified_lender'] = request.POST['agent_pre_qualified']

                    if 'working_with_buyer' in request.POST and int(request.POST['working_with_buyer']) == 1:
                        buyer_first_name = ''
                        if 'agent_buyer_first_name' in request.POST and request.POST['agent_buyer_first_name'] != "":
                            buyer_first_name = request.POST['agent_buyer_first_name']

                        buyer_last_name = ''
                        if 'agent_buyer_last_name' in request.POST and request.POST['agent_buyer_last_name'] != "":
                            buyer_last_name = request.POST['agent_buyer_last_name']

                        buyer_email = ''
                        if 'agent_buyer_user_email' in request.POST and request.POST['agent_buyer_user_email'] != "":
                            buyer_email = request.POST['agent_buyer_user_email']

                        buyer_phone_no = ''
                        if 'agent_buyer_phone_no' in request.POST and request.POST['agent_buyer_phone_no'] != "":
                            buyer_phone_no = re.sub('\D', '', request.POST['agent_buyer_phone_no'])

                        buyer_company = ''
                        if 'agent_buyer_company' in request.POST and request.POST['agent_buyer_company'] != "":
                            buyer_company = request.POST['agent_buyer_company']

                        offer_step_param['buyer_first_name'] = buyer_first_name
                        offer_step_param['buyer_last_name'] = buyer_last_name
                        offer_step_param['buyer_email'] = buyer_email
                        offer_step_param['buyer_company'] = buyer_company
                        offer_step_param['buyer_phone_no'] = buyer_phone_no
                        offer_step_param['behalf_of_buyer'] = True
                    else:
                        offer_step_param['behalf_of_buyer'] = False


            elif int(step) == 2:
                offer_price = None
                if 'offer_price' in request.POST and request.POST['offer_price'] != "":
                    offer_price = float(request.POST['offer_price'].replace(',', '').replace('$', ''))

                down_payment = 0
                if 'down_payment' in request.POST and request.POST['down_payment'].replace(',', '').replace('$', '') != "":
                    down_payment = float(request.POST['down_payment'].replace(',', '').replace('$', ''))

                earnest_deposit = None
                if 'earnest_deposit' in request.POST and request.POST['earnest_deposit'] != "":
                    earnest_deposit = float(request.POST['earnest_deposit'].replace(',', '').replace('$', ''))

                due_diligence = None
                if 'due_diligence' in request.POST and request.POST['due_diligence'] != "":
                    due_diligence = request.POST['due_diligence'].replace(',', '')

                closing_period = None
                if 'closing_period' in request.POST and request.POST['closing_period'] != "":
                    closing_period = request.POST['closing_period'].replace(',', '')
                offer_step_param = {
                    "offer_price": offer_price,
                    "earnest_money_deposit": earnest_deposit,
                    "due_diligence_period": due_diligence,
                    "closing_period": closing_period,
                    "financing": request.POST['financing'],
                    "sale_contingency": request.POST['sale_contingency'],
                    "offer_contingent": request.POST['offer_contingent'],
                    "appraisal_contingent": request.POST['appraisal_contingent'] if 'appraisal_contingent' in request.POST and  request.POST['appraisal_contingent'] else 0,
                    "closing_cost": request.POST['closing_cost'] if 'closing_cost' in request.POST and  request.POST['closing_cost'] else "",
                    "down_payment": down_payment,
                }
            elif int(step) == 3:
                offer_doc = []
                if 'offer_doc_id' in request.POST and request.POST['offer_doc_id']:
                    offer_doc_id = request.POST['offer_doc_id']
                    try:
                        offer_doc = offer_doc_id.split(',')
                    except:
                        offer_doc = []
                offer_step_param = {
                    "document_id": offer_doc,
                    "offer_comment": request.POST['offer_comment'],
                }
            elif int(step) == 4:
                offer_step_param = {
                    "terms": 1,
                }
            else:
                offer_step_param = {}
            print(offer_step_param)
            offer_param.update(offer_step_param)
            offer_url = settings.API_URL + '/api-bid/enhanced-make-loi/'
            offer_data = call_api_post_method(offer_param, offer_url, token)
            if 'error' in offer_data and offer_data['error'] == 0:
                if int(step) == 4:
                    data = {"status": 200, "msg": "Submitted Successfully.", "error": 0, "data": offer_data}
                else:
                    data = {"status": 200, "msg": "Saved Successfully.", "error": 0, "data": offer_data}
            else:
                data = {"status": 400, "msg": offer_data['msg'], "error": 1, "data": offer_data}

            return JsonResponse(data)
    except Exception as exp:
        print(exp)
        return HttpResponse("Issue in views")

@csrf_exempt
def buyer_accept_best_offer(request):
    try:
        site_detail = subdomain_site_details(request)
        site_id = site_detail['site_detail']['site_id']
        token = request.session['token']['access_token']
        user_id = request.session['user_id']

        if request.is_ajax() and request.method == 'POST':
            params = {
                'domain_id': site_id,
                'user_id': user_id,
                'property_id': request.POST['property_id'],
                'negotiation_id': request.POST['negotiated_id'],
            }
            url = settings.API_URL + '/api-bid/enhanced-buyer-accept-loi/'

            response = call_api_post_method(params, url, token)

            if 'error' in response and response['error'] == 0:
                data = {
                    'error': 0,
                    'msg': "Offer Accepted successfully."
                }
            else:
                data = {
                    'error': 1,
                    'msg': response['msg']
                }
            return JsonResponse(data)
    except Exception as exp:
        print(exp)
        return HttpResponse("Issue in views")

@csrf_exempt
def current_best_offer_details(request):
    '''
    Get Current highest and Best offer details
    '''
    try:
        site_detail = subdomain_site_details(request)
        site_id = site_detail['site_detail']['site_id']
        token = request.session['token']['access_token']
        user_id = request.session['user_id']

        if request.is_ajax() and request.method == 'POST':
            params = {
                'domain_id': site_id,
                'user_id': user_id,
                'property_id': request.POST['property_id'],
            }
            url = settings.API_URL + '/api-bid/enhanced-best-current-offer/'

            response = call_api_post_method(params, url, token)

            if 'error' in response and response['error'] == 0:
                try:
                    response_data = response['data']['data']
                except:
                    response_data = {}
                start_price = 0
                if 'offer_price' in response_data and response_data['offer_price']:
                    start_price = response_data['offer_price']
                if 'earnest_money_deposit' in response_data and response_data['earnest_money_deposit'] and start_price:
                    if 'earnest_deposit_type' in response_data and response_data['earnest_deposit_type'] == 1:
                        earnest_money_deposit = response_data['earnest_money_deposit']
                        earnest_money_deposit = format_currency(earnest_money_deposit)
                        earnest_money_deposit_percent = (float(earnest_money_deposit)*100)/float(start_price)
                        earnest_money_deposit_percent = format_currency(earnest_money_deposit_percent)
                    elif 'earnest_deposit_type' in response_data and response_data['earnest_deposit_type'] == 2:
                        earnest_money_deposit_percent = response_data['earnest_money_deposit']
                        earnest_money_deposit_percent = format_currency(earnest_money_deposit_percent)
                        earnest_money_deposit = (float(start_price)*float(response_data['earnest_money_deposit']))/float(100)
                        earnest_money_deposit = format_currency(earnest_money_deposit)
                    else:
                        earnest_money_deposit = ''
                        earnest_money_deposit_percent = ''
                else:
                    earnest_money_deposit = ''
                    earnest_money_deposit_percent = ''

                if 'sale_contingency' in response_data and response_data['sale_contingency'] == True:
                    sale_contingency = 'Yes'
                elif 'sale_contingency' in response_data and response_data['sale_contingency'] == False:
                    sale_contingency = 'No'
                else:
                    sale_contingency = ''

                if 'appraisal_contingent' in response_data and response_data['appraisal_contingent'] == True:
                    appraisal_contingent = 'Yes'
                elif 'appraisal_contingent' in response_data and response_data['appraisal_contingent'] == False:
                    appraisal_contingent = 'No'
                else:
                    appraisal_contingent = ''

                if 'closing_cost' in response_data and response_data['closing_cost'] == 1:
                    closing_cost = 'Buyer agrees to pay for all loan-related closing costs and half of the transaction closing costs.'
                elif 'closing_cost' in response_data and response_data['closing_cost'] == 2:
                    closing_cost = 'Buyer agrees to pay for all loan-related closing costs and all of the transaction closing costs.'
                elif 'closing_cost' in response_data and response_data['closing_cost'] == 3:
                    closing_cost = 'Seller to pay for all loan-related closing costs and all of the transaction closing costs.'
                else:
                    closing_cost = ''

                result = {
                    'offer_price': start_price,
                    'earnest_deposit_type': response_data['earnest_deposit_type'] if 'earnest_deposit_type' in response_data and response_data['earnest_deposit_type'] else "",
                    'earnest_money_deposit': earnest_money_deposit,
                    'earnest_money_deposit_percent': earnest_money_deposit_percent,
                    'due_diligence': response_data['due_diligence_period']  if 'due_diligence_period' in response_data and response_data['due_diligence_period'] else "",
                    'closing_date': response_data['closing_date'] if 'closing_date' in response_data and response_data['closing_date'] else "",
                    'loan_type': response_data['financing'] if 'financing' in response_data and response_data['financing'] else "",
                    'down_payment': response_data['down_payment'] if 'down_payment' in response_data and response_data['down_payment'] else "",
                    'offer_contingent': response_data['offer_contingent'] if 'offer_contingent' in response_data and response_data['offer_contingent'] else "",
                    'sale_contingency': sale_contingency,
                    'appraisal_contingent': appraisal_contingent,
                    'closing_cost': closing_cost,
                }
                data = {
                    'error': 0,
                    'msg': 'success',
                    'data': result
                }
            else:
                data = {
                    'error': 1,
                    'msg': response['msg'],
                    'data': {}
                }
            return JsonResponse(data)
    except Exception as exp:
        print(exp)
        return HttpResponse("Issue in views")

@csrf_exempt
def counter_best_offer_details(request):
    try:
        if request.is_ajax() and request.method == 'POST':
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']
            token = request.session['token']['access_token']
            user_id = request.session['user_id']

            params = {
                "domain_id": site_id,
                "property_id": request.POST['property_id'],
                "user_id": user_id
            }

            api_url = settings.API_URL + '/api-bid/best-counter-buyer-offer-detail/'

            offer_data = call_api_post_method(params, api_url, token)
            if 'error' in offer_data and offer_data['error'] == 0:
                response_data = offer_data['data']['data']
                start_price = 0
                if 'offer_price' in response_data and response_data['offer_price']:
                    start_price = response_data['offer_price']
                # if 'earnest_deposit_type' in response_data and response_data['earnest_deposit_type'] == 1:
                #     earnest_money_deposit = response_data['earnest_money_deposit']
                #     earnest_money_deposit = format_currency(earnest_money_deposit)
                #     earnest_money_deposit_percent = (float(earnest_money_deposit) * 100) / float(start_price)
                #     earnest_money_deposit_percent = format_currency(earnest_money_deposit_percent)
                # elif 'earnest_deposit_type' in response_data and response_data['earnest_deposit_type'] == 2:
                #     earnest_money_deposit_percent = response_data['earnest_money_deposit']
                #     earnest_money_deposit_percent = format_currency(earnest_money_deposit_percent)
                #     earnest_money_deposit = (float(start_price) * float(
                #         response_data['earnest_money_deposit'])) / float(100)
                #     earnest_money_deposit = format_currency(earnest_money_deposit)
                # else:
                #     earnest_money_deposit = ''
                #     earnest_money_deposit_percent = ''

                if 'earnest_money_deposit' in response_data and response_data['earnest_money_deposit'] and start_price:
                    if 'earnest_deposit_type' in response_data and response_data['earnest_deposit_type'] == 1:
                        earnest_money_deposit = response_data['earnest_money_deposit']
                        earnest_money_deposit = format_currency(earnest_money_deposit)
                        earnest_money_deposit_percent = (float(earnest_money_deposit) * 100) / float(start_price)
                        earnest_money_deposit_percent = format_currency(earnest_money_deposit_percent)
                    elif 'earnest_deposit_type' in response_data and response_data['earnest_deposit_type'] == 2:
                        earnest_money_deposit_percent = response_data['earnest_money_deposit']
                        earnest_money_deposit_percent = format_currency(earnest_money_deposit_percent)
                        earnest_money_deposit = (float(start_price) * float(response_data['earnest_money_deposit'])) / float(100)
                        earnest_money_deposit = format_currency(earnest_money_deposit)
                else:
                    earnest_money_deposit = ''
                    earnest_money_deposit_percent = ''

                if 'sale_contingency' in response_data and response_data['sale_contingency'] == True:
                    sale_contingency = 'Yes'
                elif 'sale_contingency' in response_data and response_data['sale_contingency'] == False:
                    sale_contingency = 'No'
                else:
                    sale_contingency = ''

                if 'appraisal_contingent' in response_data and response_data['appraisal_contingent'] == True:
                    appraisal_contingent = 'Yes'
                elif 'appraisal_contingent' in response_data and response_data['appraisal_contingent'] == False:
                    appraisal_contingent = 'No'
                else:
                    appraisal_contingent = ''

                if 'closing_cost' in response_data and response_data['closing_cost'] == 1:
                    closing_cost = 'Buyer agrees to pay for all loan-related closing costs and half of the transaction closing costs.'
                elif 'closing_cost' in response_data and response_data['closing_cost'] == 2:
                    closing_cost = 'Buyer agrees to pay for all loan-related closing costs and all of the transaction closing costs.'
                elif 'closing_cost' in response_data and response_data['closing_cost'] == 3:
                    closing_cost = 'Seller to pay for all loan-related closing costs and all of the transaction closing costs.'
                else:
                    closing_cost = ''

                try:
                    if 'property_asset' in response_data and int(response_data['property_asset']) == 1:
                        loan_type_list = land_loan_types
                    elif 'property_asset' in response_data and int(response_data['property_asset']) == 2:
                        loan_type_list = commerical_loan_types
                    elif 'property_asset' in response_data and int(response_data['property_asset']) == 3:
                        loan_type_list = residential_loan_types
                    else:
                        loan_type_list = []
                except Exception as exp:
                    print(exp)
                    loan_type_list = []

                offer = {
                    'property_id': response_data['property'],
                    'negotiated_id': response_data['negotiation_id'],
                    'offer_price': start_price,
                    'earnest_deposit_type': response_data[
                        'earnest_deposit_type'] if 'earnest_deposit_type' in response_data and response_data[
                        'earnest_deposit_type'] else "",
                    'earnest_money_deposit': earnest_money_deposit,
                    'earnest_money_deposit_percent': earnest_money_deposit_percent,
                    'due_diligence': response_data[
                        'due_diligence_period'] if 'due_diligence_period' in response_data and response_data[
                        'due_diligence_period'] else "",
                    'closing_date': response_data['closing_period'] if 'closing_period' in response_data and response_data[
                        'closing_period'] else "",
                    'loan_type': response_data['financing'] if 'financing' in response_data and response_data[
                        'financing'] else "",
                    'down_payment': response_data['down_payment'] if 'down_payment' in response_data and response_data[
                        'down_payment'] else "",
                    'offer_contingent': response_data['offer_contingent'] if 'offer_contingent' in response_data and
                                                                             response_data['offer_contingent'] else "",
                    'sale_contingency': sale_contingency,
                    'appraisal_contingent': appraisal_contingent,
                    'closing_cost': closing_cost,
                    'buyer_detail_info': response_data['buyer_detail']['buyer_detail'],
                    'agent_detail_info': response_data['buyer_detail']['agent_detail'],
                    'behalf_of_buyer': response_data['buyer_detail']['behalf_of_buyer'],
                    'cancel_reason': response_data['cancel_reason'] if 'cancel_reason' in response_data else None,
                    'comments': response_data['comments'] if 'comments' in response_data else None,
                    'loan_type_list': loan_type_list,
                }
                data = {'status': 200, 'data': offer, 'error': 0}
            else:
                data = {'status': 200, 'data': {}, 'error': 1,'msg': offer_data['msg']}


        else:
            data = {'status': 403, 'data': {}, 'error': 1, 'msg': 'forbidden'}

        return JsonResponse(data)
    except Exception as exp:
        print(exp)
        data = {'status': 403, 'data': {}, 'error': 1, 'msg': 'forbidden'}
        return JsonResponse(data)

@csrf_exempt
def best_offer_history_details(request):
    try:
        if request.is_ajax() and request.method == 'POST':
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']
            token = request.session['token']['access_token']
            user_id = request.session['user_id']

            params = {
                "domain_id": site_id,
                "property_id": request.POST['property_id'],
                "best_offers_id": request.POST['best_offers_id'],
                "user_id": user_id
            }

            api_url = settings.API_URL + '/api-bid/best-offer-history-detail/'

            offer_data = call_api_post_method(params, api_url, token)
            if 'error' in offer_data and offer_data['error'] == 0:
                response_data = offer_data['data']['data']
                start_price = 0
                if 'offer_price' in response_data and response_data['offer_price']:
                    start_price = response_data['offer_price']
                if 'earnest_money_deposit' in response_data and response_data['earnest_money_deposit'] and start_price:
                    if 'earnest_deposit_type' in response_data and response_data['earnest_deposit_type'] == 1:
                        earnest_money_deposit = response_data['earnest_money_deposit']
                        earnest_money_deposit = format_currency(earnest_money_deposit)
                        earnest_money_deposit_percent = (float(earnest_money_deposit) * 100) / float(start_price)
                        earnest_money_deposit_percent = format_currency(earnest_money_deposit_percent)
                    elif 'earnest_deposit_type' in response_data and response_data['earnest_deposit_type'] == 2:
                        earnest_money_deposit_percent = response_data['earnest_money_deposit']
                        earnest_money_deposit_percent = format_currency(earnest_money_deposit_percent)
                        earnest_money_deposit = (float(start_price) * float(response_data['earnest_money_deposit'])) / float(100)
                        earnest_money_deposit = format_currency(earnest_money_deposit)
                else:
                    earnest_money_deposit = ''
                    earnest_money_deposit_percent = ''

                if 'sale_contingency' in response_data and response_data['sale_contingency'] == True:
                    sale_contingency = 'Yes'
                elif 'sale_contingency' in response_data and response_data['sale_contingency'] == False:
                    sale_contingency = 'No'
                else:
                    sale_contingency = ''

                if 'appraisal_contingent' in response_data and response_data['appraisal_contingent'] == True:
                    appraisal_contingent = 'Yes'
                elif 'appraisal_contingent' in response_data and response_data['appraisal_contingent'] == False:
                    appraisal_contingent = 'No'
                else:
                    appraisal_contingent = ''

                if 'closing_cost' in response_data and response_data['closing_cost'] == 1:
                    closing_cost = 'Buyer agrees to pay for all loan-related closing costs and half of the transaction closing costs.'
                elif 'closing_cost' in response_data and response_data['closing_cost'] == 2:
                    closing_cost = 'Buyer agrees to pay for all loan-related closing costs and all of the transaction closing costs.'
                elif 'closing_cost' in response_data and response_data['closing_cost'] == 3:
                    closing_cost = 'Seller to pay for all loan-related closing costs and all of the transaction closing costs.'
                else:
                    closing_cost = ''
                offer = {
                    'property_id': response_data['property'],
                    'negotiated_id': response_data['negotiation_id'],
                    'offer_price': start_price,
                    'earnest_deposit_type': response_data[
                        'earnest_deposit_type'] if 'earnest_deposit_type' in response_data and response_data[
                        'earnest_deposit_type'] else "",
                    'earnest_money_deposit': earnest_money_deposit,
                    'earnest_money_deposit_percent': earnest_money_deposit_percent,
                    'due_diligence': response_data[
                        'due_diligence_period'] if 'due_diligence_period' in response_data and response_data[
                        'due_diligence_period'] else "",
                    'closing_date': response_data['closing_period'] if 'closing_period' in response_data and response_data[
                        'closing_period'] else "",
                    'loan_type': response_data['financing'] if 'financing' in response_data and response_data[
                        'financing'] else "",
                    'down_payment': response_data['down_payment'] if 'down_payment' in response_data and response_data[
                        'down_payment'] else "",
                    'offer_contingent': response_data['offer_contingent'] if 'offer_contingent' in response_data and
                                                                             response_data['offer_contingent'] else "",
                    'sale_contingency': sale_contingency,
                    'appraisal_contingent': appraisal_contingent,
                    'closing_cost': closing_cost,
                    'buyer_detail_info': response_data['buyer_detail']['buyer_detail'],
                    'agent_detail_info': response_data['buyer_detail']['agent_detail'],
                    'behalf_of_buyer': response_data['buyer_detail']['behalf_of_buyer'],
                    'declined_reason': response_data['declined_reason'] if 'declined_reason' in response_data and response_data['declined_reason'] else None,
                    'offer_msg': response_data['comments'] if 'comments' in response_data and response_data[
                        'comments'] else None,
                    'offer_by': response_data['offer_by'] if 'offer_by' in response_data and response_data[
                        'offer_by'] else None,
                }
                data = {'status': 200, 'data': offer, 'error': 0}
            else:
                data = {'status': 200, 'data': {}, 'error': 1,'msg': offer_data['msg']}
        else:
            data = {'status': 403, 'data': {}, 'error': 1, 'msg': 'forbidden'}
        return JsonResponse(data)
    except Exception as exp:
        print(exp)
        data = {'status': 403, 'data': {}, 'error': 1, 'msg': 'forbidden'}
        return JsonResponse(data)


@csrf_exempt
def save_chat_docs(request):
    try:
        user_id = request.session['user_id']
        site_detail = subdomain_site_details(request)
        site_id = site_detail['site_detail']['site_id']
        token = request.session['token']['access_token']
        file_urls = ""
        upload_to = ""
        uploaded_file_list = []
        file_size = request.POST['file_size']

        try:
            for key, value in request.FILES.items():
                params = {}
                doc_type = 20
                upload_to = 'chat_document'    
                file_urls = request.FILES[key]
        except:
            pass

        if int(request.POST['file_length']) > 1:
            for key, value in request.FILES.items():
                params = {}
                doc_type = 20
                upload_to = 'chat_document'

                file_res = request.FILES[key]
                response = save_to_s3(file_res, upload_to)
                if 'error' in response and response['error'] == 0:
                    try:
                        upload_param = {
                            "site_id": site_id,
                            "user_id": user_id,
                            "doc_file_name": response['file_name'],
                            "document_type": doc_type,
                            "bucket_name": upload_to,
                            "added_by": user_id,
                            "file_size": str(file_size)+'MB'
                        }
                        url = settings.API_URL + '/api-users/file-upload/'
                        upload_data = call_api_post_method(upload_param, url, token)
                        upload_id = upload_data['data']['upload_id']
                        upload_size = upload_data['data']['file_size']
                        upload_date = upload_data['data']['added_date']
                    except:
                        upload_id = 0
                        upload_size = '0MB'
                        upload_date = ''

                    params['file_name'] = response['file_name']
                    params['error'] = 0
                    params['msg'] = response['msg']
                    params['upload_id'] = upload_id
                    params['file_size'] = upload_size
                    params['upload_date'] = upload_date
                    params['upload_to'] = upload_to
                else:
                    params['file_name'] = response['file_name']
                    params['error'] = 1
                    params['msg'] = response['msg']
                    params['upload_id'] = 0
                    params['file_size'] = '0MB'
                    params['upload_date'] = ''
                    params['upload_to'] = upload_to

                uploaded_file_list.append(params)
        else:
            params = {}
            response = save_to_s3(file_urls, upload_to)
            if 'error' in response and response['error'] == 0:
                try:
                    upload_param = {
                        "site_id": site_id,
                        "user_id": user_id,
                        "doc_file_name": response['file_name'],
                        "document_type": doc_type,
                        "bucket_name": upload_to,
                        "added_by": user_id,
                        "file_size": str(file_size) + 'MB'
                    }
                    url = settings.API_URL + '/api-users/file-upload/'
                    upload_data = call_api_post_method(upload_param, url, token)
                    upload_id = upload_data['data']['upload_id']
                    upload_size = upload_data['data']['file_size']
                    upload_date = upload_data['data']['added_date']
                except Exception as exp:
                    print(exp)
                    upload_id = 0
                    upload_size = '0MB'
                    upload_date = ''

                params['file_name'] = response['file_name']
                params['error'] = 0
                params['msg'] = response['msg']
                params['upload_id'] = upload_id
                params['file_size'] = upload_size
                params['upload_date'] = upload_date
                params['upload_to'] = upload_to
            else:
                params['file_name'] = response['file_name']
                params['error'] = 1
                params['msg'] = response['msg']
                params['upload_id'] = 0
                params['file_size'] = '0MB'
                params['upload_date'] = ''
                params['upload_to'] = upload_to
            uploaded_file_list.append(params)



        return JsonResponse({'status': 200, 'uploaded_file_list': uploaded_file_list})
    except Exception as exp:
        print(exp)
        data = {'status': 403, 'msg': 'invalid request.'}
        return JsonResponse(data)


@csrf_exempt
def delete_chat_docs(request):
    try:
        return_data = {
            'article_id': request.POST['article_id'],
            'section': request.POST['section'],
            'image_id': request.POST['image_id'],
            'image_name': request.POST['image_name'],
        }
        if request.is_ajax() and request.method == 'POST':
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']
            token = request.session['token']['access_token']
            image_name = request.POST['image_name']
            bucket = request.POST['section']
            params = {}
            if request.POST['section'] == 'chat_document':
                params = {
                    "domain": site_id,
                    "user_id": request.session['user_id'],
                    "upload_type": "chat_document",
                    "upload_id": request.POST['image_id']
                }
                url = settings.API_URL + '/api-users/delete-file/'


            delete_data = call_api_post_method(params, url, token)
            if 'error' in delete_data and delete_data['error'] == 0:
                try:
                    delete = delete_s3_file(bucket+'/'+image_name, settings.AWS_BUCKET_NAME)
                except Exception as exp:
                    print(exp)

                return_data['msg'] = 'Document deleted successfully'
                return_data['error'] = 0
                return_data['status'] = 200
                data = return_data
            else:
                return_data['msg'] = 'Some error occurs, please try again'
                return_data['error'] = 1
                return_data['status'] = 403
                data = return_data
        else:
            return_data['msg'] = 'Invalid request'
            return_data['error'] = 1
            return_data['status'] = 403
            data = return_data

        return JsonResponse(data)
    except Exception as exp:
        print(exp)
        data = {'status': 403, 'msg': 'invalid request.', 'user_list': []}
        return JsonResponse(data)

@csrf_exempt
def insider_bids(request):
    try:
        try:
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']
            templete_dir = site_detail['site_detail']['theme_directory']

        except Exception as exp:
            print(exp)
            site_id = ""
            templete_dir = 'theme-1'

        static_dir = templete_dir

        token = request.session['token']['access_token']
        user_id = request.session['user_id']
        page_size = 10
        if request.is_ajax() and request.method == 'POST':
            search = ''
            if 'search' in request.POST and request.POST['search']:
                search = request.POST['search']
            page = 1
            if 'page' in request.POST and request.POST['page'] != "":
                page = request.POST['page']

            if 'page_size' in request.POST and request.POST['page_size'] != "":
                page_size = request.POST['page_size']
            params = {
                'site_id': site_id,
                'user_id': user_id,
                'page_size': page_size,
                'search': search,
                'page': page
            }
            sno = (int(page) - 1) * int(page_size) + 1
            api_url = settings.API_URL + '/api-bid/inline-bidding-listing/'
            bid_data = call_api_post_method(params, api_url, token=token)
            if 'error' in bid_data and bid_data['error'] == 0:
                bid_listing = bid_data['data']['data']
                total = bid_data['data']['total']
            else:
                bid_listing = []
                total = 0
            context = {'bid_listing': bid_listing, 'total': total, "active_menu": "insider_bid", "aws_url": settings.AWS_URL, 'sno': sno}

            bid_listing_path = 'home/{}/user-dashboard/insider_bids/bid-listing.html'.format(templete_dir)
            bid_listing_template = get_template(bid_listing_path)
            bid_listing_html = bid_listing_template.render(context)
            # ---------------Pagination--------
            pagination_path = 'home/{}/user-dashboard/insider_bids/pagination.html'.format(templete_dir)
            pagination_template = get_template(pagination_path)
            total_page = math.ceil(int(total) / int(page_size))
            pagination_data = {"no_page": int(total_page), "total_page": range(total_page), "current_page": int(page)}
            pagination_html = pagination_template.render(pagination_data)
            data = {'bid_listing_html': bid_listing_html, 'status': 200, 'msg': '', 'error': 0, 'total': total, "pagination_html": pagination_html}
            return JsonResponse(data)
        else:
            page = 1
            params = {
                'site_id': site_id,
                'user_id': user_id,
                'page_size': page_size,
                'search': '',
                'page': page
            }
            api_url = settings.API_URL + '/api-bid/inline-bidding-listing/'
            bid_data = call_api_post_method(params, api_url, token=token)

            if 'error' in bid_data and bid_data['error'] == 0:
                bid_listing = bid_data['data']['data']
                total = bid_data['data']['total']
            else:
                bid_listing = []
                total = 0
            sno = (int(page) - 1) * int(page_size) + 1
            # ---------------Pagination--------
            pagination_path = 'home/{}/user-dashboard/insider_bids/pagination.html'.format(templete_dir)
            pagination_template = get_template(pagination_path)
            total_page = math.ceil(total/page_size)
            pagination_data = {"no_page": total_page, "total_page": range(total_page), "current_page": 1}
            pagination_html = pagination_template.render(pagination_data)

            x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
            if x_forwarded_for:
                ip_address = x_forwarded_for.split(',')[0]
            else:
                ip_address = request.META.get('REMOTE_ADDR')

            context = {'bid_listing': bid_listing, 'total': total, "active_menu": "insider_bid", "pagination_html": pagination_html, "sno":sno, "node_url": settings.NODE_URL, "user_id": user_id, "domain_id": site_id, "ip_address": ip_address}

            theme_path = 'home/{}/user-dashboard/insider_bids/my-bid.html'.format(templete_dir)
            return render(request, theme_path, context)
    except Exception as exp:
        print(exp)
        return HttpResponse("Issue in views")


def property_value_estimator(request):
    """
            Method : GET/POST
            Description : Property Value Estimator page
            Url : /about-us/
        """
    try:
        total_address_ans = 0
        total_address_not_ans = 0
        total_detail_ans = 0
        total_detail_not_ans = 0
        total_doc_ans = 0
        total_doc_not_ans = 0
        total_additional_ans = 0
        total_additional_not_ans = 0

        http_host = request.META['HTTP_HOST']
        site_url = settings.URL_SCHEME + str(http_host)
        try:
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']
            templete_dir = site_detail['site_detail']['theme_directory']
        except Exception as exp:
            site_id = ""
            templete_dir = 'theme-1'

        static_dir = templete_dir

        user_id = None
        token = None
        if 'user_id' in request.session and request.session['user_id']:
            user_id = request.session['user_id']
            token = request.session['token']['access_token']

        try:
            category_params = {}
            category_url = settings.API_URL + '/api-property/property-evaluator-category/'
            category_data = call_api_post_method(category_params, category_url, token)
            category_list = category_data['data']
        except Exception as exp:
            print(exp)
            category_list = []


        try:
            question_params = {
                'domain_id': site_id,
                'user_id': user_id
            }
            question_url = settings.API_URL + '/api-property/property-evaluator-question/'
            response_data = call_api_post_method(question_params, question_url, token)
            if 'error' in response_data and response_data['error'] == 0:
                question_list = response_data['data']
                domain_name = response_data['data']['domain_name']
                address_question_data = response_data['data']['property_address']
                detail_question_data = response_data['data']['property_details']
                document_question_data = response_data['data']['photos_document']
                additional_question_data = response_data['data']['additionals_questions']
                setting_list = []
                setting_data = response_data['data']['setting']
                if setting_data:
                    for obj in setting_data:
                        if int(obj['property_type_id']) == 3:
                            setting_list.append(24)
                        elif int(obj['property_type_id']) == 2:
                            setting_list.append(23)
                        elif int(obj['property_type_id']) == 1:
                            setting_list.append(22)
                else:
                    setting_list = []
                try:
                    asset_type = int(response_data['data']['property_type'])
                    if asset_type == 24:
                        property_asset_type = 3
                    elif asset_type == 23:
                        property_asset_type = 2
                    elif asset_type == 22:
                        property_asset_type = 1
                    else:
                        property_asset_type = ''

                except:
                    property_asset_type = ''
                address_question_list = []
                detail_question_list = []
                document_question_list = []
                additional_question_list = []

                if address_question_data:
                    try:
                        for ques in address_question_data:
                            all_answer = ''
                            doc_cnt = 0
                            if ques['documents']:
                                for doc in ques['documents']:
                                    if doc_cnt == 0:
                                        all_answer = str(doc['id'])
                                    else:
                                        all_answer = all_answer + ',' + str(doc['id'])
                                    doc_cnt = doc_cnt + 1

                            ques['all_ans'] = all_answer
                    except Exception as exp:
                        print(exp)

                if address_question_data:
                    for ques in address_question_data:
                        if ques['option_type'] == 6:
                            if ques['answer'] and ques['answer'][0]['answer']:
                                address = ques['answer'][0]['answer']
                                geolocator = Nominatim(user_agent=settings.GEOLOCATOR_EMAIL, timeout=10)
                                location = geolocator.geocode(address)
                                try:
                                    ques['latitude'] = location.latitude
                                    ques['longitude'] = location.longitude
                                except:
                                    ques['latitude'] = 33.1836304229855
                                    ques['longitude'] = -117.32728034223504
                            else:
                                ques['latitude'] = 33.1836304229855
                                ques['longitude'] = -117.32728034223504
                        if ques['option_type'] == 6:
                            formatted_addr = ''
                            addr_i = 0
                            if ques['answer'] and ques['answer'][0]['answer']:
                                address_arr = ques['answer'][0]['answer'].split(',')
                                if address_arr:
                                    for addr in address_arr:
                                        if addr_i == 0:
                                            formatted_addr = addr
                                        elif addr_i == 1:
                                            if len(address_arr) == 2:
                                                formatted_addr = formatted_addr+', '+addr
                                            else:
                                                formatted_addr = formatted_addr + ', <br>' + addr
                                        elif addr_i == 2:
                                            if len(address_arr) == 3:
                                                formatted_addr = formatted_addr + ', ' + addr
                                            else:
                                                formatted_addr = formatted_addr + ', ' + addr + ', <br>'
                                        else:
                                            formatted_addr = formatted_addr+' '+addr

                                        addr_i = addr_i + 1
                                ques['formatted_addr'] = formatted_addr

                        if property_asset_type and property_asset_type == int(ques['property_type']):
                            if ques['answer'] and (ques['answer'][0]['id'] or ques['answer'][0]['answer']):
                                total_address_ans = total_address_ans + 1
                            else:
                                total_address_not_ans = total_address_not_ans + 1
                            address_question_list.append(ques)

                        if not property_asset_type:
                            if ques['answer'] and (ques['answer'][0]['id'] or ques['answer'][0]['answer']):
                                total_address_ans = total_address_ans + 1
                            else:
                                total_address_not_ans = total_address_not_ans + 1
                            address_question_list.append(ques)

                if detail_question_data:
                    try:
                        for ques in detail_question_data:
                            all_answer = ''
                            doc_cnt = 0
                            if ques['documents']:
                                for doc in ques['documents']:
                                    if doc_cnt == 0:
                                        all_answer = str(doc['id'])
                                    else:
                                        all_answer = all_answer + ',' + str(doc['id'])
                                    doc_cnt = doc_cnt + 1

                            ques['all_ans'] = all_answer
                    except Exception as exp:
                        print(exp)

                if detail_question_data:
                    for ques in detail_question_data:
                        if ques['option_type'] == 6:
                            if ques['answer'] and ques['answer'][0]['answer']:
                                address = ques['answer'][0]['answer']
                                geolocator = Nominatim(user_agent=settings.GEOLOCATOR_EMAIL, timeout=10)
                                location = geolocator.geocode(address)
                                try:
                                    ques['latitude'] = location.latitude
                                    ques['longitude'] = location.longitude
                                except:
                                    ques['latitude'] = 33.1836304229855
                                    ques['longitude'] = -117.32728034223504
                            else:
                                ques['latitude'] = 33.1836304229855
                                ques['longitude'] = -117.32728034223504

                        if ques['option_type'] == 6:
                            formatted_addr = ''
                            addr_i = 0
                            if ques['answer'] and ques['answer'][0]['answer']:
                                address_arr = ques['answer'][0]['answer'].split(',')
                                if address_arr:
                                    for addr in address_arr:
                                        if addr_i == 0:
                                            formatted_addr = addr
                                        elif addr_i == 1:
                                            if len(address_arr) == 2:
                                                formatted_addr = formatted_addr+', '+addr
                                            else:
                                                formatted_addr = formatted_addr + ', <br>' + addr
                                        elif addr_i == 2:
                                            if len(address_arr) == 3:
                                                formatted_addr = formatted_addr + ', ' + addr
                                            else:
                                                formatted_addr = formatted_addr + ', ' + addr + ', <br>'
                                        else:
                                            formatted_addr = formatted_addr+' '+addr

                                        addr_i = addr_i + 1
                                ques['formatted_addr'] = formatted_addr
                        if property_asset_type and property_asset_type == int(ques['property_type']):
                            if ques['answer'] and (ques['answer'][0]['id'] or ques['answer'][0]['answer']):
                                total_detail_ans = total_detail_ans + 1
                            else:
                                total_detail_not_ans = total_detail_not_ans + 1
                            detail_question_list.append(ques)
                        if not property_asset_type:
                            if ques['answer'] and (ques['answer'][0]['id'] or ques['answer'][0]['answer']):
                                total_detail_ans = total_detail_ans + 1
                            else:
                                total_detail_not_ans = total_detail_not_ans + 1
                            detail_question_list.append(ques)




                if document_question_data:
                    try:
                        for ques in document_question_data:
                            all_answer = ''
                            doc_cnt = 0
                            if ques['documents']:
                                for doc in ques['documents']:
                                    if doc_cnt == 0:
                                        all_answer = str(doc['id'])
                                    else:
                                        all_answer = all_answer + ',' + str(doc['id'])
                                    doc_cnt = doc_cnt + 1

                            ques['all_ans'] = all_answer
                    except Exception as exp:
                        print(exp)

                if document_question_data:
                    for ques in document_question_data:
                        if ques['option_type'] == 6:
                            if ques['answer'] and ques['answer'][0]['answer']:
                                address = ques['answer'][0]['answer']
                                geolocator = Nominatim(user_agent=settings.GEOLOCATOR_EMAIL, timeout=10)
                                location = geolocator.geocode(address)
                                try:
                                    ques['latitude'] = location.latitude
                                    ques['longitude'] = location.longitude
                                except:
                                    ques['latitude'] = 33.1836304229855
                                    ques['longitude'] = -117.32728034223504
                            else:
                                ques['latitude'] = 33.1836304229855
                                ques['longitude'] = -117.32728034223504

                        if ques['option_type'] == 6:
                            formatted_addr = ''
                            addr_i = 0
                            if ques['answer'] and ques['answer'][0]['answer']:
                                address_arr = ques['answer'][0]['answer'].split(',')
                                if address_arr:
                                    for addr in address_arr:
                                        if addr_i == 0:
                                            formatted_addr = addr
                                        elif addr_i == 1:
                                            if len(address_arr) == 2:
                                                formatted_addr = formatted_addr+', '+addr
                                            else:
                                                formatted_addr = formatted_addr + ', <br>' + addr
                                        elif addr_i == 2:
                                            if len(address_arr) == 3:
                                                formatted_addr = formatted_addr + ', ' + addr
                                            else:
                                                formatted_addr = formatted_addr + ', ' + addr + ', <br>'
                                        else:
                                            formatted_addr = formatted_addr+' '+addr

                                        addr_i = addr_i + 1
                                ques['formatted_addr'] = formatted_addr

                        if property_asset_type and property_asset_type == int(ques['property_type']):
                            if ques['answer'] and (ques['answer'][0]['id'] or ques['answer'][0]['answer']):
                                total_doc_ans = total_doc_ans + 1
                            else:
                                total_doc_not_ans = total_doc_not_ans + 1

                            document_question_list.append(ques)
                        if not property_asset_type:
                            if ques['answer'] and (ques['answer'][0]['id'] or ques['answer'][0]['answer']):
                                total_doc_ans = total_doc_ans + 1
                            else:
                                total_doc_not_ans = total_doc_not_ans + 1
                            document_question_list.append(ques)


                if additional_question_data:
                    try:
                        for ques in additional_question_data:
                            all_answer = ''
                            doc_cnt = 0
                            if ques['documents']:
                                for doc in ques['documents']:
                                    if doc_cnt == 0:
                                        all_answer = str(doc['id'])
                                    else:
                                        all_answer = all_answer + ',' + str(doc['id'])
                                    doc_cnt = doc_cnt + 1

                            ques['all_ans'] = all_answer
                    except Exception as exp:
                        print(exp)

                if additional_question_data:
                    for ques in additional_question_data:
                        if ques['option_type'] == 6:
                            if ques['answer'] and ques['answer'][0]['answer']:
                                address = ques['answer'][0]['answer']
                                geolocator = Nominatim(user_agent=settings.GEOLOCATOR_EMAIL, timeout=10)
                                location = geolocator.geocode(address)
                                try:
                                    ques['latitude'] = location.latitude
                                    ques['longitude'] = location.longitude
                                except:
                                    ques['latitude'] = 33.1836304229855
                                    ques['longitude'] = -117.32728034223504
                            else:
                                ques['latitude'] = 33.1836304229855
                                ques['longitude'] = -117.32728034223504

                        if ques['option_type'] == 6:
                            formatted_addr = ''
                            addr_i = 0
                            if ques['answer'] and ques['answer'][0]['answer']:
                                address_arr = ques['answer'][0]['answer'].split(',')
                                if address_arr:
                                    for addr in address_arr:
                                        if addr_i == 0:
                                            formatted_addr = addr
                                        elif addr_i == 1:
                                            if len(address_arr) == 2:
                                                formatted_addr = formatted_addr+', '+addr
                                            else:
                                                formatted_addr = formatted_addr + ', <br>' + addr
                                        elif addr_i == 2:
                                            if len(address_arr) == 3:
                                                formatted_addr = formatted_addr + ', ' + addr
                                            else:
                                                formatted_addr = formatted_addr + ', ' + addr + ', <br>'
                                        else:
                                            formatted_addr = formatted_addr+' '+addr

                                        addr_i = addr_i + 1
                                ques['formatted_addr'] = formatted_addr

                        if property_asset_type and property_asset_type == int(ques['property_type']):
                            if ques['answer'] and (ques['answer'][0]['id'] or ques['answer'][0]['answer']):
                                total_additional_ans = total_additional_ans + 1
                            else:
                                total_additional_not_ans = total_additional_not_ans + 1
                            additional_question_list.append(ques)

                        if not property_asset_type:
                            if ques['answer'] and (ques['answer'][0]['id'] or ques['answer'][0]['answer']):
                                total_additional_ans = total_additional_ans + 1
                            else:
                                total_additional_not_ans = total_additional_not_ans + 1
                            additional_question_list.append(ques)


            else:
                question_list = []
                address_question_list = []
                detail_question_list = []
                document_question_list = []
                additional_question_list = []
                setting_list = []
        except Exception as exp:
            print(exp)
            question_list = []
            address_question_list = []
            detail_question_list = []
            document_question_list = []
            additional_question_list = []
            setting_list = []


        context = {
            'is_home_page': True,
            'data': 'page content goes here',
            'aws_url': settings.AWS_URL,
            'sess_user_id': user_id,
            'SITE_URL': site_url,
            'question_list': question_list,
            'address_question_list': address_question_list,
            'detail_question_list': detail_question_list,
            'document_question_list': document_question_list,
            'additional_question_list': additional_question_list,
            'category_list': category_list,
            'total_address_ans': total_address_ans,
            'total_address_not_ans': total_address_not_ans,
            'total_detail_ans': total_detail_ans,
            'total_detail_not_ans': total_detail_not_ans,
            'total_doc_ans': total_doc_ans,
            'total_doc_not_ans': total_doc_not_ans,
            'total_additional_ans': total_additional_ans,
            'total_additional_not_ans': total_additional_not_ans,
            'domain_name': domain_name,
            'setting_list': setting_list,
            'num_doc': range(10),
        }
        theme_path = 'home/{}/estimator/property_value_estimator.html'.format(templete_dir)
        return render(request, theme_path, context)
    except Exception as exp:
        print(exp)
        return HttpResponse("Issue in views")

@csrf_exempt
def save_estimator_answer(request):
    try:
        if request.is_ajax() and request.method == 'POST':
            http_host = request.META['HTTP_HOST']
            site_url = settings.URL_SCHEME + str(http_host)
            total_address_ans = 0
            total_address_not_ans = 0
            total_detail_ans = 0
            total_detail_not_ans = 0
            total_doc_ans = 0
            total_doc_not_ans = 0
            total_additional_ans = 0
            total_additional_not_ans = 0

            try:
                site_detail = subdomain_site_details(request)
                site_id = site_detail['site_detail']['site_id']
                templete_dir = site_detail['site_detail']['theme_directory']
            except Exception as exp:
                site_id = ""
                templete_dir = 'theme-1'


            site_detail = subdomain_site_details(request)
            token = request.session['token']['access_token']
            user_id = request.session['user_id']
            counter_id = request.POST['counter_id']
            counter = request.POST['counter']
            question_id = request.POST['question_id']
            answer = request.POST['answer']
            section_id = request.POST['section_id'] if 'section_id' in request.POST else ""
            is_last_question = False
            if section_id == 'additional_section':
                if 'total_remains' in request.POST and int(request.POST['total_remains']) <= 1:
                    is_last_question = True
            params = {
                "domain_id": site_id,
                "question_id": question_id,
                "user_id": user_id,
                "answer": answer,
                "is_last_question": is_last_question,
            }

            api_url = settings.API_URL + '/api-property/save-property-evaluator-answer/'
            data = call_api_post_method(params, api_url, token)
            if 'error' in data and data['error'] == 0:
                try:
                    try:
                        category_params = {}
                        category_url = settings.API_URL + '/api-property/property-evaluator-category/'
                        category_data = call_api_post_method(category_params, category_url, token)
                        category_list = category_data['data']
                    except Exception as exp:
                        print(exp)
                        category_list = []

                    question_params = {
                        'domain_id': site_id,
                        'user_id': user_id
                    }
                    question_url = settings.API_URL + '/api-property/property-evaluator-question/'
                    response_data = call_api_post_method(question_params, question_url, token)

                    if 'error' in response_data and response_data['error'] == 0:
                        question_list = response_data['data']
                        domain_name = response_data['data']['domain_name']
                        address_question_data = response_data['data']['property_address']
                        detail_question_data = response_data['data']['property_details']
                        document_question_data = response_data['data']['photos_document']
                        additional_question_data = response_data['data']['additionals_questions']
                        setting_list = []
                        setting_data = response_data['data']['setting']
                        if setting_data:
                            for obj in setting_data:
                                if int(obj['property_type_id']) == 3:
                                    setting_list.append(24)
                                elif int(obj['property_type_id']) == 2:
                                    setting_list.append(23)
                                elif int(obj['property_type_id']) == 1:
                                    setting_list.append(22)
                        else:
                            setting_list = []
                        try:
                            asset_type = int(response_data['data']['property_type'])
                            if asset_type == 24:
                                property_asset_type = 3
                            elif asset_type == 23:
                                property_asset_type = 2
                            elif asset_type == 22:
                                property_asset_type = 1
                            else:
                                property_asset_type = ''

                        except:
                            property_asset_type = ''
                        address_question_list = []
                        detail_question_list = []
                        document_question_list = []
                        additional_question_list = []


                        if address_question_data:
                            try:
                                for ques in address_question_data:
                                    all_answer = ''
                                    doc_cnt = 0
                                    if ques['documents']:
                                        for doc in ques['documents']:
                                            if doc_cnt == 0:
                                                all_answer = str(doc['id'])
                                            else:
                                                all_answer = all_answer + ',' + str(doc['id'])
                                            doc_cnt = doc_cnt + 1

                                    ques['all_ans'] = all_answer
                            except Exception as exp:
                                print(exp)

                        if address_question_data:
                            for ques in address_question_data:
                                if ques['option_type'] == 6:
                                    if ques['answer'] and ques['answer'][0]['answer']:
                                        address = ques['answer'][0]['answer']
                                        geolocator = Nominatim(user_agent=settings.GEOLOCATOR_EMAIL, timeout=10)
                                        location = geolocator.geocode(address)
                                        try:
                                            ques['latitude'] = location.latitude
                                            ques['longitude'] = location.longitude
                                        except:
                                            ques['latitude'] = 33.1836304229855
                                            ques['longitude'] = -117.32728034223504
                                    else:
                                        ques['latitude'] = 33.1836304229855
                                        ques['longitude'] = -117.32728034223504
                                if ques['option_type'] == 6:
                                    formatted_addr = ''
                                    addr_i = 0
                                    if ques['answer'] and ques['answer'][0]['answer']:
                                        address_arr = ques['answer'][0]['answer'].split(',')
                                        if address_arr:
                                            for addr in address_arr:
                                                if addr_i == 0:
                                                    formatted_addr = addr
                                                elif addr_i == 1:
                                                    if len(address_arr) == 2:
                                                        formatted_addr = formatted_addr + ', ' + addr
                                                    else:
                                                        formatted_addr = formatted_addr + ', <br>' + addr
                                                elif addr_i == 2:
                                                    if len(address_arr) == 3:
                                                        formatted_addr = formatted_addr + ', ' + addr
                                                    else:
                                                        formatted_addr = formatted_addr + ', ' + addr + ', <br>'
                                                else:
                                                    formatted_addr = formatted_addr + ' ' + addr

                                                addr_i = addr_i + 1
                                        ques['formatted_addr'] = formatted_addr

                                if property_asset_type and property_asset_type == int(ques['property_type']):
                                    if ques['answer'] and (ques['answer'][0]['id'] or ques['answer'][0]['answer']):
                                        total_address_ans = total_address_ans + 1
                                    else:
                                        total_address_not_ans = total_address_not_ans + 1
                                    address_question_list.append(ques)

                                if not property_asset_type:
                                    if ques['answer'] and (ques['answer'][0]['id'] or ques['answer'][0]['answer']):
                                        total_address_ans = total_address_ans + 1
                                    else:
                                        total_address_not_ans = total_address_not_ans + 1
                                    address_question_list.append(ques)

                        if detail_question_data:
                            try:
                                for ques in detail_question_data:
                                    all_answer = ''
                                    doc_cnt = 0
                                    if ques['documents']:
                                        for doc in ques['documents']:
                                            if doc_cnt == 0:
                                                all_answer = str(doc['id'])
                                            else:
                                                all_answer = all_answer + ',' + str(doc['id'])
                                            doc_cnt = doc_cnt + 1

                                    ques['all_ans'] = all_answer
                            except Exception as exp:
                                print(exp)

                        if detail_question_data:
                            for ques in detail_question_data:
                                if ques['option_type'] == 6:
                                    if ques['answer'] and ques['answer'][0]['answer']:
                                        address = ques['answer'][0]['answer']
                                        geolocator = Nominatim(user_agent=settings.GEOLOCATOR_EMAIL, timeout=10)
                                        location = geolocator.geocode(address)
                                        try:
                                            ques['latitude'] = location.latitude
                                            ques['longitude'] = location.longitude
                                        except:
                                            ques['latitude'] = 33.1836304229855
                                            ques['longitude'] = -117.32728034223504
                                    else:
                                        ques['latitude'] = 33.1836304229855
                                        ques['longitude'] = -117.32728034223504

                                if ques['option_type'] == 6:
                                    formatted_addr = ''
                                    addr_i = 0
                                    if ques['answer'] and ques['answer'][0]['answer']:
                                        address_arr = ques['answer'][0]['answer'].split(',')
                                        if address_arr:
                                            for addr in address_arr:
                                                if addr_i == 0:
                                                    formatted_addr = addr
                                                elif addr_i == 1:
                                                    if len(address_arr) == 2:
                                                        formatted_addr = formatted_addr + ', ' + addr
                                                    else:
                                                        formatted_addr = formatted_addr + ', <br>' + addr
                                                elif addr_i == 2:
                                                    if len(address_arr) == 3:
                                                        formatted_addr = formatted_addr + ', ' + addr
                                                    else:
                                                        formatted_addr = formatted_addr + ', ' + addr + ', <br>'
                                                else:
                                                    formatted_addr = formatted_addr + ' ' + addr

                                                addr_i = addr_i + 1
                                        ques['formatted_addr'] = formatted_addr
                                if property_asset_type and property_asset_type == int(ques['property_type']):
                                    if ques['answer'] and (ques['answer'][0]['id'] or ques['answer'][0]['answer']):
                                        total_detail_ans = total_detail_ans + 1
                                    else:
                                        total_detail_not_ans = total_detail_not_ans + 1
                                    detail_question_list.append(ques)
                                if not property_asset_type:
                                    if ques['answer'] and (ques['answer'][0]['id'] or ques['answer'][0]['answer']):
                                        total_detail_ans = total_detail_ans + 1
                                    else:
                                        total_detail_not_ans = total_detail_not_ans + 1
                                    detail_question_list.append(ques)

                        if document_question_data:
                            try:
                                for ques in document_question_data:
                                    all_answer = ''
                                    doc_cnt = 0
                                    if ques['documents']:
                                        for doc in ques['documents']:
                                            if doc_cnt == 0:
                                                all_answer = str(doc['id'])
                                            else:
                                                all_answer = all_answer + ',' + str(doc['id'])
                                            doc_cnt = doc_cnt + 1

                                    ques['all_ans'] = all_answer
                            except Exception as exp:
                                print(exp)

                        if document_question_data:
                            for ques in document_question_data:
                                if ques['option_type'] == 6:
                                    if ques['answer'] and ques['answer'][0]['answer']:
                                        address = ques['answer'][0]['answer']
                                        geolocator = Nominatim(user_agent=settings.GEOLOCATOR_EMAIL, timeout=10)
                                        location = geolocator.geocode(address)
                                        try:
                                            ques['latitude'] = location.latitude
                                            ques['longitude'] = location.longitude
                                        except:
                                            ques['latitude'] = 33.1836304229855
                                            ques['longitude'] = -117.32728034223504
                                    else:
                                        ques['latitude'] = 33.1836304229855
                                        ques['longitude'] = -117.32728034223504

                                if ques['option_type'] == 6:
                                    formatted_addr = ''
                                    addr_i = 0
                                    if ques['answer'] and ques['answer'][0]['answer']:
                                        address_arr = ques['answer'][0]['answer'].split(',')
                                        if address_arr:
                                            for addr in address_arr:
                                                if addr_i == 0:
                                                    formatted_addr = addr
                                                elif addr_i == 1:
                                                    if len(address_arr) == 2:
                                                        formatted_addr = formatted_addr + ', ' + addr
                                                    else:
                                                        formatted_addr = formatted_addr + ', <br>' + addr
                                                elif addr_i == 2:
                                                    if len(address_arr) == 3:
                                                        formatted_addr = formatted_addr + ', ' + addr
                                                    else:
                                                        formatted_addr = formatted_addr + ', ' + addr + ', <br>'
                                                else:
                                                    formatted_addr = formatted_addr + ' ' + addr

                                                addr_i = addr_i + 1
                                        ques['formatted_addr'] = formatted_addr

                                if property_asset_type and property_asset_type == int(ques['property_type']):
                                    if ques['answer'] and (ques['answer'][0]['id'] or ques['answer'][0]['answer']):
                                        total_doc_ans = total_doc_ans + 1
                                    else:
                                        total_doc_not_ans = total_doc_not_ans + 1

                                    document_question_list.append(ques)
                                if not property_asset_type:
                                    if ques['answer'] and (ques['answer'][0]['id'] or ques['answer'][0]['answer']):
                                        total_doc_ans = total_doc_ans + 1
                                    else:
                                        total_doc_not_ans = total_doc_not_ans + 1
                                    document_question_list.append(ques)

                        if additional_question_data:
                            try:
                                for ques in additional_question_data:
                                    all_answer = ''
                                    doc_cnt = 0
                                    if ques['documents']:
                                        for doc in ques['documents']:
                                            if doc_cnt == 0:
                                                all_answer = str(doc['id'])
                                            else:
                                                all_answer = all_answer + ',' + str(doc['id'])
                                            doc_cnt = doc_cnt + 1

                                    ques['all_ans'] = all_answer
                            except Exception as exp:
                                print(exp)

                        if additional_question_data:
                            for ques in additional_question_data:
                                if ques['option_type'] == 6:
                                    if ques['answer'] and ques['answer'][0]['answer']:
                                        address = ques['answer'][0]['answer']
                                        geolocator = Nominatim(user_agent=settings.GEOLOCATOR_EMAIL, timeout=10)
                                        location = geolocator.geocode(address)
                                        try:
                                            ques['latitude'] = location.latitude
                                            ques['longitude'] = location.longitude
                                        except:
                                            ques['latitude'] = 33.1836304229855
                                            ques['longitude'] = -117.32728034223504
                                    else:
                                        ques['latitude'] = 33.1836304229855
                                        ques['longitude'] = -117.32728034223504

                                if ques['option_type'] == 6:
                                    formatted_addr = ''
                                    addr_i = 0
                                    if ques['answer'] and ques['answer'][0]['answer']:
                                        address_arr = ques['answer'][0]['answer'].split(',')
                                        if address_arr:
                                            for addr in address_arr:
                                                if addr_i == 0:
                                                    formatted_addr = addr
                                                elif addr_i == 1:
                                                    if len(address_arr) == 2:
                                                        formatted_addr = formatted_addr + ', ' + addr
                                                    else:
                                                        formatted_addr = formatted_addr + ', <br>' + addr
                                                elif addr_i == 2:
                                                    if len(address_arr) == 3:
                                                        formatted_addr = formatted_addr + ', ' + addr
                                                    else:
                                                        formatted_addr = formatted_addr + ', ' + addr + ', <br>'
                                                else:
                                                    formatted_addr = formatted_addr + ' ' + addr

                                                addr_i = addr_i + 1
                                        ques['formatted_addr'] = formatted_addr

                                if property_asset_type and property_asset_type == int(ques['property_type']):
                                    if ques['answer'] and (ques['answer'][0]['id'] or ques['answer'][0]['answer']):
                                        total_additional_ans = total_additional_ans + 1
                                    else:
                                        total_additional_not_ans = total_additional_not_ans + 1
                                    additional_question_list.append(ques)

                                if not property_asset_type:
                                    if ques['answer'] and (ques['answer'][0]['id'] or ques['answer'][0]['answer']):
                                        total_additional_ans = total_additional_ans + 1
                                    else:
                                        total_additional_not_ans = total_additional_not_ans + 1
                                    additional_question_list.append(ques)





                        context = {
                            'aws_url': settings.AWS_URL,
                            'sess_user_id': user_id,
                            'SITE_URL': site_url,
                            'question_list': question_list,
                            'address_question_list': address_question_list,
                            'detail_question_list': detail_question_list,
                            'document_question_list': document_question_list,
                            'additional_question_list': additional_question_list,
                            'category_list': category_list,
                            'total_address_ans': total_address_ans,
                            'total_address_not_ans': total_address_not_ans,
                            'total_detail_ans': total_detail_ans,
                            'total_detail_not_ans': total_detail_not_ans,
                            'total_doc_ans': total_doc_ans,
                            'total_doc_not_ans': total_doc_not_ans,
                            'total_additional_ans': total_additional_ans,
                            'total_additional_not_ans': total_additional_not_ans,
                            'domain_name': domain_name,
                            'setting_list': setting_list,
                            'num_doc': range(10),
                        }
                        estimator_details_path = 'home/{}/estimator/property_estimator_content.html'.format(
                            templete_dir)
                        estimator_details_template = get_template(estimator_details_path)
                        estimator_details_html = estimator_details_template.render(context)
                    else:
                        question_list = []
                        address_question_list = []
                        detail_question_list = []
                        document_question_list = []
                        additional_question_list = []
                        estimator_details_html = ''
                except:
                    estimator_details_html = ''
                if section_id == 'property_address_section':
                    if total_address_not_ans == 0:
                        scroll_top = True
                    else:
                        scroll_top = False
                elif section_id == 'property_detail_section':
                    if total_detail_not_ans == 0:
                        scroll_top = True
                    else:
                        scroll_top = False
                elif section_id == 'photo_document_section':
                    if total_doc_not_ans == 0:
                        scroll_top = True
                    else:
                        scroll_top = False
                elif section_id == 'additional_section':
                    if total_additional_not_ans == 0:
                        scroll_top = True
                    else:
                        scroll_top = False
                else:
                    scroll_top = False
                custom_data = {'counter_id': counter_id, 'counter': counter, 'question_id': question_id, 'answer': answer, 'estimator_details_html': estimator_details_html, 'scroll_top': scroll_top}
                data = {'status': 200, 'error': 0, 'msg': 'Saved Successfully.', 'data': custom_data}
            else:
                data = {'status': 403, 'error': 1, 'msg': data['msg'], 'data': {}}
        else:
            data = {'status': 403, 'error': 1, 'msg': 'Forbidden.', 'data': {}}

        return JsonResponse(data)
    except Exception as exp:
        print(exp)
        data = {'status': 403, 'error': 1, 'msg': 'Forbidden.'}
        return JsonResponse(data)

@csrf_exempt
def estimator_file_upload(request):
    try:
        if request.is_ajax() and request.method == 'POST':
            http_host = request.META['HTTP_HOST']
            site_url = settings.URL_SCHEME + str(http_host)

            try:
                site_detail = subdomain_site_details(request)
                site_id = site_detail['site_detail']['site_id']
                templete_dir = site_detail['site_detail']['theme_directory']
            except Exception as exp:
                site_id = ""
                templete_dir = 'theme-1'

            token = request.session['token']['access_token']
            user_id = request.session['user_id']
            file_urls = ""
            upload_to = ""
            uploaded_file_list = []
            uploaded_file_ids = []
            file_size = request.POST['file_size']
            question_id = request.POST['question_id']
            counter_id = request.POST['counter_id']
            counter = request.POST['counter']
            img_cnt = request.POST['img_cnt']
            section_id = request.POST['section_id']
            try:
                for key, value in request.FILES.items():
                    params = {}
                    if 'profile_image' in key.lower():
                        upload_to = 'profile_image'
                        doc_type = 22
                        file_urls = request.FILES[key]
            except:
                pass

            if int(request.POST['file_length']):
                cnt = 0
                for key, value in request.FILES.items():
                    params = {}
                    if 'profile_image' in key.lower():
                        upload_to = 'profile_image'
                        doc_type = 22
                        file_res = request.FILES[key]
                        response = save_to_s3(file_res, upload_to)
                        if 'error' in response and response['error'] == 0:
                            try:
                                upload_param = {
                                    "site_id": site_id,
                                    "user_id": user_id,
                                    "doc_file_name": response['file_name'],
                                    "document_type": doc_type,
                                    "bucket_name": upload_to,
                                    "added_by": user_id,
                                    "file_size": str(file_size)+'MB'
                                    }
                                url = settings.API_URL + '/api-users/file-upload/'
                                upload_data = call_api_post_method(upload_param, url, token)
                                upload_id = upload_data['data']['upload_id']
                                upload_size = upload_data['data']['file_size']
                                upload_date = upload_data['data']['added_date']
                            except Exception as exp:
                                print(exp)
                                upload_id = 0
                                upload_size = '0MB'
                                upload_date = ''
                            params['file_name'] = response['file_name']
                            params['error'] = 0
                            params['msg'] = response['msg']
                            params['upload_id'] = upload_id
                            params['file_size'] = upload_size
                            params['upload_date'] = upload_date
                            params['upload_to'] = upload_to
                            uploaded_file_ids.append(upload_id)
                        else:
                            params['file_name'] = response['file_name']
                            params['error'] = 1
                            params['msg'] = response['msg']
                            params['upload_id'] = 0
                            params['file_size'] = '0MB'
                            params['upload_date'] = ''
                            params['upload_to'] = upload_to
                            uploaded_file_ids.append(0)
                        uploaded_file_list.append(params)

                    cnt = cnt + 1

                custom_data = {
                    'counter_id': counter_id,
                    'counter': counter,
                    'question_id': question_id,
                    'uploaded_file_list': uploaded_file_list,
                    'img_cnt': img_cnt,
                    'section_id': section_id,
                    'aws_url': settings.AWS_URL,
                }
                data = {'status': 200, 'error': 0, 'msg': 'Saved Successfully.', 'data': custom_data}
                return JsonResponse(data)
        else:
            data = {'status': 403, 'msg': 'Invalid data', 'error': 1}
            return JsonResponse(data)

    except Exception as exp:
        print(exp)
        data = {'status': 403, 'msg': 'Invalid data', 'error': 1}
        return JsonResponse(data)

@csrf_exempt
def delete_bot_document(request):
    try:
        site_detail = subdomain_site_details(request)
        site_id = site_detail['site_detail']['site_id']
        token = request.session['token']['access_token']
        user_id = request.session['user_id']
        question_id = request.POST['question_id']
        counter_id = request.POST['counter_id']
        counter = request.POST['counter']
        upload_id = request.POST['upload_id']
        bot_doc_id = request.POST['bot_doc_id']
        img_name = request.POST['img_name']
        img_cnt = request.POST['img_cnt']
        section_id = request.POST['section_id']

        if request.is_ajax() and request.method == 'POST':
            params = {
                'domain_id': site_id,
                'user_id': user_id,
                'bot_doc_id': request.POST['bot_doc_id'],
            }
            url = settings.API_URL + '/api-property/delete-bot-doc/'

            response = call_api_post_method(params, url, token)
            if 'error' in response and response['error'] == 0:
                custom_data = {
                    'question_id': question_id,
                    'counter_id': counter_id,
                    'counter': counter,
                    'upload_id': upload_id,
                    'bot_doc_id': bot_doc_id,
                    'img_name': img_name,
                    'img_cnt': img_cnt,
                    'section_id': section_id,
                }
                data = {
                    'error': 0,
                    'msg': "Deleted successfully.",
                    'data': custom_data
                }
            else:
                data = {
                    'error': 1,
                    'msg': response['msg']
                }
            return JsonResponse(data)
    except Exception as exp:
        print(exp)
        return HttpResponse("Issue in views")

@csrf_exempt
def save_upload_answer(request):
    try:
        if request.is_ajax() and request.method == 'POST':
            http_host = request.META['HTTP_HOST']
            site_url = settings.URL_SCHEME + str(http_host)
            total_address_ans = 0
            total_address_not_ans = 0
            total_detail_ans = 0
            total_detail_not_ans = 0
            total_doc_ans = 0
            total_doc_not_ans = 0
            total_additional_ans = 0
            total_additional_not_ans = 0

            try:
                site_detail = subdomain_site_details(request)
                site_id = site_detail['site_detail']['site_id']
                templete_dir = site_detail['site_detail']['theme_directory']
            except Exception as exp:
                site_id = ""
                templete_dir = 'theme-1'


            site_detail = subdomain_site_details(request)
            token = request.session['token']['access_token']
            user_id = request.session['user_id']
            counter_id = request.POST['counter_id']
            counter = request.POST['counter']
            question_id = request.POST['question_id']
            answer = request.POST['answer']
            section_id = request.POST['section_id']
            is_last_question = False
            if section_id == 'additional_section':
                if 'total_remains' in request.POST and int(request.POST['total_remains']) <= 1:
                    is_last_question = True
            try:
                answer_arr = answer.split(",")
            except:
                answer_arr = []

            params = {
                "domain_id": site_id,
                "question_id": question_id,
                "user_id": user_id,
                "answer": answer_arr,
                "is_last_question": is_last_question,
            }
            api_url = settings.API_URL + '/api-property/save-property-evaluator-answer/'
            data = call_api_post_method(params, api_url, token)
            if 'error' in data and data['error'] == 0:
                try:
                    try:
                        category_params = {}
                        category_url = settings.API_URL + '/api-property/property-evaluator-category/'
                        category_data = call_api_post_method(category_params, category_url, token)
                        category_list = category_data['data']
                    except Exception as exp:
                        print(exp)
                        category_list = []

                    question_params = {
                        'domain_id': site_id,
                        'user_id': user_id
                    }
                    question_url = settings.API_URL + '/api-property/property-evaluator-question/'
                    response_data = call_api_post_method(question_params, question_url, token)

                    if 'error' in response_data and response_data['error'] == 0:
                        question_list = response_data['data']
                        domain_name = response_data['data']['domain_name']
                        address_question_data = response_data['data']['property_address']
                        detail_question_data = response_data['data']['property_details']
                        document_question_data = response_data['data']['photos_document']
                        additional_question_data = response_data['data']['additionals_questions']
                        setting_list = []
                        setting_data = response_data['data']['setting']
                        if setting_data:
                            for obj in setting_data:
                                if int(obj['property_type_id']) == 3:
                                    setting_list.append(24)
                                elif int(obj['property_type_id']) == 2:
                                    setting_list.append(23)
                                elif int(obj['property_type_id']) == 1:
                                    setting_list.append(22)
                        else:
                            setting_list = []
                        try:
                            asset_type = int(response_data['data']['property_type'])
                            if asset_type == 24:
                                property_asset_type = 3
                            elif asset_type == 23:
                                property_asset_type = 2
                            elif asset_type == 22:
                                property_asset_type = 1
                            else:
                                property_asset_type = ''

                        except:
                            property_asset_type = ''
                        address_question_list = []
                        detail_question_list = []
                        document_question_list = []
                        additional_question_list = []


                        if address_question_data:
                            try:
                                for ques in address_question_data:
                                    all_answer = ''
                                    doc_cnt = 0
                                    if ques['documents']:
                                        for doc in ques['documents']:
                                            if doc_cnt == 0:
                                                all_answer = str(doc['id'])
                                            else:
                                                all_answer = all_answer + ',' + str(doc['id'])
                                            doc_cnt = doc_cnt + 1

                                    ques['all_ans'] = all_answer
                            except Exception as exp:
                                print(exp)

                        if address_question_data:
                            for ques in address_question_data:
                                if ques['option_type'] == 6:
                                    if ques['answer'] and ques['answer'][0]['answer']:
                                        address = ques['answer'][0]['answer']
                                        geolocator = Nominatim(user_agent=settings.GEOLOCATOR_EMAIL, timeout=10)
                                        location = geolocator.geocode(address)
                                        try:
                                            ques['latitude'] = location.latitude
                                            ques['longitude'] = location.longitude
                                        except:
                                            ques['latitude'] = 33.1836304229855
                                            ques['longitude'] = -117.32728034223504
                                    else:
                                        ques['latitude'] = 33.1836304229855
                                        ques['longitude'] = -117.32728034223504
                                if ques['option_type'] == 6:
                                    formatted_addr = ''
                                    addr_i = 0
                                    if ques['answer'] and ques['answer'][0]['answer']:
                                        address_arr = ques['answer'][0]['answer'].split(',')
                                        if address_arr:
                                            for addr in address_arr:
                                                if addr_i == 0:
                                                    formatted_addr = addr
                                                elif addr_i == 1:
                                                    if len(address_arr) == 2:
                                                        formatted_addr = formatted_addr + ', ' + addr
                                                    else:
                                                        formatted_addr = formatted_addr + ', <br>' + addr
                                                elif addr_i == 2:
                                                    if len(address_arr) == 3:
                                                        formatted_addr = formatted_addr + ', ' + addr
                                                    else:
                                                        formatted_addr = formatted_addr + ', ' + addr + ', <br>'
                                                else:
                                                    formatted_addr = formatted_addr + ' ' + addr

                                                addr_i = addr_i + 1
                                        ques['formatted_addr'] = formatted_addr

                                if property_asset_type and property_asset_type == int(ques['property_type']):
                                    if ques['answer'] and (ques['answer'][0]['id'] or ques['answer'][0]['answer']):
                                        total_address_ans = total_address_ans + 1
                                    else:
                                        total_address_not_ans = total_address_not_ans + 1
                                    address_question_list.append(ques)

                                if not property_asset_type:
                                    if ques['answer'] and (ques['answer'][0]['id'] or ques['answer'][0]['answer']):
                                        total_address_ans = total_address_ans + 1
                                    else:
                                        total_address_not_ans = total_address_not_ans + 1
                                    address_question_list.append(ques)

                        if detail_question_data:
                            try:
                                for ques in detail_question_data:
                                    all_answer = ''
                                    doc_cnt = 0
                                    if ques['documents']:
                                        for doc in ques['documents']:
                                            if doc_cnt == 0:
                                                all_answer = str(doc['id'])
                                            else:
                                                all_answer = all_answer + ',' + str(doc['id'])
                                            doc_cnt = doc_cnt + 1

                                    ques['all_ans'] = all_answer
                            except Exception as exp:
                                print(exp)

                        if detail_question_data:
                            for ques in detail_question_data:
                                if ques['option_type'] == 6:
                                    if ques['answer'] and ques['answer'][0]['answer']:
                                        address = ques['answer'][0]['answer']
                                        geolocator = Nominatim(user_agent=settings.GEOLOCATOR_EMAIL, timeout=10)
                                        location = geolocator.geocode(address)
                                        try:
                                            ques['latitude'] = location.latitude
                                            ques['longitude'] = location.longitude
                                        except:
                                            ques['latitude'] = 33.1836304229855
                                            ques['longitude'] = -117.32728034223504
                                    else:
                                        ques['latitude'] = 33.1836304229855
                                        ques['longitude'] = -117.32728034223504

                                if ques['option_type'] == 6:
                                    formatted_addr = ''
                                    addr_i = 0
                                    if ques['answer'] and ques['answer'][0]['answer']:
                                        address_arr = ques['answer'][0]['answer'].split(',')
                                        if address_arr:
                                            for addr in address_arr:
                                                if addr_i == 0:
                                                    formatted_addr = addr
                                                elif addr_i == 1:
                                                    if len(address_arr) == 2:
                                                        formatted_addr = formatted_addr + ', ' + addr
                                                    else:
                                                        formatted_addr = formatted_addr + ', <br>' + addr
                                                elif addr_i == 2:
                                                    if len(address_arr) == 3:
                                                        formatted_addr = formatted_addr + ', ' + addr
                                                    else:
                                                        formatted_addr = formatted_addr + ', ' + addr + ', <br>'
                                                else:
                                                    formatted_addr = formatted_addr + ' ' + addr

                                                addr_i = addr_i + 1
                                        ques['formatted_addr'] = formatted_addr
                                if property_asset_type and property_asset_type == int(ques['property_type']):
                                    if ques['answer'] and (ques['answer'][0]['id'] or ques['answer'][0]['answer']):
                                        total_detail_ans = total_detail_ans + 1
                                    else:
                                        total_detail_not_ans = total_detail_not_ans + 1
                                    detail_question_list.append(ques)
                                if not property_asset_type:
                                    if ques['answer'] and (ques['answer'][0]['id'] or ques['answer'][0]['answer']):
                                        total_detail_ans = total_detail_ans + 1
                                    else:
                                        total_detail_not_ans = total_detail_not_ans + 1
                                    detail_question_list.append(ques)

                        if document_question_data:
                            try:
                                for ques in document_question_data:
                                    all_answer = ''
                                    doc_cnt = 0
                                    if ques['documents']:
                                        for doc in ques['documents']:
                                            if doc_cnt == 0:
                                                all_answer = str(doc['id'])
                                            else:
                                                all_answer = all_answer + ',' + str(doc['id'])
                                            doc_cnt = doc_cnt + 1

                                    ques['all_ans'] = all_answer
                            except Exception as exp:
                                print(exp)

                        if document_question_data:
                            for ques in document_question_data:
                                if ques['option_type'] == 6:
                                    if ques['answer'] and ques['answer'][0]['answer']:
                                        address = ques['answer'][0]['answer']
                                        geolocator = Nominatim(user_agent=settings.GEOLOCATOR_EMAIL, timeout=10)
                                        location = geolocator.geocode(address)
                                        try:
                                            ques['latitude'] = location.latitude
                                            ques['longitude'] = location.longitude
                                        except:
                                            ques['latitude'] = 33.1836304229855
                                            ques['longitude'] = -117.32728034223504
                                    else:
                                        ques['latitude'] = 33.1836304229855
                                        ques['longitude'] = -117.32728034223504

                                if ques['option_type'] == 6:
                                    formatted_addr = ''
                                    addr_i = 0
                                    if ques['answer'] and ques['answer'][0]['answer']:
                                        address_arr = ques['answer'][0]['answer'].split(',')
                                        if address_arr:
                                            for addr in address_arr:
                                                if addr_i == 0:
                                                    formatted_addr = addr
                                                elif addr_i == 1:
                                                    if len(address_arr) == 2:
                                                        formatted_addr = formatted_addr + ', ' + addr
                                                    else:
                                                        formatted_addr = formatted_addr + ', <br>' + addr
                                                elif addr_i == 2:
                                                    if len(address_arr) == 3:
                                                        formatted_addr = formatted_addr + ', ' + addr
                                                    else:
                                                        formatted_addr = formatted_addr + ', ' + addr + ', <br>'
                                                else:
                                                    formatted_addr = formatted_addr + ' ' + addr

                                                addr_i = addr_i + 1
                                        ques['formatted_addr'] = formatted_addr

                                if property_asset_type and property_asset_type == int(ques['property_type']):
                                    if ques['answer'] and (ques['answer'][0]['id'] or ques['answer'][0]['answer']):
                                        total_doc_ans = total_doc_ans + 1
                                    else:
                                        total_doc_not_ans = total_doc_not_ans + 1

                                    document_question_list.append(ques)
                                if not property_asset_type:
                                    if ques['answer'] and (ques['answer'][0]['id'] or ques['answer'][0]['answer']):
                                        total_doc_ans = total_doc_ans + 1
                                    else:
                                        total_doc_not_ans = total_doc_not_ans + 1
                                    document_question_list.append(ques)

                        if additional_question_data:
                            try:
                                for ques in additional_question_data:
                                    all_answer = ''
                                    doc_cnt = 0
                                    if ques['documents']:
                                        for doc in ques['documents']:
                                            if doc_cnt == 0:
                                                all_answer = str(doc['id'])
                                            else:
                                                all_answer = all_answer + ',' + str(doc['id'])
                                            doc_cnt = doc_cnt + 1

                                    ques['all_ans'] = all_answer
                            except Exception as exp:
                                print(exp)

                        if additional_question_data:
                            for ques in additional_question_data:
                                if ques['option_type'] == 6:
                                    if ques['answer'] and ques['answer'][0]['answer']:
                                        address = ques['answer'][0]['answer']
                                        geolocator = Nominatim(user_agent=settings.GEOLOCATOR_EMAIL, timeout=10)
                                        location = geolocator.geocode(address)
                                        try:
                                            ques['latitude'] = location.latitude
                                            ques['longitude'] = location.longitude
                                        except:
                                            ques['latitude'] = 33.1836304229855
                                            ques['longitude'] = -117.32728034223504
                                    else:
                                        ques['latitude'] = 33.1836304229855
                                        ques['longitude'] = -117.32728034223504

                                if ques['option_type'] == 6:
                                    formatted_addr = ''
                                    addr_i = 0
                                    if ques['answer'] and ques['answer'][0]['answer']:
                                        address_arr = ques['answer'][0]['answer'].split(',')
                                        if address_arr:
                                            for addr in address_arr:
                                                if addr_i == 0:
                                                    formatted_addr = addr
                                                elif addr_i == 1:
                                                    if len(address_arr) == 2:
                                                        formatted_addr = formatted_addr + ', ' + addr
                                                    else:
                                                        formatted_addr = formatted_addr + ', <br>' + addr
                                                elif addr_i == 2:
                                                    if len(address_arr) == 3:
                                                        formatted_addr = formatted_addr + ', ' + addr
                                                    else:
                                                        formatted_addr = formatted_addr + ', ' + addr + ', <br>'
                                                else:
                                                    formatted_addr = formatted_addr + ' ' + addr

                                                addr_i = addr_i + 1
                                        ques['formatted_addr'] = formatted_addr

                                if property_asset_type and property_asset_type == int(ques['property_type']):
                                    if ques['answer'] and (ques['answer'][0]['id'] or ques['answer'][0]['answer']):
                                        total_additional_ans = total_additional_ans + 1
                                    else:
                                        total_additional_not_ans = total_additional_not_ans + 1
                                    additional_question_list.append(ques)

                                if not property_asset_type:
                                    if ques['answer'] and (ques['answer'][0]['id'] or ques['answer'][0]['answer']):
                                        total_additional_ans = total_additional_ans + 1
                                    else:
                                        total_additional_not_ans = total_additional_not_ans + 1
                                    additional_question_list.append(ques)





                        context = {
                            'aws_url': settings.AWS_URL,
                            'sess_user_id': user_id,
                            'SITE_URL': site_url,
                            'question_list': question_list,
                            'address_question_list': address_question_list,
                            'detail_question_list': detail_question_list,
                            'document_question_list': document_question_list,
                            'additional_question_list': additional_question_list,
                            'category_list': category_list,
                            'total_address_ans': total_address_ans,
                            'total_address_not_ans': total_address_not_ans,
                            'total_detail_ans': total_detail_ans,
                            'total_detail_not_ans': total_detail_not_ans,
                            'total_doc_ans': total_doc_ans,
                            'total_doc_not_ans': total_doc_not_ans,
                            'total_additional_ans': total_additional_ans,
                            'total_additional_not_ans': total_additional_not_ans,
                            'domain_name': domain_name,
                            'setting_list': setting_list,
                            'num_doc': range(10),
                        }
                        estimator_details_path = 'home/{}/estimator/property_estimator_content.html'.format(
                            templete_dir)
                        estimator_details_template = get_template(estimator_details_path)
                        estimator_details_html = estimator_details_template.render(context)
                    else:
                        question_list = []
                        address_question_list = []
                        detail_question_list = []
                        document_question_list = []
                        additional_question_list = []
                        estimator_details_html = ''
                except:
                    estimator_details_html = ''

                if section_id == 'property_address_section':
                    if total_address_not_ans == 0:
                        scroll_top = True
                    else:
                        scroll_top = False
                elif section_id == 'property_detail_section':
                    if total_detail_not_ans == 0:
                        scroll_top = True
                    else:
                        scroll_top = False
                elif section_id == 'photo_document_section':
                    if total_doc_not_ans == 0:
                        scroll_top = True
                    else:
                        scroll_top = False
                elif section_id == 'additional_section':
                    if total_additional_not_ans == 0:
                        scroll_top = True
                    else:
                        scroll_top = False
                else:
                    scroll_top = False

                custom_data = {'counter_id': counter_id, 'counter': counter, 'question_id': question_id, 'answer': answer, 'estimator_details_html': estimator_details_html, 'scroll_top': scroll_top}
                data = {'status': 200, 'error': 0, 'msg': 'Saved Successfully.', 'data': custom_data}
            else:
                data = {'status': 403, 'error': 1, 'msg': data['msg'], 'data': {}}
        else:
            data = {'status': 403, 'error': 1, 'msg': 'Forbidden.', 'data': {}}

        return JsonResponse(data)
    except Exception as exp:
        print(exp)
        data = {'status': 403, 'error': 1, 'msg': 'Forbidden.'}
        return JsonResponse(data)

@csrf_exempt
def get_latitude_longitude(request):
    try:
        if request.is_ajax() and request.method == 'POST':
            address = request.POST['address']
            geolocator = Nominatim(user_agent=settings.GEOLOCATOR_EMAIL, timeout=10)
            location = geolocator.geocode(address)
            try:
                latitude = location.latitude
                longitude= location.longitude
            except Exception as exp:
                print(exp)
                latitude = 33.1836304229855
                longitude = -117.32728034223504
            data = {
                'address': address,
                'latitude': latitude,
                'longitude': longitude,
            }
            return JsonResponse(data)
    except Exception as exp:
        print(exp)
        return HttpResponse("Issue in views")


@csrf_exempt
def multiple_parcel(request):
    """
        Method : GET/POST
        Description : Get multiple parcel listing page
        Url : /multiple-parcel/
        Params:
        :param 1: page :: integer
        :param 1: page_size :: integer
        :param 1: property_id :: String
        :param 1: domain_id :: String
        :param 1: search :: String
    """
    try:
        try:
            site_detail = subdomain_site_details(request)
            domain_id = site_detail['site_detail']['site_id']
            templete_dir = site_detail['site_detail']['theme_directory']

        except Exception as exp:
            domain_id = ""
            templete_dir = 'theme-1'

        static_dir = templete_dir

        token = None
        user_id = None
        if 'user_id' in request.session and request.session['user_id']:
            user_id = request.session['user_id']
            token = request.session['token']['access_token']

        portfolio_id = request.GET.get('portfolio_id', None)

        if request.is_ajax() and request.method == 'POST':
            site_detail = subdomain_site_details(request)
            domain_id = site_detail['site_detail']['site_id']
            search = ''
            if 'search' in request.POST and request.POST['search']:
                search = request.POST['search']

            page = 1
            if 'page' in request.POST and request.POST['page']:
                page = int(request.POST['page'])

            page_size = 10
            if 'perpage' in request.POST and request.POST['perpage']:
                page_size = int(request.POST['perpage'])
            parcel_id = int(request.POST['parcel_id'])

            listing_params = {
                "parcel_id": parcel_id,
                "domain_id": domain_id,
                "page_size": page_size,
                "page": page,
                "search": search
            }
            if 'user_id' in request.session and request.session['user_id']:
                listing_params['user_id'] = user_id
            listing_url = settings.API_URL + '/api-property/parcel-detail/'
            listing_data = call_api_post_method(listing_params, listing_url, token)
            http_host = request.META['HTTP_HOST']
            site_url = settings.URL_SCHEME + str(http_host)
            if 'error' in listing_data and listing_data['error'] == 0:
                property_list = listing_data['data']['data']
                property_total = listing_data['data']['total']
                portfolio_detail = listing_data['data']['portfolio_detail']
                total_page = math.ceil(property_total / page_size)
                property_list_data_path = 'home/{}/parcel/our_listing_content.html'.format(templete_dir)

                listing_template = get_template(property_list_data_path)
                listing_html = listing_template.render(
                    {'property_list': property_list, 'total': property_total, 'aws_url': settings.AWS_URL,
                     'sess_user_id': user_id, 'site_url': site_url, 'page': page, 'no_page': total_page, 'portfolio_detail': portfolio_detail})

                pagination_path = 'home/{}/parcel/pagination.html'.format(templete_dir)
                pagination_template = get_template(pagination_path)

                pagination_data = {"no_page": total_page, "total_page": range(total_page), "current_page": page,
                                   "pagination_id": "prop_listing_pagination_list"}
                pagination_html = pagination_template.render(pagination_data)

                data = {'listing_html': listing_html, 'property_list': property_list, 'status': 200,
                        'msg': 'Successfully received', 'error': 0, 'total': property_total, 'pagination_html': pagination_html, 'no_page': total_page}
            else:
                data = {'status': 403, 'msg': 'Server error, Please try again', 'property_list': [], 'error': 1, 'total': 0, 'portfolio_detail': {}}
            return JsonResponse(data)
        else:
            page = 1
            page_size = 10
            search = ''
            listing_params = {
                "parcel_id": portfolio_id,
                "domain_id": domain_id,
                "page_size": page_size,
                "page": page,
                "search": search,
            }
            if 'user_id' in request.session and request.session['user_id']:
                listing_params['user_id'] = user_id
            listing_url = settings.API_URL + '/api-property/parcel-detail/'
            listing_data = call_api_post_method(listing_params, listing_url, token)

            if 'error' in listing_data and listing_data['error'] == 0:
                property_list = listing_data['data']['data']
                property_total = listing_data['data']['total']
                portfolio_detail = listing_data['data']['portfolio_detail']
            else:
                property_list = []
                property_total = 0
                portfolio_detail = {}

            # ---------------Pagination--------
            pagination_path = 'home/{}/parcel/pagination.html'.format(templete_dir)
            pagination_template = get_template(pagination_path)
            total_page = math.ceil(property_total / page_size)
            pagination_data = {"no_page": total_page, "total_page": range(total_page), "current_page": page,
                               "pagination_id": "prop_listing_pagination_list"}
            pagination_html = pagination_template.render(pagination_data)

            http_host = request.META['HTTP_HOST']
            site_url = settings.URL_SCHEME + str(http_host)
            # context = {
            #     'property_list': property_list,
            #     'total': property_total,
            #     'aws_url': settings.AWS_URL,
            #     'sess_user_id': user_id,
            #     'site_url': site_url,
            #     'pagination_html': pagination_html,
            #     'node_url': settings.NODE_URL,
            #     'domain_id': domain_id,
            #     'page': page,
            #     'no_page': total_page,
            # }
            context = {
                'property_list': property_list,
                'total': property_total,
                'aws_url': settings.AWS_URL,
                'sess_user_id': user_id,
                'site_url': site_url,
                'pagination_html': pagination_html,
                'node_url': settings.NODE_URL,
                'domain_id': domain_id,
                'page': page,
                'no_page': total_page,
                'parcel_id': portfolio_id,
                'portfolio_detail': portfolio_detail
            }
            theme_path = 'home/{}/parcel/our-listing.html'.format(templete_dir)
            return render(request, theme_path, context)
    except Exception as exp:
        print(exp)
        return HttpResponse("Issue in views")


def state_list(request):
    try:
        if request.is_ajax() and request.method == 'POST':
            country_id = request.POST['country_id']
            try:
                state_param = {'country_id': country_id}
                state_api_url = settings.API_URL + '/api-settings/get-state/'
                state_data = call_api_post_method(state_param, state_api_url)
                state_lists = state_data['data']
                data = {'error': 0, 'status': 200, 'state_lists': state_lists}
            except Exception as exp:
                data = {'error': 1, 'status': 403, 'state_lists': ''}
        else:
            data = {'error': 1, 'status': 403, 'state_lists': ''}
        return JsonResponse(data)
    except Exception as exp:
        data = {'error': 1, 'status': 403, 'msg': 'invalid request.'}
        return JsonResponse(data)


@csrf_exempt
def close_tour(request):
    try:
        if request.is_ajax() and request.method == 'POST':
            request.session['is_first_login'] = 0
            show_active_plan = 0
            if request.session['is_broker'] and request.session['is_free_plan']:
                show_active_plan = 1

            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']
            user_id = request.session['user_id']
            token = request.session['token']['access_token']
            # ---------------------Save Payment Detail Data------------------
            plan_price_id = int(site_detail['site_detail']['current_plan_price_id'])
            theme_id = int(site_detail['site_detail']['current_theme_id'])
            api_url = settings.API_URL + "/api-payments/create-payment-detail/"
            params = {
                "domain_id": site_id,
                "user_id": user_id,
                "plan_price_id": plan_price_id,
                "theme_id": theme_id
            }
            api_response = call_api_post_method(params, api_url, token)
            if 'error' in api_response and api_response['error'] == 0:
                # -------------Stripe Payment Credentials-------------
                api_url = settings.API_URL + "/api-payments/stripe-payment-detail/"
                params = {
                    "plan_price_id": plan_price_id,
                }
                api_response = call_api_post_method(params, api_url, token)
                if 'error' in api_response and api_response['error'] == 0:
                    data = api_response['data']
                    data = {"data": data, "email": request.session['email'], "show_active_plan": show_active_plan, 'status': 200, "error": 0, "msg": "success", "pay_redirect": request.session['is_first_admin_login']}
                else:
                    data = {'data': "", 'status': 403, 'error': 1, "msg": api_response['msg']}
            else:
                data = {'data': "", 'status': 403, 'error': 1, "msg": api_response['msg']}

        else:
            data = {'data': "", 'status': 403, 'error': 1, "msg": "Forbidden"}
        return JsonResponse(data)
    except Exception as exp:
        data = {'data': "", 'status': 403, 'error': 1, 'msg': 'invalid request.'}
        return JsonResponse(data)


def maintenance(request):
    """This function is used to maintenance

    """
    try:
        site_detail = subdomain_site_details(request)
        company_name = site_detail['site_detail']['company_name'].title() if site_detail['site_detail']['company_name'] else "Bidhom"
        theme_path = 'home/maintenance.html'
        return render(request, theme_path, context={'company_name': company_name})
    except Exception as exp:
        return HttpResponse("Issue in views")


@csrf_exempt
def check_payment(request):
    try:
        if request.is_ajax() and request.method == 'POST':
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']
            user_id = request.session['user_id']
            token = request.session['token']['access_token']

            # api_url = settings.API_URL + "/api-payments/check-payment-success/"
            api_url = settings.API_URL + "/api-payments/check-payment-subscription/"
            params = {
                "domain_id": site_id,
                "user_id": user_id,
            }
            api_response = call_api_post_method(params, api_url, token)
            if 'error' in api_response and api_response['error'] == 0:
                request.session['is_free_plan'] = False
                data = {"data": "", "error": 0, "msg": api_response['msg']}
            else:
                data = {'data': "", 'error': 1, "msg": api_response['msg']}

        else:
            data = {'data': "", 'error': 1, "msg": "Forbidden"}
        return JsonResponse(data)
    except Exception as exp:
        data = {'data': '', 'status': 403, 'error': 1, 'msg': 'invalid request.'}
        return JsonResponse(data)

def get_property_type(request):
    try:
        if request.is_ajax() and request.method == 'POST':
            asset_id = request.POST['asset_id']
            api_url = settings.API_URL + "/api-property/property-type/"
            params = {
                "asset_id": asset_id,
            }
            api_response = call_api_post_method(params, api_url)
            if 'error' in api_response and api_response['error'] == 0:
                data = {"data": api_response['data'], "error": 0, "msg": api_response['msg']}
            else:
                data = {'data': "", 'error': 1, "msg": api_response['msg']}
        else:
            data = {'data': "", 'error': 1, "msg": "Forbidden"}
        return JsonResponse(data)
    except Exception as exp:
        data = {'data': '', 'status': 403, 'error': 1, 'msg': 'invalid request.'}
        return JsonResponse(data)   


def bid_deposit_success(request):
    session_id = request.GET.get('session_id', None)
    try:
        try:
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']
            templete_dir = site_detail['site_detail']['theme_directory']
        except Exception as exp:
            site_id = ""
            templete_dir = 'theme-1'

        user_id = None
        token = None
        if 'user_id' in request.session and request.session['user_id']:
            token = request.session['token']['access_token']
            user_id = request.session['user_id']

        
        api_url = settings.API_URL + "/api-bid/bid-deposit-success/"
        params = {
            "session_id": session_id,
        }
        api_response = call_api_post_method(params, api_url, token)
        if 'error' in api_response and api_response['error'] == 0:
            property_id = api_response['data']['property_id']
        else:
            property_id = ""
        context = {'title': "Registration success", "property_id": property_id}
        theme_path = 'home/{}/listings/bid-deposit-success.html'.format(templete_dir)
        return render(request, theme_path, context)
    except Exception as exp:
        return HttpResponse("Issue in views") 


def bid_deposit_cancel(request):
    # session_id = request.GET.get('session_id', None)
    try:
        try:
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']
            templete_dir = site_detail['site_detail']['theme_directory']
        except Exception as exp:
            site_id = ""
            templete_dir = 'theme-1'
        context = {'title': "Registration Cancel"}
        theme_path = 'home/{}/listings/bid-deposit-cancel.html'.format(templete_dir)
        return render(request, theme_path, context)
    except Exception as exp:
        return HttpResponse("Issue in views")                