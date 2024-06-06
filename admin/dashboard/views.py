# -*- coding: utf-8 -*-
"""This file contains general view functions for this module
"""
import re
import json
import time
import datetime
from datetime import timedelta
from django.http import HttpResponseRedirect
from django.shortcuts import render, redirect
from django.http import HttpResponseForbidden
from django.conf import settings
from django.http import JsonResponse
from django.http import HttpResponse
# from home.services import *
from django.core.cache import cache
from django.conf import settings
from subdomain.services import call_api_get_method, call_api_post_method
from django.core.cache.backends.base import DEFAULT_TIMEOUT
from django.views.decorators.csrf import csrf_exempt
from subdomain.constants import settings_data
from packages.context_processors import subdomain_site_details, subdomain_admin_settings
from packages.globalfunction import *
from packages.constants import *
from django.template import RequestContext
from django.template.loader import get_template
from geopy.geocoders import Nominatim
import urllib.request
from urllib.request import urlopen, unquote
from urllib.parse import parse_qs, urlparse
from pyzipcode import ZipCodeDatabase
import math
import pandas as pd
import io
from io import BytesIO
from io import StringIO
from xhtml2pdf import pisa
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from django.template.defaulttags import register
CACHE_TTL = getattr(settings, 'CACHE_TTL', DEFAULT_TIMEOUT)


def dashboard(request):

    try:
        token = request.GET.get('token', None)

        try:
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']
        except:
            site_id = ""
        if token is not None and 'user_id' not in request.session:

            api_url = settings.API_URL + '/api-users/login-details/'
            payload = {
                'token': token,
                'site_id': site_id
            }
            response = call_api_post_method(payload, api_url)
            if "error" in response and response['error'] == 0:
                response = response['data']
                user_id = response['user_id']
                token = response['auth_token']['access_token']
                request.session['user_id'] = response['user_id']
                expires_in = int(response['auth_token']['expires_in'])

                expiry_date_time = datetime.datetime.now() + timedelta(seconds=expires_in)
                expiry_time = expiry_date_time.timestamp() * 1000

                request.session['token_expiry_time'] = expiry_time
                # request.session['site_id'] = response['site_id']
                request.session['token'] = response['auth_token']
                request.session['user_type'] = response['user_type']
                request.session['first_name'] = response['first_name']
                request.session['is_admin'] = response['is_admin']
                request.session['is_broker'] = response['is_broker']
                request.session['profile_image'] = response['profile_image']
                return HttpResponseRedirect('/admin/')

        if 'token' in request.session:
            token = request.session['token']['access_token']
        if 'user_id' in request.session:
            user_id = request.session['user_id']

        is_permission = check_permission(request, 5)

        if not is_permission:
            http_host = request.META['HTTP_HOST']
            redirect_url = settings.URL_SCHEME + str(http_host)
            return HttpResponseRedirect(redirect_url)

        try:
            subscription_plan_param = {
                'user_id': user_id
            }
            subscription_plan_url = settings.API_URL + '/api-settings/subscription-listing/'
            subscription_plan_data = call_api_post_method(subscription_plan_param, subscription_plan_url, token)
            subscription_plan_list = subscription_plan_data['data']
        except:
            subscription_plan_list = []

        try:
            current_plan_param = {'domain': site_id}
            current_plan_url = settings.API_URL + '/api-users/get-current-plan/'
            current_plan_data = call_api_post_method(current_plan_param, current_plan_url, token)
            current_plan_list = current_plan_data['data']
        except:
            current_plan_list = {}


        try:
            plan_detail_param = {
                'site_id': site_id,
                'user_id': user_id
            }
            plan_detail_url = settings.API_URL + '/api-users/plan-dashboard/'
            plan_detail_data = call_api_post_method(plan_detail_param, plan_detail_url, token)
            plan_details = plan_detail_data['data']
        except:
            plan_details = {}

        try:
            theme_param = {
                'user_id': user_id
            }
            theme_url = settings.API_URL + '/api-settings/theme-listing/'
            theme_data = call_api_post_method(theme_param, theme_url, token)
            theme_list = theme_data['data']
        except:
            theme_list = []

        try:
            page = 1
            page_size = 10
            plan_history_param = {
                'site_id': site_id,
                'user_id': user_id,
                'page': 1,
                'page_size': page_size
            }
            sno = (int(page) - 1) * int(page_size) + 1
            plan_history_url = settings.API_URL + '/api-users/plan-billing-history/'
            plan_history_data = call_api_post_method(plan_history_param, plan_history_url, token)
            if 'error' in plan_history_data and plan_history_data['error'] == 0:
                plan_history_list = plan_history_data['data']['data']
                total = plan_history_data['data']['total']
            else:
                plan_history_list = []
                total = 0
            # ---------------Pagination--------
            pagination_html = ''
            pagination_path = 'admin/dashboard/dashboard/history-pagination.html'
            pagination_template = get_template(pagination_path)
            total_page = math.ceil(total / page_size)

            if total_page > 1:
                pagination_data = {"no_page": total_page, "total_page": range(total_page), "current_page": 1,
                                   "pagination_id": "history_listing_pagination_list"}
                pagination_html = pagination_template.render(pagination_data)

        except Exception as exp:
            print(exp)
            plan_history_list = []
            pagination_html = ''
            total = 0


        context = {
            "data": "COMING SOON Dataa...",
            "active_menu": "site setting",
            "active_submenu": "dashboard",
            "subscription_plan_list": subscription_plan_list,
            "plan_details": plan_details,
            'theme_list': theme_list,
            'plan_history_list': plan_history_list,
            'pagination_html': pagination_html,
            'total': total,
            'stripe_publishable_key': settings.STRIPE_PUBLIC_KEY,
            'current_plan_list': current_plan_list
        }
        
        return render(request, "admin/dashboard/dashboard/dashboard.html", context)
    except Exception as exp:
        print(exp)
        return HttpResponse("Issue in views")


def dashboard_data(request):
    try:
        token = request.GET.get('token', None)

        try:
            is_free = subdomain_admin_settings(request)
            if is_free['settings_data']['is_free']:
                http_host = request.META['HTTP_HOST']
                redirect_url = settings.URL_SCHEME + str(http_host)
                return HttpResponseRedirect(redirect_url)
        except Exception as exp:
            http_host = request.META['HTTP_HOST']
            redirect_url = settings.URL_SCHEME + str(http_host)
            return HttpResponseRedirect(redirect_url)

        try:
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']
        except:
            site_id = ""

        # try:
        #     params = {
        #         "site_id": site_id
        #     }
        #     api_url = settings.API_URL + '/api-users/property-registration-graph/'
        #     property_registration_data = call_api_post_method(params, api_url, token)
        #     property_registration = property_registration_data['data']
        # except Exception as exp:
        #     property_registration = {}

        # try:
        #     params = {
        #         "site_id": site_id
        #     }
        #     api_url = settings.API_URL + '/api-users/signup-page-graph/'
        #     signup_view_data = call_api_post_method(params, api_url, token)
        #     signup_view = signup_view_data['data']
        # except Exception as exp:
        #     signup_view = {}

        # try:
        #     params = {
        #         "site_id": site_id
        #     }
        #     api_url = settings.API_URL + '/api-users/admin-dashboard/'
        #     dashboard_data = call_api_post_method(params, api_url, token)
        #     dashboard_data = dashboard_data['data']
        # except Exception as exp:
        #     dashboard_data = {}

        # try:
        #     params = {
        #         "site_id": site_id,
        #     }
        #     api_url = settings.API_URL + '/api-users/update-dashboard-map/'
        #     api_response = call_api_post_method(params, api_url, token)
        #     dashboard_map = api_response['data']
        # except Exception as exp:
        #     dashboard_map = {}

        context = {
            # "dashboard_data": dashboard_data,
            # "property_registration": property_registration,
            # "signup_view": signup_view,
            "title": "Dashboard",
            # "dashboard_map": dashboard_map
        }
        return render(request, "admin/dashboard/dashboard/dashboard_data.html", context)
    except Exception as exp:
        print(exp)
        return HttpResponse("Issue in views")


@csrf_exempt
def save_plan_change(request):
    try:
        if request.is_ajax() and request.method == 'POST':
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']
            user_id = request.session['user_id']
            token = request.session['token']['access_token']
            # ---------------------Save Payment Detail Data------------------
            plan_price_id = int(request.POST['plan_price_id'])
            theme_id = int(request.POST['theme_id'])
            api_url = settings.API_URL + "/api-payments/create-payment-detail/"
            params = {
                "domain_id": site_id,
                "user_id": user_id,
                "plan_price_id": plan_price_id,
                "theme_id": theme_id
            }
            api_response = call_api_post_method(params, api_url, token)
            if 'error' in api_response and api_response['error'] == 0:
                data = {"data": "", "error": 0, "msg": api_response['msg']}
            else:
                data = {'data': "", 'error': 1, "msg": api_response['msg']}

        else:
            data = {'data': "", 'error': 1, "msg": "Forbidden"}
        return JsonResponse(data)
    except Exception as exp:
        data = {'data': '', 'status': 403, 'error': 1, 'msg': 'invalid request.'}
        return JsonResponse(data)


def pay_now(request):
    try:
        token = request.session['token']['access_token']
        show_active_plan = 0
        if request.session['is_broker'] and request.session['is_free_plan']:
            show_active_plan = 1
        try:
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']
        except:
            site_id = ""
        create_payment_detail = None
        auto_payment_redirection = 0
        if token is not None and 'user_id' in request.session:
            if 'user_id' in request.session:
                user_id = request.session['user_id']
            else:
                user_id = ""

            if 'email' in request.session:
                email = request.session['email']
            else:
                email = ""

            try:
                params = {
                    'user_id': user_id,
                    'domain_id': site_id
                }
                api_url = settings.API_URL + "/api-payments/create-payment-data/"
                create_payment_detail = call_api_post_method(params, api_url, token)
                create_payment_detail = create_payment_detail['data']['data']
            except Exception as exp:
                create_payment_detail = {}

        context = {"data": create_payment_detail, "stripe_publishable_key": settings.STRIPE_PUBLIC_KEY, "email": email, "show_active_plan": show_active_plan}
        return render(request, "admin/dashboard/dashboard/pay-now.html", context)
    except Exception as exp:
        return HttpResponse("Issue in views")

# def login(request):
#     try:
#         token = request.GET.get('token', None)
#         if token is not None:
#             api_url = settings.API_URL + '/api-users/login-details/'
#             payload = {
#                 'token': token,
#             }
#             response = call_api_post_method(payload, api_url)
#             if "error" in response and response['error'] == 0:
#                 response = response['data']
#                 request.session['user_id'] = response['user_id']
#                 request.session['site_id'] = response['site_id']
#                 request.session['token'] = response['auth_token']
#                 request.session['first_name'] = response['first_name']
#         context = {"data": "Login Data"}
#         return render(request, "admin/home/login.html", context)
#     except Exception as exp:
#         return HttpResponse("Issue in views")


def logout(request):
    try:
        api_url = settings.API_URL + '/api-users/revoke-token/'
        payload = {
            'user_id': request.session['user_id'],
            'token': request.session['token']['access_token']
        }
        call_api_post_method(payload, api_url, request.session['token']['access_token'])
        del request.session['user_id']
        del request.session['profile_image']
        del request.session['token']
        del request.session['first_name']
        del request.session['token_expiry_time']
        request.session.modified = True
        request.session.flush()
        return HttpResponseRedirect(settings.BASE_URL)
    except Exception as exp:
        # print(exp)
        # return HttpResponse("Issue in views")
        return HttpResponseRedirect(settings.BASE_URL)


def profile(request):
    try:
        return HttpResponse("Profile Page")
    except Exception as exp:
        return HttpResponse("Issue in views")

def business_info(request):
    token = request.GET.get('token', None)

    try:
        site_detail = subdomain_site_details(request)
        site_id = site_detail['site_detail']['site_id']
    except:
        site_id = ""
    if token is not None and 'user_id' not in request.session:

        api_url = settings.API_URL + '/api-users/login-details/'
        payload = {
            'token': token,
            'site_id': site_id
        }
        response = call_api_post_method(payload, api_url)
        if "error" in response and response['error'] == 0:
            response = response['data']
            user_id = response['user_id']
            token = response['auth_token']['access_token']
            request.session['user_id'] = response['user_id']
            expires_in = int(response['auth_token']['expires_in'])

            expiry_date_time = datetime.datetime.now() + timedelta(seconds=expires_in)
            expiry_time = expiry_date_time.timestamp() * 1000

            request.session['token_expiry_time'] = expiry_time
            # request.session['site_id'] = response['site_id']
            request.session['token'] = response['auth_token']
            request.session['user_type'] = response['user_type']
            request.session['first_name'] = response['first_name']
            request.session['is_admin'] = response['is_admin']
            request.session['is_broker'] = response['is_broker']
            request.session['profile_image'] = response['profile_image']
            return HttpResponseRedirect('/admin/business-info/')
    try:
        is_permission = check_permission(request,7)
        if not is_permission:
            http_host = request.META['HTTP_HOST']
            redirect_url = settings.URL_SCHEME + str(http_host)
            return HttpResponseRedirect(redirect_url)
        try:
            business_id = ''
            token = request.session['token']['access_token']
            user_id = request.session['user_id']
            response = {
                'error': '',
                'msg': ''
            }
            try:
                site_detail = subdomain_site_details(request)
                site_id = site_detail['site_detail']['site_id']
            except:
                site_id = ""

            if request.is_ajax() and request.method == 'POST':
                addresses = []

                total = int(request.POST['total_section'])
                for i in range(total):
                    address_params = {
                        'address_first': request.POST['office_address_' + str(i)],
                        'state': request.POST['state_' + str(i)],
                        'postal_code': request.POST['zip_code_' + str(i)],
                    }
                    addresses.append(address_params)



                business_params = {
                    'site_id': site_id,
                    'user_id': user_id,
                    'first_name': request.POST['business_first_name'],
                    'last_name': request.POST['business_last_name'],
                    'company_name': request.POST['company_name'],
                    'mobile_no': re.sub('\D', '', request.POST['business_mobile']),
                    'phone_no': re.sub('\D', '', request.POST['business_phone']),
                    'email': request.POST['business_email'],
                    'address': addresses,
                    'licence_no': request.POST['broker_license_no'],
                    'company_logo': request.POST['business_logo_img_id'] if 'business_logo_img_id' in request.POST and request.POST['business_logo_img_id'] != "" else None,
                    'country': request.POST['country'],
                }
                update_business_url = settings.API_URL + '/api-users/update-business-info/'
                business_response = call_api_post_method(business_params, update_business_url, token)

                if 'error' in business_response and business_response['error'] == 0:
                    response = {
                        'error': 0,
                        'msg': 'Business info updated successfully.'
                    }
                else:
                    response = {
                        'error': 1,
                        'msg': 'Some error occur, please try again.'
                    }
                return JsonResponse(response)
            try:
                business_detail_param = {
                    'user_id': user_id
                }
                business_detail_url = settings.API_URL + '/api-users/get-business-info/'
                business_detail_data = call_api_post_method(business_detail_param, business_detail_url, token)
                business_details = business_detail_data['data']
                business_id = business_details['id']
                country_id = business_details['country']
            except:
                business_details = {}


            try:
                state_param = {"country_id": country_id}
                state_api_url = settings.API_URL + '/api-settings/get-state/'
                state_data = call_api_post_method(state_param, state_api_url, token)
                state_list = state_data['data']
            except:
                state_list = []

            try:
                country_param = {}
                country_api_url = settings.API_URL + '/api-settings/get-country/'
                country_data = call_api_post_method(country_param, country_api_url)
                country_list = country_data['data']
            except:
                country_list = []
        except:
            business_details = {}
            state_list = []
            country_list = []
        context = {"data": "COMING SOON Dataa...", "active_menu": "site setting", "active_submenu": "business", 'business_details': business_details,
                   'state_list': state_list, 'response': response, 'country_list': country_list}


        return render(request, "admin/dashboard/business/update-business-info.html", context)


    except Exception as exp:
        print(exp)
        return HttpResponse("Issue in views")

def website(request):
    is_permission = check_permission(request, 3)
    if not is_permission:
        http_host = request.META['HTTP_HOST']
        redirect_url = settings.URL_SCHEME + str(http_host)
        return HttpResponseRedirect(redirect_url)
    try:
        business_id = ''
        token = request.session['token']['access_token']
        user_id = request.session['user_id']
        website_setting_response = {
            'error': '',
            'msg': ''
        }
        try:
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']
            site_domain = site_detail['site_detail']['domain_name']
        except:
            site_id = ""
            site_domain = ""


        try:
            website_details_param = {
                'user_id': user_id,
                'site_id': site_id
            }

            website_details_url = settings.API_URL + '/api-users/subdomain-website-detail/'
            website_details_data = call_api_post_method(website_details_param, website_details_url, token)
            website_details = website_details_data['data']

            banner_img_list = website_details['banner_images']
            footer_img_list = website_details['footer_images']
            social_url_list = website_details['social_account']
            mls_type = website_details['mls_type']
            mls_configuration = website_details['mls_configuration']
            #about_img_list = website_details['about_images']
            str_banner_image_id = ''
            banner_image_ids = ''
            str_banner_image_name = ''
            banner_image_name = ''

            str_footer_image_id = ''
            footer_image_ids = ''
            str_footer_image_name = ''
            footer_image_name = ''

            str_about_image_id = ''
            about_image_ids = ''
            str_about_image_name = ''
            about_image_name = ''

            if banner_img_list:
                cnt = 0
                for banner in banner_img_list:
                    if cnt == 0:
                        str_banner_image_id = str(banner['upload_id'])
                        str_banner_image_name = str(banner['doc_file_name'])
                    else:
                        str_banner_image_id = str_banner_image_id+','+str(banner['upload_id'])
                        str_banner_image_name = str_banner_image_name+','+str(banner['doc_file_name'])
                    cnt += 1
                banner_image_ids = str_banner_image_id.rstrip(',')
                banner_image_name = str_banner_image_name.rstrip(',')
                website_details['banner_image_ids'] = banner_image_ids
                website_details['banner_image_name'] = banner_image_name
            else:
                website_details['banner_image_ids'] = banner_image_ids
                website_details['banner_image_name'] = banner_image_name

            if footer_img_list:
                cnt = 0
                for footer in footer_img_list:
                    if cnt == 0:
                        str_footer_image_id = str(footer['upload_id'])
                        str_footer_image_name = str(footer['doc_file_name'])
                    else:
                        str_footer_image_id = str_footer_image_id+','+str(footer['upload_id'])
                        str_footer_image_name = str_footer_image_name+','+str(footer['doc_file_name'])
                    cnt += 1
                footer_image_ids = str_footer_image_id.rstrip(',')
                footer_image_name = str_footer_image_name.rstrip(',')
                website_details['footer_image_ids'] = footer_image_ids
                website_details['footer_image_name'] = footer_image_name
            else:
                website_details['footer_image_ids'] = footer_image_ids
                website_details['footer_image_name'] = footer_image_name


            icon_list = [
                {
                    'id': 5,
                    'icon_name': 'residential-icon',
                    'icon_type_id': 1,
                    'icon_type_name': 'Residential'
                },
                {
                    'id': 6,
                    'icon_name': 'land-icon',
                    'icon_type_id': 2,
                    'icon_type_name': 'Land'
                },
                {
                    'id': 7,
                    'icon_name': 'commercial-icon',
                    'icon_type_id': 3,
                    'icon_type_name': 'Commercial'
                },
                {
                    'id': 8,
                    'icon_name': 'auction-icon',
                    'icon_type_id': 4,
                    'icon_type_name': 'Auction'
                }
            ]
            count_type = 0
            for exp_type in website_details['expertise']:
                filter_icon_list = []
                exp_type_id = exp_type['expertise_icon_type_id']

                try:
                    params = {"icon_type": exp_type_id}
                    url = settings.API_URL + '/api-users/get-expertise-icon/'
                    data = call_api_post_method(params, url, token)
                    icon_list = data['data']
                except:
                    icon_list = []
                exp_type['icon_list'] = icon_list
                exp_type['counter'] = count_type
                count_type = count_type+1


        except Exception as exp:
            print(exp)
            website_details = {}
            social_url_list = []


        pre_auction_type_list = []
        for i in range(6):
            type_name = ''
            if i == 0:
                type_name = 'Classic Online Auction'
            elif i == 1:
                type_name = 'Live Event Auction'
            elif i == 2:
                type_name = 'Insider Online Auction'
            elif i == 3:
                type_name = 'Multiple Parcel Online Auction'
            elif i == 4:
                type_name = 'Highest and Best Offer'
            elif i == 5:
                type_name = 'Traditional Listing'


            pre_auction_type_list.append(type_name)


        pre_expertise_list = []
        for i in range(4):
            type_name = ''
            if i == 0:
                type_name = 'House'
            elif i == 1:
                type_name = 'Land'
            elif i == 2:
                type_name = 'Commercial'
            elif i == 3:
                type_name = 'Auction'

            pre_expertise_list.append(type_name)



        pre_number_list = []
        for i in range(4):
            type_name = ''
            value = 0
            if i == 0:
                type_name = 'Best Downtown Locations'
                value = 10
            elif i == 1:
                type_name = 'Exceptional Properties'
                value = 26
            elif i == 2:
                type_name = 'Satisfied Buyer-Seller'
                value = 200
            elif i == 3:
                type_name = 'Email Subscriber'
                value = '+20K'

            pre_num_params = {
                'title': type_name,
                'value': value
            }

            pre_number_list.append(pre_num_params)


        pre_social_account_list = []
        if len(social_url_list) > 0:
            for item in social_url_list:
                li_title = ''
                li_icon = ''
                li_class_name = ''
                li_a_class_name = ''
                li_a_id = ''
                input_url_id = ''
                input_position_id = ''
                default_url = ''
                default_position = ''
                account_type = ''
                if item['account_type'] == 1:
                    li_title = 'facebook'
                    li_icon = 'fab fa-facebook-square'
                    li_class_name = 'facebook'
                    li_a_class_name = 'social_account'
                    li_a_id = 'fb_account'
                    input_url_id = 'facebook_url'
                    input_position_id = 'facebook_position'
                    default_url = item['url'] if 'url' in item and item['url'] else ""
                    default_position = item['position'] if 'position' in item and item['position'] else 1
                    account_type = item['account_type'] if 'account_type' in item and item['account_type'] == 1 else ""

                elif item['account_type'] == 2:
                    li_title = 'twitter'
                    li_icon = 'fab fa-twitter'
                    li_class_name = 'twitter'
                    li_a_class_name = 'social_account'
                    li_a_id = 'twitter_account'
                    input_url_id = 'twitter_url'
                    input_position_id = 'twitter_position'
                    default_url = item['url'] if 'url' in item and item['url'] is not None else ""
                    default_position = item['position'] if 'position' in item and item['position'] is not None else 2
                    account_type = item['account_type'] if 'account_type' in item and item['account_type'] == 2 else ""

                elif item['account_type'] == 3:
                    li_title = 'youtube'
                    li_icon = 'fab fa-youtube'
                    li_class_name = 'youtube'
                    li_a_class_name = 'social_account'
                    li_a_id = 'youtube_account'
                    input_url_id = 'youtube_url'
                    input_position_id = 'youtube_position'
                    default_url = item['url'] if 'url' in item and item['url'] is not None else ""
                    default_position = item['position'] if 'position' in item and item['position'] is not None else 3
                    account_type = item['account_type'] if 'account_type' in item and item['account_type'] == 3 else ""

                elif item['account_type'] == 4:
                    li_title = 'linkedin'
                    li_icon = 'fab fa-linkedin-in'
                    li_class_name = 'linkedin'
                    li_a_class_name = 'social_account'
                    li_a_id = 'linkedin_account'
                    input_url_id = 'linkedin_url'
                    input_position_id = 'linkedin_position'
                    default_url = item['url'] if 'url' in item and item['url'] is not None else ""
                    default_position = item['position'] if 'position' in item and item['position'] is not None else 4
                    account_type = item['account_type'] if 'account_type' in item and item['account_type'] == 4 else ""
                elif item['account_type'] == 5:
                    li_title = 'instagram'
                    li_icon = 'fab fa-instagram'
                    li_class_name = 'instagram'
                    li_a_class_name = 'social_account'
                    li_a_id = 'instagram_account'
                    input_url_id = 'instagram_url'
                    input_position_id = 'instagram_position'
                    default_url = item['url'] if 'url' in item and item['url'] is not None else ""
                    default_position = item['position'] if 'position' in item and item['position'] is not None else 5
                    account_type = item['account_type'] if 'account_type' in item and item['account_type'] == 5 else ""

                pre_social_params = {
                    'li_title': li_title,
                    'li_icon': li_icon,
                    'li_class_name': li_class_name,
                    'li_a_class_name': li_a_class_name,
                    'li_a_id': li_a_id,
                    'input_url_id': input_url_id,
                    'input_position_id': input_position_id,
                    'default_url': default_url,
                    'default_position': default_position,
                    'account_type': account_type,
                }

                pre_social_account_list.append(pre_social_params)

        else:
            for i in range(1, 6):
                li_title = ''
                li_icon = ''
                li_class_name = ''
                li_a_class_name = ''
                li_a_id = ''
                input_url_id = ''
                input_position_id = ''
                default_url = ''
                default_position = ''
                account_type = ''
                if i == 1:
                    li_title = 'facebook'
                    li_icon = 'fab fa-facebook-square'
                    li_class_name = 'facebook'
                    li_a_class_name = 'social_account'
                    li_a_id = 'fb_account'
                    input_url_id = 'facebook_url'
                    input_position_id = 'facebook_position'
                    default_url = 'https://www.facebook.com/'
                    default_position = i
                    account_type = 1
                elif i == 2:
                    li_title = 'twitter'
                    li_icon = 'fab fa-twitter'
                    li_class_name = 'twitter'
                    li_a_class_name = 'social_account'
                    li_a_id = 'twitter_account'
                    input_url_id = 'twitter_url'
                    input_position_id = 'twitter_position'
                    default_url = 'https://www.twitter.com/'
                    default_position = i
                    account_type = 2
                elif i == 3:
                    li_title = 'youtube'
                    li_icon = 'fab fa-youtube'
                    li_class_name = 'youtube'
                    li_a_class_name = 'social_account'
                    li_a_id = 'youtube_account'
                    input_url_id = 'youtube_url'
                    input_position_id = 'youtube_position'
                    default_url = 'https://www.youtube.com/'
                    default_position = i
                    account_type = 3
                elif i == 4:
                    li_title = 'linkedin'
                    li_icon = 'fab fa-linkedin-in'
                    li_class_name = 'linkedin'
                    li_a_class_name = 'social_account'
                    li_a_id = 'linkedin_account'
                    input_url_id = 'linkedin_url'
                    input_position_id = 'linkedin_position'
                    default_url = 'https://www.linkedin.com/'
                    default_position = i
                    account_type = 4
                elif i == 5:
                    li_title = 'instagram'
                    li_icon = 'fab fa-instagram'
                    li_class_name = 'instagram'
                    li_a_class_name = 'social_account'
                    li_a_id = 'instagram_account'
                    input_url_id = 'instagram_url'
                    input_position_id = 'instagram_position'
                    default_url = 'https://www.instagram.com/'
                    default_position = i
                    account_type = 5

                pre_social_params = {
                    'li_title': li_title,
                    'li_icon': li_icon,
                    'li_class_name': li_class_name,
                    'li_a_class_name': li_a_class_name,
                    'li_a_id': li_a_id,
                    'input_url_id': input_url_id,
                    'input_position_id': input_position_id,
                    'default_url': default_url,
                    'default_position': i,
                    'account_type': account_type,
                }

                pre_social_account_list.append(pre_social_params)


        asset_type = []
        if 'bot_setting' in website_details and website_details['bot_setting']:
            for obj in website_details['bot_setting']:
                asset_type.append(obj['property_type_id'])

        context = {"data": "COMING SOON Data...","active_menu": "site setting", "active_submenu": "website", 'response': website_setting_response, 'website_details': website_details, 'aws_url': settings.AWS_URL, 'pre_auction_type_list': pre_auction_type_list, 'pre_expertise_list': pre_expertise_list, 'pre_number_list': pre_number_list, 'pre_social_account_list': pre_social_account_list, 'asset_type': asset_type, 'mls_type': mls_type, 'mls_configuration': mls_configuration}
        return render(request, "admin/dashboard/website/website-setting.html", context)
    except Exception as exp:
        print(exp)
        return HttpResponse("Issue in views")

@csrf_exempt
def agents(request):
    try:
        is_permission = check_permission(request, 1)
        if not is_permission:
            http_host = request.META['HTTP_HOST']
            redirect_url = settings.URL_SCHEME + str(http_host)
            return HttpResponseRedirect(redirect_url)

        try:
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']

        except Exception as exp:
            print(exp)
            site_id = ""

        user_id = None
        token = None
        if 'user_id' in request.session and request.session['user_id']:
            token = request.session['token']['access_token']
            user_id = request.session['user_id']

        page_size = 10
        page = 1
        if request.is_ajax() and request.method == 'POST':

            agent_search = ''
            if 'search' in request.POST and request.POST['search']:
                agent_search = request.POST['search']

            page = 1
            if 'page' in request.POST and request.POST['page'] != "":
                page = request.POST['page']

            page_size = 10
            if 'perpage' in request.POST and request.POST['perpage']:
                page_size = request.POST['perpage']

            if 'status' in request.POST and request.POST['status'] and request.POST['status'] and request.POST[
                'status'].lower() == 'active':
                status = [1]
            elif 'status' in request.POST and request.POST['status'] and request.POST['status'] and request.POST[
                'status'].lower() == 'inactive':
                status = [2]
            else:
                status = [2, 1]

            list_param = {
                'site_id': site_id,
                "page": page,
                "user_id": user_id,
                "page_size": page_size,
                "status": status,
                "search": agent_search
            }
            sno = (int(page) - 1) * int(page_size) + 1
            list_url = settings.API_URL + '/api-users/subdomain-agent-listing/'
            list_data = call_api_post_method(list_param, list_url, token)

            if 'error' in list_data and list_data['error'] == 0:
                agent_list = list_data['data']['data']
                total = list_data['data']['total']
            else:
                agent_list = []
                total = 0
            context = {'agent_list': agent_list, 'total': total, "aws_url": settings.AWS_URL, 'sno': sno}

            agent_listing_path = 'admin/dashboard/agents/agent-listing-content.html'
            agent_listing_template = get_template(agent_listing_path)
            agent_listing_html = agent_listing_template.render(context)
            # ---------------Pagination--------
            pagination_path = 'admin/dashboard/agents/pagination.html'
            pagination_template = get_template(pagination_path)
            total_page = math.ceil(int(total) / int(page_size))
            pagination_html = ''
            if total_page > 1:
                pagination_data = {"no_page": int(total_page), "total_page": range(total_page), "current_page": int(page),
                                   "pagination_id": "agent_listing_pagination_list"}
                pagination_html = pagination_template.render(pagination_data)

            data = {'agent_listing_html': agent_listing_html, 'status': 200, 'msg': '', 'error': 0, 'total': total,
                    "pagination_html": pagination_html, 'pagination_id': 'agent_listing_pagination_list'}
            return JsonResponse(data)
        else:
            list_param = {
                'site_id': site_id,
                "page": 1,
                "user_id": user_id,
                "page_size": page_size,
                "status": [1],
                "search": ''
            }
            sno = (int(page) - 1) * int(page_size) + 1
            list_url = settings.API_URL + '/api-users/subdomain-agent-listing/'
            list_data = call_api_post_method(list_param, list_url, token)

            if 'error' in list_data and list_data['error'] == 0:
                agent_list = list_data['data']['data']
                total = list_data['data']['total']
            else:
                agent_list = []
                total = 0
            # ---------------Pagination--------
            pagination_html = ''
            pagination_path = 'admin/dashboard/agents/pagination.html'
            pagination_template = get_template(pagination_path)
            total_page = math.ceil(total / page_size)
            if total_page > 1:
                pagination_data = {"no_page": total_page, "total_page": range(total_page), "current_page": 1,
                                   "pagination_id": "user_listing_pagination_list"}
                pagination_html = pagination_template.render(pagination_data)
            context = {'agent_list': agent_list, 'total': total, "pagination_html": pagination_html,
                       "pagination_id": "agent_listing_pagination_list", "active_menu": "user setting", "active_submenu": "agent", "sno": sno}


            return render(request, "admin/dashboard/agents/view-agents.html", context)
    except Exception as exp:
        print(exp)
        return HttpResponse("Issue in views")

@csrf_exempt
def sub_admin(request):
    try:
        is_permission = check_permission(request, 1)
        if not is_permission:
            http_host = request.META['HTTP_HOST']
            redirect_url = settings.URL_SCHEME + str(http_host)
            return HttpResponseRedirect(redirect_url)

        try:
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']

        except Exception as exp:
            print(exp)
            site_id = ""

        user_id = None
        token = None
        if 'user_id' in request.session and request.session['user_id']:
            token = request.session['token']['access_token']
            user_id = request.session['user_id']

        page_size = 10
        page = 1
        if request.is_ajax() and request.method == 'POST':

            agent_search = ''
            if 'search' in request.POST and request.POST['search']:
                agent_search = request.POST['search']

            page = 1
            if 'page' in request.POST and request.POST['page'] != "":
                page = request.POST['page']

            page_size = 10
            if 'perpage' in request.POST and request.POST['perpage']:
                page_size = request.POST['perpage']

            if 'status' in request.POST and request.POST['status'] and request.POST['status'] and request.POST[
                'status'].lower() == 'active':
                status = [1]
            elif 'status' in request.POST and request.POST['status'] and request.POST['status'] and request.POST[
                'status'].lower() == 'inactive':
                status = [2]
            else:
                status = [2, 1]

            list_param = {
                'site_id': site_id,
                "page": page,
                "user_id": user_id,
                "page_size": page_size,
                "status": status,
                "search": agent_search
            }
            sno = (int(page) - 1) * int(page_size) + 1
            list_url = settings.API_URL + '/api-users/sub-admin-listing/'
            list_data = call_api_post_method(list_param, list_url, token)

            if 'error' in list_data and list_data['error'] == 0:
                agent_list = list_data['data']['data']
                total = list_data['data']['total']
            else:
                agent_list = []
                total = 0
            context = {'agent_list': agent_list, 'total': total, "aws_url": settings.AWS_URL, 'sno': sno}

            agent_listing_path = 'admin/dashboard/sub-admin/sub-admin-listing-content.html'
            agent_listing_template = get_template(agent_listing_path)
            agent_listing_html = agent_listing_template.render(context)
            # ---------------Pagination--------
            pagination_path = 'admin/dashboard/sub-admin/pagination.html'
            pagination_template = get_template(pagination_path)
            total_page = math.ceil(int(total) / int(page_size))
            pagination_html = ''
            if total_page > 1:
                pagination_data = {"no_page": int(total_page), "total_page": range(total_page), "current_page": int(page),
                                   "pagination_id": "agent_listing_pagination_list"}
                pagination_html = pagination_template.render(pagination_data)

            data = {'agent_listing_html': agent_listing_html, 'status': 200, 'msg': '', 'error': 0, 'total': total,
                    "pagination_html": pagination_html, 'pagination_id': 'agent_listing_pagination_list'}
            return JsonResponse(data)
        else:
            list_param = {
                'site_id': site_id,
                "page": 1,
                "user_id": user_id,
                "page_size": page_size,
                "status": [1],
                "search": ''
            }
            sno = (int(page) - 1) * int(page_size) + 1
            list_url = settings.API_URL + '/api-users/sub-admin-listing/'
            list_data = call_api_post_method(list_param, list_url, token)

            if 'error' in list_data and list_data['error'] == 0:
                agent_list = list_data['data']['data']
                total = list_data['data']['total']
            else:
                agent_list = []
                total = 0
            # ---------------Pagination--------
            pagination_html = ''
            pagination_path = 'admin/dashboard/agents/pagination.html'
            pagination_template = get_template(pagination_path)
            total_page = math.ceil(total / page_size)
            if total_page > 1:
                pagination_data = {"no_page": total_page, "total_page": range(total_page), "current_page": 1,
                                   "pagination_id": "user_listing_pagination_list"}
                pagination_html = pagination_template.render(pagination_data)
            context = {'agent_list': agent_list, 'total': total, "pagination_html": pagination_html,
                       "pagination_id": "agent_listing_pagination_list", "active_menu": "user setting", "active_submenu": "sub-admin", "sno": sno}


            return render(request, "admin/dashboard/sub-admin/view-sub-admin.html", context)
    except Exception as exp:
        print(exp)
        return HttpResponse("Issue in views")        

@csrf_exempt
def users(request):
    try:
        is_permission = check_permission(request, 4)
        if not is_permission:
            http_host = request.META['HTTP_HOST']
            redirect_url = settings.URL_SCHEME + str(http_host)
            return HttpResponseRedirect(redirect_url)

        try:
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']

        except Exception as exp:
            print(exp)
            site_id = ""

        user_id = None
        token = None
        if 'user_id' in request.session and request.session['user_id']:
            token = request.session['token']['access_token']
            user_id = request.session['user_id']

        try:
            state_param = {}
            state_api_url = settings.API_URL + '/api-settings/get-state/'
            state_data = call_api_post_method(state_param, state_api_url, token)
            state_list = state_data['data']
        except:
            state_list = []

        page_size = 10
        page = 1
        if request.is_ajax() and request.method == 'POST':

            user_search = ''
            if 'search' in request.POST and request.POST['search']:
                user_search = request.POST['search']

            page = 1
            if 'page' in request.POST and request.POST['page'] != "":
                page = request.POST['page']


            if 'perpage' in request.POST and request.POST['perpage']:
                page_size = request.POST['perpage']

            if 'status' in request.POST and request.POST['status'] and request.POST['status'].lower() == 'active':
                status = [1]
            elif 'status' in request.POST and request.POST['status'] and request.POST['status'] and request.POST[
                'status'].lower() == 'inactive':
                status = [2]
            else:
                status = [2, 1]

            list_param = {
                'site_id': site_id,
                "page": page,
                "page_size": page_size,
                "status": status,
                "search": user_search
            }
            sno = (int(page) - 1) * int(page_size) + 1
            list_url = settings.API_URL + '/api-users/subdomain-user-listing/'
            list_data = call_api_post_method(list_param, list_url, token)

            if 'error' in list_data and list_data['error'] == 0:
                user_listing = list_data['data']['data']
                total = list_data['data']['total']
            else:
                user_listing = []
                total = 0
            context = {'user_list': user_listing, 'total': total, "aws_url": settings.AWS_URL, 'sno': sno}

            user_listing_path = 'admin/dashboard/users/user-list-content.html'
            user_listing_template = get_template(user_listing_path)
            user_listing_html = user_listing_template.render(context)
            # ---------------Pagination--------
            pagination_path = 'admin/dashboard/users/pagination.html'
            pagination_template = get_template(pagination_path)
            total_page = math.ceil(int(total) / int(page_size))
            pagination_html = ''
            if total_page > 1:
                pagination_data = {"no_page": int(total_page), "total_page": range(total_page), "current_page": int(page),
                                   "pagination_id": "user_listing_pagination_list"}
                pagination_html = pagination_template.render(pagination_data)

            data = {'user_listing_html': user_listing_html, 'status': 200, 'msg': '', 'error': 0, 'total': total,
                    "pagination_html": pagination_html, 'pagination_id': 'user_listing_pagination_list'}
            return JsonResponse(data)
        else:
            list_param = {
                'site_id': site_id,
                "page": 1,
                "page_size": page_size,
                "status": [1],
                "search": ''
            }
            sno = (int(page) - 1) * int(page_size) + 1
            list_url = settings.API_URL + '/api-users/subdomain-user-listing/'
            list_data = call_api_post_method(list_param, list_url, token)

            if 'error' in list_data and list_data['error'] == 0:
                user_listing = list_data['data']['data']
                total = list_data['data']['total']
            else:
                user_listing = []
                total = 0
            # ---------------Pagination--------
            pagination_html = ''
            pagination_path = 'admin/dashboard/users/pagination.html'
            pagination_template = get_template(pagination_path)
            total_page = math.ceil(total / page_size)
            if total_page > 1:
                pagination_data = {"no_page": total_page, "total_page": range(total_page), "current_page": 1,
                                   "pagination_id": "user_listing_pagination_list"}
                pagination_html = pagination_template.render(pagination_data)
            context = {'user_list': user_listing, 'total': total, "pagination_html": pagination_html,
                       "pagination_id": "user_listing_pagination_list", "active_menu": "user setting", "active_submenu": "users", "state_list": state_list, "sno": sno}


            return render(request, "admin/dashboard/users/view-users.html", context)
    except Exception as exp:
        print(exp)
        return HttpResponse("Issue in views")

@csrf_exempt
def listing(request):
    try:
        listing_type = request.GET.get('auction_type', None)

        is_permission = check_permission(request, 6)
        if not is_permission:
            http_host = request.META['HTTP_HOST']
            redirect_url = settings.URL_SCHEME + str(http_host)
            return HttpResponseRedirect(redirect_url)

        try:
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']

        except Exception as exp:
            print(exp)
            site_id = ""

        user_id = None
        token = None
        is_broker = 0
        if 'user_id' in request.session and request.session['user_id']:
            token = request.session['token']['access_token']
            user_id = request.session['user_id']
            is_broker = 1 if request.session['is_broker'] == True else 0


        try:
            auction_type_param = {}
            auction_type_url = settings.API_URL + '/api-settings/subdomain-auction-type/'
            auction_type_data = call_api_post_method(auction_type_param, auction_type_url, token)
            auction_type_list = auction_type_data['data']
        except:
            auction_type_list = []

        try:
            asset_listing_params = {}

            asset_listing_url = settings.API_URL + '/api-property/asset-listing/'

            asset_listing_data = call_api_post_method(asset_listing_params, asset_listing_url, token)
            asset_listing = asset_listing_data['data']
        except:
            asset_listing = []

        try:
            params = {"domain_id": site_id, "user_id": user_id}
            api_url = settings.API_URL + '/api-users/agent-listing/'
            auction_type_data = call_api_post_method(params, api_url, token)
            agent_list = auction_type_data['data']
        except:
            agent_list = []

        try:
            property_type_params = {}

            property_type_url = settings.API_URL + '/api-property/property-type-listing/'

            property_type_data = call_api_post_method(property_type_params, property_type_url, token)
            property_type_listing = property_type_data['data']
        except:
            property_type_listing = []

        try:
            status_param = {'object_id': 9}
            status_url = settings.API_URL + '/api-settings/lookup-status-listing/'
            status_data = call_api_post_method(status_param, status_url, token)
            status_list = status_data['data']
        except:
            status_list = []

        page_size = 10
        if request.is_ajax() and request.method == 'POST':
            search = ''
            if 'search' in request.POST and request.POST['search']:
                search = request.POST['search']

            page = 1
            if 'page' in request.POST and request.POST['page'] != "":
                page = int(request.POST['page'])

            page_size = 10
            if 'perpage' in request.POST and request.POST['perpage']:
                page_size = int(request.POST['perpage'])

            asset_type = ''
            if 'asset_type' in request.POST and request.POST['asset_type']:
                asset_type = int(request.POST['asset_type'])
            auction_type = ''
            if 'auction_type' in request.POST and request.POST['auction_type']:
                auction_type = int(request.POST['auction_type'])
            property_type = ''
            if 'property_type' in request.POST and request.POST['property_type']:
                property_type = int(request.POST['property_type'])

            status = ""
            if 'status' in request.POST and request.POST['status']:
                status = request.POST['status']
            
            agent = ""
            if 'agent' in request.POST and request.POST['agent']:
                agent = request.POST['agent']

            list_param = {
                "page": page,
                "page_size": page_size,
                "site_id": site_id,
                "user_id": user_id,
                "auction_id": auction_type,
                "asset_id": asset_type,
                "property_type": property_type,
                "search": search,
                "status": status,
                "agent_id":agent,
            }

            list_url = settings.API_URL + '/api-property/property-listing/'

            list_data = call_api_post_method(list_param, list_url, token)

            if 'error' in list_data and list_data['error'] == 0:
                property_listing = list_data['data']['data']
                total = list_data['data']['total'] if 'total' in list_data['data'] else 0
                user_domain = list_data['data']['user_domain']
            else:
                property_listing = []
                total = 0
            sno = (int(page) - 1) * int(page_size) + 1
            context = {'property_list': property_listing, 'total': total, "aws_url": settings.AWS_URL, "is_broker": is_broker, 'status_list': status_list, "sno": sno, "auction_id": auction_type, "user_domain": user_domain}

            property_listing_path = 'admin/dashboard/listings/property_listing_content.html'
            property_listing_template = get_template(property_listing_path)
            property_listing_html = property_listing_template.render(context)
            # ---------------Pagination--------
            pagination_path = 'admin/dashboard/listings/property-pagination.html'
            pagination_template = get_template(pagination_path)
            total_page = math.ceil(int(total) / int(page_size))
            pagination_html = ''
            if total_page > 1:
                pagination_data = {"no_page": int(total_page), "total_page": range(total_page), "current_page": int(page),
                                   "pagination_id": "prop_listing_pagination_list"}
                pagination_html = pagination_template.render(pagination_data)

            data = {'property_listing_html': property_listing_html, 'status': 200, 'msg': '', 'error': 0, 'total': total,
                    "pagination_html": pagination_html, 'pagination_id': 'prop_listing_pagination_list', "sno": sno, "page": page, "auction_id": auction_type, 'status': status}
            return JsonResponse(data)
        else:
            page = 1
            asset_type = request.GET.get('asset_type', '')
            page_size = request.GET.get('record_per_page', 10)
            search = request.GET.get('search', '')
            status = request.GET.get('status', 1)
            agent = request.GET.get('agent', '')
            list_param = {
                "page": page,
                "page_size": page_size,
                "site_id": site_id,
                "user_id": user_id,
                "auction_id": "",
                "asset_id": "",
                "property_type": "",
                "search": search,
                "status": status,
                "agent_id": agent,
            }
            if listing_type and listing_type.lower() == 'traditional offer':
                auction_id = 4
                list_param['auction_id'] = 4
            elif listing_type and listing_type.lower() == 'highest offer':
                auction_id = 7
                list_param['auction_id'] = 7
            elif listing_type and listing_type.lower() == 'live offer':
                auction_id = 6
                list_param['auction_id'] = 6
            elif listing_type and listing_type.lower() == 'insider auction':
                auction_id = 2
                list_param['auction_id'] = 2
            elif listing_type and listing_type.lower() == 'classic online auction':
                auction_id = 1
                list_param['auction_id'] = 1
            else:
                auction_id = ""
                list_param['auction_id'] = ""

            if asset_type and asset_type.lower() == 'land':
                asset_type_id = 1
                list_param['asset_id'] = 1
            elif asset_type and asset_type.lower() == 'commercial':
                asset_type_id = 2
                list_param['asset_id'] = 2
            elif asset_type and asset_type.lower() == 'residential':
                asset_type_id = 3
                list_param['asset_id'] = 3
            else:
                asset_type_id = ''
                list_param['asset_id'] = ''

            list_url = settings.API_URL + '/api-property/property-listing/'

            list_data = call_api_post_method(list_param, list_url, token)
            if 'error' in list_data and list_data['error'] == 0:
                property_listing = list_data['data']['data']
                total = list_data['data']['total'] if 'total' in list_data['data'] else 0
                user_domain = list_data['data']['user_domain']

            else:
                property_listing = []
                total = 0
            # ---------------Pagination--------
            pagination_html = ''
            pagination_path = pagination_path = 'admin/dashboard/listings/property-pagination.html'
            pagination_template = get_template(pagination_path)

            total_page = math.ceil(total / int(page_size))
            sno = (int(page) - 1) * int(page_size) + 1
            if total_page > 1:
                pagination_data = {"no_page": total_page, "total_page": range(total_page), "current_page": 1,
                                   "pagination_id": "prop_listing_pagination_list"}
                pagination_html = pagination_template.render(pagination_data)



            context = {
                "data": "Listing COMING SOON...",
                "active_menu": "listing setting",
                "active_submenu": "listing",
                "asset_listing": asset_listing,
                "auction_type_list": auction_type_list,
                "property_type_listing": property_type_listing,
                "status_list": status_list,
                "pagination_html": pagination_html,
                "property_list": property_listing,
                "is_broker": is_broker,
                "sno": sno,
                "agent_list": agent_list,
                "auction_id": auction_id,
                "status": int(status) if status else "",
                "asset_type_id": asset_type_id,
                "search": search,
                "page_size": int(page_size),
                "agent": int(agent) if agent else "",
                "user_domain": user_domain
            }
            return render(request, "admin/dashboard/listings/property-listing.html", context)
    except Exception as exp:
        print(exp)
        return HttpResponse("Issue in views")


def save_user_plan(request):
    try:
        if request.is_ajax() and request.method == 'POST':
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']
            user_id = request.session['user_id']
            params = {
                'site_id': site_id,
                'user_id': request.session['user_id'],
                'opted_plan': request.POST['plan_id'],
                'theme_id': request.POST['theme_id'],
                'cc_ac_no': request.POST['card_number'],
                'cc_ac_type_name': request.POST['card_user_name']
            }
            token = request.session['token']['access_token']

            url = settings.API_URL + '/api-users/change-plan/'
            data = call_api_post_method(params, url, token)

        else:
            data = {'status': 403, 'msg': 'Forbidden'}

        return JsonResponse(data)
    except Exception as exp:
        print(exp)
        data = {'status': 403, 'msg': 'invalid request.'}
        return JsonResponse(data)

@csrf_exempt
def check_user_exists(request):
    try:
        if request.is_ajax() and request.method == 'POST':
            if 'email' in request.POST and request.POST['email'] != "":
                email = request.POST['email']
            elif 'user_email' in request.POST and request.POST['user_email'] != "":
                email = request.POST['user_email']
            elif 'business_email' in request.POST and request.POST['business_email'] != "":
                email = request.POST['business_email']
            elif 'usr_email' in request.POST and request.POST['usr_email'] != "":
                email = request.POST['usr_email']
            else:
                email = ""

            if 'phone' in request.POST and request.POST['phone'] != "":
                phone = re.sub('\D', '', request.POST['phone'])
            elif 'business_phone' in request.POST and request.POST['business_phone'] != "":
                phone = re.sub('\D', '', request.POST['business_phone'])
            elif 'usr_phone_no' in request.POST and request.POST['usr_phone_no'] != "":
                phone = re.sub('\D', '', request.POST['usr_phone_no'])
            elif 'user_phone_no' in request.POST and request.POST['user_phone_no'] != "":
                phone = re.sub('\D', '', request.POST['user_phone_no'])
            else:
                phone = ""
            user_id = request.session['user_id']
            if 'user_id' in request.POST:
                user_id = request.POST['user_id']

            params = {
                'user_id': user_id,
                'email': email,
                'phone': phone,
                'check_type': request.POST['check_type']
            }
            url = settings.API_URL + '/api-users/check-user-exists/'
            data = call_api_post_method(params, url)
        else:
            data = {'status': 403, 'msg': 'Forbidden'}


        return JsonResponse(data)
    except Exception as exp:
        print(exp)
        data = {'status': 403, 'msg': 'invalid request.'}
        return JsonResponse(data)

@csrf_exempt
def user_details(request):
    try:
        user_id = request.GET.get('user_id', None)
        if not user_id:
            http_host = request.META['HTTP_HOST']
            redirect_url = settings.URL_SCHEME + str(http_host)
            return HttpResponseRedirect(redirect_url)

        try:
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']
        except Exception as exp:
            print(exp)
            site_id = ""
        
        token = None
        if 'user_id' in request.session and request.session['user_id']:
            token = request.session['token']['access_token']

        page = 1
        page_size = 10
        if request.is_ajax() and request.method == 'POST':
            
            if 'page' in request.POST and request.POST['page'] != "":
                page = request.POST['page']

            if 'perpage' in request.POST and request.POST['perpage'] != "":
                page_size = request.POST['perpage']
            
            param = {
                "site_id": site_id,
                "page": page,
                "page_size": page_size,
                "user_id": user_id
            }

            sno = (int(page) - 1) * int(page_size) + 1
            api_url = settings.API_URL + '/api-users/subdomain-user-registration/'
            response_data = call_api_post_method(param, api_url, token)

            if 'error' in response_data and response_data['error'] == 0:
                bid_list = response_data['data']['data']
                user_detail = response_data['data']['user_detail']
                total = response_data['data']['total']
            else:
                bid_list = []
                user_detail = {}
                total = 0
            
            context = {"bid_list": bid_list, "total": total, "aws_url": settings.AWS_URL, "sno": sno, "user_detail": user_detail}
            user_details_bid_path = 'admin/dashboard/users/view-user-details-content.html'
            user_details_bid_template = get_template(user_details_bid_path)
            user_details_bid_html = user_details_bid_template.render(context)
            # ---------------Pagination--------
            pagination_path = 'admin/dashboard/users/user_details_pagination.html'
            pagination_template = get_template(pagination_path)
            total_page = math.ceil(int(total) / int(page_size))
            pagination_html = ''
            if total_page > 1:
                pagination_data = {"no_page": int(total_page), "total_page": range(total_page), "current_page": int(page),
                                   "pagination_id": "user_bid_listing_pagination_list"}
                pagination_html = pagination_template.render(pagination_data)

            data = {'user_details_bid_html': user_details_bid_html, 'status': 200, 'msg': '', 'error': 0, 'total': total,
                    "pagination_html": pagination_html, 'pagination_id': 'user_bid_listing_pagination_list'}
            return JsonResponse(data)
        else:
            param = {
                "site_id": site_id,
                "user_id": user_id,
                "page": page,
                "page_size": page_size
            }
            sno = (int(page) - 1) * int(page_size) + 1    
            api_url = settings.API_URL + '/api-users/subdomain-user-registration/'                
            response_data = call_api_post_method(param, api_url, token)
            
            if 'error' in response_data and response_data['error'] == 0:
                bid_list = response_data['data']['data']
                user_detail = response_data['data']['user_detail']
                total = response_data['data']['total']
            else:
                bid_list = []
                user_detail = {}
                total = 0
            # ---------------Pagination--------
            pagination_html = ''
            pagination_path = 'admin/dashboard/users/user_details_pagination.html'
            pagination_template = get_template(pagination_path)
            total_page = math.ceil(total / page_size)
            if total_page > 1:
                pagination_data = {"no_page": total_page, "total_page": range(total_page), "current_page": 1,
                                   "pagination_id": "user_bid_listing_pagination_list"}
                pagination_html = pagination_template.render(pagination_data)

            context = {"bid_list": bid_list, "total": total, "user_detail": user_detail, "pagination_html": pagination_html,
                       "pagination_id": "user_bid_listing_pagination_list", "active_menu": "user setting", "active_submenu": "users", "sno": sno, "user_id": user_id}
            return render(request, "admin/dashboard/users/view-user-details.html", context)
    except Exception as exp:
        print(exp)
        return HttpResponse("Issue in views")

def save_personal_info(request):
    try:
        if request.is_ajax() and request.method == 'POST':
            try:
                site_detail = subdomain_site_details(request)
                site_id = site_detail['site_detail']['site_id']
            except Exception as exp:
                print(exp)
                site_id = ""

            user_id = None
            token = None
            if 'user_id' in request.session and request.session['user_id']:
                user_id = request.session['user_id']
                token = request.session['token']['access_token']

            upload_id = request.POST['loggedin_user_img_id'] if 'loggedin_user_img_id' in request.POST and request.POST['loggedin_user_img_id'] != "" else None
            upload_name = request.POST['loggedin_user_img_name'] if 'loggedin_user_img_name' in request.POST and request.POST['loggedin_user_img_name'] != "" else None
            params = {
                'user_id': user_id,
                'first_name': request.POST['first_name'],
                'last_name': request.POST['last_name'],
                'phone_no': request.POST['usr_phone_no'],
                'email': request.POST['user_email'],
                'upload_id': request.POST['loggedin_user_img_id'] if 'loggedin_user_img_id' in request.POST and request.POST['loggedin_user_img_id'] != "" else None
            }
            url = settings.API_URL + '/api-users/update-personal-info/'
            data = call_api_post_method(params, url, token)

            if upload_id and upload_name and "error" in data and data['error'] == 0:
                api_url = settings.API_URL + '/api-users/login-details/'
                payload = {
                    'token': token,
                    'site_id': site_id
                }
                response = call_api_post_method(payload, api_url)
                if "error" in response and response['error'] == 0:
                    response = response['data']
                    request.session.modified = True
                    request.session['profile_image'] = response['profile_image']

        else:
            data = {'status': 403, 'msg': 'Forbidden'}

        return JsonResponse(data)
    except Exception as exp:
        print(exp)
        data = {'status': 403, 'msg': 'invalid request.'}
        return JsonResponse(data)

def change_admin_password(request):
    try:
        if request.is_ajax() and request.method == 'POST':
            params = {
                'user_id': request.session['user_id'],
                'password': request.POST['current_password'],
                'new_password': request.POST['new_password']
            }
            token = request.session['token']['access_token']
            url = settings.API_URL + '/api-users/user-change-password/'
            data = call_api_post_method(params, url, token)

        else:
            data = {'status': 403, 'msg': 'Forbidden'}

        return JsonResponse(data)
    except Exception as exp:
        print(exp)
        data = {'status': 403, 'msg': 'invalid request.'}
        return JsonResponse(data)

@csrf_exempt
def delete_website_user(request):
    try:
        if request.is_ajax() and request.method == 'POST':

            page_size = 10
            user_search = ''
            if 'search' in request.POST and request.POST['search']:
                user_search = request.POST['search']

            page = 1
            if 'page' in request.POST and request.POST['page'] != "":
                page = request.POST['page']

            if 'perpage' in request.POST and request.POST['perpage']:
                page_size = request.POST['perpage']

            if 'status' in request.POST and request.POST['status'] and request.POST['status'].lower() == 'active':
                status = [1]
            elif 'status' in request.POST and request.POST['status'] and request.POST['status'] and request.POST[
                'status'].lower() == 'inactive':
                status = [2]
            else:
                status = [2, 1]


            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']
            params = {
                'site_id': site_id,
                'user_id': request.POST['user_id']
            }
            token = request.session['token']['access_token']
            url = settings.API_URL + '/api-users/subdomain-delete-user/'
            data = call_api_post_method(params, url, token)
            user_list = []
            if 'error' in data and data['error'] == 0:
                list_param = {
                    'site_id': site_id,
                    "page": page,
                    "page_size": page_size,
                    "status": status,
                    "search": user_search
                }
                list_url = settings.API_URL + '/api-users/subdomain-user-listing/'
                list_data = call_api_post_method(list_param, list_url, token)

                if 'error' in list_data and list_data['error'] == 0:
                    user_listing = list_data['data']['data']
                    total = list_data['data']['total']
                else:
                    user_listing = []
                    total = 0
                sno = (int(page) - 1) * int(page_size) + 1
                context = {'user_list': user_listing, 'total': total, "aws_url": settings.AWS_URL, 'sno': sno}

                user_listing_path = 'admin/dashboard/users/user-list-content.html'
                user_listing_template = get_template(user_listing_path)
                user_listing_html = user_listing_template.render(context)
                # ---------------Pagination--------
                pagination_path = 'admin/dashboard/users/pagination.html'
                pagination_template = get_template(pagination_path)
                total_page = math.ceil(int(total) / int(page_size))
                pagination_html = ''
                if total_page > 1:
                    pagination_data = {"no_page": int(total_page), "total_page": range(total_page),
                                       "current_page": int(page),
                                       "pagination_id": "user_listing_pagination_list"}
                    pagination_html = pagination_template.render(pagination_data)

                data = {'error': 0, 'status': 200, 'user_listing_html': user_listing_html, 'data': data, 'total': total, "pagination_html": pagination_html,
                           "pagination_id": "user_listing_pagination_list"}
            else:
                data = {'error': 1, 'status': 403, 'msg': 'Server error, Please try again', 'user_listing_html': '', 'data': data, 'total': 0, "pagination_html": ''}
        else:
            data = {'error': 1, 'status': 403, 'msg': 'Forbidden', 'user_listing_html': '', 'data': {}, 'total': 0, "pagination_html": ''}

        return JsonResponse(data)
    except Exception as exp:
        print(exp)
        data = {'status': 403, 'msg': 'invalid request.', 'user_list': []}
        return JsonResponse(data)

@csrf_exempt
def save_images(request):
    try:
        user_id = request.session['user_id']
        site_detail = subdomain_site_details(request)
        site_id = site_detail['site_detail']['site_id']
        token = request.session['token']['access_token']
        file_urls = ""
        upload_to = ""
        uploaded_file_list = []
        file_size = request.POST['file_size']

        try:
            for key, value in request.FILES.items():
                params = {}

                if 'banner_img' in key.lower():
                    upload_to = 'banner_img'
                    doc_type = 1
                elif 'about_us_img' in key.lower():
                    upload_to = 'about_us_img'
                    doc_type = 3
                elif 'company_partner_img' in key.lower():
                    upload_to = 'company_partner_img'
                    doc_type = 8
                elif 'favicon_image' in key.lower():
                    doc_type = 4
                    upload_to = 'favicons'
                    file_urls = request.FILES['favicon_image']
                elif 'website_logo' in key.lower():
                    doc_type = 5
                    upload_to = 'website_logo'
                    file_urls = request.FILES['website_logo']

                elif 'testimonial_img' in key.lower():
                    doc_type = 6
                    upload_to = 'testimonial_img'
                    file_urls = request.FILES['testimonial_img']
                elif 'testimonial_author_img' in key.lower():
                    doc_type = 7
                    upload_to = 'testimonial_author_img'
                    file_urls = request.FILES['testimonial_author_img']
                elif 'agent_image' in key.lower():
                    doc_type = 9
                    upload_to = 'profile_image'
                    file_urls = request.FILES['agent_image']
                elif 'user_image' in key.lower():
                    doc_type = 9
                    upload_to = 'profile_image'
                    file_urls = request.FILES['user_image']
                elif 'loggedin_usr_image' in key.lower():
                    doc_type = 9
                    upload_to = 'profile_image'
                    file_urls = request.FILES['loggedin_usr_image']
                elif 'home_auction_image' in key.lower():
                    doc_type = 10
                    #upload_to = 'home_auction_image'
                    upload_to = 'home_auctioin'
                    file_urls = request.FILES['home_auction_image']
                elif 'expertise_image' in key.lower():
                    doc_type = 11
                    #upload_to = 'expertise_image'
                    upload_to = 'home_expertise'
                    file_urls = request.FILES['expertise_image']
                elif 'property_image' in key.lower():
                    doc_type = 13
                    upload_to = 'property_image'
                    #file_urls = request.FILES['property_image']
                elif 'property_document' in key.lower():
                    doc_type = 12
                    upload_to = 'property_document'
                    #file_urls = request.FILES['property_document']
                elif 'agent_logo_image' in key.lower():
                    doc_type = 14
                    upload_to = 'company_logo'
                    file_urls = request.FILES['agent_logo_image']
                elif 'business_logo_image' in key.lower():
                    doc_type = 14
                    upload_to = 'company_logo'
                    file_urls = request.FILES['business_logo_image']
                elif 'chat_document' in key.lower():
                    doc_type = 20
                    upload_to = 'chat_document'
                elif 'portfolio' in key.lower():
                    doc_type = 23
                    upload_to = 'portfolio'
                    # file_urls = request.FILES['chat_document']
                file_urls = request.FILES[key]
        except:
            pass


        if int(request.POST['file_length']) > 1:
            for key, value in request.FILES.items():
                params = {}
                if 'banner_img' in key.lower():
                    upload_to = 'banner_img'
                    doc_type = 1
                elif 'about_us_img' in key.lower():
                    upload_to = 'about_us_img'
                    doc_type = 3
                elif 'property_image' in key.lower():
                    doc_type = 13
                    upload_to = 'property_image'
                elif 'property_document' in key.lower():
                    doc_type = 12
                    upload_to = 'property_document'
                elif 'chat_document' in key.lower():
                    doc_type = 20
                    upload_to = 'chat_document'
                elif 'portfolio' in key.lower():
                    doc_type = 23
                    upload_to = 'portfolio'
                else:
                    upload_to = 'company_partner_img'
                    doc_type = 8

                file_res = request.FILES[key]
                response = save_to_s3(file_res, upload_to)
                if 'error' in response and response['error'] == 0:
                    try:
                        upload_param = {
                            "site_id": "",
                            "user_id": user_id,
                            "doc_file_name": response['file_name'],
                            "document_type": doc_type,
                            "bucket_name": upload_to,
                            "added_by": user_id,
                            "file_size": str(file_size)+'MB'
                        }
                        url = settings.API_URL + '/api-users/file-upload/'
                        upload_data = call_api_post_method(upload_param, url, token)
                        upload_id = upload_data['data']['upload_id']
                        upload_size = upload_data['data']['file_size']
                        upload_date = upload_data['data']['added_date']
                    except:
                        upload_id = 0
                        upload_size = '0MB'
                        upload_date = ''

                    params['file_name'] = response['file_name']
                    params['error'] = 0
                    params['msg'] = response['msg']
                    params['upload_id'] = upload_id
                    params['file_size'] = upload_size
                    params['upload_date'] = upload_date
                    params['upload_to'] = upload_to
                else:
                    params['file_name'] = response['file_name']
                    params['error'] = 1
                    params['msg'] = response['msg']
                    params['upload_id'] = 0
                    params['file_size'] = '0MB'
                    params['upload_date'] = ''
                    params['upload_to'] = upload_to

                uploaded_file_list.append(params)
        else:
            params = {}
            response = save_to_s3(file_urls, upload_to)
            if 'error' in response and response['error'] == 0:
                try:
                    print("oneeeee")
                    upload_param = {
                        "site_id": "",
                        "user_id": user_id,
                        "doc_file_name": response['file_name'],
                        "document_type": doc_type,
                        "bucket_name": upload_to,
                        "added_by": user_id,
                        "file_size": str(file_size) + 'MB'
                    }
                    url = settings.API_URL + '/api-users/file-upload/'
                    upload_data = call_api_post_method(upload_param, url, token)
                    print(upload_data)
                    upload_id = upload_data['data']['upload_id']
                    upload_size = upload_data['data']['file_size']
                    upload_date = upload_data['data']['added_date']
                except Exception as exp:
                    print("twoooooo")
                    print(exp)
                    upload_id = 0
                    upload_size = '0MB'
                    upload_date = ''

                params['file_name'] = response['file_name']
                params['error'] = 0
                params['msg'] = response['msg']
                params['upload_id'] = upload_id
                params['file_size'] = upload_size
                params['upload_date'] = upload_date
                params['upload_to'] = upload_to
            else:
                params['file_name'] = response['file_name']
                params['error'] = 1
                params['msg'] = response['msg']
                params['upload_id'] = 0
                params['file_size'] = '0MB'
                params['upload_date'] = ''
                params['upload_to'] = upload_to
            uploaded_file_list.append(params)



        return JsonResponse({'status': 200, 'uploaded_file_list': uploaded_file_list})
    except Exception as exp:
        print(exp)
        data = {'status': 403, 'msg': 'invalid request.'}
        return JsonResponse(data)

@csrf_exempt
def delete_images(request):
    try:
        return_data = {
            'article_id': request.POST['article_id'],
            'section': request.POST['section'],
            'image_id': request.POST['image_id'],
            'image_name': request.POST['image_name'],
        }
        if request.is_ajax() and request.method == 'POST':
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']
            token = request.session['token']['access_token']
            image_name = request.POST['image_name']
            bucket = request.POST['section']
            params = {}
            if request.POST['article_id'] != "" and request.POST['section'] == 'testimonial_img':

                params = {
                    'site_id': site_id,
                    'user_id': request.session['user_id'],
                    'article_id': request.POST['article_id'],
                    'upload_id': request.POST['image_id'],
                    'upload_type': 'article_image'
                }
                url = settings.API_URL + '/api-users/delete-article-file/'
            elif request.POST['article_id'] != "" and request.POST['section'] == 'testimonial_author_img' and request.POST['delete_for'] != 'testimonial_image':
                params = {
                    'site_id': site_id,
                    'user_id': request.session['user_id'],
                    'article_id': request.POST['article_id'],
                    'upload_id': request.POST['image_id'],
                    'upload_type': 'author_image'
                }
                print("-----------------------")
                url = settings.API_URL + '/api-users/delete-article-file/'
            elif request.POST['section'] == 'testimonial_author_img' and request.POST['delete_for'] == 'testimonial_image':
                params = {
                    'site_id': site_id,
                    'user_id': request.session['user_id'],
                    'testimonial_id': request.POST['article_id'],
                    'upload_id': request.POST['image_id'],
                }
                print("======================")
                url = settings.API_URL + '/api-users/testimonial-image-delete/'
            elif request.POST['section'] == 'banner_img':
                params = {
                    'site_id': site_id,
                    'user_id': request.session['user_id'],
                    'upload_id': request.POST['image_id'],
                    'upload_type': 'banner_image'
                }
                url = settings.API_URL + '/api-users/delete-settings-file/'

            elif request.POST['section'] == 'company_partner_img':
                params = {
                    'site_id': site_id,
                    'user_id': request.session['user_id'],
                    'upload_id': request.POST['image_id'],
                    'upload_type': 'footer_company'
                }
                url = settings.API_URL + '/api-users/delete-settings-file/'

            elif request.POST['section'] == 'website_logo':
                params = {
                    'site_id': site_id,
                    'user_id': request.session['user_id'],
                    'upload_id': request.POST['image_id'],
                    'upload_type': 'logo'
                }
                url = settings.API_URL + '/api-users/delete-settings-file/'
            elif request.POST['section'] == 'favicons':
                params = {
                    'site_id': site_id,
                    'user_id': request.session['user_id'],
                    'upload_id': request.POST['image_id'],
                    'upload_type': 'fabicon'
                }
                url = settings.API_URL + '/api-users/delete-settings-file/'
            elif request.POST['section'] == 'profile_image':
                if 'agent_id' in request.POST and request.POST['agent_id'] != "":
                    user_id = request.POST['agent_id']
                elif 'user_id' in request.POST and request.POST['user_id'] != "":
                    user_id = request.POST['user_id']
                elif 'loggedin_user_id' in request.POST and request.POST['loggedin_user_id'] != "" and 'request_from' in request.POST and request.POST['request_from'] != "":
                    user_id = request.session['user_id']
                else:
                    user_id = request.session['user_id']
                params = {
                    'site_id': site_id,
                    'user_id': user_id,
                    'upload_id': request.POST['image_id'],
                    'upload_type': 'profile_image'
                }

                url = settings.API_URL + '/api-users/delete-file/'
            elif request.POST['section'] == 'home_auctioin':
                params = {
                    'upload_id': request.POST['image_id'],
                    'upload_type': 'auction_image'
                }
                url = settings.API_URL + '/api-users/delete-file/'
            elif request.POST['section'] == 'company_logo':
                params = {
                    'upload_id': request.POST['image_id'],
                    'upload_type': 'company_logo'
                }
                url = settings.API_URL + '/api-users/delete-file/'
            elif request.POST['section'] == 'home_expertise':
                params = {
                    'upload_id': request.POST['image_id'],
                    'upload_type': 'expertise_image'
                }
                url = settings.API_URL + '/api-users/delete-file/'
            elif request.POST['section'] == 'property_image':
                params = {
                    'site_id': site_id,
                    'user_id': request.session['user_id'],
                    'property_id': request.POST['article_id'],
                    'upload_id': request.POST['image_id'],
                    'upload_type': 'property_image'
                }
                url = settings.API_URL + '/api-property/subdomain-property-document-delete/'
            elif request.POST['section'] == 'property_document':
                params = {
                    'site_id': site_id,
                    'user_id': request.session['user_id'],
                    'property_id': request.POST['article_id'],
                    'upload_id': request.POST['image_id'],
                    'upload_type': 'property_document'
                }
                url = settings.API_URL + '/api-property/subdomain-property-document-delete/'
            elif request.POST['section'] == 'property_video':
                params = {
                    'site_id': site_id,
                    'user_id': request.session['user_id'],
                    'property_id': request.POST['article_id'],
                    'upload_id': request.POST['image_id'],
                    'upload_type': 'property_video'
                }
                url = settings.API_URL + '/api-property/subdomain-property-document-delete/'
            elif request.POST['section'] == 'chat_document':
                params = {
                    "domain": site_id,
                    "user_id": request.session['user_id'],
                    "upload_type": "chat_document",
                    "upload_id": request.POST['image_id']
                }
                url = settings.API_URL + '/api-users/delete-file/'

            elif request.POST['section'] == 'portfolio':
                params = {
                    "domain_id": site_id,
                    "user_id": request.session['user_id'],
                    "upload_type": "portfolio",
                    "upload_id": request.POST['image_id']
                }
                url = settings.API_URL + '/api-property/portfolio-delete-image/'



            delete_data = call_api_post_method(params, url, token)
            user_list = []
            if 'error' in delete_data and delete_data['error'] == 0:
                try:
                    delete = delete_s3_file(bucket+'/'+image_name, settings.AWS_BUCKET_NAME)

                except Exception as exp:
                    print(exp)

                try:
                    if 'loggedin_user_id' in request.POST and request.POST['loggedin_user_id'] != "" and 'request_from' in request.POST and request.POST['request_from'] != "":
                        api_url = settings.API_URL + '/api-users/login-details/'
                        payload = {
                            'token': token,
                            'site_id': site_id
                        }
                        response = call_api_post_method(payload, api_url)
                        if "error" in response and response['error'] == 0:
                            response = response['data']
                            request.session.modified = True
                            request.session['profile_image'] = response['profile_image']
                except:
                    pass
                return_data['msg'] = 'Image deleted successfully'
                return_data['error'] = 0
                return_data['status'] = 200
                data = return_data
            else:
                return_data['msg'] = 'Some error occurs, please try again'
                return_data['error'] = 1
                return_data['status'] = 403
                data = return_data
        else:
            return_data['msg'] = 'Invalid request'
            return_data['error'] = 1
            return_data['status'] = 403
            data = return_data

        return JsonResponse(data)
    except Exception as exp:
        print(exp)
        data = {'status': 403, 'msg': 'invalid request.', 'user_list': []}
        return JsonResponse(data)

def save_website_setting(request):
    try:
        site_detail = subdomain_site_details(request)
        site_id = site_detail['site_detail']['site_id']
        token = request.session['token']['access_token']
        user_id = request.session['user_id']

        if request.is_ajax() and request.method == 'POST':
            banner_image_list = []
            if request.POST['banner_image_id'] != "":
                banner_image_list = request.POST['banner_image_id'].split(',')

            company_parter_logo_list = []
            if request.POST['footer_company_parter_logo_id'] != "":
                company_parter_logo_list = request.POST['footer_company_parter_logo_id'].split(',')

            # about_us_image_list = []
            # if request.POST['about_us_image_id'] != "":
            #     about_us_image_list = request.POST['about_us_image_id'].split(',')

            total_count = int(request.POST['total_section'])

            article_list = []
            # for x in range(total_count):
            #     article_params = {
            #         'title': request.POST['testimonial_title_' + str(x)],
            #         'author_name': request.POST['author_name_' + str(x)],
            #         'description': request.POST['testimonial_description_' + str(x)],
            #         'author_image': request.POST['testimonial_author_image_id_' + str(x)],
            #         'upload': request.POST['testimonial_image_id_' + str(x)]
            #     }
            #     article_list.append(article_params)

            domain_name_url = request.POST['primary_domain'].split('.')[0]
            if int(settings.HTTPS) == 1:
                domain_name = domain_name_url.replace('https://', '')
            else:
                domain_name = domain_name_url.replace('http://', '')
            auction_list = []
            for i in range(6):
                # auction_params = {
                #     'auction_name': request.POST['auction_type_' + str(i)],
                #     'upload': request.POST['auction_image_id_' + str(i)],
                #     'status': request.POST['selected_auction_type_' + str(i)],
                # }
                auction_params = {
                    'auction_name': request.POST['auction_type_' + str(i)],
                    'status': request.POST['selected_auction_type_' + str(i)],
                }
                auction_list.append(auction_params)

            expertise_list = []

            for j in range(4):
                expertise_params = {
                    'expertise_name': request.POST['expertise_type_' + str(j)],
                    #'upload': request.POST['expertise_image_id_' + str(j)],
                    'expertise_icon_type': request.POST['expertise_icon_type_' + str(j)],
                    'expertise_icon': request.POST['expertise_icon_' + str(j)],
                    'status': request.POST['selected_expertise_' + str(j)],
                }
                expertise_list.append(expertise_params)


            number_list = []
            for k in range(1,5):
                number_params = {
                    'title': request.POST['number_title_' + str(k)],
                    'value': request.POST['number_value_' + str(k)],
                }
                number_list.append(number_params)

            social_url_list = []

            for cnt in range(1, 6):
                if request.POST['account_type_' + str(cnt)] == '1':
                    account_type = 1
                    account_url = request.POST['facebook_url'] if request.POST['facebook_url'] != "" else ""
                    account_position = request.POST['facebook_position'] if request.POST['facebook_position'] != "" else k
                if request.POST['account_type_' + str(cnt)] == '2':
                    account_type = 2
                    account_url = request.POST['twitter_url'] if request.POST['twitter_url'] != "" else ""
                    account_position = request.POST['twitter_position'] if request.POST['twitter_position'] != "" else k
                if request.POST['account_type_' + str(cnt)] == '3':
                    account_type = 3
                    account_url = request.POST['youtube_url'] if request.POST['youtube_url'] != "" else ""
                    account_position = request.POST['youtube_position'] if request.POST['youtube_position'] != "" else k
                if request.POST['account_type_' + str(cnt)] == '4':
                    account_type = 4
                    account_url = request.POST['linkedin_url'] if request.POST['linkedin_url'] != "" else ""
                    account_position = request.POST['linkedin_position'] if request.POST['linkedin_position'] != "" else k
                if request.POST['account_type_' + str(cnt)] == '5':
                    account_type = 5
                    account_url = request.POST['instagram_url'] if request.POST['instagram_url'] != "" else ""
                    account_position = request.POST['instagram_position'] if request.POST['instagram_position'] != "" else k

                social_params = {
                    'account_type': account_type,
                    'url': account_url,
                    'position': account_position,
                }
                social_url_list.append(social_params)

            asset_type = []
            if 'asset_type' in request.POST and request.POST['asset_type']:
                asset_type = request.POST.getlist('asset_type')
            website_params = {
                'site_id': site_id,
                'user_id': user_id,
                'domain_name': domain_name,
                'domain_url': request.POST['primary_domain'],
                'favicon': request.POST['favicon_img_id'],
                'website_title': request.POST['website_title'],
                'website_logo': request.POST['website_logo_id'],
                'banner_headline': request.POST['banner_headline'],
                'website_name': request.POST['website_name'],
                'call_to_action': request.POST['call_action'],
                'extended_description': request.POST['extend_description'],
                # 'headline': request.POST['headline'],
                # 'best_downtown_locations': request.POST['best_downtown_location'],
                # 'exceptional_properties': request.POST['exp_property'],
                # 'satisfied_buyer_seller': request.POST['satified_buyer_seller'],
                # 'email_subscriber': request.POST['email_subscriber'],
                'about_title': request.POST['about_us_title'],
                'about_description': request.POST['about_us_description'],
                'banner_images': banner_image_list,
                'footer_images': company_parter_logo_list,
                'auctions': auction_list,
                'expertise': expertise_list,
                'dashboard_numbers': number_list,
                'social_account': social_url_list,
                'bot_setting': asset_type,
                'mls_type': request.POST['mls_type'],
                'api_key': request.POST['api_key'],
                'originating_system': request.POST['originating_system'],
                #'is_search_icon': request.POST['enable_search'] if 'enable_search' in request.POST and request.POST['enable_search'] else 0,

            }
            website_setting_url = settings.API_URL + '/api-users/subdomain-website-update/'
            website_setting_response = call_api_post_method(website_params, website_setting_url, token)
            if 'error' in website_setting_response and website_setting_response['error'] == 0:
                response = {
                    'error': 0,
                    'msg': 'Website setting updated successfully.'
                }
            else:
                response = {
                    'error': 1,
                    'msg': 'Some error occur, please try again.'
                }
        return JsonResponse(response)
    except Exception as exp:
        print(exp)
        data = {'status': 403, 'msg': 'invalid request.'}
        return JsonResponse(data)

def add_agent(request):
    is_permission = check_permission(request, 1)
    if not is_permission:
        http_host = request.META['HTTP_HOST']
        redirect_url = settings.URL_SCHEME + str(http_host)
        return HttpResponseRedirect(redirect_url)
    agent_id = request.GET.get('id', None)
    if not request.is_ajax() and agent_id is None and request.session['is_broker'] == False:
        http_host = request.META['HTTP_HOST']
        redirect_agent = settings.URL_SCHEME +str(http_host)+'/admin/agents/'
        return HttpResponseRedirect(redirect_agent)
    try:
        site_detail = subdomain_site_details(request)
        site_id = site_detail['site_detail']['site_id']
        token = request.session['token']['access_token']
        user_id = request.session['user_id']
        try:
            state_param = {}
            state_api_url = settings.API_URL + '/api-settings/get-state/'
            state_data = call_api_post_method(state_param, state_api_url, token)
            state_list = state_data['data']
        except:
            state_list = []

        try:
            permission_param = {}
            perm_api_url = settings.API_URL + '/api-settings/agent-permission-listing/'
            permission_data = call_api_post_method(permission_param, perm_api_url, token)
            permission_list = permission_data['data']
        except:
            permission_list = []

        try:
            agent_detail_param = {
                'site_id': site_id,
                'user_id': agent_id
            }
            agent_detail_url = settings.API_URL + '/api-users/subdomain-agent-detail/'
            agent_detail_data = call_api_post_method(agent_detail_param, agent_detail_url, token)
            if 'error' in agent_detail_data and agent_detail_data['error'] == 1 and agent_id:

                http_host = request.META['HTTP_HOST']
                redirect_url = settings.URL_SCHEME + str(http_host)
                return HttpResponseRedirect(redirect_url+'/admin/agents/')

            agent_details = agent_detail_data['data']
            agent_id = agent_details['id']

        except:
            agent_details = {}

        if request.is_ajax() and request.method == 'POST':
            permission_list = []
            total_permission = int(request.POST['total_permission'])
            for x in range(total_permission):
                permission_params = {
                    'permission_id': request.POST['permission_id_' + str(x)],
                    'is_permission': request.POST['is_permission_' + str(x)],
                }
                permission_list.append(permission_params)


            agent_params = {
                'site_id': site_id,
                'first_name': request.POST['agent_first_name'],
                'last_name': request.POST['agent_last_name'],
                'email': request.POST['user_email'],
                'phone_no': re.sub('\D', '', request.POST['usr_phone_no']),
                'company_name': request.POST['agent_company'],
                'address_first': request.POST['agent_address'],
                'postal_code': request.POST['zip_code'],
                'licence_no': request.POST['agent_license_no'],
                'state': request.POST['agent_state'],
                'status': request.POST['agent_status'],
                'profile_image': request.POST['agent_img_id'] if 'agent_img_id' in request.POST and request.POST['agent_img_id'] != "" else "",
                'permission': permission_list,
                'company_logo': request.POST['agent_logo_img_id'] if 'agent_logo_img_id' in request.POST and request.POST['agent_logo_img_id'] != "" else "",
            }

            if request.POST['agent_id']:
                agent_params['user_id'] = request.POST['agent_id']
                agent_url = settings.API_URL + '/api-users/subdomain-update-agent/'
            else:
                agent_url = settings.API_URL + '/api-users/create-agent/'




            agent_response = call_api_post_method(agent_params, agent_url, token)

            if 'error' in agent_response and agent_response['error'] == 0:
                response = {
                    'error': 0,
                    'msg': agent_response['msg']
                }
            else:
                response = {
                    'error': 1,
                    'msg': agent_response['msg']
                }
            return JsonResponse(response)

        already_checked_permission = [4, 7, 11, 6]
        context = {"active_menu": "agent", "state_list": state_list, 'agent_details': agent_details, 'permission_list': permission_list, 'checked_permission': already_checked_permission}
        return render(request, "admin/dashboard/agents/add-agent.html", context)
    except Exception as exp:
        print(exp)
        return HttpResponse("Issue in views")

def add_sub_admin(request):
    is_permission = check_permission(request, 1)
    if not is_permission:
        http_host = request.META['HTTP_HOST']
        redirect_url = settings.URL_SCHEME + str(http_host)
        return HttpResponseRedirect(redirect_url)
    agent_id = request.GET.get('id', None)
    if not request.is_ajax() and agent_id is None and request.session['is_broker'] == False:
        http_host = request.META['HTTP_HOST']
        redirect_agent = settings.URL_SCHEME +str(http_host)+'/admin/sub-admin/'
        return HttpResponseRedirect(redirect_agent)
    try:
        site_detail = subdomain_site_details(request)
        site_id = site_detail['site_detail']['site_id']
        token = request.session['token']['access_token']
        user_id = request.session['user_id']
        try:
            state_param = {}
            state_api_url = settings.API_URL + '/api-settings/get-state/'
            state_data = call_api_post_method(state_param, state_api_url, token)
            state_list = state_data['data']
        except:
            state_list = []

        try:
            permission_param = {}
            perm_api_url = settings.API_URL + '/api-settings/agent-permission-listing/'
            permission_data = call_api_post_method(permission_param, perm_api_url, token)
            permission_list = permission_data['data']
        except:
            permission_list = []

        try:
            agent_detail_param = {
                'site_id': site_id,
                'user_id': agent_id
            }
            agent_detail_url = settings.API_URL + '/api-users/sub-admin-detail/'
            agent_detail_data = call_api_post_method(agent_detail_param, agent_detail_url, token)
            if 'error' in agent_detail_data and agent_detail_data['error'] == 1 and agent_id:

                http_host = request.META['HTTP_HOST']
                redirect_url = settings.URL_SCHEME + str(http_host)
                return HttpResponseRedirect(redirect_url+'/admin/agents/')

            agent_details = agent_detail_data['data']
            agent_id = agent_details['id']

        except:
            agent_details = {}

        if request.is_ajax() and request.method == 'POST':
            permission_list = []
            total_permission = int(request.POST['total_permission'])
            for x in range(total_permission):
                permission_params = {
                    'permission_id': request.POST['permission_id_' + str(x)],
                    'is_permission': request.POST['is_permission_' + str(x)],
                }
                permission_list.append(permission_params)


            agent_params = {
                'site_id': site_id,
                'first_name': request.POST['agent_first_name'],
                'last_name': request.POST['agent_last_name'],
                'email': request.POST['user_email'],
                'phone_no': re.sub('\D', '', request.POST['usr_phone_no']),
                # 'company_name': request.POST['agent_company'],
                'address_first': request.POST['agent_address'],
                'postal_code': request.POST['zip_code'],
                # 'licence_no': request.POST['agent_license_no'],
                'state': request.POST['agent_state'],
                'status': request.POST['agent_status'],
                'profile_image': request.POST['agent_img_id'] if 'agent_img_id' in request.POST and request.POST['agent_img_id'] != "" else "",
                'permission': permission_list,
                # 'company_logo': request.POST['agent_logo_img_id'] if 'agent_logo_img_id' in request.POST and request.POST['agent_logo_img_id'] != "" else "",
            }

            if request.POST['agent_id']:
                agent_params['user_id'] = request.POST['agent_id']
                agent_url = settings.API_URL + '/api-users/sub-admin-update/'
            else:
                # agent_url = settings.API_URL + '/api-users/create-agent/'
                agent_url = settings.API_URL + '/api-users/create-sub-admin/'

            agent_response = call_api_post_method(agent_params, agent_url, token)
            print(agent_response)

            if 'error' in agent_response and agent_response['error'] == 0:
                response = {
                    'error': 0,
                    'msg': agent_response['msg']
                }
            else:
                response = {
                    'error': 1,
                    'msg': agent_response['msg']
                }
            return JsonResponse(response)

        already_checked_permission = [4, 7, 11, 6]
        context = {"active_menu": "agent", "state_list": state_list, 'agent_details': agent_details, 'permission_list': permission_list, 'checked_permission': already_checked_permission}
        return render(request, "admin/dashboard/sub-admin/add-sub-admin.html", context)
    except Exception as exp:
        print(exp)
        return HttpResponse("Issue in views")        

@csrf_exempt
def delete_agent(request):
    try:
        if request.is_ajax() and request.method == 'POST':
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']
            params = {
                'site_id': site_id,
                'user_id': request.POST['user_id'],
                'deleted_by': request.session['user_id']
            }
            agent_search = ''
            if 'search' in request.POST and request.POST['search']:
                agent_search = request.POST['search']
            page = 1
            if 'page' in request.POST and request.POST['page'] != "":
                page = request.POST['page']

            page_size = 10
            if 'perpage' in request.POST and request.POST['perpage']:
                page_size = request.POST['perpage']

            if 'status' in request.POST and request.POST['status'] and request.POST['status'] and request.POST[
                'status'].lower() == 'active':
                status = [1]
            elif 'status' in request.POST and request.POST['status'] and request.POST['status'] and request.POST[
                'status'].lower() == 'inactive':
                status = [2]
            else:
                status = [2, 1]

            token = request.session['token']['access_token']
            user_id = request.session['user_id']
            url = settings.API_URL + '/api-users/subdomain-delete-agent/'
            data = call_api_post_method(params, url, token)
            user_list = []
            if 'error' in data and data['error'] == 0:


                list_param = {
                    'site_id': site_id,
                    "page": page,
                    "user_id": user_id,
                    "page_size": page_size,
                    "status": status,
                    "search": agent_search
                }
                sno = (int(page) - 1) * int(page_size) + 1
                list_url = settings.API_URL + '/api-users/subdomain-agent-listing/'
                list_data = call_api_post_method(list_param, list_url, token)

                if 'error' in list_data and list_data['error'] == 0:
                    agent_list = list_data['data']['data']
                    total = list_data['data']['total']
                else:
                    agent_list = []
                    total = 0
                context = {'agent_list': agent_list, 'total': total, "aws_url": settings.AWS_URL, 'sno': sno}

                agent_listing_path = 'admin/dashboard/agents/agent-listing-content.html'
                agent_listing_template = get_template(agent_listing_path)
                agent_listing_html = agent_listing_template.render(context)
                # ---------------Pagination--------
                pagination_path = 'admin/dashboard/agents/pagination.html'
                pagination_template = get_template(pagination_path)
                total_page = math.ceil(int(total) / int(page_size))
                pagination_html = ''
                if total_page > 1:
                    pagination_data = {"no_page": int(total_page), "total_page": range(total_page),
                                       "current_page": int(page),
                                       "pagination_id": "agent_listing_pagination_list"}
                    pagination_html = pagination_template.render(pagination_data)

                data = {'agent_listing_html': agent_listing_html, 'status': 200, 'msg': 'Deleted successfully', 'error': 0, 'total': total,
                        "pagination_html": pagination_html, 'pagination_id': 'agent_listing_pagination_list'}

            else:
                data = {'error': 1, 'status': 403, 'msg': 'Server error, Please try again'}
        else:
            data = {'error': 1, 'status': 403, 'msg': 'Forbidden'}

        return JsonResponse(data)
    except Exception as exp:
        print(exp)
        data = {'status': 403, 'msg': 'invalid request.', 'user_list': []}
        return JsonResponse(data)

@csrf_exempt
def get_agent_details(request):
    try:
        if request.is_ajax() and request.method == 'POST':
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']
            params = {
                'user_id': request.POST['user_id'],
                'site_id': site_id
            }
            token = request.session['token']['access_token']
            url = settings.API_URL + '/api-users/make-agent-detail/'
            data = call_api_post_method(params, url, token)

        else:
            data = {'status': 403, 'msg': 'Forbidden'}

        return JsonResponse(data)
    except Exception as exp:
        print(exp)
        data = {'status': 403, 'msg': 'invalid request.'}
        return JsonResponse(data)

@csrf_exempt
def upgrade_user_to_agent(request):
    try:
        if request.is_ajax() and request.method == 'POST':
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']
            token = request.session['token']['access_token']
            permission_list = []
            total_permission = int(request.POST['total_permission'])
            for x in range(total_permission):
                permission_params = {
                    'permission_id': request.POST['permission_id_' + str(x)],
                    'is_permission': request.POST['is_permission_' + str(x)],
                }
                permission_list.append(permission_params)
            agent_params = {
                'site_id': site_id,
                'user_id': request.POST['agent_id'],
                'first_name': request.POST['agent_first_name'],
                'last_name': request.POST['agent_last_name'],
                'email': request.POST['user_email'],
                'phone_no': re.sub('\D', '', request.POST['usr_phone_no']),
                'company_name': request.POST['agent_company'],
                'address_first': request.POST['agent_address'],
                'postal_code': request.POST['zip_code'],
                'licence_no': request.POST['agent_license_no'],
                'state': request.POST['agent_state'],
                'permission': permission_list,
            }

            agent_url = settings.API_URL + '/api-users/make-agent/'

            response = call_api_post_method(agent_params, agent_url, token)
            if 'error' in response and response['error'] == 0:
                page = 1
                page_size = 10
                if 'page_size' in request.POST and request.POST['page_size'] != "":
                    page_size = request.POST['page_size']

                if 'status' in request.POST and request.POST['status'] and request.POST['status'].lower() == 'active':
                    status = [1]
                elif 'status' in request.POST and request.POST['status'] and request.POST['status'] and request.POST[
                    'status'].lower() == 'inactive':
                    status = [2]
                else:
                    status = [2, 1]

                user_search = ''
                if 'user_search' in request.POST and request.POST['user_search'] != "":
                    user_search = request.POST['user_search']
                list_param = {
                    'site_id': site_id,
                    "page": page,
                    "page_size": page_size,
                    "status": status,
                    "search": user_search,
                }
                sno = (int(page) - 1) * int(page_size) + 1
                list_url = settings.API_URL + '/api-users/subdomain-user-listing/'
                list_data = call_api_post_method(list_param, list_url, token)
                if 'error' in list_data and list_data['error'] == 0:
                    user_listing = list_data['data']['data']
                    total = list_data['data']['total']
                else:
                    user_listing = []
                    total = 0
                context = {'user_list': user_listing, 'total': total, "aws_url": settings.AWS_URL, 'sno': sno}

                user_listing_path = 'admin/dashboard/users/user-list-content.html'
                user_listing_template = get_template(user_listing_path)
                user_listing_html = user_listing_template.render(context)
                # ---------------Pagination--------
                pagination_path = 'admin/dashboard/users/pagination.html'
                pagination_template = get_template(pagination_path)
                total_page = math.ceil(int(total) / int(page_size))
                pagination_html = ''
                if total_page > 1:
                    pagination_data = {"no_page": int(total_page), "total_page": range(total_page),
                                       "current_page": int(page),
                                       "pagination_id": "user_listing_pagination_list"}
                    pagination_html = pagination_template.render(pagination_data)

                data = {'user_listing_html': user_listing_html, 'status': 200, 'error': 0, 'total': total,
                        "pagination_html": pagination_html, 'pagination_id': 'user_listing_pagination_list', 'msg': response['msg']}
            else:
                data = {'status': 400, 'error': 1, 'data': response['msg']}

        else:
            data = {'status': 403, 'msg': 'Forbidden'}

        return JsonResponse(data)
    except Exception as exp:
        print(exp)
        data = {'status': 403, 'msg': 'invalid request.'}
        return JsonResponse(data)

def change_user_password(request):
    try:
        if request.is_ajax() and request.method == 'POST':
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']
            params = {
                'site_id': site_id,
                'user_id': request.POST['change_pass_user_id'],
                'password': request.POST['user_current_password'],
                'new_password': request.POST['user_new_password'],
                'admin_id': request.session['user_id']
            }
            token = request.session['token']['access_token']
            url = settings.API_URL + '/api-users/subdomain-user-change-password/'
            data = call_api_post_method(params, url, token)

        else:
            data = {'status': 403, 'msg': 'Forbidden'}

        return JsonResponse(data)
    except Exception as exp:
        print(exp)
        data = {'status': 403, 'msg': 'invalid request.'}
        return JsonResponse(data)

def reset_user_password(request):
    try:
        if request.is_ajax() and request.method == 'POST':
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']
            params = {
                'site_id': site_id,
                'user_id': request.POST['reset_user_id'],
                'password': request.POST['user_pass'],
                'admin_id': request.session['user_id']
            }
            token = request.session['token']['access_token']
            url = settings.API_URL + '/api-users/subdomain-user-reset-password/'
            data = call_api_post_method(params, url, token)
        else:
            data = {'status': 403, 'msg': 'Forbidden'}

        return JsonResponse(data)
    except Exception as exp:
        print(exp)
        data = {'status': 403, 'msg': 'invalid request.'}
        return JsonResponse(data)

def add_article(request):
    is_permission = check_permission(request, 11)
    if not is_permission:
        http_host = request.META['HTTP_HOST']
        redirect_url = settings.URL_SCHEME + str(http_host)
        return HttpResponseRedirect(redirect_url)
    article_id = request.GET.get('id', None)
    try:
        site_detail = subdomain_site_details(request)
        site_id = site_detail['site_detail']['site_id']
        token = request.session['token']['access_token']
        user_id = request.session['user_id']

        try:
            article_detail_param = {
                'site_id': site_id,
                'article_id': article_id
            }
            article_detail_url = settings.API_URL + '/api-users/article-detail/'
            article_detail_data = call_api_post_method(article_detail_param, article_detail_url, token)
            if 'error' in article_detail_data and article_detail_data['error'] == 1 and agent_id:

                http_host = request.META['HTTP_HOST']
                redirect_url = settings.URL_SCHEME + str(http_host)
                return HttpResponseRedirect(redirect_url+'/admin/articles/')
            article_details = article_detail_data['data']
            article_id = article_details['id']

        except:
            article_details = {}

        
        try:
            # api_url = settings.API_URL + '/api-property/asset-listing/'
            api_url = settings.API_URL + '/api-blog/blog-category/'
            asset_list = call_api_post_method({}, api_url, token)
            if 'error' in asset_list and asset_list['error'] == 1 and agent_id:

                http_host = request.META['HTTP_HOST']
                redirect_url = settings.URL_SCHEME + str(http_host)
                return HttpResponseRedirect(redirect_url+'/admin/articles/')
            asset_list = asset_list['data']
        except:
            asset_list = {}

        if request.is_ajax() and request.method == 'POST':

            article_params = {
                'site_id': site_id,
                'user_id': user_id,
                'title': request.POST['article_title'],
                'description': request.POST['article_description'],
                'author_name': request.POST['author_name'],
                'author_image': request.POST['article_author_image_id'],
                'upload': request.POST['article_image_id'],
                'publish_date': request.POST['publish_date'] if 'publish_date' in request.POST and request.POST['publish_date'] != "" else None,
                'status': request.POST['article_status'],
                'asset': request.POST['asset']
            }
            if request.POST['article_id']:
                article_params['article_id'] = request.POST['article_id']
            article_url = settings.API_URL + '/api-users/add-article/'


            article_response = call_api_post_method(article_params, article_url, token)
            if 'error' in article_response and article_response['error'] == 0:
                response = {
                    'error': 0,
                    'msg': 'Blog saved successfully'
                }
            else:
                response = {
                    'error': 1,
                    'msg': article_response['msg']
                }
            return JsonResponse(response)

        context = {"active_menu": "articles", "article_details": article_details, "asset_list": asset_list}
        return render(request, "admin/dashboard/blogs/add-article.html", context)
    except Exception as exp:
        print(exp)
        return HttpResponse("Issue in views")

@csrf_exempt
def articles(request):
    is_permission = check_permission(request, 11)
    if not is_permission:
        http_host = request.META['HTTP_HOST']
        redirect_url = settings.URL_SCHEME + str(http_host)
        return HttpResponseRedirect(redirect_url)
    page_size = request.GET.get('perpage', 10)
    status = request.GET.get('status', None)
    search = request.GET.get('search', None)
    page = request.GET.get('page', 1)
    try:
        site_detail = subdomain_site_details(request)
        site_id = site_detail['site_detail']['site_id']
        token = request.session['token']['access_token']
        user_id = request.session['user_id']
        if request.is_ajax() and request.method == 'POST':
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']
            agent_search = ''
            if request.POST['search']:
                agent_search = request.POST['search']

            if request.POST['perpage']:
                page_size = request.POST['perpage']
            else:
                page_size = 10

            if request.POST['page']:
                page = request.POST['page']
            else:
                page: 1    
            if request.POST['status'] and request.POST['status'].lower() == 'active':
                status = [1]
            elif request.POST['status'] and request.POST['status'].lower() == 'inactive':
                status = [2]
            else:
                status = [2, 1]

            list_param = {
                'site_id': site_id,
                "page": page,
                "page_size": page_size,
                "status": status,
                "search": agent_search
            }
            list_url = settings.API_URL + '/api-users/article-listing/'
            list_data = call_api_post_method(list_param, list_url, token)
            if 'error' in list_data and list_data['error'] == 0:
                article_list = list_data['data']['data']
                total = list_data['data']['total']
                # ---------------Pagination--------
                pagination_path = 'admin/dashboard/blogs/pagination.html'
                pagination_template = get_template(pagination_path)
                total_page = math.ceil(int(total) / int(page_size))
                pagination_html = ''
                if total_page > 1:
                    pagination_data = {"no_page": int(total_page), "total_page": range(total_page), "current_page": int(page),"pagination_id": "article_listing_pagination_list"}
                    pagination_html = pagination_template.render(pagination_data)
                data = {'article_list': article_list, 'status': 200, 'msg': '', 'error': 0, "pagination_html": pagination_html, "pagination_id": "article_listing_pagination_list", "total": total}
            else:
                data = {'status': 403, 'msg': 'Server error, Please try again', 'article_list': [], 'error': 0}
            return JsonResponse(data)

        try:

            article_list_param = {
                'site_id': site_id,
                "status": [1],
                'page': 1,
                'page_size': 10
            }
            sno = (int(page) - 1) * int(page_size) + 1
            article_list_url = settings.API_URL + '/api-users/article-listing/'
            article_list_data = call_api_post_method(article_list_param, article_list_url, token)
            if 'error' in article_list_data and article_list_data['error'] == 0:
                article_list = article_list_data['data']['data']
                total = article_list_data['data']['total']
            else:
                article_list = []
                total = 0
            # ---------------Pagination--------
            pagination_html = ''
            pagination_path = 'admin/dashboard/blogs/pagination.html'
            pagination_template = get_template(pagination_path)
            total_page = math.ceil(total / page_size)

            if total_page > 1:
                pagination_data = {"no_page": total_page, "total_page": range(total_page), "current_page": 1,
                                   "pagination_id": "article_listing_pagination_list"}
                pagination_html = pagination_template.render(pagination_data)
    
        except:
            article_list = []

        context = {"active_menu": "pages", "active_submenu": "articles", "article_list": article_list, "total": total, "pagination_html": pagination_html, "pagination_id": "article_listing_pagination_list", "sno": sno}
        return render(request, "admin/dashboard/blogs/view-article.html", context)
    except Exception as exp:
        print(exp)
        return HttpResponse("Issue in views")

@csrf_exempt
def edit_user_details(request):
    try:
        if request.is_ajax() and request.method == 'POST':
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']
            params = {
                'user_id': request.POST['user_id'],
                'site_id': site_id
            }
            token = request.session['token']['access_token']
            url = settings.API_URL + '/api-users/subdomain-user-detail/'
            data = call_api_post_method(params, url, token)

        else:
            data = {'status': 403, 'msg': 'Forbidden'}

        return JsonResponse(data)
    except Exception as exp:
        print(exp)
        data = {'status': 403, 'msg': 'invalid request.'}
        return JsonResponse(data)

def save_user_details(request):
    try:
        if request.is_ajax() and request.method == 'POST':
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']
            token = request.session['token']['access_token']

            user_params = {
                'site_id': site_id,
                'user_id': request.POST['update_user_id'],
                'first_name': request.POST['user_first_name'],
                'last_name': request.POST['user_last_name'],
                'email': request.POST['usr_email'],
                'phone_no': re.sub('\D', '', request.POST['user_phone_no']),
                'profile_image': request.POST['user_img_id']
            }

            user_url = settings.API_URL + '/api-users/subdomain-user-update'

            data = call_api_post_method(user_params, user_url, token)
        else:
            data = {'status': 403, 'msg': 'Forbidden'}

        return JsonResponse(data)
    except Exception as exp:
        print(exp)
        data = {'status': 403, 'msg': 'invalid request.'}
        return JsonResponse(data)

@csrf_exempt
def delete_article(request):
    try:
        if request.is_ajax() and request.method == 'POST':
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']
            params = {
                'site_id': site_id,
                'user_id': request.POST['user_id']
            }
            search = ''
            if request.POST['search']:
                search = request.POST['search']

            if request.POST['perpage']:
                page_size =  request.POST['perpage']
            else:
                page_size = 10

            if request.POST['status'] and request.POST['status'].lower() == 'active':
                status = [1]
            elif request.POST['status'] and request.POST['status'].lower() == 'inactive':
                status = [2]
            else:
                status = [2, 1]

            token = request.session['token']['access_token']
            url = settings.API_URL + '/api-users/subdomain-delete-user/'
            data = call_api_post_method(params, url, token)
            user_list = []
            if 'error' in data and data['error'] == 0:
                article_list_param = {
                    'site_id': site_id,
                    "page": 1,
                    "page_size": page_size,
                    "status": status,
                    "search": search
                }
                article_list_url = settings.API_URL + '/api-users/article-listing/'
                article_list_data = call_api_post_method(article_list_param, article_list_url, token)

                article_list = article_list_data['data']['data']
                data = {'article_list': article_list, 'data': data}
            else:
                data = {'status': 403, 'msg': 'Server error, Please try again', 'article_list': []}
        else:
            data = {'status': 403, 'msg': 'Forbidden', 'user_list': []}

        return JsonResponse(data)
    except Exception as exp:
        print(exp)
        data = {'status': 403, 'msg': 'invalid request.', 'user_list': []}
        return JsonResponse(data)

@csrf_exempt
def bidder_registration(request):
    try:
        try:
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']

        except Exception as exp:
            print(exp)
            site_id = ""

        user_id = None
        token = None
        if 'user_id' in request.session and request.session['user_id']:
            token = request.session['token']['access_token']
            user_id = request.session['user_id']

        page_size = 10
        page = 1
        if request.is_ajax() and request.method == 'POST':

            search = ''
            if 'search' in request.POST and request.POST['search']:
                search = request.POST['search']


            page = 1
            if 'page' in request.POST and request.POST['page'] != "":
                page = request.POST['page']

            if 'page_size' in request.POST and request.POST['page_size'] != "":
                page_size = request.POST['page_size']

            category_id = ''
            if 'category_id' in request.POST and request.POST['category_id'] != "":
                category_id = request.POST['category_id']
            asset_type = ''
            asset_type_name = 'all-list'
            if 'asset_type' in request.POST and request.POST['asset_type'] != "":
                asset_type = request.POST['asset_type']
                if int(asset_type) == 3:
                    asset_type_name = 'residential-list'
                elif int(asset_type) == 2:
                    asset_type_name = 'commercial-list'
                elif int(asset_type) == 1:
                    asset_type_name = 'land-list'


            filter_bidder_status = ''
            if 'filter_bidder_status' in request.POST and request.POST['filter_bidder_status'] != "":
                filter_bidder_status = request.POST['filter_bidder_status']


            params = {
                'site_id': site_id,
                'user_id': user_id,
                'page_size': page_size,
                'search': search,
                'page': page,
                'asset_type': asset_type,
                'filter_data': filter_bidder_status,
            }

            api_url = settings.API_URL + '/api-bid/subdomain-bid-registration-listing/'
            bidder_data = call_api_post_method(params, api_url, token=token)

            if 'error' in bidder_data and bidder_data['error'] == 0:
                bidder_list = bidder_data['data']['data']
                total = bidder_data['data']['total']
            else:
                bidder_list = []
                total = 0
            sno = (int(page) - 1) * int(page_size) + 1
            context = {'bidder_list': bidder_list, 'total': total, "aws_url": settings.AWS_URL, 'sno': sno}

            bidder_listing_path = 'admin/dashboard/bidder/bidder-listing.html'
            bidder_listing_template = get_template(bidder_listing_path)
            bidder_listing_html = bidder_listing_template.render(context)
            # ---------------Pagination--------
            pagination_path = 'admin/dashboard/bidder/pagination.html'
            pagination_template = get_template(pagination_path)
            total_page = math.ceil(int(total) / int(page_size))
            pagination_html = ''
            if total_page > 1:
                pagination_data = {"no_page": int(total_page), "total_page": range(total_page), "current_page": int(page),"pagination_id": "bid_listing_pagination_list", "asset_type": asset_type, "asset_type_name": asset_type_name}
                pagination_html = pagination_template.render(pagination_data)

            data = {'bidder_listing_html': bidder_listing_html, 'status': 200, 'msg': '', 'error': 0, 'total': total,
                    "pagination_html": pagination_html, 'pagination_id': 'bid_listing_pagination_list', 'asset_type': asset_type, 'asset_type_name': asset_type_name}
            return JsonResponse(data)
        else:
            params = {
                'site_id': site_id,
                'user_id': user_id,
                'page_size': page_size,
                'page': 1,
                'asset_type': '',
                'filter_data': 1,
                'search': '',
            }
            sno = (int(page) - 1) * int(page_size) + 1
            api_url = settings.API_URL + '/api-bid/subdomain-bid-registration-listing/'
            bidder_data = call_api_post_method(params, api_url, token=token)
            if 'error' in bidder_data and bidder_data['error'] == 0:
                bidder_list = bidder_data['data']['data']
                total = bidder_data['data']['total']
            else:
                bidder_list = []
                total = 0
            # ---------------Pagination--------
            pagination_html = ''
            pagination_path = 'admin/dashboard/bidder/pagination.html'
            pagination_template = get_template(pagination_path)
            total_page = math.ceil(total / page_size)

            if total_page > 1:
                pagination_data = {"no_page": total_page, "total_page": range(total_page), "current_page": 1,
                                   "pagination_id": "bid_listing_pagination_list"}
                pagination_html = pagination_template.render(pagination_data)
            context = {'bidder_list': bidder_list, 'total': total, "pagination_html": pagination_html,
                       "pagination_id": "bid_listing_pagination_list", "active_menu": "regbidder", "sno": sno}


            return render(request, 'admin/dashboard/bidder/bidder-registration-list.html', context)
    except Exception as exp:
        print(exp)
        return HttpResponse("Issue in views")

@csrf_exempt
def user_search_suggestion(request):
    try:
        if request.is_ajax() and request.method == 'POST':
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']
            token = request.session['token']['access_token']

            user_params = {
                'site_id': site_id,
                'search': request.POST['search']
            }

            user_url = settings.API_URL + '/api-users/subdomain-user-suggestion/'

            suggestion_data = call_api_post_method(user_params, user_url, token)
            if 'error' in suggestion_data and suggestion_data['error'] == 0:
                data = {'status': 200, 'suggestion_list': suggestion_data['data'], 'error': 0}
            else:
                data = {'status': 403, 'suggestion_list': [], 'error': 1}
        else:
            data = {'status': 403, 'suggestion_list': [], 'error': 1}

        return JsonResponse(data)
    except Exception as exp:
        print(exp)
        data = {'status': 403, 'suggestion_list': [], 'error': 1}
        return JsonResponse(data)

@csrf_exempt
def agent_search_suggestion(request):
    try:
        if request.is_ajax() and request.method == 'POST':
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']
            token = request.session['token']['access_token']

            user_params = {
                'site_id': site_id,
                'search': request.POST['search']
            }

            user_url = settings.API_URL + '/api-users/subdomain-agent-suggestion/'

            suggestion_data = call_api_post_method(user_params, user_url, token)
            if 'error' in suggestion_data and suggestion_data['error'] == 0:
                data = {'status': 200, 'suggestion_list': suggestion_data['data'], 'error': 0}
            else:
                data = {'status': 403, 'suggestion_list': [], 'error': 1}
        else:
            data = {'status': 403, 'suggestion_list': [], 'error': 1}

        return JsonResponse(data)
    except Exception as exp:
        print(exp)
        data = {'status': 403, 'suggestion_list': [], 'error': 1}
        return JsonResponse(data)

@csrf_exempt
def sub_admin_search_suggestion(request):
    try:
        if request.is_ajax() and request.method == 'POST':
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']
            token = request.session['token']['access_token']

            user_params = {
                'site_id': site_id,
                'search': request.POST['search']
            }

            user_url = settings.API_URL + '/api-users/sub-admin-suggestion/'

            suggestion_data = call_api_post_method(user_params, user_url, token)
            if 'error' in suggestion_data and suggestion_data['error'] == 0:
                data = {'status': 200, 'suggestion_list': suggestion_data['data'], 'error': 0}
            else:
                data = {'status': 403, 'suggestion_list': [], 'error': 1}
        else:
            data = {'status': 403, 'suggestion_list': [], 'error': 1}

        return JsonResponse(data)
    except Exception as exp:
        print(exp)
        data = {'status': 403, 'suggestion_list': [], 'error': 1}
        return JsonResponse(data)        

def listing_property_info(request):
    property_id = request.GET.get('property_id', None)
    is_permission = check_permission(request, 6)
    if not is_permission:
        http_host = request.META['HTTP_HOST']
        redirect_url = settings.URL_SCHEME + str(http_host)
        return HttpResponseRedirect(redirect_url)
    try:
        token = request.session['token']['access_token']
        site_detail = subdomain_site_details(request)
        site_id = site_detail['site_detail']['site_id']
        try:
            asset_listing_params = {}

            asset_listing_url = settings.API_URL + '/api-property/asset-listing/'

            asset_listing_data = call_api_post_method(asset_listing_params, asset_listing_url, token)
            asset_listing = asset_listing_data['data']
        except:
            asset_listing = []

        # try:
        #     state_param = {}
        #     state_api_url = settings.API_URL + '/api-settings/get-state/'
        #     state_data = call_api_post_method(state_param, state_api_url, token)
        #     state_list = state_data['data']
        # except:
        #     state_list = []


        try:
            auction_type_param = {}
            auction_type_url = settings.API_URL + '/api-settings/subdomain-auction-type/'
            auction_type_data = call_api_post_method(auction_type_param, auction_type_url, token)
            auction_type_list = auction_type_data['data']
        except:
            auction_type_list = []

        try:
            auction_status_param = {'object_id': 12}
            auction_status_url = settings.API_URL + '/api-settings/lookup-status-listing/'
            auction_status_data = call_api_post_method(auction_status_param, auction_status_url, token)
            auction_status_list = auction_status_data['data']

        except Exception as exp:
            print(exp)
            auction_status_list = []

        try:
            closing_status_param = {'object_id': 14}
            closing_status_url = settings.API_URL + '/api-settings/lookup-status-listing/'
            closing_status_data = call_api_post_method(closing_status_param, closing_status_url, token)
            closing_status_list = closing_status_data['data']

        except Exception as exp:
            print(exp)
            closing_status_list = []

        try:
            timezone_param = {}
            timezone_url = settings.API_URL + '/api-settings/timezone-listing/'
            timezone_data = call_api_post_method(timezone_param, timezone_url, token)
            timezone_list = timezone_data['data']
        except:
            timezone_list = []
        status_list = []
        if property_id is not None:
            property_detail_param = {"site_id": site_id, "property_id": property_id, "step_id": 1}
            property_detail_url = settings.API_URL + '/api-property/subdomain-property-detail/'
            property_detail_data = call_api_post_method(property_detail_param, property_detail_url, token)

            try:
                property_details = property_detail_data['data']
                print(property_details)
                country_id = property_details['country']
                asset_id = property_details['asset_id']

                property_details['selected_property_subtype'] = [int(item['feature_id']) for item in
                                                                 property_details['property_subtype'] if
                                                                 'property_subtype' in property_details and len(
                                                                     property_details['property_subtype']) > 0]
                property_details['selected_terms_accepted'] = [int(item['feature_id']) for item in
                                                               property_details['terms_accepted'] if
                                                               'terms_accepted' in property_details and len(
                                                                   property_details['terms_accepted']) > 0]
                property_details['selected_occupied_by'] = [int(item['feature_id']) for item in
                                                               property_details['occupied_by'] if
                                                               'occupied_by' in property_details and len(
                                                                   property_details['occupied_by']) > 0]
                property_details['selected_ownership'] = [int(item['feature_id']) for item in
                                                            property_details['ownership'] if
                                                            'ownership' in property_details and len(
                                                                property_details['ownership']) > 0]
                property_details['selected_possession'] = [int(item['feature_id']) for item in
                                                          property_details['possession'] if
                                                          'possession' in property_details and len(
                                                              property_details['possession']) > 0]
                property_details['selected_style'] = [int(item['feature_id']) for item in
                                                           property_details['style'] if
                                                           'style' in property_details and len(
                                                               property_details['style']) > 0]
                property_details['selected_cooling'] = [int(item['feature_id']) for item in
                                                      property_details['cooling'] if
                                                      'cooling' in property_details and len(
                                                          property_details['cooling']) > 0]
                property_details['selected_stories'] = [int(item['feature_id']) for item in
                                                        property_details['stories'] if
                                                        'stories' in property_details and len(
                                                            property_details['stories']) > 0]
                property_details['selected_heating'] = [int(item['feature_id']) for item in
                                                        property_details['heating'] if
                                                        'heating' in property_details and len(
                                                            property_details['heating']) > 0]
                property_details['selected_electric'] = [int(item['feature_id']) for item in
                                                        property_details['electric'] if
                                                        'electric' in property_details and len(
                                                            property_details['electric']) > 0]
                property_details['selected_gas'] = [int(item['feature_id']) for item in
                                                         property_details['gas'] if
                                                         'gas' in property_details and len(
                                                             property_details['gas']) > 0]
                property_details['selected_recent_updates'] = [int(item['feature_id']) for item in
                                                    property_details['recent_updates'] if
                                                    'recent_updates' in property_details and len(
                                                        property_details['recent_updates']) > 0]
                property_details['selected_water'] = [int(item['feature_id']) for item in
                                                               property_details['water'] if
                                                               'water' in property_details and len(
                                                                   property_details['water']) > 0]
                property_details['selected_security_features'] = [int(item['feature_id']) for item in
                                                      property_details['security_features'] if
                                                      'security_features' in property_details and len(
                                                          property_details['security_features']) > 0]
                property_details['selected_sewer'] = [int(item['feature_id']) for item in
                                                                  property_details['sewer'] if
                                                                  'sewer' in property_details and len(
                                                                      property_details['sewer']) > 0]
                property_details['selected_tax_exemptions'] = [int(item['feature_id']) for item in
                                                      property_details['tax_exemptions'] if
                                                      'tax_exemptions' in property_details and len(
                                                          property_details['tax_exemptions']) > 0]
                property_details['selected_zoning'] = [int(item['feature_id']) for item in
                                                               property_details['zoning'] if
                                                               'zoning' in property_details and len(
                                                                   property_details['zoning']) > 0]
                property_details['selected_hoa_amenities'] = [int(item['feature_id']) for item in
                                                       property_details['hoa_amenities'] if
                                                       'hoa_amenities' in property_details and len(
                                                           property_details['hoa_amenities']) > 0]
                property_details['selected_kitchen_features'] = [int(item['feature_id']) for item in
                                                              property_details['kitchen_features'] if
                                                              'kitchen_features' in property_details and len(
                                                                  property_details['kitchen_features']) > 0]
                property_details['selected_appliances'] = [int(item['feature_id']) for item in
                                                                 property_details['appliances'] if
                                                                 'appliances' in property_details and len(
                                                                     property_details['appliances']) > 0]
                property_details['selected_flooring'] = [int(item['feature_id']) for item in
                                                           property_details['flooring'] if
                                                           'flooring' in property_details and len(
                                                               property_details['flooring']) > 0]
                property_details['selected_windows'] = [int(item['feature_id']) for item in
                                                         property_details['windows'] if
                                                         'windows' in property_details and len(
                                                             property_details['windows']) > 0]
                property_details['selected_bedroom_features'] = [int(item['feature_id']) for item in
                                                         property_details['bedroom_features'] if
                                                         'bedroom_features' in property_details and len(
                                                             property_details['bedroom_features']) > 0]
                property_details['selected_other_rooms'] = [int(item['feature_id']) for item in
                                                                 property_details['other_rooms'] if
                                                                 'other_rooms' in property_details and len(
                                                                     property_details['other_rooms']) > 0]
                property_details['selected_bathroom_features'] = [int(item['feature_id']) for item in
                                                            property_details['bathroom_features'] if
                                                            'bathroom_features' in property_details and len(
                                                                property_details['bathroom_features']) > 0]
                property_details['selected_other_features'] = [int(item['feature_id']) for item in
                                                                  property_details['other_features'] if
                                                                  'other_features' in property_details and len(
                                                                      property_details['other_features']) > 0]
                property_details['selected_master_bedroom_features'] = [int(item['feature_id']) for item in
                                                               property_details['master_bedroom_features'] if
                                                               'master_bedroom_features' in property_details and len(
                                                                   property_details['master_bedroom_features']) > 0]
                property_details['selected_fireplace_type'] = [int(item['feature_id']) for item in
                                                                        property_details['fireplace_type'] if
                                                                        'fireplace_type' in property_details and len(
                                                                            property_details[
                                                                                'fireplace_type']) > 0]
                property_details['selected_basement_features'] = [int(item['feature_id']) for item in
                                                               property_details['basement_features'] if
                                                               'basement_features' in property_details and len(
                                                                   property_details[
                                                                       'basement_features']) > 0]
                property_details['selected_handicap_amenities'] = [int(item['feature_id']) for item in
                                                                  property_details['handicap_amenities'] if
                                                                  'handicap_amenities' in property_details and len(
                                                                      property_details[
                                                                          'handicap_amenities']) > 0]
                property_details['selected_construction'] = [int(item['feature_id']) for item in
                                                                   property_details['construction'] if
                                                                   'construction' in property_details and len(
                                                                       property_details[
                                                                           'construction']) > 0]
                property_details['selected_garage_parking'] = [int(item['feature_id']) for item in
                                                             property_details['garage_parking'] if
                                                             'garage_parking' in property_details and len(
                                                                 property_details[
                                                                     'garage_parking']) > 0]
                property_details['selected_exterior_features'] = [int(item['feature_id']) for item in
                                                               property_details['exterior_features'] if
                                                               'exterior_features' in property_details and len(
                                                                   property_details[
                                                                       'exterior_features']) > 0]
                property_details['selected_garage_features'] = [int(item['feature_id']) for item in
                                                                  property_details['garage_features'] if
                                                                  'garage_features' in property_details and len(
                                                                      property_details[
                                                                          'garage_features']) > 0]
                property_details['selected_roof'] = [int(item['feature_id']) for item in
                                                                property_details['roof'] if
                                                                'roof' in property_details and len(
                                                                    property_details[
                                                                        'roof']) > 0]
                property_details['selected_outbuildings'] = [int(item['feature_id']) for item in
                                                     property_details['outbuildings'] if
                                                     'outbuildings' in property_details and len(
                                                         property_details[
                                                             'outbuildings']) > 0]
                property_details['selected_foundation'] = [int(item['feature_id']) for item in
                                                             property_details['foundation'] if
                                                             'foundation' in property_details and len(
                                                                 property_details[
                                                                     'foundation']) > 0]
                property_details['selected_location_features'] = [int(item['feature_id']) for item in
                                                           property_details['location_features'] if
                                                           'location_features' in property_details and len(
                                                               property_details[
                                                                   'location_features']) > 0]
                property_details['selected_fence'] = [int(item['feature_id']) for item in
                                                                  property_details['fence'] if
                                                                  'fence' in property_details and len(
                                                                      property_details[
                                                                          'fence']) > 0]
                property_details['selected_road_frontage'] = [int(item['feature_id']) for item in
                                                      property_details['road_frontage'] if
                                                      'road_frontage' in property_details and len(
                                                          property_details[
                                                              'road_frontage']) > 0]
                property_details['selected_pool'] = [int(item['feature_id']) for item in
                                                              property_details['pool'] if
                                                              'pool' in property_details and len(
                                                                  property_details[
                                                                      'pool']) > 0]
                property_details['selected_property_faces'] = [int(item['feature_id']) for item in
                                                     property_details['property_faces'] if
                                                     'property_faces' in property_details and len(
                                                         property_details[
                                                             'property_faces']) > 0]
                property_details['selected_lease_type'] = [int(item['feature_id']) for item in
                                                               property_details['lease_type'] if
                                                               'lease_type' in property_details and len(
                                                                   property_details[
                                                                       'lease_type']) > 0]
                property_details['selected_tenant_pays'] = [int(item['feature_id']) for item in
                                                           property_details['tenant_pays'] if
                                                           'tenant_pays' in property_details and len(
                                                               property_details[
                                                                   'tenant_pays']) > 0]
                property_details['selected_inclusions'] = [int(item['feature_id']) for item in
                                                            property_details['inclusions'] if
                                                            'inclusions' in property_details and len(
                                                                property_details[
                                                                    'inclusions']) > 0]
                property_details['selected_building_class'] = [int(item['feature_id']) for item in
                                                           property_details['building_class'] if
                                                           'building_class' in property_details and len(
                                                               property_details[
                                                                   'building_class']) > 0]
                property_details['selected_interior_features'] = [int(item['feature_id']) for item in
                                                               property_details['interior_features'] if
                                                               'interior_features' in property_details and len(
                                                                   property_details[
                                                                       'interior_features']) > 0]
                property_details['selected_mineral_rights'] = [int(item['feature_id']) for item in
                                                                  property_details['mineral_rights'] if
                                                                  'mineral_rights' in property_details and len(
                                                                      property_details[
                                                                          'mineral_rights']) > 0]
                property_details['selected_easements'] = [int(item['feature_id']) for item in
                                                               property_details['easements'] if
                                                               'easements' in property_details and len(
                                                                   property_details[
                                                                       'easements']) > 0]
                property_details['selected_survey'] = [int(item['feature_id']) for item in
                                                          property_details['survey'] if
                                                          'survey' in property_details and len(
                                                              property_details[
                                                                  'survey']) > 0]
                property_details['selected_utilities'] = [int(item['feature_id']) for item in
                                                       property_details['utilities'] if
                                                       'utilities' in property_details and len(
                                                           property_details[
                                                               'utilities']) > 0]
                property_details['selected_improvements'] = [int(item['feature_id']) for item in
                                                       property_details['improvements'] if
                                                       'improvements' in property_details and len(
                                                           property_details[
                                                               'improvements']) > 0]
                property_details['selected_topography'] = [int(item['feature_id']) for item in
                                                             property_details['topography'] if
                                                             'topography' in property_details and len(
                                                                 property_details[
                                                                     'topography']) > 0]
                property_details['selected_wildlife'] = [int(item['feature_id']) for item in
                                                           property_details['wildlife'] if
                                                           'wildlife' in property_details and len(
                                                               property_details[
                                                                   'wildlife']) > 0]
                property_details['selected_fish'] = [int(item['feature_id']) for item in
                                                         property_details['fish'] if
                                                         'fish' in property_details and len(
                                                             property_details[
                                                                 'fish']) > 0]
                property_details['selected_irrigation_system'] = [int(item['feature_id']) for item in
                                                     property_details['irrigation_system'] if
                                                     'irrigation_system' in property_details and len(
                                                         property_details[
                                                             'irrigation_system']) > 0]
                property_details['selected_recreation'] = [int(item['feature_id']) for item in
                                                                  property_details['recreation'] if
                                                                  'recreation' in property_details and len(
                                                                      property_details[
                                                                          'recreation']) > 0]

                insider_start_price = 0
                try:
                    if property_details['property_auction_data'] and 'start_price' in property_details['property_auction_data'][
                        0] and property_details['property_auction_data'][0]['start_price']:
                        insider_start_price = float(property_details['property_auction_data'][0]['start_price'])

                    if property_details['property_auction_data'] and 'insider_price_decrease' in property_details['property_auction_data'][
                        0] and property_details['property_auction_data'][0]['insider_price_decrease']:
                        insider_price_decrease = float(property_details['property_auction_data'][0]['insider_price_decrease'])

                    if insider_start_price and insider_price_decrease:
                        insider_decreased_price = (insider_start_price - ((insider_start_price*insider_price_decrease)/100))
                        property_details['insider_decreased_price'] = insider_decreased_price

                except:
                    pass

                if int(property_details['sale_by_type']) == 1:
                    status_param = {'object_id': 10}
                elif int(property_details['sale_by_type']) == 6:
                    status_param = {'object_id': 13}
                elif int(property_details['sale_by_type']) in [4, 7]:
                    status_param = {'object_id': 11}
                else:
                    status_param = {}


                try:
                    status_url = settings.API_URL + '/api-settings/lookup-status-listing/'
                    status_data = call_api_post_method(status_param, status_url, token)

                    status_list = status_data['data']
                except:
                    status_list = []



                try:
                    property_feature_listing_param = {"asset_id": asset_id}
                    property_feature_listing_url = settings.API_URL + '/api-property/property-features-listing/'
                    property_feature_listing_data = call_api_post_method(property_feature_listing_param, property_feature_listing_url, token)
                    property_features_listing = property_feature_listing_data['data']
                except:
                    property_features_listing = []

            except Exception as exp:
                print(exp)
                property_details = {}
                property_features_listing = []


        else:
            property_details = {}
            property_features_listing = []
            country_id = 1

        try:
            state_param = {'country_id': country_id}
            state_api_url = settings.API_URL + '/api-settings/get-state/'
            state_data = call_api_post_method(state_param, state_api_url, token)
            state_list = state_data['data']
        except:
            state_list = []

        res_comm_in = [2,3]
        comm_land_in = [1,2]
        res_land_in = [1,3]
        try:
            country_param = {}
            country_api_url = settings.API_URL + '/api-settings/get-country/'
            country_data = call_api_post_method(country_param, country_api_url)
            country_list = country_data['data']
        except:
            country_list = []
        context = {
            "data": "Listing COMING SOON...",
            "active_menu": "listing setting",
            "active_submenu": "add_listing",
            "asset_listing": asset_listing,
            "state_list": state_list,
            "property_details": property_details,
            "property_features_listing": property_features_listing,
            "auction_type_list": auction_type_list,
            "res_comm_in": res_comm_in,
            "comm_land_in": comm_land_in,
            "res_land_in": res_land_in,
            "status_list": status_list,
            "auction_status_list": auction_status_list,
            "timezone_list": timezone_list,
            "closing_status_list": closing_status_list,
            "country_list": country_list,
            "country_id": country_id
        }

        return render(request, "admin/dashboard/listings/add-listing-property-info.html", context)
    except Exception as exp:
        print(exp)
        return HttpResponse("Issue in views")

def listing_settings(request):
    is_permission = check_permission(request, 6)
    if not is_permission:
        http_host = request.META['HTTP_HOST']
        redirect_url = settings.URL_SCHEME + str(http_host)
        return HttpResponseRedirect(redirect_url)
    try:
        try:
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']
            token = request.session['token']['access_token']
            user_id = request.session['user_id']
        except:
            settings_value = {}
            user_id = ''
            site_id = ''

        try:
            is_broker = request.session['is_broker']
            if is_broker == True:
                is_broker = 1
                is_agent = 0
            else:
                is_broker = 0
                is_agent = 1
            params = {"site_id": site_id, 'is_global': 1, 'is_broker': is_broker, 'is_agent': is_agent}

            url = settings.API_URL + '/api-property/get-property-setting/'
            data = call_api_post_method(params, url, token)
            listing_setting = data['data']
        except:
            listing_setting = {}


        context = {"data": "Listing COMING SOON...", "active_menu": "listing", "listing_setting": listing_setting}
        return render(request, "admin/dashboard/listings/listing-settings.html", context)
    except Exception as exp:
        print(exp)
        return HttpResponse("Issue in views")

@csrf_exempt
def get_property_info_data(request):
    try:
        if request.is_ajax() and request.method == 'POST':
            # site_detail = subdomain_site_details(request)
            # site_id = site_detail['site_detail']['site_id']
            token = request.session['token']['access_token']
            data = {}

            # one api for all features
            try:
                property_features_params = {
                    'asset_id': request.POST['asset_id']
                }

                property_features_url = settings.API_URL + '/api-property/property-features-listing/'

                property_features_data = call_api_post_method(property_features_params, property_features_url, token)
                all_property_features = property_features_data['data']
                try:
                    property_type_listing = all_property_features['property_type']
                except:
                    property_type_listing = []
                try:
                    property_sub_type_listing = all_property_features['property_subtype']
                except:
                    property_sub_type_listing = []

                try:
                    lot_size_units = all_property_features['lot_size']
                except:
                    lot_size_units = []

                try:
                    term_accepted = all_property_features['terms_accepted']
                except:
                    term_accepted = []
                try:
                    occupied_by = all_property_features['occupied_by']
                except:
                    occupied_by = []
                try:
                    ownership = all_property_features['ownership']
                except:
                    ownership = []
                try:
                    possession = all_property_features['possession']
                except:
                    possession = []
                try:
                    property_styles = all_property_features['style']
                except:
                    property_styles = []
                try:
                    property_stories = all_property_features['stories']
                except:
                    property_stories = []
                try:
                    recent_updates = all_property_features['recent_updates']
                except:
                    recent_updates = []
                try:
                    security_features = all_property_features['security_features']
                except:
                    security_features = []
                try:
                    property_cooling = all_property_features['cooling']
                except:
                    property_cooling = []
                try:
                    property_heating = all_property_features['heating']
                except:
                    property_heating = []
                try:
                    property_electric = all_property_features['electric']
                except:
                    property_electric = []
                try:
                    property_gas = all_property_features['gas']
                except:
                    property_gas = []
                try:
                    property_water = all_property_features['water']
                except:
                    property_water = []
                try:
                    property_sewer = all_property_features['sewer']
                except:
                    property_sewer = []
                try:
                    property_zoning = all_property_features['zoning']
                except:
                    property_zoning = []
                try:
                    tax_exemptions = all_property_features['tax_exemptions']
                except:
                    tax_exemptions = []
                try:
                    hoa_amenties = all_property_features['amenities']
                except:
                    hoa_amenties = []
                try:
                    kitchen_features = all_property_features['kitchen_features']
                except:
                    kitchen_features = []
                try:
                    appliances = all_property_features['appliances']
                except:
                    appliances = []
                try:
                    property_flooring = all_property_features['flooring']
                except:
                    property_flooring = []
                try:
                    property_windows = all_property_features['windows']
                except:
                    property_windows = []
                try:
                    bedroom_features = all_property_features['bedroom_features']
                except:
                    bedroom_features = []
                try:
                    bathroom_features = all_property_features['bathroom_features']
                except:
                    bathroom_features = []
                try:
                    master_bedroom_features = all_property_features['master_bedroom_features']
                except:
                    master_bedroom_features = []
                try:
                    basement_features = all_property_features['basement_features']
                except:
                    basement_features = []
                try:
                    other_rooms = all_property_features['other_rooms']
                except:
                    other_rooms = []
                try:
                    other_features = all_property_features['other_features']
                except:
                    other_features = []
                try:
                    fire_place_units = all_property_features['fireplace_type']
                except:
                    fire_place_units = []
                try:
                    handicap_amenities = all_property_features['handicap_amenities']
                except:
                    handicap_amenities = []
                try:
                    property_construction = all_property_features['construction']
                except:
                    property_construction = []
                try:
                    exterior_features = all_property_features['exterior_features']
                except:
                    exterior_features = []
                try:
                    roofs = all_property_features['roof']
                except:
                    roofs = []
                try:
                    foundations = all_property_features['foundation']
                except:
                    foundations = []
                try:
                    fence_list = all_property_features['fence']
                except:
                    fence_list = []
                try:
                    pools = all_property_features['pool']
                except:
                    pools = []
                try:
                    garage_parkings = all_property_features['garage_parking']
                except:
                    garage_parkings = []
                try:
                    garage_features = all_property_features['garage_features']
                except:
                    garage_features = []
                try:
                    out_buildings = all_property_features['outbuildings']
                except:
                    out_buildings = []
                try:
                    location_features = all_property_features['location_features']
                except:
                    location_features = []
                try:
                    road_frontages = all_property_features['road_frontage']
                except:
                    road_frontages = []
                try:
                    property_faces = all_property_features['property_faces']
                except:
                    property_faces = []
                try:
                    lease_types = all_property_features['lease_type']
                except:
                    lease_types = []
                try:
                    tenant_pays = all_property_features['tenant_pays']
                except:
                    tenant_pays = []
                try:
                    inclusions = all_property_features['inclusions']
                except:
                    inclusions = []
                try:
                    building_classes = all_property_features['building_class']
                except:
                    building_classes = []

                try:
                    interior_features = all_property_features['interior_features']
                except:
                    interior_features = []
                try:
                    mineral_rights = all_property_features['mineral_rights']
                except:
                    mineral_rights = []
                try:
                    easements = all_property_features['easements']
                except:
                    easements = []
                try:
                    survey = all_property_features['survey']
                except:
                    survey = []
                try:
                    utilities = all_property_features['utilities']
                except:
                    utilities = []
                try:
                    improvements = all_property_features['improvements']
                except:
                    improvements = []

                try:
                    irrigation_system = all_property_features['irrigation_system']
                except:
                    irrigation_system = []
                try:
                    topography = all_property_features['topography']
                except:
                    topography = []
                try:
                    wildlife = all_property_features['wildlife']
                except:
                    wildlife = []
                try:
                    fish = all_property_features['fish']
                except:
                    fish = []
                try:
                    recreation = all_property_features['recreation']
                except:
                    recreation = []

            except:
                property_type_listing = []
                property_sub_type_listing = []
                lot_size_units = []
                term_accepted = []
                occupied_by = []
                ownership = []
                possession = []
                property_styles = []
                property_stories = []
                recent_updates = []
                security_features = []
                property_cooling = []
                property_heating = []
                property_electric = []
                property_gas = []
                property_water = []
                property_sewer = []
                property_zoning = []
                tax_exemptions = []

                hoa_amenties = []
                kitchen_features = []
                appliances = []
                property_flooring = []
                property_windows = []
                bedroom_features = []
                bathroom_features = []
                master_bedroom_features = []
                basement_features = []
                other_rooms = []
                other_features = []
                fire_place_units = []
                handicap_amenities = []
                property_construction = []
                exterior_features = []
                roofs = []
                foundations = []
                fence_list = []
                pools = []
                garage_parkings = []
                garage_features = []
                out_buildings = []
                location_features = []
                road_frontages = []
                property_faces = []
                lease_types = []
                tenant_pays = []
                inclusions = []
                building_classes = []
                interior_features = []
                mineral_rights = []
                easements = []
                survey = []
                utilities = []
                improvements = []
                irrigation_system = []
                topography = []
                wildlife = []
                fish = []
                recreation = []


            data = {
                'status': 200,
                'error': 0,
                'property_type_listing': property_type_listing,
                'property_sub_type_listing': property_sub_type_listing,
                'lot_size_units': lot_size_units,
                'term_accepted': term_accepted,
                'occupied_by': occupied_by,
                'ownership': ownership,
                'possession': possession,
                'property_styles': property_styles,
                'property_stories': property_stories,
                'recent_updates': recent_updates,
                'security_features': security_features,
                'property_cooling': property_cooling,
                'property_heating': property_heating,
                'property_electric': property_electric,
                'property_gas': property_gas,
                'property_water': property_water,
                'property_sewer': property_sewer,
                'property_zoning': property_zoning,
                'tax_exemptions': tax_exemptions,
                'hoa_amenties': hoa_amenties,
                'kitchen_features': kitchen_features,
                'appliances': appliances,
                'property_flooring': property_flooring,
                'property_windows': property_windows,
                'bedroom_features': bedroom_features,
                'bathroom_features': bathroom_features,
                'master_bedroom_features': master_bedroom_features,
                'basement_features': basement_features,
                'other_rooms': other_rooms,
                'other_features': other_features,
                'fire_place_units': fire_place_units,
                'handicap_amenities': handicap_amenities,
                'property_construction': property_construction,
                'exterior_features': exterior_features,
                'roofs': roofs,
                'foundations': foundations,
                'fence_list': fence_list,
                'pools': pools,
                'garage_parkings': garage_parkings,
                'garage_features': garage_features,
                'out_buildings': out_buildings,
                'location_features': location_features,
                'road_frontages': road_frontages,
                'property_faces': property_faces,
                'lease_types': lease_types,
                'tenant_pays': tenant_pays,
                'inclusions': inclusions,
                'interior_features': interior_features,
                'mineral_rights': mineral_rights,
                'easements': easements,
                'survey': survey,
                'utilities': utilities,
                'improvements': improvements,
                'irrigation_system': irrigation_system,
                'topography': topography,
                'wildlife': wildlife,
                'fish': fish,
                'recreation': recreation,
                'building_classes': building_classes,
            }
        else:
            data = {'status': 403, 'property_type_listing': [], 'error': 1}

        return JsonResponse(data)
    except Exception as exp:
        print(exp)
        data = {'status': 403, 'suggestion_list': [], 'error': 1}
        return JsonResponse(data)

def property_map_view(request):
    property_id = request.GET.get('property_id', None)
    is_permission = check_permission(request, 6)
    if not is_permission:
        http_host = request.META['HTTP_HOST']
        redirect_url = settings.URL_SCHEME + str(http_host)
        return HttpResponseRedirect(redirect_url)
    elif not property_id:
        return HttpResponseRedirect('/admin/listing/')
    try:
        token = request.session['token']['access_token']
        site_detail = subdomain_site_details(request)
        site_id = site_detail['site_detail']['site_id']
        try:
            asset_listing_params = {}

            asset_listing_url = settings.API_URL + '/api-property/asset-listing/'

            asset_listing_data = call_api_post_method(asset_listing_params, asset_listing_url, token)
            asset_listing = asset_listing_data['data']
        except:
            asset_listing = []

        if property_id is not None:
            property_detail_param = {"site_id": site_id, "property_id": property_id, "step_id": 2}
            property_detail_url = settings.API_URL + '/api-property/subdomain-property-detail/'
            property_detail_data = call_api_post_method(property_detail_param, property_detail_url, token)
            property_address = ''
            try:
                property_details = property_detail_data['data']
                if 'address_one' in property_details and property_details['address_one'] != "":
                    property_address = property_details['address_one']
                if 'address_two' in property_details and property_details['address_two'] is not None and property_details['address_two'] != "":
                    property_address = property_address+','+property_details['address_two']
                if 'city' in property_details and property_details['city'] is not None and property_details['city'] != "":
                    property_address = property_address+','+property_details['city']
                if 'state' in property_details and property_details['state'] is not None and property_details['state'] != "":
                    property_address = property_address+','+property_details['state']
                if 'postal_code' in property_details and property_details['postal_code'] is not None and property_details['postal_code'] != "":
                    property_address = property_address+','+property_details['postal_code']
            except:
                property_details = {}

            if 'latitude' not in property_details or not property_details['latitude'] or 'longitude' not in property_details or not property_details['longitude']:
                try:
                    geolocator = Nominatim(user_agent=settings.GEOLOCATOR_EMAIL, timeout=10)
                    location = geolocator.geocode(property_address)
                    try:
                        property_details['latitude'] = location.latitude
                        property_details['longitude'] = location.longitude
                    except:
                        property_details['latitude'] = 0
                        property_details['longitude'] = 0
                except:
                    geolocator = Nominatim(user_agent=settings.GEOLOCATOR_EMAIL, timeout=10)
                    location = geolocator.geocode("47 W 13th St, New York, NY 10011, USA")
                    try:
                        property_details['latitude'] = location.latitude
                        property_details['longitude'] = location.longitude
                    except:
                        property_details['latitude'] = 0
                        property_details['longitude'] = 0

        else:
            property_details = {}

        context = {
            "data": "Listing COMING SOON...",
            "active_menu": "listing",
            "asset_listing": asset_listing,
            "property_details": property_details,
        }
        return render(request, "admin/dashboard/listings/property-map-street-view.html", context)
    except Exception as exp:
        print(exp)
        return HttpResponse("Issue in views")

def property_photo_video(request):
    property_id = request.GET.get('property_id', None)
    is_permission = check_permission(request, 6)
    if not is_permission:
        http_host = request.META['HTTP_HOST']
        redirect_url = settings.URL_SCHEME + str(http_host)
        return HttpResponseRedirect(redirect_url)
    elif not property_id:
        return HttpResponseRedirect('/admin/listing/')
    try:
        token = request.session['token']['access_token']
        site_detail = subdomain_site_details(request)
        site_id = site_detail['site_detail']['site_id']
        try:
            asset_listing_params = {}

            asset_listing_url = settings.API_URL + '/api-property/asset-listing/'

            asset_listing_data = call_api_post_method(asset_listing_params, asset_listing_url, token)
            asset_listing = asset_listing_data['data']
        except:
            asset_listing = []

        if property_id is not None:
            property_detail_param = {"site_id": site_id, "property_id": property_id, "step_id": 3}
            property_detail_url = settings.API_URL + '/api-property/subdomain-property-detail/'
            property_detail_data = call_api_post_method(property_detail_param, property_detail_url, token)
            try:
                property_details = property_detail_data['data']
                str_prop_image_id = ''
                prop_image_id = ''
                str_prop_image_name = ''
                prop_image_name = ''
                str_prop_video_id = ''
                prop_video_id = ''
                str_prop_video_name = ''
                prop_video_name = ''
                property_img_list = property_details['photo']
                property_video_list = property_details['video']
                if property_img_list:
                    cnt = 0
                    for item in property_img_list:
                        if cnt == 0:
                            str_prop_image_id = str(item['upload_id'])
                            str_prop_image_name = str(item['doc_file_name'])
                        else:
                            str_prop_image_id = str_prop_image_id + ',' + str(item['upload_id'])
                            str_prop_image_name = str_prop_image_name + ',' + str(item['doc_file_name'])
                        cnt += 1
                    prop_image_id = str_prop_image_id.rstrip(',')
                    prop_image_name = str_prop_image_name.rstrip(',')
                    property_details['property_image_ids'] = prop_image_id
                    property_details['property_image_name'] = prop_image_name
                else:
                    property_details['property_image_ids'] = prop_image_id
                    property_details['property_image_name'] = prop_image_name
                if property_video_list:
                    cnt = 0
                    for item in property_video_list:
                        if cnt == 0:
                            str_prop_video_id = str(item['upload_id'])
                            str_prop_video_name = str(item['doc_file_name'])
                        else:
                            str_prop_video_id = str_prop_video_id + ',' + str(item['upload_id'])
                            str_prop_video_name = str_prop_video_name + ',' + str(item['doc_file_name'])
                        cnt += 1
                    prop_video_id = str_prop_video_id.rstrip(',')
                    prop_video_name = str_prop_video_name.rstrip(',')
                    property_details['property_video_ids'] = prop_video_id
                    property_details['property_video_name'] = prop_video_name
                else:
                    property_details['property_video_ids'] = prop_video_id
                    property_details['property_video_name'] = prop_video_name
            except:
                property_details = {}
        else:
            property_details = {}
        context = {
            "data": "Listing COMING SOON...",
            "active_menu": "listing",
            "asset_listing": asset_listing,
            "property_details": property_details,
        }
        return render(request, "admin/dashboard/listings/property-photo-video.html", context)
    except Exception as exp:
        print(exp)
        return HttpResponse("Issue in views")

def property_document(request):
    property_id = request.GET.get('property_id', None)
    is_permission = check_permission(request, 6)
    if not is_permission:
        http_host = request.META['HTTP_HOST']
        redirect_url = settings.URL_SCHEME + str(http_host)
        return HttpResponseRedirect(redirect_url)
    elif not property_id:
        return HttpResponseRedirect('/admin/listing/')
    try:
        token = request.session['token']['access_token']
        site_detail = subdomain_site_details(request)
        site_id = site_detail['site_detail']['site_id']
        try:
            asset_listing_params = {}

            asset_listing_url = settings.API_URL + '/api-property/asset-listing/'

            asset_listing_data = call_api_post_method(asset_listing_params, asset_listing_url, token)
            asset_listing = asset_listing_data['data']
        except:
            asset_listing = []

        if property_id is not None:
            property_detail_param = {"site_id": site_id, "property_id": property_id, "step_id": 4}
            property_detail_url = settings.API_URL + '/api-property/subdomain-property-detail/'
            property_detail_data = call_api_post_method(property_detail_param, property_detail_url, token)
            try:
                property_details = property_detail_data['data']
                str_prop_doc_id = ''
                prop_doc_id = ''
                str_prop_doc_name = ''
                prop_doc_name = ''
                property_doc_list = property_details['documents']
                if property_doc_list:
                    cnt = 0
                    for item in property_doc_list:
                        if cnt == 0:
                            str_prop_doc_id = str(item['upload_id'])
                            str_prop_doc_name = str(item['doc_file_name'])
                        else:
                            str_prop_doc_id = str_prop_doc_id + ',' + str(item['upload_id'])
                            str_prop_doc_name = str_prop_doc_name + ',' + str(item['doc_file_name'])
                        cnt += 1
                    prop_doc_id = str_prop_doc_id.rstrip(',')
                    prop_doc_name = str_prop_doc_name.rstrip(',')
                    property_details['property_doc_ids'] = prop_doc_id
                    property_details['property_doc_name'] = prop_doc_name
                else:
                    property_details['property_doc_ids'] = prop_video_id
                    property_details['property_doc_name'] = prop_doc_name
            except:
                property_details = {}
        else:
            property_details = {}
        context = {
            "data": "Listing COMING SOON...",
            "active_menu": "listing",
            "asset_listing": asset_listing,
            "property_details": property_details,
        }
        return render(request, "admin/dashboard/listings/property-document-vault.html", context)
    except Exception as exp:
        print(exp)
        return HttpResponse("Issue in views")


def save_property(request):
    try:
        if request.is_ajax() and request.method == 'POST':
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']
            token = request.session['token']['access_token']
            user_id = request.session['user_id']
            next_url = request.POST['next_url'] if 'next_url' in request.POST and request.POST['next_url'] != "" else ""
            property_id = request.POST['property_id'] if 'property_id' in request.POST and request.POST['property_id'] != "" else ""
            is_featured = request.POST['is_featured'] if 'is_featured' in request.POST and request.POST['is_featured'] != "" else 0
            property_param = {}
            open_house_dates = []
            step = request.POST['step']
            is_reset_offer = 0
            auction_type = ''
            # print(request.POST)
            if int(request.POST['step']) == 1:
                is_reset_offer = request.POST['is_reset_offer']
                auction_type = request.POST['auction_type']

                buyers_premium_percentage = None
                if 'buyers_premium_percentage' in request.POST and request.POST['buyers_premium_percentage'] != "" and request.POST['buyers_premium_percentage'] != "$":
                    buyers_premium_percentage = float(request.POST['buyers_premium_percentage'].replace(',', '').replace('$', ''))

                buyers_premium_min_amount = None
                if 'buyers_premium_min_amount' in request.POST and request.POST['buyers_premium_min_amount'] != "" and request.POST['buyers_premium_min_amount'] != "$":
                    buyers_premium_min_amount = float(request.POST['buyers_premium_min_amount'].replace(',', '').replace('$', ''))

                deposit_amount = None
                if 'deposit_amount' in request.POST and request.POST['deposit_amount'] != "" and request.POST['deposit_amount'] != "$":
                    deposit_amount = float(request.POST['deposit_amount'].replace(',', '').replace('$', '')) 

                is_deposit_required = 0
                if 'is_deposit_required' in request.POST and request.POST['is_deposit_required'] != "":
                    is_deposit_required = request.POST['is_deposit_required']      

                bidding_min_price = None
                if 'bidding_min_price' in request.POST and request.POST['bidding_min_price'] != "" and request.POST['bidding_min_price'] != "$":
                    bidding_min_price = float(request.POST['bidding_min_price'].replace(',', '').replace('$', ''))

                reserve_amount = None
                if 'reserve_amount' in request.POST and request.POST['reserve_amount'] and int(auction_type) != 4:
                    reserve_amount = int(request.POST['reserve_amount'].replace(',', '').replace('$', ''))

                offer_amount = None
                if 'offer_amount' in request.POST and request.POST['offer_amount'] != "" and int(request.POST['auction_type']) == 7:
                    offer_amount = float(request.POST['offer_amount'].replace(',', '').replace('$', ''))

                bid_increments = None
                insider_bid_increment = None
                if 'bid_increments' in request.POST and request.POST['bid_increments'] != "":
                    bid_increments = float(request.POST['bid_increments'].replace(',', '').replace('$', ''))

                earnest_deposit = None
                earnest_deposit_type = None
                highest_best_format = None
                insider_start_price = None
                if 'auction_type' in request.POST and request.POST['auction_type'] != "" and int(request.POST['auction_type']) == 7:
                    earnest_deposit = request.POST['earnest_deposit'].replace(',', '').replace('$', '').replace('%', '')
                    earnest_deposit_type = request.POST['earnest_deposit_type']
                    highest_best_format = request.POST['highest_offer_format'] if 'highest_offer_format' in request.POST else None
                if 'auction_type' in request.POST and request.POST['auction_type'] != "" and int(request.POST['auction_type']) == 2:
                    if 'insider_bid_increment' in request.POST and request.POST['insider_bid_increment'] != "":
                        insider_bid_increment = float(request.POST['insider_bid_increment'].replace(',', '').replace('$', ''))
                    if 'insider_start_price' in request.POST and request.POST['insider_start_price'] != "" and request.POST[
                        'insider_start_price'] != "$":
                        insider_start_price = float(request.POST['insider_start_price'].replace(',', '').replace('$', ''))
                    property_auction_data = {
                        'bid_increments': insider_bid_increment,
                        'time_zone': request.POST['bidding_time_zone'] if 'bidding_time_zone' in request.POST and request.POST['bidding_time_zone'] != "" else None,
                        'start_price': insider_start_price,
                        'auction_status': request.POST['auction_status'] if 'auction_status' in request.POST and request.POST['auction_status'] != "" else None,
                        'start_date': request.POST['dutch_bidding_start_date'] if 'dutch_bidding_start_date' in request.POST and request.POST['dutch_bidding_start_date'] != "" else None,
                        'dutch_end_time': request.POST['dutch_bidding_end_date'] if 'dutch_bidding_end_date' in request.POST and request.POST['dutch_bidding_end_date'] != "" else None,
                        'insider_price_decrease': request.POST['price_decrease_rate'] if 'price_decrease_rate' in request.POST and request.POST['price_decrease_rate'] != "" else None,
                        'dutch_time': request.POST['dutch_auction_time'] if 'dutch_auction_time' in request.POST and request.POST['dutch_auction_time'] != "" else None,
                        'dutch_pause_time': request.POST['dutch_pause_time'] if 'dutch_pause_time' in request.POST and request.POST['dutch_pause_time'] != "" else None,
                        'sealed_start_time': request.POST['sealed_bidding_start_date'] if 'sealed_bidding_start_date' in request.POST and request.POST['sealed_bidding_start_date'] != "" else None,
                        'sealed_end_time': request.POST['sealed_bidding_end_date'] if 'sealed_bidding_end_date' in request.POST and request.POST['sealed_bidding_end_date'] != "" else None,
                        'sealed_pause_time': request.POST['sealed_pause_time'] if 'sealed_pause_time' in request.POST and request.POST['sealed_pause_time'] != "" else None,
                        'sealed_time': request.POST['sealed_auction_time'] if 'sealed_auction_time' in request.POST and request.POST['sealed_auction_time'] != "" else None,
                        'english_start_time': request.POST['english_bidding_start_date'] if 'english_bidding_start_date' in request.POST and request.POST['english_bidding_start_date'] != "" else None,
                        'end_date': request.POST['english_bidding_end_date'] if 'english_bidding_end_date' in request.POST and request.POST['english_bidding_end_date'] != "" else None,
                        'english_time': request.POST['english_auction_time'] if 'english_auction_time' in request.POST and request.POST['english_auction_time'] != "" else None,
                    }
                else:
                    property_auction_data = {
                        'start_date': request.POST['bidding_start_date'] if 'bidding_start_date' in request.POST and request.POST['bidding_start_date'] != "" else None,
                        'end_date': request.POST['bidding_end_date'] if 'bidding_end_date' in request.POST and request.POST['bidding_start_date'] != "" else None,
                        'bid_increments': bid_increments,
                        'reserve_amount': reserve_amount,
                        'time_zone': request.POST['bidding_time_zone'] if 'bidding_time_zone' in request.POST and request.POST['bidding_time_zone'] != "" else None,
                        'start_price': bidding_min_price,
                        'auction_status': request.POST['auction_status'] if 'auction_status' in request.POST and request.POST['auction_status'] != "" else None,
                        #'open_house_start_date': request.POST['open_house_start_date'] if 'open_house_start_date' in request.POST and request.POST['open_house_start_date'] != "" else None,
                        #'open_house_end_date': request.POST['open_house_end_date'] if 'open_house_end_date' in request.POST and request.POST['open_house_end_date'] != "" else None,
                        'offer_amount': offer_amount,
                    }

                if 'asset_type' in request.POST and int(request.POST['asset_type']) == 3:
                    total = int(request.POST['total_section'])
                    for i in range(total):
                        open_house_dates_params = {
                            'start_date': request.POST['open_house_start_date_' + str(i)],
                            'end_date': request.POST['open_house_end_date_' + str(i)],
                        }
                        open_house_dates.append(open_house_dates_params)

                    property_param = {
                        'site_id': site_id,
                        'user_id': user_id,
                        'property_id': property_id,
                        'property_asset': request.POST['asset_type'] if 'asset_type' in request.POST and request.POST['asset_type'] != "" else None,
                        'step': request.POST['step'],
                        'description': request.POST['property_desc'] if 'property_desc' in request.POST and request.POST['property_desc'] != "" else None,
                        'address_one': request.POST['property_address_one'] if 'property_address_one' in request.POST and request.POST['property_address_one'] != "" else None,
                        'city': request.POST['property_city'] if 'property_city' in request.POST and request.POST['property_city'] != "" else None,
                        'state': request.POST['property_state'] if 'property_state' in request.POST and request.POST['property_state'] != "" else None,
                        'postal_code': request.POST['property_zip_code'] if 'property_zip_code' in request.POST and request.POST['property_zip_code'] != "" else None,
                        'property_type': request.POST['property_type'] if 'property_type' in request.POST and request.POST['property_type'] != "" else None,
                        'property_subtype': request.POST.getlist('property_sub_type') if 'property_sub_type' in request.POST else [],
                        'beds': request.POST['num_beds'] if 'num_beds' in request.POST and request.POST['num_beds'] != "" else None,
                        'baths': request.POST['num_bath'] if 'num_bath' in request.POST and request.POST['num_bath'] != "" else None,
                        'year_built': request.POST['year_build'] if 'year_build' in request.POST and request.POST['year_build'] != "" else None,
                        'year_renovated': request.POST['year_renovated'] if 'year_renovated' in request.POST and request.POST['year_renovated'] != "" else None,
                        'square_footage': request.POST['square_footage'] if 'square_footage' in request.POST and request.POST['square_footage'] != "" else None,
                        'lot_size': request.POST['lot_size'] if 'lot_size' in request.POST and request.POST['lot_size'] != "" else None,
                        'lot_size_unit': request.POST['lot_size_unit'] if 'lot_size_unit' in request.POST and request.POST['lot_size_unit'] and int(request.POST['lot_size_unit']) > 0 else None,
                        'home_warranty': request.POST['home_warranty'] if 'home_warranty' in request.POST and request.POST['home_warranty'] != "" else 0,
                        'lot_dimensions': request.POST['lot_dimension'] if 'lot_dimension' in request.POST and request.POST['lot_dimension'] != "" else None,
                        'broker_co_op': request.POST['broker_co_op'] if 'broker_co_op' in request.POST and request.POST['broker_co_op'] != "" else 0,
                        'terms_accepted': request.POST.getlist('term_accepted') if 'term_accepted' in request.POST else [],
                        'occupied_by': request.POST.getlist('occupied_by') if 'occupied_by' in request.POST else [],
                        'ownership': request.POST.getlist('ownership') if 'ownership' in request.POST else [],
                        'possession': request.POST.getlist('possession') if 'possession' in request.POST else [],
                        'style': request.POST.getlist('property_style') if 'property_style' in request.POST else [],
                        'cooling': request.POST.getlist('property_cooling') if 'property_cooling' in request.POST else [],
                        'stories': request.POST.getlist('property_stories') if 'property_stories' in request.POST else [],
                        'heating': request.POST.getlist('property_heating') if 'property_heating' in request.POST else [],
                        'garage_spaces': request.POST['garage_space'] if 'garage_space' in request.POST and request.POST['garage_space'] != "" else None,
                        'electric': request.POST.getlist('property_electric') if 'property_electric' in request.POST else [],
                        'basement': request.POST['property_basement'] if 'property_basement' in request.POST and request.POST['property_basement'] != "" else 0,
                        'gas': request.POST.getlist('property_gas') if 'property_gas' in request.POST else [],
                        'recent_updates': request.POST.getlist('recent_update') if 'recent_update' in request.POST else [],
                        'water': request.POST.getlist('property_water') if 'property_water' in request.POST else [],
                        'security_features': request.POST.getlist('security_feature') if 'security_feature' in request.POST else [],
                        'sewer': request.POST.getlist('property_sewer') if 'property_sewer' in request.POST else [],
                        'property_taxes': request.POST['property_taxes'] if 'property_taxes' in request.POST and request.POST['property_taxes'] != "" else None,
                        'special_assessment_tax': request.POST['special_assessment_taxes'] if 'special_assessment_taxes' in request.POST and request.POST['special_assessment_taxes'] != "" else None,
                        'county': request.POST['property_country'] if 'property_country' in request.POST and request.POST['property_country'] != "" else None,
                        'tax_exemptions': request.POST.getlist('tax_exemption') if 'tax_exemption' in request.POST else [],
                        'zoning': request.POST.getlist('property_zoning') if 'property_zoning' in request.POST else [],
                        'hoa_fee': request.POST['hoa_fee'] if 'hoa_fee' in request.POST and request.POST['hoa_fee'] != "" else None,
                        'hoa_fee_type': request.POST['hoa_unit'] if 'hoa_unit' in request.POST and request.POST['hoa_unit'] != "" else None,
                        'subdivision': request.POST['property_subdivision'] if 'property_country' in request.POST and request.POST['property_country'] != "" else None,
                        'hoa_amenities': request.POST.getlist('hoa_amenties') if 'hoa_amenties' in request.POST else [],
                        'school_district': request.POST['school_district'] if 'school_district' in request.POST and request.POST['school_district'] != "" else None,
                        'upper_floor_area': request.POST['sqft_upper_floor'] if 'sqft_upper_floor' in request.POST and request.POST['sqft_upper_floor'] != "" else None,
                        'main_floor_area': request.POST['sqft_main_floor'] if 'sqft_main_floor' in request.POST and request.POST['sqft_main_floor'] != "" else None,
                        'basement_area': request.POST['sqft_basement'] if 'sqft_basement' in request.POST and request.POST['sqft_basement'] != "" else None,
                        'upper_floor_bedroom': request.POST['bed_rooms_upper_floor'] if 'bed_rooms_upper_floor' in request.POST and request.POST['bed_rooms_upper_floor'] != "" else None,
                        'main_floor_bedroom': request.POST['bed_rooms_main_floor'] if 'bed_rooms_main_floor' in request.POST and request.POST['bed_rooms_main_floor'] != "" else None,
                        'basement_bedroom': request.POST['bed_rooms_basement'] if 'bed_rooms_basement' in request.POST and request.POST['bed_rooms_basement'] != "" else None,
                        'upper_floor_bathroom': request.POST['bath_rooms_upper_floor'] if 'bath_rooms_upper_floor' in request.POST and request.POST['bath_rooms_upper_floor'] != "" else None,
                        'main_floor_bathroom': request.POST['bath_rooms_main_floor'] if 'bath_rooms_main_floor' in request.POST and request.POST['bath_rooms_main_floor'] != "" else None,
                        'basement_bathroom': request.POST['bath_rooms_basement'] if 'bath_rooms_basement' in request.POST and request.POST['bath_rooms_basement'] != "" else None,
                        'kitchen_features': request.POST.getlist('kitchen_features') if 'kitchen_features' in request.POST else [],
                        'appliances': request.POST.getlist('appliances') if 'appliances' in request.POST else [],
                        'flooring': request.POST.getlist('property_flooring') if 'property_flooring' in request.POST else [],
                        'windows': request.POST.getlist('property_windows') if 'property_windows' in request.POST else [],
                        'bedroom_features': request.POST.getlist('bedroom_features') if 'bedroom_features' in request.POST else [],
                        'other_rooms': request.POST.getlist('other_rooms') if 'other_rooms' in request.POST else [],
                        'bathroom_features': request.POST.getlist('bathroom_features') if 'bathroom_features' in request.POST else [],
                        'other_features': request.POST.getlist('other_features') if 'other_features' in request.POST else [],
                        'master_bedroom_features': request.POST.getlist('master_bedroom_features') if 'master_bedroom_features' in request.POST else [],
                        'fireplace': request.POST['fire_place'] if 'fire_place' in request.POST and request.POST['fire_place'] != "" else None,
                        'fireplace_type': request.POST.getlist('fire_place_unit') if 'fire_place_unit' in request.POST else [],
                        'basement_features': request.POST.getlist('basement_features') if 'basement_features' in request.POST else [],
                        'handicap_amenities': request.POST.getlist('handicap_amenities') if 'handicap_amenities' in request.POST else [],
                        'construction': request.POST.getlist('property_construction') if 'property_construction' in request.POST else [],
                        'garage_parking': request.POST.getlist('garage_parking') if 'garage_parking' in request.POST else [],
                        'exterior_features': request.POST.getlist('exterior_features') if 'exterior_features' in request.POST else [],
                        'garage_features': request.POST.getlist('garage_features') if 'garage_features' in request.POST else [],
                        'roof': request.POST.getlist('roofs') if 'roofs' in request.POST else [],
                        'outbuildings': request.POST.getlist('out_buildings') if 'out_buildings' in request.POST else [],
                        'foundation': request.POST.getlist('foundations') if 'foundations' in request.POST else [],
                        'location_features': request.POST.getlist('location_features') if 'location_features' in request.POST else [],
                        'fence': request.POST.getlist('fence') if 'fence' in request.POST else [],
                        'road_frontage': request.POST.getlist('road_frontages') if 'road_frontages' in request.POST else [],
                        'pool': request.POST.getlist('pools') if 'pools' in request.POST else [],
                        'property_faces': request.POST.getlist('property_faces') if 'property_faces' in request.POST else [],
                        'sale_terms': request.POST['term_of_sale'] if 'term_of_sale' in request.POST and request.POST['term_of_sale'] != "" else None,
                        'property_auction_data': property_auction_data,
                        'sale_by_type': request.POST['auction_type'] if 'auction_type' in request.POST and request.POST[
                            'auction_type'] != "" else None,
                        'is_featured': is_featured,
                        'status': request.POST['prop_listing_status'] if 'prop_listing_status' in request.POST and request.POST['prop_listing_status'] else "",
                        'property_opening_dates': open_house_dates,
                        'auction_location': request.POST['auction_location'] if 'auction_location' in request.POST and
                                                                         request.POST['auction_location'] else "",
                        'closing_status': request.POST['closing_status'] if 'closing_status' in request.POST and
                                                                         request.POST['closing_status'] else "",
                        'financing_available': request.POST['financing_available'] if 'financing_available' in request.POST and request.POST['financing_available'] else "",
                        'country': request.POST['prop_country'] if 'prop_country' in request.POST and request.POST['prop_country'] != "" else None,
                        'buyers_premium': int(request.POST['buyers_premium']) if 'buyers_premium' in request.POST and request.POST['buyers_premium'] != "" else None,
                        'buyers_premium_percentage': buyers_premium_percentage,
                        'buyers_premium_min_amount': buyers_premium_min_amount,

                    }

                if 'asset_type' in request.POST and int(request.POST['asset_type']) == 2:
                    total = int(request.POST['total_section'])
                    for i in range(total):
                        open_house_dates_params = {
                            'start_date': request.POST['open_house_start_date_' + str(i)],
                            'end_date': request.POST['open_house_end_date_' + str(i)],
                        }
                        open_house_dates.append(open_house_dates_params)
                    property_param = {
                        "site_id": site_id,
                        "property_asset": request.POST['asset_type'] if 'asset_type' in request.POST and request.POST['asset_type'] != "" else None,
                        "step": request.POST['step'],
                        "user_id": user_id,
                        "property_id": property_id,
                        "description": request.POST['property_desc'] if 'property_desc' in request.POST and request.POST['property_desc'] != "" else None,
                        "address_one": request.POST['property_address_one'] if 'property_address_one' in request.POST and request.POST['property_address_one'] != "" else None,
                        "city": request.POST['property_city'] if 'property_city' in request.POST and request.POST['property_city'] != "" else None,
                        "state": request.POST['property_state'] if 'property_state' in request.POST and request.POST['property_state'] != "" else None,
                        "postal_code": request.POST['property_zip_code'] if 'property_zip_code' in request.POST and request.POST['property_zip_code'] != "" else None,
                        "property_type": request.POST['property_type'] if 'property_type' in request.POST and request.POST['property_type'] != "" else None,
                        "broker_co_op": request.POST['broker_co_op'] if 'broker_co_op' in request.POST and request.POST['broker_co_op'] != "" else 0,
                        "property_subtype": request.POST.getlist('property_sub_type') if 'property_sub_type' in request.POST else [],
                        "terms_accepted": request.POST.getlist('term_accepted') if 'term_accepted' in request.POST else [],
                        "year_built": request.POST['year_build'] if 'year_build' in request.POST and request.POST['year_build'] != "" else None,
                        "year_renovated": request.POST['year_renovated'] if 'year_renovated' in request.POST and request.POST['year_renovated'] != "" else None,
                        "square_footage": request.POST['square_footage'] if 'square_footage' in request.POST and request.POST['square_footage'] != "" else None,
                        "lot_size": request.POST['lot_size'] if 'lot_size' in request.POST and request.POST['lot_size'] != "" else None,
                        "lot_size_unit": request.POST['lot_size_unit'] if 'lot_size_unit' in request.POST and request.POST['lot_size_unit'] and int(request.POST['lot_size_unit']) > 0 else None,
                        "lot_dimensions": request.POST['lot_dimension'] if 'lot_dimension' in request.POST and request.POST['lot_dimension'] != "" else None,
                        "lease_type": request.POST.getlist('lease_type') if 'lease_type' in request.POST else [],
                        "lease_expiration": request.POST['lease_exp_date'] if 'lease_exp_date' in request.POST and request.POST['lease_exp_date'] != "" else None,
                        "occupied_by": request.POST.getlist('occupied_by') if 'occupied_by' in request.POST else [],
                        "ownership": request.POST.getlist('ownership') if 'ownership' in request.POST else [],
                        "possession": request.POST.getlist('possession') if 'possession' in request.POST else [],
                        "total_buildings": request.POST['total_buildings'] if 'total_buildings' in request.POST and request.POST['total_buildings'] != "" else None,
                        "total_units": request.POST['total_units'] if 'total_units' in request.POST and request.POST['total_units'] != "" else None,
                        "net_operating_income": request.POST['net_operating_income'] if 'net_operating_income' in request.POST and request.POST['net_operating_income'] != "" else None,
                        "occupancy": request.POST['occupancy'] if 'occupancy' in request.POST and request.POST['occupancy'] != "" else None,
                        "total_floors": request.POST['total_floors'] if 'total_floors' in request.POST and request.POST['total_floors'] != "" else None,
                        "garage_spaces": request.POST['garage_space'] if 'garage_space' in request.POST and request.POST['garage_space'] != "" else None,
                        "cap_rate": request.POST['cap_rate'] if 'cap_rate' in request.POST and request.POST['cap_rate'] != "" else None,
                        "average_monthly_rate": request.POST['avg_rent_unit'] if 'avg_rent_unit' in request.POST and request.POST['avg_rent_unit'] != "" else None,
                        "basement": request.POST['property_basement'] if 'property_basement' in request.POST and request.POST['property_basement'] != "" else 0,
                        "tenant_pays": request.POST.getlist('tenant_pays') if 'tenant_pays' in request.POST else [],
                        "recent_updates": request.POST.getlist('recent_update') if 'recent_update' in request.POST else [],
                        "inclusions": request.POST.getlist('inclusions') if 'inclusions' in request.POST else [],
                        "security_features": request.POST.getlist('security_feature') if 'security_feature' in request.POST else [],
                        "building_class": request.POST.getlist('building_class') if 'building_class' in request.POST else [],
                        "cooling": request.POST.getlist('property_cooling') if 'property_cooling' in request.POST else [],
                        "county": request.POST['property_country'] if 'property_country' in request.POST and request.POST['property_country'] != "" else None,
                        "heating": request.POST.getlist('property_heating') if 'property_heating' in request.POST else [],
                        "zoning": request.POST.getlist('property_zoning') if 'property_zoning' in request.POST else [],
                        "electric": request.POST.getlist('property_electric') if 'property_electric' in request.POST else [],
                        "subdivision": request.POST['property_subdivision'] if 'property_subdivision' in request.POST and request.POST['property_subdivision'] != "" else None,
                        "gas": request.POST.getlist('property_gas') if 'property_gas' in request.POST else [],
                        "property_taxes": request.POST['property_taxes'] if 'property_taxes' in request.POST and request.POST['property_taxes'] != "" else None,
                        "special_assessment_tax": request.POST['special_assessment_taxes'] if 'special_assessment_taxes' in request.POST and request.POST['special_assessment_taxes'] != "" else None,
                        "water": request.POST.getlist('property_water') if 'property_water' in request.POST else [],
                        "tax_exemptions": request.POST.getlist('tax_exemption') if 'tax_exemption' in request.POST else [],
                        "sewer": request.POST.getlist('property_sewer') if 'property_sewer' in request.POST else [],
                        "hoa_fee": request.POST['hoa_fee'] if 'hoa_fee' in request.POST and request.POST['hoa_fee'] != "" else None,
                        "hoa_fee_type": request.POST['hoa_unit'] if 'hoa_unit' in request.POST and request.POST['hoa_unit'] != "" else None,
                        "hoa_amenities": request.POST.getlist('hoa_amenties') if 'hoa_amenties' in request.POST else [],
                        "total_rooms": request.POST['total_rooms'] if 'total_rooms' in request.POST and request.POST['total_rooms'] != "" else None,
                        "flooring": request.POST.getlist('property_flooring') if 'property_flooring' in request.POST else [],
                        "total_bedrooms": request.POST['total_bedrooms'] if 'total_bedrooms' in request.POST and request.POST['total_bedrooms'] != "" else None,
                        "appliances": request.POST.getlist('appliances') if 'appliances' in request.POST else [],
                        "total_bathrooms": request.POST['total_bathrooms'] if 'total_bathrooms' in request.POST and request.POST['total_bathrooms'] != "" else None,
                        "windows": request.POST.getlist('property_windows') if 'property_windows' in request.POST else [],
                        "total_public_restrooms": request.POST['total_public_rest_rooms'] if 'total_public_rest_rooms' in request.POST and request.POST['total_public_rest_rooms'] != "" else None,
                        "interior_features": request.POST.getlist('interior_features') if 'interior_features' in request.POST else [],
                        "ceiling_height": request.POST['ceiling_height'] if 'ceiling_height' in request.POST and request.POST['ceiling_height'] != "" else None,
                        "basement_features": request.POST.getlist('basement_features') if 'basement_features' in request.POST else [],
                        "handicap_amenities": request.POST.getlist('handicap_amenities') if 'handicap_amenities' in request.POST else [],
                        "construction": request.POST.getlist('property_construction') if 'property_construction' in request.POST else [],
                        "garage_parking": request.POST.getlist('garage_parking') if 'garage_parking' in request.POST else [],
                        "exterior_features": request.POST.getlist('exterior_features') if 'exterior_features' in request.POST else [],
                        "outbuildings": request.POST.getlist('out_buildings') if 'out_buildings' in request.POST else [],
                        "roof": request.POST.getlist('roofs') if 'roofs' in request.POST else [],
                        "location_features": request.POST.getlist('location_features') if 'location_features' in request.POST else [],
                        "foundation": request.POST.getlist('foundations') if 'foundations' in request.POST else [],
                        "road_frontage": request.POST.getlist('road_frontages') if 'road_frontages' in request.POST else [],
                        "fence": request.POST.getlist('fence') if 'fence' in request.POST else [],
                        "property_faces": request.POST.getlist('property_faces') if 'property_faces' in request.POST else [],
                        "pool": request.POST.getlist('pools') if 'pools' in request.POST else [],
                        "property_auction_data": property_auction_data,
                        "sale_terms": request.POST['term_of_sale'] if 'term_of_sale' in request.POST and request.POST['term_of_sale'] != "" else None,
                        'sale_by_type': request.POST['auction_type'] if 'auction_type' in request.POST and request.POST[
                            'auction_type'] != "" else None,
                        'is_featured': is_featured,
                        'status': request.POST['prop_listing_status'] if 'prop_listing_status' in request.POST and
                                                                         request.POST['prop_listing_status'] else "",
                        'property_opening_dates': open_house_dates,
                        'auction_location': request.POST['auction_location'] if 'auction_location' in request.POST and
                                                                         request.POST['auction_location'] else "",
                        'closing_status': request.POST['closing_status'] if 'closing_status' in request.POST and
                                                                            request.POST['closing_status'] else "",
                        'country': request.POST['prop_country'] if 'prop_country' in request.POST and request.POST['prop_country'] != "" else None,
                        'buyers_premium': int(request.POST['buyers_premium']) if 'buyers_premium' in request.POST and request.POST['buyers_premium'] != "" else None,
                        'buyers_premium_percentage': buyers_premium_percentage,
                        'buyers_premium_min_amount': buyers_premium_min_amount,
                    }
                if 'asset_type' in request.POST and int(request.POST['asset_type']) == 1:
                    property_param = {
                        "site_id": site_id,
                        "property_asset": request.POST['asset_type'] if 'asset_type' in request.POST and request.POST['asset_type'] != "" else None,
                        "step": request.POST['step'],
                        "user_id": user_id,
                        "property_id": property_id,
                        "description": request.POST['property_desc'] if 'property_desc' in request.POST and request.POST['property_desc'] != "" else None,
                        "address_one": request.POST['property_address_one'] if 'property_address_one' in request.POST and request.POST['property_address_one'] != "" else None,
                        "city": request.POST['property_city'] if 'property_city' in request.POST and request.POST['property_city'] != "" else None,
                        "state": request.POST['property_state'] if 'property_state' in request.POST and request.POST['property_state'] != "" else None,
                        "postal_code": request.POST['property_zip_code'] if 'property_zip_code' in request.POST and request.POST['property_zip_code'] != "" else None,
                        "property_type": request.POST['property_type'] if 'property_type' in request.POST and request.POST['property_type'] != "" else None,
                        "broker_co_op": request.POST['broker_co_op'] if 'broker_co_op' in request.POST and request.POST['broker_co_op'] != "" else 0,
                        "property_subtype": request.POST.getlist('property_sub_type') if 'property_sub_type' in request.POST else [],
                        "terms_accepted": request.POST.getlist('term_accepted') if 'term_accepted' in request.POST else [],
                        "total_acres": request.POST['total_land_acres'] if 'total_land_acres' in request.POST and request.POST['total_land_acres'] != "" else None,
                        "ownership": request.POST.getlist('ownership') if 'ownership' in request.POST else [],
                        "dryland_acres": request.POST['dry_land_acres'] if 'dry_land_acres' in request.POST and request.POST['dry_land_acres'] != "" else None,
                        "irrigated_acres": request.POST['irrigated_acres'] if 'irrigated_acres' in request.POST and request.POST['irrigated_acres'] != "" else None,
                        "possession": request.POST.getlist('possession') if 'possession' in request.POST else [],
                        "grass_acres": request.POST['grass_acres'] if 'grass_acres' in request.POST and request.POST['grass_acres'] != "" else None,
                        "pasture_fenced_acres": request.POST['pasture_fence_acres'] if 'pasture_fence_acres' in request.POST and request.POST['pasture_fence_acres'] != "" else None,
                        "lease_expiration": request.POST['lease_exp_date'] if 'lease_exp_date' in request.POST and request.POST['lease_exp_date'] != "" else None,
                        "crp_acres": request.POST['crp_acres'] if 'crp_acres' in request.POST and request.POST['crp_acres'] != "" else None,
                        "timber_acres": request.POST['timber_acres'] if 'timber_acres' in request.POST and request.POST['timber_acres'] != "" else None,
                        "mineral_rights": request.POST.getlist('mineral_rights') if 'mineral_rights' in request.POST else [],
                        "lot_acres": request.POST['lot_acres'] if 'lot_acres' in request.POST and request.POST['lot_acres'] != "" else None,
                        "balance_other_acres": request.POST['balence_other_acres'] if 'balence_other_acres' in request.POST and request.POST['balence_other_acres'] != "" else None,
                        "easements": request.POST.getlist('easements') if 'easements' in request.POST else [],
                        "survey": request.POST.getlist('survey') if 'survey' in request.POST else [],
                        "water": request.POST.getlist('property_water') if 'property_water' in request.POST else [],
                        "fsa_information": request.POST['fsa_info'] if 'fsa_info' in request.POST and request.POST['fsa_info'] != "" else None,
                        "utilities": request.POST.getlist('utilities') if 'utilities' in request.POST else [],
                        "sewer": request.POST.getlist('property_sewer') if 'property_sewer' in request.POST else [],
                        "crop_yield_history": request.POST['crop_yield_history'] if 'crop_yield_history' in request.POST and request.POST['crop_yield_history'] != "" else None,
                        "ponds": request.POST['num_ponds'] if 'num_ponds' in request.POST and request.POST['num_ponds'] != "" else None,
                        "wells": request.POST['num_wells'] if 'num_wells' in request.POST and request.POST['num_wells'] != "" else None,
                        "soil_productivity_rating": request.POST['soil_productivity_rating'] if 'soil_productivity_rating' in request.POST and request.POST['soil_productivity_rating'] != "" else None,
                        "improvements": request.POST.getlist('improvements') if 'improvements' in request.POST else [],
                        "road_frontage": request.POST.getlist('road_frontages') if 'road_frontages' in request.POST else [],
                        "outbuildings": request.POST.getlist('out_buildings') if 'out_buildings' in request.POST else [],
                        "topography": request.POST.getlist('topography') if 'topography' in request.POST else [],
                        "other_features": request.POST.getlist('other_features_land') if 'other_features_land' in request.POST else [],
                        "wildlife": request.POST.getlist('wildlife') if 'wildlife' in request.POST else [],
                        "fence": request.POST.getlist('fence') if 'fence' in request.POST else [],
                        "fish": request.POST.getlist('fish') if 'fish' in request.POST else [],
                        "irrigation_system": request.POST.getlist('irrigation_system') if 'irrigation_system' in request.POST else [],
                        "recreation": request.POST.getlist('recreation') if 'recreation' in request.POST else [],
                        "livestock_carrying_capacity": request.POST['livestock_carrying_capacity'] if 'livestock_carrying_capacity' in request.POST and request.POST['livestock_carrying_capacity'] != "" else None,
                        "annual_payment": request.POST['crop_retirement_program_payment'] if 'crop_retirement_program_payment' in request.POST and request.POST['crop_retirement_program_payment'] != "" else None,
                        "contract_expire": request.POST['crp_exp_date'] if 'crp_exp_date' in request.POST and request.POST['crp_exp_date'] != "" else None,
                        "county": request.POST['property_country'] if 'property_country' in request.POST and request.POST['property_country'] != "" else None,
                        "property_taxes": request.POST['property_taxes'] if 'property_taxes' in request.POST and request.POST['property_taxes'] != "" else None,
                        "special_assessment_tax": request.POST['special_assessment_taxes'] if 'special_assessment_taxes' in request.POST and request.POST['special_assessment_taxes'] != "" else None,
                        "zoning": request.POST.getlist('property_zoning') if 'property_zoning' in request.POST else [],
                        "tax_exemptions": request.POST.getlist('tax_exemption') if 'tax_exemption' in request.POST else [],
                        "subdivision": request.POST['property_subdivision'] if 'property_subdivision' in request.POST and request.POST['property_subdivision'] != "" else None,
                        "property_auction_data": property_auction_data,
                        "sale_terms": request.POST['term_of_sale'] if 'term_of_sale' in request.POST and request.POST['term_of_sale'] != "" else None,
                        'sale_by_type': request.POST['auction_type'] if 'auction_type' in request.POST and request.POST[
                            'auction_type'] != "" else None,
                        'is_featured': is_featured,
                        'status': request.POST['prop_listing_status'] if 'prop_listing_status' in request.POST and
                                                                         request.POST['prop_listing_status'] else "",
                        'auction_location': request.POST['auction_location'] if 'auction_location' in request.POST and
                                                                         request.POST['auction_location'] else "",
                        'closing_status': request.POST['closing_status'] if 'closing_status' in request.POST and
                                                                            request.POST['closing_status'] else "",
                        'country': request.POST['prop_country'] if 'prop_country' in request.POST and request.POST['prop_country'] != "" else None,
                        'buyers_premium': int(request.POST['buyers_premium']) if 'buyers_premium' in request.POST and request.POST['buyers_premium'] != "" else None,
                        'buyers_premium_percentage': buyers_premium_percentage,
                        'buyers_premium_min_amount': buyers_premium_min_amount,

                    }
                # ------------------Deposit listings-----------------
                if int(auction_type) in [1, 2]:
                    property_param['deposit_amount'] = deposit_amount if is_deposit_required else 0
                    property_param['is_deposit_required'] = is_deposit_required
                else:
                    property_param['deposit_amount'] = 0 
                    property_param['is_deposit_required'] = 0

                if int(auction_type) == 7:
                    property_param['due_diligence_period'] = request.POST['due_diligence_period'] if 'due_diligence_period' in request.POST and request.POST['due_diligence_period'] != "" else None
                    property_param['escrow_period'] = request.POST['escrow_period'] if 'escrow_period' in request.POST and request.POST['escrow_period'] != "" else None
                    property_param['earnest_deposit'] = earnest_deposit if earnest_deposit else None
                    property_param['earnest_deposit_type'] = earnest_deposit_type
                    property_param['highest_best_format'] = highest_best_format
                    property_param['un_priced'] = 1 if 'offer_unpriced' in request.POST and request.POST['offer_unpriced'] else 0
                    property_param['required_all'] = 1 if 'required_all' in request.POST and request.POST['required_all'] else 0



            elif int(request.POST['step']) == 2:
                property_param = {
                    'site_id': site_id,
                    'user_id': user_id,
                    'property_id': request.POST['property_id'],
                    'step': request.POST['step'],
                    'is_map_view': request.POST['enable_map_view'] if 'enable_map_view' in request.POST else 0,
                    'is_street_view': request.POST['enable_street_view'] if 'enable_street_view' in request.POST else 0,
                    'is_arial_view': request.POST['enable_arial_view'] if 'enable_arial_view' in request.POST else 0,
                    'map_url': request.POST['map_url'] if 'map_url' in request.POST and request.POST['map_url']  else "",
                }
                property_param['latitude'] = request.POST['latitude'] if 'latitude' in request.POST and \
                                                                           request.POST['latitude'] else ""
                property_param['longitude'] = request.POST['longitude'] if 'longitude' in request.POST and \
                                                                            request.POST['longitude'] else ""

            elif int(request.POST['step']) == 3:
                property_image_list = []
                if request.POST['property_image_id'] != "":
                    property_image_list = request.POST['property_image_id'].split(',')

                property_video_list = []
                if 'property_video_id' in request.POST and request.POST['property_video_id'] != "":
                    property_video_list = request.POST['property_video_id'].split(',')

                property_param = {
                    'site_id': site_id,
                    'user_id': user_id,
                    'property_id': request.POST['property_id'],
                    'step': request.POST['step'],
                    'property_pic': property_image_list,
                    'property_video': property_video_list,
                }
            elif int(request.POST['step']) == 4:
                property_doc_list = []
                if request.POST['property_doc_id'] != "":
                    property_doc_list = request.POST['property_doc_id'].split(',')
                property_param = {
                    'site_id': site_id,
                    'user_id': user_id,
                    'property_id': request.POST['property_id'],
                    'step': request.POST['step'],
                    'property_documents': property_doc_list,
                }
            property_url = settings.API_URL + '/api-property/add-property/'
            # print(property_param)
            property_data = call_api_post_method(property_param, property_url, token)
            if 'error' in property_data and property_data['error'] == 0:

                if property_id and int(is_reset_offer) == 1 and int(step) == 1 and int(auction_type) == 4:

                    try:
                        clear_offer_param = {
                            "domain_id": site_id,
                            "property_id": property_id,
                            "user_id": user_id
                        }
                        clear_offer_url = settings.API_URL + '/api-bid/clear-offer/'

                        clear_offer_data = call_api_post_method(clear_offer_param, clear_offer_url, token)
                        # if 'error' in clear_offer_data and clear_offer_data['error'] == 0:
                        #     print(clear_offer_data)
                    except Exception as exp:
                        print(exp)
                data = {'status': 200, 'data': property_data, 'error': 0, "next_url": next_url, "msg": "Property saved successfully"}

            else:
                data = {'status': 403, 'data': property_data, 'error': 1, "next_url": "", "msg": "Some error occurs, please try again"}

        else:
            data = {'status': 403, 'data': {}, 'error': 1, "next_url": "", "msg": "Some error occurs, please try again"}

        return JsonResponse(data)
    except Exception as exp:
        print(exp)
        data = {'status': 403, 'data': {}, 'error': 1, "next_url": "", "msg": "Some error occurs, please try again"}
        return JsonResponse(data)

@csrf_exempt
def save_property_video(request):
    try:
        if request.is_ajax() and request.method == 'POST':
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']
            token = request.session['token']['access_token']
            user_id = request.session['user_id']
            new_video_url = ''
            video_params = {}
            if 'video_url' in request.POST and request.POST['video_url']:
                video_id = parse_qs(urlparse(request.POST['video_url']).query)['v'][0]
                new_video_url = 'https://www.youtube.com/embed/'+video_id
                video_params = {
                    'site_id': site_id,
                    'user_id': user_id,
                    'video_url': new_video_url
                }

            video_url = settings.API_URL + '/api-property/subdomain-add-property-video/'

            video_data = call_api_post_method(video_params, video_url, token)
            if 'error' in video_data and video_data['error'] == 0:
                data = {'status': 200, 'data': video_data['data'], 'error': 0, 'msg': 'Video saved successfully', 'video_url': new_video_url}
            else:
                data = {'status': 403, 'data': {}, 'error': 1, 'msg': 'Some error occur, please try again.', 'video_url': ''}
        else:
            data = {'status': 403, 'data': {}, 'error': 1, 'msg': 'Some error occur, please try again.', 'video_url': ''}

        return JsonResponse(data)
    except Exception as exp:
        print(exp)
        data = {'status': 403, 'data': {}, 'error': 1, 'msg': 'Some error occur, please try again.'}
        return JsonResponse(data)

@csrf_exempt
def get_property_types(request):
    try:
        if request.is_ajax() and request.method == 'POST':
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']
            token = request.session['token']['access_token']
            user_id = request.session['user_id']
            property_type_params = {
                'asset_id': request.POST['asset_id']
            }

            property_type_url = settings.API_URL + '/api-property/property-type-listing/'

            property_type_data = call_api_post_method(property_type_params, property_type_url, token)


            if 'error' in property_type_data and property_type_data['error'] == 0:
                property_type_listing = property_type_data['data']
                data = {'status': 200, 'property_type_listing': property_type_listing, 'error': 0}
            else:
                data = {'status': 403, 'data': {}, 'error': 1, 'msg': 'Some error occur, please try again.', 'property_type_listing': []}
        else:
            data = {'status': 403, 'data': {}, 'error': 1, 'msg': 'Some error occur, please try again.', 'property_type_listing': []}

        return JsonResponse(data)
    except Exception as exp:
        print(exp)
        data = {'status': 403, 'data': {}, 'error': 1, 'msg': 'Some error occur, please try again.', 'property_type_listing': []}
        return JsonResponse(data)

@csrf_exempt
def add_property_features(request):
    try:
        if request.is_ajax() and request.method == 'POST':
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']
            token = request.session['token']['access_token']
            user_id = request.session['user_id']
            feature_name = request.POST['feature_name']
            feature_type = request.POST['feature_type']
            feature_type_name = get_feature_type(request.POST['feature_type'])
            add_feature_params = {
                'asset_id': request.POST['asset_id'],
                'feature_type': feature_type_name,
                'name': feature_name,
            }

            add_feature_url = settings.API_URL + '/api-property/add-property-features/'

            feature_data = call_api_post_method(add_feature_params, add_feature_url, token)

            if 'error' in feature_data and feature_data['error'] == 0:

                feature_id = feature_data['data']['feature_id']
                feature_name = feature_name
                data = {'status': 200, 'feature_id': feature_id, 'feature_name': feature_name, 'error': 0, 'section': feature_type}
            else:
                data = {'status': 403, 'feature_id': '', 'feature_name': '', 'error': 1, 'section': feature_type}
        else:
            data = {'status': 403, 'feature_id': '', 'feature_name': '', 'error': 1, 'section': ''}

        return JsonResponse(data)
    except Exception as exp:
        print(exp)
        data = {'status': 403, 'feature_id': '', 'feature_name': '', 'error': 1, 'section': ''}
        return JsonResponse(data)


@csrf_exempt
def change_property_status(request):
    try:
        if request.is_ajax() and request.method == 'POST':
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']
            token = request.session['token']['access_token']
            user_id = request.session['user_id']
            status_id = request.POST['status_id']
            status_name = request.POST['status_name']
            property_id = request.POST['property_id']
            change_status_params = {
                "site_id": site_id,
                "user_id": user_id,
                "property_id": request.POST['property_id'],
                "status": status_id
            }

            change_status_url = settings.API_URL + '/api-property/property-status-change/'

            change_status_data = call_api_post_method(change_status_params, change_status_url, token)

            if 'error' in change_status_data and change_status_data['error'] == 0:
                data = {'status': 200, 'status_id': status_id, 'status_name': status_name, 'property_id': property_id, 'error': 0, 'msg': 'Status changed successfully.'}
            else:
                data = {'status': 403, 'status_id': status_id, 'status_name': status_name, 'property_id': property_id, 'error': 1, 'msg': 'Some error occurs, please try again'}
        else:
            data = {'status': 403, 'status_id': '', 'status_name': '', 'error': 1, 'msg': 'Some error occurs, please try again'}

        return JsonResponse(data)
    except Exception as exp:
        print(exp)
        data = {'status': 403, 'status_id': '', 'status_name': '', 'error': 1, 'msg': 'Some error occurs, please try again'}
        return JsonResponse(data)

@csrf_exempt
def change_approval_status(request):
    try:
        if request.is_ajax() and request.method == 'POST':
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']
            token = request.session['token']['access_token']
            user_id = request.session['user_id']
            approval_id = request.POST['approval_id']
            approval_name = request.POST['approval_name']
            property_id = request.POST['property_id']
            change_status_params = {
                "site_id": site_id,
                "user_id": user_id,
                "property_id": request.POST['property_id'],
                "is_approved": approval_id
            }

            change_status_url = settings.API_URL + '/api-property/property-approval-change/'

            change_status_data = call_api_post_method(change_status_params, change_status_url, token)
            if 'error' in change_status_data and change_status_data['error'] == 0:
                data = {'status': 200, 'approval_id': approval_id, 'approval_name': approval_name, 'property_id': property_id, 'error': 0, 'msg': 'Approval changed successfully.'}
            else:
                data = {'status': 403, 'approval_id': approval_id, 'approval_name': approval_name, 'property_id': property_id, 'error': 1, 'msg': 'Some error occurs, please try again'}
        else:
            data = {'status': 403, 'approval_id': '', 'approval_name': '', 'error': 1, 'msg': 'Some error occurs, please try again'}

        return JsonResponse(data)
    except Exception as exp:
        print(exp)
        data = {'status': 403, 'approval_id': '', 'approval_name': '', 'error': 1, 'msg': 'Some error occurs, please try again'}
        return JsonResponse(data)

@csrf_exempt
def get_user_personal_info(request):
    try:
        if request.is_ajax() and request.method == 'POST':
            user_id = request.session['user_id']
            token = request.session['token']['access_token']
            param = {
                'user_id': user_id
            }
            url = settings.API_URL + '/api-users/get-personal-info/'
            user_personal_info_data = call_api_post_method(param, url, token)

            if 'error' in user_personal_info_data and user_personal_info_data['error'] == 0:
                user_personal_info = user_personal_info_data['data']
                data = {'status': 200, 'user_personal_info': user_personal_info, 'msg': '', 'error': 0}
            else:
                data = {'status': 403, 'user_personal_info': {}, 'msg': 'Some Error Occurs, please try again', 'error': 1}

        else:
            data = {'status': 403, 'msg': 'Forbidden'}

        return JsonResponse(data)
    except Exception as exp:
        print(exp)
        data = {'status': 403, 'msg': 'invalid request.'}
        return JsonResponse(data)

@csrf_exempt
def get_icons_by_type(request):
    try:
        if request.is_ajax() and request.method == 'POST':
            icon_type_id = request.POST['icon_type_id']
            user_id = request.session['user_id']
            token = request.session['token']['access_token']
            position = request.POST['position']
            try:
                params = {"icon_type": icon_type_id}
                url = settings.API_URL + '/api-users/get-expertise-icon/'
                data = call_api_post_method(params, url, token)
                icon_list = data['data']
            except:
                icon_list = []

            data = {'error': 0, 'status': 200, 'icon_list': icon_list, 'position': position}
        else:
            data = {'error': 0, 'status': 403, 'msg': 'Forbidden', 'position': ''}

        return JsonResponse(data)
    except Exception as exp:
        print(exp)
        data = {'status': 403, 'msg': 'invalid request.'}
        return JsonResponse(data)

@csrf_exempt
def get_address_detail_by_zipcode(request):
    try:
        if request.is_ajax() and request.method == 'POST':
            zip_code = request.POST['zip_code']
            try:
                zcdb = ZipCodeDatabase()
                address = zcdb[zip_code]
                data = {'error': 0, 'status': 200, 'state': address.state, 'city': address.city, 'zip_code': zip_code}
            except:
                data = {'error': 1, 'status': 403, 'state': '', 'city': '', 'zip_code': zip_code}
        else:
            data = {'error': 0, 'status': 403, 'msg': 'Forbidden', 'state': '', 'city': ''}

        return JsonResponse(data)
    except Exception as exp:
        print(exp)
        data = {'status': 403, 'msg': 'invalid request.'}
        return JsonResponse(data)


@csrf_exempt
def contact_listing(request):
    try:
        is_permission = check_permission(request, 12)
        if not is_permission:
            http_host = request.META['HTTP_HOST']
            redirect_url = settings.URL_SCHEME + str(http_host)
            return HttpResponseRedirect(redirect_url)
        site_detail = subdomain_site_details(request)
        site_id = site_detail['site_detail']['site_id']
        token = request.session['token']['access_token']
        user_id = request.session['user_id']
        page_size = 10
        page = 1
        if request.is_ajax() and request.method == 'POST':
            search = ''
            if 'search' in request.POST and request.POST['search']:
                search = request.POST['search']
            page = 1
            if 'page' in request.POST and request.POST['page'] != "":
                page = request.POST['page']

            if 'page_size' in request.POST and request.POST['page_size'] != "":
                page_size = request.POST['page_size']

            user_type = ''
            if 'user_type' in request.POST and request.POST['user_type'] != "":
                user_type = request.POST['user_type']
            params = {
                'site_id': site_id,
                'user_id': user_id,
                'page_size': page_size,
                'search': search,
                'page': page,
                'user_type': user_type,
            }
            sno = (int(page) - 1) * int(page_size) + 1
            contact_url = settings.API_URL + '/api-users/admin-contact-us-listing/'
            contact_data = call_api_post_method(params, contact_url, token=token)

            if 'error' in contact_data and contact_data['error'] == 0:
                contact_list = contact_data['data']['data']
                contact_total = contact_data['data']['total']
            else:
                contact_list = []
                contact_total = 0
            context = {'contact_list': contact_list, 'contact_total': contact_total, "active_menu": "contact_listing", "sno": sno}
            contact_data_path = 'admin/dashboard/contacts/contact-list.html'
            contact_template = get_template(contact_data_path)
            contact_listing_html = contact_template.render(context)
            # ---------------Pagination--------
            pagination_path = 'admin/dashboard/contacts/pagination.html'
            pagination_template = get_template(pagination_path)
            total_page = math.ceil(int(contact_total) / int(page_size))
            pagination_data = {"no_page": int(total_page), "total_page": range(total_page), "current_page": int(page)}
            pagination_html = pagination_template.render(pagination_data)
            data = {'contact_listing_html': contact_listing_html, 'status': 200, 'msg': '', 'error': 0, 'total': contact_total, "pagination_html": pagination_html}
            return JsonResponse(data)
        else:
            user_type = request.GET.get('user_type', '')
            params = {
                'site_id': site_id,
                'user_id': user_id,
                'page': page,
                'page_size': page_size,
                'search': '',
                'user_type': user_type,
            }
            sno = (int(page) - 1) * int(page_size) + 1
            contact_url = settings.API_URL + '/api-users/admin-contact-us-listing/'
            contact_data = call_api_post_method(params, contact_url, token=token)
            if 'error' in contact_data and contact_data['error'] == 0:
                contact_list = contact_data['data']['data']
                contact_total = contact_data['data']['total']
            else:
                contact_list = []
                contact_total = 0
            # ---------------Pagination--------
            pagination_path = 'admin/dashboard/contacts/pagination.html'
            pagination_template = get_template(pagination_path)
            total_page = math.ceil(contact_total/page_size)
            pagination_data = {"no_page": total_page, "total_page": range(total_page), "current_page": 1}
            pagination_html = pagination_template.render(pagination_data)

            context = {'contact_list': contact_list, 'contact_total': contact_total, "active_menu": "pages", "active_submenu": "contact_listing", "pagination_html": pagination_html, "sno": sno, "user_type": user_type}
            return render(request, "admin/dashboard/contacts/view-contact.html", context)
    except Exception as exp:
        print(exp)
        return HttpResponse("Issue in views")


@csrf_exempt
def contact_detail(request):
    try:
        site_detail = subdomain_site_details(request)
        site_id = site_detail['site_detail']['site_id']
        token = request.session['token']['access_token']
        user_id = request.session['user_id']
        if request.is_ajax() and request.method == 'POST':
            contact_id = ''
            if 'contact_id' in request.POST and request.POST['contact_id']:
                contact_id = request.POST['contact_id']
            params = {
                'site_id': site_id,
                'user_id': user_id,
                'contact_id': contact_id
            }
            contact_url = settings.API_URL + '/api-users/admin-contact-us-detail/'
            contact_data = call_api_post_method(params, contact_url, token=token)
            if 'error' in contact_data and contact_data['error'] == 0:
                data = contact_data['data']
                data["error"] = 0
                data["msg"] = contact_data["msg"]
            else:
                data = {"error": 1, "msg": contact_data["msg"]}
            return JsonResponse(data)
    except Exception as exp:
        return HttpResponse("Issue in views")

@csrf_exempt
def listing_search_suggestion(request):
    try:
        if request.is_ajax() and request.method == 'POST':
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']
            token = request.session['token']['access_token']

            params = {
                'site_id': site_id,
                'search': request.POST['search']
            }

            api_url = settings.API_URL + '/api-property/subdomain-property-suggestion/'

            suggestion_data = call_api_post_method(params, api_url, token)
            if 'error' in suggestion_data and suggestion_data['error'] == 0:
                data = {'status': 200, 'suggestion_list': suggestion_data['data'], 'error': 0}
            else:
                data = {'status': 403, 'suggestion_list': [], 'error': 1}
        else:
            data = {'status': 403, 'suggestion_list': [], 'error': 1}

        return JsonResponse(data)
    except Exception as exp:
        print(exp)
        data = {'status': 403, 'suggestion_list': [], 'error': 1}
        return JsonResponse(data)

@csrf_exempt
def contact_search_suggestion(request):
    try:
        if request.is_ajax() and request.method == 'POST':
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']
            token = request.session['token']['access_token']

            params = {
                'site_id': site_id,
                'search': request.POST['search']
            }

            api_url = settings.API_URL + '/api-cms/subdomain-contact-suggestion/'

            suggestion_data = call_api_post_method(params, api_url, token)
            if 'error' in suggestion_data and suggestion_data['error'] == 0:
                data = {'status': 200, 'suggestion_list': suggestion_data['data'], 'error': 0}
            else:
                data = {'status': 403, 'suggestion_list': [], 'error': 1}
        else:
            data = {'status': 403, 'suggestion_list': [], 'error': 1}

        return JsonResponse(data)
    except Exception as exp:
        print(exp)
        data = {'status': 403, 'suggestion_list': [], 'error': 1}
        return JsonResponse(data)

@csrf_exempt
def blog_search_suggestion(request):
    try:
        if request.is_ajax() and request.method == 'POST':
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']
            token = request.session['token']['access_token']

            params = {
                'site_id': site_id,
                'search': request.POST['search']
            }

            api_url = settings.API_URL + '/api-users/subdomain-article-suggestion/'

            suggestion_data = call_api_post_method(params, api_url, token)
            if 'error' in suggestion_data and suggestion_data['error'] == 0:
                data = {'status': 200, 'suggestion_list': suggestion_data['data'], 'error': 0}
            else:
                data = {'status': 403, 'suggestion_list': [], 'error': 1}
        else:
            data = {'status': 403, 'suggestion_list': [], 'error': 1}

        return JsonResponse(data)
    except Exception as exp:
        print(exp)
        data = {'status': 403, 'suggestion_list': [], 'error': 1}
        return JsonResponse(data)

@csrf_exempt
def delete_property(request):
    try:
        if request.is_ajax() and request.method == 'POST':
            page = 1
            page_size = 10
            search = ''
            if 'search' in request.POST and request.POST['search']:
                search = request.POST['search']

            page = 1
            if 'page' in request.POST and request.POST['page'] != "":
                page = int(request.POST['page'])

            page_size = 10
            if 'perpage' in request.POST and request.POST['perpage']:
                page_size = int(request.POST['perpage'])

            asset_type = ''
            if 'asset_type' in request.POST and request.POST['asset_type']:
                asset_type = int(request.POST['asset_type'])
            auction_type = ''
            if 'auction_type' in request.POST and request.POST['auction_type']:
                auction_type = int(request.POST['auction_type'])
            property_type = ''
            if 'property_type' in request.POST and request.POST['property_type']:
                property_type = int(request.POST['property_type'])

            status = ""
            if 'status' in request.POST and request.POST['status']:
                status = request.POST['status']


            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']
            token = request.session['token']['access_token']
            user_id = request.session['user_id']
            is_broker = request.session['is_broker']
            params = {
                'site_id': site_id,
                'user_id': request.session['user_id'],
                'property_id': request.POST['property_id'],
            }


            url = settings.API_URL + '/api-property/delete-property/'
            data = call_api_post_method(params, url, token)
            user_list = []
            if 'error' in data and data['error'] == 0:
                try:
                    status_param = {}
                    status_url = settings.API_URL + '/api-settings/lookup-status-listing/'
                    status_data = call_api_post_method(status_param, status_url, token)
                    status_list = status_data['data']
                except:
                    status_list = []
                list_param = {
                    "page": page,
                    "page_size": page_size,
                    "site_id": site_id,
                    "user_id": user_id,
                    "auction_id": auction_type,
                    "asset_id": asset_type,
                    "property_type": property_type,
                    "search": search,
                    "status": status
                }

                list_url = settings.API_URL + '/api-property/property-listing/'

                list_data = call_api_post_method(list_param, list_url, token)
                
                if 'error' in list_data and list_data['error'] == 0:
                    property_listing = list_data['data']['data']
                    total = list_data['data']['total'] if 'total' in list_data['data'] else 0
                else:
                    property_listing = []
                    total = 0
                sno = (int(page) - 1) * int(page_size) + 1
                context = {'property_list': property_listing, 'total': total, "aws_url": settings.AWS_URL,
                           "is_broker": is_broker, 'status_list': status_list, 'sno': sno}

                property_listing_path = 'admin/dashboard/listings/property_listing_content.html'
                property_listing_template = get_template(property_listing_path)
                property_listing_html = property_listing_template.render(context)
                # ---------------Pagination--------
                pagination_path = 'admin/dashboard/listings/property-pagination.html'
                pagination_template = get_template(pagination_path)
                total_page = math.ceil(int(total) / int(page_size))
                pagination_html = ''
                if total_page > 1:
                    pagination_data = {"no_page": int(total_page), "total_page": range(total_page),
                                       "current_page": int(page),
                                       "pagination_id": "prop_listing_pagination_list"}
                    pagination_html = pagination_template.render(pagination_data)
                data = {'error': 0, 'status': 200, 'msg': "deleted successfully", "property_listing_html": property_listing_html, "pagination_html": pagination_html}
            else:
                data = {'error': 1, 'status': 403, 'msg': 'Server error, Please try again'}
        else:
            data = {'error': 1, 'status': 403, 'msg': 'Forbidden'}

        return JsonResponse(data)
    except Exception as exp:
        print(exp)
        data = {'status': 403, 'msg': 'invalid request.', 'user_list': []}
        return JsonResponse(data)

@csrf_exempt
def save_listing_settings(request):
    try:
        if request.is_ajax() and request.method == 'POST':
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']
            token = request.session['token']['access_token']
            user_id = request.session['user_id']
            params = {
                "site_id": site_id,
                "user_id": user_id,
                "property_id": request.POST['property_id'] if 'property_id' in request.POST and request.POST['property_id'] != "" else "",
                "time_flash": request.POST['timer_flash'],
                "auto_approval": request.POST['auto_approval'],
                "is_deposit_required": request.POST['is_deposit_required'],
                "show_reverse_not_met": request.POST['reserve_not_met'],
                "is_log_time_extension": request.POST['is_log_time_extension'],
                "log_time_extension": request.POST['log_time_extension'],
                "remain_time_to_add_extension": request.POST['remain_time_to_add_extension']
            }
            if int(request.POST['auto_approval']) == 1:
                params['bid_limit'] = request.POST['bid_limit']

            if int(request.POST['is_deposit_required']) == 1:
                params['deposit_amount'] = request.POST['listing_deposit_amount']    

            if 'property_id' in request.POST and request.POST['property_id'] != "":
                url = settings.API_URL + '/api-property/save-property-setting/'
            else:
                url = settings.API_URL + '/api-property/save-global-property-setting/'
            data = call_api_post_method(params, url, token)
            if 'error' in data and data['error'] == 0:
                response = data['data']
                data = {"error": 0, "status": 200, "msg": "Saved successfully", "data": response}
            else:
                data = {"error": 1, "status": 403, "msg": data['msg'], "data": {}}
        else:
            data = {"error": 1, "status": 403, "msg": "Forbidden", "data": {}}

        return JsonResponse(data)
    except Exception as exp:
        print(exp)
        data = {'status': 403, 'msg': 'invalid request.', 'user_list': []}
        return JsonResponse(data)


@csrf_exempt
def get_listing_settings(request):
    try:
        if request.is_ajax() and request.method == 'POST':
            try:
                site_detail = subdomain_site_details(request)
                site_id = site_detail['site_detail']['site_id']
                token = request.session['token']['access_token']
                user_id = request.session['user_id']
            except:
                settings_value = {}
                user_id = ''
                site_id = ''
            property_id = request.POST['property_id']
            params = {"site_id": site_id, 'property_id': request.POST['property_id'], 'is_global': 0}

            url = settings.API_URL + '/api-property/get-property-setting/'
            data = call_api_post_method(params, url, token)

            if 'error' in data and data['error'] == 0:
                listing_setting = data['data']
                data = {'error': 0, 'status': 200, 'msg': "", 'listing_setting': listing_setting, 'property_id': property_id}
            else:
                data = {'error': 1, 'status': 403, 'msg': 'Server error, Please try again', 'listing_setting': {}, 'property_id': property_id}
        else:
            data = {'error': 1, 'status': 403, 'msg': 'Forbidden', 'listing_setting': {}}

        return JsonResponse(data)
    except Exception as exp:
        print(exp)
        data = {'status': 403, 'msg': 'invalid request.', 'user_list': []}
        return JsonResponse(data)


@csrf_exempt
def article_search_suggestion(request):
    try:
        if request.is_ajax() and request.method == 'POST':
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']
            token = request.session['token']['access_token']

            user_params = {
                'site_id': site_id,
                'search': request.POST['search']
            }

            user_url = settings.API_URL + '/api-users/subdomain-article-suggestion/'

            suggestion_data = call_api_post_method(user_params, user_url, token)
            if 'error' in suggestion_data and suggestion_data['error'] == 0:
                data = {'status': 200, 'suggestion_list': suggestion_data['data'], 'error': 0}
            else:
                data = {'status': 403, 'suggestion_list': [], 'error': 1}
        else:
            data = {'status': 403, 'suggestion_list': [], 'error': 1}

        return JsonResponse(data)
    except Exception as exp:
        print(exp)
        data = {'status': 403, 'suggestion_list': [], 'error': 1}
        return JsonResponse(data)


@csrf_exempt
def chat_list(request):
    try:
        site_detail = subdomain_site_details(request)
        site_id = site_detail['site_detail']['site_id']
        context = {
            "active_menu": "chat",
            "site_id": site_id
        }
        return render(request, 'admin/dashboard/chat/chat.html', context)
    except Exception as exp:
        print(exp)
        return HttpResponse("Issue in views")

# @csrf_exempt
# def chat_list(request):
#     try:

#         site_detail = subdomain_site_details(request)
#         site_id = site_detail['site_detail']['site_id']

#         token = request.session['token']['access_token']
#         user_id = request.session['user_id']

#         page_size = 10
#         is_broker = request.session['is_broker']
#         if request.is_ajax() and request.method == 'POST':
#             page = 1
#             if 'page' in request.POST and request.POST['page'] != "":
#                 page = request.POST['page']

#             if 'page_size' in request.POST and request.POST['page_size'] != "":
#                 page_size = request.POST['page_size']

#             last_msg_id = ''
#             if 'last_msg_id' in request.POST and request.POST['last_msg_id'] != "":
#                 last_msg_id = request.POST['last_msg_id']

#             filter_chat = ''
#             if 'filter_chat' in request.POST and request.POST['filter_chat'] != "":
#                 filter_chat = request.POST['filter_chat']

#             params = {
#                 "site_id": site_id,
#                 "user_id": user_id,
#                 "page": page,
#                 "page_size": page_size,
#                 "last_msg_id": last_msg_id,
#                 "filter_data": filter_chat,
#             }
#             if 'msg_type' in request.POST and request.POST['msg_type'] != "":
#                 msg_type = request.POST['msg_type']
#                 params['msg_type'] = msg_type

#             api_url = settings.API_URL + '/api-contact/subdomain-chat-master-listing/'
#             if is_broker:
#                 api_url = settings.API_URL + '/api-contact/broker-chat-master-listing/'


#             chat_data = call_api_post_method(params, api_url, token=token)

#             if 'error' in chat_data and chat_data['error'] == 0:

#                 chat_listing = chat_data['data']['data']
#                 try:
#                     length = len(chat_listing)
#                     first_master_id = chat_listing[0]['id']
#                     last_master_id = chat_listing[length-1]['id']
#                 except:
#                     first_master_id = ''
#                     last_master_id = ''

#                 total = chat_data['data']['total']
#                 page = chat_data['data']['page']
#                 page_size = chat_data['data']['page_size']
#             else:
#                 chat_listing = []
#                 total = 0
#                 page = page
#                 page_size = page_size
#                 first_master_id = ''
#                 last_master_id = ''

#             context = {'chat_listing': chat_listing, 'total': total, "aws_url": settings.AWS_URL, 'is_broker': is_broker, 'sess_user_id': user_id}

#             chat_listing_path = 'admin/dashboard/chat/chat-listing.html'
#             chat_listing_template = get_template(chat_listing_path)
#             chat_listing_html = chat_listing_template.render(context)


#             data = {'chat_listing_html': chat_listing_html, 'status': 200, 'msg': '', 'error': 0, 'total': total, "page": page, "page_size": page_size, "last_master_id": last_master_id, "first_master_id": first_master_id}
#             return JsonResponse(data)

#         else:
#             try:
#                 params = {"site_id": site_id, "user_id": user_id, "page": 1, "page_size": page_size}
#                 api_url = settings.API_URL + '/api-contact/subdomain-chat-master-listing/'
#                 if is_broker:
#                     api_url = settings.API_URL + '/api-contact/broker-chat-master-listing/'


#                 chat_data = call_api_post_method(params, api_url, token=token)

#                 if 'error' in chat_data and chat_data['error'] == 0:
#                     chat_listing = chat_data['data']['data']
#                     length = len(chat_listing)
#                     try:
#                         first_master_id = chat_listing[0]['id']
#                         last_master_id = chat_listing[length-1]['id']
#                     except:
#                         first_master_id = ''
#                         last_master_id = ''
#                     total = chat_data['data']['total']
#                     page = chat_data['data']['page']
#                     page_size = chat_data['data']['page_size']

#                 else:
#                     chat_listing = []
#                     total = 0
#                     page = 1
#                     first_master_id = ''
#                     last_master_id = ''
#                     page_size = page_size
#             except:
#                 chat_listing = []
#                 total = 0
#                 page = 1
#                 page_size = page_size
#                 first_master_id = ''
#                 last_master_id = ''

#             context = {
#                 "active_menu": "chat",
#                 "chat_listing": chat_listing,
#                 "total": total,
#                 "page": page,
#                 "page_size": page_size,
#                 "last_master_id": last_master_id,
#                 "first_master_id": first_master_id
#             }
#             return render(request, 'admin/dashboard/chat/chat.html', context)
#     except Exception as exp:
#         print(exp)
#         return HttpResponse("Issue in views")

# @csrf_exempt
# def chat_history(request):
#     try:
#         if request.is_ajax() and request.method == 'POST':

#             site_detail = subdomain_site_details(request)
#             site_id = site_detail['site_detail']['site_id']

#             user_id = None
#             token = None
#             page = 1
#             page_size = 10
#             is_broker = False
#             if 'user_id' in request.session:
#                 user_id = request.session['user_id']
#                 token = request.session['token']['access_token']
#                 is_broker = request.session['is_broker']


#             master_id = request.POST['master_id']
#             post_last_msg = request.POST['last_msg_id']
#             params = {
#                 "site_id": site_id,
#                 "user_id": user_id,
#                 "master_id": master_id,
#                 "page": page,
#                 "page_size": page_size,
#                 "last_msg_id": request.POST['last_msg_id'],
#             }
#             if 'msg_type' in request.POST and request.POST['msg_type'] != "":
#                 msg_type = request.POST['msg_type']
#                 params['msg_type'] = msg_type

#             api_url = settings.API_URL + '/api-contact/subdomain-chat-listing/'
#             chat_data = call_api_post_method(params, api_url, token=token)
#             if 'error' in chat_data and chat_data['error'] == 0:
#                 try:
#                     read_msg_params = {
#                         "site_id": site_id,
#                         "user_id": user_id,
#                         "master_id": master_id
#                     }
#                     read_api_url = settings.API_URL + '/api-contact/mark-chat-read/'
#                     read_chat = call_api_post_method(read_msg_params, read_api_url, token=token)
#                 except:
#                     pass
#                 chat_listing = chat_data['data']['data']
#                 length = len(chat_listing)
#                 last_msg_id = chat_listing[length-1]['id']
#                 first_msg_id = chat_listing[0]['id']
#                 sender_id = chat_listing[0]['sender_id']
#                 receiver_id = chat_listing[0]['receiver_id']
#                 total = chat_data['data']['total']
#                 page = page
#                 page_size = page_size

#             else:
#                 chat_listing = []
#                 total = 0
#                 page = page
#                 page_size = page_size
#                 last_msg_id = ''
#                 first_msg_id = ''
#                 sender_id = ''
#                 receiver_id = ''

#             context = {'chat_listing': chat_listing, 'total': total, "aws_url": settings.AWS_URL, 'sess_user_id': user_id, 'is_broker': is_broker}
#             if receiver_id != user_id:
#                 context['disable_chat'] = True
#             chat_listing_path = 'admin/dashboard/chat/message-history.html'

#             chat_listing_template = get_template(chat_listing_path)
#             chat_listing_html = chat_listing_template.render(context)

#             bottom_fixed_path = 'admin/dashboard/chat/chat-bottom.html'
#             bottom_fixed_template = get_template(bottom_fixed_path)
#             bottom_fixed_html = bottom_fixed_template.render({})
#             data = {'chat_listing_html': chat_listing_html, 'total': total, 'last_msg_id': last_msg_id, 'master_id': master_id, 'error': 0, "first_msg_id": first_msg_id, "bottom_fixed_html": bottom_fixed_html}
#             if receiver_id != user_id:
#                 data['disable_chat'] = True

#         else:
#             data = {'chat_listing_html': '', 'total': 0, 'last_msg_id': '', 'master_id': master_id, 'error': 1, 'disable_chat': False}

#         return JsonResponse(data)
#     except Exception as exp:
#         print(exp)
#         data = {'status': 403, 'suggestion_list': [], 'error': 1}
#         return JsonResponse(data)

@csrf_exempt
def save_chat_message(request):
    try:
        if request.is_ajax() and request.method == 'POST':
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']
            user_id = None
            token = None
            page = 1
            page_size = 10
            chat_list = []
            if 'user_id' in request.session:
                user_id = request.session['user_id']
                token = request.session['token']['access_token']

            master_id = request.POST['master_id']
            message = request.POST['message']

            send_params = {
                "site_id": site_id,
                "user_id": user_id,
                "master_id": master_id,
                "message": message
            }
            new_api_url = settings.API_URL + '/api-contact/subdomain-send-chat/'
            new_chat_data = call_api_post_method(send_params, new_api_url, token=token)

            if 'error' in new_chat_data and new_chat_data['error'] == 0:
                msg_data = new_chat_data['data'] if 'data' in new_chat_data and new_chat_data['data'] != "" else ""
                last_msg_id = msg_data['id']
                if msg_data:
                    chat_list.append(msg_data)
            else:
                chat_list = []
                last_msg_id = ''


            context = {'chat_listing': chat_list, 'total': 1, "aws_url": settings.AWS_URL, 'sess_user_id': user_id}

            chat_listing_path = 'admin/dashboard/chat/message-history.html'
            chat_listing_template = get_template(chat_listing_path)
            chat_listing_html = chat_listing_template.render(context)

            data = {'chat_listing_html': chat_listing_html, 'total': 1, 'last_msg_id': last_msg_id, 'master_id': master_id, 'error': 0}

        else:
            data = {'chat_listing_html': '', 'total': 0, 'last_msg_id': '', 'master_id': master_id, 'error': 1}

        return JsonResponse(data)
    except Exception as exp:
        print(exp)
        data = {'status': 403, 'suggestion_list': [], 'error': 1}
        return JsonResponse(data)

def preview_theme(request):
    theme = request.GET.get('theme', None)
    is_permission = check_permission(request, 5)
    if not is_permission:
        http_host = request.META['HTTP_HOST']
        redirect_url = settings.URL_SCHEME + str(http_host)
        return HttpResponseRedirect(redirect_url)
    try:
        context = {"data": "Listing COMING SOON...", "active_menu": "dashboard"}
        theme_path = 'admin/dashboard/dashboard/layout-1.html'
        if int(theme) == 5:
            theme_path = 'admin/dashboard/dashboard/layout-2.html'
        if int(theme) == 4:
            theme_path = 'admin/dashboard/dashboard/layout-3.html'
        if int(theme) == 7:
            theme_path = 'admin/dashboard/dashboard/layout-4.html'
        return render(request, theme_path, context)
    except Exception as exp:
        print(exp)
        return HttpResponse("Issue in views")

@csrf_exempt
def get_chat_count(request):
    try:
        if request.is_ajax() and request.method == 'POST':
            try:
                site_detail = subdomain_site_details(request)
                site_id = site_detail['site_detail']['site_id']
            except Exception as exp:
                print(exp)
                site_id = ""

            user_id = None
            token = None
            if 'user_id' in request.session and request.session['user_id']:
                token = request.session['token']['access_token']
                user_id = request.session['user_id']

            params = {
                "site_id": site_id,
                "user_id": user_id
            }
            api_url = settings.API_URL + '/api-settings/chat-count/'
            chat_data = call_api_post_method(params, api_url, token=token)
            if 'error' in chat_data and chat_data['error'] == 0:
                chat = chat_data['data']
                data = {'chat_count': chat['admin_msg_cnt'], 'error': 0, 'msg':'success'}
            else:
                chat = {}
                data = {'chat_count': 0, 'error': 1, 'msg':'failure'}

            return JsonResponse(data)

    except Exception as exp:
        print(exp)
        return HttpResponse("Issue in views")

@csrf_exempt
def bidder_registration_details(request):
    try:
        bidder_id = request.GET.get('id', None)
        try:
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']

        except Exception as exp:
            print(exp)
            site_id = ""


        user_id = None
        token = None
        upload_ids = ''
        upload_names = ''
        if 'user_id' in request.session and request.session['user_id']:
            token = request.session['token']['access_token']
            user_id = request.session['user_id']
        if request.is_ajax() and request.method == 'POST':
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']

            bidder_doc_id = request.POST['bidder_doc_id']
            bidder_doc = []
            try:
                bidder_doc = bidder_doc_id.split(',')

            except:
                bidder_doc = []

            approval_limit = None
            if 'approval_limit' in request.POST and request.POST['approval_limit']:
                approval_limit = request.POST['approval_limit'].replace(',', '').replace('$', '')

            bidder_params = {
                "domain": site_id,
                "registration_id": request.POST['reg_id'],
                "user": user_id,
                "uploads": bidder_doc,
                "is_reviewed": request.POST['review_status'],
                "is_approved": request.POST['apprvoal_status'],
                "seller_comment": request.POST['note_for_buyer'],
                "approval_limit": approval_limit,
            }

            bidder_api_url = settings.API_URL + '/api-bid/update-subdomain-bid-registration/'

            save_data = call_api_post_method(bidder_params, bidder_api_url, token=token)
            if 'error' in save_data and save_data['error'] == 0:
                data = {'status': 200, 'msg': 'Registered Successfully.', 'error': 0, 'data': save_data}
            else:
                data = {'status': 403, 'msg': 'Some error occurs, please try again.', 'error': 1}
            return JsonResponse(data)

        params = {
            'site_id': site_id,
            'registration_id': bidder_id,
            'user_id': user_id
        }
        api_url = settings.API_URL + '/api-bid/subdomain-bid-registration-detail/'
        bidder_data = call_api_post_method(params, api_url, token=token)

        if 'error' in bidder_data and bidder_data['error'] == 0:
            bidder = bidder_data['data']
            try:
                for doc in bidder['upload_information']:
                    upload_ids = upload_ids+','+str(doc['upload_id'])
                    try:
                        ext = doc['doc_file_name'].split('.')
                        extension = ext[-1].lower()
                    except Exception as exp:
                        extension = 'pdf'
                    doc['file_type'] = extension
                    upload_names = upload_names+','+doc['doc_file_name']
            except Exception as exp:
                print(exp)
        else:
            bidder = {}
        upload_ids = upload_ids.lstrip(',')
        upload_names = upload_names.lstrip(',')
        bidder['upload_ids'] = upload_ids
        bidder['upload_names'] = upload_names

        context = {'bidder': bidder, 'active_menu': 'regbidder', 'aws_url': settings.AWS_URL}
        return render(request, 'admin/dashboard/bidder/bidder-registration-details.html', context)
    except Exception as exp:
        print(exp)
        return HttpResponse("Issue in views")

@csrf_exempt
def save_bidder_document(request):
    try:
        user_id = request.session['user_id']
        site_detail = subdomain_site_details(request)
        site_id = site_detail['site_detail']['site_id']
        token = request.session['token']['access_token']
        file_urls = ""
        upload_to = ""
        uploaded_file_list = []
        file_size = request.POST['file_size']

        try:
            for key, value in request.FILES.items():
                params = {}
                if 'bidder_document' in key.lower():
                    upload_to = 'bidder_document'
                    doc_type = 16
                file_urls = request.FILES[key]
        except:
            pass

        if int(request.POST['file_length']) > 1:
            for key, value in request.FILES.items():
                params = {}
                upload_to = 'bidder_document'
                doc_type = 16

                file_res = request.FILES[key]
                response = save_to_s3(file_res, upload_to)
                if 'error' in response and response['error'] == 0:
                    try:
                        upload_param = {
                            "site_id": site_id,
                            "user_id": user_id,
                            "doc_file_name": response['file_name'],
                            "document_type": doc_type,
                            "bucket_name": upload_to,
                            "added_by": user_id,
                            "file_size": str(file_size) + 'MB'
                        }
                        url = settings.API_URL + '/api-users/file-upload/'
                        upload_data = call_api_post_method(upload_param, url, token)
                        upload_id = upload_data['data']['upload_id']
                        upload_size = upload_data['data']['file_size']
                        upload_date = upload_data['data']['added_date']
                    except:
                        upload_id = 0
                        upload_size = '0MB'
                        upload_date = ''

                    params['file_name'] = response['file_name']
                    params['error'] = 0
                    params['msg'] = response['msg']
                    params['upload_id'] = upload_id
                    params['file_size'] = upload_size
                    params['upload_date'] = upload_date
                    params['upload_to'] = upload_to
                else:
                    params['file_name'] = response['file_name']
                    params['error'] = 1
                    params['msg'] = response['msg']
                    params['upload_id'] = 0
                    params['file_size'] = '0MB'
                    params['upload_date'] = ''
                    params['upload_to'] = upload_to

                uploaded_file_list.append(params)
        else:
            params = {}
            response = save_to_s3(file_urls, upload_to)
            if 'error' in response and response['error'] == 0:
                try:
                    upload_param = {
                        "site_id": site_id,
                        "user_id": user_id,
                        "doc_file_name": response['file_name'],
                        "document_type": doc_type,
                        "bucket_name": upload_to,
                        "added_by": user_id,
                        "file_size": str(file_size) + 'MB'
                    }
                    url = settings.API_URL + '/api-users/file-upload/'
                    upload_data = call_api_post_method(upload_param, url, token)
                    upload_id = upload_data['data']['upload_id']
                    upload_size = upload_data['data']['file_size']
                    upload_date = upload_data['data']['added_date']
                except Exception as exp:
                    print(exp)
                    upload_id = 0
                    upload_size = '0MB'
                    upload_date = ''

                params['file_name'] = response['file_name']
                params['error'] = 0
                params['msg'] = response['msg']
                params['upload_id'] = upload_id
                params['file_size'] = upload_size
                params['upload_date'] = upload_date
                params['upload_to'] = upload_to
            else:
                params['file_name'] = response['file_name']
                params['error'] = 1
                params['msg'] = response['msg']
                params['upload_id'] = 0
                params['file_size'] = '0MB'
                params['upload_date'] = ''
                params['upload_to'] = upload_to
            uploaded_file_list.append(params)

        return JsonResponse({'status': 200, 'uploaded_file_list': uploaded_file_list})
    except Exception as exp:
        print(exp)
        data = {'status': 403, 'msg': 'invalid request.'}
        return JsonResponse(data)


@csrf_exempt
def delete_bidder_documents(request):
    try:
        return_data = {
            'section': request.POST['section'],
            'image_id': request.POST['image_id'],
            'image_name': request.POST['image_name'],
        }
        if request.is_ajax() and request.method == 'POST':
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']
            token = request.session['token']['access_token']
            image_name = request.POST['image_name']
            bucket = request.POST['section']

            params = {
                "domain": site_id,
                "user": request.session['user_id'],
                "registration_id": request.POST['reg_id'],
                "upload_id": request.POST['image_id']
            }
            url = settings.API_URL + '/api-bid/subdomain-delete-registration-upload/'

            delete_data = call_api_post_method(params, url, token)

            user_list = []
            if 'error' in delete_data and delete_data['error'] == 0:
                try:
                    delete = delete_s3_file(image_name, setting.AWS_BUCKET_NAME + '/' + bucket)
                except Exception as exp:
                    print(exp)
                return_data['msg'] = 'Image deleted successfully'
                return_data['error'] = 0
                return_data['status'] = 200
                data = return_data
            else:
                return_data['msg'] = 'Some error occurs, please try again'
                return_data['error'] = 1
                return_data['status'] = 403
                data = return_data
        else:
            return_data['msg'] = 'Invalid request'
            return_data['error'] = 1
            return_data['status'] = 403
            data = return_data

        return JsonResponse(data)
    except Exception as exp:
        print(exp)
        data = {'status': 403, 'msg': 'invalid request.', 'user_list': []}
        return JsonResponse(data)

@csrf_exempt
def delete_bidder_reg(request):
    try:
        if request.is_ajax() and request.method == 'POST':
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']
            token = request.session['token']['access_token']
            params = {
                "domain": site_id,
                "registration_id": request.POST['row_id'],
                "user": request.session['user_id']
            }
            token = request.session['token']['access_token']
            url = settings.API_URL + '/api-bid/subdomain-delete-bid-registration/'
            data = call_api_post_method(params, url, token)
            if 'error' in data and data['error'] == 0:
                data = {'status': 200, 'msg': 'Server error, Please try again', 'error': 0, 'data': data}
            else:
                data = {'status': 403, 'msg': data['msg'], 'error': 1}
        else:
            data = {'status': 403, 'msg': 'Forbidden', 'error': 1}

        return JsonResponse(data)
    except Exception as exp:
        print(exp)
        data = {'status': 403, 'msg': 'invalid request.', 'error': 1}
        return JsonResponse(data)

@csrf_exempt
def bidder_registration_search_suggestion(request):
    try:
        if request.is_ajax() and request.method == 'POST':
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']
            token = request.session['token']['access_token']
            user_id = request.session['user_id']

            params = {
                'domain': site_id,
                'user_id': user_id,
                'search': request.POST['search']
            }

            api_url = settings.API_URL + '/api-bid/subdomain-bid-registration-suggestion/'

            suggestion_data = call_api_post_method(params, api_url, token)
            if 'error' in suggestion_data and suggestion_data['error'] == 0:
                data = {'status': 200, 'suggestion_list': suggestion_data['data'], 'error': 0}
            else:
                data = {'status': 403, 'suggestion_list': [], 'error': 1}
        else:
            data = {'status': 403, 'suggestion_list': [], 'error': 1}

        return JsonResponse(data)
    except Exception as exp:
        print(exp)
        data = {'status': 403, 'suggestion_list': [], 'error': 1}
        return JsonResponse(data)


@csrf_exempt
def property_bidder_registration(request):
    try:
        try:
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']

        except Exception as exp:
            print(exp)
            site_id = ""

        user_id = None
        token = None
        if 'user_id' in request.session and request.session['user_id']:
            token = request.session['token']['access_token']
            user_id = request.session['user_id']

        page_size = 10
        if request.is_ajax() and request.method == 'POST':

            search = ''
            if 'search' in request.POST and request.POST['search']:
                search = request.POST['search']


            page = 1
            if 'page' in request.POST and request.POST['page'] != "":
                page = request.POST['page']

            if 'page_size' in request.POST and request.POST['page_size'] != "":
                page_size = request.POST['page_size']

            category_id = ''
            if 'category_id' in request.POST and request.POST['category_id'] != "":
                category_id = request.POST['category_id']
            asset_type = ''
            if 'asset_type' in request.POST and request.POST['asset_type'] != "":
                asset_type = request.POST['asset_type']

            filter_bidder_status = ''
            if 'filter_bidder_status' in request.POST and request.POST['filter_bidder_status'] != "":
                filter_bidder_status = request.POST['filter_bidder_status']

            property_id = request.POST['property_id']
            params = {
                'site_id': site_id,
                'user_id': user_id,
                'page_size': page_size,
                'page': page,
                'asset_type': asset_type,
                'filter_data': filter_bidder_status,
                'property_id': property_id,
                'search': search,
            }


            api_url = settings.API_URL + '/api-bid/subdomain-bid-registration-listing/'
            bidder_data = call_api_post_method(params, api_url, token=token)
            try:
                address = bidder_data['data']['property_address']
                image = bidder_data['data']['property_image']
                if image and image['image'] and image['image'] != "":
                    image_url = settings.AWS_URL + image['bucket_name'] + '/' + image['image']
                else:
                    image_url = ''
                property_address = address['address_one']
                property_city = address['city']
                property_state = address['state']
                property_postal_code = address['postal_code']
                property_image = image_url
            except:
                property_address = ''
                property_city = ''
                property_state = ''
                property_postal_code = ''
                property_image = ''
            
            if 'error' in bidder_data and bidder_data['error'] == 0:
                bidder_list = bidder_data['data']['data']
                total = bidder_data['data']['total']

            else:
                bidder_list = []
                total = 0


            context = {'bidder_list': bidder_list, 'total': total, "aws_url": settings.AWS_URL}

            bidder_listing_path = 'admin/dashboard/listings/property-bid-listings.html'
            bidder_listing_template = get_template(bidder_listing_path)
            bidder_listing_html = bidder_listing_template.render(context)
            # ---------------Pagination--------
            pagination_path = 'admin/dashboard/listings/bidder-list-pagination.html'
            pagination_template = get_template(pagination_path)
            total_page = math.ceil(int(total) / int(page_size))
            pagination_html = ''
            if total_page > 1:
                pagination_data = {"no_page": int(total_page), "total_page": range(total_page), "current_page": int(page),
                                   "pagination_id": "bid_popup_listing_pagination_list", "property_id": property_id}
                pagination_html = pagination_template.render(pagination_data)

            data = {
                'bidder_listing_html': bidder_listing_html,
                'status': 200,
                'msg': '',
                'error': 0,
                'total': total,
                "pagination_html": pagination_html,
                'pagination_id': 'bid_popup_listing_pagination_list',
                'asset_type': asset_type,
                'property_address': property_address,
                'property_city': property_city,
                'property_state': property_state,
                'property_postal_code': property_postal_code,
                'property_image': property_image,
                'page': page,
                'property_id': property_id,
            }
        else:
            data = {'status': 403, 'msg': 'Forbidden', 'error': 1}
        return JsonResponse(data)
    except Exception as exp:
        print(exp)
        return HttpResponse("Issue in views")


@csrf_exempt
def property_bid_history(request):
    try:
        user_id = user_id = None
        page_size = 10
        try:
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']

        except Exception as exp:
            print(exp)
            site_id = ""

        if 'user_id' in request.session and request.session['user_id']:
            token = request.session['token']['access_token']
            user_id = request.session['user_id']

        if request.is_ajax() and request.method == 'POST':
            page = 1
            if 'page' in request.POST and request.POST['page'] != "":
                page = request.POST['page']

            if 'page_size' in request.POST and request.POST['page_size'] != "":
                page_size = request.POST['page_size']

            search = ''
            if 'search' in request.POST and request.POST['search'] != "":
                search = request.POST['search']

            property_id = request.POST['property_id']
            register_user =  request.POST['register_user'] if 'register_user' in request.POST and request.POST['register_user'] else ''
            params = {
                'site_id': site_id,
                'user_id': user_id,
                'page_size': page_size,
                'page': page,
                'property_id': property_id,
                'register_user':register_user,
                'search': search,
            }
            api_url = settings.API_URL + '/api-bid/subdomain-bid-history/'
            bid_history = call_api_post_method(params, api_url, token=token)
            try:
                prop_detail = bid_history['data']['property_detail']
                image = prop_detail['property_image']
                if image and image['image'] and image['image'] != "":
                    image_url = settings.AWS_URL + image['bucket_name'] + '/' + image['image']
                else:
                    image_url = ''
                property_address = prop_detail['address_one']
                property_city = prop_detail['city']
                property_state = prop_detail['state']
                property_postal_code = prop_detail['postal_code']
                property_image = image_url
                auction_type = prop_detail['auction_type']
                bid_increment = prop_detail['bid_increment']
                property_type = prop_detail['property_type']
            except:
                property_address = property_city = property_state = property_postal_code = property_image=auction_type=bid_increment=property_type=''

            if 'error' in bid_history and bid_history['error'] == 0:
                total = bid_history['data']['total']
                new_bid_history = bid_history['data']['new_data']
                # bid_history = bid_history['data']['data']
            else:
                # bid_history = []
                new_bid_history = []
                total = 0

            # context = {'bid_history': bid_history, 'new_bid_history': new_bid_history, 'total': total, "aws_url": settings.AWS_URL, 'start_index': (int(page) - 1) * int(page_size)}
            context = {'new_bid_history': new_bid_history, 'total': total, "aws_url": settings.AWS_URL, 'start_index': (int(page) - 1) * int(page_size)}

            # bidder_listing_path = 'admin/dashboard/listings/property-bid-history.html'
            bidder_listing_path = 'admin/dashboard/listings/new_property-bid-history.html'
            bidder_listing_template = get_template(bidder_listing_path)
            bid_history_html = bidder_listing_template.render(context)
            # ---------------Pagination--------
            # pagination_path = 'admin/dashboard/listings/bid-history-pagination.html'
            pagination_path = 'admin/dashboard/listings/new-bid-history-pagination.html'
            pagination_template = get_template(pagination_path)
            total_page = math.ceil(int(total) / int(page_size))
            pagination_html = ''
            if total_page > 1:
                pagination_data = {"no_page": int(total_page), "total_page": range(total_page), "current_page": int(page),
                                   "pagination_id": "bidHistoryPaginationList", "property_id": property_id}
                pagination_html = pagination_template.render(pagination_data)

            data = {
                'bid_history_html': bid_history_html,
                'status': 200,
                'msg': '',
                'error': 0,
                'total': total,
                "pagination_html": pagination_html,
                'pagination_id': 'bidHistoryPaginationList',
                'property_address': property_address,
                'property_city': property_city,
                'property_state': property_state,
                'property_postal_code': property_postal_code,
                'property_image': property_image,
                'auction_type': auction_type,
                'property_id': property_id,
                'page': page,
                'page_size': page_size,
                'bid_increment': f"{bid_increment:,}",
                'property_type': property_type
            }
        else:
            data = {'status': 403, 'msg': 'Forbidden', 'error': 1}
        return JsonResponse(data)
    except Exception as exp:
        return HttpResponse("Issue in views")


@csrf_exempt
def popup_bidder_registration_search_suggestion(request):
    try:
        if request.is_ajax() and request.method == 'POST':
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']
            token = request.session['token']['access_token']
            user_id = request.session['user_id']

            params = {
                'property_id': request.POST['property_id'],
                'domain': site_id,
                'user_id': user_id,
                'search': request.POST['search']
            }

            api_url = settings.API_URL + '/api-bid/subdomain-property-registration-suggestion/'

            suggestion_data = call_api_post_method(params, api_url, token)
            if 'error' in suggestion_data and suggestion_data['error'] == 0:
                data = {'status': 200, 'suggestion_list': suggestion_data['data'], 'error': 0}
            else:
                data = {'status': 403, 'suggestion_list': [], 'error': 1}
        else:
            data = {'status': 403, 'suggestion_list': [], 'error': 1}

        return JsonResponse(data)
    except Exception as exp:
        print(exp)
        data = {'status': 403, 'suggestion_list': [], 'error': 1}
        return JsonResponse(data)

@csrf_exempt
def user_bid_reg_list(request):
    try:
        if request.is_ajax() and request.method == 'POST':
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']
            params = {
                'user_id': request.POST['user_id'],
                'site_id': site_id
            }
            token = request.session['token']['access_token']
            url = settings.API_URL + '/api-users/subdomain-user-registration/'
            data_list = call_api_post_method(params, url, token)
            if 'error' in data_list and data_list['error'] == 0:
                list_data =  data_list['data']['data']
            else:
                list_data = []
            context = {"list_data": list_data}
            listing_path = 'admin/dashboard/users/user-bid-reg-list.html'
            listing_template = get_template(listing_path)
            listing_html = listing_template.render(context)
            data = {'status': 200, 'msg': 'Data fetched', 'error': 0, 'listing_html': listing_html}
        else:
            data = {'status': 403, 'msg': 'Forbidden'}
        return JsonResponse(data)
    except Exception as exp:
        print(exp)
        data = {'status': 403, 'msg': 'invalid request.'}
        return JsonResponse(data)


@csrf_exempt
def schedule_tour_list(request):
    try:
        site_detail = subdomain_site_details(request)
        site_id = site_detail['site_detail']['site_id']
        token = request.session['token']['access_token']
        user_id = request.session['user_id']
        page_size = 10
        page = 1
        if request.is_ajax() and request.method == 'POST':
            search = ''
            if 'search' in request.POST and request.POST['search']:
                search = request.POST['search']
            page = 1
            if 'page' in request.POST and request.POST['page'] != "":
                page = request.POST['page']

            if 'page_size' in request.POST and request.POST['page_size'] != "":
                page_size = request.POST['page_size']
            params = {
                'domain': site_id,
                'user_id': user_id,
                'page_size': page_size,
                'search': search,
                'page': page
            }
            sno = (int(page) - 1) * int(page_size) + 1
            api_url = settings.API_URL + '/api-property/subdomain-schedule-tour-listing/'
            response_data = call_api_post_method(params, api_url, token=token)
            if 'error' in response_data and response_data['error'] == 0:
                schedule_tour_list = response_data['data']['data']
                schedule_tour_total = response_data['data']['total']
            else:
                schedule_tour_list = []
                schedule_tour_total = 0
            context = {'schedule_tour_list': schedule_tour_list, 'schedule_tour_total': schedule_tour_total, "active_menu": "schedule_tours", "sno": sno}
            schedule_tour_data_path = 'admin/dashboard/schedule/schedule-tour-list.html'
            schedule_tour_template = get_template(schedule_tour_data_path)
            schedule_tour_listing_html = schedule_tour_template.render(context)
            # ---------------Pagination--------
            pagination_path = 'admin/dashboard/schedule/pagination.html'
            pagination_template = get_template(pagination_path)
            total_page = math.ceil(int(schedule_tour_total) / int(page_size))
            pagination_data = {"no_page": int(total_page), "total_page": range(total_page), "current_page": int(page)}
            pagination_html = pagination_template.render(pagination_data)
            data = {'schedule_tour_listing_html': schedule_tour_listing_html, 'status': 200, 'msg': '', 'error': 0, 'total': schedule_tour_total, "pagination_html": pagination_html}
            return JsonResponse(data)
        else:
            params = {
                'domain': site_id,
                'user_id': user_id,
                'page_size': page_size,
                'page': page,
                'search': ''
            }
            sno = (int(page) - 1) * int(page_size) + 1
            api_url = settings.API_URL + '/api-property/subdomain-schedule-tour-listing/'
            response_data = call_api_post_method(params, api_url, token=token)
            if 'error' in response_data and response_data['error'] == 0:
                schedule_tour_list = response_data['data']['data']
                schedule_tour_total = response_data['data']['total']
            else:
                schedule_tour_list = []
                schedule_tour_total = 0
            # ---------------Pagination--------
            pagination_path = 'admin/dashboard/schedule/pagination.html'
            pagination_template = get_template(pagination_path)
            total_page = math.ceil(schedule_tour_total/page_size)
            pagination_data = {"no_page": total_page, "total_page": range(total_page), "current_page": 1}
            pagination_html = pagination_template.render(pagination_data)
            context = {'schedule_tour_list': schedule_tour_list, 'schedule_tour_total': schedule_tour_total, "active_menu": "pages", "active_submenu": "schedule_tours", "pagination_html": pagination_html, "sno": sno}
            return render(request, "admin/dashboard/schedule/view-schedule-tour.html", context)
    except Exception as exp:
        print(exp)
        return HttpResponse("Issue in views")


@csrf_exempt
def property_listing_ordering(request):
    try:
        if request.is_ajax() and request.method == 'POST':
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']
            user_id = request.session['user_id']

            try:
                my_list = json.loads(request.POST['reorder'])
                ordering = {}
                for obj in my_list:
                    ordering[str(obj['property_id'])] = obj['reorder_id']
                
                params = {
                "domain": site_id,
                "user": request.session['user_id'],
                "ordering": ordering 
                }
                token = request.session['token']['access_token']
                url = settings.API_URL + '/api-property/property-listing-ordering/'
                data = call_api_post_method(params, url, token)    
            except Exception as exp:
                print(exp)
                data = {'error': 1, 'data': 'Not api call', 'msg': 'error'}
        else:
            data = {'status': 403, 'msg': 'Forbidden'}
        return JsonResponse(data)
    except Exception as exp:
        print(exp)
        data = {'status': 403, 'msg': 'invalid request.'}
        return JsonResponse(data)


@csrf_exempt
def schedule_search_suggestion(request):
    try:
        if request.is_ajax() and request.method == 'POST':
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']
            token = request.session['token']['access_token']
            user_id = request.session['user_id']
            user_params = {
                'site_id': site_id,
                'search': request.POST['search'],
                'user': user_id
            }

            user_url = settings.API_URL + '/api-property/subdomain-schedule-tour-suggestion/'

            suggestion_data = call_api_post_method(user_params, user_url, token)
            if 'error' in suggestion_data and suggestion_data['error'] == 0:
                data = {'status': 200, 'suggestion_list': suggestion_data['data'], 'error': 0}
            else:
                data = {'status': 403, 'suggestion_list': [], 'error': 1}
        else:
            data = {'status': 403, 'suggestion_list': [], 'error': 1}

        return JsonResponse(data)
    except Exception as exp:
        print(exp)
        data = {'status': 403, 'suggestion_list': [], 'error': 1}
        return JsonResponse(data)

@csrf_exempt
def get_plan_history(request):
    try:
        is_permission = check_permission(request, 5)
        if not is_permission:
            http_host = request.META['HTTP_HOST']
            redirect_url = settings.URL_SCHEME + str(http_host)
            return HttpResponseRedirect(redirect_url)

        try:
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']

        except Exception as exp:
            print(exp)
            site_id = ""

        user_id = None
        token = None
        if 'user_id' in request.session and request.session['user_id']:
            token = request.session['token']['access_token']
            user_id = request.session['user_id']


        if request.is_ajax() and request.method == 'POST':

            user_search = ''
            # if 'search' in request.POST and request.POST['search']:
            #     user_search = request.POST['search']

            page = 1
            page_size = 10
            if 'page' in request.POST and request.POST['page'] != "":
                page = request.POST['page']


            if 'perpage' in request.POST and request.POST['perpage']:
                page_size = request.POST['perpage']

            # if 'status' in request.POST and request.POST['status'] and request.POST['status'].lower() == 'active':
            #     status = [1]
            # elif 'status' in request.POST and request.POST['status'] and request.POST['status'] and request.POST[
            #     'status'].lower() == 'inactive':
            #     status = [2]
            # else:
            #     status = [2, 1]

            list_param = {
                'site_id': site_id,
                'user_id': user_id,
                'page': page,
                'page_size': page_size
            }
            sno = (int(page) - 1) * int(page_size) + 1
            list_url = settings.API_URL + '/api-users/plan-billing-history/'
            list_data = call_api_post_method(list_param, list_url, token)

            if 'error' in list_data and list_data['error'] == 0:
                plan_history_list = list_data['data']['data']
                total = list_data['data']['total']
            else:
                plan_history_list = []
                total = 0
            context = {'plan_history_list': plan_history_list, 'total': total, "aws_url": settings.AWS_URL, 'sno': sno}

            plan_history_path = 'admin/dashboard/dashboard/plan-history-content.html'
            plan_history_template = get_template(plan_history_path)
            plan_history_html = plan_history_template.render(context)
            # ---------------Pagination--------
            pagination_path = 'admin/dashboard/dashboard/history-pagination.html'
            pagination_template = get_template(pagination_path)
            total_page = math.ceil(int(total) / int(page_size))
            pagination_html = ''
            if total_page > 1:
                pagination_data = {"no_page": int(total_page), "total_page": range(total_page), "current_page": int(page),
                                   "pagination_id": "history_listing_pagination_list"}
                pagination_html = pagination_template.render(pagination_data)

            data = {'history_html': plan_history_html, 'status': 200, 'msg': '', 'error': 0, 'total': total,
                    "pagination_html": pagination_html, 'pagination_id': 'history_listing_pagination_list'}

        else:
            data = {'history_html': '', 'status': 403, 'msg': '', 'error': 1, 'total': total,
                    "pagination_html": '', 'pagination_id': 'history_listing_pagination_list'}

        return JsonResponse(data)
    except Exception as exp:
        print(exp)
        return HttpResponse("Issue in views")

def about_us(request):
    is_permission = check_permission(request, 16)
    if not is_permission:
        http_host = request.META['HTTP_HOST']
        redirect_url = settings.URL_SCHEME + str(http_host)
        return HttpResponseRedirect(redirect_url)
    cms_id = request.GET.get('id', None)
    try:
        site_detail = subdomain_site_details(request)
        site_id = site_detail['site_detail']['site_id']
        token = request.session['token']['access_token']
        user_id = request.session['user_id']


        try:
            cms_detail_param = {
                'domain': site_id,
                'user_id': user_id,
                'slug': 'about-us'
            }
            cms_detail_url = settings.API_URL + '/api-cms/subdomain-cms/'
            cms_detail_data = call_api_post_method(cms_detail_param, cms_detail_url, token)
            cms_details = cms_detail_data['data']

        except:
            cms_details = {}

        try:
            status_param = {}
            status_url = settings.API_URL + '/api-settings/lookup-status-listing/'
            status_data = call_api_post_method(status_param, status_url, token)
            status_list = status_data['data']
        except:
            status_list = []

        context = {"active_menu": "cms", "active_submenu":"about", "status_list": status_list, 'cms': cms_details}
        return render(request, "admin/dashboard/cms/about-us.html", context)
    except Exception as exp:
        print(exp)
        return HttpResponse("Issue in views")

def contact_us(request):
    is_permission = check_permission(request, 16)
    if not is_permission:
        http_host = request.META['HTTP_HOST']
        redirect_url = settings.URL_SCHEME + str(http_host)
        return HttpResponseRedirect(redirect_url)
    cms_id = request.GET.get('id', None)
    try:
        site_detail = subdomain_site_details(request)
        site_id = site_detail['site_detail']['site_id']
        token = request.session['token']['access_token']
        user_id = request.session['user_id']


        try:
            cms_detail_param = {
                'domain': site_id,
                'user_id': user_id,
                'slug': 'contact-us'
            }
            cms_detail_url = settings.API_URL + '/api-cms/subdomain-cms/'
            cms_detail_data = call_api_post_method(cms_detail_param, cms_detail_url, token)
            cms_details = cms_detail_data['data']

        except:
            cms_details = {}

        try:
            status_param = {}
            status_url = settings.API_URL + '/api-settings/lookup-status-listing/'
            status_data = call_api_post_method(status_param, status_url, token)
            status_list = status_data['data']
        except:
            status_list = []

        context = {"active_menu": "cms", "active_submenu":"cms_contact", "status_list": status_list, 'cms': cms_details}
        return render(request, "admin/dashboard/cms/contact-us.html", context)
    except Exception as exp:
        print(exp)
        return HttpResponse("Issue in views")

def terms(request):
    is_permission = check_permission(request, 16)
    if not is_permission:
        http_host = request.META['HTTP_HOST']
        redirect_url = settings.URL_SCHEME + str(http_host)
        return HttpResponseRedirect(redirect_url)
    cms_id = request.GET.get('id', None)
    try:
        site_detail = subdomain_site_details(request)
        site_id = site_detail['site_detail']['site_id']
        token = request.session['token']['access_token']
        user_id = request.session['user_id']


        try:
            cms_detail_param = {
                'domain': site_id,
                'user_id': user_id,
                'slug': 'terms'
            }
            cms_detail_url = settings.API_URL + '/api-cms/subdomain-cms/'
            cms_detail_data = call_api_post_method(cms_detail_param, cms_detail_url, token)
            cms_details = cms_detail_data['data']

        except:
            cms_details = {}

        try:
            status_param = {}
            status_url = settings.API_URL + '/api-settings/lookup-status-listing/'
            status_data = call_api_post_method(status_param, status_url, token)
            status_list = status_data['data']
        except:
            status_list = []

        context = {"active_menu": "cms", "active_submenu":"terms", "status_list": status_list, 'cms': cms_details}
        return render(request, "admin/dashboard/cms/terms.html", context)
    except Exception as exp:
        print(exp)
        return HttpResponse("Issue in views")

def privacy_policy(request):
    is_permission = check_permission(request, 16)
    if not is_permission:
        http_host = request.META['HTTP_HOST']
        redirect_url = settings.URL_SCHEME + str(http_host)
        return HttpResponseRedirect(redirect_url)
    cms_id = request.GET.get('id', None)
    try:
        site_detail = subdomain_site_details(request)
        site_id = site_detail['site_detail']['site_id']
        token = request.session['token']['access_token']
        user_id = request.session['user_id']


        try:
            cms_detail_param = {
                'domain': site_id,
                'user_id': user_id,
                'slug': 'privacy-policy'
            }
            cms_detail_url = settings.API_URL + '/api-cms/subdomain-cms/'
            cms_detail_data = call_api_post_method(cms_detail_param, cms_detail_url, token)
            cms_details = cms_detail_data['data']

        except:
            cms_details = {}

        try:
            status_param = {}
            status_url = settings.API_URL + '/api-settings/lookup-status-listing/'
            status_data = call_api_post_method(status_param, status_url, token)
            status_list = status_data['data']
        except:
            status_list = []

        context = {"active_menu": "cms", "active_submenu":"privacy", "status_list": status_list, 'cms': cms_details}
        return render(request, "admin/dashboard/cms/privacy-policy.html", context)
    except Exception as exp:
        print(exp)
        return HttpResponse("Issue in views")

@csrf_exempt
def save_cms(request):
    try:
        if request.is_ajax() and request.method == 'POST':
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']
            token = request.session['token']['access_token']
            user_id = request.session['user_id']
            cms_params = {
                'domain': site_id,
                'user_id': user_id,
                'status': request.POST['cms_status'],
                'meta_key_word': request.POST['meta_keywords'],
                'meta_title': request.POST['meta_title'],
                'meta_description': request.POST['meta_desc'],
                'page_title': request.POST['page_title'],
                'slug': request.POST['slug'],
                'page_content': request.POST['page_content'],
            }
            if request.POST['cms_id']:
                cms_params['cms_id'] = request.POST['cms_id']

            cms_url = settings.API_URL + '/api-cms/subdomain-cms-update/'

            cms_response = call_api_post_method(cms_params, cms_url, token)

            if 'error' in cms_response and cms_response['error'] == 0:
                data = {
                    'error': 0,
                    'status': 200,
                    'msg': cms_response['msg']
                }
            else:
                data = {
                    'error': 1,
                    'status': 403,
                    'msg': cms_response['msg']
                }
        else:
            data = {'status': 403, 'msg': " invalid request", 'error': 1}

        return JsonResponse(data)
    except Exception as exp:
        print(exp)
        data = {'status': 403, 'msg': "forbidden", 'error': 1}
        return JsonResponse(data)


@csrf_exempt
def auction_dashboard(request):
    try:
        site_detail = subdomain_site_details(request)
        site_id = site_detail['site_detail']['site_id']
        token = request.session['token']['access_token']
        user_id = request.session['user_id']
        page_size = 10

        if request.is_ajax() and request.method == 'POST':

            search = ''
            if 'search' in request.POST and request.POST['search']:
                search = request.POST['search']

            page = 1
            if 'page' in request.POST and request.POST['page'] != "":
                page = int(request.POST['page'])

            page_size = 10
            if 'perpage' in request.POST and request.POST['perpage']:
                page_size = int(request.POST['perpage'])

            asset_type = ''
            if 'asset_type' in request.POST and request.POST['asset_type']:
                asset_type = int(request.POST['asset_type'])
            auction_type = ''
            if 'auction_type' in request.POST and request.POST['auction_type']:
                auction_type = int(request.POST['auction_type'])
            property_type = ''
            if 'property_type' in request.POST and request.POST['property_type']:
                property_type = int(request.POST['property_type'])

            status = ""
            if 'status' in request.POST and request.POST['status']:
                status = request.POST['status']
            
            agent = ''
            if 'agent' in request.POST and request.POST['agent']:
                agent = int(request.POST['agent'])

            list_param = {
                "page": page,
                "page_size": page_size,
                "domain_id": site_id,
                "user_id": user_id,
                "auction_id": auction_type,
                "asset_id": asset_type,
                "property_type": property_type,
                "search": search,
                "agent_id":agent,
                "status": status
            }
            api_url = settings.API_URL + '/api-property/property-auction-dashboard/'

            auction_data = call_api_post_method(list_param, api_url, token)

            if 'error' in auction_data and auction_data['error'] == 0:
                auction_listings = auction_data['data']['data']
                total = auction_data['data']['total'] if 'total' in auction_data['data'] else 0
            else:
                auction_listings = []
                total = 0
            sno = (int(page) - 1) * int(page_size) + 1

            context = {
                "auction_data": auction_listings,
                'total': total,
                "aws_url": settings.AWS_URL,
                "is_broker": 1 if request.session['is_broker'] == True else 0,
                "sno": sno
            }
            prop_auction_path = 'admin/dashboard/auction/auction-listing-content.html'
            property_listing_template = get_template(prop_auction_path)
            property_auction_listings = property_listing_template.render(context)
            # ---------------Pagination--------
            pagination_html = ''
            pagination_path = pagination_path = 'admin/dashboard/auction/auction-pagination.html'
            pagination_template = get_template(pagination_path)
            total_page = math.ceil(total / page_size)
            
            if total_page > 1:
                pagination_data = {"no_page": int(total_page), "total_page": range(total_page), "current_page": int(page),
                    "pagination_id": "prop_auction_pagination_list"}
                pagination_html = pagination_template.render(pagination_data)
            


            return JsonResponse({
                'property_auction_listings': property_auction_listings,
                'status': 200,
                'msg': '',
                'error': 0,
                'total': total,
                "pagination_html": pagination_html,
                'pagination_id': 'prop_listing_pagination_list',
                "sno": sno
            })
        else:
            page = 1
            try:
                params = {"domain_id": site_id, "user_id": user_id}
                api_url = settings.API_URL + '/api-users/agent-listing/'
                auction_type_data = call_api_post_method(params, api_url, token)
                agent_list = auction_type_data['data']
            except:
                agent_list = []
            try:
                auction_type_param = {}
                auction_type_url = settings.API_URL + '/api-settings/subdomain-auction-type/'
                auction_type_data = call_api_post_method(auction_type_param, auction_type_url, token)
                auction_type_list = auction_type_data['data']
            except:
                auction_type_list = []

            try:
                asset_listing_params = {}

                asset_listing_url = settings.API_URL + '/api-property/asset-listing/'

                asset_listing_data = call_api_post_method(asset_listing_params, asset_listing_url, token)
                asset_listing = asset_listing_data['data']
            except:
                asset_listing = []

            try:
                property_type_params = {}

                property_type_url = settings.API_URL + '/api-property/property-type-listing/'

                property_type_data = call_api_post_method(property_type_params, property_type_url, token)
                property_type_listing = property_type_data['data']
            except:
                property_type_listing = []

            # try:
            #     status_param = {'object_id': 9}
            #     status_url = settings.API_URL + '/api-settings/lookup-status-listing/'
            #     status_data = call_api_post_method(status_param, status_url, token)
            #     status_list = status_data['data']
            # except:
            #     status_list = []
            
            params = {"domain_id": site_id, "user_id": user_id, "page": page, "page_size": page_size, "status": 1 }

            api_url = settings.API_URL + '/api-property/property-auction-dashboard/'

            auction_data = call_api_post_method(params, api_url, token)

            if 'error' in auction_data and auction_data['error'] == 0:
                auction_listings = auction_data['data']['data']
                total = auction_data['data']['total'] if 'total' in auction_data['data'] else 0
            else:
                auction_listings = []
                total = 0
            # ---------------Pagination--------
            pagination_html = ''
            pagination_path = pagination_path = 'admin/dashboard/auction/auction-pagination.html'
            pagination_template = get_template(pagination_path)
            total_page = math.ceil(total / page_size)
            sno = (int(page) - 1) * int(page_size) + 1
            if total_page > 1:
                pagination_data = {"no_page": total_page, "total_page": range(total_page), "current_page": 1,
                    "pagination_id": "prop_auction_pagination_list"}
                pagination_html = pagination_template.render(pagination_data)
            
            x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
            if x_forwarded_for:
                ip_address = x_forwarded_for.split(',')[0]
            else:
                ip_address = request.META.get('REMOTE_ADDR')

            context = {
                "active_menu": "listing_monitor",
                "active_submenu": "auction-dashboard",
                "auction_data": auction_listings,
                "auction_type_list": auction_type_list,
                "property_type_listing": property_type_listing,
                # "status_list": status_list,
                "asset_listing": asset_listing,
                "agent_list":agent_list,
                "pagination_html": pagination_html,
                "is_broker": 1 if request.session['is_broker'] == True else 0,
                "sno": sno,
                "ip_address": ip_address,
                "node_url": settings.NODE_URL,
                "user_id": user_id,
                "domain_id": site_id
            }

            return render(request, "admin/dashboard/auction/index.html", context)
    except Exception as exp:
        print(exp)
        return HttpResponse("Issue in views")


@csrf_exempt
def auction_search_suggestion(request):
    try:
        if request.is_ajax() and request.method == 'POST':
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']
            token = request.session['token']['access_token']

            user_params = {
                'site_id': site_id,
                'search': request.POST['search'],
                'user_id':request.session['user_id']
            }

            user_url = settings.API_URL + '/api-property/property-auction-suggestion/'

            suggestion_data = call_api_post_method(user_params, user_url, token)
            if 'error' in suggestion_data and suggestion_data['error'] == 0:
                data = {'status': 200, 'suggestion_list': suggestion_data['data'], 'error': 0}
            else:
                data = {'status': 403, 'suggestion_list': [], 'error': 1}
        else:
            data = {'status': 403, 'suggestion_list': [], 'error': 1}

        return JsonResponse(data)
    except Exception as exp:
        print(exp)
        data = {'status': 403, 'suggestion_list': [], 'error': 1}
        return JsonResponse(data)


@csrf_exempt
def update_reserve_price(request):
    try:
        if request.is_ajax() and request.method == 'POST':
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']
            token = request.session['token']['access_token']

            user_params = {
                'domain_id': site_id,
                'user_id':request.session['user_id'],
                "property_id": request.POST['listing_id'],
                "reserve_amount": request.POST['new_price']
            }

            user_url = settings.API_URL + '/api-property/update-reserve-amount/'

            data = call_api_post_method(user_params, user_url, token)
            if 'error' in data and data['error'] == 0:
                data = {'status': 200, 'msg': data['msg'], 'error': 0}
            else:
                data = {'status': 403, 'msg': 'Something went wrong!', 'error': 1}
        else:
            data = {'status': 403, 'msg': 'Forbidden', 'error': 1}
        return JsonResponse(data)
    except Exception as exp:
        print(exp)
        data = {'status': 403, 'msg': 'Something went wrong!', 'error': 1}
        return JsonResponse(data)


@csrf_exempt
def update_bid_increment(request):
    try:
        if request.is_ajax() and request.method == 'POST':
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']
            token = request.session['token']['access_token']

            user_params = {
                'domain_id': site_id,
                'user_id':request.session['user_id'],
                "property_id": request.POST['listing_id'],
                "bid_increments": request.POST['new_price']
            }

            user_url = settings.API_URL + '/api-property/update-bid-increment/'

            data = call_api_post_method(user_params, user_url, token)
            if 'error' in data and data['error'] == 0:
                data = {'status': 200, 'msg': data['msg'], 'error': 0}
            else:
                data = {'status': 403, 'msg': 'Something went wrong!', 'error': 1}
        else:
            data = {'status': 403, 'msg': 'Forbidden', 'error': 1}
        return JsonResponse(data)
    except Exception as exp:
        print(exp)
        data = {'status': 403, 'msg': 'Something went wrong!', 'error': 1}
        return JsonResponse(data)


@csrf_exempt
def start_stop_bid_auction(request):
    try:
        if request.is_ajax() and request.method == 'POST':
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']
            token = request.session['token']['access_token']

            user_params = {
                'domain_id': site_id,
                'user_id':request.session['user_id'],
                "property_id": request.POST['listing_id'],
            }

            response = settings.API_URL + '/api-property/start-stop-auction/'

            data = call_api_post_method(user_params, response, token)
            if 'error' in data and data['error'] == 0:
                data = {'status': 200, 'msg': data['msg'], 'error': 0}
            else:
                data = {'status': 403, 'msg': 'Something went wrong!', 'error': 1}
        else:
            data = {'status': 403, 'msg': 'Forbidden', 'error': 1}
        return JsonResponse(data)
    except Exception as exp:
        print(exp)
        data = {'status': 403, 'msg': 'Something went wrong!', 'error': 1}
        return JsonResponse(data)


@csrf_exempt
def edit_auction(request):
    try:
        if request.is_ajax() and request.method == 'POST':
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']
            token = request.session['token']['access_token']

            user_params = {
                'domain_id': site_id,
                'user_id':request.session['user_id'],
                "property_id": request.POST['listing_id'],
                "start_date": request.POST['start_date'],
                "end_date": request.POST['end_date'],
            }

            response = settings.API_URL + '/api-property/update-auction-date/'

            data = call_api_post_method(user_params, response, token)
            if 'error' in data and data['error'] == 0:
                data = {'status': 200, 'msg': data['msg'], 'error': 0}
            else:
                data = {'status': 403, 'msg': 'Something went wrong!', 'error': 1}
        else:
            data = {'status': 403, 'msg': 'Forbidden', 'error': 1}
        return JsonResponse(data)
    except Exception as exp:
        print(exp)
        data = {'status': 403, 'msg': 'Something went wrong!', 'error': 1}
        return JsonResponse(data)


@csrf_exempt
def email_auction_users(request):
    try:
        if request.is_ajax() and request.method == 'POST':
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']
            token = request.session['token']['access_token']

            user_params = {
                'domain_id': site_id,
                'user_id':request.session['user_id'],
                "property_id": request.POST['listing_id'],
                "subject": request.POST['subject'],
                "message": request.POST['message'],
                "email_for": request.POST['email_for']
            }
            response = settings.API_URL + '/api-users/send-custom-email/'

            data = call_api_post_method(user_params, response, token)
            if 'error' in data and data['error'] == 0:
                data = {'status': 200, 'msg': data['msg'], 'error': 0}
            else:
                data = {'status': 403, 'msg': 'Something went wrong!', 'error': 1}
        else:
            data = {'status': 403, 'msg': 'Forbidden', 'error': 1}
        return JsonResponse(data)
    except Exception as exp:
        print(exp)
        data = {'status': 403, 'msg': 'Something went wrong!', 'error': 1}
        return JsonResponse(data)


@csrf_exempt
def auction_bid_history(request):
    try:
        user_id = None
        page_size = 10
        try:
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']

        except Exception as exp:
            print(exp)
            site_id = ""

        if 'user_id' in request.session and request.session['user_id']:
            token = request.session['token']['access_token']
            user_id = request.session['user_id']

        if request.is_ajax() and request.method == 'POST':
            page = 1
            if 'page' in request.POST and request.POST['page'] != "":
                page = request.POST['page']

            if 'page_size' in request.POST and request.POST['page_size'] != "":
                page_size = request.POST['page_size']

            property_id = request.POST['property_id']
            params = {
                'site_id': site_id,
                'user_id': user_id,
                'page_size': page_size,
                'page': page,
                'property_id': property_id,
            }

            api_url = settings.API_URL + '/api-bid/auction-total-bids/'
            bid_history = call_api_post_method(params, api_url, token=token)
            try:
                prop_detail = bid_history['data']['property_detail']
                image = prop_detail['property_image']
                if image and image['image'] and image['image'] != "":
                    image_url = settings.AWS_URL + image['bucket_name'] + '/' + image['image']
                else:
                    image_url = ''
                property_address = prop_detail['address_one']
                property_city = prop_detail['city']
                property_state = prop_detail['state']
                property_postal_code = prop_detail['postal_code']
                property_image = image_url
            except:
                property_address = property_city = property_state = property_postal_code = property_image=''

            
            if 'error' in bid_history and bid_history['error'] == 0:
                total = bid_history['data']['total']
                bid_history = bid_history['data']['data']

            else:
                bid_history = []
                total = 0


            context = {'bid_history': bid_history, 'total': total, "aws_url": settings.AWS_URL}

            bidder_listing_path = 'admin/dashboard/auction/bid-history.html'
            bidder_listing_template = get_template(bidder_listing_path)
            bid_history_html = bidder_listing_template.render(context)
            # ---------------Pagination--------
            pagination_path = 'admin/dashboard/auction/bid-history-pagination.html'
            pagination_template = get_template(pagination_path)
            total_page = math.ceil(int(total) / int(page_size))
            pagination_html = ''
            if total_page > 1:
                pagination_data = {"no_page": int(total_page), "total_page": range(total_page), "current_page": int(page),
                                   "pagination_id": "bidHistoryPaginationList", "property_id": property_id}
                pagination_html = pagination_template.render(pagination_data)

            data = {
                'bid_history_html': bid_history_html,
                'status': 200,
                'msg': '',
                'error': 0,
                'total': total,
                "pagination_html": pagination_html,
                'pagination_id': 'bidHistoryPaginationList',
                'property_address': property_address,
                'property_city': property_city,
                'property_state': property_state,
                'property_postal_code': property_postal_code,
                'property_image': property_image,
                'page': page,
                'property_id': property_id,
            }
        else:
            data = {'status': 403, 'msg': 'Forbidden', 'error': 1}
        return JsonResponse(data)
    except Exception as exp:
        print(exp)
        return HttpResponse("Issue in views")


@csrf_exempt
def auction_bidder_history(request):
    try:
        user_id = None
        page_size = 10
        try:
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']

        except Exception as exp:
            print(exp)
            site_id = ""

        if 'user_id' in request.session and request.session['user_id']:
            token = request.session['token']['access_token']
            user_id = request.session['user_id']

        if request.is_ajax() and request.method == 'POST':
            page = 1
            if 'page' in request.POST and request.POST['page'] != "":
                page = request.POST['page']

            if 'page_size' in request.POST and request.POST['page_size'] != "":
                page_size = request.POST['page_size']

            property_id = request.POST['property_id']
            params = {
                'site_id': site_id,
                'user_id': user_id,
                'page_size': page_size,
                'page': page,
                'property_id': property_id,
            }

            api_url = settings.API_URL + '/api-bid/auction-bidders/'
            bidder_history = call_api_post_method(params, api_url, token=token)
            try:
                prop_detail = bidder_history['data']['property_detail']
                image = prop_detail['property_image']
                if image and image['image'] and image['image'] != "":
                    image_url = settings.AWS_URL + image['bucket_name'] + '/' + image['image']
                else:
                    image_url = ''
                property_address = prop_detail['address_one']
                property_city = prop_detail['city']
                property_state = prop_detail['state']
                property_postal_code = prop_detail['postal_code']
                property_image = image_url
            except:
                property_address = property_city = property_state = property_postal_code = property_image=''

            
            if 'error' in bidder_history and bidder_history['error'] == 0:
                total = bidder_history['data']['total']
                bidder_history = bidder_history['data']['data']

            else:
                bidder_history = []
                total = 0


            context = {'bidder_history': bidder_history, 'total': total, "aws_url": settings.AWS_URL, 'start_index': (int(page) - 1) * int(page_size)}
            bidder_listing_path = 'admin/dashboard/auction/bidder-history.html'
            bidder_listing_template = get_template(bidder_listing_path)
            bidder_history_html = bidder_listing_template.render(context)
            # ---------------Pagination--------
            pagination_path = 'admin/dashboard/auction/bidder-history-pagination.html'
            pagination_template = get_template(pagination_path)
            total_page = math.ceil(int(total) / int(page_size))
            pagination_html = ''
            if total_page > 1:
                pagination_data = {"no_page": int(total_page), "total_page": range(total_page), "current_page": int(page),
                                   "pagination_id": "bidderHistoryPaginationList", "property_id": property_id}
                pagination_html = pagination_template.render(pagination_data)

            data = {
                'bidder_history_html': bidder_history_html,
                'status': 200,
                'msg': '',
                'error': 0,
                'total': total,
                "pagination_html": pagination_html,
                'pagination_id': 'bidderHistoryPaginationList',
                'property_address': property_address,
                'property_city': property_city,
                'property_state': property_state,
                'property_postal_code': property_postal_code,
                'property_image': property_image,
                'page': page,
                'property_id': property_id,
            }
        else:
            data = {'status': 403, 'msg': 'Forbidden', 'error': 1}
        return JsonResponse(data)
    except Exception as exp:
        print(exp)
        return HttpResponse("Issue in views")


@csrf_exempt
def auction_watcher_history(request):
    try:
        user_id = None
        page_size = 8
        try:
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']

        except Exception as exp:
            print(exp)
            site_id = ""

        if 'user_id' in request.session and request.session['user_id']:
            token = request.session['token']['access_token']
            user_id = request.session['user_id']

        if request.is_ajax() and request.method == 'POST':
            page = 1
            if 'page' in request.POST and request.POST['page'] != "":
                page = request.POST['page']

            if 'page_size' in request.POST and request.POST['page_size'] != "":
                page_size = request.POST['page_size']

            property_id = request.POST['property_id']
            params = {
                'site_id': site_id,
                'user_id': user_id,
                'page_size': page_size,
                'page': page,
                'property_id': property_id,
            }

            api_url = settings.API_URL + '/api-bid/auction-total-watching/'
            watcher_history = call_api_post_method(params, api_url, token=token)
            try:
                prop_detail = watcher_history['data']['property_detail']
                image = prop_detail['property_image']
                if image and image['image'] and image['image'] != "":
                    image_url = settings.AWS_URL + image['bucket_name'] + '/' + image['image']
                else:
                    image_url = ''
                property_address = prop_detail['address_one']
                property_city = prop_detail['city']
                property_state = prop_detail['state']
                property_postal_code = prop_detail['postal_code']
                property_image = image_url
            except Exception as e:
                print(e)
                property_address = property_city = property_state = property_postal_code = property_image=''

            
            if 'error' in watcher_history and watcher_history['error'] == 0:
                total = watcher_history['data']['total']
                no_anonymous_watcher = watcher_history['data']['no_anonymous_watcher']
                total_watcher = watcher_history['data']['total_watcher']
                watcher_history = watcher_history['data']['data']

            else:
                watcher_history = []
                total=no_anonymous_watcher=total_watcher=0


            context = {'watcher_history': watcher_history, 'total': total, "aws_url": settings.AWS_URL, 'start_index': (int(page) - 1) * int(page_size)}
            bidder_listing_path = 'admin/dashboard/auction/watcher-history.html'
            watcher_listing_template = get_template(bidder_listing_path)
            watcher_history_html = watcher_listing_template.render(context)
            # ---------------Pagination--------
            pagination_path = 'admin/dashboard/auction/watcher-history-pagination.html'
            pagination_template = get_template(pagination_path)
            total_page = math.ceil(int(total) / int(page_size))
            pagination_html = ''
            if total_page > 1:
                pagination_data = {"no_page": int(total_page), "total_page": range(total_page), "current_page": int(page),
                                   "pagination_id": "watcherHistoryPaginationList", "property_id": property_id}
                pagination_html = pagination_template.render(pagination_data)

            data = {
                'watcher_history_html': watcher_history_html,
                'status': 200,
                'msg': '',
                'error': 0,
                'total': total,
                "no_anonymous_watcher": no_anonymous_watcher,
                "total_watcher":total_watcher,
                "pagination_html": pagination_html,
                'pagination_id': 'watcherHistoryPaginationList',
                'property_address': property_address,
                'property_city': property_city,
                'property_state': property_state,
                'property_postal_code': property_postal_code,
                'property_image': property_image,
                'page': page,
                'property_id': property_id,
            }
        else:
            data = {'status': 403, 'msg': 'Forbidden', 'error': 1}
        return JsonResponse(data)
    except Exception as exp:
        print(exp)
        return HttpResponse("Issue in views")


@csrf_exempt
def new_bid_checked(request):
    try:
        if request.is_ajax() and request.method == 'POST':
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']
            token = request.session['token']['access_token']
            params = {
                'domain_id': site_id,
                'user_id':request.session['user_id'],
                "property_id": request.POST['property_id'],
            }
            response = settings.API_URL + '/api-property/auction-listing-read/'
            data = call_api_post_method(params, response, token)
            if 'error' in data and data['error'] == 0:
                data = {'status': 200, 'msg': data['msg'], 'error': 0}
            else:
                data = {'status': 403, 'msg': 'Something went wrong!', 'error': 1}
        else:
            data = {'status': 403, 'msg': 'Forbidden', 'error': 1}
        return JsonResponse(data)
    except Exception as exp:
        print(exp)
        data = {'status': 403, 'msg': 'Something went wrong!', 'error': 1}
        return JsonResponse(data)


@csrf_exempt
def email_template(request):
    try:
        is_permission = check_permission(request, 15)
        if not is_permission:
            http_host = request.META['HTTP_HOST']
            redirect_url = settings.URL_SCHEME + str(http_host)
            return HttpResponseRedirect(redirect_url)

        try:
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']

        except Exception as exp:
            print(exp)
            site_id = ""

        user_id = None
        token = None
        if 'user_id' in request.session and request.session['user_id']:
            token = request.session['token']['access_token']
            user_id = request.session['user_id']

        page_size = 10
        page = 1
        try:
            status_param = {}
            status_url = settings.API_URL + '/api-settings/lookup-status-listing/'
            status_data = call_api_post_method(status_param, status_url, token)
            status_list = status_data['data']
        except:
            status_list = []

        if request.is_ajax() and request.method == 'POST':

            template_search = ''
            if 'search' in request.POST and request.POST['search']:
                template_search = request.POST['search']

            page = 1
            if 'page' in request.POST and request.POST['page'] != "":
                page = request.POST['page']

            page_size = 10
            if 'perpage' in request.POST and request.POST['perpage']:
                page_size = request.POST['perpage']

            if 'status' in request.POST and request.POST['status'] and request.POST['status'] and request.POST[
                'status'].lower() == 'active':
                status = 1
            elif 'status' in request.POST and request.POST['status'] and request.POST['status'] and request.POST[
                'status'].lower() == 'inactive':
                status = 2
            else:
                status = ""

            list_param = {
                "domain_id": site_id,
                "status": status,
                "user_id": user_id,
                "search": template_search,
                "page": page,
                "page_size": page_size,
            }
            sno = (int(page) - 1) * int(page_size) + 1
            list_url = settings.API_URL + '/api-notifications/subdomain-template-listing/'
            list_data = call_api_post_method(list_param, list_url, token)

            if 'error' in list_data and list_data['error'] == 0:
                template_list = list_data['data']['data']
                total = list_data['data']['total']
            else:
                template_list = []
                total = 0
            context = {'template_list': template_list, 'total': total, "aws_url": settings.AWS_URL, 'sno': sno}

            template_listing_path = 'admin/dashboard/email_template/template-listing-content.html'
            template_listing_template = get_template(template_listing_path)
            template_listing_html = template_listing_template.render(context)
            # ---------------Pagination--------
            pagination_path = 'admin/dashboard/email_template/pagination.html'
            pagination_template = get_template(pagination_path)
            total_page = math.ceil(int(total) / int(page_size))
            pagination_html = ''
            if total_page > 1:
                pagination_data = {"no_page": int(total_page), "total_page": range(total_page), "current_page": int(page),
                                   "pagination_id": "template_pagination_list"}
                pagination_html = pagination_template.render(pagination_data)

            data = {'template_listing_html': template_listing_html, 'status': 200, 'msg': '', 'error': 0, 'total': total,
                    "pagination_html": pagination_html, 'pagination_id': 'template_pagination_list'}
            return JsonResponse(data)
        else:
            list_param = {
                'domain_id': site_id,
                "page": 1,
                "user_id": user_id,
                "page_size": page_size,
                "status": '',
                "search": ''
            }
            sno = (int(page) - 1) * int(page_size) + 1
            list_url = settings.API_URL + '/api-notifications/subdomain-template-listing/'
            list_data = call_api_post_method(list_param, list_url, token)

            if 'error' in list_data and list_data['error'] == 0:
                template_list = list_data['data']['data']
                total = list_data['data']['total']
            else:
                template_list = []
                total = 0
            # ---------------Pagination--------
            pagination_html = ''
            pagination_path = 'admin/dashboard/email_template/pagination.html'
            pagination_template = get_template(pagination_path)
            total_page = math.ceil(total / page_size)
            if total_page > 1:
                pagination_data = {"no_page": total_page, "total_page": range(total_page), "current_page": 1,
                                   "pagination_id": "template_pagination_list"}
                pagination_html = pagination_template.render(pagination_data)
            context = {'template_list': template_list, 'total': total, "pagination_html": pagination_html,
                       "pagination_id": "template_pagination_list", "active_menu": "cms", "active_submenu": "template", "sno": sno}


            return render(request, "admin/dashboard/email_template/template-listing.html", context)
    except Exception as exp:
        print(exp)
        return HttpResponse("Issue in views")

def add_template(request):
    is_permission = check_permission(request, 15)
    if not is_permission:
        http_host = request.META['HTTP_HOST']
        redirect_url = settings.URL_SCHEME + str(http_host)
        return HttpResponseRedirect(redirect_url)
    template_id = request.GET.get('id', None)
    try:
        site_detail = subdomain_site_details(request)
        site_id = site_detail['site_detail']['site_id']
        token = request.session['token']['access_token']
        user_id = request.session['user_id']

        try:
            event_param = {}
            event_url = settings.API_URL + '/api-settings/active-event-listing/'
            event_data = call_api_post_method(event_param, event_url, token)
            event_list = event_data['data']
        except:
            event_list = []

        try:
            template_param = {
                'site': site_id,
                'user_id': user_id,
                'template_id': template_id,
            }
            template_url = settings.API_URL + '/api-notifications/subdomain-template-detail/'
            template_data = call_api_post_method(template_param, template_url, token)
            if 'error' in template_data and template_data['error'] == 1 and template_id:

                http_host = request.META['HTTP_HOST']
                redirect_url = settings.URL_SCHEME + str(http_host)
                return HttpResponseRedirect(redirect_url+'/admin/email-template/')

            template_details = template_data['data']
            template_id = template_details['id']

        except Exception as exp:
            print(exp)
            template_details = {}

        if request.is_ajax() and request.method == 'POST':
            add_template_params = {
                'site': site_id,
                'user_id': user_id,
                'event': request.POST['event_name'],
                'email_subject': request.POST['subject'],
                'email_content': request.POST['email_content'],
                'status': request.POST['template_status'],
                'template_id': request.POST['template_id']
            }
            add_template_url = settings.API_URL + '/api-notifications/subdomain-add-template/'

            add_template_response = call_api_post_method(add_template_params, add_template_url, token)

            if 'error' in add_template_response and add_template_response['error'] == 0:
                response = {
                    'error': 0,
                    'msg': add_template_response['msg']
                }
            else:
                response = {
                    'error': 1,
                    'msg': add_template_response['msg']
                }
            return JsonResponse(response)


        context = {"active_menu": "template", "event_list": event_list, 'template_details': template_details}
        return render(request, "admin/dashboard/email_template/add-template.html", context)
    except Exception as exp:
        print(exp)
        return HttpResponse("Issue in views")

@csrf_exempt
def template_search_suggestion(request):
    try:
        if request.is_ajax() and request.method == 'POST':
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']
            token = request.session['token']['access_token']

            params = {
                'domain': site_id,
                'search': request.POST['search']
            }

            api_url = settings.API_URL + '/api-notifications/subdomain-template-suggestion/'

            suggestion_data = call_api_post_method(params, api_url, token)
            if 'error' in suggestion_data and suggestion_data['error'] == 0:
                data = {'status': 200, 'suggestion_list': suggestion_data['data'], 'error': 0}
            else:
                data = {'status': 403, 'suggestion_list': [], 'error': 1}
        else:
            data = {'status': 403, 'suggestion_list': [], 'error': 1}

        return JsonResponse(data)
    except Exception as exp:
        print(exp)
        data = {'status': 403, 'suggestion_list': [], 'error': 1}
        return JsonResponse(data)

@csrf_exempt
def get_auction_type_status(request):
    try:
        if request.is_ajax() and request.method == 'POST':
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']
            token = request.session['token']['access_token']
            if 'auction_type' in request.POST and int(request.POST['auction_type']) == 1:
                status_param = {'object_id': 10}
            elif 'auction_type' in request.POST and (int(request.POST['auction_type']) == 4 or int(request.POST['auction_type']) == 7):
                status_param = {'object_id': 11}
            elif 'auction_type' in request.POST and int(request.POST['auction_type']) == 6:
                status_param = {'object_id': 13}
            else:
                status_param = {}


            status_url = settings.API_URL + '/api-settings/lookup-status-listing/'
            status_data = call_api_post_method(status_param, status_url, token)

            if 'error' in status_data and status_data['error'] == 0:
                status_list = status_data['data']
                data = {'status': 200, 'status_list': status_list, 'error': 0}
            else:
                data = {'status': 403, 'status_list': [], 'error': 1}
        else:
            data = {'status': 403, 'status_list': [], 'error': 1}

        return JsonResponse(data)
    except Exception as exp:
        print(exp)
        data = {'status': 403, 'status_list': [], 'error': 1}
        return JsonResponse(data)


@csrf_exempt
def property_offer_list(request):
    try:
        user_id = user_id = None
        page_size = 10
        try:
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']

        except Exception as exp:
            print(exp)
            site_id = ""

        if 'user_id' in request.session and request.session['user_id']:
            token = request.session['token']['access_token']
            user_id = request.session['user_id']

        if request.is_ajax() and request.method == 'POST':
            page = 1
            if 'page' in request.POST and request.POST['page'] != "":
                page = request.POST['page']

            if 'page_size' in request.POST and request.POST['page_size'] != "":
                page_size = request.POST['page_size']

            property_id = request.POST['property_id']
            params = {
                'domain_id': site_id,
                'user_id': user_id,
                'property_id': property_id,
            }

            api_url = settings.API_URL + '/api-bid/seller-offer-listing/'
            seller_offer_list = call_api_post_method(params, api_url, token=token)


            if 'error' in seller_offer_list and seller_offer_list['error'] == 0:
                seller_offers = seller_offer_list['data']['data']

            else:
                seller_offers = []

            context = {'seller_offer_list': seller_offers, "aws_url": settings.AWS_URL}

            offer_listing_path = 'admin/dashboard/listings/offer_listing_content.html'
            offer_listing_template = get_template(offer_listing_path)
            offer_history_html = offer_listing_template.render(context)

            try:
                if len(seller_offers) > 0:
                    negotiated_id = seller_offers[0]['id']
                    params = {
                        "domain_id": site_id,
                        "negotiated_id": negotiated_id,
                        "user_id": user_id
                    }

                    api_url = settings.API_URL + '/api-bid/seller-offer-detail/'
                    offer_data = call_api_post_method(params, api_url, token)
                    if 'error' in offer_data and offer_data['error'] == 0:
                        offer_details = offer_data['data']['data']
                        context = {'offer': offer_details, "aws_url": settings.AWS_URL, "user_id": user_id}

                        offer_details_path = 'admin/dashboard/listings/offer_details_content.html'
                        offer_details_template = get_template(offer_details_path)
                        offer_details_html = offer_details_template.render(context)
                    else:
                        offer_details_html = ''
                else:
                    offer_details_html = ''
            except Exception as exp:
                print(exp)
                offer_details_html= ''

            data = {
                'offer_history_html': offer_history_html,
                'offer_details_html': offer_details_html,
                'status': 200,
                'msg': '',
                'error': 0
            }
        else:
            data = {'status': 403, 'msg': 'Forbidden', 'error': 1}
        return JsonResponse(data)
    except Exception as exp:
        print(exp)
        return HttpResponse("Issue in views")

@csrf_exempt
def property_offer_details(request):
    try:
        if request.is_ajax() and request.method == 'POST':
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']
            token = request.session['token']['access_token']
            user_id = request.session['user_id']

            params = {
                "domain_id": site_id,
                "negotiated_id": request.POST['negotiated_id'],
                "user_id": user_id
            }

            api_url = settings.API_URL + '/api-bid/seller-offer-detail/'

            offer_data = call_api_post_method(params, api_url, token)

            if 'error' in offer_data and offer_data['error'] == 0:
                offer_details = offer_data['data']['data']
                context = {'offer': offer_details, "aws_url": settings.AWS_URL,"session_user_id": user_id}

                offer_details_path = 'admin/dashboard/listings/offer_details_content.html'
                offer_details_template = get_template(offer_details_path)
                offer_details_html = offer_details_template.render(context)
            else:
                offer_details_html = ''

            data = {'status': 200, 'offer_details_html': offer_details_html, 'error': 0}
        else:
            data = {'status': 403, 'offer_details_html': '', 'error': 1}

        return JsonResponse(data)
    except Exception as exp:
        print(exp)
        data = {'status': 403, 'offer_details_html': '', 'error': 1}
        return JsonResponse(data)


@csrf_exempt
def seller_accept_offer(request):
    try:
        site_detail = subdomain_site_details(request)
        site_id = site_detail['site_detail']['site_id']
        token = request.session['token']['access_token']
        user_id = request.session['user_id']

        if request.is_ajax() and request.method == 'POST':
            params = {
                'domain_id': site_id,
                'user_id': user_id,
                'property_id': request.POST['property_id'],
                'negotiation_id': request.POST['negotiated_id'],
            }
            url = settings.API_URL + '/api-bid/seller-accept-offer/'

            response = call_api_post_method(params, url, token)

            if 'error' in response and response['error'] == 0:
                data = {
                    'error': 0,
                    'msg': "Offer Accepted successfully."
                }
            else:
                data = {
                    'error': 1,
                    'msg': response['msg']
                }
            return JsonResponse(data)
    except Exception as exp:
        print(exp)
        return HttpResponse("Issue in views")

@csrf_exempt
def seller_reject_offer(request):
    try:
        site_detail = subdomain_site_details(request)
        site_id = site_detail['site_detail']['site_id']
        token = request.session['token']['access_token']
        user_id = request.session['user_id']

        if request.is_ajax() and request.method == 'POST':
            params = {
                'domain_id': site_id,
                'user_id': user_id,
                'property_id': request.POST['property_id'],
                'negotiation_id': request.POST['negotiated_id'],
            }
            url = settings.API_URL + '/api-bid/seller-reject-offer/'

            response = call_api_post_method(params, url, token)

            if 'error' in response and response['error'] == 0:
                data = {
                    'error': 0,
                    'msg': "Offer Rejected successfully."
                }
            else:
                data = {
                    'error': 1,
                    'msg': response['msg']
                }
            return JsonResponse(data)
    except Exception as exp:
        print(exp)
        return HttpResponse("Issue in views")

@csrf_exempt
def seller_counter_offer(request):
    try:
        site_detail = subdomain_site_details(request)
        site_id = site_detail['site_detail']['site_id']
        token = request.session['token']['access_token']
        user_id = request.session['user_id']

        if request.is_ajax() and request.method == 'POST':
            if 'offer_price' in request.POST and request.POST['offer_price']:
                offer_price = request.POST['offer_price'].replace(',', '').replace('$', '')
            else:
                offer_price = ''

            property_id = request.POST['property_id']
            negotiated_id = request.POST['negotiated_id']

            params = {
                "domain_id": site_id,
                "property_id": request.POST['property_id'],
                "negotiation_id": request.POST['negotiated_id'],
                "offer_price": offer_price,
                "comment": request.POST['offer_comment'],
                "user_id": user_id
            }
            url = settings.API_URL + '/api-bid/seller-counter-offer/'

            response = call_api_post_method(params, url, token)

            if 'error' in response and response['error'] == 0:
                try:
                    page = 1
                    page_size = 10

                    params = {
                        'domain_id': site_id,
                        'user_id': user_id,
                        'property_id': property_id,
                    }

                    api_url = settings.API_URL + '/api-bid/seller-offer-listing/'
                    seller_offer_list = call_api_post_method(params, api_url, token=token)


                    if 'error' in seller_offer_list and seller_offer_list['error'] == 0:
                        seller_offers = seller_offer_list['data']['data']

                    else:
                        seller_offers = []

                    context = {'seller_offer_list': seller_offers, "aws_url": settings.AWS_URL}

                    offer_listing_path = 'admin/dashboard/listings/offer_listing_content.html'
                    offer_listing_template = get_template(offer_listing_path)
                    offer_history_html = offer_listing_template.render(context)

                    try:
                        if len(seller_offers) > 0:
                            params = {
                                "domain_id": site_id,
                                "negotiated_id": negotiated_id,
                                "user_id": user_id
                            }

                            api_url = settings.API_URL + '/api-bid/seller-offer-detail/'
                            offer_data = call_api_post_method(params, api_url, token)
                            if 'error' in offer_data and offer_data['error'] == 0:
                                offer_details = offer_data['data']['data']
                                context = {'offer': offer_details, "aws_url": settings.AWS_URL, "user_id": user_id,"session_user_id": user_id}

                                offer_details_path = 'admin/dashboard/listings/offer_details_content.html'
                                offer_details_template = get_template(offer_details_path)
                                offer_details_html = offer_details_template.render(context)
                            else:
                                offer_details_html = ''
                        else:
                            offer_details_html = ''
                    except Exception as exp:
                        print(exp)
                        offer_details_html = ''

                    offer_data = {
                        'offer_history_html': offer_history_html,
                        'offer_details_html': offer_details_html,
                    }
                except:
                    offer_data = {
                        'offer_history_html': '',
                        'offer_details_html': '',
                    }
                data = {
                    'error': 0,
                    'msg': "Counter Offer successfully.",
                    'status': 200,
                    'offer': offer_data,
                    'negotiated_id': negotiated_id
                }
            else:
                data = {
                    'error': 1,
                    'msg': response['msg'],
                    'status': 400
                }
            return JsonResponse(data)
    except Exception as exp:
        print(exp)
        return HttpResponse("Issue in views")

@csrf_exempt
def counter_offer_details(request):
    try:
        if request.is_ajax() and request.method == 'POST':
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']
            token = request.session['token']['access_token']
            user_id = request.session['user_id']

            params = {
                "domain_id": site_id,
                "negotiated_id": request.POST['negotiated_id'],
                "user_id": user_id
            }

            api_url = settings.API_URL + '/api-bid/seller-offer-detail/'


            offer_data = call_api_post_method(params, api_url, token)


            if 'error' in offer_data and offer_data['error'] == 0:
                offer_details = offer_data['data']['data']
                offer = {
                    'property_id': offer_details['property_detail']['id'],
                    'negotiated_id': offer_details['master_offer'],
                    'offer_amount': offer_details['offer_price']
                }
                data = {'status': 200, 'offer': offer, 'error': 0}
            else:
                data = {'status': 200, 'offer': {}, 'error': 1,'msg': offer_data['msg']}


        else:
            data = {'status': 403, 'offer_details': {}, 'error': 1, 'msg': 'forbidden'}

        return JsonResponse(data)
    except Exception as exp:
        print(exp)
        data = {'status': 403, 'offer_details': {}, 'error': 1, 'msg': 'forbidden'}
        return JsonResponse(data)

@csrf_exempt
def get_offer_document_details(request):
    try:
        if request.is_ajax() and request.method == 'POST':
            try:
                site_detail = subdomain_site_details(request)
                site_id = site_detail['site_detail']['site_id']
                templete_dir = site_detail['site_detail']['theme_directory']

            except Exception as exp:
                print(exp)
                site_id = ""
                templete_dir = 'theme-1'

            user_id = None
            token = None
            if 'user_id' in request.session and request.session['user_id']:
                token = request.session['token']['access_token']
                user_id = request.session['user_id']

            params = {
                "domain": site_id,
                "property_id": request.POST['property_id'],
                "negotiation_id": request.POST['negotiated_id'],
            }
            api_url = settings.API_URL + '/api-bid/get-offer-documents/'
            document_data = call_api_post_method(params, api_url, token=token)
            if 'error' in document_data and document_data['error'] == 0:
                document_list = document_data['data']
                if document_list:
                    for doc in document_list:
                        document_name = doc['doc_file_name']
                        original_doc_arr = document_name.split('_')
                        original_doc_ext = document_name.split('.')[-1]
                        doc_name_length = len(original_doc_arr)
                        # original_doc_name = document_name.split('_')[1]
                        original_doc_name = ''
                        for i in range(doc_name_length):
                            if i > 0:
                                original_doc_name = original_doc_name + '_' + original_doc_arr[i]
                        original_doc_name = original_doc_name.lstrip('_')
                        doc['original_doc_name'] = original_doc_name
                        doc['extension'] = original_doc_ext
                context = {'document_list': document_list,"aws_url": settings.AWS_URL}

                doc_listing_path = 'admin/dashboard/listings/traditional_offer_document_content.html'.format(templete_dir)
                doc_listing_template = get_template(doc_listing_path)
                doc_listing_html = doc_listing_template.render(context)
                data = {'doc_listing_html': doc_listing_html, 'error': 0, 'msg':'success'}
            else:
                data = {'doc_listing_html': '', 'error': 1, 'msg':'failure'}

            return JsonResponse(data)

    except Exception as exp:
        print(exp)
        return HttpResponse("Issue in views")

def chat_to_buyer(request):
    try:
        site_detail = subdomain_site_details(request)
        site_id = site_detail['site_detail']['site_id']
        token = request.session['token']['access_token']
        user_id = request.session['user_id']

        if request.is_ajax() and request.method == 'POST':
            params = {
                'site_id': site_id,
                'seller_id': user_id,
                'property_id': request.POST['property_id'],
                'user_id': request.POST['user_id'],
                'message': request.POST['user_message'],
            }
            url = settings.API_URL + '/api-contact/chat-to-buyer/'

            response = call_api_post_method(params, url, token)

            if 'error' in response and response['error'] == 0:
                data = {
                    'error': 0,
                    'msg': "Message Send successfully."
                }
            else:
                data = {
                    'error': 1,
                    'msg': response['msg']
                }
            return JsonResponse(data)
    except Exception as exp:
        print(exp)
        return HttpResponse("Issue in views")

@csrf_exempt
def download_listing(request):
    try:
        """
            Downloads all listings as Excel file with a single worksheet
        """
        try:
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']

        except Exception as exp:
            print(exp)
            site_id = ""

        user_id = None
        token = None
        is_broker = 0
        if 'user_id' in request.session and request.session['user_id']:
            token = request.session['token']['access_token']
            user_id = request.session['user_id']
            is_broker = 1 if request.session['is_broker'] == True else 0

        search = request.GET.get('search', '')
        page = request.GET.get('page', 1)
        page_size = request.GET.get('page_size', 10)
        asset_type = request.GET.get('asset_type', '')
        auction_type = request.GET.get('auction_type', '')
        property_type = request.GET.get('property_type', '')
        status = request.GET.get('status', '')
        agent = request.GET.get('agent', '')


        list_param = {
            "page": page,
            "page_size": page_size,
            "site_id": site_id,
            "user_id": user_id,
            "auction_id": auction_type,
            "asset_id": asset_type,
            "property_type": property_type,
            "search": search,
            "status": status,
            "agent_id": agent,
        }

        list_url = settings.API_URL + '/api-property/property-listing/'

        list_data = call_api_post_method(list_param, list_url, token)

        if 'error' in list_data and list_data['error'] == 0:
            property_listing = list_data['data']['data']
            total = list_data['data']['total'] if 'total' in list_data['data'] else 0
        else:
            property_listing = []
            total = 0
        sno = (int(page) - 1) * int(page_size) + 1


        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename={date}-listing.xlsx'.format(
            date=datetime.datetime.now().strftime('%Y-%m-%d'),
        )
        workbook = Workbook()

        # Get active worksheet/tab
        worksheet = workbook.active
        worksheet.title = 'Listings'

        # Define the titles for columns
        columns = [
            'ID',
            'Property',
            'Auction Type',
            'No. of view',
        ]
        row_num = 1
        header_font = Font(name='Calibri', bold=True)
        # Assign the titles for each cell of the header
        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title
            cell.font = header_font

        # Iterate through all movies
        count = 0
        for property in property_listing:
            row_num += 1


            # Define the data for each cell in the row
            row = [
                sno,
                property['name'],
                property['auction_type'],
                property['no_view'],
            ]
            sno += 1

            # Assign the data for each cell of the row
            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value


        workbook.save(response)

        return response
    except Exception as exp:
        print(exp)
        return HttpResponse("Issue in views")

@csrf_exempt
def property_best_offer_list(request):
    try:
        user_id = user_id = None
        page_size = 10
        try:
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']

        except Exception as exp:
            print(exp)
            site_id = ""

        if 'user_id' in request.session and request.session['user_id']:
            token = request.session['token']['access_token']
            user_id = request.session['user_id']

        if request.is_ajax() and request.method == 'POST':
            page = 1
            if 'page' in request.POST and request.POST['page'] != "":
                page = request.POST['page']

            if 'page_size' in request.POST and request.POST['page_size'] != "":
                page_size = request.POST['page_size']

            property_id = request.POST['property_id']
            params = {
                'domain_id': site_id,
                'user_id': user_id,
                'property_id': property_id,
            }

            api_url = settings.API_URL + '/api-bid/seller-offer-listing/'
            seller_offer_list = call_api_post_method(params, api_url, token=token)

            if 'error' in seller_offer_list and seller_offer_list['error'] == 0:
                seller_offers = seller_offer_list['data']['data']

            else:
                seller_offers = []

            context = {'seller_offer_list': seller_offers, "aws_url": settings.AWS_URL}

            offer_listing_path = 'admin/dashboard/listings/best_offer_listing_content.html'
            offer_listing_template = get_template(offer_listing_path)
            offer_history_html = offer_listing_template.render(context)

            try:
                if len(seller_offers) > 0:
                    negotiated_id = seller_offers[0]['id']
                    params = {
                        "domain_id": site_id,
                        "negotiated_id": negotiated_id,
                        "user_id": user_id
                    }

                    api_url = settings.API_URL + '/api-bid/seller-offer-detail/'
                    offer_data = call_api_post_method(params, api_url, token)
                    if 'error' in offer_data and offer_data['error'] == 0:
                        offer_details = offer_data['data']['data']
                        context = {'offer': offer_details, "aws_url": settings.AWS_URL, "user_id": user_id}

                        offer_details_path = 'admin/dashboard/listings/best_offer_details_content.html'
                        offer_details_template = get_template(offer_details_path)
                        offer_details_html = offer_details_template.render(context)
                    else:
                        offer_details_html = ''
                else:
                    offer_details_html = ''
            except Exception as exp:
                print(exp)
                offer_details_html= ''


            data = {
                'offer_history_html': offer_history_html,
                'offer_details_html': offer_details_html,
                'status': 200,
                'msg': '',
                'error': 0
            }
        else:
            data = {'status': 403, 'msg': 'Forbidden', 'error': 1}
        return JsonResponse(data)
    except Exception as exp:
        print(exp)
        return HttpResponse("Issue in views")

@csrf_exempt
def property_best_offer_details(request):
    try:
        if request.is_ajax() and request.method == 'POST':
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']
            token = request.session['token']['access_token']
            user_id = request.session['user_id']

            params = {
                "domain_id": site_id,
                "negotiated_id": request.POST['negotiated_id'],
                "user_id": user_id
            }

            api_url = settings.API_URL + '/api-bid/enhanced-best-seller-offer-detail/'

            offer_data = call_api_post_method(params, api_url, token)

            if 'error' in offer_data and offer_data['error'] == 0:
                offer_details = offer_data['data']['data']
                try:
                    if 'offer_detail' in offer_details and offer_details['offer_detail']:
                        offer_start_price = 0
                        if 'offer_price' in offer_details and offer_details['offer_price']:
                            offer_start_price = offer_details['offer_price']

                        if 'earnest_deposit_type' in offer_details['offer_detail'] and offer_details['offer_detail'][
                            'earnest_deposit_type'] == 1:
                            offer_earnest_deposit = offer_details['offer_detail']['earnest_money_deposit']
                            current_earnest_money_deposit = format_currency(offer_earnest_deposit)
                            offer_money_deposit_percent = (float(offer_earnest_deposit) * 100) / float(
                                offer_start_price)
                            current_earnest_money_deposit_percent = format_currency(offer_money_deposit_percent)
                        elif 'earnest_deposit_type' in offer_details['offer_detail'] and offer_details['offer_detail'][
                            'earnest_deposit_type'] == 2:
                            offer_money_deposit_percent = offer_details['offer_detail']['earnest_money_deposit']
                            current_earnest_money_deposit_percent = format_currency(offer_money_deposit_percent)
                            offer_earnest_deposit = (float(offer_start_price) * float(
                                offer_money_deposit_percent)) / float(100)
                            current_earnest_money_deposit = format_currency(offer_earnest_deposit)
                        else:
                            current_earnest_money_deposit = ''
                            current_earnest_money_deposit_percent = ''

                        offer_details['offer_detail'][
                            'current_earnest_money_deposit'] = current_earnest_money_deposit
                        offer_details['offer_detail'][
                            'current_earnest_money_deposit_percent'] = current_earnest_money_deposit_percent
                except:
                    pass
                context = {'offer': offer_details, "aws_url": settings.AWS_URL, 'session_user_id': user_id}

                offer_details_path = 'admin/dashboard/listings/best_offer_details_content.html'
                offer_details_template = get_template(offer_details_path)
                offer_details_html = offer_details_template.render(context)
            else:
                offer_details_html = ''

            data = {'status': 200, 'offer_details_html': offer_details_html, 'error': 0}
        else:
            data = {'status': 403, 'offer_details_html': '', 'error': 1}

        return JsonResponse(data)
    except Exception as exp:
        print(exp)
        data = {'status': 403, 'offer_details_html': '', 'error': 1}
        return JsonResponse(data)

@csrf_exempt
def export_bid_history(request):
    try:
        """
            Downloads bid history as Excel file with a single worksheet
        """
        try:
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']

        except Exception as exp:
            print(exp)
            site_id = ""

        user_id = None
        token = None
        is_broker = 0
        if 'user_id' in request.session and request.session['user_id']:
            token = request.session['token']['access_token']
            user_id = request.session['user_id']
            is_broker = 1 if request.session['is_broker'] == True else 0

        search = request.GET.get('search', '')
        page = request.GET.get('page', 1)
        page_size = request.GET.get('page_size', 10)
        property_id = request.GET.get('property', '')
        timezone = request.GET.get('timezone', '')
        register_user = request.GET.get('register_user', '')

        list_param = {
            'site_id': site_id,
            'user_id': user_id,
            'page_size': page_size,
            'page': page,
            'property_id': property_id,
            'search': search,
            'register_user': register_user,
        }

        # list_url = settings.API_URL + '/api-bid/subdomain-bid-history/'
        list_url = settings.API_URL + '/api-bid/user-bid-history/'

        list_data = call_api_post_method(list_param, list_url, token=token)
        if 'error' in list_data and list_data['error'] == 0:
            bid_history = list_data['data']['data']
            # bid_history = list_data['data']['new_data']
            total = list_data['data']['total'] if 'total' in list_data['data'] else 0
        else:
            bid_history = []
            total = 0
        sno = (int(page) - 1) * int(page_size) + 1


        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename={date}-bid-history.xlsx'.format(
            date=datetime.datetime.now().strftime('%Y-%m-%d'),
        )
        workbook = Workbook()

        # Get active worksheet/tab
        worksheet = workbook.active
        worksheet.title = 'Bid History'

        # Define the titles for columns
        # columns = [
        #     '#',
        #     'Bidder Name',
        #     'Email',
        #     'Phone',
        #     'Start Bids',
        #     'High Bids',
        #     'Bids',
        #     'IP Address',
        #     'Bidding Date',
        #     'Approval Status',
        # ]
        columns = [
            '#',
            'Bidder Name',
            'Email',
            'Phone',
            'Bid Amount',
            'IP Address',
            'Bidding Date',
        ]

        row_num = 1
        header_font = Font(name='Calibri', bold=True)
        # Assign the titles for each cell of the header
        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title
            cell.font = header_font

        # Iterate through all history
        count = 0
        for bids in bid_history:
            row_num += 1

            phone_no = bids['bidder_detail']['phone_no']
            # bid_amount = bids['bid_amount']
            formatted_phone_no = format_phone_number(phone_no)
            # formatted_amount = format_currency(bid_amount)
            # bid_date = bids['bid_time']
            bid_date = bids['bid_date']
            bid_date_time = ''
            if timezone:
                try:
                    added_on_time = time.mktime(
                        datetime.datetime.strptime(bid_date, "%Y-%m-%dT%H:%M:%SZ").timetuple())
                except ValueError:
                    added_on_time = time.mktime(
                        datetime.datetime.strptime(bid_date, "%Y-%m-%dT%H:%M:%S.%fZ").timetuple())
                except:
                    added_on_time = 0
            if added_on_time:
                added_on_time = float(added_on_time) - (float(timezone)*60)
                bid_date_time = datetime.datetime.fromtimestamp(added_on_time)
                bid_date_time = datetime.datetime.strftime(bid_date_time, "%m-%d-%Y %I:%M %p")
            # Define the data for each cell in the row
            # row = [
            #     sno,
            #     bids['bidder_detail']['first_name']+' '+bids['bidder_detail']['last_name'],
            #     bids['bidder_detail']['email'],
            #     formatted_phone_no,
            #     '$'+str(f"{formatted_amount:,}"),
            #     bids['ip_address'],
            #     bid_date_time,
            # ]
            row = [
                sno,
                bids['bidder_detail']['first_name'] + ' ' + bids['bidder_detail']['last_name'],
                bids['bidder_detail']['email'],
                formatted_phone_no,
                # '$' + str(f"{bids['start_bid']:,}"),
                # '$' + str(f"{bids['max_bid']:,}"),
                # bids['bids'],
                bids['bid_amount'],
                # bids['bidder_detail']['ip_address'],
                bids['ip_address'],
                bid_date_time,
                # 'Approved'
            ]
            sno += 1

            # Assign the data for each cell of the row
            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value


        workbook.save(response)

        return response
    except Exception as exp:
        print(exp)
        return HttpResponse("Issue in views")

@csrf_exempt
def delete_current_bid(request):
    try:
        site_detail = subdomain_site_details(request)
        site_id = site_detail['site_detail']['site_id']
        token = request.session['token']['access_token']
        user_id = request.session['user_id']
        property_id = request.POST['property_id']

        if request.is_ajax() and request.method == 'POST':
            params = {
                'domain': site_id,
                'user_id': user_id,
                'property_id': request.POST['property_id'],
            }
            url = settings.API_URL + '/api-bid/bid-delete/'

            response = call_api_post_method(params, url, token)

            if 'error' in response and response['error'] == 0:
                auction_id = response['data']['auction_id']
                property_id = response['data']['property_id']
                data = {
                    'error': 0,
                    'site_id': site_id,
                    'property_id': property_id,
                    'auction_id': auction_id,
                    'user_id': user_id,
                    'msg': "Deleted successfully."
                }
            else:
                data = {
                    'error': 1,
                    'msg': response['msg']
                }
            return JsonResponse(data)
    except Exception as exp:
        print(exp)
        return HttpResponse("Issue in views")

@csrf_exempt
def download_loi(request):
    try:
        """
            Downloads LOI
        """
        try:
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']
            templete_dir = site_detail['site_detail']['theme_directory']

        except Exception as exp:
            print(exp)
            site_id = ""
            templete_dir = 'theme-1'

        user_id = None
        token = None
        if 'user_id' in request.session and request.session['user_id']:
            token = request.session['token']['access_token']
            user_id = request.session['user_id']

        property = request.GET.get('property', '')
        user = request.GET.get('user', '')
        loi_param = {
            "domain": site_id,
            "user": user,
            "property": property
        }
        loi_url = settings.API_URL + '/api-bid/get-loi/'
        loi_data = call_api_post_method(loi_param, loi_url, token)
        offer_details = loi_data['data']
        try:
            if 'offer_detail' in offer_details and offer_details['offer_detail']:
                offer_start_price = 0

                if 'last_offer_price' in offer_details['offer_detail'] and \
                        offer_details['offer_detail']['last_offer_price']:
                    offer_start_price = offer_details['offer_detail']['last_offer_price']

                if 'earnest_deposit_type' in offer_details['property_detail'] and offer_details['property_detail'][
                    'earnest_deposit_type'] == 1:
                    offer_earnest_deposit = offer_details['offer_detail']['earnest_money_deposit']
                    current_earnest_money_deposit = format_currency(offer_earnest_deposit)
                    offer_money_deposit_percent = (float(offer_earnest_deposit) * 100) / float(offer_start_price)
                    current_earnest_money_deposit_percent = format_currency(offer_money_deposit_percent)
                elif 'earnest_deposit_type' in offer_details['property_detail'] and offer_details['property_detail'][
                    'earnest_deposit_type'] == 2:
                    offer_money_deposit_percent = offer_details['offer_detail']['earnest_money_deposit']
                    current_earnest_money_deposit_percent = format_currency(offer_money_deposit_percent)
                    offer_earnest_deposit = (float(offer_start_price) * float(offer_money_deposit_percent)) / float(100)
                    current_earnest_money_deposit = format_currency(offer_earnest_deposit)
                else:
                    current_earnest_money_deposit = ''
                    current_earnest_money_deposit_percent = ''

                offer_details['offer_detail']['current_earnest_money_deposit'] = current_earnest_money_deposit
                offer_details['offer_detail'][
                    'current_earnest_money_deposit_percent'] = current_earnest_money_deposit_percent

        except:
            pass
        context = {"offer_details": offer_details}

        offer_listing_path = 'home/download/loi_pdf.html'
        offer_listing_template = get_template(offer_listing_path)
        offer_listing_html = offer_listing_template.render(context)
        result = BytesIO()

        # This part will create the pdf.
        pdf = pisa.pisaDocument(BytesIO(offer_listing_html.encode("UTF-8")), result)
        if not pdf.err:
            #return HttpResponse(result.getvalue(), content_type='application/pdf')
            response = HttpResponse(result.getvalue(), content_type='application/pdf')
            response['Content-Disposition'] = 'attachment; filename={date}-offer.pdf'.format(
                date=datetime.datetime.now().strftime('%Y-%m-%d'),
            )
            return response
        return None

    except Exception as exp:
        print(exp)
        http_host = request.META['HTTP_HOST']
        redirect_url = settings.URL_SCHEME + str(http_host) + '/admin/user-best-offer-list/?id=' + property
        return HttpResponseRedirect(redirect_url)

@csrf_exempt
def seller_send_loi(request):
    try:
        site_detail = subdomain_site_details(request)
        site_id = site_detail['site_detail']['site_id']
        token = request.session['token']['access_token']
        user_id = request.session['user_id']

        if request.is_ajax() and request.method == 'POST':
            params = {
                "domain": site_id,
                "user_id": user_id,
                "negotiation_id": request.POST['negotiated_id'],
                "property_id": request.POST['property_id'],
                "email": request.POST['loi_email'],
                "message": request.POST['loi_comment']
            }

            url = settings.API_URL + '/api-bid/send-loi/'

            response = call_api_post_method(params, url, token)

            if 'error' in response and response['error'] == 0:
                data = {
                    'error': 0,
                    'msg': "Loi Send successfully."
                }
            else:
                data = {
                    'error': 1,
                    'msg': response['msg']
                }
            return JsonResponse(data)
    except Exception as exp:
        print(exp)
        return HttpResponse("Issue in views")

@csrf_exempt
def download_payment_invoice(request):
    try:
        """
            Downloads LOI
        """
        try:
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']
            templete_dir = site_detail['site_detail']['theme_directory']

        except Exception as exp:
            print(exp)
            site_id = ""
            templete_dir = 'theme-1'

        user_id = None
        token = None
        if 'user_id' in request.session and request.session['user_id']:
            token = request.session['token']['access_token']
            user_id = request.session['user_id']

        subscription_id = request.GET.get('id', '')
        timezone = request.GET.get('timezone', '')
        invoice_param = {
            "site_id": site_id,
            "user_id": user_id,
            "subscription_id": subscription_id
        }

        invoice_url = settings.API_URL + '/api-users/plan-billing-detail/'
        invoice_data = call_api_post_method(invoice_param, invoice_url, token)
        http_host = request.META['HTTP_HOST']
        base_url = settings.URL_SCHEME + str(http_host)
        try:
            invoice_details_data = invoice_data['data']['data']
            payment_date_time = ''
            if timezone:
                payment_date = invoice_details_data['payment_date']
                try:
                    added_on_time = time.mktime(
                        datetime.datetime.strptime(payment_date, "%Y-%m-%dT%H:%M:%SZ").timetuple())
                except ValueError:
                    added_on_time = time.mktime(
                        datetime.datetime.strptime(payment_date, "%Y-%m-%dT%H:%M:%S.%fZ").timetuple())
                except:
                    added_on_time = 0
            if added_on_time:
                added_on_time = float(added_on_time) - (float(timezone) * 60)
                payment_date_time = datetime.datetime.fromtimestamp(added_on_time)
                payment_date_time = datetime.datetime.strftime(payment_date_time, "%m-%d-%Y")
                invoice_details_data['timezone_payment_date'] = payment_date_time

            invoice_details = {'invoice': invoice_details_data,'base_url': base_url}
        except:
            invoice_details = {'invoice': {},'base_url': base_url}

        invoice_path = 'admin/invoice/invoice_pdf.html'
        invoice_template = get_template(invoice_path)
        invoice_html = invoice_template.render(invoice_details)

        result = BytesIO()

        # This part will create the pdf.
        pdf = pisa.pisaDocument(BytesIO(invoice_html.encode("ISO-8859-1")), result)
        if not pdf.err:
            #return HttpResponse(result.getvalue(), content_type='application/pdf')
            response = HttpResponse(result.getvalue(), content_type='application/pdf')
            response['Content-Disposition'] = 'attachment; filename={date}-invoice.pdf'.format(
                date=datetime.datetime.now().strftime('%Y-%m-%d'),
            )
            return response
        return None

        #return render(request, "admin/invoice/invoice_pdf.html", {'invoice': invoice_details_data,'base_url': base_url})

    except Exception as exp:
        print(exp)
        return HttpResponse("Issue in views")

def user_best_offer_list(request):
    is_permission = check_permission(request, 6)
    if not is_permission:
        http_host = request.META['HTTP_HOST']
        redirect_url = settings.URL_SCHEME + str(http_host)
        return HttpResponseRedirect(redirect_url)
    property_id = request.GET.get('id', None)
    try:
        site_detail = subdomain_site_details(request)
        site_id = site_detail['site_detail']['site_id']
        token = request.session['token']['access_token']
        user_id = request.session['user_id']

        params = {
            'domain_id': site_id,
            'user_id': user_id,
            'property_id': property_id,
        }

        api_url = settings.API_URL + '/api-bid/enhanced-best-seller-offer-listing/'
        seller_offer_list = call_api_post_method(params, api_url, token=token)

        if 'error' in seller_offer_list and seller_offer_list['error'] == 0:
            seller_offers = seller_offer_list['data']['data']
            current_highest_id = 0
            current_highest_price = 0
            if seller_offers:
                for offer in seller_offers:
                    if 'offer_price_detail' in offer and 'best_offer_is_accept' in offer['offer_price_detail'] and offer['offer_price_detail']['best_offer_is_accept'] == True:
                        if current_highest_price == 0:
                            current_highest_price = offer['offer_price_detail']['price']
                            current_highest_id = offer['id']
                        else:
                            if int(offer['offer_price_detail']['price']) >= int(current_highest_price):
                                current_highest_price = offer['offer_price_detail']['price']
                                current_highest_id = offer['id']

        else:
            seller_offers = []

        try:
            if len(seller_offers) > 0:
                negotiated_id = seller_offers[0]['id']
                params = {
                    "domain_id": site_id,
                    "negotiated_id": negotiated_id,
                    "user_id": user_id
                }

                api_url = settings.API_URL + '/api-bid/enhanced-best-seller-offer-detail/'
                offer_data = call_api_post_method(params, api_url, token)
                if 'error' in offer_data and offer_data['error'] == 0:
                    offer_details = offer_data['data']['data']
                    try:
                        if 'offer_detail' in offer_details and offer_details['offer_detail']:
                            offer_start_price = 0
                            if 'offer_price' in offer_details and offer_details['offer_price']:
                                offer_start_price = offer_details['offer_price']

                            if 'earnest_deposit_type' in offer_details['offer_detail'] and offer_details['offer_detail'][
                                'earnest_deposit_type'] == 1:
                                offer_earnest_deposit = offer_details['offer_detail']['earnest_money_deposit']
                                current_earnest_money_deposit = format_currency(offer_earnest_deposit)
                                offer_money_deposit_percent = (float(offer_earnest_deposit) * 100) / float(
                                    offer_start_price)
                                current_earnest_money_deposit_percent = format_currency(offer_money_deposit_percent)
                            elif 'earnest_deposit_type' in offer_details['offer_detail'] and offer_details['offer_detail'][
                                'earnest_deposit_type'] == 2:
                                offer_money_deposit_percent = offer_details['offer_detail']['earnest_money_deposit']
                                current_earnest_money_deposit_percent = format_currency(offer_money_deposit_percent)
                                offer_earnest_deposit = (float(offer_start_price) * float(
                                    offer_money_deposit_percent)) / float(100)
                                current_earnest_money_deposit = format_currency(offer_earnest_deposit)
                            else:
                                current_earnest_money_deposit = ''
                                current_earnest_money_deposit_percent = ''

                            offer_details['offer_detail'][
                                'current_earnest_money_deposit'] = current_earnest_money_deposit
                            offer_details['offer_detail'][
                                'current_earnest_money_deposit_percent'] = current_earnest_money_deposit_percent
                    except:
                        pass
                else:
                    offer_details = {}
            else:
                offer_details = {}
        except Exception as exp:
            print(exp)
            offer_details = {}

        context = {
            "active_menu": "listing setting",
            "active_submenu": "listing",
            'seller_offer_list': seller_offers,
            "aws_url": settings.AWS_URL,
            "offer": offer_details,
            "current_highest_id": current_highest_id,
            "current_highest_price": current_highest_price,
            "user_id": user_id,
            "session_user_id": user_id,
        }
        return render(request, "admin/dashboard/listings/user-best-offer-list.html", context)
    except Exception as exp:
        print(exp)
        return HttpResponse("Issue in views")

@csrf_exempt
def seller_approve_best_offer(request):
    try:
        site_detail = subdomain_site_details(request)
        site_id = site_detail['site_detail']['site_id']
        token = request.session['token']['access_token']
        user_id = request.session['user_id']

        if request.is_ajax() and request.method == 'POST':
            property_id = request.POST['property_id']
            negotiated_id = request.POST['negotiated_id']
            params = {
                'domain_id': site_id,
                'user_id': user_id,
                'property_id': request.POST['property_id'],
                'negotiation_id': request.POST['negotiated_id'],
            }
            url = settings.API_URL + '/api-bid/enhanced-seller-accept-loi/'

            response = call_api_post_method(params, url, token)

            if 'error' in response and response['error'] == 0:
                try:
                    page = 1
                    page_size = 10

                    params = {
                        'domain_id': site_id,
                        'user_id': user_id,
                        'property_id': property_id,
                    }

                    api_url = settings.API_URL + '/api-bid/enhanced-best-seller-offer-listing/'
                    seller_offer_list = call_api_post_method(params, api_url, token=token)

                    current_highest_id = 0
                    current_highest_price = 0
                    if 'error' in seller_offer_list and seller_offer_list['error'] == 0:
                        seller_offers = seller_offer_list['data']['data']
                        if seller_offers:
                            for offer in seller_offers:
                                if 'offer_price_detail' in offer and 'best_offer_is_accept' in offer['offer_price_detail'] and offer['offer_price_detail']['best_offer_is_accept'] == True:
                                    if current_highest_price == 0:
                                        current_highest_price = offer['offer_price_detail']['price']
                                        current_highest_id = offer['id']
                                    else:
                                        if int(offer['offer_price_detail']['price']) >= int(current_highest_price):
                                            current_highest_price = offer['offer_price_detail']['price']
                                            current_highest_id = offer['id']
                    else:
                        seller_offers = []

                    context = {'seller_offer_list': seller_offers, "aws_url": settings.AWS_URL,
                               "current_highest_id": current_highest_id, "current_highest_price": current_highest_price,"session_user_id": user_id}

                    offer_listing_path = 'admin/dashboard/listings/best_offer_listing_content.html'
                    offer_listing_template = get_template(offer_listing_path)
                    offer_history_html = offer_listing_template.render(context)
                except:
                    offer_history_html = ''
                try:
                    params = {
                        "domain_id": site_id,
                        "negotiated_id": negotiated_id,
                        "user_id": user_id
                    }

                    api_url = settings.API_URL + '/api-bid/enhanced-best-seller-offer-detail/'

                    offer_data = call_api_post_method(params, api_url, token)

                    if 'error' in offer_data and offer_data['error'] == 0:
                        offer_details = offer_data['data']['data']
                        try:
                            if 'offer_detail' in offer_details and offer_details['offer_detail']:
                                offer_start_price = 0
                                if 'offer_price' in offer_details and offer_details['offer_price']:
                                    offer_start_price = offer_details['offer_price']

                                if 'earnest_deposit_type' in offer_details['offer_detail'] and \
                                        offer_details['offer_detail'][
                                            'earnest_deposit_type'] == 1:
                                    offer_earnest_deposit = offer_details['offer_detail']['earnest_money_deposit']
                                    current_earnest_money_deposit = format_currency(offer_earnest_deposit)
                                    offer_money_deposit_percent = (float(offer_earnest_deposit) * 100) / float(
                                        offer_start_price)
                                    current_earnest_money_deposit_percent = format_currency(offer_money_deposit_percent)
                                elif 'earnest_deposit_type' in offer_details['offer_detail'] and \
                                        offer_details['offer_detail'][
                                            'earnest_deposit_type'] == 2:
                                    offer_money_deposit_percent = offer_details['offer_detail']['earnest_money_deposit']
                                    current_earnest_money_deposit_percent = format_currency(offer_money_deposit_percent)
                                    offer_earnest_deposit = (float(offer_start_price) * float(
                                        offer_money_deposit_percent)) / float(100)
                                    current_earnest_money_deposit = format_currency(offer_earnest_deposit)
                                else:
                                    current_earnest_money_deposit = ''
                                    current_earnest_money_deposit_percent = ''

                                offer_details['offer_detail'][
                                    'current_earnest_money_deposit'] = current_earnest_money_deposit
                                offer_details['offer_detail'][
                                    'current_earnest_money_deposit_percent'] = current_earnest_money_deposit_percent
                        except:
                            pass
                        context = {'offer': offer_details, "aws_url": settings.AWS_URL,"session_user_id": user_id}

                        offer_details_path = 'admin/dashboard/listings/best_offer_details_content.html'
                        offer_details_template = get_template(offer_details_path)
                        offer_details_html = offer_details_template.render(context)
                    else:
                        offer_details_html = ''
                except:
                    offer_details_html = ''
                data = {
                    'error': 0,
                    'msg': "Offer Accepted successfully.",
                    'offer_history_html': offer_history_html,
                    'offer_details_html': offer_details_html,
                    'negotiated_id': negotiated_id,
                }
            else:
                data = {
                    'error': 1,
                    'msg': response['msg']
                }
            return JsonResponse(data)
    except Exception as exp:
        print(exp)
        return HttpResponse("Issue in views")

@csrf_exempt
def seller_reject_best_offer(request):
    try:
        site_detail = subdomain_site_details(request)
        site_id = site_detail['site_detail']['site_id']
        token = request.session['token']['access_token']
        user_id = request.session['user_id']

        if request.is_ajax() and request.method == 'POST':
            property_id = request.POST['property_id']
            negotiated_id = request.POST['negotiated_id']
            params = {
                'domain_id': site_id,
                'user_id': user_id,
                'property_id': request.POST['property_id'],
                'negotiation_id': request.POST['negotiated_id'],
                'declined_reason': request.POST['reason'] if 'reason' in request.POST and request.POST['reason'] != "" else "",
            }
            url = settings.API_URL + '/api-bid/enhanced-seller-reject-offer/'

            response = call_api_post_method(params, url, token)
            if 'error' in response and response['error'] == 0:
                try:
                    page = 1
                    page_size = 10

                    params = {
                        'domain_id': site_id,
                        'user_id': user_id,
                        'property_id': property_id,
                    }

                    api_url = settings.API_URL + '/api-bid/enhanced-best-seller-offer-listing/'
                    seller_offer_list = call_api_post_method(params, api_url, token=token)

                    current_highest_id = 0
                    current_highest_price = 0
                    if 'error' in seller_offer_list and seller_offer_list['error'] == 0:
                        seller_offers = seller_offer_list['data']['data']
                        if seller_offers:
                            for offer in seller_offers:
                                if 'offer_price_detail' in offer and 'best_offer_is_accept' in offer['offer_price_detail'] and offer['offer_price_detail']['best_offer_is_accept'] == True:
                                    if current_highest_price == 0:
                                        current_highest_price = offer['offer_price_detail']['price']
                                        current_highest_id = offer['id']
                                    else:
                                        if int(offer['offer_price_detail']['price']) >= int(current_highest_price):
                                            current_highest_price = offer['offer_price_detail']['price']
                                            current_highest_id = offer['id']
                    else:
                        seller_offers = []

                    context = {'seller_offer_list': seller_offers, "aws_url": settings.AWS_URL,
                               "current_highest_id": current_highest_id, "current_highest_price": current_highest_price,"session_user_id": user_id}

                    offer_listing_path = 'admin/dashboard/listings/best_offer_listing_content.html'
                    offer_listing_template = get_template(offer_listing_path)
                    offer_history_html = offer_listing_template.render(context)
                except:
                    offer_history_html = ''

                try:
                    params = {
                        "domain_id": site_id,
                        "negotiated_id": negotiated_id,
                        "user_id": user_id
                    }

                    api_url = settings.API_URL + '/api-bid/enhanced-best-seller-offer-detail/'

                    offer_data = call_api_post_method(params, api_url, token)

                    if 'error' in offer_data and offer_data['error'] == 0:
                        offer_details = offer_data['data']['data']
                        try:
                            if 'offer_detail' in offer_details and offer_details['offer_detail']:
                                offer_start_price = 0
                                if 'offer_price' in offer_details and offer_details['offer_price']:
                                    offer_start_price = offer_details['offer_price']

                                if 'earnest_deposit_type' in offer_details['offer_detail'] and \
                                        offer_details['offer_detail'][
                                            'earnest_deposit_type'] == 1:
                                    offer_earnest_deposit = offer_details['offer_detail']['earnest_money_deposit']
                                    current_earnest_money_deposit = format_currency(offer_earnest_deposit)
                                    offer_money_deposit_percent = (float(offer_earnest_deposit) * 100) / float(
                                        offer_start_price)
                                    current_earnest_money_deposit_percent = format_currency(offer_money_deposit_percent)
                                elif 'earnest_deposit_type' in offer_details['offer_detail'] and \
                                        offer_details['offer_detail'][
                                            'earnest_deposit_type'] == 2:
                                    offer_money_deposit_percent = offer_details['offer_detail']['earnest_money_deposit']
                                    current_earnest_money_deposit_percent = format_currency(offer_money_deposit_percent)
                                    offer_earnest_deposit = (float(offer_start_price) * float(
                                        offer_money_deposit_percent)) / float(100)
                                    current_earnest_money_deposit = format_currency(offer_earnest_deposit)
                                else:
                                    current_earnest_money_deposit = ''
                                    current_earnest_money_deposit_percent = ''

                                offer_details['offer_detail'][
                                    'current_earnest_money_deposit'] = current_earnest_money_deposit
                                offer_details['offer_detail'][
                                    'current_earnest_money_deposit_percent'] = current_earnest_money_deposit_percent
                        except:
                            pass
                        context = {'offer': offer_details, "aws_url": settings.AWS_URL,"session_user_id": user_id}

                        offer_details_path = 'admin/dashboard/listings/best_offer_details_content.html'
                        offer_details_template = get_template(offer_details_path)
                        offer_details_html = offer_details_template.render(context)
                    else:
                        offer_details_html = ''
                except:
                    offer_details_html = ''
                data = {
                    'error': 0,
                    'msg': "Offer Rejected successfully.",
                    'offer_history_html': offer_history_html,
                    'offer_details_html': offer_details_html,
                    'negotiated_id': negotiated_id,
                }
            else:
                data = {
                    'error': 1,
                    'msg': response['msg']
                }
            return JsonResponse(data)
    except Exception as exp:
        print(exp)
        return HttpResponse("Issue in views")

@csrf_exempt
def seller_counter_best_offer(request):
    try:
        site_detail = subdomain_site_details(request)
        site_id = site_detail['site_detail']['site_id']
        token = request.session['token']['access_token']
        user_id = request.session['user_id']

        if request.is_ajax() and request.method == 'POST':
            property_id = request.POST['property_id']
            negotiated_id = request.POST['negotiated_id']
            if 'offer_price' in request.POST and request.POST['offer_price']:
                offer_price = float(request.POST['offer_price'].replace('$', '').replace(',', ''))
            else:
                offer_price = ''

            if 'earnest_deposit' in request.POST and request.POST['earnest_deposit'] and request.POST['earnest_deposit'] != "$":
                earnest_deposit = float(request.POST['earnest_deposit'].replace('$', '').replace(',', ''))
            else:
                earnest_deposit = ''

            if 'down_payment' in request.POST and request.POST['down_payment'].replace(',', '').replace('$', '') != "" and request.POST['down_payment'] != "$":
                down_payment = float(request.POST['down_payment'].replace('$', '').replace(',', ''))
            else:
                down_payment = ''

            params = {
                "domain_id": site_id,
                "property_id": request.POST['property_id'],
                "negotiation_id": request.POST['negotiated_id'],
                "offer_price": offer_price,
                "offer_comment": request.POST['offer_comment'],
                "user_id": user_id,
                "earnest_money_deposit": earnest_deposit,
                "down_payment": down_payment,
                "due_diligence_period": request.POST['due_diligence'],
                "closing_period": request.POST['closing_period'],
                "financing": request.POST['financing'],
                "offer_contingent": request.POST['offer_contingent'] if "offer_contingent" in request.POST else "",
                "sale_contingency": request.POST['sale_contingency'] if "sale_contingency" in request.POST else "",
                "appraisal_contingent": request.POST['appraisal_contingent'] if "appraisal_contingent" in request.POST else "",
                "closing_cost": request.POST['closing_cost'] if "closing_cost" in request.POST else "",
            }
            url = settings.API_URL + '/api-bid/enhanced-best-seller-counter-offer/'

            response = call_api_post_method(params, url, token)
            if 'error' in response and response['error'] == 0:
                try:
                    page = 1
                    page_size = 10

                    params = {
                        'domain_id': site_id,
                        'user_id': user_id,
                        'property_id': property_id,
                    }

                    api_url = settings.API_URL + '/api-bid/enhanced-best-seller-offer-listing/'
                    seller_offer_list = call_api_post_method(params, api_url, token=token)

                    current_highest_id = 0
                    current_highest_price = 0
                    if 'error' in seller_offer_list and seller_offer_list['error'] == 0:
                        seller_offers = seller_offer_list['data']['data']
                        if seller_offers:
                            for offer in seller_offers:
                                if 'offer_price_detail' in offer and 'best_offer_is_accept' in offer['offer_price_detail'] and offer['offer_price_detail']['best_offer_is_accept'] == True:
                                    if current_highest_price == 0:
                                        current_highest_price = offer['offer_price_detail']['price']
                                        current_highest_id = offer['id']
                                    else:
                                        if int(offer['offer_price_detail']['price']) >= int(current_highest_price):
                                            current_highest_price = offer['offer_price_detail']['price']
                                            current_highest_id = offer['id']
                    else:
                        seller_offers = []

                    context = {'seller_offer_list': seller_offers, "aws_url": settings.AWS_URL, "current_highest_id": current_highest_id, "current_highest_price": current_highest_price}

                    offer_listing_path = 'admin/dashboard/listings/best_offer_listing_content.html'
                    offer_listing_template = get_template(offer_listing_path)
                    offer_history_html = offer_listing_template.render(context)

                    try:
                        if len(seller_offers) > 0:
                            params = {
                                "domain_id": site_id,
                                "negotiated_id": negotiated_id,
                                "user_id": user_id
                            }

                            api_url = settings.API_URL + '/api-bid/enhanced-best-seller-offer-detail/'
                            offer_data = call_api_post_method(params, api_url, token)
                            if 'error' in offer_data and offer_data['error'] == 0:
                                offer_details = offer_data['data']['data']
                                try:
                                    if 'offer_detail' in offer_details and offer_details['offer_detail']:
                                        offer_start_price = 0
                                        if 'offer_price' in offer_details and offer_details['offer_price']:
                                            offer_start_price = offer_details['offer_price']

                                        if 'earnest_deposit_type' in offer_details['offer_detail'] and \
                                                offer_details['offer_detail'][
                                                    'earnest_deposit_type'] == 1:
                                            offer_earnest_deposit = offer_details['offer_detail'][
                                                'earnest_money_deposit']
                                            current_earnest_money_deposit = format_currency(offer_earnest_deposit)
                                            offer_money_deposit_percent = (float(offer_earnest_deposit) * 100) / float(
                                                offer_start_price)
                                            current_earnest_money_deposit_percent = format_currency(
                                                offer_money_deposit_percent)
                                        elif 'earnest_deposit_type' in offer_details['offer_detail'] and \
                                                offer_details['offer_detail'][
                                                    'earnest_deposit_type'] == 2:
                                            offer_money_deposit_percent = offer_details['offer_detail'][
                                                'earnest_money_deposit']
                                            current_earnest_money_deposit_percent = format_currency(
                                                offer_money_deposit_percent)
                                            offer_earnest_deposit = (float(offer_start_price) * float(
                                                offer_money_deposit_percent)) / float(100)
                                            current_earnest_money_deposit = format_currency(offer_earnest_deposit)
                                        else:
                                            current_earnest_money_deposit = ''
                                            current_earnest_money_deposit_percent = ''

                                        offer_details['offer_detail'][
                                            'current_earnest_money_deposit'] = current_earnest_money_deposit
                                        offer_details['offer_detail'][
                                            'current_earnest_money_deposit_percent'] = current_earnest_money_deposit_percent
                                except:
                                    pass
                                context = {'offer': offer_details, "aws_url": settings.AWS_URL, "user_id": user_id}

                                offer_details_path = 'admin/dashboard/listings/best_offer_details_content.html'
                                offer_details_template = get_template(offer_details_path)
                                offer_details_html = offer_details_template.render(context)
                            else:
                                offer_details_html = ''
                        else:
                            offer_details_html = ''
                    except Exception as exp:
                        print(exp)
                        offer_details_html = ''

                    offer_data = {
                        'offer_history_html': offer_history_html,
                        'offer_details_html': offer_details_html,
                    }
                except:
                    offer_data = {
                        'offer_history_html': '',
                        'offer_details_html': '',
                    }
                data = {
                    'error': 0,
                    'msg': "Counter Offer successfully.",
                    'status': 200,
                    'offer': offer_data,
                    'negotiated_id': negotiated_id
                }
            else:
                data = {
                    'error': 1,
                    'msg': response['msg'],
                    'status': 400
                }
            return JsonResponse(data)
    except Exception as exp:
        print(exp)
        return HttpResponse("Issue in views")

def user_traditional_offer_list(request):
    is_permission = check_permission(request, 6)
    if not is_permission:
        http_host = request.META['HTTP_HOST']
        redirect_url = settings.URL_SCHEME + str(http_host)
        return HttpResponseRedirect(redirect_url)
    property_id = request.GET.get('id', None)
    try:
        site_detail = subdomain_site_details(request)
        site_id = site_detail['site_detail']['site_id']
        token = request.session['token']['access_token']
        user_id = request.session['user_id']

        params = {
            'domain_id': site_id,
            'user_id': user_id,
            'property_id': property_id,
        }

        api_url = settings.API_URL + '/api-bid/seller-offer-listing/'
        seller_offer_list = call_api_post_method(params, api_url, token=token)

        if 'error' in seller_offer_list and seller_offer_list['error'] == 0:
            seller_offers = seller_offer_list['data']['data']

        else:
            seller_offers = []

        try:
            if len(seller_offers) > 0:
                negotiated_id = seller_offers[0]['id']
                params = {
                    "domain_id": site_id,
                    "negotiated_id": negotiated_id,
                    "user_id": user_id
                }

                api_url = settings.API_URL + '/api-bid/seller-offer-detail/'
                offer_data = call_api_post_method(params, api_url, token)
                if 'error' in offer_data and offer_data['error'] == 0:
                    offer_details = offer_data['data']['data']
                else:
                    offer_details = {}
            else:
                offer_details = {}
        except Exception as exp:
            print(exp)
            offer_details = {}

        context = {
            "active_menu": "listing setting",
            "active_submenu": "listing",
            'seller_offer_list': seller_offers,
            "aws_url": settings.AWS_URL,
            "offer": offer_details,
            "user_id": user_id,
            "session_user_id": user_id,
        }
        return render(request, "admin/dashboard/listings/user-traditional-offer-list.html", context)
    except Exception as exp:
        print(exp)
        return HttpResponse("Issue in views")

@csrf_exempt
def change_theme(request):
    try:
        site_detail = subdomain_site_details(request)
        site_id = site_detail['site_detail']['site_id']
        token = request.session['token']['access_token']
        user_id = request.session['user_id']

        if request.is_ajax() and request.method == 'POST':
            theme_id = request.POST['theme_id']
            params = {
                'domain': site_id,
                'user': user_id,
                'theme': request.POST['theme_id'],
            }
            url = settings.API_URL + '/api-users/change-theme/'

            response = call_api_post_method(params, url, token)

            if 'error' in response and response['error'] == 0:
                data = {
                    'error': 0,
                    'msg': "Theme Updated successfully.",
                }
            else:
                data = {
                    'error': 1,
                    'msg': response['msg']
                }
            return JsonResponse(data)
    except Exception as exp:
        print(exp)
        return HttpResponse("Issue in views")

@csrf_exempt
def seller_accept_best_offer(request):
    try:
        site_detail = subdomain_site_details(request)
        site_id = site_detail['site_detail']['site_id']
        token = request.session['token']['access_token']
        user_id = request.session['user_id']

        if request.is_ajax() and request.method == 'POST':
            property_id = request.POST['property_id']
            negotiated_id = request.POST['negotiated_id']
            params = {
                'domain_id': site_id,
                'user_id': user_id,
                'property_id': request.POST['property_id'],
                'negotiation_id': request.POST['negotiated_id'],
            }
            url = settings.API_URL + '/api-bid/enhanced-seller-accept-best-offer/'

            response = call_api_post_method(params, url, token)

            if 'error' in response and response['error'] == 0:
                try:
                    page = 1
                    page_size = 10

                    params = {
                        'domain_id': site_id,
                        'user_id': user_id,
                        'property_id': property_id,
                    }

                    api_url = settings.API_URL + '/api-bid/enhanced-best-seller-offer-listing/'
                    seller_offer_list = call_api_post_method(params, api_url, token=token)

                    current_highest_id = 0
                    current_highest_price = 0
                    if 'error' in seller_offer_list and seller_offer_list['error'] == 0:
                        seller_offers = seller_offer_list['data']['data']
                        if seller_offers:
                            for offer in seller_offers:
                                if 'offer_price_detail' in offer and 'best_offer_is_accept' in offer['offer_price_detail'] and offer['offer_price_detail']['best_offer_is_accept'] == True:
                                    if current_highest_price == 0:
                                        current_highest_price = offer['offer_price_detail']['price']
                                        current_highest_id = offer['id']
                                    else:
                                        if int(offer['offer_price_detail']['price']) >= int(current_highest_price):
                                            current_highest_price = offer['offer_price_detail']['price']
                                            current_highest_id = offer['id']
                    else:
                        seller_offers = []

                    context = {'seller_offer_list': seller_offers, "aws_url": settings.AWS_URL,
                               "current_highest_id": current_highest_id, "current_highest_price": current_highest_price,"session_user_id": user_id}

                    offer_listing_path = 'admin/dashboard/listings/best_offer_listing_content.html'
                    offer_listing_template = get_template(offer_listing_path)
                    offer_history_html = offer_listing_template.render(context)
                except:
                    offer_history_html = ''
                try:
                    params = {
                        "domain_id": site_id,
                        "negotiated_id": negotiated_id,
                        "user_id": user_id
                    }

                    api_url = settings.API_URL + '/api-bid/enhanced-best-seller-offer-detail/'

                    offer_data = call_api_post_method(params, api_url, token)

                    if 'error' in offer_data and offer_data['error'] == 0:
                        offer_details = offer_data['data']['data']
                        try:
                            if 'offer_detail' in offer_details and offer_details['offer_detail']:
                                offer_start_price = 0
                                if 'offer_price' in offer_details and offer_details['offer_price']:
                                    offer_start_price = offer_details['offer_price']

                                if 'earnest_deposit_type' in offer_details['offer_detail'] and \
                                        offer_details['offer_detail'][
                                            'earnest_deposit_type'] == 1:
                                    offer_earnest_deposit = offer_details['offer_detail']['earnest_money_deposit']
                                    current_earnest_money_deposit = format_currency(offer_earnest_deposit)
                                    offer_money_deposit_percent = (float(offer_earnest_deposit) * 100) / float(
                                        offer_start_price)
                                    current_earnest_money_deposit_percent = format_currency(offer_money_deposit_percent)
                                elif 'earnest_deposit_type' in offer_details['offer_detail'] and \
                                        offer_details['offer_detail'][
                                            'earnest_deposit_type'] == 2:
                                    offer_money_deposit_percent = offer_details['offer_detail']['earnest_money_deposit']
                                    current_earnest_money_deposit_percent = format_currency(offer_money_deposit_percent)
                                    offer_earnest_deposit = (float(offer_start_price) * float(
                                        offer_money_deposit_percent)) / float(100)
                                    current_earnest_money_deposit = format_currency(offer_earnest_deposit)
                                else:
                                    current_earnest_money_deposit = ''
                                    current_earnest_money_deposit_percent = ''

                                offer_details['offer_detail'][
                                    'current_earnest_money_deposit'] = current_earnest_money_deposit
                                offer_details['offer_detail'][
                                    'current_earnest_money_deposit_percent'] = current_earnest_money_deposit_percent
                        except:
                            pass
                        context = {'offer': offer_details, "aws_url": settings.AWS_URL,"session_user_id": user_id}

                        offer_details_path = 'admin/dashboard/listings/best_offer_details_content.html'
                        offer_details_template = get_template(offer_details_path)
                        offer_details_html = offer_details_template.render(context)
                    else:
                        offer_details_html = ''
                except:
                    offer_details_html = ''
                data = {
                    'error': 0,
                    'msg': "Offer Accepted successfully.",
                    'offer_history_html': offer_history_html,
                    'offer_details_html': offer_details_html,
                    'negotiated_id': negotiated_id,
                }
            else:
                data = {
                    'error': 1,
                    'msg': response['msg']
                }
            return JsonResponse(data)
    except Exception as exp:
        print(exp)
        return HttpResponse("Issue in views")

@csrf_exempt
def counter_best_offer_details(request):
    try:
        if request.is_ajax() and request.method == 'POST':
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']
            token = request.session['token']['access_token']
            user_id = request.session['user_id']

            params = {
                "domain_id": site_id,
                "property_id": request.POST['property_id'],
                "negotiation_id": request.POST['negotiated_id'],
                "user_id": user_id
            }

            api_url = settings.API_URL + '/api-bid/best-counter-seller-offer-detail/'

            offer_data = call_api_post_method(params, api_url, token)
            if 'error' in offer_data and offer_data['error'] == 0:
                response_data = offer_data['data']['data']
                start_price = 0
                if 'offer_price' in response_data and response_data['offer_price']:
                    start_price = response_data['offer_price']
                if 'earnest_money_deposit' in response_data and response_data['earnest_money_deposit'] and start_price:
                    if 'earnest_deposit_type' in response_data and response_data['earnest_deposit_type'] == 1:
                        earnest_money_deposit = response_data['earnest_money_deposit']
                        earnest_money_deposit = format_currency(earnest_money_deposit)
                        earnest_money_deposit_percent = (float(earnest_money_deposit) * 100) / float(start_price)
                        earnest_money_deposit_percent = format_currency(earnest_money_deposit_percent)
                    elif 'earnest_deposit_type' in response_data and response_data['earnest_deposit_type'] == 2:
                        earnest_money_deposit_percent = response_data['earnest_money_deposit']
                        earnest_money_deposit_percent = format_currency(earnest_money_deposit_percent)
                        earnest_money_deposit = (float(start_price) * float(response_data['earnest_money_deposit'])) / float(100)
                        earnest_money_deposit = format_currency(earnest_money_deposit)
                else:
                    earnest_money_deposit = ''
                    earnest_money_deposit_percent = ''

                if 'sale_contingency' in response_data and response_data['sale_contingency'] == True:
                    sale_contingency = 'Yes'
                elif 'sale_contingency' in response_data and response_data['sale_contingency'] == False:
                    sale_contingency = 'No'
                else:
                    sale_contingency = ''

                if 'appraisal_contingent' in response_data and response_data['appraisal_contingent'] == True:
                    appraisal_contingent = 'Yes'
                elif 'appraisal_contingent' in response_data and response_data['appraisal_contingent'] == False:
                    appraisal_contingent = 'No'
                else:
                    appraisal_contingent = ''

                if 'closing_cost' in response_data and response_data['closing_cost'] == 1:
                    closing_cost = 'Buyer agrees to pay for all loan-related closing costs and half of the transaction closing costs.'
                elif 'closing_cost' in response_data and response_data['closing_cost'] == 2:
                    closing_cost = 'Buyer agrees to pay for all loan-related closing costs and all of the transaction closing costs.'
                elif 'closing_cost' in response_data and response_data['closing_cost'] == 3:
                    closing_cost = 'Seller to pay for all loan-related closing costs and all of the transaction closing costs.'
                else:
                    closing_cost = ''

                try:
                    if 'property_asset' in response_data and int(response_data['property_asset']) == 1:
                        loan_type_list = land_loan_types
                    elif 'property_asset' in response_data and int(response_data['property_asset']) == 2:
                        loan_type_list = commerical_loan_types
                    elif 'property_asset' in response_data and int(response_data['property_asset']) == 3:
                        loan_type_list = residential_loan_types
                    else:
                        loan_type_list = []
                except Exception as exp:
                    print(exp)
                    loan_type_list = []

                offer = {
                    'property_id': response_data['property'],
                    'negotiated_id': response_data['negotiation_id'],
                    'offer_price': start_price,
                    'earnest_deposit_type': response_data[
                        'earnest_deposit_type'] if 'earnest_deposit_type' in response_data and response_data[
                        'earnest_deposit_type'] else "",
                    'earnest_money_deposit': earnest_money_deposit,
                    'earnest_money_deposit_percent': earnest_money_deposit_percent,
                    'due_diligence': response_data[
                        'due_diligence_period'] if 'due_diligence_period' in response_data and response_data[
                        'due_diligence_period'] else "",
                    'closing_date': response_data['closing_period'] if 'closing_period' in response_data and response_data[
                        'closing_period'] else "",
                    'loan_type': response_data['financing'] if 'financing' in response_data and response_data[
                        'financing'] else "",
                    'down_payment': response_data['down_payment'] if 'down_payment' in response_data and response_data[
                        'down_payment'] else "",
                    'offer_contingent': response_data['offer_contingent'] if 'offer_contingent' in response_data and
                                                                             response_data['offer_contingent'] else "",
                    'sale_contingency': sale_contingency,
                    'appraisal_contingent': appraisal_contingent,
                    'closing_cost': closing_cost,
                    'buyer_detail_info': response_data['buyer_detail']['buyer_detail'],
                    'agent_detail_info': response_data['buyer_detail']['agent_detail'],
                    'behalf_of_buyer': response_data['buyer_detail']['behalf_of_buyer'],
                    'loan_type_list': loan_type_list,
                }
                data = {'status': 200, 'data': offer, 'error': 0}
            else:
                data = {'status': 200, 'data': {}, 'error': 1,'msg': offer_data['msg']}


        else:
            data = {'status': 403, 'data': {}, 'error': 1, 'msg': 'forbidden'}

        return JsonResponse(data)
    except Exception as exp:
        print(exp)
        data = {'status': 403, 'data': {}, 'error': 1, 'msg': 'forbidden'}
        return JsonResponse(data)

@csrf_exempt
def best_offer_history_details(request):
    try:
        if request.is_ajax() and request.method == 'POST':
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']
            token = request.session['token']['access_token']
            user_id = request.session['user_id']

            params = {
                "domain_id": site_id,
                "property_id": request.POST['property_id'],
                "best_offers_id": request.POST['best_offers_id'],
                "user_id": user_id
            }

            api_url = settings.API_URL + '/api-bid/best-offer-history-detail/'

            offer_data = call_api_post_method(params, api_url, token)
            if 'error' in offer_data and offer_data['error'] == 0:
                response_data = offer_data['data']['data']
                start_price = 0
                if 'offer_price' in response_data and response_data['offer_price']:
                    start_price = response_data['offer_price']
                if 'earnest_money_deposit' in response_data and response_data['earnest_money_deposit'] and start_price:
                    if 'earnest_deposit_type' in response_data and response_data['earnest_deposit_type'] == 1:
                        earnest_money_deposit = response_data['earnest_money_deposit']
                        earnest_money_deposit = format_currency(earnest_money_deposit)
                        earnest_money_deposit_percent = (float(earnest_money_deposit) * 100) / float(start_price)
                        earnest_money_deposit_percent = format_currency(earnest_money_deposit_percent)
                    elif 'earnest_deposit_type' in response_data and response_data['earnest_deposit_type'] == 2:
                        earnest_money_deposit_percent = response_data['earnest_money_deposit']
                        earnest_money_deposit_percent = format_currency(earnest_money_deposit_percent)
                        earnest_money_deposit = (float(start_price) * float(response_data['earnest_money_deposit'])) / float(100)
                        earnest_money_deposit = format_currency(earnest_money_deposit)
                else:
                    earnest_money_deposit = ''
                    earnest_money_deposit_percent = ''

                if 'sale_contingency' in response_data and response_data['sale_contingency'] == True:
                    sale_contingency = 'Yes'
                elif 'sale_contingency' in response_data and response_data['sale_contingency'] == False:
                    sale_contingency = 'No'
                else:
                    sale_contingency = ''

                if 'appraisal_contingent' in response_data and response_data['appraisal_contingent'] == True:
                    appraisal_contingent = 'Yes'
                elif 'appraisal_contingent' in response_data and response_data['appraisal_contingent'] == False:
                    appraisal_contingent = 'No'
                else:
                    appraisal_contingent = ''

                if 'closing_cost' in response_data and response_data['closing_cost'] == 1:
                    closing_cost = 'Buyer agrees to pay for all loan-related closing costs and half of the transaction closing costs.'
                elif 'closing_cost' in response_data and response_data['closing_cost'] == 2:
                    closing_cost = 'Buyer agrees to pay for all loan-related closing costs and all of the transaction closing costs.'
                elif 'closing_cost' in response_data and response_data['closing_cost'] == 3:
                    closing_cost = 'Seller to pay for all loan-related closing costs and all of the transaction closing costs.'
                else:
                    closing_cost = ''
                offer = {
                    'property_id': response_data['property'],
                    'negotiated_id': response_data['negotiation_id'],
                    'offer_price': start_price,
                    'earnest_deposit_type': response_data[
                        'earnest_deposit_type'] if 'earnest_deposit_type' in response_data and response_data[
                        'earnest_deposit_type'] else "",
                    'earnest_money_deposit': earnest_money_deposit,
                    'earnest_money_deposit_percent': earnest_money_deposit_percent,
                    'due_diligence': response_data[
                        'due_diligence_period'] if 'due_diligence_period' in response_data and response_data[
                        'due_diligence_period'] else "",
                    'closing_date': response_data['closing_period'] if 'closing_period' in response_data and response_data[
                        'closing_period'] else "",
                    'loan_type': response_data['financing'] if 'financing' in response_data and response_data[
                        'financing'] else "",
                    'down_payment': response_data['down_payment'] if 'down_payment' in response_data and response_data[
                        'down_payment'] else "",
                    'offer_contingent': response_data['offer_contingent'] if 'offer_contingent' in response_data and
                                                                             response_data['offer_contingent'] else "",
                    'sale_contingency': sale_contingency,
                    'appraisal_contingent': appraisal_contingent,
                    'closing_cost': closing_cost,
                    'buyer_detail_info': response_data['buyer_detail']['buyer_detail'],
                    'agent_detail_info': response_data['buyer_detail']['agent_detail'],
                    'behalf_of_buyer': response_data['buyer_detail']['behalf_of_buyer'],
                    'declined_reason': response_data['declined_reason'] if 'declined_reason' in response_data and response_data['declined_reason'] else None,
                    'offer_msg': response_data['comments'] if 'comments' in response_data and response_data['comments'] else None,
                    'offer_by': response_data['offer_by'] if 'offer_by' in response_data and response_data['offer_by'] else None,
                }
                data = {'status': 200, 'data': offer, 'error': 0}
            else:
                data = {'status': 200, 'data': {}, 'error': 1,'msg': offer_data['msg']}
        else:
            data = {'status': 403, 'data': {}, 'error': 1, 'msg': 'forbidden'}
        return JsonResponse(data)
    except Exception as exp:
        print(exp)
        data = {'status': 403, 'data': {}, 'error': 1, 'msg': 'forbidden'}
        return JsonResponse(data)

@csrf_exempt
def check_user_pass(request):
    try:
        site_detail = subdomain_site_details(request)
        site_id = site_detail['site_detail']['site_id']
        token = request.session['token']['access_token']
        user_id = request.session['user_id']

        if request.is_ajax() and request.method == 'POST':
            params = {
                'domain': site_id,
                'user_id': user_id,
                'password': request.POST['user_pass'],
            }
            url = settings.API_URL + '/api-users/password-verify/'

            response = call_api_post_method(params, url, token)

            if 'error' in response and response['error'] == 0:
                data = {
                    'error': 0,
                    'msg': "Password Matched.",
                }
            else:
                data = {
                    'error': 1,
                    'msg': response['msg']
                }
            return JsonResponse(data)
    except Exception as exp:
        print(exp)
        return HttpResponse("Issue in views")

@csrf_exempt
def export_bidder_list(request):
    try:
        """
            Downloads bidder list as Excel file with a single worksheet from listing page
        """
        try:
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']

        except Exception as exp:
            print(exp)
            site_id = ""

        user_id = None
        token = None
        is_broker = 0
        if 'user_id' in request.session and request.session['user_id']:
            token = request.session['token']['access_token']
            user_id = request.session['user_id']
            is_broker = 1 if request.session['is_broker'] == True else 0

        search = request.GET.get('search', '')
        status = request.GET.get('status', '')
        page = request.GET.get('page', 1)
        page_size = request.GET.get('page_size', 10)
        property_id = request.GET.get('property', '')
        timezone = request.GET.get('timezone', '')

        list_param = {
            'site_id': site_id,
            'user_id': user_id,
            'page_size': page_size,
            'page': page,
            'filter_data': status,
            'property_id': property_id,
            'search': search,
        }

        list_url = settings.API_URL + '/api-bid/subdomain-bid-registration-listing/'

        list_data = call_api_post_method(list_param, list_url, token=token)

        if 'error' in list_data and list_data['error'] == 0:
            bid_history = list_data['data']['data']
            total = list_data['data']['total'] if 'total' in list_data['data'] else 0
        else:
            bid_history = []
            total = 0
        sno = (int(page) - 1) * int(page_size) + 1


        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename={date}-bidder-list.xlsx'.format(
            date=datetime.datetime.now().strftime('%Y-%m-%d'),
        )
        workbook = Workbook()

        # Get active worksheet/tab
        worksheet = workbook.active
        worksheet.title = 'Bidder List'

        # Define the titles for columns
        columns = [
            'ID',
            'Buyer',
            'Email',
            'Phone',
            'Reviewed',
            'Approval',
            'IP Address',
            'Registration Date',
            'Last Updated',
        ]
        row_num = 1
        header_font = Font(name='Calibri', bold=True)
        # Assign the titles for each cell of the header
        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title
            cell.font = header_font

        # Iterate through all history
        count = 0
        for bidder in bid_history:
            row_num += 1

            phone_no = bidder['phone_no']
            formatted_phone_no = format_phone_number(phone_no)
            bid_date = bidder['added_on']
            bid_updated_date = bidder['updated_on']
            bid_date_time = ''
            bid_updated_date_time = ''
            added_on_time = 0
            updated_on_time = 0
            if timezone:
                try:
                    added_on_time = time.mktime(
                        datetime.datetime.strptime(bid_date, "%Y-%m-%dT%H:%M:%SZ").timetuple())
                except ValueError:
                    added_on_time = time.mktime(
                        datetime.datetime.strptime(bid_date, "%Y-%m-%dT%H:%M:%S.%fZ").timetuple())
                except:
                    added_on_time = 0

                try:
                    updated_on_time = time.mktime(
                        datetime.datetime.strptime(bid_updated_date, "%Y-%m-%dT%H:%M:%SZ").timetuple())
                except ValueError:
                    updated_on_time = time.mktime(
                        datetime.datetime.strptime(bid_updated_date, "%Y-%m-%dT%H:%M:%S.%fZ").timetuple())
                except:
                    updated_on_time = 0

            if added_on_time:
                added_on_time = float(added_on_time) - (float(timezone)*60)
                bid_date_time = datetime.datetime.fromtimestamp(added_on_time)
                bid_date_time = datetime.datetime.strftime(bid_date_time, "%m-%d-%Y %I:%M %p")

            if updated_on_time:
                updated_on_time = float(updated_on_time) - (float(timezone)*60)
                bid_updated_date_time = datetime.datetime.fromtimestamp(updated_on_time)
                bid_updated_date_time = datetime.datetime.strftime(bid_updated_date_time, "%m-%d-%Y %I:%M %p")
            # Define the data for each cell in the row
            row = [
                sno,
                bidder['registrant'],
                bidder['email'],
                formatted_phone_no,
                bidder['is_reviewed'],
                bidder['is_approved'],
                bidder['ip_address'],
                bid_date_time,
                bid_updated_date_time,
            ]
            sno += 1

            # Assign the data for each cell of the row
            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value


        workbook.save(response)

        return response
    except Exception as exp:
        print(exp)
        return HttpResponse("Issue in views")

@csrf_exempt
def export_auction_bidders(request):
    try:
        """
            Downloads bidders as Excel file with a single worksheet from auction monitor
        """
        try:
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']

        except Exception as exp:
            print(exp)
            site_id = ""

        user_id = None
        token = None
        is_broker = 0
        if 'user_id' in request.session and request.session['user_id']:
            token = request.session['token']['access_token']
            user_id = request.session['user_id']
            is_broker = 1 if request.session['is_broker'] == True else 0

        search = request.GET.get('search', '')
        page = request.GET.get('page', 1)
        page_size = request.GET.get('page_size', 10)
        property_id = request.GET.get('property', '')
        timezone = request.GET.get('timezone', '')

        list_param = {
            'site_id': site_id,
            'user_id': user_id,
            'page_size': page_size,
            'page': page,
            'property_id': property_id,
        }

        list_url = settings.API_URL + '/api-bid/auction-bidders/'

        list_data = call_api_post_method(list_param, list_url, token=token)


        if 'error' in list_data and list_data['error'] == 0:
            bidder_list = list_data['data']['data']
            total = list_data['data']['total'] if 'total' in list_data['data'] else 0
        else:
            bidder_list = []
            total = 0
        sno = (int(page) - 1) * int(page_size) + 1


        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename={date}-bidders.xlsx'.format(
            date=datetime.datetime.now().strftime('%Y-%m-%d'),
        )
        workbook = Workbook()

        # Get active worksheet/tab
        worksheet = workbook.active
        worksheet.title = 'Bidders'

        # Define the titles for columns
        columns = [
            '#',
            'Bidder Name',
            'Company',
            'Email',
            'Phone',
            'Approved Bid Limit',
            'Approval Status',
            'Due Diligence Vault',
            # 'CA Signed',
            'Registration Approval',
        ]
        row_num = 1
        header_font = Font(name='Calibri', bold=True)
        # Assign the titles for each cell of the header
        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title
            cell.font = header_font

        # Iterate through all history
        count = 0
        for bidder in bidder_list:
            row_num += 1

            bid_limit = bidder['bid_limit']
            phone_no = bidder['bidder_detail']['phone_no']
            formatted_phone_no = format_phone_number(phone_no)
            formatted_bid_limit = format_currency(bid_limit)
            approval_date = bidder['approval_date']
            approval_date_time = ''
            approval_time = 0
            if timezone:
                try:
                    approval_time = time.mktime(
                        datetime.datetime.strptime(approval_date, "%Y-%m-%dT%H:%M:%SZ").timetuple())
                except ValueError:
                    approval_time = time.mktime(
                        datetime.datetime.strptime(approval_date, "%Y-%m-%dT%H:%M:%S.%fZ").timetuple())
                except:
                    approval_time = 0



            if approval_time:
                approval_time = float(approval_time) - (float(timezone)*60)
                approval_date_time = datetime.datetime.fromtimestamp(approval_time)
                approval_date_time = datetime.datetime.strftime(approval_date_time, "%m-%d-%Y %I:%M %p")


            # Define the data for each cell in the row
            row = [
                sno,
                bidder['bidder_detail']['first_name']+' '+bidder['bidder_detail']['last_name'],
                bidder['company_name'],
                bidder['bidder_detail']['email'],
                formatted_phone_no,
                '$'+str(f"{formatted_bid_limit:,}"),
                bidder['approval_status'],
                bidder['ca_signed'],
                approval_date_time,
            ]
            sno += 1

            # Assign the data for each cell of the row
            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value


        workbook.save(response)

        return response
    except Exception as exp:
        print(exp)
        return HttpResponse("Issue in views")

@csrf_exempt
def export_auction_bids(request):
    try:
        """
            Downloads bids as Excel file with a single worksheet from auction monitor
        """
        try:
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']

        except Exception as exp:
            print(exp)
            site_id = ""

        user_id = None
        token = None
        is_broker = 0
        if 'user_id' in request.session and request.session['user_id']:
            token = request.session['token']['access_token']
            user_id = request.session['user_id']
            is_broker = 1 if request.session['is_broker'] == True else 0

        search = request.GET.get('search', '')
        page = request.GET.get('page', 1)
        page_size = request.GET.get('page_size', 10)
        property_id = request.GET.get('property', '')
        timezone = request.GET.get('timezone', '')

        list_param = {
            'site_id': site_id,
            'user_id': user_id,
            'page_size': page_size,
            'page': page,
            'property_id': property_id,
        }

        list_url = settings.API_URL + '/api-bid/auction-total-bids/'

        list_data = call_api_post_method(list_param, list_url, token=token)


        if 'error' in list_data and list_data['error'] == 0:
            bid_list = list_data['data']['data']
            total = list_data['data']['total'] if 'total' in list_data['data'] else 0
        else:
            bid_list = []
            total = 0
        sno = (int(page) - 1) * int(page_size) + 1


        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename={date}-bidders.xlsx'.format(
            date=datetime.datetime.now().strftime('%Y-%m-%d'),
        )
        workbook = Workbook()

        # Get active worksheet/tab
        worksheet = workbook.active
        worksheet.title = 'Bids'

        # Define the titles for columns
        columns = [
            'Bidder',
            'Start Bid',
            'High Bids',
            'Bids',
            'Bidding Date',
        ]
        row_num = 1
        header_font = Font(name='Calibri', bold=True)
        # Assign the titles for each cell of the header
        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title
            cell.font = header_font

        # Iterate through all history
        count = 0
        for bid in bid_list:
            row_num += 1

            start_bid = bid['start_bid']
            max_bid = bid['max_bid']
            phone_no = bid['bidder_detail']['phone_no']
            formatted_phone_no = format_phone_number(phone_no)
            formatted_start_bid = format_currency(start_bid)
            formatted_max_bid = format_currency(max_bid)
            bid_date = bid['bid_time']
            bid_date_time = ''
            bid_time = 0
            formatted_bidder_data = ''
            try:
                bidder_name = bid['bidder_detail']['first_name']+' '+bid['bidder_detail']['last_name']
            except:
                bidder_name = ''
            try:
                bidder_address = bid['bidder_detail']['address_first']+', '+bid['bidder_detail']['city']+', '+bid['bidder_detail']['state']+', '+bid['bidder_detail']['postal_code']
            except:
                bidder_address = ''
            try:
                bidder_ip = bid['bidder_detail']['ip_address']
            except:
                bidder_ip = ''

            try:
                bidder_email = bid['bidder_detail']['email']
            except:
                bidder_email = ''

            if bidder_name or formatted_phone_no or bidder_address or bidder_ip or bidder_email:
                formatted_bidder_data = bidder_name+', '+bidder_address+', '+formatted_phone_no+', '+bidder_email+', '+bidder_ip


            if timezone:
                try:
                    bid_time = time.mktime(
                        datetime.datetime.strptime(bid_date, "%Y-%m-%dT%H:%M:%SZ").timetuple())
                except ValueError:
                    bid_time = time.mktime(
                        datetime.datetime.strptime(bid_date, "%Y-%m-%dT%H:%M:%S.%fZ").timetuple())
                except:
                    bid_time = 0



            if bid_time:
                bid_time = float(bid_time) - (float(timezone)*60)
                bid_date_time = datetime.datetime.fromtimestamp(bid_time)
                bid_date_time = datetime.datetime.strftime(bid_date_time, "%m-%d-%Y %I:%M %p")


            # Define the data for each cell in the row
            row = [
                formatted_bidder_data,
                '$'+str(f"{formatted_start_bid:,}"),
                '$'+str(f"{formatted_max_bid:,}"),
                bid['bids'],
                bid_date_time,
                # 'Accepted',
            ]
            sno += 1

            # Assign the data for each cell of the row
            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value


        workbook.save(response)

        return response
    except Exception as exp:
        print(exp)
        return HttpResponse("Issue in views")

@csrf_exempt
def export_watchers(request):
    try:
        """
            Downloads watchers as Excel file with a single worksheet from auction monitor
        """
        try:
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']

        except Exception as exp:
            print(exp)
            site_id = ""

        user_id = None
        token = None
        is_broker = 0
        if 'user_id' in request.session and request.session['user_id']:
            token = request.session['token']['access_token']
            user_id = request.session['user_id']
            is_broker = 1 if request.session['is_broker'] == True else 0

        search = request.GET.get('search', '')
        page = request.GET.get('page', 1)
        page_size = request.GET.get('page_size', 10)
        property_id = request.GET.get('property', '')
        timezone = request.GET.get('timezone', '')

        list_param = {
            'site_id': site_id,
            'user_id': user_id,
            'page_size': page_size,
            'page': page,
            'property_id': property_id,
        }

        list_url = settings.API_URL + '/api-bid/auction-total-watching/'

        list_data = call_api_post_method(list_param, list_url, token=token)


        if 'error' in list_data and list_data['error'] == 0:
            watcher_list = list_data['data']['data']
            total = list_data['data']['total'] if 'total' in list_data['data'] else 0
        else:
            watcher_list = []
            total = 0
        sno = (int(page) - 1) * int(page_size) + 1


        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename={date}-watchers.xlsx'.format(
            date=datetime.datetime.now().strftime('%Y-%m-%d'),
        )
        workbook = Workbook()

        # Get active worksheet/tab
        worksheet = workbook.active
        worksheet.title = 'Bids'

        # Define the titles for columns
        columns = [
            'Watchers',
        ]
        row_num = 1
        header_font = Font(name='Calibri', bold=True)
        # Assign the titles for each cell of the header
        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title
            cell.font = header_font

        # Iterate through all history
        count = 0
        for watcher in watcher_list:
            row_num += 1

            phone_no = watcher['watcher_detail']['phone_no']
            formatted_phone_no = format_phone_number(phone_no)

            formatted_watcher_data = ''
            try:
                watcher_name = watcher['watcher_detail']['first_name']+' '+watcher['watcher_detail']['last_name']
            except:
                watcher_name = ''
            try:
                watcher_address = watcher['watcher_detail']['address_first']+', '+watcher['watcher_detail']['city']+', '+watcher['watcher_detail']['state']+', '+watcher['watcher_detail']['postal_code']
            except:
                watcher_address = ''
            try:
                watcher_ip = watcher['watcher_detail']['ip_address']
            except:
                watcher_ip = ''

            try:
                watcher_email = watcher['watcher_detail']['email']
            except:
                watcher_email = ''

            if watcher_name or formatted_phone_no or watcher_address or watcher_ip or watcher_email:
                formatted_watcher_data = watcher_name+', '+watcher_address+', '+formatted_phone_no+', '+watcher_email
                if watcher_ip:
                    formatted_watcher_data = formatted_watcher_data+', '+watcher_ip





            # Define the data for each cell in the row
            row = [
                formatted_watcher_data,
            ]
            sno += 1

            # Assign the data for each cell of the row
            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value


        workbook.save(response)

        return response
    except Exception as exp:
        print(exp)
        return HttpResponse("Issue in views")


@csrf_exempt
def insider_property_bid_history(request):
    try:
        user_id = user_id = None
        page_size = 10
        try:
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']

        except Exception as exp:
            print(exp)
            site_id = ""

        if 'user_id' in request.session and request.session['user_id']:
            token = request.session['token']['access_token']
            user_id = request.session['user_id']

        if request.is_ajax() and request.method == 'POST':
            page = 1
            if 'page' in request.POST and request.POST['page'] != "":
                page = request.POST['page']

            if 'page_size' in request.POST and request.POST['page_size'] != "":
                page_size = request.POST['page_size']

            search = ''
            if 'search' in request.POST and request.POST['search'] != "":
                search = request.POST['search']

            property_id = request.POST['property_id']
            step = request.POST['step'] if 'step' in  request.POST else "dutch"
            #register_user = request.POST['register_user'] if 'register_user' in request.POST and request.POST['register_user'] else ''
            params = {
                'domain_id': site_id,
                'user_id': user_id,
                'page_size': page_size,
                'page': page,
                'property_id': property_id,
                'search': search,
                'step': step,
            }

            api_url = settings.API_URL + '/api-bid/insider-bidder/'
            bid_history = call_api_post_method(params, api_url, token=token)
            try:
                prop_detail = bid_history['data']['property_address']
                image = bid_history['data']['property_image']
                if image and image['image'] and image['image'] != "":
                    image_url = settings.AWS_URL + image['bucket_name'] + '/' + image['image']
                else:
                    image_url = ''
                property_address = prop_detail['address_one']
                property_city = prop_detail['city']
                property_state = prop_detail['state']
                property_postal_code = prop_detail['postal_code']
                property_image = image_url
                #auction_type = prop_detail['auction_type']
                #bid_increment = prop_detail['bid_increment']
            except Exception as exp:
                print(exp)
                property_address = property_city = property_state = property_postal_code = property_image = auction_type = bid_increment = ''

            if 'error' in bid_history and bid_history['error'] == 0:
                total = bid_history['data']['total']
                bid_history = bid_history['data']['data']
            else:
                bid_history = []
                total = 0

            context = {'bid_history': bid_history, 'total': total, "aws_url": settings.AWS_URL,
                       'start_index': (int(page) - 1) * int(page_size), "step": step}

            bidder_listing_path = 'admin/dashboard/listings/insider-property-bid-history.html'
            bidder_listing_template = get_template(bidder_listing_path)
            bid_history_html = bidder_listing_template.render(context)
            # ---------------Pagination--------
            pagination_path = 'admin/dashboard/listings/insider-bid-history-pagination.html'
            pagination_template = get_template(pagination_path)
            total_page = math.ceil(int(total) / int(page_size))
            pagination_html = ''
            if total_page > 1:
                pagination_data = {"no_page": int(total_page), "total_page": range(total_page),
                                   "current_page": int(page),
                                   "pagination_id": "bidHistoryPaginationList", "property_id": property_id, "step": step}
                pagination_html = pagination_template.render(pagination_data)

            data = {
                'bid_history_html': bid_history_html,
                'status': 200,
                'msg': '',
                'error': 0,
                'total': total,
                "pagination_html": pagination_html,
                'pagination_id': 'bidHistoryPaginationList',
                'property_address': property_address,
                'property_city': property_city,
                'property_state': property_state,
                'property_postal_code': property_postal_code,
                'property_image': property_image,
                'property_id': property_id,
                'page': page,
                'page_size': page_size,
                'step': step,
            }
        else:
            data = {'status': 403, 'msg': 'Forbidden', 'error': 1}
        return JsonResponse(data)
    except Exception as exp:
        print(exp)
        return HttpResponse("Issue in views")

@csrf_exempt
def export_insider_bid_history(request):
    try:
        """
            Downloads insider bid history as Excel file with a single worksheet
        """
        try:
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']

        except Exception as exp:
            print(exp)
            site_id = ""

        user_id = None
        token = None
        is_broker = 0
        if 'user_id' in request.session and request.session['user_id']:
            token = request.session['token']['access_token']
            user_id = request.session['user_id']
            is_broker = 1 if request.session['is_broker'] == True else 0

        search = request.GET.get('search', '')
        page = request.GET.get('page', 1)
        page_size = request.GET.get('page_size', 10)
        property_id = request.GET.get('property', '')
        timezone = request.GET.get('timezone', '')
        register_user = request.GET.get('register_user', '')
        step = request.GET.get('step', 'dutch')


        list_param = {
            'domain_id': site_id,
            'user_id': user_id,
            'page_size': page_size,
            'page': page,
            'property_id': property_id,
            'search': search,
            'step': step,
        }

        list_url = settings.API_URL + '/api-bid/insider-bidder/'

        list_data = call_api_post_method(list_param, list_url, token=token)
        if 'error' in list_data and list_data['error'] == 0:
            bid_history = list_data['data']['data']
            total = list_data['data']['total'] if 'total' in list_data['data'] else 0
        else:
            bid_history = []
            total = 0
        sno = (int(page) - 1) * int(page_size) + 1


        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename={date}-bid-history.xlsx'.format(
            date=datetime.datetime.now().strftime('%Y-%m-%d'),
        )
        workbook = Workbook()

        # Get active worksheet/tab
        worksheet = workbook.active
        worksheet.title = 'Bid History'

        # Define the titles for columns
        columns = [
            '#',
            'Bidder Name',
            'User Type',
            'Email',
            'Phone',
            'Starting Bid',
            'Bid Amount',
            'IP Address',
            'Bidding Start Time',
            'Bidding End Time',
        ]
        row_num = 1
        header_font = Font(name='Calibri', bold=True)
        # Assign the titles for each cell of the header
        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title
            cell.font = header_font

        # Iterate through all history
        count = 0
        for bids in bid_history:
            row_num += 1

            phone_no = bids['phone_no']
            start_price = bids['start_price']
            bid_amount = bids['bid_amount']
            formatted_phone_no = format_phone_number(phone_no)
            formatted_start_price = format_currency(start_price)
            formatted_amount = format_currency(bid_amount)
            bid_start_date = bids['bidding_start_time']
            bid_end_date = bids['bidding_end_time']
            bid_start_date_time = ''
            bid_end_date_time = ''
            if timezone:
                try:
                    added_on_time = time.mktime(
                        datetime.datetime.strptime(bid_start_date, "%Y-%m-%dT%H:%M:%SZ").timetuple())
                except ValueError:
                    added_on_time = time.mktime(
                        datetime.datetime.strptime(bid_start_date, "%Y-%m-%dT%H:%M:%S.%fZ").timetuple())
                except:
                    added_on_time = 0

                try:
                    end_on_time = time.mktime(
                        datetime.datetime.strptime(bid_end_date, "%Y-%m-%dT%H:%M:%SZ").timetuple())
                except ValueError:
                    end_on_time = time.mktime(
                        datetime.datetime.strptime(bid_end_date, "%Y-%m-%dT%H:%M:%S.%fZ").timetuple())
                except:
                    end_on_time = 0

            if added_on_time:
                added_on_time = float(added_on_time) - (float(timezone)*60)
                bid_start_date_time = datetime.datetime.fromtimestamp(added_on_time)
                bid_start_date_time = datetime.datetime.strftime(bid_start_date_time, "%m-%d-%Y %I:%M %p")

            if end_on_time:
                end_on_time = float(end_on_time) - (float(timezone)*60)
                bid_end_date_time = datetime.datetime.fromtimestamp(end_on_time)
                bid_end_date_time = datetime.datetime.strftime(bid_end_date_time, "%m-%d-%Y %I:%M %p")
            # Define the data for each cell in the row
            row = [
                sno,
                bids['first_name']+' '+bids['last_name'],
                bids['user_type'],
                bids['email'],
                formatted_phone_no,
                '$'+str(f"{formatted_start_price:,}"),
                '$'+str(f"{formatted_amount:,}"),
                bids['ip_address'],
                bid_start_date_time,
                bid_end_date_time,
            ]
            sno += 1

            # Assign the data for each cell of the row
            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value


        workbook.save(response)

        return response
    except Exception as exp:
        print(exp)
        return HttpResponse("Issue in views")

@csrf_exempt
def subdomain_property_estimator_list(request):
    try:
        # is_permission = check_permission(request, 1)
        # if not is_permission:
        #     http_host = request.META['HTTP_HOST']
        #     redirect_url = settings.URL_SCHEME + str(http_host)
        #     return HttpResponseRedirect(redirect_url)

        try:
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']

        except Exception as exp:
            print(exp)
            site_id = ""

        user_id = None
        token = None
        is_broker = 0
        if 'user_id' in request.session and request.session['user_id']:
            token = request.session['token']['access_token']
            user_id = request.session['user_id']
            is_broker = 1 if request.session['is_broker'] == True else 0

        page_size = 10
        page = 1
        if request.is_ajax() and request.method == 'POST':

            estimator_search = ''
            if 'search' in request.POST and request.POST['search']:
                estimator_search = request.POST['search']

            page = 1
            if 'page' in request.POST and request.POST['page'] != "":
                page = request.POST['page']

            page_size = 10
            if 'perpage' in request.POST and request.POST['perpage']:
                page_size = request.POST['perpage']

            if 'status' in request.POST and request.POST['status'] and request.POST['status'] and request.POST[
                'status'].lower() == 'active':
                status = [1]
            elif 'status' in request.POST and request.POST['status'] and request.POST['status'] and request.POST[
                'status'].lower() == 'inactive':
                status = [2]
            else:
                status = [2, 1]

            list_param = {
                'domain_id': site_id,
                "page": page,
                "user_id": user_id,
                "page_size": page_size,
                "status": "",
                "search": estimator_search
            }
            sno = (int(page) - 1) * int(page_size) + 1
            list_url = settings.API_URL + '/api-property/subdomain-property-evaluator/'
            list_data = call_api_post_method(list_param, list_url, token)

            if 'error' in list_data and list_data['error'] == 0:
                estimator_list = list_data['data']['data']
                total = list_data['data']['total']
            else:
                estimator_list = []
                total = 0
            try:
                agent_list = list_data['data']['agent_list']
            except:
                agent_list = []

            try:
                status_param = {'object_id': 22}
                status_url = settings.API_URL + '/api-settings/lookup-status-listing/'
                status_data = call_api_post_method(status_param, status_url, token)
                status_list = status_data['data']
            except:
                status_list = []
            context = {'estimator_list': estimator_list, 'agent_list': agent_list, 'total': total, "aws_url": settings.AWS_URL, 'sno': sno, 'is_broker': is_broker, 'status_list': status_list}

            estimator_listing_path = 'admin/dashboard/estimator/estimator-content.html'
            estimator_listing_template = get_template(estimator_listing_path)
            estimator_listing_html = estimator_listing_template.render(context)
            # ---------------Pagination--------
            pagination_path = 'admin/dashboard/estimator/pagination.html'
            pagination_template = get_template(pagination_path)
            total_page = math.ceil(int(total) / int(page_size))
            pagination_html = ''
            if total_page > 1:
                pagination_data = {"no_page": int(total_page), "total_page": range(total_page), "current_page": int(page),
                                   "pagination_id": "estimator_listing_pagination_list"}
                pagination_html = pagination_template.render(pagination_data)

            data = {'estimator_listing_html': estimator_listing_html, 'status': 200, 'msg': '', 'error': 0, 'total': total,
                    "pagination_html": pagination_html, 'pagination_id': 'estimator_listing_pagination_list'}
            return JsonResponse(data)
        else:
            list_param = {
                'domain_id': site_id,
                "page": page,
                "user_id": user_id,
                "page_size": page_size,
                "status": "",
                "search": ""
            }
            sno = (int(page) - 1) * int(page_size) + 1
            list_url = settings.API_URL + '/api-property/subdomain-property-evaluator/'
            list_data = call_api_post_method(list_param, list_url, token)
            if 'error' in list_data and list_data['error'] == 0:
                estimator_list = list_data['data']['data']
                total = list_data['data']['total']
            else:
                estimator_list = []
                total = 0
            try:
                agent_list = list_data['data']['agent_list']
            except:
                agent_list = []

            try:
                status_param = {'object_id': 22}
                status_url = settings.API_URL + '/api-settings/lookup-status-listing/'
                status_data = call_api_post_method(status_param, status_url, token)
                status_list = status_data['data']
            except:
                status_list = []
            # ---------------Pagination--------
            pagination_html = ''
            pagination_path = 'admin/dashboard/estimator/pagination.html'
            pagination_template = get_template(pagination_path)
            total_page = math.ceil(total / page_size)
            if total_page > 1:
                pagination_data = {"no_page": total_page, "total_page": range(total_page), "current_page": 1,
                                   "pagination_id": "estimator_listing_pagination_list"}
                pagination_html = pagination_template.render(pagination_data)
            context = {"estimator_list": estimator_list, "agent_list": agent_list, "status_list": status_list, "total": total, "pagination_html": pagination_html,
                       "pagination_id": "estimator_listing_pagination_list", "active_menu": "estimator_list", "active_submenu": "estimator_list", "sno": sno, "is_broker": is_broker}

            
            return render(request, "admin/dashboard/estimator/view-estimators.html", context)
    except Exception as exp:
        print(exp)
        return HttpResponse("Issue in views")


@csrf_exempt
def insider_auction_dashboard(request):
    try:
        site_detail = subdomain_site_details(request)
        site_id = site_detail['site_detail']['site_id']
        token = request.session['token']['access_token']
        user_id = request.session['user_id']
        page_size = 10

        if request.is_ajax() and request.method == 'POST':

            search = ''
            if 'search' in request.POST and request.POST['search']:
                search = request.POST['search']

            page = 1
            if 'page' in request.POST and request.POST['page'] != "":
                page = int(request.POST['page'])

            page_size = 10
            if 'perpage' in request.POST and request.POST['perpage']:
                page_size = int(request.POST['perpage'])

            asset_type = ''
            if 'asset_type' in request.POST and request.POST['asset_type']:
                asset_type = int(request.POST['asset_type'])
            auction_type = ''
            if 'auction_type' in request.POST and request.POST['auction_type']:
                auction_type = int(request.POST['auction_type'])
            property_type = ''
            if 'property_type' in request.POST and request.POST['property_type']:
                property_type = int(request.POST['property_type'])

            status = ""
            if 'status' in request.POST and request.POST['status']:
                status = request.POST['status']
            
            agent = ''
            if 'agent' in request.POST and request.POST['agent']:
                agent = int(request.POST['agent'])

            list_param = {
                "page": page,
                "page_size": page_size,
                "domain_id": site_id,
                "user_id": user_id,
                "auction_id": auction_type,
                "asset_id": asset_type,
                "property_type": property_type,
                "search": search,
                "agent_id":agent,
                "status": status
            }
            api_url = settings.API_URL + '/api-bid/subdomain-inline-bidding-monitor/'

            auction_data = call_api_post_method(list_param, api_url, token)

            if 'error' in auction_data and auction_data['error'] == 0:
                auction_listings = auction_data['data']['data']
                total = auction_data['data']['total'] if 'total' in auction_data['data'] else 0
            else:
                auction_listings = []
                total = 0
            sno = (int(page) - 1) * int(page_size) + 1

            context = {
                "auction_data": auction_listings,
                'total': total,
                "aws_url": settings.AWS_URL,
                "is_broker": 1 if request.session['is_broker'] == True else 0,
                "sno": sno
            }
            prop_auction_path = 'admin/dashboard/auction/insider-auction-listing-content.html'
            property_listing_template = get_template(prop_auction_path)
            property_auction_listings = property_listing_template.render(context)
            # ---------------Pagination--------
            pagination_html = ''
            pagination_path = pagination_path = 'admin/dashboard/auction/auction-pagination.html'
            pagination_template = get_template(pagination_path)
            total_page = math.ceil(total / page_size)
            
            if total_page > 1:
                pagination_data = {"no_page": int(total_page), "total_page": range(total_page), "current_page": int(page),
                    "pagination_id": "prop_auction_pagination_list"}
                pagination_html = pagination_template.render(pagination_data)
            


            return JsonResponse({
                'property_auction_listings': property_auction_listings,
                'status': 200,
                'msg': '',
                'error': 0,
                'total': total,
                "pagination_html": pagination_html,
                'pagination_id': 'prop_listing_pagination_list',
                "sno": sno
            })
        else:
            page = 1
            try:
                params = {"domain_id": site_id, "user_id": user_id}
                api_url = settings.API_URL + '/api-users/agent-listing/'
                auction_type_data = call_api_post_method(params, api_url, token)
                agent_list = auction_type_data['data']
            except:
                agent_list = []
            try:
                auction_type_param = {}
                auction_type_url = settings.API_URL + '/api-settings/subdomain-auction-type/'
                auction_type_data = call_api_post_method(auction_type_param, auction_type_url, token)
                auction_type_list = auction_type_data['data']
            except:
                auction_type_list = []

            try:
                asset_listing_params = {}

                asset_listing_url = settings.API_URL + '/api-property/asset-listing/'

                asset_listing_data = call_api_post_method(asset_listing_params, asset_listing_url, token)
                asset_listing = asset_listing_data['data']
            except:
                asset_listing = []

            try:
                property_type_params = {}

                property_type_url = settings.API_URL + '/api-property/property-type-listing/'

                property_type_data = call_api_post_method(property_type_params, property_type_url, token)
                property_type_listing = property_type_data['data']
            except:
                property_type_listing = []

            # try:
            #     status_param = {'object_id': 9}
            #     status_url = settings.API_URL + '/api-settings/lookup-status-listing/'
            #     status_data = call_api_post_method(status_param, status_url, token)
            #     status_list = status_data['data']
            # except:
            #     status_list = []
            
            params = {"domain_id": site_id, "user_id": user_id, "page": page, "page_size": page_size, "status": 1 }

            api_url = settings.API_URL + '/api-bid/subdomain-inline-bidding-monitor/'

            auction_data = call_api_post_method(params, api_url, token)

            if 'error' in auction_data and auction_data['error'] == 0:
                auction_listings = auction_data['data']['data']
                total = auction_data['data']['total'] if 'total' in auction_data['data'] else 0
            else:
                auction_listings = []
                total = 0
            # ---------------Pagination--------
            pagination_html = ''
            pagination_path = pagination_path = 'admin/dashboard/auction/auction-pagination.html'
            pagination_template = get_template(pagination_path)
            total_page = math.ceil(total / page_size)
            sno = (int(page) - 1) * int(page_size) + 1
            if total_page > 1:
                pagination_data = {"no_page": total_page, "total_page": range(total_page), "current_page": 1,
                    "pagination_id": "prop_auction_pagination_list"}
                pagination_html = pagination_template.render(pagination_data)
            
            x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
            if x_forwarded_for:
                ip_address = x_forwarded_for.split(',')[0]
            else:
                ip_address = request.META.get('REMOTE_ADDR')

            context = {
                "active_menu": "listing_monitor",
                "active_submenu": "insider-auction-dashboard",
                "auction_data": auction_listings,
                "auction_type_list": auction_type_list,
                "property_type_listing": property_type_listing,
                # "status_list": status_list,
                "asset_listing": asset_listing,
                "agent_list":agent_list,
                "pagination_html": pagination_html,
                "is_broker": 1 if request.session['is_broker'] == True else 0,
                "sno": sno,
                "ip_address": ip_address,
                "node_url": settings.NODE_URL,
                "user_id": user_id,
                "domain_id": site_id
            }

            return render(request, "admin/dashboard/auction/insider-index.html", context)
    except Exception as exp:
        print(exp)
        return HttpResponse("Issue in views")
@csrf_exempt
def estimator_assign_agent(request):
    try:
        if request.is_ajax() and request.method == 'POST':
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']
            token = request.session['token']['access_token']
            user_id = request.session['user_id']
            assign_to_id = request.POST['assign_to_id']
            assign_to_name = request.POST['assign_to_name']
            estimator_id = request.POST['estimator_id']
            change_status_params = {
                "domain_id": site_id,
                "user_id": user_id,
                "bot_id": estimator_id,
                "assign_to": assign_to_id,
            }

            change_agent_url = settings.API_URL + '/api-property/assign-property-evaluator/'

            change_agent_data = call_api_post_method(change_status_params, change_agent_url, token)

            if 'error' in change_agent_data and change_agent_data['error'] == 0:
                data = {'status': 200, 'assign_to_id': assign_to_id, 'assign_to_name': assign_to_name, 'estimator_id': estimator_id, 'error': 0, 'msg': 'Agent assigned successfully.'}
            else:
                data = {'status': 403, 'assign_to_id': assign_to_id, 'assign_to_name': assign_to_name, 'estimator_id': estimator_id, 'error': 1, 'msg': change_agent_data['msg']}
        else:
            data = {'status': 403, 'status_id': '', 'status_name': '', 'error': 1, 'msg': 'Some error occurs, please try again'}

        return JsonResponse(data)
    except Exception as exp:
        print(exp)
        data = {'status': 403, 'status_id': '', 'status_name': '', 'error': 1, 'msg': 'Some error occurs, please try again'}
        return JsonResponse(data)

@csrf_exempt
def domain_property_estimator_details(request):
    try:
        bot_id = request.GET.get('id', None)
        address_question_list = []
        try:
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']

        except Exception as exp:
            print(exp)
            site_id = ""


        user_id = None
        token = None
        if 'user_id' in request.session and request.session['user_id']:
            token = request.session['token']['access_token']
            user_id = request.session['user_id']

        user_first_name = None
        user_profile_image = None

        try:
            params = {
                "domain_id": site_id,
                "bot_id": bot_id,
                "category_id": 1,
                "user_id": user_id
            }
            api_url = settings.API_URL + '/api-property/property-evaluator-detail/'
            estimator_data = call_api_post_method(params, api_url, token=token)
            address_question_list = estimator_data['data']

            if address_question_list:
                for ques in address_question_list:
                    if user_first_name is None:
                        user_first_name = ques['user_detail']['first_name']

                    if user_profile_image is None:
                        user_profile_image = ques['user_detail']['profile_image']

                    if ques['option_type'] == 6:
                        if ques['answer'] and ques['answer']['answer']:
                            address = ques['answer']['answer']
                            geolocator = Nominatim(user_agent=settings.GEOLOCATOR_EMAIL, timeout=10)
                            location = geolocator.geocode(address)
                            try:
                                ques['latitude'] = location.latitude
                                ques['longitude'] = location.longitude
                            except:
                                ques['latitude'] = 33.1836304229855
                                ques['longitude'] = -117.32728034223504
                        else:
                            ques['latitude'] = 33.1836304229855
                            ques['longitude'] = -117.32728034223504

                    if ques['option_type'] == 6:
                        formatted_addr = ''
                        addr_i = 0
                        if ques['answer'] and ques['answer']['answer']:
                            address_arr = ques['answer']['answer'].split(',')
                            if address_arr:
                                for addr in address_arr:
                                    if addr_i == 0:
                                        formatted_addr = addr
                                    elif addr_i == 1:
                                        if len(address_arr) == 2:
                                            formatted_addr = formatted_addr + ', ' + addr
                                        else:
                                            formatted_addr = formatted_addr + ', <br>' + addr
                                    elif addr_i == 2:
                                        if len(address_arr) == 3:
                                            formatted_addr = formatted_addr + ', ' + addr
                                        else:
                                            formatted_addr = formatted_addr + ', ' + addr + ', <br>'
                                    else:
                                        formatted_addr = formatted_addr + ' ' + addr

                                    addr_i = addr_i + 1
                            ques['formatted_addr'] = formatted_addr
        except:
            address_question_list = []

        try:
            params = {
                "domain_id": site_id,
                "bot_id": bot_id,
                "category_id": 2,
                "user_id": user_id
            }
            api_url = settings.API_URL + '/api-property/property-evaluator-detail/'
            estimator_data = call_api_post_method(params, api_url, token=token)
            details_question_list = estimator_data['data']
            if details_question_list:
                for ques in details_question_list:
                    if user_first_name is None:
                        user_first_name = ques['user_detail']['first_name']

                    if user_profile_image is None:
                        user_profile_image = ques['user_detail']['profile_image']

                    if ques['option_type'] == 6:
                        if ques['answer'] and ques['answer']['answer']:
                            address = ques['answer']['answer']
                            geolocator = Nominatim(user_agent=settings.GEOLOCATOR_EMAIL, timeout=10)
                            location = geolocator.geocode(address)
                            try:
                                ques['latitude'] = location.latitude
                                ques['longitude'] = location.longitude
                            except:
                                ques['latitude'] = 33.1836304229855
                                ques['longitude'] = -117.32728034223504
                        else:
                            ques['latitude'] = 33.1836304229855
                            ques['longitude'] = -117.32728034223504

                    if ques['option_type'] == 6:
                        formatted_addr = ''
                        addr_i = 0
                        if ques['answer'] and ques['answer']['answer']:
                            address_arr = ques['answer']['answer'].split(',')
                            if address_arr:
                                for addr in address_arr:
                                    if addr_i == 0:
                                        formatted_addr = addr
                                    elif addr_i == 1:
                                        if len(address_arr) == 2:
                                            formatted_addr = formatted_addr + ', ' + addr
                                        else:
                                            formatted_addr = formatted_addr + ', <br>' + addr
                                    elif addr_i == 2:
                                        if len(address_arr) == 3:
                                            formatted_addr = formatted_addr + ', ' + addr
                                        else:
                                            formatted_addr = formatted_addr + ', ' + addr + ', <br>'
                                    else:
                                        formatted_addr = formatted_addr + ' ' + addr

                                    addr_i = addr_i + 1
                            ques['formatted_addr'] = formatted_addr
        except:
            details_question_list = []

        try:
            params = {
                "domain_id": site_id,
                "bot_id": bot_id,
                "category_id": 3,
                "user_id": user_id
            }
            api_url = settings.API_URL + '/api-property/property-evaluator-detail/'
            estimator_data = call_api_post_method(params, api_url, token=token)
            doc_question_list = estimator_data['data']
            if doc_question_list:
                if user_first_name is None:
                    user_first_name = ques['user_detail']['first_name']

                if user_profile_image is None:
                    user_profile_image = ques['user_detail']['profile_image']

                for ques in doc_question_list:
                    if ques['option_type'] == 6:
                        if ques['answer'] and ques['answer']['answer']:
                            address = ques['answer']['answer']
                            geolocator = Nominatim(user_agent=settings.GEOLOCATOR_EMAIL, timeout=10)
                            location = geolocator.geocode(address)
                            try:
                                ques['latitude'] = location.latitude
                                ques['longitude'] = location.longitude
                            except:
                                ques['latitude'] = 33.1836304229855
                                ques['longitude'] = -117.32728034223504
                        else:
                            ques['latitude'] = 33.1836304229855
                            ques['longitude'] = -117.32728034223504
                    if ques['option_type'] == 6:
                        formatted_addr = ''
                        addr_i = 0
                        if ques['answer'] and ques['answer']['answer']:
                            address_arr = ques['answer']['answer'].split(',')
                            if address_arr:
                                for addr in address_arr:
                                    if addr_i == 0:
                                        formatted_addr = addr
                                    elif addr_i == 1:
                                        if len(address_arr) == 2:
                                            formatted_addr = formatted_addr + ', ' + addr
                                        else:
                                            formatted_addr = formatted_addr + ', <br>' + addr
                                    elif addr_i == 2:
                                        if len(address_arr) == 3:
                                            formatted_addr = formatted_addr + ', ' + addr
                                        else:
                                            formatted_addr = formatted_addr + ', ' + addr + ', <br>'
                                    else:
                                        formatted_addr = formatted_addr + ' ' + addr

                                    addr_i = addr_i + 1
                            ques['formatted_addr'] = formatted_addr
        except:
            doc_question_list = []

        try:
            params = {
                "domain_id": site_id,
                "bot_id": bot_id,
                "category_id": 4,
                "user_id": user_id
            }
            api_url = settings.API_URL + '/api-property/property-evaluator-detail/'
            estimator_data = call_api_post_method(params, api_url, token=token)
            additional_question_list = estimator_data['data']
            if additional_question_list:
                for ques in additional_question_list:
                    if user_first_name is None:
                        user_first_name = ques['user_detail']['first_name']

                    if user_profile_image is None:
                        user_profile_image = ques['user_detail']['profile_image']

                    if ques['option_type'] == 6:
                        formatted_addr = ''
                        addr_i = 0
                        if ques['answer'] and ques['answer']['answer']:
                            address_arr = ques['answer']['answer'].split(',')
                            if address_arr:
                                for addr in address_arr:
                                    if addr_i == 0:
                                        formatted_addr = addr
                                    elif addr_i == 1:
                                        if len(address_arr) == 2:
                                            formatted_addr = formatted_addr + ', ' + addr
                                        else:
                                            formatted_addr = formatted_addr + ', <br>' + addr
                                    elif addr_i == 2:
                                        if len(address_arr) == 3:
                                            formatted_addr = formatted_addr + ', ' + addr
                                        else:
                                            formatted_addr = formatted_addr + ', ' + addr + ', <br>'
                                    else:
                                        formatted_addr = formatted_addr + ' ' + addr

                                    addr_i = addr_i + 1
                            ques['formatted_addr'] = formatted_addr
        except:
            additional_question_list = []

        context = {
            "address_question_list": address_question_list,
            "details_question_list": details_question_list,
            "doc_question_list": doc_question_list,
            "additional_question_list": additional_question_list,
            "active_menu": "estimator_list",
            "active_submenu": "estimator_list",
            "aws_url": settings.AWS_URL,
            "user_first_name": user_first_name,
            "user_profile_image": user_profile_image,
        }
        return render(request, "admin/dashboard/estimator/estimator-details.html", context)
    except Exception as exp:
        print(exp)
        return HttpResponse("Issue in views")

@csrf_exempt
def estimator_change_status(request):
    try:
        if request.is_ajax() and request.method == 'POST':
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']
            token = request.session['token']['access_token']
            user_id = request.session['user_id']
            status_id = request.POST['status_id']
            status_name = request.POST['status_name']
            estimator_id = request.POST['estimator_id']
            change_status_params = {
                "domain_id": site_id,
                "user_id": user_id,
                "bot_id": estimator_id,
                "status": status_id,
            }

            change_status_url = settings.API_URL + '/api-property/property-evaluator-status-change/'

            change_status_data = call_api_post_method(change_status_params, change_status_url, token)

            if 'error' in change_status_data and change_status_data['error'] == 0:
                data = {'status': 200, 'status_id': status_id, 'status_name': status_name, 'estimator_id': estimator_id, 'error': 0, 'msg': 'Status updated successfully.'}
            else:
                data = {'status': 403, 'status_id': status_id, 'status_name': status_name, 'estimator_id': estimator_id, 'error': 1, 'msg': change_status_data['msg']}
        else:
            data = {'status': 403, 'status_id': '', 'status_name': '', 'error': 1, 'msg': 'Some error occurs, please try again'}

        return JsonResponse(data)
    except Exception as exp:
        print(exp)
        data = {'status': 403, 'status_id': '', 'status_name': '', 'error': 1, 'msg': 'Some error occurs, please try again'}
        return JsonResponse(data)

@csrf_exempt
def estimator_send_message(request):
    try:
        if request.is_ajax() and request.method == 'POST':
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']
            token = request.session['token']['access_token']
            user_id = request.session['user_id']
            estimator_id = request.POST['estimator_id']
            estimator_msg = request.POST['message']
            params = {
                "domain_id": site_id,
                "user_id": user_id,
                "bot_id": estimator_id,
                "msg": estimator_msg
            }

            api_url = settings.API_URL + '/api-property/property-evaluator-save-msg/'

            change_msg_data = call_api_post_method(params, api_url, token)

            if 'error' in change_msg_data and change_msg_data['error'] == 0:
                data = {'status': 200, 'estimator_id': estimator_id, 'error': 0, 'msg': 'Message send successfully.'}
            else:
                data = {'status': 403, 'estimator_id': estimator_id, 'error': 1, 'msg': change_msg_data['msg']}
        else:
            data = {'status': 403, 'error': 1, 'msg': 'Some error occurs, please try again'}

        return JsonResponse(data)
    except Exception as exp:
        print(exp)
        data = {'status': 403, 'error': 1, 'msg': 'Some error occurs, please try again'}
        return JsonResponse(data)


@csrf_exempt
def portfolio_list(request):
    try:
        # is_permission = check_permission(request, 6)
        # if not is_permission:
        #     http_host = request.META['HTTP_HOST']
        #     redirect_url = settings.URL_SCHEME + str(http_host)
        #     return HttpResponseRedirect(redirect_url)

        try:
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']

        except Exception as exp:
            print(exp)
            site_id = ""

        user_id = None
        token = None
        is_broker = 0
        if 'user_id' in request.session and request.session['user_id']:
            token = request.session['token']['access_token']
            user_id = request.session['user_id']
            is_broker = 1 if request.session['is_broker'] == True else 0

        page_size = 10
        page = 1
        if request.is_ajax() and request.method == 'POST':

            portfolio_search = ''
            if 'search' in request.POST and request.POST['search']:
                portfolio_search = request.POST['search']

            page = 1
            if 'page' in request.POST and request.POST['page'] != "":
                page = request.POST['page']

            page_size = 10
            if 'perpage' in request.POST and request.POST['perpage']:
                page_size = request.POST['perpage']

            list_param = {
                'domain_id': site_id,
                "page": page,
                "user_id": user_id,
                "page_size": page_size,
                "search": portfolio_search
            }
            sno = (int(page) - 1) * int(page_size) + 1
            list_url = settings.API_URL + '/api-property/portfolio-listing/'
            list_data = call_api_post_method(list_param, list_url, token)

            if 'error' in list_data and list_data['error'] == 0:
                portfolio_list = list_data['data']['data']
                total = list_data['data']['total']
            else:
                portfolio_list = []
                total = 0

            context = {'portfolio_list': portfolio_list, 'total': total,
                       "aws_url": settings.AWS_URL, 'sno': sno, 'is_broker': is_broker}

            portfolio_listing_path = 'admin/dashboard/portfolio/portfolio-content.html'
            portfolio_listing_template = get_template(portfolio_listing_path)
            portfolio_listing_html = portfolio_listing_template.render(context)
            # ---------------Pagination--------
            pagination_path = 'admin/dashboard/portfolio/pagination.html'
            pagination_template = get_template(pagination_path)
            total_page = math.ceil(int(total) / int(page_size))
            pagination_html = ''
            if total_page > 1:
                pagination_data = {"no_page": int(total_page), "total_page": range(total_page),
                                   "current_page": int(page),
                                   "pagination_id": "portfolio_pagination_list"}
                pagination_html = pagination_template.render(pagination_data)

            data = {'portfolio_listing_html': portfolio_listing_html, 'status': 200, 'msg': '', 'error': 0,
                    'total': total,
                    "pagination_html": pagination_html, 'pagination_id': 'portfolio_pagination_list'}
            return JsonResponse(data)
        else:
            list_param = {
                'domain_id': site_id,
                "page": page,
                "user_id": user_id,
                "page_size": page_size,
                "search": ""
            }
            sno = (int(page) - 1) * int(page_size) + 1
            list_url = settings.API_URL + '/api-property/portfolio-listing/'
            list_data = call_api_post_method(list_param, list_url, token)
            if 'error' in list_data and list_data['error'] == 0:
                portfolio_list = list_data['data']['data']
                total = list_data['data']['total']
            else:
                portfolio_list = []
                total = 0

            # ---------------Pagination--------
            pagination_html = ''
            pagination_path = 'admin/dashboard/portfolio/pagination.html'
            pagination_template = get_template(pagination_path)
            total_page = math.ceil(total / page_size)
            if total_page > 1:
                pagination_data = {"no_page": total_page, "total_page": range(total_page), "current_page": 1,
                                   "pagination_id": "estimator_listing_pagination_list"}
                pagination_html = pagination_template.render(pagination_data)


            context = {"portfolio_list": portfolio_list,
                       "total": total, "pagination_html": pagination_html,
                       "pagination_id": "portfolio_pagination_list", "active_menu": "portfolio_list",
                       "active_submenu": "portfolio_list", "sno": sno, "is_broker": is_broker}

            return render(request, "admin/dashboard/portfolio/view-portfolio.html", context)
    except Exception as exp:
        print(exp)
        return HttpResponse("Issue in views")

def add_portfolio(request):
    # is_permission = check_permission(request, 1)
    # if not is_permission:
    #     http_host = request.META['HTTP_HOST']
    #     redirect_url = settings.URL_SCHEME + str(http_host)
    #     return HttpResponseRedirect(redirect_url)
    portfolio_id = request.GET.get('id', '')
    # if not request.is_ajax() and agent_id is None and request.session['is_broker'] == False:
    #     http_host = request.META['HTTP_HOST']
    #     redirect_agent = settings.URL_SCHEME +str(http_host)+'/admin/agents/'
    #     return HttpResponseRedirect(redirect_agent)
    try:
        site_detail = subdomain_site_details(request)
        site_id = site_detail['site_detail']['site_id']
        token = request.session['token']['access_token']
        user_id = request.session['user_id']
        try:
            prop_param = {"domain_id": site_id, "user_id": user_id, "portfolio_id": portfolio_id}
            prop_url = settings.API_URL + '/api-property/portfolio-property-list/'
            prop_data = call_api_post_method(prop_param, prop_url, token)
            property_list = prop_data['data']
        except:
            property_list = []

        try:
            portfolio_detail_param = {
                'domain_id': site_id,
                'user_id': user_id,
                'portfolio_id': portfolio_id,
            }
            portfolio_detail_url = settings.API_URL + '/api-property/portfolio-detail/'
            portfolio_detail_data = call_api_post_method(portfolio_detail_param, portfolio_detail_url, token)


            portfolio_detail = portfolio_detail_data['data']
            portfolio_detail['property_id_in'] = [int(item['property']) for item in
                                                             portfolio_detail['property'] if
                                                             'property' in portfolio_detail and len(
                                                                 portfolio_detail['property']) > 0]
            portfolio_id = portfolio_detail['id']
            str_prop_image_id = ''
            prop_image_id = ''
            str_prop_image_name = ''
            prop_image_name = ''
            portfolio_img_list = portfolio_detail['images']
            if portfolio_img_list:
                cnt = 0
                for item in portfolio_img_list:
                    if cnt == 0:
                        str_prop_image_id = str(item['doc_id'])
                        str_prop_image_name = str(item['doc_file_name'])
                    else:
                        str_prop_image_id = str_prop_image_id + ',' + str(item['doc_id'])
                        str_prop_image_name = str_prop_image_name + ',' + str(item['doc_file_name'])
                    cnt += 1
                prop_image_id = str_prop_image_id.rstrip(',')
                prop_image_name = str_prop_image_name.rstrip(',')
                portfolio_detail['portfolio_image_ids'] = prop_image_id
                portfolio_detail['portfolio_image_name'] = prop_image_name
            else:
                portfolio_detail['portfolio_image_ids'] = prop_image_id
                portfolio_detail['portfolio_image_name'] = prop_image_name

        except:
            portfolio_detail = {}

        if request.is_ajax() and request.method == 'POST':
            portfolio_image_list = []
            if request.POST['portfolio_image_id'] != "":
                portfolio_image_list = request.POST['portfolio_image_id'].split(',')

            portfolio_params = {
                "domain": site_id,
                "user": user_id,
                "name": request.POST['portfolio_name'],
                "details": request.POST['portfolio_details'] if 'portfolio_details' in request.POST and request.POST['portfolio_details'] != "" else "",
                "terms": request.POST['portfolio_terms'] if 'portfolio_terms' in request.POST and request.POST['portfolio_terms'] != "" else "",
                "contact": request.POST['portfolio_contacts'] if 'portfolio_contacts' in request.POST and request.POST['portfolio_contacts'] != "" else "",
                "status": request.POST['portfolio_status'],
                "property_id": request.POST.getlist('property_name') if 'property_name' in request.POST else [],
                "portfolio_image": portfolio_image_list
            }

            if request.POST['portfolio_id']:
                portfolio_params['portfolio_id'] = request.POST['portfolio_id']

            portfolio_url = settings.API_URL + '/api-property/add-portfolio/'
            portfolio_response = call_api_post_method(portfolio_params, portfolio_url, token)
            if 'error' in portfolio_response and portfolio_response['error'] == 0:
                response = {
                    'error': 0,
                    'msg': portfolio_response['msg']
                }
            else:
                response = {
                    'error': 1,
                    'msg': portfolio_response['msg']
                }
            return JsonResponse(response)

        context = {"active_menu": "portfolio_list", "property_list": property_list, 'portfolio_detail': portfolio_detail}
        return render(request, "admin/dashboard/portfolio/add-portfolio.html", context)
    except Exception as exp:
        print(exp)
        return HttpResponse("Issue in views")


@csrf_exempt
def send_reset_password_link(request):
    try:
        if request.is_ajax() and request.method == 'POST':
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']
            params = {
                'domain_id': site_id,
                'user_id': request.POST['reset_user_id'],
            }
            token = request.session['token']['access_token']
            url = settings.API_URL + '/api-users/send-reset-password-link/'
            data = call_api_post_method(params, url, token)
            return JsonResponse(data)
        else:
            data = {'status': 403, 'msg': 'Forbidden'}
        return JsonResponse(data)
    except Exception as exp:
        data = {'status': 403, 'msg': 'invalid request.'}
        return JsonResponse(data)


@csrf_exempt
def property_total_view(request):
    try:
        user_id = None
        page_size = 10
        try:
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']

        except Exception as exp:
            site_id = ""

        if 'user_id' in request.session and request.session['user_id']:
            token = request.session['token']['access_token']
            user_id = request.session['user_id']

        if request.is_ajax() and request.method == 'POST':
            page = 1
            if 'page' in request.POST and request.POST['page'] != "":
                page = request.POST['page']

            if 'page_size' in request.POST and request.POST['page_size'] != "":
                page_size = request.POST['page_size']

            search = ''
            if 'search' in request.POST and request.POST['search'] != "":
                search = request.POST['search']

            property_id = request.POST['property_id']
            params = {
                'site_id': site_id,
                'user_id': user_id,
                'page_size': page_size,
                'page': page,
                'property_id': property_id,
                'search': search,
            }

            api_url = settings.API_URL + '/api-property/property-total-view/'
            total_property_view = call_api_post_method(params, api_url, token=token)
            try:
                prop_detail = total_property_view['data']['property_detail']
                image = prop_detail['property_image']
                if image and image['image'] and image['image'] != "":
                    image_url = settings.AWS_URL + image['bucket_name'] + '/' + image['image']
                else:
                    image_url = ''
                property_address = prop_detail['address_one']
                property_city = prop_detail['city']
                property_state = prop_detail['state']
                property_postal_code = prop_detail['postal_code']
                property_image = image_url
                auction_type = prop_detail['auction_type']
                bid_increment = prop_detail['bid_increment'] if prop_detail['bid_increment'] else ""
            except:
                property_address = property_city = property_state = property_postal_code = property_image = auction_type = bid_increment = ''

            if 'error' in total_property_view and total_property_view['error'] == 0:
                total = total_property_view['data']['total']
                total_property_view = total_property_view['data']['data']

            else:
                bid_history = []
                total = 0

            context = {'total_property_view': total_property_view, 'total': total, "aws_url": settings.AWS_URL,
                       'start_index': (int(page) - 1) * int(page_size)}
            html_path = 'admin/dashboard/listings/property-total-view.html'
            html_template = get_template(html_path)
            html = html_template.render(context)
            # ---------------Pagination--------
            pagination_path = 'admin/dashboard/listings/property-total-view-pagination.html'
            pagination_template = get_template(pagination_path)
            total_page = math.ceil(int(total) / int(page_size))
            pagination_html = ''
            if total_page > 1:
                pagination_data = {"no_page": int(total_page), "total_page": range(total_page),
                                   "current_page": int(page),
                                   "pagination_id": "propertyTotalViewPaginationList", "property_id": property_id}
                pagination_html = pagination_template.render(pagination_data)

            data = {
                'html': html,
                'status': 200,
                'msg': '',
                'error': 0,
                'total': total,
                "pagination_html": pagination_html,
                'pagination_id': 'propertyTotalViewPaginationList',
                'property_address': property_address,
                'property_city': property_city,
                'property_state': property_state,
                'property_postal_code': property_postal_code,
                'property_image': property_image,
                'auction_type': auction_type,
                'property_id': property_id,
                'page': page,
                'page_size': page_size,
                'bid_increment': f"{bid_increment:,}" if bid_increment != "" else "",
            }
        else:
            data = {'status': 403, 'msg': 'Forbidden', 'error': 1}
        return JsonResponse(data)
    except Exception as exp:
        return HttpResponse("Issue in views")


@csrf_exempt
def export_property_total_view(request):
    try:
        """
            Downloads property total view  as Excel file with a single worksheet
        """
        try:
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']

        except Exception as exp:
            print(exp)
            site_id = ""

        user_id = None
        token = None
        is_broker = 0
        if 'user_id' in request.session and request.session['user_id']:
            token = request.session['token']['access_token']
            user_id = request.session['user_id']
            is_broker = 1 if request.session['is_broker'] == True else 0

        search = request.GET.get('search', '')
        page = request.GET.get('page', 1)
        page_size = request.GET.get('page_size', 10)
        property_id = request.GET.get('property', '')
        timezone = request.GET.get('timezone', '')

        list_param = {
            'site_id': site_id,
            'user_id': user_id,
            'page_size': page_size,
            'page': page,
            'property_id': property_id,
            'search': search,
        }

        list_url = settings.API_URL + '/api-property/property-total-view/'

        list_data = call_api_post_method(list_param, list_url, token=token)
        if 'error' in list_data and list_data['error'] == 0:
            property_total_view = list_data['data']['data']
            total = list_data['data']['total'] if 'total' in list_data['data'] else 0
        else:
            property_total_view = []
            total = 0
        sno = (int(page) - 1) * int(page_size) + 1


        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename={date}-property-total-view.xlsx'.format(
            date=datetime.datetime.now().strftime('%Y-%m-%d'),
        )
        workbook = Workbook()

        # Get active worksheet/tab
        worksheet = workbook.active
        worksheet.title = 'Property Total Viewer'

        # Define the titles for columns
        columns = [
            '#',
            'Name',
            'Email',
            'Phone',
            'Date',
        ]
        row_num = 1
        header_font = Font(name='Calibri', bold=True)
        # Assign the titles for each cell of the header
        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title
            cell.font = header_font

        # Iterate through all history
        count = 0
        for viewer in property_total_view:
            row_num += 1
            phone_no = viewer['phone_no']
            formatted_phone_no = format_phone_number(phone_no)
            added_on = viewer['added_on']
            added_on_time = ''
            if timezone:
                try:
                    added_on_time = time.mktime(
                        datetime.datetime.strptime(added_on, "%Y-%m-%dT%H:%M:%SZ").timetuple())
                except ValueError:
                    added_on_time = time.mktime(datetime.datetime.strptime(added_on, "%Y-%m-%dT%H:%M:%S.%fZ").timetuple())
                except:
                    added_on_time = 0
            if added_on_time:
                added_on_time = float(added_on_time) - (float(timezone)*60)
                added_on_date_time = datetime.datetime.fromtimestamp(added_on_time)
                added_on_date_time = datetime.datetime.strftime(added_on_date_time, "%m-%d-%Y %I:%M %p")
            # Define the data for each cell in the row
            row = [
                sno, viewer['first_name']+' '+viewer['last_name'], viewer['email'], formatted_phone_no, added_on_date_time,
            ]
            sno += 1

            # Assign the data for each cell of the row
            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value
        workbook.save(response)
        return response
    except Exception as exp:
        return HttpResponse("Issue in views")


@csrf_exempt
def email_all_property_viewer(request):
    try:
        if request.is_ajax() and request.method == 'POST':
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']
            token = request.session['token']['access_token']

            user_params = {
                'domain_id': site_id,
                'user_id': request.session['user_id'],
                "property_id": request.POST['property_id'],
                "subject": request.POST['subject'],
                "message": request.POST['message'],
                "email_for": request.POST['email_for']
            }
            response = settings.API_URL + '/api-users/send-custom-email/'

            data = call_api_post_method(user_params, response, token)
            if 'error' in data and data['error'] == 0:
                data = {'status': 200, 'msg': data['msg'], 'error': 0}
            else:
                data = {'status': 403, 'msg': 'Something went wrong!', 'error': 1}
        else:
            data = {'status': 403, 'msg': 'Forbidden', 'error': 1}
        return JsonResponse(data)
    except Exception as exp:
        data = {'status': 403, 'msg': 'Something went wrong!', 'error': 1}
        return JsonResponse(data)


@csrf_exempt
def property_total_watcher(request):
    try:
        user_id = None
        page_size = 10
        try:
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']

        except Exception as exp:
            site_id = ""

        if 'user_id' in request.session and request.session['user_id']:
            token = request.session['token']['access_token']
            user_id = request.session['user_id']

        if request.is_ajax() and request.method == 'POST':
            page = 1
            if 'page' in request.POST and request.POST['page'] != "":
                page = request.POST['page']

            if 'page_size' in request.POST and request.POST['page_size'] != "":
                page_size = request.POST['page_size']

            search = ''
            if 'search' in request.POST and request.POST['search'] != "":
                search = request.POST['search']

            property_id = request.POST['property_id']
            params = {
                'site_id': site_id,
                'user_id': user_id,
                'page_size': page_size,
                'page': page,
                'property_id': property_id,
                'search': search,
            }

            api_url = settings.API_URL + '/api-property/property-total-watcher/'
            all_data = call_api_post_method(params, api_url, token=token)
            try:
                prop_detail = all_data['data']['property_detail']
                image = prop_detail['property_image']
                if image and image['image'] and image['image'] != "":
                    image_url = settings.AWS_URL + image['bucket_name'] + '/' + image['image']
                else:
                    image_url = ''
                property_address = prop_detail['address_one']
                property_city = prop_detail['city']
                property_state = prop_detail['state']
                property_postal_code = prop_detail['postal_code']
                property_image = image_url
                auction_type = prop_detail['auction_type']
                bid_increment = prop_detail['bid_increment']
            except:
                property_address = property_city = property_state = property_postal_code = property_image = auction_type = bid_increment = ''

            if 'error' in all_data and all_data['error'] == 0:
                total = all_data['data']['total']
                all_data = all_data['data']['data']

            else:
                bid_history = []
                total = 0

            context = {'all_data': all_data, 'total': total, "aws_url": settings.AWS_URL,
                       'start_index': (int(page) - 1) * int(page_size)}
            html_path = 'admin/dashboard/listings/property-total-watcher.html'
            html_template = get_template(html_path)
            html = html_template.render(context)
            # ---------------Pagination--------
            pagination_path = 'admin/dashboard/listings/property-total-watcher-pagination.html'
            pagination_template = get_template(pagination_path)
            total_page = math.ceil(int(total) / int(page_size))
            pagination_html = ''
            if total_page > 1:
                pagination_data = {"no_page": int(total_page), "total_page": range(total_page),
                                   "current_page": int(page),
                                   "pagination_id": "propertyTotalWatcherPaginationList", "property_id": property_id}
                pagination_html = pagination_template.render(pagination_data)
            data = {
                'html': html,
                'status': 200,
                'msg': '',
                'error': 0,
                'total': total,
                "pagination_html": pagination_html,
                'pagination_id': 'propertyTotalWatcherPaginationList',
                'property_address': property_address,
                'property_city': property_city,
                'property_state': property_state,
                'property_postal_code': property_postal_code,
                'property_image': property_image,
                'auction_type': auction_type,
                'property_id': property_id,
                'page': page,
                'page_size': page_size,
                # 'bid_increment': f"{bid_increment:,}",
                'bid_increment': "{:,}".format(bid_increment) if bid_increment is not None and bid_increment > 0 else "",
            }

        else:
            data = {'status': 403, 'msg': 'Forbidden', 'error': 1}
        return JsonResponse(data)
    except Exception as exp:
        return HttpResponse("Issue in views")


@csrf_exempt
def export_property_total_watcher(request):
    try:
        """
            Downloads property total watcher  as Excel file with a single worksheet
        """
        try:
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']

        except Exception as exp:
            print(exp)
            site_id = ""

        user_id = None
        token = None
        is_broker = 0
        if 'user_id' in request.session and request.session['user_id']:
            token = request.session['token']['access_token']
            user_id = request.session['user_id']
            is_broker = 1 if request.session['is_broker'] == True else 0

        search = request.GET.get('search', '')
        page = request.GET.get('page', 1)
        page_size = request.GET.get('page_size', 10)
        property_id = request.GET.get('property', '')
        timezone = request.GET.get('timezone', '')

        list_param = {
            'site_id': site_id,
            'user_id': user_id,
            'page_size': page_size,
            'page': page,
            'property_id': property_id,
            'search': search,
        }

        list_url = settings.API_URL + '/api-property/property-total-watcher/'

        list_data = call_api_post_method(list_param, list_url, token=token)
        if 'error' in list_data and list_data['error'] == 0:
            property_total_view = list_data['data']['data']
            total = list_data['data']['total'] if 'total' in list_data['data'] else 0
        else:
            property_total_view = []
            total = 0
        sno = (int(page) - 1) * int(page_size) + 1


        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename={date}-property-total-watcher.xlsx'.format(
            date=datetime.datetime.now().strftime('%Y-%m-%d'),
        )
        workbook = Workbook()

        # Get active worksheet/tab
        worksheet = workbook.active
        worksheet.title = 'Property Total Viewer'

        # Define the titles for columns
        columns = [
            '#',
            'Name',
            'Email',
            'Phone',
            'Date',
        ]
        row_num = 1
        header_font = Font(name='Calibri', bold=True)
        # Assign the titles for each cell of the header
        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title
            cell.font = header_font

        # Iterate through all history
        count = 0
        for viewer in property_total_view:
            row_num += 1
            phone_no = viewer['phone_no']
            formatted_phone_no = format_phone_number(phone_no)
            added_on = viewer['added_on']
            added_on_time = ''
            if timezone:
                try:
                    added_on_time = time.mktime(
                        datetime.datetime.strptime(added_on, "%Y-%m-%dT%H:%M:%SZ").timetuple())
                except ValueError:
                    added_on_time = time.mktime(datetime.datetime.strptime(added_on, "%Y-%m-%dT%H:%M:%S.%fZ").timetuple())
                except:
                    added_on_time = 0
            if added_on_time:
                added_on_time = float(added_on_time) - (float(timezone)*60)
                added_on_date_time = datetime.datetime.fromtimestamp(added_on_time)
                added_on_date_time = datetime.datetime.strftime(added_on_date_time, "%m-%d-%Y %I:%M %p")
            # Define the data for each cell in the row
            row = [
                sno, viewer['first_name']+' '+viewer['last_name'], viewer['email'], formatted_phone_no, added_on_date_time,
            ]
            sno += 1

            # Assign the data for each cell of the row
            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value
        workbook.save(response)
        return response
    except Exception as exp:
        return HttpResponse("Issue in views")


@csrf_exempt
def user_property_bid_history(request):
    try:
        user_id = user_id = None
        page_size = 10
        try:
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']

        except Exception as exp:
            print(exp)
            site_id = ""

        if 'user_id' in request.session and request.session['user_id']:
            token = request.session['token']['access_token']
            user_id = request.session['user_id']

        if request.is_ajax() and request.method == 'POST':
            page = 1
            if 'page' in request.POST and request.POST['page'] != "":
                page = request.POST['page']

            if 'page_size' in request.POST and request.POST['page_size'] != "":
                page_size = request.POST['page_size']

            search = ''
            if 'search' in request.POST and request.POST['search'] != "":
                search = request.POST['search']

            property_id = request.POST['property_id']
            register_user =  request.POST['register_user'] if 'register_user' in request.POST and request.POST['register_user'] else ''
            params = {
                'site_id': site_id,
                'user_id': user_id,
                'page_size': page_size,
                'page': page,
                'property_id': property_id,
                'register_user':register_user,
                'search': search,
            }
            api_url = settings.API_URL + '/api-bid/user-bid-history/'
            bid_history = call_api_post_method(params, api_url, token=token)
            try:
                prop_detail = bid_history['data']['property_detail']
                image = prop_detail['property_image']
                if image and image['image'] and image['image'] != "":
                    image_url = settings.AWS_URL + image['bucket_name'] + '/' + image['image']
                else:
                    image_url = ''
                property_address = prop_detail['address_one']
                property_city = prop_detail['city']
                property_state = prop_detail['state']
                property_postal_code = prop_detail['postal_code']
                property_image = image_url
                auction_type = prop_detail['auction_type']
                bid_increment = prop_detail['bid_increment']
            except:
                property_address = property_city = property_state = property_postal_code = property_image=auction_type=bid_increment=''

            if 'error' in bid_history and bid_history['error'] == 0:
                total = bid_history['data']['total']
                # new_bid_history = bid_history['data']['new_data']
                bid_history = bid_history['data']['data']
            else:
                bid_history = []
                # new_bid_history = []
                total = 0

            context = {'bid_history': bid_history, 'total': total, "aws_url": settings.AWS_URL, 'start_index': (int(page) - 1) * int(page_size)}
            # context = {'new_bid_history': new_bid_history, 'total': total, "aws_url": settings.AWS_URL, 'start_index': (int(page) - 1) * int(page_size)}

            bidder_listing_path = 'admin/dashboard/listings/property-bid-history.html'
            # bidder_listing_path = 'admin/dashboard/listings/new_property-bid-history.html'
            bidder_listing_template = get_template(bidder_listing_path)
            bid_history_html = bidder_listing_template.render(context)
            # ---------------Pagination--------
            pagination_path = 'admin/dashboard/listings/user-bid-history-pagination.html'
            # pagination_path = 'admin/dashboard/listings/new-bid-history-pagination.html'
            pagination_template = get_template(pagination_path)
            total_page = math.ceil(int(total) / int(page_size))
            pagination_html = ''
            if total_page > 1:
                pagination_data = {"no_page": int(total_page), "total_page": range(total_page), "current_page": int(page),
                                   "pagination_id": "bidHistoryPaginationList", "property_id": property_id}
                pagination_html = pagination_template.render(pagination_data)

            data = {
                'bid_history_html': bid_history_html,
                'status': 200,
                'msg': '',
                'error': 0,
                'total': total,
                "pagination_html": pagination_html,
                'pagination_id': 'bidHistoryPaginationList',
                'property_address': property_address,
                'property_city': property_city,
                'property_state': property_state,
                'property_postal_code': property_postal_code,
                'property_image': property_image,
                'auction_type': auction_type,
                'property_id': property_id,
                'page': page,
                'page_size': page_size,
                'bid_increment': f"{bid_increment:,}",
            }
        else:
            data = {'status': 403, 'msg': 'Forbidden', 'error': 1}
        return JsonResponse(data)
    except Exception as exp:
        print(exp)
        return HttpResponse("Issue in views")


@csrf_exempt
def property_total_bid(request):
    try:
        user_id = None
        page_size = 10
        try:
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']

        except Exception as exp:
            site_id = ""

        if 'user_id' in request.session and request.session['user_id']:
            token = request.session['token']['access_token']
            user_id = request.session['user_id']

        if request.is_ajax() and request.method == 'POST':
            page = 1
            if 'page' in request.POST and request.POST['page'] != "":
                page = request.POST['page']

            if 'page_size' in request.POST and request.POST['page_size'] != "":
                page_size = request.POST['page_size']

            # search = ''
            # if 'search' in request.POST and request.POST['search'] != "":
            #     search = request.POST['search']

            bid_id = request.POST['bid_id']
            params = {
                'site_id': site_id,
                'user_id': user_id,
                'page_size': page_size,
                'page': page,
                'bid_id': bid_id,
                # 'search': search,
            }

            api_url = settings.API_URL + '/api-bid/total-bid-history/'
            all_data = call_api_post_method(params, api_url, token=token)
            try:
                prop_detail = all_data['data']['property_detail']
                image = prop_detail['property_image']
                if image and image['image'] and image['image'] != "":
                    image_url = settings.AWS_URL + image['bucket_name'] + '/' + image['image']
                else:
                    image_url = ''
                property_address = prop_detail['address_one']
                property_city = prop_detail['city']
                property_state = prop_detail['state']
                property_postal_code = prop_detail['postal_code']
                property_image = image_url
                auction_type = prop_detail['auction_type']
                bid_increment = prop_detail['bid_increment']
                property_id = prop_detail['id']
                property_type = prop_detail['property_type']
            except:
                property_address = property_city = property_state = property_postal_code = property_image = auction_type = bid_increment = ''

            if 'error' in all_data and all_data['error'] == 0:
                total = all_data['data']['total']
                all_data = all_data['data']['data']

            else:
                all_data = []
                total = 0

            context = {'all_data': all_data, 'total': total, "aws_url": settings.AWS_URL,
                       'start_index': (int(page) - 1) * int(page_size)}
            html_path = 'admin/dashboard/listings/property-total-bid.html'
            html_template = get_template(html_path)
            html = html_template.render(context)
            # ---------------Pagination--------
            pagination_path = 'admin/dashboard/listings/property-total-bid-pagination.html'
            pagination_template = get_template(pagination_path)
            total_page = math.ceil(int(total) / int(page_size))
            pagination_html = ''
            if total_page > 1:
                pagination_data = {"no_page": int(total_page), "total_page": range(total_page),
                                   "current_page": int(page),
                                   "pagination_id": "propertyTotalBidPaginationList", "bid_id": bid_id }
                pagination_html = pagination_template.render(pagination_data)

            data = {
                'html': html,
                'status': 200,
                'msg': '',
                'error': 0,
                'total': total,
                "pagination_html": pagination_html,
                'pagination_id': 'propertyTotalBidPaginationList',
                'property_address': property_address,
                'property_city': property_city,
                'property_state': property_state,
                'property_postal_code': property_postal_code,
                'property_image': property_image,
                'auction_type': auction_type,
                'property_id': property_id,
                'page': page,
                'page_size': page_size,
                'bid_increment': f"{bid_increment:,}",
                'property_type': property_type
            }
        else:
            data = {'status': 403, 'msg': 'Forbidden', 'error': 1}
        return JsonResponse(data)
    except Exception as exp:
        return HttpResponse("Issue in views")


@csrf_exempt
def send_message_to_agent(request):
    try:
        if request.is_ajax() and request.method == 'POST':
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']
            user_id = None
            token = None
            if 'user_id' in request.session:
                token = request.session['token']['access_token']
                user_id = request.session['user_id']

            params = {
                "site_id": site_id,
                "agent_id": request.POST['agent_id'],
                "user_id": user_id,
                "message": request.POST['user_message']
            }

            api_url = settings.API_URL + '/api-contact/chat-to-agent/'

            enq_data = call_api_post_method(params, api_url, token=token)
            if 'error' in enq_data and enq_data['error'] == 0:
                data = {'status': 200, 'msg': enq_data['msg'], 'error': 0}
            else:
                data = {'status': 403, 'msg': enq_data['msg'], 'error': 1}
        else:
            data = {'status': 403, 'msg': 'Forbidden', 'error': 1}

        return JsonResponse(data)
    except Exception as exp:
        data = {'status': 403, 'msg': 'Invalid request', 'error': 1}
        return JsonResponse(data)


@csrf_exempt
def update_subscription_plan(request):
    try:
        if request.is_ajax() and request.method == 'POST':
            site_detail = subdomain_site_details(request)
            domain_id = site_detail['site_detail']['site_id']
            token = request.session['token']['access_token']
            user_id = request.session['user_id']
            plan_price_id = request.POST['plan_price_id']
            params = {
                "domain_id": domain_id,
                "user_id": user_id,
                "plan_price_id": plan_price_id
            }

            api_url = settings.API_URL + '/api-payments/change-plan-subscription/'
            api_response = call_api_post_method(params, api_url, token=token)
            if 'error' in api_response and api_response['error'] == 0:
                data = {'status': 200, 'msg': api_response['msg'], 'error': 0}
            else:
                data = {'status': 403, 'msg': api_response['msg'], 'error': 1}
        else:
            data = {'status': 403, 'msg': 'Forbidden', 'error': 1}

        return JsonResponse(data)
    except Exception as exp:
        data = {'status': 403, 'msg': 'Invalid request', 'error': 1}
        return JsonResponse(data)


@csrf_exempt
def check_payment(request):
    try:
        if request.is_ajax() and request.method == 'POST':
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']
            user_id = request.session['user_id']
            token = request.session['token']['access_token']

            api_url = settings.API_URL + "/api-payments/check-payment-success/"
            params = {
                "domain_id": site_id,
                "user_id": user_id,
            }
            api_response = call_api_post_method(params, api_url, token)
            if 'error' in api_response and api_response['error'] == 0:
                request.session['is_free_plan'] = False
                data = {"data": "", "error": 0, "msg": api_response['msg']}
            else:
                data = {'data': "", 'error': 1, "msg": api_response['msg']}

        else:
            data = {'data': "", 'error': 1, "msg": "Forbidden"}
        return JsonResponse(data)
    except Exception as exp:
        data = {'data': '', 'status': 403, 'error': 1, 'msg': 'invalid request.'}
        return JsonResponse(data)


@csrf_exempt
def testimonials(request):
    is_permission = check_permission(request, 11)
    if not is_permission:
        http_host = request.META['HTTP_HOST']
        redirect_url = settings.URL_SCHEME + str(http_host)
        return HttpResponseRedirect(redirect_url)
    page_size = request.GET.get('perpage', 10)
    status = request.GET.get('status', None)
    search = request.GET.get('search', None)
    page = request.GET.get('page', 1)
    try:
        site_detail = subdomain_site_details(request)
        site_id = site_detail['site_detail']['site_id']
        token = request.session['token']['access_token']
        user_id = request.session['user_id']
        if request.is_ajax() and request.method == 'POST':
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']
            agent_search = ''
            if request.POST['search']:
                agent_search = request.POST['search']

            if request.POST['perpage']:
                page_size = request.POST['perpage']
            else:
                page_size = 10

            if request.POST['page']:
                page = request.POST['page']
            else:
                page: 1
            if request.POST['status'] and request.POST['status'].lower() == 'active':
                status = [1]
            elif request.POST['status'] and request.POST['status'].lower() == 'inactive':
                status = [2]
            else:
                status = [2, 1]

            list_param = {
                'site_id': site_id,
                "page": page,
                "page_size": page_size,
                "status": status,
                "search": agent_search
            }
            list_url = settings.API_URL + '/api-users/testimonial-listing/'
            list_data = call_api_post_method(list_param, list_url, token)
            if 'error' in list_data and list_data['error'] == 0:
                testimonial_list = list_data['data']['data']
                total = list_data['data']['total']
                # ---------------Pagination--------
                pagination_path = 'admin/dashboard/newtork_testimonials/pagination.html'
                pagination_template = get_template(pagination_path)
                total_page = math.ceil(int(total) / int(page_size))
                pagination_html = ''
                if total_page > 1:
                    pagination_data = {"no_page": int(total_page), "total_page": range(total_page),
                                       "current_page": int(page), "pagination_id": "article_listing_pagination_list"}
                    pagination_html = pagination_template.render(pagination_data)
                data = {'article_list': testimonial_list, 'status': 200, 'msg': '', 'error': 0,
                        "pagination_html": pagination_html, "pagination_id": "article_listing_pagination_list",
                        "total": total}
            else:
                data = {'status': 403, 'msg': 'Server error, Please try again', 'article_list': [], 'error': 0}
            return JsonResponse(data)

        try:

            testimonial_list_param = {
                'site_id': site_id,
                "status": [1],
                'page': 1,
                'page_size': 10
            }
            sno = (int(page) - 1) * int(page_size) + 1
            testimonial_list_url = settings.API_URL + '/api-users/testimonial-listing/'
            testimonial_list_data = call_api_post_method(testimonial_list_param, testimonial_list_url, token)

            if 'error' in testimonial_list_data and testimonial_list_data['error'] == 0:
                testimonial_list = testimonial_list_data['data']['data']
                total = testimonial_list_data['data']['total']
            else:
                testimonial_list = []
                total = 0
            # ---------------Pagination--------
            pagination_html = ''
            pagination_path = 'admin/dashboard/newtork_testimonials/pagination.html'
            pagination_template = get_template(pagination_path)
            total_page = math.ceil(total / page_size)

            if total_page > 1:
                pagination_data = {"no_page": total_page, "total_page": range(total_page), "current_page": 1,
                                   "pagination_id": "testimonials_listing_pagination_list"}
                pagination_html = pagination_template.render(pagination_data)

        except Exception as exp:
            print(exp)
            testimonial_list = []
            total = 0
            pagination_html = ""

        context = {"active_menu": "cms", "active_submenu": "testimonials", "testimonial_list": testimonial_list, "total": total,
                   "pagination_html": pagination_html, "pagination_id": "article_listing_pagination_list", "sno": sno}
        return render(request, "admin/dashboard/newtork_testimonials/view-testimonials.html", context)
    except Exception as exp:
        print(exp)
        return HttpResponse("Issue in views")


def add_testimonials(request):
    is_permission = check_permission(request, 11)
    if not is_permission:
        http_host = request.META['HTTP_HOST']
        redirect_url = settings.URL_SCHEME + str(http_host)
        return HttpResponseRedirect(redirect_url)
    testimonial_id = request.GET.get('id', None)
    try:
        site_detail = subdomain_site_details(request)
        site_id = site_detail['site_detail']['site_id']
        token = request.session['token']['access_token']
        user_id = request.session['user_id']

        try:
            testimonial_detail_param = {
                'site_id': site_id,
                'testimonial_id': testimonial_id
            }
            testimonial_detail_url = settings.API_URL + '/api-users/testimonial-detail/'
            testimonial_detail_data = call_api_post_method(testimonial_detail_param, testimonial_detail_url, token)
            if 'error' in testimonial_detail_data and testimonial_detail_data['error'] == 1 and agent_id:
                http_host = request.META['HTTP_HOST']
                redirect_url = settings.URL_SCHEME + str(http_host)
                return HttpResponseRedirect(redirect_url + '/admin/testimonials/')
            testimonial_details = testimonial_detail_data['data']
            # article_id = testimonial_details['id']

        except:
            testimonial_details = {}

        if request.is_ajax() and request.method == 'POST':
            testimonial_params = {
                'site_id': site_id,
                'user_id': user_id,
                'description': request.POST['testimonial_description'],
                'author_name': request.POST['author_name'],
                'author_image': request.POST['article_author_image_id'],
                'status': request.POST['testimonial_status'],
                'type': request.POST['type']
            }
            if request.POST['testimonial_id']:
                testimonial_params['testimonial_id'] = request.POST['testimonial_id']
            testimonial_url = settings.API_URL + '/api-users/add-testimonial/'
            print(token)
            testimonial_response = call_api_post_method(testimonial_params, testimonial_url, token)
            if 'error' in testimonial_response and testimonial_response['error'] == 0:
                response = {
                    'error': 0,
                    'msg': 'Testimonial saved successfully'
                }
            else:
                response = {
                    'error': 1,
                    'msg': article_response['msg']
                }
            return JsonResponse(response)

        context = {"active_menu": "testimonials", "testimonial_details": testimonial_details, "asset_list": {}}
        return render(request, "admin/dashboard/newtork_testimonials/add-testimonials.html", context)
    except Exception as exp:
        print(exp)
        return HttpResponse("Issue in views")


@csrf_exempt
def testimonials_search_suggestion(request):
    try:
        if request.is_ajax() and request.method == 'POST':
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']
            token = request.session['token']['access_token']

            user_params = {
                'site_id': site_id,
                'search': request.POST['search']
            }

            user_url = settings.API_URL + '/api-users/subdomain-testimonials-suggestion/'

            suggestion_data = call_api_post_method(user_params, user_url, token)
            if 'error' in suggestion_data and suggestion_data['error'] == 0:
                data = {'status': 200, 'suggestion_list': suggestion_data['data'], 'error': 0}
            else:
                data = {'status': 403, 'suggestion_list': [], 'error': 1}
        else:
            data = {'status': 403, 'suggestion_list': [], 'error': 1}

        return JsonResponse(data)
    except Exception as exp:
        print(exp)
        data = {'status': 403, 'suggestion_list': [], 'error': 1}
        return JsonResponse(data)


def state_list(request):
    try:
        if request.is_ajax() and request.method == 'POST':
            country_id = request.POST['country_id']
            try:
                state_param = {'country_id': country_id}
                state_api_url = settings.API_URL + '/api-settings/get-state/'
                state_data = call_api_post_method(state_param, state_api_url)
                state_lists = state_data['data']
                data = {'error': 0, 'status': 200, 'state_lists': state_lists}
            except Exception as exp:
                data = {'error': 1, 'status': 403, 'state_lists': ''}
        else:
            data = {'error': 1, 'status': 403, 'state_lists': ''}
        return JsonResponse(data)
    except Exception as exp:
        data = {'error': 1, 'status': 403, 'msg': 'invalid request.'}
        return JsonResponse(data)


def update_dashboard(request):
    try:
        if request.is_ajax() and request.method == 'POST':
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']
            token = request.session['token']['access_token']

            start_date = request.POST.get('start_date', None)
            end_date = request.POST.get('end_date', None)
            if start_date and end_date:
                start_year_data = int(start_date.split("-")[0])
                end_year_data = int(end_date.split("-")[0])

                if start_year_data == end_year_data:
                    start_month_data = int(start_date.split("-")[1])
                    end_month_data = int(end_date.split("-")[1])
                    label_index = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "July", "Aug", "Sep", "Oct", "Nov", "Dec"]
                    label = [label_index[i] for i in range(start_month_data-1, end_month_data)]
                else:
                    label = [i for i in range(start_year_data, end_year_data+1)]
            else:
                label = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "July", "Aug", "Sep", "Oct", "Nov", "Dec"]

            api_url = settings.API_URL + '/api-users/signup-page-graph/'
            params = {
                "site_id": site_id,
                "start_date": start_date,
                "end_date": end_date
            }
            api_response = call_api_post_method(params, api_url, token)
            data = {"error": 0, "msg": "Success", "label": label}
            if 'error' in api_response and api_response['error'] == 0:
                data["signup_view"] = api_response['data']
            else:
                data["signup_view"] = {}
                data["error"] = 1
                data["msg"] = api_response['msg']

            api_url = settings.API_URL + '/api-users/property-registration-graph/'
            params = {
                "site_id": site_id,
                "start_date": start_date,
                "end_date": end_date
            }
            api_response = call_api_post_method(params, api_url, token)

            if 'error' in api_response and api_response['error'] == 0:
                data['property_registration'] = api_response['data']
            else:
                data['property_registration'] = {}
                data["error"] = 1
                data["msg"] = api_response['msg']

            api_url = settings.API_URL + '/api-users/admin-dashboard/'
            params = {
                "site_id": site_id,
                "start_date": start_date,
                "end_date": end_date
            }
            api_response = call_api_post_method(params, api_url, token)

            if 'error' in api_response and api_response['error'] == 0:
                data['dashboard_data'] = api_response['data']
            else:
                data['dashboard_data'] = ""
                data["error"] = 1
                data["msg"] = api_response['msg']

            api_url = settings.API_URL + '/api-users/update-dashboard-map/'
            params = {
                "site_id": site_id,
                "start_date": start_date,
                "end_date": end_date
            }
            api_response = call_api_post_method(params, api_url, token)
            if 'error' in api_response and api_response['error'] == 0:
                data['dashboard_map'] = api_response['data']
            else:
                data['dashboard_map'] = {}
                data["error"] = 1
                data["msg"] = api_response['msg']

        else:
            data = {'data': "", 'label': [], 'error': 1, "msg": "Forbidden"}
        return JsonResponse(data)
    except Exception as exp:
        print(exp)
        data = {'data': '', 'label': [], 'status': 403, 'error': 1, 'msg': 'invalid request.'}
        return JsonResponse(data)


def update_signup_view_graph(request):
    try:
        if request.is_ajax() and request.method == 'POST':
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']
            token = request.session['token']['access_token']

            start_date = request.POST.get('start_date', None)
            end_date = request.POST.get('end_date', None)
            if start_date and end_date:
                start_year_data = int(start_date.split("-")[0])
                end_year_data = int(end_date.split("-")[0])
                if start_year_data == end_year_data:
                    start_month_data = int(start_date.split("-")[1])
                    end_month_data = int(end_date.split("-")[1])
                    label_index = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "July", "Aug", "Sep", "Oct", "Nov", "Dec"]
                    label = [label_index[i] for i in range(start_month_data - 1, end_month_data)]
                else:
                    label = [i for i in range(start_year_data, end_year_data+1)]
            else:
                label = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "July", "Aug", "Sep", "Oct", "Nov", "Dec"]

            api_url = settings.API_URL + '/api-users/signup-page-graph/'
            params = {
                "site_id": site_id,
                "start_date": start_date,
                "end_date": end_date
            }
            api_response = call_api_post_method(params, api_url, token)

            if 'error' in api_response and api_response['error'] == 0:
                data = {"signup_view": api_response['data'], "label": label, "error": 0, "msg": api_response['msg']}
            else:
                data = {"signup_view": {}, "label": label, "error": 1, "msg": api_response['msg']}
        else:
            data = {'signup_view': {}, "label": label, 'error': 1, "msg": "Forbidden"}
        return JsonResponse(data)
    except Exception as exp:
        print(exp)
        data = {'signup_view': {}, "label": [], 'status': 403, 'error': 1, 'msg': 'invalid request.'}
        return JsonResponse(data)


def update_property_registration_graph(request):
    try:
        if request.is_ajax() and request.method == 'POST':
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']
            token = request.session['token']['access_token']

            start_date = request.POST.get('start_date', None)
            end_date = request.POST.get('end_date', None)

            if start_date and end_date:
                start_year_data = int(start_date.split("-")[0])
                end_year_data = int(end_date.split("-")[0])
                if start_year_data == end_year_data:
                    start_month_data = int(start_date.split("-")[1])
                    end_month_data = int(end_date.split("-")[1])
                    label_index = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "July", "Aug", "Sep", "Oct", "Nov", "Dec"]
                    label = [label_index[i] for i in range(start_month_data - 1, end_month_data)]
                else:
                    label = [i for i in range(start_year_data, end_year_data+1)]
            else:
                label = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "July", "Aug", "Sep", "Oct", "Nov", "Dec"]

            api_url = settings.API_URL + '/api-users/property-registration-graph/'
            params = {
                "site_id": site_id,
                "start_date": start_date,
                "end_date": end_date
            }
            api_response = call_api_post_method(params, api_url, token)

            if 'error' in api_response and api_response['error'] == 0:
                data = {"property_registration": api_response['data'], "label": label, "error": 0, "msg": api_response['msg']}
            else:
                data = {"property_registration": {}, "label": label, "error": 1, "msg": api_response['msg']}
        else:
            data = {'property_registration': {}, "label": label, 'error': 1, "msg": "Forbidden"}
        return JsonResponse(data)
    except Exception as exp:
        data = {'property_registration': {}, "label": [], 'status': 403, 'error': 1, 'msg': 'invalid request.'}
        return JsonResponse(data)


def update_dashboard_map(request):
    try:
        if request.is_ajax() and request.method == 'POST':
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']
            token = request.session['token']['access_token']

            start_date = request.POST.get('start_date', None)
            end_date = request.POST.get('end_date', None)

            api_url = settings.API_URL + '/api-users/update-dashboard-map/'
            params = {
                "site_id": site_id,
                "start_date": start_date,
                "end_date": end_date
            }
            api_response = call_api_post_method(params, api_url, token)

            if 'error' in api_response and api_response['error'] == 0:
                data = {"data": api_response['data'], "error": 0, "msg": api_response['msg']}
            else:
                data = {"data": [], "error": 1, "msg": api_response['msg']}
        else:
            data = {'data': [], 'error': 1, "msg": "Forbidden"}
        return JsonResponse(data)
    except Exception as exp:
        data = {'data': [], 'status': 403, 'error': 1, 'msg': 'invalid request.'}
        return JsonResponse(data)


@csrf_exempt
def close_tour(request):
    try:
        if request.is_ajax() and request.method == 'POST':
            request.session['is_first_admin_login'] = 0
            show_active_plan = 0
            if request.session['is_broker'] and request.session['is_free_plan']:
                show_active_plan = 1
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']
            user_id = request.session['user_id']
            token = request.session['token']['access_token']
            # ---------------------Save Payment Detail Data------------------
            plan_price_id = int(site_detail['site_detail']['current_plan_price_id'])
            theme_id = int(site_detail['site_detail']['current_theme_id'])
            api_url = settings.API_URL + "/api-payments/create-payment-detail/"
            params = {
                "domain_id": site_id,
                "user_id": user_id,
                "plan_price_id": plan_price_id,
                "theme_id": theme_id
            }
            api_response = call_api_post_method(params, api_url, token)
            if 'error' in api_response and api_response['error'] == 0:
                # -------------Stripe Payment Credentials-------------
                api_url = settings.API_URL + "/api-payments/stripe-payment-detail/"
                params = {
                    "plan_price_id": plan_price_id,
                }
                api_response = call_api_post_method(params, api_url, token)
                if 'error' in api_response and api_response['error'] == 0:
                    data = api_response['data']
                    data = {"data": data, "email": request.session['email'], "show_active_plan": show_active_plan, 'status': 200, "error": 0, "msg": "success", "pay_redirect": request.session['is_first_login']}
                    # print(data)
                else:
                    data = {'data': "", 'status': 403, 'error': 1, "msg": api_response['msg']}
            else:
                data = {'data': "", 'status': 403, 'error': 1, "msg": api_response['msg']}
        else:
            data = {'data': "", 'status': 403, 'error': 1, "msg": "Forbidden"}
        return JsonResponse(data)
    except Exception as exp:
        data = {'data': "", 'status': 403, 'error': 1, 'msg': 'invalid request.'}
        return JsonResponse(data)


@csrf_exempt
def send_email_verification_link(request):
    try:
        if request.is_ajax() and request.method == 'POST':
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']
            user_id = request.session['user_id']
            token = request.session['token']['access_token']

            api_url = settings.API_URL + "/api-users/send-email-verification-link/"
            params = {
                "domain_id": site_id,
                # "user_id": user_id,
            }
            api_response = call_api_post_method(params, api_url, token)
            if 'error' in api_response and api_response['error'] == 0:
                data = {'msg': api_response['msg'], 'status': 201, 'error': 0}
            else:
                data = {'msg': api_response['msg'], 'status': 403, 'error': 1}
        else:
            data = {'msg': 'Forbidden', 'status': 403, 'error': 1}
        return JsonResponse(data)
    except Exception as exp:
        print(exp)
        data = {'msg': 'invalid request.', 'status': 403, 'error': 1}
        return JsonResponse(data)


def upload_csv(request):
    try:
        if request.is_ajax() and request.method == 'POST':
            user_id = request.session['user_id']
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']
            token = request.session['token']['access_token']
            upload_to = 'property_image'
            file_urls = request.FILES.get('bulk_upload')
            response = save_to_s3(file_urls, upload_to)
            if 'error' in response and response['error'] == 0:
                upload_param = {
                    'user_id': user_id,
                    'domain_id': site_id,
                    'csv_url': settings.AWS_URL + upload_to + "/" + response['file_name']
                }
                url = settings.API_URL + '/api-property/bulk-property-upload/'
                upload_data = call_api_post_method(upload_param, url, token)
                if upload_data['error'] == 1:
                    return JsonResponse({'status': 404, 'error': 1, "msg": upload_data['msg']})
            return JsonResponse({'status': 200, 'error': 0, 'success': 1})
        else:
            return JsonResponse({'response': 403, 'msg': 'Forbidden', 'error': 1})
    except Exception as exp:
        print(exp)
        data = {'status': 403, 'msg': 'invalid request.', 'error': 1}
        return JsonResponse(data)


def payment_settings(request):
    try:
        site_detail = subdomain_site_details(request)
        site_id = site_detail['site_detail']['site_id']
        token = request.session['token']['access_token']
        user_id = request.session['user_id']
        user_url = settings.API_URL + '/api-users/get-payment-settings/'
        params = {
            'user_id': user_id,
            'domain_id': site_id,
        }
        response = call_api_post_method(params, user_url, token)
        if response['error'] == 0:
            data = response['data']
        else:
            data = {}
        context = {"title": "Payment Settings", "active_menu": "site setting", "active_submenu": "Payment Settings", "data": data}
        return render(request, "admin/dashboard/payment_settings/payment-settings.html", context)
    except Exception as exp:
        return HttpResponse("Issue in views")


def save_payment_settings(request):
    try:
        if request.is_ajax() and request.method == 'POST':
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']
            token = request.session['token']['access_token']
            user_id = request.session['user_id']
            user_url = settings.API_URL + '/api-users/save-payment-settings/'
            params = {
                'user': user_id,
                'domain': site_id,
                'stripe_public_key': request.POST['stripe_public_key'],
                'stripe_secret_key': request.POST['stripe_secret_key'],
                'payment_id': request.POST['payment_id']
            }
            response = call_api_post_method(params, user_url, token)
            if response['error'] == 0:
                data = {'status': 200, 'msg': response['msg'], 'error': 0}
            else:
                data = {'status': 403, 'msg': response['msg'], 'error': 1}
        else:
            data = {'status': 403, 'msg': 'Forbidden', 'error': 1}

        return JsonResponse(data)
    except Exception as exp:
        data = {'status': 403, 'msg': 'invalid request.', 'error': 1}
        return JsonResponse(data)  


# Cms Management
def cms_list(request):
    """ Use to list CMS
    """
    try:
        domain_list = get_network_domain_list(request)
        status_list = get_lookup_status_list(request, 16)

        return render(request, "admin/cms/cms-list.html", {"domain_list": domain_list, "status_list": status_list, "active_submenu": "add-cms"})
    except Exception as exp:
        print(exp)
        return HttpResponse("Issue in views")


@csrf_exempt
def ajax_cms_list(request):
    """ Use to load cms list from ajax
    """
    cms_data = {}
    pagination = {}
    page_size = 20
    page_size = int(request.POST['count']
                    ) if 'count' in request.POST and request.POST['count'] else 20
    current_page = int(request.POST['page']) if 'page' in request.POST else 1
    api_url = settings.API_URL + '/api-cms/admin-cms-listing/'
    params = {
        'page': current_page,
        'page_size': page_size,
        'search': request.POST['search'] if 'search' in request.POST else '',
        'site_id': request.POST.getlist('site_id[]'),
        'status': request.POST.getlist('status[]')
    }
    try:
        response = call_api_post_method(
            params, api_url, request.session['token']['access_token'])
    except Exception as exp:
        # msg = err
        response = {'msg': exp, 'status': 422}
    else:
        if "error" in response and response['error'] == 0:
            cms_data = response['data']
            total_contacts = response['data']['total']
            total_pages = total_contacts / page_size
            if total_contacts % page_size != 0:
                total_pages += 1  # adding one more page if the last page
                # will contains less contacts

            pagination = make_pagination_html(current_page, total_pages,
                                              'cms', 'cms_list')

    context = {
        'data': cms_data,
        'pagination': pagination,
        'start_index': (current_page - 1) * page_size
    }
    return render(request, 'admin/cms/ajax-cms-list.html', context)


def active_inactive_cms(request):
    """ Use to active/in-active cms
    """
    try:
        error = 1
        if request.method == 'POST':
            api_url = settings.API_URL + '/api-cms/cms-change-status/'
            payload = {
                'cms_id': request.POST['cms_id'],
                'status': request.POST['status']
            }
            try:
                response = call_api_post_method(
                    payload, api_url, request.session['token']['access_token'])
                msg = response['msg']
                error = response['error']
            except Exception as exp:
                msg = exp
        else:
            msg = 'Get action not allowed'

        return JsonResponse({"msg": msg, "error": error})
    except Exception as exp:
        print(exp)
        return JsonResponse({"msg": exp, "error": 1})


def add_cms(request):
    """ Use to add cms from admin end
    """
    try:
        domain_list = status_list = []
        if request.method == 'GET':  # add user
            domain_list = get_network_domain_list(request)
            status_list = get_lookup_status_list(request, 16)
        else:
            pass

        context = {
            'domain_list': domain_list,
            'status_list': status_list
        }
        return render(request, 'admin/cms/add-cms.html', context)
    except Exception as exp:
        print(exp)
        return


def ajax_add_cms(request):
    """ add cms from ajax
    """
    try:
        error = 1
        try:
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']
        except:
            site_id = ""
        if request.method == 'POST':

            api_url = settings.API_URL + '/api-cms/add-cms/'

            payload = {
                "site": None,
                "page_title": request.POST['page_title'],
                "meta_key_word": request.POST['meta_key_word'],
                "meta_description": request.POST['meta_description'],
                "meta_title": request.POST['meta_title'],
                "page_content": request.POST['page_content'],
                "added_by": request.session['user_id'],
                "status": request.POST['status'],
                'slug': request.POST['slug']
            }

            # check if edit cms
            if 'cms_id' in request.POST and request.POST:
                payload['cms_id'] = int(request.POST['cms_id'])

            try:
                response = call_api_post_method(payload, api_url, request.session['token']['access_token'])
                msg = response['msg']
                error = response['error']
            except Exception as exp:
                msg = exp

        else:
            msg = 'Get action not allowed'

        return JsonResponse({"msg": msg, "error": error})
    except Exception as exp:
        print(exp)
        return JsonResponse({"msg": exp, "error": 1})


def edit_cms(request, id):
    """ Use to add user from admin end
    """
    try:

        api_url = settings.API_URL + '/api-cms/cms-detail/'
        data = domain_list = status_list = []
        try:
            response = call_api_post_method({'cms_id': id}, api_url, request.session['token']['access_token'])
            if "error" in response and response['error'] == 0:
                data = response['data']
            else:  # Redirect if user detail not found
                # messages.error(request, response['msg'])
                return redirect('admin-cms-list')

        except Exception as exp:
            print(exp)
            # messages.error(request, response['msg'])
            return redirect('admin-cms-list')

        domain_list = get_network_domain_list(request)
        status_list = get_lookup_status_list(request, 16)
        context = {
            "data": data,
            "domain_list": domain_list,
            'status_list': status_list,
        }
        return render(request, 'admin/cms/edit-cms.html', context)
    except Exception as exp:
        print(exp)
        return


# FAQ Management
def faq_list(request):
    """ Use to list faq's
    """
    try:
        domain_list = status_list = []
        if request.method == 'GET':  # add blog
            domain_list = get_network_domain_list(request)
            status_list = get_lookup_status_list(request, 18)
        else:
            pass

        return render(request, "admin/faq/faq-list.html", {"domain_list": domain_list, "status_list": status_list, "active_submenu": "faq-list"})
    except Exception as exp:
        print(exp)
        return HttpResponse("Issue in views")


@csrf_exempt
def ajax_faq_list(request):
    """ Use to load faq list from ajax
    """
    cms_data = {}
    pagination = {}
    page_size = 20
    page_size = int(request.POST['count']
                    ) if 'count' in request.POST and request.POST['count'] else 20
    current_page = int(request.POST['page']) if 'page' in request.POST else 1
    api_url = settings.API_URL + '/api-faq/super-admin-faq-listing/'
    params = {
        'page': current_page,
        'page_size': page_size,
        'search': request.POST['search'] if 'search' in request.POST else '',
        'site_id': request.POST.getlist('site_id[]'),
        'status': request.POST.getlist('status[]')
    }
    try:
        response = call_api_post_method(
            params, api_url, request.session['token']['access_token'])
    except Exception as exp:
        # msg = err
        response = {'msg': exp, 'status': 422}
    else:
        if "error" in response and response['error'] == 0:
            cms_data = response['data']
            total_contacts = response['data']['total']
            total_pages = total_contacts / page_size
            if total_contacts % page_size != 0:
                total_pages += 1  # adding one more page if the last page
                # will contains less contacts

            pagination = make_pagination_html(current_page, total_pages,
                                              'faq', 'faq_list')

    context = {
        'data': cms_data,
        'pagination': pagination,
        'start_index': (current_page - 1) * page_size
    }
    return render(request, 'admin/faq/ajax-faq-list.html', context)


def add_faq(request):
    """ Use to add faq from admin end
    """
    try:
        domain_list = status_list = []
        if request.method == 'GET':  # add user
            domain_list = get_network_domain_list(request)
            status_list = get_lookup_status_list(request, 18)
        else:
            pass

        context = {
            'domain_list': domain_list,
            'status_list': status_list
        }
        return render(request, 'admin/faq/add-faq.html', context)
    except Exception as exp:
        print(exp)
        return


def ajax_add_faq(request):
    """ add faq from ajax
    """
    try:
        error = 1
        if request.method == 'POST':

            api_url = settings.API_URL + '/api-faq/super-admin-add-faq/'

            payload = {
                "domain": "",
                "question": request.POST['question'],
                "answer": request.POST['answer'],
                "added_by": request.session['user_id'],
                "status": request.POST['status'],
                "user_type": request.POST['user_type'],
            }

            # check if edit cms
            if 'faq_id' in request.POST and request.POST:
                payload['faq_id'] = int(request.POST['faq_id'])

            try:
                response = call_api_post_method(
                    payload, api_url, request.session['token']['access_token'])
                msg = response['msg']
                error = response['error']
            except Exception as exp:
                msg = exp

        else:
            msg = 'Get action not allowed'

        return JsonResponse({"msg": msg, "error": error})
    except Exception as exp:
        print(exp)
        return JsonResponse({"msg": exp, "error": 1})


def edit_faq(request, id):
    """ Use to add user from admin end
    """
    try:

        api_url = settings.API_URL + '/api-faq/super-admin-faq-detail/'
        data = domain_list = status_list = []
        try:
            response = call_api_post_method(
                {'faq_id': id}, api_url, request.session['token']['access_token'])
            if "error" in response and response['error'] == 0:
                data = response['data']
            else:  # Redirect if user detail not found
                # messages.error(request, response['msg'])
                return redirect('admin-faq-list')

        except Exception as exp:
            # print(exp)
            # messages.error(request, response['msg'])
            return redirect('admin-faq-list')

        domain_list = get_network_domain_list(request)
        status_list = get_lookup_status_list(request, 18)

        context = {
            "data": data,
            "domain_list": domain_list,
            'status_list': status_list,
        }

        return render(request, 'admin/faq/edit-faq.html', context)
    except Exception as exp:
        # print(exp)
        return


# Blog Management
def blog_list(request):
    """ Use to list Blog
    """
    try:
        domain_list = status_list = []
        if request.method == 'GET':  # add blog
            domain_list = get_network_domain_list(request)
            status_list = get_lookup_status_list(request, 17)
        else:
            pass

        try:
            api_url = settings.API_URL + '/api-blog/blog-category/'
            asset_list = call_api_post_method(
                {}, api_url, request.session['token']['access_token'])
            if 'error' in asset_list and asset_list['error'] == 1:
                # messages.error(request, asset_list['msg'])
                return redirect('admin-article-list')
            asset_list = asset_list['data']
        except:
            asset_list = {}

        return render(request, "admin/article/article-list.html", {"domain_list": domain_list, "status_list": status_list, "asset_list": asset_list, "active_submenu": "blog-list"})
    except Exception as exp:
        print(exp)
        return HttpResponse("Issue in views")


@csrf_exempt
def ajax_blog_list(request):
    """ Use to load Blog list from ajax
    """
    article_data = {}
    pagination = {}
    page_size = 20
    page_size = int(request.POST['count']
                    ) if 'count' in request.POST and request.POST['count'] else 20
    current_page = int(request.POST['page']) if 'page' in request.POST else 1
    api_url = settings.API_URL + '/api-cms/admin-article-listing/'
    params = {
        'page': current_page,
        'page_size': page_size,
        'search': request.POST['search'] if 'search' in request.POST else '',
        'site_id': request.POST.getlist('site_id[]'),
        'status': request.POST.getlist('status[]'),
        'asset_type':request.POST.getlist('asset_type[]'),
    }
    try:
        # print(params)
        response = call_api_post_method(
            params, api_url, request.session['token']['access_token'])
        # print(response)
    except Exception as exp:
        # msg = err
        response = {'msg': exp, 'status': 422}
    else:
        if "error" in response and response['error'] == 0:
            article_data = response['data']
            total_contacts = response['data']['total']
            total_pages = total_contacts / page_size
            if total_contacts % page_size != 0:
                total_pages += 1  # adding one more page if the last page
                # will contains less contacts

            pagination = make_pagination_html(current_page, total_pages,
                                              'article', 'article_list')

    context = {
        'data': article_data,
        'pagination': pagination,
        'aws_url': settings.AWS_URL,
        'start_index': (current_page - 1) * page_size
    }
    return render(request, 'admin/article/ajax-article-list.html', context)


def active_inactive_blog(request):
    """ Use to active/in-active blog
    """
    try:
        error = 1
        if request.method == 'POST':
            api_url = settings.API_URL + '/api-cms/admin-article-change-status/'
            payload = {
                'article_id': request.POST['article_id'],
                'status': request.POST['status']
            }
            try:
                response = call_api_post_method(
                    payload, api_url, request.session['token']['access_token'])
                msg = response['msg']
                error = response['error']
            except Exception as exp:
                msg = exp
        else:
            msg = 'Get action not allowed'

        return JsonResponse({"msg": msg, "error": error})
    except Exception as exp:
        print(exp)
        return JsonResponse({"msg": exp, "error": 1})


def add_blog(request):
    """ Use to add blog from admin end
    """
    try:
        domain_list = status_list = []
        if request.method == 'GET':  # add blog
            domain_list = get_network_domain_list(request)
            status_list = get_lookup_status_list(request, 17)
        else:
            pass

        try:
            api_url = settings.API_URL + '/api-blog/blog-category/'
            asset_list = call_api_post_method(
                {}, api_url, request.session['token']['access_token'])
            if 'error' in asset_list and asset_list['error'] == 1:
                # messages.error(request, asset_list['msg'])
                return redirect('admin-article-list')
            asset_list = asset_list['data']
        except:
            asset_list = {}

        context = {
            'domain_list': domain_list,
            'status_list': status_list,
            'aws_url': settings.AWS_URL,
            'asset_list': asset_list

        }
        return render(request, 'admin/article/add-article.html', context)
    except Exception as exp:
        print(exp)
        return


def ajax_add_blog(request):
    """ add blog from ajax
    """
    try:
        error = 1
        if request.method == 'POST':

            api_url = settings.API_URL + '/api-cms/admin-add-article/'

            payload = {
                "site_id": None,
                "user_id": request.session['user_id'],
                "title": request.POST['title'],
                "description": request.POST['description'],
                "author_name": request.POST['author_name'],
                "author_image": request.POST['author_image'],
                "upload": request.POST['upload'],
                "status": request.POST['status'],
                "asset": request.POST['asset'],
                "publish_date": request.POST['publish_date']
            }

            # check if edit blog
            if 'article_id' in request.POST and request.POST:
                payload['article_id'] = int(request.POST['article_id'])

            try:
                response = call_api_post_method(
                    payload, api_url, request.session['token']['access_token'])
                msg = response['msg']
                error = response['error']
            except Exception as exp:
                msg = exp

        else:
            msg = 'Get action not allowed'

        return JsonResponse({"msg": msg, "error": error})
    except Exception as exp:
        print(exp)
        return JsonResponse({"msg": exp, "error": 1})


def edit_blog(request, id):
    """ Use to add user from admin end
    """
    try:

        api_url = settings.API_URL + '/api-cms/admin-article-detail/'
        data = domain_list = status_list = []
        try:
            response = call_api_post_method(
                {'article_id': id}, api_url, request.session['token']['access_token'])
            if "error" in response and response['error'] == 0:
                data = response['data']
            else:  # Redirect if user detail not found
                # messages.error(request, response['msg'])
                return redirect('admin-article-list')

        except Exception as exp:
            print(exp)
            # messages.error(request, response['msg'])
            return redirect('admin-article-list')

        domain_list = get_network_domain_list(request)
        status_list = get_lookup_status_list(request, 17)

        try:
            api_url = settings.API_URL + '/api-blog/blog-category/'
            asset_list = call_api_post_method(
                {}, api_url, request.session['token']['access_token'])
            if 'error' in asset_list and asset_list['error'] == 1:
                # messages.error(request, asset_list['msg'])
                return redirect('admin-article-list')
            asset_list = asset_list['data']
        except:
            asset_list = {}

        context = {
            "data": data,
            "domain_list": domain_list,
            'status_list': status_list,
            'aws_url': settings.AWS_URL,
            "asset_list": asset_list
        }

        return render(request, 'admin/article/edit-article.html', context)
    except Exception as exp:
        print(exp)
        return


# Video Tutorials
def video_list(request):
    """ Use to list video
    """
    try:
        domain_list = status_list = []
        if request.method == 'GET':  # add blog
            domain_list = get_network_domain_list(request)
            status_list = get_lookup_status_list(request, 19)
        else:
            pass

        return render(request, "admin/video/video-list.html", {"domain_list": domain_list, "status_list": status_list, "active_submenu": "video-list"})
    except Exception as exp:
        print(exp)
        return HttpResponse("Issue in views")


@csrf_exempt
def ajax_video_list(request):
    """ Use to load video list from ajax
    """
    article_data = {}
    pagination = {}
    page_size = 20
    page_size = int(request.POST['count']) if 'count' in request.POST and request.POST['count'] else 20
    current_page = int(request.POST['page']) if 'page' in request.POST else 1
    current_page = int(request.POST['page']) if 'page' in request.POST else 1
    api_url = settings.API_URL + '/api-cms/super-admin-video-tutorials-listing/'
    params = {
        'page': current_page,
        'page_size': page_size,
        'search': request.POST['search'] if 'search' in request.POST else '',
        'site_id': request.POST.getlist('site_id[]'),
        'status': request.POST.getlist('status[]'),
        "user_id": request.session['user_id'],
    }
    try:
        response = call_api_post_method(params, api_url, request.session['token']['access_token'])
        print(response)
    except Exception as exp:
        # msg = err
        response = {'msg': exp, 'status': 422}
    else:
        if "error" in response and response['error'] == 0:
            article_data = response['data']
            total_contacts = response['data']['total']
            total_pages = total_contacts / page_size
            if total_contacts % page_size != 0:
                total_pages += 1  # adding one more page if the last page
                # will contains less contacts

            pagination = make_pagination_html(current_page, total_pages, 'video', 'video_list')

    context = {
        'data': article_data,
        'pagination': pagination,
        'aws_url': settings.AWS_URL,
        'start_index': (current_page - 1) * page_size
    }
    return render(request, 'admin/video/ajax-video-list.html', context)


# def active_inactive_video(request):
#     """ Use to active/in-active video
#     """
#     try:
#         error = 1
#         if request.method == 'POST':
#             api_url = settings.API_URL + '/api-cms/admin-article-change-status/'
#             payload = {
#                 'article_id': request.POST['article_id'],
#                 'status': request.POST['status']
#             }
#             try:
#                 response = call_api_post_method(payload, api_url, request.session['admin_token']['access_token'])
#                 msg = response['msg']
#                 error = response['error']
#             except Exception as exp:
#                 msg = exp
#         else:
#             msg = 'Get action not allowed'

#         return JsonResponse({"msg": msg, "error": error})
#     except Exception as exp:
#         print(exp)
#         return JsonResponse({"msg": exp, "error": 1})


def ajax_add_video(request):
    """ add/update video tutorials from ajax
    """
    try:
        error = 1
        if request.method == 'POST':

            api_url = settings.API_URL + '/api-cms/super-admin-add-video-tutorials/'

            payload = {
                "site_id": "",
                "user_id": request.session['user_id'],
                "title": request.POST['title'],
                "description": request.POST['description'],
                "video_url": request.POST['video_url'],
                "status": request.POST['status'],
            }

            # check if edit blog
            if 'video_id' in request.POST and request.POST:
                payload['video_id'] = int(request.POST['video_id'])

            try:
                response = call_api_post_method(
                    payload, api_url, request.session['token']['access_token'])
                msg = response['msg']
                error = response['error']
            except Exception as exp:
                msg = exp

        else:
            msg = 'Get action not allowed'

        return JsonResponse({"msg": msg, "error": error})
    except Exception as exp:
        # print(exp)
        return JsonResponse({"msg": exp, "error": 1})


@csrf_exempt
def ajax_video_details(request):
    """ get video details with id from ajax
    """
    try:
        data = []
        if request.method == 'POST':
            api_url = settings.API_URL + '/api-cms/super-admin-video-tutorials-detail/'
            payload = {
                'video_id': request.POST['id'],
                'user_id': request.session['user_id'],
            }
            try:
                response = call_api_post_method(
                    payload, api_url, request.session['token']['access_token'])
            except Exception as exp:
                response = {'data': {}, 'error': 1, 'msg': exp}

        else:
            msg = 'Get action not allowed'
            response = {'data': {}, 'error': 1, 'msg': msg}

        return JsonResponse(response)
    except Exception as exp:
        data = {'status': 403, 'data': {}, 'error': 1, 'msg': exp}
        return JsonResponse(data)


# Email Template Management
def email_template_list(request):
    """ Use to list email templates
    """
    try:
        domain_list = get_network_domain_list(request)
        status_list = get_lookup_status_list(request, 20)
        return render(request, "admin/email-template/template-list.html", {"domain_list": domain_list, "status_list": status_list, "active_submenu": "email-template-list"})
    except Exception as exp:
        print(exp)
        return HttpResponse("Issue in views")


@csrf_exempt
def ajax_email_template_list(request):
    """ Use to load email template list from ajax
    """
    template_data = {}
    pagination = {}
    page_size = 20
    page_size = int(request.POST['count']
                    ) if 'count' in request.POST and request.POST['count'] else 20
    current_page = int(request.POST['page']) if 'page' in request.POST else 1
    current_page = int(
        request.POST['page']) if 'page' in request.POST and request.POST['page'] else 1
    api_url = settings.API_URL + '/api-notifications/template-listing/'
    params = {
        'page': current_page,
        'page_size': page_size,
        'search': request.POST['search'] if 'search' in request.POST else '',
        'site_id': request.POST.getlist('site_id[]'),
        'status': request.POST.getlist('status[]')
    }
    try:
        response = call_api_post_method(
            params, api_url, request.session['token']['access_token'])
    except Exception as exp:
        # msg = err
        response = {'msg': exp, 'status': 422}
    else:
        if "error" in response and response['error'] == 0:
            template_data = response['data']
            total_templates = response['data']['total']
            total_pages = total_templates / page_size
            if total_templates % page_size != 0:
                total_pages += 1  # adding one more page if the last page
                # will contains less contacts

            pagination = make_pagination_html(current_page, total_pages,
                                              'email-template', 'email_template_list')

    context = {
        'data': template_data,
        'pagination': pagination,
        'start_index': (current_page - 1) * page_size
    }
    return render(request, 'admin/email-template/ajax-template-list.html', context)


def active_inactive_email_template(request):
    """ Use to active/in-active email template
    """
    try:
        error = 1
        if request.method == 'POST':
            api_url = settings.API_URL + '/api-notifications/template-change-status/'
            payload = {
                'template_id': request.POST['template_id'],
                'status': request.POST['status']
            }
            try:
                response = call_api_post_method(
                    payload, api_url, request.session['token']['access_token'])
                msg = response['msg']
                error = response['error']
            except Exception as exp:
                msg = exp
        else:
            msg = 'Get action not allowed'

        return JsonResponse({"msg": msg, "error": error})
    except Exception as exp:
        # print(exp)
        return JsonResponse({"msg": exp, "error": 1})


def add_email_template(request):
    """ Use to add email template
    """
    try:
        domain_list = status_list = event_list = []
        if request.method == 'GET':
            domain_list = get_network_domain_list(request)
            status_list = get_lookup_status_list(request, 20)
            event_list = get_event_list(request)
        else:
            pass

        context = {
            'domain_list': domain_list,
            'event_list': event_list,
            'status_list': status_list
        }
        return render(request, 'admin/email-template/add-template.html', context)
    except Exception as exp:
        print(exp)
        return


def ajax_add_email_template(request):
    """ add cms from ajax
    """
    try:
        error = 1
        if request.method == 'POST':

            api_url = settings.API_URL + '/api-notifications/add-template/'

            payload = {
                "user_id": request.session['user_id'],
                "site": "",
                "event": request.POST['event'],
                "email_subject": request.POST['email_subject'],
                "email_content": request.POST['email_content'],
                "status": request.POST['status'],
            }

            # check if edit cms
            if 'template_id' in request.POST and request.POST:
                payload['template_id'] = int(request.POST['template_id'])

            try:
                response = call_api_post_method(
                    payload, api_url, request.session['token']['access_token'])
                msg = response['msg']
                error = response['error']
            except Exception as exp:
                msg = exp

        else:
            msg = 'Get action not allowed'

        return JsonResponse({"msg": msg, "error": error})
    except Exception as exp:
        print(exp)
        return JsonResponse({"msg": exp, "error": 1})


def edit_email_template(request, id):
    """ Used to edit email template
    """
    """ Use to add user from admin end
    """
    try:

        api_url = settings.API_URL + '/api-notifications/template-detail/'
        data = domain_list = status_list = event_list = []
        try:
            response = call_api_post_method(
                {'template_id': id}, api_url, request.session['token']['access_token'])
            if "error" in response and response['error'] == 0:
                data = response['data']
            else:  # Redirect if user detail not found
                # messages.error(request, response['msg'])
                return redirect('admin-cms-list')

        except Exception as exp:
            # print(exp)
            # messages.error(request, response['msg'])
            return redirect('admin-cms-list')

        domain_list = get_network_domain_list(request)
        status_list = get_lookup_status_list(request, 20)
        event_list = get_event_list(request)

        context = {
            "data": data,
            "domain_list": domain_list,
            'status_list': status_list,
            'event_list': event_list
        }
        return render(request, 'admin/email-template/edit-template.html', context)
    except Exception as exp:
        print(exp)
        return


@csrf_exempt
def estimator_category_list(request):
    """ Use to list/App/Update available category for
    property bot value estimator
    """
    try:
        admin_id = None
        token = None
        if 'user_id' in request.session and request.session['user_id']:
            token = request.session['token']['access_token']
            admin_id = request.session['user_id']

        page_size = 20
        page = 1
        if request.is_ajax() and request.method == 'POST':

            estimator_search = ''
            if 'search' in request.POST and request.POST['search']:
                estimator_search = request.POST['search']

            page = 1
            if 'page' in request.POST and request.POST['page'] != "":
                page = request.POST['page']

            page_size = 20
            if 'page_size' in request.POST and request.POST['page_size']:
                page_size = request.POST['page_size']

            if 'status' in request.POST and request.POST['status'] and request.POST['status'] and request.POST[
                'status'].lower() == 'active':
                status = [1]
            elif 'status' in request.POST and request.POST['status'] and request.POST['status'] and request.POST[
                'status'].lower() == 'inactive':
                status = [2]
            else:
                status = [2, 1]

            list_param = {
                "admin_id": admin_id,
                "search": estimator_search,
                "status":status,
                "page_size": page_size,
                "page": page
            }
            sno = (int(page) - 1) * int(page_size) + 1
            list_url = settings.API_URL + '/api-property/property-evaluator-category-list/'
            list_data = call_api_post_method(list_param, list_url, token)

            if 'error' in list_data and list_data['error'] == 0:
                estimator_category_list = list_data['data']['data']
                total = list_data['data']['total']
            else:
                estimator_category_list = []
                total = 0


            context = {'estimator_category_list': estimator_category_list, 'total': total, 'sno': sno}

            estimator_listing_path = 'admin/property-estimator/category/category-list-content.html'
            estimator_listing_template = get_template(estimator_listing_path)
            estimator_listing_html = estimator_listing_template.render(context)
            try:
                total_pages = int(total) / int(page_size)
            except:
                total_pages = 0

            if int(total) % int(page_size) != 0:
                total_pages = int(total_pages) + 1  # adding one more page if the last page
                # will contains less contacts
            pagination_html = make_pagination_html(page, total_pages,
                                              'estimator_category', 'estimator_category_list')

            data = {'estimator_listing_html': estimator_listing_html, 'status': 200, 'msg': '', 'error': 0, 'total': total,
                    "pagination_html": pagination_html, 'pagination_id': 'estimator_pagination_list'}
            return JsonResponse(data)
        else:
            list_param = {
                "admin_id": admin_id,
                "search": "",
                "status":[1,2],
                "page_size": page_size,
                "page": page
            }

            list_url = settings.API_URL + '/api-property/property-evaluator-category-list/'
            list_data = call_api_post_method(list_param, list_url, token)

            sno = (int(page) - 1) * int(page_size) + 1
            if 'error' in list_data and list_data['error'] == 0:
                estimator_category_list = list_data['data']['data']
                total = list_data['data']['total']
            else:
                estimator_category_list = []
                total = 0
            # ---------------Pagination--------

            try:
                total_pages = int(total) / int(page_size)
            except:
                total_pages = 0

            if int(total) % int(page_size) != 0:
                total_pages = int(total_pages) + 1  # adding one more page if the last page
                # will contains less contacts

            pagination_html = make_pagination_html(page, total_pages,
                                              'estimator_category', 'estimator_category_list')


            context = {'estimator_category_list': estimator_category_list, 'total': total, "pagination_html": pagination_html,
                       "pagination_id": "estimator_pagination_list", "active_menu": "estimator",
                       "active_submenu": "estimator_list", "sno": sno}

            return render(request, "admin/property-estimator/category/category-list.html", context)
    except Exception as exp:
        print(exp)
        return HttpResponse("Issue in views")

@csrf_exempt
def get_estimator_category_details(request):
    """ Use to get category property
        bot value estimator category details
        """
    try:
        if request.is_ajax() and request.method == 'POST':
            admin_id = None
            token = None
            if 'user_id' in request.session and request.session['user_id']:
                token = request.session['token']['access_token']
                admin_id = request.session['user_id']

            params = {"admin_id": admin_id, 'category_id': request.POST['category_id']}

            url = settings.API_URL + '/api-property/property-evaluator-category-detail/'
            data = call_api_post_method(params, url, token)

            if 'error' in data and data['error'] == 0:
                category_details = data['data']
                data = {'error': 0, 'status': 200, 'msg': '', 'category_details': category_details}
            else:
                data = {'error': 1, 'status': 403, 'msg': 'Server error, Please try again', 'category_details': {}}
        else:
            data = {'error': 1, 'status': 403, 'msg': 'Forbidden', 'category_details': {}}

        return JsonResponse(data)
    except Exception as exp:
        print(exp)
        data = {'status': 403, 'msg': 'invalid request.', 'category_details': {}}
        return JsonResponse(data)


def save_estimator_category(request):
    try:
        if request.is_ajax() and request.method == 'POST':
            admin_id = None
            token = None
            if 'user_id' in request.session and request.session['user_id']:
                token = request.session['token']['access_token']
                admin_id = request.session['user_id']

            params = {
                'admin_id': admin_id,
                'name': request.POST['category_name'],
                'status': request.POST['is_active'],
            }
            if 'estimator_category_id' in request.POST and request.POST['estimator_category_id']:
                params['category_id'] = int(request.POST['estimator_category_id'])
            url = settings.API_URL + '/api-property/add-property-evaluator-category/'

            data = call_api_post_method(params, url, token)
            if 'error' in data and data['error'] == 0:
                data = {
                    'status': 200, 'data': data['data'], 'error': 0, 'msg': 'Category saved successfully'}
            else:
                data = {'status': 403, 'data': {}, 'error': 1, 'msg': data['msg']}
        else:
            data = {'status': 403, 'data': {}, 'error': 1, 'msg': 'Only POST methods are allowed'}

        return JsonResponse(data)
    except Exception as exp:
        print(exp)
        data = {'status': 403, 'data': {}, 'error': 1, 'msg': exp}
        return JsonResponse(data)


@csrf_exempt
def estimator_question_list(request):
    """ Use to list available for
    property bot value estimator
    """
    try:
        admin_id = None
        token = None
        if 'user_id' in request.session and request.session['user_id']:
            token = request.session['token']['access_token']
            admin_id = request.session['user_id']

        page_size = 20
        page = 1
        if request.is_ajax() and request.method == 'POST':

            estimator_search = ''
            if 'search' in request.POST and request.POST['search']:
                estimator_search = request.POST['search']

            page = 1
            if 'page' in request.POST and request.POST['page'] != "":
                page = request.POST['page']

            page_size = 20
            if 'page_size' in request.POST and request.POST['page_size']:
                page_size = request.POST['page_size']

            if 'status' in request.POST and request.POST['status'] and request.POST['status'] and request.POST[
                'status'].lower() == 'active':
                status = [1]
            elif 'status' in request.POST and request.POST['status'] and request.POST['status'] and request.POST[
                'status'].lower() == 'inactive':
                status = [2]
            else:
                status = [2, 1]

            list_param = {
                "admin_id": admin_id,
                "search": estimator_search,
                "status":status,
                "page_size": page_size,
                "page": page
            }
            sno = (int(page) - 1) * int(page_size) + 1
            list_url = settings.API_URL + '/api-property/property-evaluator-question-list/'
            list_data = call_api_post_method(list_param, list_url, token)

            if 'error' in list_data and list_data['error'] == 0:
                estimator_question_list = list_data['data']['data']
                total = list_data['data']['total']
            else:
                estimator_question_list = []
                total = 0


            context = {'estimator_question_list': estimator_question_list, 'total': total, 'sno': sno}

            estimator_listing_path = 'admin/property-estimator/estimator/question-list-content.html'
            estimator_listing_template = get_template(estimator_listing_path)
            estimator_listing_html = estimator_listing_template.render(context)
            try:
                total_pages = int(total) / int(page_size)
            except:
                total_pages = 0

            if int(total) % int(page_size) != 0:
                total_pages = int(total_pages) + 1  # adding one more page if the last page
                # will contains less contacts
            pagination_html = make_pagination_html(page, total_pages,
                                              'estimator_category', 'estimator_category_list')

            data = {'estimator_listing_html': estimator_listing_html, 'status': 200, 'msg': '', 'error': 0, 'total': total,
                    "pagination_html": pagination_html, 'pagination_id': 'estimator_pagination_list'}
            return JsonResponse(data)
        else:
            list_param = {
                "admin_id": admin_id,
                "search": "",
                "status":[1,2],
                "page_size": page_size,
                "page": page
            }

            list_url = settings.API_URL + '/api-property/property-evaluator-question-list/'
            list_data = call_api_post_method(list_param, list_url, token)

            sno = (int(page) - 1) * int(page_size) + 1
            if 'error' in list_data and list_data['error'] == 0:
                estimator_question_list = list_data['data']['data']
                total = list_data['data']['total']
            else:
                estimator_question_list = []
                total = 0
            # ---------------Pagination--------

            try:
                total_pages = int(total) / int(page_size)
            except:
                total_pages = 0

            if int(total) % int(page_size) != 0:
                total_pages = int(total_pages) + 1  # adding one more page if the last page
                # will contains less contacts

            pagination_html = make_pagination_html(page, total_pages,
                                              'estimator_question', 'estimator_question_list')



            context = {'estimator_question_list': estimator_question_list, 'total': total, "pagination_html": pagination_html,
                       "pagination_id": "estimator_pagination_list", "active_menu": "estimator",
                       "active_submenu": "estimator_list", "sno": sno}

            return render(request, "admin/property-estimator/estimator/question-list.html", context)
    except Exception as exp:
        print(exp)
        return HttpResponse("Issue in views")

@csrf_exempt
def save_estimator_question(request):
    question_id = request.GET.get('id', None)
    try:
        admin_id = None
        token = None
        if 'user_id' in request.session and request.session['user_id']:
            token = request.session['token']['access_token']
            admin_id = request.session['user_id']

        try:
            category_param = {
                "admin_id": admin_id,
                "search": "",
                "status": [1],
            }
            category_url = settings.API_URL + '/api-property/property-evaluator-category-list/'
            category_data = call_api_post_method(category_param, category_url, token)
            category_list = category_data['data']['data']
        except:
            category_list = []

        try:
            asset_listing = get_admin_property_types(request)
        except:
            asset_listing = []

        question_type_list = [
            {
                "id": 1,
                "name": "Text Area"
            },
            {
                "id": 2,
                "name": "Radio"
            },
            {
                "id": 3,
                "name": "DropDown"
            },
            {
                "id": 4,
                "name": "Image"
            },
            {
                "id": 5,
                "name": "Rating"
            },
            {
                "id": 6,
                "name": "Map"
            },

        ]
        status_list = [
            {
                "id": 1,
                "name": "Active"
            },
            {
                "id": 2,
                "name": "Inactive"
            }
        ]

        try:
            question_param = {
                'admin_id': admin_id,
                'question_id': question_id
            }
            question_url = settings.API_URL + '/api-property/property-evaluator-question-detail/'
            question_data = call_api_post_method(question_param, question_url, token)
            if 'error' in question_data and question_data['error'] == 1 and question_id:

                http_host = request.META['HTTP_HOST']
                redirect_url = settings.URL_SCHEME + str(http_host)
                return HttpResponseRedirect(redirect_url+'/admin/estimator-question-list/')

            question_details = question_data['data']
            question_id = question_details['id']

        except:
            question_details = {}

        if request.is_ajax() and request.method == 'POST':
            option_list = []
            if int(request.POST['option_type']) in [2,3]:
                total_option = int(request.POST['total_option'])
                for x in range(total_option):
                    # option_params = {
                    #     'question_option': request.POST['option_value_' + str(x)],
                    # }
                    #option_param = request.POST['option_value_' + str(x)]
                    option_list.append(request.POST['option_value_' + str(x)])


            ques_params = {
                "admin_id": 1,
                "category":request.POST['question_category'],
                "question": request.POST['question'],
                "option_type": request.POST['option_type'],
                "property_type": request.POST['asset_type'] if 'asset_type' in request.POST and request.POST['asset_type'] else "",
                "status": request.POST['question_status'],
                "question_option": option_list,
                "placeholder": request.POST['text_placeholder'] if 'text_placeholder' in request.POST and request.POST['text_placeholder'] else "",
            }

            if request.POST['question_id']:
                ques_params['question_id'] = request.POST['question_id']

            ques_url = settings.API_URL + '/api-property/add-property-evaluator-question/'
            ques_response = call_api_post_method(ques_params, ques_url, token)
            
            if 'error' in ques_response and ques_response['error'] == 0:
                response = {
                    'error': 0,
                    'msg': ques_response['msg']
                }
            else:
                response = {
                    'error': 1,
                    'msg': ques_response['msg']
                }
            return JsonResponse(response)

        # print(question_details)
        context = {"category_list": category_list, 'question': question_details, 'question_type_list': question_type_list, 'asset_listing': asset_listing, 'status_list': status_list}
        return render(request, "admin/property-estimator/estimator/add-question.html", context)
    except Exception as exp:
        print(exp)
        return HttpResponse("Issue in views")
# @csrf_exempt
# def new_bid_checked(request):
#     try:
#         if request.is_ajax() and request.method == 'POST':
#             token = request.session['admin_token']['access_token']
#             params = {
#                 'user_id':request.session['admin_id'],
#                 "property_id": request.POST['property_id'],
#             }
#             response = settings.API_URL + '/api-property/auction-listing-read/'
#             data = call_api_post_method(params, response, token)
#             if 'error' in data and data['error'] == 0:
#                 data = {'status': 200, 'msg': data['msg'], 'error': 0}
#             else:
#                 data = {'status': 403, 'msg': 'Something went wrong!', 'error': 1}
#         else:
#             data = {'status': 403, 'msg': 'Forbidden', 'error': 1}
#         return JsonResponse(data)
#     except Exception as exp:
#         print(exp)
#         data = {'status': 403, 'msg': 'Something went wrong!', 'error': 1}
#         return JsonResponse(data)

@csrf_exempt
def property_estimator_list(request):
    """ Use to list available for
    property bot value estimator
    """
    try:
        admin_id = None
        token = None
        if 'user_id' in request.session and request.session['user_id']:
            token = request.session['token']['access_token']
            admin_id = request.session['user_id']

        page_size = 20
        page = 1
        domain_list = get_network_domain_list(request)
        if request.is_ajax() and request.method == 'POST':

            estimator_search = ''
            if 'search' in request.POST and request.POST['search']:
                estimator_search = request.POST['search']

            page = 1
            if 'page' in request.POST and request.POST['page'] != "":
                page = request.POST['page']

            page_size = 20
            if 'page_size' in request.POST and request.POST['page_size']:
                page_size = request.POST['page_size']

            domain_id = []
            domain_id = request.POST.getlist('domain_id[]') if 'domain_id[]' in request.POST else request.POST['domain_id']
            list_param = {
                "admin_id": admin_id,
                "page_size": page_size,
                "page": page,
                "domain_id": domain_id,
                "search": estimator_search
            }
            sno = (int(page) - 1) * int(page_size) + 1
            list_url = settings.API_URL + '/api-property/superadmin-property-evaluator-list/'
            list_data = call_api_post_method(list_param, list_url, token)
            if 'error' in list_data and list_data['error'] == 0:
                estimator_list = list_data['data']['data']
                total = list_data['data']['total']
            else:
                estimator_list = []
                total = 0


            context = {'estimator_list': estimator_list, 'total': total, 'sno': sno}

            estimator_listing_path = 'admin/property-estimator/property-bot/property-estimator-listing-content.html'
            estimator_listing_template = get_template(estimator_listing_path)
            estimator_listing_html = estimator_listing_template.render(context)
            try:
                total_pages = int(total) / int(page_size)
            except:
                total_pages = 0

            if int(total) % int(page_size) != 0:
                total_pages = int(total_pages) + 1  # adding one more page if the last page
                # will contains less contacts
            pagination_html = make_pagination_html(page, total_pages,
                                              'property_estimator', 'property_estimator_list')

            data = {'estimator_listing_html': estimator_listing_html, 'status': 200, 'msg': '', 'error': 0, 'total': total,
                    "pagination_html": pagination_html, 'pagination_id': 'estimator_pagination_list'}
            return JsonResponse(data)
        else:
            list_param = {
                "admin_id": admin_id,
                "page_size": page_size,
                "page": page,
                "domain_id": [],
                "search": ""
            }

            list_url = settings.API_URL + '/api-property/superadmin-property-evaluator-list/'
            list_data = call_api_post_method(list_param, list_url, token)

            sno = (int(page) - 1) * int(page_size) + 1
            if 'error' in list_data and list_data['error'] == 0:
                estimator_list = list_data['data']['data']
                total = list_data['data']['total']
            else:
                estimator_list = []
                total = 0
            # ---------------Pagination--------

            try:
                total_pages = int(total) / int(page_size)
            except:
                total_pages = 0

            if int(total) % int(page_size) != 0:
                total_pages = int(total_pages) + 1  # adding one more page if the last page
                # will contains less contacts

            pagination_html = make_pagination_html(page, total_pages,
                                              'property_estimator', 'property_estimator_list')



            context = {'estimator_list': estimator_list, 'total': total, "pagination_html": pagination_html,
                       "pagination_id": "estimator_pagination_list", "active_menu": "property_estimator",
                       "active_submenu": "property_estimator_list", "sno": sno, 'domain_list': domain_list}

            return render(request, "admin/property-estimator/property-bot/property-estimator-list.html", context)
    except Exception as exp:
        print(exp)
        return HttpResponse("Issue in views")

@csrf_exempt
def property_estimator_details(request):
    try:
        bot_id = request.GET.get('id', None)
        address_question_list = []
        admin_id = None
        token = None
        if 'user_id' in request.session and request.session['user_id']:
            token = request.session['token']['access_token']
            admin_id = request.session['user_id']

        try:
            params = {
                "admin_id": admin_id,
                "bot_id": bot_id,
                "category_id": 1,
            }
            api_url = settings.API_URL + '/api-property/superadmin-property-evaluator-detail/'
            estimator_data = call_api_post_method(params, api_url, token=token)
            address_question_list = estimator_data['data']

            if address_question_list:
                for ques in address_question_list:
                    if ques['option_type'] == 6:
                        if ques['answer'] and ques['answer']['answer']:
                            address = ques['answer']['answer']
                            geolocator = Nominatim(user_agent="arvindm@clavax.com", timeout=10)
                            location = geolocator.geocode(address)
                            try:
                                ques['latitude'] = location.latitude
                                ques['longitude'] = location.longitude
                            except:
                                ques['latitude'] = 33.1836304229855
                                ques['longitude'] = -117.32728034223504
                        else:
                            ques['latitude'] = 33.1836304229855
                            ques['longitude'] = -117.32728034223504

                    if ques['option_type'] == 6:
                        formatted_addr = ''
                        addr_i = 0
                        if ques['answer'] and ques['answer']['answer']:
                            address_arr = ques['answer']['answer'].split(',')
                            if address_arr:
                                for addr in address_arr:
                                    if addr_i == 0:
                                        formatted_addr = addr
                                    elif addr_i == 1:
                                        if len(address_arr) == 2:
                                            formatted_addr = formatted_addr + ', ' + addr
                                        else:
                                            formatted_addr = formatted_addr + ', <br>' + addr
                                    elif addr_i == 2:
                                        if len(address_arr) == 3:
                                            formatted_addr = formatted_addr + ', ' + addr
                                        else:
                                            formatted_addr = formatted_addr + ', ' + addr + ', <br>'
                                    else:
                                        formatted_addr = formatted_addr + ' ' + addr

                                    addr_i = addr_i + 1
                            ques['formatted_addr'] = formatted_addr
        except:
            address_question_list = []

        try:
            params = {
                "admin_id": admin_id,
                "bot_id": bot_id,
                "category_id": 2,
            }
            api_url = settings.API_URL + '/api-property/superadmin-property-evaluator-detail/'
            estimator_data = call_api_post_method(params, api_url, token=token)
            details_question_list = estimator_data['data']
            if details_question_list:
                for ques in details_question_list:
                    if ques['option_type'] == 6:
                        if ques['answer'] and ques['answer']['answer']:
                            address = ques['answer']['answer']
                            geolocator = Nominatim(user_agent="arvindm@clavax.com", timeout=10)
                            location = geolocator.geocode(address)
                            try:
                                ques['latitude'] = location.latitude
                                ques['longitude'] = location.longitude
                            except:
                                ques['latitude'] = 33.1836304229855
                                ques['longitude'] = -117.32728034223504
                        else:
                            ques['latitude'] = 33.1836304229855
                            ques['longitude'] = -117.32728034223504

                    if ques['option_type'] == 6:
                        formatted_addr = ''
                        addr_i = 0
                        if ques['answer'] and ques['answer']['answer']:
                            address_arr = ques['answer']['answer'].split(',')
                            if address_arr:
                                for addr in address_arr:
                                    if addr_i == 0:
                                        formatted_addr = addr
                                    elif addr_i == 1:
                                        if len(address_arr) == 2:
                                            formatted_addr = formatted_addr + ', ' + addr
                                        else:
                                            formatted_addr = formatted_addr + ', <br>' + addr
                                    elif addr_i == 2:
                                        if len(address_arr) == 3:
                                            formatted_addr = formatted_addr + ', ' + addr
                                        else:
                                            formatted_addr = formatted_addr + ', ' + addr + ', <br>'
                                    else:
                                        formatted_addr = formatted_addr + ' ' + addr

                                    addr_i = addr_i + 1
                            ques['formatted_addr'] = formatted_addr
        except:
            details_question_list = []

        try:
            params = {
                "admin_id": admin_id,
                "bot_id": bot_id,
                "category_id": 3,
            }
            api_url = settings.API_URL + '/api-property/superadmin-property-evaluator-detail/'
            estimator_data = call_api_post_method(params, api_url, token=token)
            doc_question_list = estimator_data['data']
            if doc_question_list:
                for ques in doc_question_list:
                    if ques['option_type'] == 6:
                        if ques['answer'] and ques['answer']['answer']:
                            address = ques['answer']['answer']
                            geolocator = Nominatim(user_agent="arvindm@clavax.com", timeout=10)
                            location = geolocator.geocode(address)
                            try:
                                ques['latitude'] = location.latitude
                                ques['longitude'] = location.longitude
                            except:
                                ques['latitude'] = 33.1836304229855
                                ques['longitude'] = -117.32728034223504
                        else:
                            ques['latitude'] = 33.1836304229855
                            ques['longitude'] = -117.32728034223504
                    if ques['option_type'] == 6:
                        formatted_addr = ''
                        addr_i = 0
                        if ques['answer'] and ques['answer']['answer']:
                            address_arr = ques['answer']['answer'].split(',')
                            if address_arr:
                                for addr in address_arr:
                                    if addr_i == 0:
                                        formatted_addr = addr
                                    elif addr_i == 1:
                                        if len(address_arr) == 2:
                                            formatted_addr = formatted_addr + ', ' + addr
                                        else:
                                            formatted_addr = formatted_addr + ', <br>' + addr
                                    elif addr_i == 2:
                                        if len(address_arr) == 3:
                                            formatted_addr = formatted_addr + ', ' + addr
                                        else:
                                            formatted_addr = formatted_addr + ', ' + addr + ', <br>'
                                    else:
                                        formatted_addr = formatted_addr + ' ' + addr

                                    addr_i = addr_i + 1
                            ques['formatted_addr'] = formatted_addr
        except:
            doc_question_list = []

        try:
            params = {
                "admin_id": admin_id,
                "bot_id": bot_id,
                "category_id": 4,
            }
            api_url = settings.API_URL + '/api-property/superadmin-property-evaluator-detail/'
            estimator_data = call_api_post_method(params, api_url, token=token)
            additional_question_list = estimator_data['data']
            if additional_question_list:
                for ques in additional_question_list:
                    if ques['option_type'] == 6:
                        formatted_addr = ''
                        addr_i = 0
                        if ques['answer'] and ques['answer']['answer']:
                            address_arr = ques['answer']['answer'].split(',')
                            if address_arr:
                                for addr in address_arr:
                                    if addr_i == 0:
                                        formatted_addr = addr
                                    elif addr_i == 1:
                                        if len(address_arr) == 2:
                                            formatted_addr = formatted_addr + ', ' + addr
                                        else:
                                            formatted_addr = formatted_addr + ', <br>' + addr
                                    elif addr_i == 2:
                                        if len(address_arr) == 3:
                                            formatted_addr = formatted_addr + ', ' + addr
                                        else:
                                            formatted_addr = formatted_addr + ', ' + addr + ', <br>'
                                    else:
                                        formatted_addr = formatted_addr + ' ' + addr

                                    addr_i = addr_i + 1
                            ques['formatted_addr'] = formatted_addr
        except:
            additional_question_list = []

        context = {
            "address_question_list": address_question_list,
            "details_question_list": details_question_list,
            "doc_question_list": doc_question_list,
            "additional_question_list": additional_question_list,
            "active_menu": "property_estimator",
            "active_submenu": "property_estimator_list",
            "aws_url": settings.AWS_URL,
        }
        return render(request, "admin/property-estimator/property-bot/estimator-details.html", context)
    except Exception as exp:
        print(exp)
        return HttpResponse("Issue in views")

@csrf_exempt
def get_latitude_longitude(request):
    try:
        if request.is_ajax() and request.method == 'POST':
            address = request.POST['address']
            geolocator = Nominatim(user_agent="arvindm@clavax.com", timeout=10)
            location = geolocator.geocode(address)
            try:
                latitude = location.latitude
                longitude= location.longitude
            except Exception as exp:
                print(exp)
                latitude = 33.1836304229855
                longitude = -117.32728034223504
            data = {
                'address': address,
                'latitude': latitude,
                'longitude': longitude,
            }
            return JsonResponse(data)
    except Exception as exp:
        print(exp)
        return HttpResponse("Issue in views")            


@register.filter(name='phone_split')
def phone_split(x):
    """Convert a 10 character string into (xxx) xxx-xxxx."""
    try: 
        first = x[0:3]
        second = x[3:6]
        third = x[6:10]
        return '(' + first + ')' + ' ' + second + '-' + third
    except Exception as exp:
        print(exp)
        return x 


@register.filter(name='escape_char')
def escape_char(value):
    esc_char = ''
    if value:
        esc_char =  value.replace("'","\\'")

    return esc_char  


# Advertisement List
def advertisement_list(request):
    """ Use to list advertisements
    """
    try:
        domain_list = get_network_domain_list(request)
        status_list = get_lookup_status_list(request, 15)

        return render(
            request,
            "admin/advertisement/list.html",
            {
                'domain_list': domain_list,
                'status_list': status_list,
                'aws_url': settings.AWS_URL
            }
        )
    except Exception as exp:
        print(exp)
        return HttpResponse("Issue in views")


@csrf_exempt
def ajax_advertisement_list(request):
    """ Use to load advertisements list from ajax
    """
    advertisements = {}
    pagination = {}
    page_size = int(request.POST['count']
                    ) if 'count' in request.POST and request.POST['count'] else 20
    current_page = int(
        request.POST['page']) if 'page' in request.POST and request.POST['page'] else 1
    api_url = settings.API_URL + '/api-advertisement/super-admin-advertisement-listing/'
    params = {
        "user_id": request.session['user_id'],
        'page': current_page,
        'page_size': page_size,
        'search': request.POST['search'] if 'search' in request.POST else '',
        'domain': request.POST.getlist('site_id[]'),
        'status': request.POST.getlist('status[]'),
    }
    try:
        response = call_api_post_method(
            params, api_url, request.session['token']['access_token'])
    except Exception as exp:
        # msg = err
        response = {'msg': exp, 'status': 422}
    else:
        if "error" in response and response['error'] == 0:
            advertisements = response['data']
            total_contacts = response['data']['total']
            total_pages = total_contacts / page_size
            if total_contacts % page_size != 0:
                total_pages += 1  # adding one more page if the last page
                # will contains less contacts

            pagination = make_pagination_html(current_page, total_pages,
                                              'advertisement', 'advertisement_list')

    context = {
        'data': advertisements,
        'pagination': pagination,
        'aws_url': settings.AWS_URL,
        'start_index': (current_page - 1) * page_size
    }
    return render(request, 'admin/advertisement/ajax-list.html', context)


@csrf_exempt
def save_advertisement(request):
    try:
        if request.is_ajax() and request.method == 'POST':
            token = request.session['token']['access_token']
            try:
                site_detail = subdomain_site_details(request)
                site_id = site_detail['site_detail']['site_id']
            except:
                site_id = ""
            params = {
                'domain': "",
                'company_name': request.POST['company_name'],
                'url': request.POST['url'],
                'image': request.POST['upload'],
                'status': request.POST['status'],
                'added_by': request.session['user_id'],
            }
            if 'advertisement_id' in request.POST and request.POST['advertisement_id']:
                params['advertisement_id'] = int(
                    request.POST['advertisement_id'])
            url = settings.API_URL + '/api-advertisement/add-advertisement/'

            data = call_api_post_method(params, url, token)
            print("--------------")
            print(data)
            if 'error' in data and data['error'] == 0:
                data = {
                    'status': 200, 'data': data['data'], 'error': 0, 'msg': 'Ad saved successfully'}
            else:
                data = {'status': 403, 'data': {},
                        'error': 1, 'msg': data['msg']}
        else:
            data = {'status': 403, 'data': {}, 'error': 1,
                    'msg': 'Only POST methods are allowed'}

        return JsonResponse(data)
    except Exception as exp:
        print(exp)
        data = {'status': 403, 'data': {}, 'error': 1, 'msg': exp}
        return JsonResponse(data)


@csrf_exempt
def ajax_advertisement_details(request):
    """ get advertisement details with id from ajax
    """
    try:
        data = []
        if request.method == 'POST':
            api_url = settings.API_URL + '/api-advertisement/super-admin-advertisement-detail/'
            payload = {
                'advertisement_id': request.POST['id'],
                'user_id': request.session['user_id'],
            }
            try:
                response = call_api_post_method(
                    payload, api_url, request.session['token']['access_token'])
            except Exception as exp:
                response = {'data': {}, 'error': 1, 'msg': exp}

        else:
            msg = 'Get action not allowed'
            response = {'data': {}, 'error': 1, 'msg': msg}

        return JsonResponse(response)
    except Exception as exp:
        data = {'status': 403, 'data': {}, 'error': 1, 'msg': exp}
        return JsonResponse(data)   


@csrf_exempt
def admin_save_images(request):
    try:
        user_id = request.session['user_id']
        site_id = ""
        token = request.session['token']['access_token']
        file_urls = ""
        upload_to = ""
        uploaded_file_list = []
        file_size = request.POST['file_size']

        try:
            for key, value in request.FILES.items():
                params = {}
                if 'banner_img' in key.lower():
                    upload_to = 'banner_img'
                    doc_type = 1
                elif 'about_us_img' in key.lower():
                    upload_to = 'about_us_img'
                    doc_type = 3
                elif 'company_partner_img' in key.lower():
                    upload_to = 'company_partner_img'
                    doc_type = 8
                elif 'favicon_image' in key.lower():
                    doc_type = 4
                    upload_to = 'favicons'
                    file_urls = request.FILES['favicon_image']
                elif 'website_logo' in key.lower():
                    doc_type = 5
                    upload_to = 'website_logo'
                    file_urls = request.FILES['website_logo']

                elif 'testimonial_img' in key.lower():
                    doc_type = 6
                    upload_to = 'testimonial_img'
                    file_urls = request.FILES['testimonial_img']
                elif 'testimonial_author_img' in key.lower():
                    doc_type = 7
                    upload_to = 'testimonial_author_img'
                    file_urls = request.FILES['testimonial_author_img']
                elif 'agent_image' in key.lower():
                    doc_type = 9
                    upload_to = 'profile_image'
                    file_urls = request.FILES['agent_image']
                elif 'home_auction_image' in key.lower():
                    doc_type = 10
                    #upload_to = 'home_auction_image'
                    upload_to = 'home_auctioin'
                    file_urls = request.FILES['home_auction_image']
                elif 'expertise_image' in key.lower():
                    doc_type = 11
                    #upload_to = 'expertise_image'
                    upload_to = 'home_expertise'
                    file_urls = request.FILES['expertise_image']
                elif 'property_image' in key.lower():
                    doc_type = 13
                    upload_to = 'property_image'
                    #file_urls = request.FILES['property_image']
                elif 'property_document' in key.lower():
                    doc_type = 12
                    upload_to = 'property_document'
                elif 'ad_image' in key.lower():
                    doc_type = 17
                    upload_to = 'ad_image'
                    #file_urls = request.FILES['property_document']
                file_urls = request.FILES[key]
        except:
            pass

        if int(request.POST['file_length']) > 1:
            for key, value in request.FILES.items():
                params = {}
                if 'banner_img' in key.lower():
                    upload_to = 'banner_img'
                    doc_type = 1
                elif 'about_us_img' in key.lower():
                    upload_to = 'about_us_img'
                    doc_type = 3
                elif 'property_image' in key.lower():
                    doc_type = 13
                    upload_to = 'property_image'
                elif 'property_document' in key.lower():
                    doc_type = 12
                    upload_to = 'property_document'
                else:
                    upload_to = 'company_partner_img'
                    doc_type = 8

                file_res = request.FILES[key]
                response = save_to_s3(file_res, upload_to)
                if 'error' in response and response['error'] == 0:
                    try:
                        upload_param = {
                            "site_id": site_id,
                            "user_id": user_id,
                            "doc_file_name": response['file_name'],
                            "document_type": doc_type,
                            "bucket_name": upload_to,
                            "added_by": user_id,
                            "file_size": str(file_size)+'MB',
                            "is_admin": 1
                        }
                        url = settings.API_URL + '/api-users/file-upload/'
                        upload_data = call_api_post_method(
                            upload_param, url, token)
                        upload_id = upload_data['data']['upload_id']
                        upload_size = upload_data['data']['file_size']
                        upload_date = upload_data['data']['added_date']
                    except:
                        upload_id = 0
                        upload_size = '0MB'
                        upload_date = ''

                    params['file_name'] = response['file_name']
                    params['error'] = 0
                    params['msg'] = response['msg']
                    params['upload_id'] = upload_id
                    params['file_size'] = upload_size
                    params['upload_date'] = upload_date
                    params['upload_to'] = upload_to
                else:
                    params['file_name'] = response['file_name']
                    params['error'] = 1
                    params['msg'] = response['msg']
                    params['upload_id'] = 0
                    params['file_size'] = '0MB'
                    params['upload_date'] = ''
                    params['upload_to'] = upload_to

                uploaded_file_list.append(params)
        else:
            params = {}
            response = save_to_s3(file_urls, upload_to)
            if 'error' in response and response['error'] == 0:
                try:
                    upload_param = {
                        "site_id": site_id,
                        "user_id": user_id,
                        "doc_file_name": response['file_name'],
                        "document_type": doc_type,
                        "bucket_name": upload_to,
                        "added_by": user_id,
                        "file_size": str(file_size) + 'MB'
                    }
                    url = settings.API_URL + '/api-users/file-upload/'
                    upload_data = call_api_post_method(
                        upload_param, url, token)
                    upload_id = upload_data['data']['upload_id']
                    upload_size = upload_data['data']['file_size']
                    upload_date = upload_data['data']['added_date']
                except Exception as exp:
                    print(exp)
                    upload_id = 0
                    upload_size = '0MB'
                    upload_date = ''

                params['file_name'] = response['file_name']
                params['error'] = 0
                params['msg'] = response['msg']
                params['upload_id'] = upload_id
                params['file_size'] = upload_size
                params['upload_date'] = upload_date
                params['upload_to'] = upload_to
            else:
                params['file_name'] = response['file_name']
                params['error'] = 1
                params['msg'] = response['msg']
                params['upload_id'] = 0
                params['file_size'] = '0MB'
                params['upload_date'] = ''
                params['upload_to'] = upload_to
            uploaded_file_list.append(params)

        return JsonResponse({'status': 200, 'uploaded_file_list': uploaded_file_list})
    except Exception as exp:
        print(exp)
        data = {'status': 403, 'msg': 'invalid request.'}
        return JsonResponse(data) 


def verify_email(request):
    try:
        if request.is_ajax() and request.method == 'POST':
            site_detail = subdomain_site_details(request)
            site_id = site_detail['site_detail']['site_id']
            user_id = request.session['user_id']
            token = request.session['token']['access_token']
            # ---------------------Save Payment Detail Data------------------
            params_data = json.loads(request.POST['params_data'])
            api_url = settings.API_URL + "/api-notifications/verify-email/"
            params = {
                "domain_id": site_id,
                "params_data": params_data,
                "template_id": int(request.POST['template_list']),
                "email_to": request.POST['email_to'],
            }
            api_response = call_api_post_method(params, api_url, token)
            if 'error' in api_response and api_response['error'] == 0:
                data = {"data": "", "error": 0, "msg": api_response['msg']}
            else:
                data = {'data': "", 'error': 1, "msg": api_response['msg']}
            return JsonResponse(data)
        else:
            token = request.session['token']['access_token']
            url = settings.API_URL + '/api-notifications/template-list/'
            template_list = call_api_post_method({}, url, token)
            all_data = []
            if template_list['error'] == 0 and template_list['code'] == 1:
                all_data = template_list['data']
            context = {"all_data": all_data}
            return render(request, "admin/email-template/check-email.html", context)
    except Exception as exp:
        print("====================")
        print(exp)
        data = {'status': 403, 'msg': 'invalid request.'}
        return JsonResponse(data)                                                                               

